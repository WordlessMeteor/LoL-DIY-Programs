from lcu_driver import Connector
from lcu_driver.connection import Connection
import copy, os, json, time, pandas, re, requests
from openpyxl import load_workbook, Workbook
from typing import Any
from openpyxl.worksheet.worksheet import Worksheet
from src.utils.summoner import print_summoner_info, get_info_name
from src.utils.format import getISOTime, addDefaultStyle, optimize_bool_display, pyobj2json
from src.utils.excel_workbook import create_workbook_win32, sort_worksheet
from src.core.config.servers import set_platform_folder, set_summonerInfo_folder, save_platform_info
from src.core.dataframes.inventory import build_inventory_map
from src.localization.languages.zh_CN import inventoryType_dict, ownershipTypes, subInventoryTypes

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2026/09/01
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

session: requests.Session = requests.Session()
LoLChampions: dict[int, dict[str, Any]] = {}
catalogDict: dict[str, list[dict[str, Any]]] = {}
catalogList: list[dict[str, Any]] = []
store: list[dict[str, Any]] = []
storeDict: dict[str, list[dict[str, Any]]] = {}
collection: list[dict[str, Any]] = []

connector: Connector = Connector()

#-----------------------------------------------------------------------------
# 获取商品（Capture items in the store）
#-----------------------------------------------------------------------------
async def prepare_data_resources(connection: Connection) -> None: #准备数据资源（Prepare data resources）
    '''
    准备全局数据资源。<br>Prepare global data resources.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    '''
    global LoLChampions, catalogDict, catalogList, store, storeDict, collection
    #英雄（Champion）
    LoLChampions_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/champion-summary.json")).json()
    LoLChampions = {int(LoLChampion_iter["id"]): LoLChampion_iter for LoLChampion_iter in LoLChampions_source}
    #道具类型（Inventory type）
    lolinventorytype_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/lolinventorytype.json")).json()
    inventoryTypes: list[str] = sorted(list(map(lambda x: x["inventoryTypeId"], lolinventorytype_source)))
    #inventoryTypes: list[str] = ["ACHIEVEMENT_BANNER_ACCENT", "ACHIEVEMENT_TITLE", "ANNOUNCER_PACK", "ARAM_BOON", "AUGMENT", "AUGMENT_SLOT", "BOOST", "BUNDLES", "CHAMPION", "CHAMPION_SKIN", "CHERRY_BOON", "COMPANION", "CURRENCY", "EMOTE", "EVENT_PASS", "FANPASS", "GIFT", "HEXTECH_CRAFTING", "MODE_PROGRESSION_REWARD", "MYSTERY", "NEXUS_FINISHER", "OPAL_ACHIEVEMENT", "PREMIUM_CLUB_MEMBERSHIP", "PROGRESSION", "PROVIEW_PASS", "PVE_RELIC", "PVE_SUMMONER_PACKAGE", "PVE_UPGRADE", "QUEUE_ENTRY", "REGALIA_BANNER", "REGALIA_BORDER", "REGALIA_CREST", "RP", "RUNE", "SKIN_AUGMENT", "SKIN_BORDER", "SKIN_UPGRADE_GEAR", "SKIN_UPGRADE_HOME_GUARD", "SKIN_UPGRADE_RECALL", "SKIN_UPGRADE_SPAWN", "SPELL_BOOK_PAGE", "STATSTONE", "STRAWBERRY_BOON", "STRAWBERRY_LOADOUT_ITEM", "STRAWBERRY_MAP", "SUMMONER_CUSTOMIZATION", "SUMMONER_ICON", "TEAMPASS", "TEAM_SKIN_PURCHASE", "TFT_DAMAGE_SKIN", "TFT_EVENT_PVE_BUDDY", "TFT_EVENT_PVE_DIFFICULTY", "TFT_EVENT_SKILLS", "TFT_MAP_SKIN", "TFT_PLAYBOOK", "TFT_ZOOM_SKIN", "TOURNAMENT_FLAG", "TOURNAMENT_FRAME", "TOURNAMENT_LOGO", "TOURNAMENT_TROPHY", "TRANSFER", "WARD_SKIN"]
    #道具目录（Catalog）
    catalogDict = {} #该变量并未投入使用，只是用于观察时分类（This variable isn't put to use. It's only intended for classifcation during inspection）
    catalogList = []
    for inventoryType in inventoryTypes:
        catalogDict[inventoryType] = await (await connection.request("GET", "/lol-catalog/v1/items/" + inventoryType)).json()
        catalogDict[inventoryType] = sorted(catalogDict[inventoryType], key = lambda x: x["itemId"])
        catalogList += catalogDict[inventoryType]
    #商店（Store）
    store = await (await connection.request("GET", "/lol-store/v1/catalog")).json()
    storeDict = {} #该变量并未投入使用，只是用于观察时分类（This variable isn't put to use. It's only intended for classifcation during inspection）
    for inventoryType in sorted(set(map(lambda x: x["inventoryType"], store))):
        storeDict[inventoryType] = []
    for item in store:
        storeDict[item["inventoryType"]].append(item)
    for inventoryType in storeDict:
        storeDict[inventoryType] = sorted(storeDict[inventoryType], key = lambda x: x["itemId"]) #将商店中道具按照道具类型和道具序号升序排列（Sort the items by inventoryType and itemId）
    #藏品（Collection）
    #collection = await (await connection.request("GET", "/lol-inventory/v1/inventory", data = inventoryTypes)).json()
    collection = await (await connection.request("GET", "/lol-inventory/v1/inventory?inventoryTypes=%s" %(json.dumps(inventoryTypes).replace(" ", "")))).json()
    #collection = await (await connection.request("GET", '/lol-inventory/v1/inventory?inventoryTypes=["ACHIEVEMENT_BANNER_ACCENT","ACHIEVEMENT_TITLE","ANNOUNCER_PACK","ARAM_BOON","AUGMENT","AUGMENT_SLOT","BOOST","BUNDLES","CHAMPION","CHAMPION_SKIN","CHERRY_BOON","COMPANION","CURRENCY","EMOTE","EVENT_PASS","FANPASS","GIFT","HEXTECH_CRAFTING","MODE_PROGRESSION_REWARD","MYSTERY","NEXUS_FINISHER","OPAL_ACHIEVEMENT","PREMIUM_CLUB_MEMBERSHIP","PROGRESSION","PROVIEW_PASS","PVE_RELIC","PVE_SUMMONER_PACKAGE","PVE_UPGRADE","QUEUE_ENTRY","REGALIA_BANNER","REGALIA_BORDER","REGALIA_CREST","RP","RUNE","SKIN_AUGMENT","SKIN_BORDER","SKIN_UPGRADE_GEAR","SKIN_UPGRADE_HOME_GUARD","SKIN_UPGRADE_RECALL","SKIN_UPGRADE_SPAWN","SPELL_BOOK_PAGE","STATSTONE","STRAWBERRY_BOON","STRAWBERRY_LOADOUT_ITEM","STRAWBERRY_MAP","SUMMONER_CUSTOMIZATION","SUMMONER_ICON","TEAMPASS","TEAM_SKIN_PURCHASE","TFT_DAMAGE_SKIN","TFT_EVENT_HEALTH_BADGE","TFT_EVENT_PLAYER_TAG","TFT_EVENT_PVE_BUDDY","TFT_EVENT_PVE_DIFFICULTY","TFT_EVENT_RIBBON","TFT_EVENT_SKILLS","TFT_MAP_SKIN","TFT_PLAYBOOK","TFT_ZOOM_SKIN","TOURNAMENT_FLAG","TOURNAMENT_FRAME","TOURNAMENT_LOGO","TOURNAMENT_TROPHY","TRANSFER","WARD_SKIN"]')).json()

def sort_catalog_items(catalogList: list[dict[str, Any]], hashtable_dicts: dict[str, dict[Any, dict[str, str]]]) -> pandas.DataFrame:
    '''
    将道具目录数据整理成数据框。<br>Sort out catalog data into a dataframe.
    
    :param catalogList: 道具目录列表。<br>A list of catalog items.
    :type catalogList: list[dict[str, Any]]
    :param hashtable_dicts: 综合索引字典，通过create_hashtable函数得到。<br>Universal index dictionary obtained by `create_hashtable` function.
    :type hashtable_dicts: dict[str, dict[Any, dict[str, str]]]
    :return: 道具目录数据框。<br>A catalog dataframe.
    :rtype: pandas.DataFrame
    '''
    #定义商品数据结构（Define the store item data structure）
    catalog_header: dict[str, str] = {
        "active": "可用性",
        "description": "简介",
        "duration": "持续时间",
        "imagePath": "缩略图路径",
        "inactiveDate": "停止销售时间戳",
        "inventoryType": "道具类型",
        "itemId": "序号",
        "itemInstanceId": "识别码",
        "loadScreenPath": "加载界面路径",
        "maxQuantity": "最大持有数量",
        "metadata": "元数据",
        "name": "名称",
        "offerId": "交易代码",
        "owned": "已拥有",
        "ownershipType": "拥有状态",
        "purchaseDate": "购买时间戳",
        "questSkinInfo": "任务皮肤信息",
        "rarity": "稀有度",
        "releaseDate": "发布时间戳",
        "sale": "销售信息",
        "subInventoryType": "次级道具类型",
        "subTitle": "副标题",
        "taggedChampionsIds": "标签英雄序号",
        "tags": "搜索关键词",
        "tilePath": "方块图像路径",
        "inactiveTime": "停止销售时间",
        "purchaseTime": "购买时间",
        "releaseTime": "发布时间",
        "taggedChampionsNames": "标签英雄称号",
        "IP_cost": "原价（蓝色精萃）",
        "IP_costType": "支付类型（蓝色精萃）",
        "RP_cost": "原价（点券）",
        "RP_costType": "支付类型（点券）",
        "sale IP_cost": "售价（蓝色精萃）",
        "sale IP_discount": "销售折扣（蓝色精萃）",
        "sale IP_endDate": "停止售卖时间（蓝色精萃）",
        "sale IP_startDate": "开放售卖时间（蓝色精萃）",
        "sale RP_cost": "售价（点券）",
        "sale RP_discount": "销售折扣（点券）",
        "sale RP_endDate": "停止售卖时间（点券）",
        "sale RP_startDate": "开放售卖时间（点券）"
    }
    catalog_header_keys: list[str] = list(catalog_header.keys())
    catalog_data: dict[str, list[Any]] = {key: [] for key in catalog_header_keys}
    catalog_data_json: dict[str, list[Any]] = copy.deepcopy(catalog_data)
    #数据整理核心部分（Data organization - core part）
    print("商品信息整理进度（Catalog data organization process）：")
    for item_index in range(len(catalogList)):
        item: dict[str, Any] = catalogList[item_index]
        priceDict: dict[str, dict[str, int]] = {}
        for price in item["prices"]:
            priceDict[price["currency"]] = price
        sale_priceDict: dict[str, dict[str, int]] = {}
        for price in item["prices"]:
            if price["sale"] != None:
                sale_priceDict[price["currency"]] = price["sale"]
        for i in range(len(catalog_header)):
            key: str = catalog_header_keys[i]
            if i <= 28:
                if i == 1: #简介（`description`）
                    if item[key] != "" and (len(set(list(item[key]))) != 1 or item[key][0] != " "):
                        to_append = item[key]
                    elif item["inventoryType"] in hashtable_dicts:
                        if item["inventoryType"] in {"STRAWBERRY_BOON", "STRAWBERRY_LOADOUT_ITEM", "STRAWBERRY_MAP"} and item["itemInstanceId"] in hashtable_dicts[item["inventoryType"]]: #PVE模式相关的索引是识别码（PVE mode hashtable index is itemInstanceId）
                            to_append = hashtable_dicts[item["inventoryType"]][item["itemInstanceId"]]["description"] #为了简化这里的代码，所有索引字典的描述键必须是“description”（To simplify the code here, description keys of all hashtable dictionaries must be "description"）
                        elif item["itemId"] in hashtable_dicts[item["inventoryType"]]: #道具序号为111007的炫彩皮肤没有收录在CommunityDragon数据库中（The skin chroma with the itemId 111007 isn't archived in CommunityDragon database）
                            to_append = hashtable_dicts[item["inventoryType"]][item["itemId"]]["description"]
                        else:
                            to_append = ""
                    else:
                        to_append = ""
                elif i == 5: #道具类型（`inventoryType`）
                    to_append = inventoryType_dict[item[key]]
                elif i == 11: #名称（`name`）
                    if item[key] != "":
                        to_append = item[key]
                    elif item["inventoryType"] in hashtable_dicts:
                        if item["inventoryType"] in {"STRAWBERRY_BOON", "STRAWBERRY_LOADOUT_ITEM", "STRAWBERRY_MAP"} and item["itemInstanceId"] in hashtable_dicts[item["inventoryType"]]:
                            to_append = hashtable_dicts[item["inventoryType"]][item["itemInstanceId"]]["name"] #为了简化这里的代码，所有索引字典的名称键必须是“name”（To simplify the code here, name keys of all hashtable dictionaries must be "name"）
                        elif item["itemId"] in hashtable_dicts[item["inventoryType"]]:
                            to_append = hashtable_dicts[item["inventoryType"]][item["itemId"]]["name"]
                        else:
                            to_append = ""
                    else:
                        to_append = ""
                elif i == 14: #拥有状态（`ownershipType`）
                    to_append = ownershipTypes[item[key]]
                elif i == 20: #次级道具类型（`subInventoryType`）
                    to_append = subInventoryTypes[item[key]]
                elif i >= 25 and i <= 27: #时间戳相关键（Timestamp-related keys）
                    subkey = "inactiveDate" if i == 25 else "purchaseDate" if i == 26 else "releaseDate"
                    to_append = "" if item[subkey] == 0 else "∞" if item[subkey] == 18446744073709551615 else getISOTime(item[subkey])
                elif i == 28: #标签英雄称号（`taggedChampionsNames`）
                    taggedChampionIds: list[int] = item["taggedChampionsIds"]
                    if taggedChampionIds == [0]:
                        to_append = ""
                    else:
                        taggedChampionNames: list[str] = list(map(lambda x: LoLChampions[x]["name"] if x in LoLChampions else x, taggedChampionIds))
                        to_append = json.dumps(taggedChampionNames, indent = 4, ensure_ascii = False)
                else:
                    to_append = item[key]
            elif i <= 32:
                currency, subkey = key.split("_")
                if currency in priceDict and subkey in priceDict[currency]:
                    to_append = priceDict[currency][subkey]
                else:
                    to_append = ""
            else:
                currency, subkey = key.split(" ")[1].split("_")
                if currency in sale_priceDict and subkey in sale_priceDict[currency]:
                    to_append = sale_priceDict[currency][subkey]
                else:
                    to_append = ""
            catalog_data[key].append(to_append)
            catalog_data_json[key].append(pyobj2json(to_append))
        print("[%d/%d](%s, %d)" %(item_index + 1, len(catalogList), item["inventoryType"], item["itemId"]), end = "\r")
    #数据框列序整理（Dataframe column ordering）
    catalog_statistics_output_order: list[int] = [11, 21, 1, 6, 0, 9, 5, 20, 17, 10, 7, 27, 25, 2, 29, 30, 31, 32, 19, 33, 34, 36, 35, 37, 38, 40, 39, 13, 14, 26, 16, 12, 23, 28, 3, 8, 24]
    catalog_data_organized: dict[str, list[Any]] = {catalog_header_keys[i]: catalog_data_json[catalog_header_keys[i]] for i in catalog_statistics_output_order}
    catalog_df: pandas.DataFrame = pandas.DataFrame(data = catalog_data_organized)
    optimize_bool_display(catalog_df)
    catalog_df = pandas.concat([pandas.DataFrame([catalog_header])[catalog_df.columns], catalog_df], ignore_index = True)
    return catalog_df

def sort_store_items(store: list[dict[str, Any]], locale: str, collection_hashtable: dict[tuple[str, int], str], hashtable_dicts: dict[str, dict[Any, dict[str, str]]]) -> pandas.DataFrame:
    '''
    将商店数据整理成数据框。<br>Sort out store data into a dataframe.
    
    :param store: 商品列表。<br>A list of store items.
    :type store: list[dict[str, Any]]
    :param locale: 语言文化代码。主要用于确定使用那个语言的本地化名称和描述。<br>Language code. Mainly used to determine which language of name or description to use.
    :type locale: str
    :param collection_hashtable: 藏品索引字典。键是由道具类型和道具序号组成的元组，值是道具名称。<br>Collection index dictionary, whose keys are tuples composed of an inventory type and an itemId and values are item names.
    :type collection_hashtable: dict[tuple[str, int], str]
    :param hashtable_dicts: 综合索引字典，通过create_hashtable函数得到。<br>Universal index dictionary obtained by `create_hashtable` function.
    :type hashtable_dicts: dict[str, dict[Any, dict[str, str]]]
    :return: 商店数据框。<br>Store dataframe.
    :rtype: pandas.DataFrame
    '''
    #定义商店道具数据结构（Define the store item data structure）
    store_header: dict[str, str] = {
        "active": "可用性",
        "bundled": "附赠信息",
        "duration": "持续时间",
        "iconUrl": "图标链接",
        "inactiveDate": "禁用日期",
        "inventoryType": "道具类型",
        "itemId": "序号",
        "itemInstanceId": "识别码",
        "itemRequirements": "购买要求",
        "maxQuantity": "最大购买数量",
        "metadata": "元数据",
        "offerId": "交易代码",
        "releaseDate": "发布日期",
        "subInventoryType": "次级道具类型",
        "tags": "关键词",
        "name": "名称",
        "description": "简介",
        "IP_cost": "原价（蓝色精萃）",
        "IP_discount": "折扣（蓝色精萃）",
        "RP_cost": "原价（点券）",
        "RP_discount": "折扣（点券）",
        "sale endDate": "停止售卖时间",
        "sale startDate": "开放售卖时间",
        "sale IP_cost": "售价（蓝色精萃）",
        "sale IP_discount": "销售折扣（蓝色精萃）",
        "sale RP_cost": "售价（点券）",
        "sale RP_discount": "销售折扣（点券）"
    }
    store_header_keys: list[str] = list(store_header.keys())
    store_data: dict[str, list[Any]] = {key: [] for key in store_header_keys}
    store_data_json: dict[str, list[Any]] = copy.deepcopy(store_data)
    #数据整理核心部分（Data organization - core part）
    print("商店信息整理进度（Store data organization process）：")
    for item_index in range(len(store)):
        item: dict[str, Any] = store[item_index]
        priceDict: dict[str, dict[str, int]] = {} #应用于“i <= 20”的场景（Applies when "i <= 20"）
        for price in item["prices"]:
            priceDict[price["currency"]] = price
        sale_priceDict: dict[str, dict[str, int]] = {} #应用与“i >= 22”的场景（Applies when "i >= 22"）
        if item["sale"] != None:
            for price in item["sale"]["prices"]:
                sale_priceDict[price["currency"]] = price
        for i in range(len(store_header)):
            key: str = store_header_keys[i]
            if i <= 14:
                if i == 5: #道具类型（`inventoryType`）
                    to_append = inventoryType_dict[item[key]]
                elif i == 8: #购买要求（`itemRequirements`）
                    itemRequirements: list[str] = []
                    if item[key] != None:
                        for requirement in item[key]:
                            requirement["name"] = collection_hashtable.get((requirement["inventoryType"], requirement["itemId"]), "")
                            itemRequirements.append(requirement)
                    to_append = itemRequirements
                    if to_append == []:
                        to_append = ""
                elif i == 13: #次级道具类型（`subInventoryType`）
                    to_append = subInventoryTypes[item[key]]
                else:
                    to_append = item[key]
            elif i <= 16:
                value: str = ""
                if item["localizations"] != None and locale in item["localizations"] and key in item["localizations"][locale]:
                    value = item["localizations"][locale][key]
                if value == "": #当商店中没有给出一件道具的名称和描述时，从索引字典中获取（When the store doesn't provide an item's name and description, get them from the hashtable dictionaries）
                    if item["inventoryType"] in hashtable_dicts:
                        if item["inventoryType"] in {"STRAWBERRY_BOON", "STRAWBERRY_LOADOUT_ITEM", "STRAWBERRY_MAP"} and item["itemInstanceId"] in hashtable_dicts[item["inventoryType"]]:
                            value = hashtable_dicts[item["inventoryType"]][item["itemInstanceId"]][key]
                        elif item["itemId"] in hashtable_dicts[item["inventoryType"]]:
                            value = hashtable_dicts[item["inventoryType"]][item["itemId"]][key]
                to_append = value
            elif i <= 20:
                currency, subkey = key.split("_")
                if currency in priceDict and subkey in priceDict[currency]:
                    to_append = priceDict[currency][subkey]
                else:
                    to_append = ""
            elif i <= 22:
                to_append = item["sale"][key.split()[1]] if item["sale"] != None else ""
            else:
                currency, subkey = key.split(" ")[1].split("_")
                if currency in sale_priceDict and subkey in sale_priceDict[currency]:
                    to_append = sale_priceDict[currency][subkey]
                else:
                    to_append = ""
            store_data[key].append(to_append)
            store_data_json[key].append(pyobj2json(to_append))
        print("[%d/%d](%s, %d)" %(item_index + 1, len(store), item["inventoryType"], item["itemId"]), end = "\r")
    #数据框列序整理（Dataframe column ordering）
    store_statistics_output_order: list[int] = [15, 16, 6, 0, 5, 13, 10, 1, 7, 12, 4, 2, 17, 18, 19, 20, 9, 8, 11, 22, 21, 23, 24, 25, 26, 14, 3]
    store_data_organized: dict[str, list[Any]] = {store_header_keys[i]: store_data_json[store_header_keys[i]] for i in store_statistics_output_order}
    store_df: pandas.DataFrame = pandas.DataFrame(data = store_data_organized)
    optimize_bool_display(store_df)
    store_df = pandas.concat([pandas.DataFrame([store_header])[store_df.columns], store_df], ignore_index = True)
    return store_df

def sort_collection_items(collection: list[dict[str, Any]], collection_hashtable: dict[tuple[str, int], str], hashtable_dicts: dict[str, dict[Any, dict[str, str]]]) -> pandas.DataFrame:
    '''
    将藏品数据整理成数据框。<br>Sort out collection data into a dataframe.
    
    :param collection: 藏品列表。<br>A list of collection items.
    :type collection: list[dict[str, Any]]
    :param collection_hashtable: 藏品索引字典。键是由道具类型和道具序号组成的元组，值是道具名称。<br>Collection index dictionary, whose keys are tuples composed of an inventory type and an itemId and values are item names.
    :type collection_hashtable: dict[tuple[str, int], str]
    :param hashtable_dicts: 综合索引字典，通过create_hashtable函数得到。<br>Universal index dictionary obtained by `create_hashtable` function.
    :type hashtable_dicts: dict[str, dict[Any, dict[str, str]]]
    :return: 藏品数据框。<br>Collection dataframe.
    :rtype: pandas.DataFrame
    '''
    #定义藏品数据结构（Define the collection item data structure）
    collection_header: dict[str, str] = {"expirationDate": "租赁到期时间", "f2p": "免费使用", "inventoryType": "道具类型", "itemId": "序号", "loyalty": "奖励计划", "loyaltySources": "奖励计划来源", "owned": "已拥有", "ownershipType": "拥有权", "purchaseDate": "购买时间", "quantity": "数量", "rental": "租借中", "usedInGameDate": "上次使用时间", "uuid": "唯一识别码", "wins": "使用该道具可获得增益的胜场数", "isVintage": "典藏皮肤", "name": "名称"}
    collection_header_keys: list[str] = list(collection_header.keys())
    collection_data: dict[str, list[Any]] = {key: [] for key in collection_header_keys}
    collection_data_json: dict[str, list[Any]] = copy.deepcopy(collection_data)
    #数据整理核心部分（Data organization - core part）
    print("藏品信息整理进度（Collection data organization process）：")
    for item_index in range(len(collection)):
        item: dict[str, Any] = collection[item_index]
        for i in range(len(collection_header)):
            key: str = collection_header_keys[i]
            if i in {0, 8, 11}: #时间字符串相关键（Time string-related keys）
                if item[key] == "":
                    to_append = ""
                elif "-" in item[key] and ":" in item[key]:
                    to_append = "%s-%s-%s %s-%s-%s" %(item[key][:4], item[key][5:7], item[key][8:10], item[key][11:13], item[key][14:16], item[key][17:19])
                else:
                    to_append = "%s-%s-%s %s-%s-%s" %(item[key][:4], item[key][4:6], item[key][6:8], item[key][9:11], item[key][11:13], item[key][13:15])
            elif i == 2: #道具类型（`inventoryType`）
                to_append = inventoryType_dict[item["inventoryType"]]
            elif i == 5: #奖励计划来源（`loyaltySources`）
                to_append = item["loyaltySources"]
                if to_append == []:
                    to_append = ""
            elif i == 7: #拥有权（`ownershipType`）
                to_append = ownershipTypes[item["ownershipType"]]
            elif i == 14: #典藏皮肤（带边框）（`isVintage`）
                to_append = item["payload"] and "isVintage" in item["payload"] and item["payload"]["isVintage"] #没有“是否典藏”选项的默认不是典藏（An item without the "isVintage" key can't be vintage）
            elif i == 15: #名称（`name`）
                if (item["inventoryType"], item["itemId"]) in collection_hashtable:
                    name: str = collection_hashtable[(item["inventoryType"], item["itemId"])]
                else:
                    name = ""
                if name == "":
                    if item["inventoryType"] in hashtable_dicts: #商品中可能不包含藏品（A collection item may not be contained in the collection）
                        if item["inventoryType"] in {"STRAWBERRY_BOON", "STRAWBERRY_LOADOUT_ITEM", "STRAWBERRY_MAP"} and item["uuid"] in hashtable_dicts[item["inventoryType"]]:
                            name = hashtable_dicts[item["inventoryType"]][item["uuid"]]["name"]
                        elif item["itemId"] in hashtable_dicts[item["inventoryType"]]:
                            name = hashtable_dicts[item["inventoryType"]][item["itemId"]]["name"]
                        else:
                            name = ""
                    else:
                        name = ""
                to_append = name
            else:
                to_append = item[key]
            collection_data[key].append(to_append)
            collection_data_json[key].append(pyobj2json(to_append))
        print("[%d/%d](%s, %d)" %(item_index + 1, len(collection), item["inventoryType"], item["itemId"]), end = "\r")
    #数据框列序整理（Dataframe column ordering）
    collection_statistics_output_order: list[int] = [15, 9, 3, 2, 12, 6, 10, 1, 4, 5, 7, 8, 0, 14, 13, 11]
    collection_data_organized: dict[str, list[Any]] = {collection_header_keys[i]: collection_data_json[collection_header_keys[i]] for i in collection_statistics_output_order}
    collection_df: pandas.DataFrame = pandas.DataFrame(data = collection_data_organized)
    optimize_bool_display(collection_df)
    collection_df = pandas.concat([pandas.DataFrame([collection_header])[collection_df.columns], collection_df], ignore_index = True)
    return collection_df

async def fetch_store(connection: Connection) -> None:
    #获取大区信息，用于设置工作簿保存位置和工作表名称和获取相应的CommunityDragon数据资源（Get server information to set up workbook saving directory and sheet name and fetch the adaptive CommunityDragon data resources）
    current_info: dict[str, Any] = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    displayName: str = get_info_name(current_info)
    current_party: dict[str, Any] = await (await connection.request("GET", "/lol-lobby/v1/parties/player")).json()
    platformId: str = current_party["platformId"]
    riot_client_info: dict[str, Any] = await (await connection.request("GET", "/riotclient/command-line-args")).json()
    region_locale: dict[str, str] = await (await connection.request("GET", "/riotclient/region-locale")).json()
    region: str = region_locale["region"]
    locale: str = region_locale["locale"]
    platform_folder: str = set_platform_folder(region, platformId)
    folder: str = set_summonerInfo_folder(region, platformId, current_info)
    common_data: dict[str, Any] = await (await connection.request("GET", "/telemetry/v1/common-data")).json()
    version: str = common_data["common.application_version"]
    #下面准备数据资源（The following code prepare the data resource）
    print("正在加载数据资源……\nLoading data resources ...")
    await prepare_data_resources(connection)
    #保存原始数据（Save the raw data）
    json1name: str = f"Catalog - {displayName}.json"
    os.makedirs(folder, exist_ok = True)
    with open(os.path.join(folder, json1name), "w", encoding = "utf-8") as fp: #从`lol-catalog`接口获取的商品信息含有个人拥有信息，因此放到召唤师信息文件夹里（Item information obtained from `lol-catalog` API contains personal information like ownership, so it's saved into the summoner information folder）
        json.dump(catalogDict, fp, indent = 4, ensure_ascii = False)
        # json.dump(catalogList, fp, indent = 4, ensure_ascii = False)
    print('道具目录信息已保存为“%s”。\nCatalog information is saved as "%s".\n' %(os.path.join(folder, json1name), os.path.join(folder, json1name)))
    json2name: str = "Store.json"
    os.makedirs(platform_folder, exist_ok = True)
    with open(os.path.join(platform_folder, json2name), "w", encoding = "utf-8") as fp: #从`lol-store`接口获取的商品信息是服务器特定的，因此放到服务器文件夹里（Item information obtained from `lol-store` API is server-specific, so it's saved into the platform folder）
        json.dump(storeDict, fp, indent = 4, ensure_ascii = False)
    print('商店信息已保存为“%s”。\nStore data are saved as "%s".\n' %(os.path.join(folder, json2name), os.path.join(folder, json2name)))
    json3name: str = f"Collection - {displayName}.json"
    os.makedirs(folder, exist_ok = True)
    with open(os.path.join(folder, json3name), "w", encoding = "utf-8") as fp:
        json.dump(collection, fp, indent = 4, ensure_ascii = False)
    print('藏品信息已保存为“%s”。\nCollection information is saved as "%s".\n' %(os.path.join(folder, json3name), os.path.join(folder, json3name)))
    print("正在创建索引……\nCreating index ...\n")
    collection_hashtable: dict[tuple[str, int], str] = {(item["inventoryType"], item["itemId"]): item["name"] for item in catalogList} | {(item["inventoryType"], item["itemId"]): item["localizations"][locale]["name"] for item in store if item["localizations"] != None} #原本的藏品信息中没有记录名称，所以需要借用商品信息中的名称。之所以不考虑使用识别码作为键，是因为在从`lol-store`接口获取的商品信息中，存在识别码重复的两件商品，而道具类型和道具序号的组合应当能够唯一确定一件商品。另外，从`lol-catalog`和`lol-store`接口获取的商品信息可以互相补充（The original collection information doesn't contain the names, so they're cited from the catalog information. The reason why `itemInstanceId` isn't taken as the key is that there're two items with the same `itemInstanceId` in the items obtaned from `lol-store` API. However, the combination of `inventoryType` and `itemId` should uniquely correspond to an item. Besides, item information obtained from `lol-catalog` API and that from `lol-store` API can supplement each other）
    hashtable_dicts = await build_inventory_map(connection)
    #整理数据（Oranize data）
    print("开始整理数据。\nBegin to organize data ...")
    catalog_df: pandas.DataFrame = sort_catalog_items(catalogList, hashtable_dicts)
    store_df: pandas.DataFrame = sort_store_items(store, locale, collection_hashtable, hashtable_dicts)
    collection_df: pandas.DataFrame = sort_collection_items(collection, collection_hashtable, hashtable_dicts)
    #保存文件（Save file）
    print("开始导出到工作簿。\nBegin to export to the workbook.\n")
    excel_name: str = f"Store - {platformId}.xlsx"
    wb1Path: str = os.path.join(platform_folder, excel_name)
    os.makedirs(platform_folder, exist_ok = True)
    if not os.path.exists(wb1Path):
        wb1CreateFlag: bool = create_workbook_win32(os.path.abspath(wb1Path))
    workbook1_exist: bool = os.path.exists(wb1Path)
    while True:
        try:
            if workbook1_exist:
                with (pandas.ExcelWriter(path = wb1Path, mode = "a", if_sheet_exists = "replace") if workbook1_exist else pandas.ExcelWriter(path = wb1Path)) as writer:
                    currentTime: str = time.strftime("%Y-%m-%d", time.localtime(time.time()))
                    addDefaultStyle(store_df).to_excel(excel_writer = writer, sheet_name = f"Store - {currentTime}_{platformId}_{locale}")
                    worksheet: Worksheet = writer.sheets[f"Store - {currentTime}_{platformId}_{locale}"]
                    worksheet.cell(row = 1, column = 1, value = version)
        except PermissionError:
            print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
            input()
        else:
            print('商店数据已保存为“%s”！\nStore data are saved as "%s"!' %(wb1Path, wb1Path))
            break
    excel_name: str = "Catalog and Collections - %s.xlsx" %displayName
    wb2Path: str = os.path.join(folder, excel_name)
    excel_name_sorted: str = "Catalog and Collections - %s (sorted).xlsx" %displayName
    os.makedirs(folder, exist_ok = True)
    if not os.path.exists(wb2Path):
        wb2CreateFlag: bool = create_workbook_win32(os.path.abspath(wb2Path))
    workbook2_exist: bool = os.path.exists(wb2Path)
    while True:
        try:
            with (pandas.ExcelWriter(path = wb2Path, mode = "a", if_sheet_exists = "replace") if workbook2_exist else pandas.ExcelWriter(path = wb2Path)) as writer:
                currentTime: str = time.strftime("%Y-%m-%d", time.localtime(time.time()))
                addDefaultStyle(catalog_df).to_excel(excel_writer = writer, sheet_name = f"Catalog - {currentTime}_{platformId}_{locale}")
                addDefaultStyle(collection_df).to_excel(excel_writer = writer, sheet_name = f"Collections - {currentTime}_{platformId}")
                for sheet_name in [f"Catalog - {currentTime}_{platformId}_{locale}", f"Collections - {currentTime}_{platformId}"]:
                    worksheet: Worksheet = writer.sheets[sheet_name]
                    worksheet.cell(row = 1, column = 1, value = version)
        except PermissionError:
            print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
            input()
        else:
            print('道具目录和藏品信息已保存为“%s”！\nCatalog and collections information is saved as "%s"!' %(wb2Path, wb2Path))
            break
    #工作表排序（Worksheet ordering）
    if workbook2_exist:
        print("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致工作表顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of worksheet order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
        sort: bool = bool(input())
        if sort:
            print("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
            while True:
                try:
                    wb: Workbook = load_workbook(wb2Path)
                except FileNotFoundError:
                    print('商品藏品信息工作簿读取失败！请确保“%s”文件夹内含有名为“%s”的工作簿。如果需要退出程序，请输入“0”。\nERROR reading the Catalog and Collections workbook! Please make sure the workbook "%s" is in the folder "%s". If you want to exit the program, please submit "0".' %(folder, excel_name, excel_name, folder))
                    store_reload: str = input()
                    if store_reload == "0":
                        break
                else:
                    sheetnames: list[str] = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                    print("请选择排序方式：\nPlease select an ordering pattern:\n☆1\t时间优先（Time in priority）\n2\t类别优先（Type in priority）")
                    op: str = input()
                    print("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                    date_re: re.Pattern[str] = re.compile(r"\d{4}-\d{2}-\d{2}") #设置正则表达式识别日期（Define a regular expression to identify a date pattern）
                    if op == "" or op[0] != "2": #按照时间优先的原则对工作表进行排序，时间相同则商品工作表在前，藏品工作表在后（Sort the sheets by time in priority. If the times are the same, then the store sheet is arranged in front of the collection sheet）
                        sheetname_date_list: list[str] = list(map(lambda x: date_re.search(x).group(), sheetnames)) #从工作表名称提取日期信息形成列表（Extract the dates from the sheetnames to form a list）
                        sheetname_type_list: list[str] = list(map(lambda x: x.split()[0], sheetnames)) #从工作表名称提取数据类型信息形成列表（Extract the data types from the sheetnames to form a list）
                        sheetname_platform_list: list[str] = list(map(lambda x: x.split("_")[1], sheetnames)) #从工作表名称提取大区信息形成列表（Extract the platformId from the sheetnames to form a list）
                        sheetname_tmpDf: pandas.DataFrame = pandas.DataFrame(data = [sheetnames, sheetname_date_list, sheetname_type_list, sheetname_platform_list]).transpose() #创建一个四列数据框，各列分别是完整工作表名、日期信息、数据类型信息和大区信息（Create a 4-column dataframe whose columns are the complete sheetname, date, data type and platformId）
                        sheetnames_sorted: list[str] = sheetname_tmpDf.sort_values(by = [1, 2, 3], ascending = [True, False, True]).iloc[:, 0].tolist() #将工作表名按照第一关键字——日期信息正序排列，第二关键字——数据类型信息倒序排列（先商品后藏品），第三关键字——大区信息正序排列（Order the sheetnames according to the ascending order of the first keyword - date, the descending order of the second keyword - data type and the ascending order of the third keyword - platformId）
                    else:
                        sheets_Store: list[str] = [sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Store") or sheet_iter.startswith("Catalog")] #提取商品类型的工作表名称（Extract the names of the sheets containing Store data）
                        sheets_Collections: list[str] = [sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Collections")] #提取藏品类型的工作表名称（Extract the names of the sheets containing Collection data）
                        sheets_Store = sorted(sheets_Store, key = lambda x: date_re.search(x).group()) #按照日期正序排列商品类型的工作表名称（Order the Store sheetnames according to the ascending order of dates）
                        sheets_Collections = sorted(sheets_Collections, key = lambda x: date_re.search(x).group()) #按照日期正序排列藏品类型的工作表名称（Order the Collection sheetnames according to the ascending order of dates）
                        sheetnames_sorted: list[str] = sheets_Store + sheets_Collections #合并列表得到先按类别排列、再按日期排列的工作表名称（Combine the lists to get the sheetname list ordered firstly by data type and secondly by date）
                    #下面排列所有工作表（The following code arrange all sheets）
                    print("正在排序……\nOrdering ...")
                    sort_worksheet(wb, sheetnames_sorted)
                    print('正在保存中……\nSaving the ordered workbook ...')
                    wb.save(os.path.join(folder, excel_name_sorted))
                    print('排序完成！排好序的工作簿已保存为“%s”。请按任意键退出。\nOrdering finished! The ordered workbook is saved as "%s". Press any key to exit ...\n' %(excel_name_sorted, excel_name_sorted))
                    wb.close()
                    input()
                    break

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection: Connection) -> None:
    await print_summoner_info(connection)
    await save_platform_info(connection)
    await fetch_store(connection)

@connector.close
async def disconnect(connection: Connection) -> None:
    print("已从英雄联盟客户端断开连接。\nDisconnected from the League Client.")

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------
connector.start()
