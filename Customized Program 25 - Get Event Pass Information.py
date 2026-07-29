from lcu_driver import Connector
from lcu_driver.connection import Connection
import json, os, pandas, re, time
from openpyxl import load_workbook, Workbook
from src.utils.summoner import print_summoner_info, get_info_name
from src.core.config.servers import set_summonerInfo_folder, save_platform_info
from src.core.config.headers import event_info_header, event_narrative_header, event_pass_chapter_header, event_pass_bundle_header, event_reward_item_header, token_bundle_header, tokenShop_categoryOffer_header
from src.core.config.localization import inventoryType_dict, subInventoryTypes, eventPassTypes, rewardTag_dict, lolEventHubRewardTrackItemStates, cardSizes, celebrationTypes, lolEventHubRewardTrackItemHeaderTypes, lolEventHubOfferCategories, lolEventHubOfferStates
from src.utils.format import getISOTime, addDefaultStyle
from src.utils.excel_workbook import create_workbook_win32, sort_worksheet
from typing import Any
from openpyxl.worksheet.worksheet import Worksheet

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2026/07/30
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

gameQueues: dict[int, dict[str, Any]] = {}
championSkins: dict[int, dict[str, Any]] = {}
connector: Connector = Connector()

#-----------------------------------------------------------------------------
# 整理通行证信息（Organize pass information）
#-----------------------------------------------------------------------------
async def prepare_data_resources(connection: Connection) -> None:
    '''
    准备全局数据资源。<br>Prepare global data resources.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    '''
    global gameQueues, championSkins
    gameQueues_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-queues/v1/queues")).json()
    gameQueues = {queue["id"]: queue for queue in gameQueues_source}
    championSkins_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/skins.json")).json()
    championSkins = {}
    for skin in championSkins_source.values():
        championSkins[skin["id"]] = skin
        if "chromas" in skin:
            for chroma in skin["chromas"]:
                championSkins[chroma["id"]] = chroma
        if "questSkinInfo" in skin:
            for tier in skin["questSkinInfo"]["tiers"]:
                if not tier["id"] in championSkins: #圣堂皮肤和终极皮肤中的系列与主皮肤存在重复的序号（There're redundant ids between the tier and the parent ultimate skin）
                    championSkins[tier["id"]] = tier

async def organize_pass_information(connection: Connection) -> None:
    '''
    调用多个事件通行证接口，将返回内容整理成若干个数据框，导出到Excel工作簿中。<br>Call event pass related endpoints, organize the returned content into several dataframes and export them into an Excel workbook.
    
    调用以下接口：<br>The following endpoints are called:
    - GET /lol-event-hub/v1/events
    - GET /lol-event-hub/v1/events/{eventId}/chapters
    - GET /lol-event-hub/v1/events/{eventId}/event-details-data
    - GET /lol-event-hub/v1/events/{eventId}/info #目前并未投入使用（Unused for now）
    - GET /lol-event-hub/v1/events/{eventId}/is-grace-period
    - GET /lol-event-hub/v1/events/{eventId}/narrative
    - GET /lol-event-hub/v1/events/{eventId}/objectives-banner
    - GET /lol-event-hub/v1/events/{eventId}/pass-background-data
    - GET /lol-event-hub/v1/events/{eventId}/pass-bundles
    - GET /lol-event-hub/v1/events/{eventId}/progress-info-data
    - GET /lol-event-hub/v1/events/{eventId}/progression-purchase-data
    - GET /lol-event-hub/v1/events/{eventId}/reward-track/bonus-items
    - GET /lol-event-hub/v1/events/{eventId}/reward-track/bonus-progress #目前并未投入使用（Unused for now）
    - GET /lol-event-hub/v1/events/{eventId}/reward-track/items
    - GET /lol-event-hub/v1/events/{eventId}/reward-track/progress #目前并未投入使用（Unused for now）
    - GET /lol-event-hub/v1/events/{eventId}/reward-track/unclaimed-rewards
    - GET /lol-event-hub/v1/events/{eventId}/reward-track/xp
    - GET /lol-event-hub/v1/events/{eventId}/token-shop
    - GET /lol-event-hub/v1/events/{eventId}/token-shop/categories-offer
    - GET /lol-event-hub/v1/events/{eventId}/token-shop/token-balance #目前并未投入使用（Unused for now）
    
    生成以下工作表：<br>The following worksheets are generated:
    - 基础信息（Basic info）
    - 叙事章节（Narratives）
    - 章节（Chapters）
    - 通行证道具包（Pass bundles）
    - 奖励（Rewards）
    - 代币道具包（Token bundles）
    - 交易类别（Offer categories）
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    '''
    current_party: dict[str, Any] = await (await connection.request("GET", "/lol-lobby/v1/parties/player")).json()
    platformId: str = current_party["platformId"]
    riot_client_info: list[str] = await (await connection.request("GET", "/riotclient/command-line-args")).json()
    client_info: dict[str, str] = {}
    for i in range(len(riot_client_info)):
        try:
            client_info[riot_client_info[i].split("=")[0]] = riot_client_info[i].split("=")[1]
        except IndexError:
            pass
    region: str = client_info["--region"]
    common_data: dict[str, Any] = await (await connection.request("GET", "/telemetry/v1/common-data")).json()
    version: str = common_data["common.application_version"]
    #设置输出路径（Set the output directory）
    current_info: dict[str, Any] = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    displayName: str = get_info_name(current_info)
    folder: str = set_summonerInfo_folder(region, platformId, current_info)
    #初始化数据结构（Initialize data structures）
    ##事件信息（Event information）
    event_info_header_keys: list[str] = list(event_info_header.keys())
    event_info_data: dict[str, list[Any]] = {key: [] for key in event_info_header_keys}
    ##叙事章节（Narrative chapter）
    event_narrative_header_keys: list[str] = list(event_narrative_header.keys())
    event_narrative_data: dict[str, list[Any]] = {key: [] for key in event_narrative_header_keys}
    ##通行证章节（Pass chapter）
    event_pass_chapter_header_keys: list[str] = list(event_pass_chapter_header.keys())
    event_pass_chapter_data: dict[str, list[Any]] = {key: [] for key in event_pass_chapter_header_keys}
    ##通行证购买道具包（Pass purchase bundle）
    event_pass_bundle_header_keys: list[str] = list(event_pass_bundle_header.keys())
    event_pass_bundle_data: dict[str, list[Any]] = {key: [] for key in event_pass_bundle_header_keys}
    ##通行证奖励（Pass reward）
    event_reward_item_header_keys: list[str] = list(event_reward_item_header.keys())
    event_reward_item_data: dict[str, list[Any]] = {key: [] for key in event_reward_item_header_keys}
    ##代币购买道具包（Token purchase bundle）
    token_bundle_header_keys: list[str] = list(token_bundle_header.keys())
    token_bundle_data: dict[str, list[Any]] = {key: [] for key in token_bundle_header_keys}
    ##代币商城交易类别（Token shop offer category）
    tokenShop_categoryOffer_header_keys: list[str] = list(tokenShop_categoryOffer_header.keys())
    tokenShop_categoryOffer_data: dict[str, list[Any]] = {key: [] for key in tokenShop_categoryOffer_header_keys}
    #整理数据（Organize data）
    events: list[dict[str, Any]] = await (await connection.request("GET", "/lol-event-hub/v1/events")).json()
    for event in events:
        eventId: str = event["eventId"] #下面假设`eventInfo`，即`event`不可能是异常信息。如果它是异常信息，这一句会报错，从而不会执行任何操作。作了这个假设后，事件基础信息之外的数据框在向“事件代码”和“事件名称”字段中追加值时会方便很多（In the following context, we assume `eventInfo`, namely `event` can't be erroneous. Otherwise, this statement will throw an error, which prevents this program from performing any operation. When this assumption holds, it's convenient to append values into "eventInfo" and "eventName" fields of dataframes excluding the event basic info one）
        eventInfo: dict[str, Any] = event["eventInfo"] #`GET /lol-event-hub/v1/events/{eventId}/info`
        chapters: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/chapters")).json()
        event_details_data: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/event-details-data")).json()
        is_grace_period: bool | dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/is-grace-period")).json()
        if is_grace_period == {"errorCode": "RPC_ERROR", "httpStatus": 404, "implementationDetails": {}, "message": f"Event {eventId} does not have GracePeriod data"}:
            is_grace_period = False
        narratives: list[dict[str, Any]] | dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/narrative")).json()
        objectives_banner: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/objectives-banner")).json()
        pass_background_data: dict[str, str] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/pass-background-data")).json()
        pass_bundles: list[dict[str, Any]] | dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/pass-bundles")).json()
        progress_info_data: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/progress-info-data")).json()
        progression_purchase_data: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/progression-purchase-data")).json()
        reward_track_bonus_items: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/reward-track/bonus-items")).json()
        # reward_track_bonus_progress: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/reward-track/bonus-progress")).json()
        reward_track_items: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/reward-track/items")).json()
        # reward_track_progress: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/reward-track/progress")).json()
        reward_track_unclaimed_rewards: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/reward-track/unclaimed-rewards")).json()
        reward_track_xp: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/reward-track/xp")).json()
        token_shop: dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/token-shop")).json()
        token_shop_categories_offer: list[dict[str, Any]] | dict[str, Any] = await (await connection.request("GET", f"/lol-event-hub/v1/events/{eventId}/token-shop/categories-offer")).json()
        ##事件信息（Event information）
        for i in range(len(event_info_header_keys)):
            key: str = event_info_header_keys[i]
            if i <= 27:
                event_data: dict[str, Any] = eventInfo
            elif i == 28:
                event_data = chapters
            elif i <= 42:
                event_data = event_details_data
            elif i <= 83:
                event_data = objectives_banner
            elif i == 84:
                event_data = pass_background_data
            elif i == 85 or i == 86:
                event_data = progress_info_data
            elif i <= 91:
                event_data = progression_purchase_data
            elif i <= 95:
                event_data = reward_track_unclaimed_rewards
            elif i <= 100:
                event_data = reward_track_xp
            else:
                event_data = token_shop
            if "errorCode" in event_data:
                if i in {8, 9, 27, 85}:
                    to_append: Any = False
                else:
                    to_append = ""
            else:
                if i == 25: #游戏模式名称（`gameModeName`）
                    queueId: int = eventInfo["queueId"]
                    to_append = "" if queueId == 0 else gameQueues[queueId]["name"]
                elif i == 26: #上次未领取奖励时间（`dateOfLastUnclaimedReward`）
                    timeOfLastUnclaimedReward: int = eventInfo["timeOfLastUnclaimedReward"]
                    to_append = "" if timeOfLastUnclaimedReward == 0 or timeOfLastUnclaimedReward == -1 else getISOTime(timeOfLastUnclaimedReward)
                elif i == 27: #商城购买宽限期（`isGracePeriod`）
                    to_append = is_grace_period
                elif i == 42: #事件细节：焦点皮肤名称（`details spotlightSkinName`）
                    spotlightSkinId: int = event_details_data["spotlightSkinId"]
                    to_append = championSkins.get(spotlightSkinId, spotlightSkinId)
                elif i == 95: #通行证里程进度：未领取奖励追踪：上次未领取奖励时间（`rewardTrack_unclaimedRewards dateOfLastUnclaimedReward`）
                    timeOfLastUnclaimedReward: int = reward_track_unclaimed_rewards["timeOfLastUnclaimedReward"]
                    to_append = "" if timeOfLastUnclaimedReward == 0 or timeOfLastUnclaimedReward == -1 else getISOTime(timeOfLastUnclaimedReward)
                else:
                    subkey_list: list[str] = key.split()[(0 if i <= 27 else 1):] #事件信息接口中的键未设置抬头（No header is set for keys in data from `eventInfo`）
                    tmp_ptr: Any = event_data
                    for subkey in subkey_list:
                        tmp_ptr = tmp_ptr[subkey]
                    else:
                        if i == 7: #事件类型（`eventType`）
                            to_append: Any = eventPassTypes[eventInfo["eventType"]]
                        elif i == 87: #进度购买：道具类型（`progressionPurchaseData inventoryType`）
                            to_append = inventoryType_dict[tmp_ptr]
                        else:
                            to_append = tmp_ptr
            event_info_data[key].append(to_append)
        ##叙事章节（Narrative chapter）
        if not "errorCode" in narratives:
            for narrative_index in range(len(narratives)):
                narrative: dict[str, Any] = narratives[narrative_index]
                for i in range(len(event_narrative_header_keys)):
                    key: str = event_narrative_header_keys[i]
                    if i == 0 or i == 1:
                        to_append: Any = eventInfo[key]
                    elif i == 2: #记叙章节序号（`narrative_index`）
                        to_append = narrative_index + 1
                    elif i <= 6:
                        to_append = narrative[key]
                    else:
                        to_append = narrative["narrativeVideo"][key.split()[1]]
                event_narrative_data[key].append(to_append)
        ##通行证章节（Pass chapter）
        if not "errorCode" in chapters:
            for chapter_index in range(len(chapters["chapters"])):
                chapter: dict[str, Any] = chapters["chapters"][chapter_index]
                for i in range(len(event_pass_chapter_header_keys)):
                    key: str = event_pass_chapter_header_keys[i]
                    if i == 0 or i == 1:
                        to_append: Any = eventInfo[key]
                    else:
                        to_append = chapter[key]
                    event_pass_chapter_data[key].append(to_append)
        ##通行证购买道具包（Pass purchase bundle）
        if not "errorCode" in pass_bundles:
            if "errorCode" in progress_info_data:
                pass_bundle_itemId_map: dict[int, dict[str, Any]] = {entry["itemId"]: entry for entry in progress_info_data["eventPassBundlesCatalogEntry"]}
            else:
                pass_bundle_itemId_map = {}
            for bundle_index in range(len(pass_bundles)):
                bundle: dict[str, Any] = pass_bundles[bundle_index]
                bundledItems: list[dict[str, Any]] = bundle["bundledItems"]
                bundledItems_empty: bool = len(bundledItems) == 0 #记录道具包内是否没有商品。无论有无商品（Records whether no item is in the bundle）
                for bundledItem_index in range(max(1, len(bundledItems))):
                    bundledItem: dict[str, Any] = bundledItems[bundledItem_index]
                    for i in range(len(event_pass_bundle_header_keys)):
                        key: str = event_pass_bundle_header_keys[i]
                        if i == 0 or i == 1: #事件相关键（Event-related keys）
                            if bundle_index == 0 and bundledItem_index == 0: #事件信息只在同事件内只追加一次（Event data are appended once per event）
                                to_append: Any = eventInfo[key]
                            else:
                                to_append = ""
                        elif i <= 19: #道具包相关键（Bundle-related keys）
                            if bundledItem_index == 0: #道具包信息只在同道具包内追加一次（Bundle data are appended once per bundle）
                                if i == 2: #道具包序号（`bundle_index`）
                                    to_append = bundle_index + 1
                                elif i <= 7:
                                    if i == 7: #可购买（`isPurchasable`）
                                        to_append = "√" if bundle["isPurchasable"] else ""
                                    else:
                                        to_append = bundle[key]
                                elif i <= 16:
                                    if i == 10: #套装道具类型（`details inventoryType`）
                                        to_append = inventoryType_dict[bundle["details"]["inventoryType"]]
                                    elif i == 13: #套装已拥有（`details owned`）
                                        to_append = "√" if bundle["details"]["owned"] else ""
                                    elif i == 16: #套装子道具类型（`details subInventoryType`）
                                        to_append = subInventoryTypes[bundle["details"]["subInventoryType"]]
                                    else:
                                        to_append = bundle["details"][key.split()[1]]
                                else:
                                    itemId: int = bundle["details"]["itemId"]
                                    if itemId in pass_bundle_itemId_map:
                                        to_append = bundle["details"][key.split()[1]]
                                    else:
                                        to_append = ""
                            else:
                                to_append = ""
                        else: #套装内商品相关键（Bundled item-related keys）
                            if bundledItems_empty: #道具包内没有商品时，为了保存道具包原本的信息，至少也应当添加一次商品信息（When no item is in the bundle, to save the information of the bundle itself, the dataframe should at least append the item information once）
                                to_append = ""
                            else:
                                if i == 20: #套装内商品序号（`bundledItem_index`）
                                    to_append = bundledItem_index + 1
                                elif i == 23: #套装内商品：道具类型（`bundledItem inventoryType`）
                                    to_append = inventoryType_dict[bundledItem["inventoryType"]]
                                elif i == 26: #套装内商品：已拥有`bundledItem owned`）
                                    to_append = "√" if bundledItem["owned"] else ""
                                elif i == 29: #套装内商品：子道具类型（`bundledItem subInventoryType`）
                                    to_append = subInventoryTypes[bundledItem["subInventoryType"]]
                                else:
                                    to_append = bundledItem[key.split()[1]]
                        event_pass_bundle_data[key].append(to_append)
        ##通行证奖励（Pass reward）
        event_pass_reward_track_items: list[dict[str, Any]] = []
        if not "errorCode" in reward_track_items:
            event_pass_reward_track_items.extend(reward_track_items)
        if not "errorCode" in reward_track_bonus_items:
            event_pass_reward_track_items.extend(reward_track_bonus_items)
        for item_index in range(len(event_pass_reward_track_items)):
            rewardItem: dict[str, Any] = event_pass_reward_track_items[item_index]
            rewardOptions: list[dict[str, Any]] = rewardItem["rewardOptions"]
            rewardOption_empty: bool = len(rewardOptions) == 0
            for rewardOption_index in range(len(rewardOptions)):
                rewardOption: dict[str, Any] = rewardOptions[rewardOption_index]
                for i in range(len(event_reward_item_header_keys)):
                    key: str = event_reward_item_header_keys[i]
                    if i == 0 or i == 1: #事件相关键（Event-related keys）
                        if item_index == 0 and rewardOption_index == 0: #事件信息只在同事件内只追加一次（Event data are appended once per event）
                            to_append: Any = eventInfo[key]
                        else:
                            to_append = ""
                    elif i <= 6: #奖励商品相关键（Reward item-related keys）
                        if rewardOption_index == 0: #奖励商品信息只在同商品内只追加一次（Reward item data are appended once per item）
                            if i == 2: #装备序号（`item_index`）
                                to_append = item_index + 1
                            elif i == 4: #奖励标签（`rewardTags`）
                                rewardTags_trans: list[str] = list(map(lambda x: rewardTag_dict[x], rewardItem["rewardTags"]))
                                to_append = json.dumps(rewardTags_trans, ensure_ascii = False)
                            elif i == 5: #状态（`state`）
                                to_append = lolEventHubRewardTrackItemStates[rewardItem["state"]]
                            else:
                                to_append = rewardItem[key]
                        else:
                            to_append = ""
                    else: #奖励选项相关键（Reward option-related keys）
                        if rewardOption_empty:
                            to_append = ""
                        else:
                            if i == 7: #奖励序号（`rewardOption_index`）
                                to_append = rewardOption_index + 1
                            elif i == 8: #奖励：卡片大小（`rewardOption cardSize`）
                                to_append = cardSizes[rewardOption["cardSize"]]
                            elif i == 9: #奖励：庆祝类型（`rewardOption celebrationType`）
                                to_append = celebrationTypes[rewardOption["celebrationType"]]
                            elif i == 10: #奖励：标题类型（`rewardOption headerType`）
                                to_append = lolEventHubRewardTrackItemHeaderTypes[rewardOption["headerType"]]
                            elif i == 15: #奖励：道具类型（`rewardOption rewardInventoryTypes`）
                                rewardInventoryTypes_trans: list[str] = list(map(lambda x: inventoryType_dict[x], rewardOption["rewardInventoryTypes"]))
                                to_append = json.dumps(rewardInventoryTypes_trans, ensure_ascii = False)
                            elif i == 19: #奖励：已选择（`rewardOption selected`）
                                to_append = "√" if rewardOption["selected"] else ""
                            elif i == 21: #奖励：状态（`rewardOption state`）
                                to_append = lolEventHubRewardTrackItemStates[rewardOption["state"]]
                            else:
                                to_append = rewardOption[key.split()[1]]
                    event_reward_item_data[key].append(to_append)
        ##代币购买道具包（Token purchase bundle）
        tokenBundles: list[dict[str, Any]] = eventInfo["tokenBundles"]
        for bundle_index in range(len(tokenBundles)):
            bundle: dict[str, Any] = tokenBundles[bundle_index]
            for i in range(len(token_bundle_header_keys)):
                key: str = token_bundle_header_keys[i]
                if i == 0 or i == 1: #事件相关键（Event-related keys）
                    if bundle_index == 0:
                        to_append: Any = eventInfo[key]
                    else:
                        to_append = ""
                else: #道具包相关键（Bundle-related keys）
                    if i == 2: #道具包序号（`bundle_index`）
                        to_append = bundle_index + 1
                    else:
                        to_append = bundle[key]
                token_bundle_data[key].append(to_append)
        ##代币商城交易类别（Token shop offer category）
        if not "errorCode" in token_shop_categories_offer:
            for category_index in range(len(token_shop_categories_offer)):
                offerCategory: dict[str, Any] = token_shop_categories_offer[category_index]
                for offer_index in range(len(offerCategory["offers"])):
                    offer: dict[str, Any] = offerCategory["offers"][offer_index]
                    offer_items: list[dict[str, Any]] = offer["items"]
                    offer_items_empty: bool = len(offer_items) == 0
                    for item_index in range(max(1, len(offer_items))):
                        offer_item: dict[str, Any] = offer_items[item_index]
                        for i in range(len(tokenShop_categoryOffer_header_keys)):
                            key: str = tokenShop_categoryOffer_header_keys[i]
                            if i == 0 or i == 1: #事件相关键（Event-related keys）
                                if category_index == 0 and offer_index == 0 and item_index == 0:
                                    to_append: Any = eventInfo[key]
                                else:
                                    to_append = ""
                            elif i <= 4: #交易分类相关键（Offer category-related keys）
                                if offer_index == 0 and item_index == 0:
                                    if i == 2: #分类序号（`category_index`）
                                        to_append = category_index + 1
                                    elif i == 3: #类别（`category`）
                                        to_append = lolEventHubOfferCategories[offerCategory["category"]]
                                    else: #类别图标路径（`categoryIconPath`）
                                        to_append = offerCategory["categoryIconPath"]
                                else:
                                    to_append = ""
                            elif i <= 13: #交易相关键（Offer-related keys）
                                if item_index == 0:
                                    if i == 10: #交易已突出显示（`offer highlighted`）
                                        to_append = "√" if offer["highlighted"] else ""
                                    elif i == 11: #交易状态（`offer offerState`）
                                        to_append = lolEventHubOfferStates[offer["offerState"]]
                                    else:
                                        to_append = offer[key.split()[1]]
                                else:
                                    to_append = ""
                            else: #交易商品相关键（Offer item-related keys）
                                if offer_items_empty:
                                    to_append = ""
                                else:
                                    if i == 14: #交易道具序号（`offer item_index`）
                                        to_append = item_index + 1
                                    elif i == 16: #交易道具类型（`offer item inventoryType`）
                                        to_append = inventoryType_dict[offer_item["inventoryType"]]
                                    else:
                                        to_append = offer_item[key.split()[2]]
                            tokenShop_categoryOffer_data[key].append(to_append)
    #构建数据框和排序（Build dataframes and sort the keys and values）
    ##事件信息（Event information）
    event_info_statistics_output_order: list[int] = [4, 5, 10, 13, 7, 18, 9, 8, 17, 25, 19, 16, 2, 27, 22, 23, 24, 53, 51, 28, 54, 55, 56, 57, 61, 58, 64, 62, 21, 20, 26, 59, 70, 98, 60, 63, 65, 72, 73, 71, 74, 75, 66, 67, 35, 41, 42, 87, 88, 89, 90, 91, 101, 103, 1, 14, 104, 0, 3, 6, 12, 15, 86, 11]
    event_info_data_organized: dict[str, list[Any]] = {event_info_header_keys[i]: event_info_data[event_info_header_keys[i]] for i in event_info_statistics_output_order}
    event_info_df: pandas.DataFrame = pandas.DataFrame(data = event_info_data_organized)
    event_info_df = pandas.concat([pandas.DataFrame([event_info_header])[event_info_df.columns], event_info_df], ignore_index = True)
    ##叙事章节（Narrative chapter）
    event_narrative_statistics_output_order: list[int] = [1, 0, 2, 3, 4, 6, 5, 9, 12, 11, 8, 10, 7]
    event_narrative_data_organized: dict[str, list[Any]] = {event_narrative_header_keys[i]: event_narrative_data[event_narrative_header_keys[i]] for i in event_narrative_statistics_output_order}
    event_narrative_df: pandas.DataFrame = pandas.DataFrame(data = event_narrative_data_organized)
    event_narrative_df = pandas.concat([pandas.DataFrame([event_narrative_header])[event_narrative_df.columns], event_narrative_df], ignore_index = True)
    ##通行证章节（Pass chapter）
    event_pass_chapter_statistics_output_order: list[int] = [1, 0, 5, 6, 4, 9, 8, 7, 2, 3, 10]
    event_pass_chapter_data_organized: dict[str, list[Any]] = {event_pass_chapter_header_keys[i]: event_pass_chapter_data[event_pass_chapter_header_keys[i]] for i in event_pass_chapter_statistics_output_order}
    event_pass_chapter_df: pandas.DataFrame = pandas.DataFrame(data = event_pass_chapter_data_organized)
    event_pass_chapter_df = pandas.concat([pandas.DataFrame([event_pass_chapter_header])[event_pass_chapter_df.columns], event_pass_chapter_df], ignore_index = True)
    ##通行证购买道具包（Pass purchase bundle）
    event_pass_bundle_statistics_output_order: list[int] = [1, 0, 2, 12, 9, 10, 16, 11, 17, 18, 19, 14, 13, 7, 6, 4, 3, 5, 8, 15, 20, 25, 22, 23, 29, 24, 27, 26, 21, 28]
    event_pass_bundle_data_organized: dict[str, list[Any]] = {event_pass_bundle_header_keys[i]: event_pass_bundle_data[event_pass_bundle_header_keys[i]] for i in event_pass_bundle_statistics_output_order}
    event_pass_bundle_df: pandas.DataFrame = pandas.DataFrame(data = event_pass_bundle_data_organized)
    event_pass_bundle_df = pandas.concat([pandas.DataFrame([event_pass_bundle_header])[event_pass_bundle_df.columns], event_pass_bundle_df], ignore_index = True)
    ##通行证奖励（Pass reward）
    event_reward_item_statistics_output_order: list[int] = [1, 0, 2, 3, 4, 5, 6, 7, 18, 12, 15, 16, 17, 9, 10, 19, 21, 11, 13, 8, 14, 20, 22]
    event_reward_item_data_organized: dict[str, list[Any]] = {event_reward_item_header_keys[i]: event_reward_item_data[event_reward_item_header_keys[i]] for i in event_reward_item_statistics_output_order}
    event_reward_item_df: pandas.DataFrame = pandas.DataFrame(data = event_reward_item_data_organized)
    event_reward_item_df = pandas.concat([pandas.DataFrame([event_reward_item_header])[event_reward_item_df.columns], event_reward_item_df], ignore_index = True)
    ##代币购买道具包（Token purchase bundle）
    token_bundle_statistics_output_order: list[int] = [1, 0, 2, 3, 4, 5, 6]
    token_bundle_data_organized: dict[str, list[Any]] = {token_bundle_header_keys[i]: token_bundle_data[token_bundle_header_keys[i]] for i in token_bundle_statistics_output_order}
    token_bundle_df: pandas.DataFrame = pandas.DataFrame(data = token_bundle_data_organized)
    token_bundle_df = pandas.concat([pandas.DataFrame([token_bundle_header])[token_bundle_df.columns], token_bundle_df], ignore_index = True)
    ##代币商城交易类别（Token shop offer category）
    tokenShop_categoryOffer_statistics_output_order: list[int] = [1, 0, 2, 3, 4, 5, 7, 8, 6, 11, 13, 12, 10, 9, 14, 15, 16, 17, 18]
    tokenShop_categoryOffer_data_organized: dict[str, list[Any]] = {tokenShop_categoryOffer_header_keys[i]: tokenShop_categoryOffer_data[tokenShop_categoryOffer_header_keys[i]] for i in tokenShop_categoryOffer_statistics_output_order}
    tokenShop_categoryOffer_df: pandas.DataFrame = pandas.DataFrame(data = tokenShop_categoryOffer_data_organized)
    tokenShop_categoryOffer_df = pandas.concat([pandas.DataFrame([tokenShop_categoryOffer_header])[tokenShop_categoryOffer_df.columns], tokenShop_categoryOffer_df], ignore_index = True)
    #保存文件（Save file）
    print("开始导出到工作簿。\nBegin to export to the workbook.\n")
    excel_name: str = "Event Pass - %s.xlsx" %displayName
    excel_name_sorted: str = "Event Pass - %s (sorted).xlsx" %displayName
    currentTime: str = time.strftime("%Y-%m-%d %H-%M", time.localtime(time.time()))
    wbPath: str = os.path.join(folder, excel_name).replace("\\", "/")
    os.makedirs(folder, exist_ok = True)
    if not os.path.exists(wbPath):
        wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath), sheet1_name = f"Info - {currentTime}")
    workbook_exist: bool = os.path.exists(wbPath)
    while True:
        try:
            with (pandas.ExcelWriter(path = wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = wbPath)) as writer:
                addDefaultStyle(event_info_df.transpose()).to_excel(excel_writer = writer, sheet_name = f"Info - {currentTime}")
                if len(event_narrative_df) > 1:
                    addDefaultStyle(event_narrative_df).to_excel(excel_writer = writer, sheet_name = f"Narrative - {currentTime}")
                if len(event_pass_chapter_df) > 1:
                    addDefaultStyle(event_pass_chapter_df).to_excel(excel_writer = writer, sheet_name = f"PassChapter - {currentTime}")
                if len(event_pass_bundle_df) > 1:
                    addDefaultStyle(event_pass_bundle_df).to_excel(excel_writer = writer, sheet_name = f"PassBundle - {currentTime}")
                if len(event_reward_item_df) > 1:
                    addDefaultStyle(event_reward_item_df).to_excel(excel_writer = writer, sheet_name = f"RewardItem - {currentTime}")
                if len(token_bundle_df) > 1:
                    addDefaultStyle(token_bundle_df).to_excel(excel_writer = writer, sheet_name = f"TokenBundle - {currentTime}")
                if len(tokenShop_categoryOffer_df) > 1:
                    addDefaultStyle(tokenShop_categoryOffer_df).to_excel(excel_writer = writer, sheet_name = f"OfferCat - {currentTime}") #全名（Full name）： OfferCategory
                for sheet_name in [f"Info - {currentTime}", f"Narrative - {currentTime}", f"PassChapter - {currentTime}", f"PassBundle - {currentTime}", f"RewardItem - {currentTime}", f"TokenBundle - {currentTime}", f"OfferCat - {currentTime}"]:
                    if sheet_name in writer.sheets:
                        worksheet: Worksheet = writer.sheets[sheet_name]
                        if worksheet.calculate_dimension() != "A1:A1":
                            worksheet.cell(row = 1, column = 1, value = version) #在A1单元格填充数据所在版本（Fill in A0 cell with the data version）
        except PermissionError:
            print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
            input()
        else:
            print('事件通行证信息已保存为“%s”！\nEvent pass information is saved as "%s"!' %(wbPath, wbPath))
            break
    if workbook_exist:
        print("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致工作表顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of worksheet order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
        sort: bool = bool(input())
        if sort:
            print("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
            while True:
                try:
                    wb: Workbook = load_workbook(wbPath)
                except FileNotFoundError:
                    print('商品藏品信息工作簿读取失败！请确保“%s”文件夹内含有名为“%s”的工作簿。如果需要退出程序，请输入“0”。\nERROR reading the Catalog and Collections workbook! Please make sure the workbook "%s" is in the folder "%s". If you want to exit the program, please submit "0".' %(folder, excel_name, excel_name, folder))
                    store_reload: str = input()
                    if store_reload == "0":
                        break
                else:
                    sheetnames: list[str] = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                    print("请选择排序方式：\nPlease select an ordering pattern:\n☆1\t时间优先（Time in priority）\n2\t类别优先（Type in priority）")
                    op: str = input()
                    print("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                    date_re: re.Pattern[str] = re.compile(r"\d{4}-\d{2}-\d{2} \d{2}-\d{2}") #设置正则表达式识别日期（Define a regular expression to identify a date pattern）
                    dOrder: list[str] = ["Info", "Narrative", "PassChapter", "PassBundle", "RewardItem", "TokenBundle", "OfferCat"] #存储数据类型的排列顺序（Store the order of data types）
                    dOrder_type_map: dict[str, int] = {_: dOrder.index(_) for _ in dOrder} #定义数据类型权重字典，用于排序数据类型（Define a data type weight dictionary to order the data types）
                    sheetname_date_list: list[str] = list(map(lambda x: date_re.search(x).group(), sheetnames)) #从工作表名称提取日期信息形成列表（Extract the dates from the sheetnames to form a list）
                    sheetname_type_list: list[str] = list(map(lambda x: x.split()[0], sheetnames)) #从工作表名称提取数据类型信息形成列表（Extract the data types from the sheetnames to form a list）
                    sheetname_type_weight_list: list[int] = list(map(lambda x: dOrder_type_map.get(x, len(dOrder) + 1), sheetname_type_list)) #将数据类型列表转换为数据类型权重列表（Transform the data type list into the data type weight list）
                    sheetname_tmpDf: pandas.DataFrame = pandas.DataFrame(data = [sheetnames, sheetname_date_list, sheetname_type_list, sheetname_type_weight_list]).transpose() #创建一个四列数据框，各列分别是完整工作表名、日期信息、数据类型信息和大区信息（Create a 4-column dataframe whose columns are the complete sheetname, date, data type and platformId）
                    if op == "" or op[0] != "2": #按照时间优先的原则对工作表进行排序，时间相同则商品工作表在前，藏品工作表在后（Sort the sheets by time in priority. If the times are the same, then the store sheet is arranged in front of the collection sheet）
                        sheetnames_sorted: list[str] = sheetname_tmpDf.sort_values(by = [1, 3], ascending = True).iloc[:, 0].tolist() #将工作表名按照第一关键字——日期信息正序排列，第二关键字——数据类型权重正序排列（Order the sheetnames according to the ascending order of the first keyword - date and the ascending order of the second keyword - data type weight）
                    else:
                        sheetnames_sorted: list[str] = sheetname_tmpDf.sort_values(by = [3, 1], ascending = True).iloc[:, 0].tolist() #将工作表名按照第一关键字——数据类型权重正序排列，第二关键字——日期信息正序排列（Order the sheetnames according to the ascending order of the first keyword - data type weight and the ascending order of the second keyword - date）
                    #下面排列所有工作表（The following code arrange all sheets）
                    print("正在排序……\nOrdering ...")
                    sort_worksheet(wb, sheetnames_sorted)
                    print('正在保存中……\nSaving the ordered workbook ...')
                    wb.save(os.path.join(folder, excel_name_sorted))
                    print('排序完成！排好序的工作簿已保存为“%s”。请按任意键退出。\nOrdering finished! The ordered workbook is saved as "%s". Press any key to exit ...\n' %(excel_name_sorted, excel_name_sorted))
                    wb.close()
                    input()
                    break

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection: Connection) -> None:
    await print_summoner_info(connection)
    await save_platform_info(connection)
    await prepare_data_resources(connection)
    await organize_pass_information(connection)

@connector.close
async def disconnect(connection: Connection) -> None:
    print("已从英雄联盟客户端断开连接。\nDisconnected from the League Client.")

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

connector.start()
