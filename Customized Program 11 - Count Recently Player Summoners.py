from lcu_driver import Connector
import argparse, copy, json, os, pandas, pickle, pyperclip, re, requests, shutil, time, traceback, unicodedata, uuid, _io
from urllib.parse import quote, unquote
from wcwidth import wcswidth
import matplotlib.pyplot as plt
from openpyxl.styles import Color, numbers, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser()
parser.add_argument("-r", "--reserve", help = "在对局不包含主玩家的情况下仍然加载该对局（Load a match even if it doesn't contain the main player）", action = "store_true")
parser.add_argument("-ss", "--save_self", help = "在对局包含主玩家的情况下仍然保存其数据（Save the main summoner's data even if they're contained in a match）", action = "store_true")
args = parser.parse_args()

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2025/08/21
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

log_folder = "日志（Logs）/Customized Program 11 - Count Recently Played Summoners"
os.makedirs(log_folder, exist_ok = True)
currentTime = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
log = open(os.path.join(log_folder, currentTime + ".log"), "a+", encoding = "utf-8")

error_header = {"errorCode": "异常代码", "httpStatus": "HTTP状态码", "implementationDetails": "细节", "message": "消息"}
error_header_keys = list(error_header.keys())
connector = Connector()

async def get_summoner_data(connection):
    data = await connection.request('GET', '/lol-summoner/v1/current-summoner')
    summoner = await data.json()
    print("displayName:    %s" %(summoner["gameName"] + "#" + summoner["tagLine"]))
    print("summonerId:     %s" %(summoner["summonerId"]))
    print("puuid:          %s" %(summoner["puuid"]))
    print("-")


#-----------------------------------------------------------------------------
#  lockfile
#-----------------------------------------------------------------------------
async def update_lockfile(connection):
    path = os.path.join(connection.installation_path.encode('gb18030').decode('utf-8'), 'lockfile')
    if os.path.isfile(path):
        file = open(path, 'w+')
        text = "LeagueClient:%d:%d:%s:%s" %(connection.pid, connection.port, connection.auth_key, connection.protocols[0])
        file.write(text)
        file.close()
    return None

async def get_lockfile(connection):
    path = os.path.join(connection.installation_path.encode('gb18030').decode('utf-8'), 'lockfile')
    if os.path.isfile(path):
        file = open(path, 'r')
        text = file.readline().split(':')
        file.close()
        print(connection.address)
        print(f'riot    {connection.auth_key}')
        return connection.auth_key
    return None

#-----------------------------------------------------------------------------
# 查找最近一起并肩作战的召唤师并给出统计信息（Find recently played summoners and give statistics of it）
#-----------------------------------------------------------------------------
def logInput(prompt: str = "", log: _io.TextIOWrapper = log, write_time: bool = True):
    s = input(prompt)
    if isinstance(log, _io.TextIOWrapper):
        if write_time:
            currentTime = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
            log.write("[%s]%s\n" %(currentTime, prompt + s))
        else:
            log.write(prompt + s + "\n")
    return s

def logPrint(s: str = "", log: _io.TextIOWrapper = log, end: str = "\n", print_time: bool = False, write_time: bool = True):
    currentTime = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    if print_time:
        print("[%s]%s" %(currentTime, s), end = end, flush = end == "\r")
    else:
        print(s, end = end, flush = end == "\r")
    if isinstance(log, _io.TextIOWrapper):
        if write_time:
            log.write("[%s]%s%s" %(currentTime, str(s), "\n" if end == "\r" else end))
        else:
            log.write("%s%s" %(str(s), "\n" if end == "\r" else end))

def format_json(origin = '''{"customGameLobby": {"configuration": {"gameMode": "PRACTICETOOL","gameMutator": "","gameServerRegion": "","mapId": 11,"mutators": {"id": 4},"spectatorPolicy": "AllAllowed","teamSize": 5},"lobbyName": "WordlessMeteor's Game","lobbyPassword": null},"isCustom": true}'''): #对字符串origin进行格式化
    temp = list(str(origin))
    brace = 0 # brace用来根据花括号的级别输出对应数量的水平制表符(brace is used to input the corresponding number of horizontal tabs based on the hierachy of the curly brackets）
    for i in range(len(temp)): #j遍历temp列表
        if temp[i] == "{":
            square_bracket = 0
            brace += 1
            temp[i] = "{\n" + brace * "\t"
        elif temp[i] == ":" and temp[i + 1] != " ":
            temp[i] = ": "
        elif temp[i] == "}":
            brace -= 1
            temp[i] = "\n" + brace * "\t" + "}"
        elif temp[i] == "[":
            square_bracket = 1
            temp[i] = "["
        elif temp[i] == "]":
            square_bracket = 0
            temp[i] = "]"
        elif temp[i] == "," and not square_bracket:
            temp[i] = ",\n" + brace * "\t"
    result = "".join(temp)
    return result

def count_nonASCII(s: str): #统计一个字符串中占用命令行2个宽度单位的字符个数（Count the number of characters that take up 2 width unit in CMD）
    return sum([unicodedata.east_asian_width(character) in ("F", "W") for character in list(str(s))])

def rm_ctrl_char(s: str): #移除一个字符串中的所有C0和C1字符（Remove all C0 and C1 characters from a string）
    return "".join(ch for ch in s if unicodedata.category(ch) != "Cc")

def format_df(df: pandas.DataFrame, width_exceed_ask: bool = True, direct_print: bool = False, print_header: bool = True, print_index: bool = False, reserve_index = False, start_index = 0, header_align: str = "^", align: str = "^", align_replicate_rule: str = "all"): #按照每列最长字符串的命令行宽度加上2，再根据每个数据的中文字符数量决定最终格式化输出的字符串宽度（Get the width of the longest string of each column, add it by 2, and substract it by the number of each cell string's Chinese characters to get the final width for each cell to print using `format` function）
    df = df.copy(deep = True)
    old_index = df.index
    df.index = range(start_index, len(df) + start_index)
    maxLens = {}
    maxWidth = shutil.get_terminal_size()[0]
    fields = df.columns.tolist()
    for field in fields:
        maxLens[field] = max(0 if len(df) == 0 else max(map(lambda x: wcswidth(rm_ctrl_char(str(x))), df[field])), wcswidth(rm_ctrl_char(field))) + 2
    index_len = 0 if len(df) == 0 else max(map(lambda x: len(str(x)), old_index)) if reserve_index else max(len(str(start_index)), len(str(start_index + len(df) - 1)))
    if sum(maxLens.values()) + 2 * (len(fields) - 1) > maxWidth or print_index and index_len + sum(maxLens.values()) + 2 * len(fields) > maxWidth:
        if width_exceed_ask:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！是否继续？（输入任意键继续，否则直接打印该数据框。）\nThe output width of each record string exceeds the current width of the terminal window! Continue? (Input anything to continue, or null to directly print this dataframe.)")
            if not bool(input()):
                #print(df)
                result = str(df)
                return (result, maxLens)
        elif direct_print:
            # print("单行数据字符串输出宽度超过当前终端窗口宽度！将直接打印该数据框！\nThe output width of each record string exceeds the current width of the terminal window! The program is going to directly print this dataframe!")
            result = str(df)
            return (result, maxLens)
        # else:
        #     print("单行数据字符串输出宽度超过当前终端窗口宽度！将继续格式化输出！\nThe output width of each record string exceeds the current width of the terminal window! The program is going on formatted printing!")
    result = ""
    #确定各列的排列方向（Determine the alignments of all columns）
    if isinstance(header_align, str) and isinstance(align, str):
        if not all(map(lambda x: x in {"<", "^", ">"}, header_align)) or not all(map(lambda x: x in {"<", "^", ">"}, align)):
            print('排列方式字符串参数错误！排列方式必须是“<”“^”或者“>”中的一个。请修改排列方式字符串参数。\nParameter ERROR of the alignment string! The alignment value must be one of {"<", "^", ">"}. Please change the alignment string parameter.')
        if len(header_align) == 0: #指定为空字符串，即默认居中输出（Specifying it as a null string means output centered by default）
            header_alignments = ["^"] * df.shape[1]
        elif len(header_align) == 1:
            header_alignments = [header_align] * df.shape[1]
        else:
            header_alignments_tmp = list(header_align)
            if len(header_align) < df.shape[1]:
                if align_replicate_rule == "last":
                    header_alignments = header_alignments_tmp + [header_alignments_tmp[-1]] * len(df.shape[1] - len(header_align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    header_alignments = header_alignments_tmp * (df.shape[1] // len(header_align)) + header_alignments_tmp[:df.shape[1] % len(header_align)]
            else:
                header_alignments = header_alignments_tmp[:df.shape[1]]
        if len(align) == 0:
            alignments = ["^"] * df.shape[1]
        elif len(align) == 1:
            alignments = [align] * df.shape[1]
        else:
            alignments_tmp = list(align)
            if len(align) < df.shape[1]:
                if align_replicate_rule == "last":
                    alignments = alignments_tmp + [alignments_tmp[-1]] * len(df.shape[1] - len(align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    alignments = alignments_tmp * (df.shape[1] // len(align)) + alignments_tmp[:df.shape[1] % len(align)]
            else:
                alignments = alignments_tmp[:df.shape[1]]
        if print_header:
            if print_index:
                result += " " * (index_len + 2)
            for i in range(df.shape[1]):
                field = fields[i]
                tmp = "{0:{align}{w}}".format(rm_ctrl_char(field), align = header_alignments[i], w = maxLens[field] - count_nonASCII(field))
                result += tmp
                #print(tmp, end = "")
                if i != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            result += "\n"
            #print()
        index = start_index
        for i in range(df.shape[0]):
            if print_index:
                result += "{0:>{w}}".format(old_index[index - start_index] if reserve_index else index, w = index_len) + "  "
            for j in range(df.shape[1]):
                field = fields[j]
                cell = str(list(df[field])[i])
                tmp = "{0:{align}{w}}".format(rm_ctrl_char(cell), align = alignments[j], w = maxLens[field] - count_nonASCII(cell))
                result += tmp
                #print(tmp, end = "")
                if j != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            if i != df.shape[0] - 1:
                result += "\n"
            #print() #注意这里的缩进和上一行不同（Note that here the indentation is different from the last line）
            index += 1
    else:
        print("排列方式参数错误！请传入字符串。\nAlignment parameter ERROR! Please pass a string instead.")
    return (result, maxLens)

def lcuTimestamp(timestamp): #根据对局时间轴的时间戳返回对局时间（Return the time according to the timestamp in match timeline）
    min = timestamp // 60
    sec = timestamp % 60
    return str(min) + ":" + "{0:0>2}".format(str(sec))

def patch_compare(patch1, patch2): #比较两个版本号的先后顺序。当patch1 < patch2时，返回True，否则返回False。用于比较DataDragon数据库中未收录的版本和收录的最新版本的关系。如果未收录的版本小于收录的最新版本，那么该版本是美测服的临时版本，后来被合并更新了，如正式服将13.2和13.3合并更新了，因此DataDragon数据库中未收录13.2版本的数据；如果未收录的版本大于收录的最新版本，那么该版本是美测服的当前版本，但是仍处于开发状态，尚未完全确定，所以DataDragon数据库尚未收录，将以最新版本代替该版本；二者不可能相等，因为如果相等的话，就不会引发报错而调用此函数（Compare the time order of two patches. When patch1 < patch2, return True and vice versa. Designed to compare a patch not archived in DataDragon database with the latest patch archived in DataDragon database. If the unarchived patch is less than the latest archived patch, then this patch must be the intermediate patch and be merged into the update of its successive patch, such as Patch 13.2 merged into the update of Patch 13.3, so that DataDragon database doesn't archive the data of Patch 13.2; If the unarchived patch is greater than the latest archived patch, then this patch must be the current patch on PBE but is under development and improvement, so that DataDragon database doesn't archive this patch, either, in which case the latest patch will be used to substitute this unarchived patch; The two patches can't be the same, for suppose they're same, then the error to cause the call of this function won't be triggered）
    if not isinstance(patch1, str):
        patch1 = str(patch1)
    if not isinstance(patch2, str):
        patch2 = str(patch2)
    lst1, lst2 = patch1.split("."), patch2.split(".")
    try:
        lst1 = list(map(int, lst1))
    except ValueError:
        if lst1[0] != "pbe":
            print("第1个版本字符串不合法！请输入用半角句号连接的正整数，如13.15.1、10.10.3216176。\nThe first patch variable is illegal! Please pass the integers concatenated by dot, such as 13.15.1 and 10.10.3216176.")
        return False
    try:
        lst2 = list(map(int, lst2))
    except ValueError:
        if lst1[0] != "pbe":
            print("第2个版本字符串不合法！请输入用半角句号连接的正整数，如13.15.1、10.10.3216176。\nThe second patch variable is illegal! Please pass the integers concatenated by dot, such as 13.15.1 and 10.10.3216176.")
            return False
        else:
            return True
    for i in range(min(len(lst1), len(lst2))):
        if lst1[i] < lst2[i]:
            return True
        elif lst1[i] > lst2[i]:
            return False
        else:
            continue
    if len(lst1) < len(lst2):
        return True
    else:
        return False #这里将两个版本相同视为假，暗示了在本程序用得到的地方，两个版本不可能相同（Here the case where the two patches are the same is regarded as False, which indicates that the two patches can't be same within its use in this program）

def FindPostPatch(patch, patchList): #二分查找某个版本号在DataDragon数据库的后一个版本（Binary search for the precedent patch of a given patch in the patch list archived in DataDragon database）
    leftIndex, rightIndex = 0, len(patchList) - 1
    mid = (leftIndex + rightIndex) // 2
    count = 0 #函数调试阶段的保护机制（A protecion mechanism during rebugging this function）
    #print("[" + str(count) + "]", leftIndex, mid, rightIndex)
    while leftIndex < rightIndex:
        count += 1
        if patch_compare(patch, patchList[mid]):
            leftIndex = mid + 1
            mid = (leftIndex + rightIndex) // 2
        elif patch_compare(patchList[mid], patch):
            rightIndex = mid
            mid = (leftIndex + rightIndex) // 2
        else:
            return patchList[mid - 1]
        #print("[" + str(count) + "]", leftIndex, mid, rightIndex)
        if count >= 15:
            print("程序即将进入死循环！请检查算法！\nThe program is stepping into a dead loop! Please check the algorithm!")
            return 1
    if mid >= 1:
        return patchList[mid - 1]
    else:
        print("该版本为美测服最新版本，暂未收录在DataDragon数据库中。\nThis version is the latest version on PBE and isn't archived in DataDragon database for now.")
        return "pbe"

async def get_info(connection, name: str, searchType: str | int = "riotId"):
    #searchTypes = {0: "selfCheck", 1: "riotId", 2: "puuid", 3: "summonerId"}
    current_info = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    result = {"searchType": "riotId", "endpoint": "/lol-summoner/v2/summoners/puuid/{puuid}", "info_got": False, "network_error": False, "body": {}, "message": "", "selfInfo": False}
    try:
        name = int(name)
    except ValueError:
        if name == "current-summoner":
            result = {"searchType": "selfCheck", "endpoint": "/lol-summoner/v1/current-summoner", "info_got": True, "network_error": False, "body": current_info, "message": "", "selfInfo": True}
        elif name.count("-") == 4 and len(name.replace(" ", "")) > 22: #拳头规定的玩家昵称不超过16个字符，昵称编号不超过5个字符（Riot game name can't exceed 16 characters. The tagline can't exceed 5 characters）
            result["searchType"] = "puuid"
            result["endpoint"] = "/lol-summoner/v2/summoners/puuid/{puuid}"
            info = await (await connection.request("GET", f"/lol-summoner/v2/summoners/puuid/{name}")).json()
            result["body"] = info
            if "errorCode" in info:
                if info["httpStatus"] == 400:
                    result["message"] = "您输入的玩家通用唯一识别码格式有误！请重新输入！\nPUUID wasn't in UUID format! Please try again!"
                elif info["httpStatus"] == 404:
                    result["message"] = "未找到玩家通用唯一识别码为%s的玩家；请核对识别码并稍后再试。\nA player with puuid %s was not found; verify the puuid and try again." %(name, name)
                else:
                    result["network_error"] = True
                    result["message"] = "网络异常。\nNetwork Error."
            else:
                result["info_got"] = True
                result["selfInfo"] = info["puuid"] == current_info["puuid"]
        else:
            result["searchType"] = "riotId"
            result["endpoint"] = "/lol-summoner/v1/summoners?name={name}"
            if name.count("#") == 0:
                result["message"] = '召唤师名称已变更为拳头ID。请以“{玩家昵称}#{昵称编号}”的格式输入。\nSummoner name has been replaced with Riot ID. Please input the name in this format: "{gameName}#{tagLine}", e.g. "%s#%s".' %(current_info["gameName"], current_info["tagLine"])
            elif name.count("#") > 1:
                result["message"] = "该玩家名字包含了无效字符。\nThis player name contains invalid characters."
            else:
                gameName, tagLine = name.split("#")
                if len(gameName) == 0:
                    result["message"] = "缺少玩家昵称。\nGame name is missing."
                elif len(tagLine) == 0:
                    result["message"] = "缺少昵称编号。\nTagline is missing."
                elif len(gameName) < 3:
                    result["message"] = "召唤师昵称过短。\nRiot ID is too short."
                elif len(gameName.replace(" ", "")) > 16:
                    result["message"] = "召唤师昵称过长。\nRiot ID is too long."
                else:
                    info = await (await connection.request("GET", "/lol-summoner/v1/summoners?name=" + quote(name))).json()
                    result["body"] = info
                    if "errorCode" in info:
                        if info["httpStatus"] == 404:
                            result["message"] = "未找到%s；请核对下名字并稍后再试。\n%s was not found; verify the name and try again." %(name, name)
                        else:
                            result["network_error"] = True
                            result["message"] = "网络异常。\nNetwork Error."
                    else:
                        result["info_got"] = True
                        result["selfInfo"] = info["puuid"] == current_info["puuid"]
    else:
        result["searchType"] = "summonerId"
        result["endpoint"] = "/lol-summoner/v1/summoners/{id}"
        info = await (await connection.request("GET", f"/lol-summoner/v1/summoners/{name}")).json()
        result["body"] = info
        if "errorCode" in info:
            if info["httpStatus"] == 400:
                if info["message"] == "Value %d for 'id' of type uint64 is out of range":
                    result["message"] = "您输入的召唤师序号格式有误！请重新输入！\nValue for 'id' of type uint64 is out of range! Please try again!"
                else:
                    result["message"] = "未找到召唤师序号为%s的玩家；请核对召唤师序号并稍后再试。\nA player with summonerId %s was not found; verify the summonerId and try again." %(name, name)
            elif info["httpStatus"] == 404:
                result["message"] = "未找到召唤师序号为%s的玩家；请核对召唤师序号并稍后再试。\nA player with summonerId %s was not found; verify the summonerId and try again." %(name, name)
            else:
                result["network_error"] = True
                result["message"] = "网络异常。\nNetwork Error."
        else:
            result["info_got"] = True
            result["selfInfo"] = info["puuid"] == current_info["puuid"]
    return result

def get_info_name(info: dict, mode = 1) -> str:
    if not isinstance(info, dict) or not all(i in info for i in ["displayName", "gameName", "tagLine"]):
        print("您的召唤师信息格式有误！\nERROR format of summoner information!")
        name = ""
    else:
        if info["displayName"] or info["gameName"]:
            if info["gameName"] and info["tagLine"]:
                name = info["gameName"] + "#" + info["tagLine"]
            elif not info["tagLine"] and info["gameName"]:
                name = info["gameName"]
            else:
                name = info["displayName"]
        else: #新玩家属于这种类型（This case matches new players）
            if mode == 1:
                name = str(info["puuid"])
            elif mode == 2: #仅用于设置召唤师数据保存路径（Designed to set the summoner name directory）
                name = "0. 新玩家\\" + str(info["puuid"])
            elif mode == 3: #仅用于设置召唤师数据保存路径（Designed to set the summoner name directory）
                name = "0. New Player\\" + str(info["puuid"])
    return name

def verify_uuid(s: str) -> bool:
    try:
        return s == str(uuid.UUID(s))
    except ValueError:
        return False

async def search_recent_players(connection):
    platform_config = await (await connection.request("GET", "/lol-platform-config/v1/namespaces")).json()
    platformId = platform_config["LoginDataPacket"]["platformId"]
    logPrint("请选择召唤师技能和装备的输出语言【默认为中文（中国）】：\nPlease select a language to output the summoner spells and items (the default option is zh_CN):") #本来考虑把可用CDragon数据版本放在第三列，但是后来发现表头名字太长了，索性放在最后了（I had considered putting "Applicable CDragon Data Patches" at the third column, but then found the header was too long. So I put it at the last column）
    language_ddragon = {1: {"CODE": "ar_AE", "LANGUAGE (EN)": "Arabic (United Arab Emirates)", "LANGUAGE (ZH)": "阿拉伯语（阿拉伯联合酋长国）", "Applicable CDragon Data Patches": "9.20～10.1, 13.20+"}, 2: {"CODE": "cs_CZ", "LANGUAGE (EN)": "Czech (Czech Republic)", "LANGUAGE (ZH)": "捷克语（捷克共和国）", "Applicable CDragon Data Patches": "7.1+"}, 3: {"CODE": "el_GR", "LANGUAGE (EN)": "Greek (Greece)", "LANGUAGE (ZH)": "希腊语（希腊）", "Applicable CDragon Data Patches": "9.1+"}, 4: {"CODE": "pl_PL", "LANGUAGE (EN)": "Polish (Poland)", "LANGUAGE (ZH)": "波兰语（波兰）", "Applicable CDragon Data Patches": "9.1+"}, 5: {"CODE": "ro_RO", "LANGUAGE (EN)": "Romanian (Romania)", "LANGUAGE (ZH)": "罗马尼亚语（罗马尼亚）", "Applicable CDragon Data Patches": "9.1+"}, 6: {"CODE": "hu_HU", "LANGUAGE (EN)": "Hungarian (Hungary)", "LANGUAGE (ZH)": "匈牙利语（匈牙利）", "Applicable CDragon Data Patches": "9.1+"}, 7: {"CODE": "en_GB", "LANGUAGE (EN)": "English (United Kingdom)", "LANGUAGE (ZH)": "英语（英国）", "Applicable CDragon Data Patches": "9.1+"}, 8: {"CODE": "de_DE", "LANGUAGE (EN)": "German (Germany)", "LANGUAGE (ZH)": "德语（德国）", "Applicable CDragon Data Patches": "7.1+"}, 9: {"CODE": "es_ES", "LANGUAGE (EN)": "Spanish (Spain)", "LANGUAGE (ZH)": "西班牙语（西班牙）", "Applicable CDragon Data Patches": "9.1+"}, 10: {"CODE": "it_IT", "LANGUAGE (EN)": "Italian (Italy)", "LANGUAGE (ZH)": "意大利语（意大利）", "Applicable CDragon Data Patches": "9.1+"}, 11: {"CODE": "fr_FR", "LANGUAGE (EN)": "French (France)", "LANGUAGE (ZH)": "法语（法国）", "Applicable CDragon Data Patches": "9.1+"}, 12: {"CODE": "ja_JP", "LANGUAGE (EN)": "Japanese (Japan)", "LANGUAGE (ZH)": "日语（日本）", "Applicable CDragon Data Patches": "9.1+"}, 13: {"CODE": "ko_KR", "LANGUAGE (EN)": "Korean (Korea)", "LANGUAGE (ZH)": "朝鲜语（韩国）", "Applicable CDragon Data Patches": "9.7+"}, 14: {"CODE": "es_MX", "LANGUAGE (EN)": "Spanish (Mexico)", "LANGUAGE (ZH)": "西班牙语（墨西哥）", "Applicable CDragon Data Patches": "9.1+"}, 15: {"CODE": "es_AR", "LANGUAGE (EN)": "Spanish (Argentina)", "LANGUAGE (ZH)": "西班牙语（阿根廷）", "Applicable CDragon Data Patches": "9.7+"}, 16: {"CODE": "pt_BR", "LANGUAGE (EN)": "Portuguese (Brazil)", "LANGUAGE (ZH)": "葡萄牙语（巴西）", "Applicable CDragon Data Patches": "9.1+"}, 17: {"CODE": "en_US", "LANGUAGE (EN)": "English (United States)", "LANGUAGE (ZH)": "英语（美国）", "Applicable CDragon Data Patches": "9.1+"}, 18: {"CODE": "en_AU", "LANGUAGE (EN)": "English (Australia)", "LANGUAGE (ZH)": "英语（澳大利亚）", "Applicable CDragon Data Patches": "9.1+"}, 19: {"CODE": "ru_RU", "LANGUAGE (EN)": "Russian (Russia)", "LANGUAGE (ZH)": "俄语（俄罗斯）", "Applicable CDragon Data Patches": "9.1+"}, 20: {"CODE": "tr_TR", "LANGUAGE (EN)": "Turkish (Turkey)", "LANGUAGE (ZH)": "土耳其语（土耳其）", "Applicable CDragon Data Patches": "9.1+"}, 21: {"CODE": "ms_MY", "LANGUAGE (EN)": "Malay (Malaysia)", "LANGUAGE (ZH)": "马来语（马来西亚）", "Applicable CDragon Data Patches": ""}, 22: {"CODE": "en_PH", "LANGUAGE (EN)": "English (Republic of the Philippines)", "LANGUAGE (ZH)": "英语（菲律宾共和国）", "Applicable CDragon Data Patches": "10.5+"}, 23: {"CODE": "en_SG", "LANGUAGE (EN)": "English (Singapore)", "LANGUAGE (ZH)": "英语（新加坡）", "Applicable CDragon Data Patches": "10.5+"}, 24: {"CODE": "th_TH", "LANGUAGE (EN)": "Thai (Thailand)", "LANGUAGE (ZH)": "泰语（泰国）", "Applicable CDragon Data Patches": "9.7+"}, 25: {"CODE": "vn_VN", "LANGUAGE (EN)": "Vietnamese (Viet Nam)", "LANGUAGE (ZH)": "越南语（越南）", "Applicable CDragon Data Patches": "9.7～13.9"}, 26: {"CODE": "vi_VN", "LANGUAGE (EN)": "Vietnamese (Viet Nam)", "LANGUAGE (ZH)": "越南语（越南）", "Applicable CDragon Data Patches": "12.17+"}, 27: {"CODE": "id_ID", "LANGUAGE (EN)": "Indonesian (Indonesia)", "LANGUAGE (ZH)": "印度尼西亚语（印度尼西亚）", "Applicable CDragon Data Patches": ""}, 28: {"CODE": "zh_MY", "LANGUAGE (EN)": "Chinese (Malaysia)", "LANGUAGE (ZH)": "中文（马来西亚）", "Applicable CDragon Data Patches": "10.5+"}, 29: {"CODE": "zh_CN", "LANGUAGE (EN)": "Chinese (China)", "LANGUAGE (ZH)": "中文（中国）", "Applicable CDragon Data Patches": "9.7+"}, 30: {"CODE": "zh_TW", "LANGUAGE (EN)": "Chinese (Taiwan)", "LANGUAGE (ZH)": "中文（台湾）", "Applicable CDragon Data Patches": "9.7+"}}
    language_cdragon = {}
    for i in language_ddragon:
        if language_ddragon[i]["CODE"] == "en_US":
            language_cdragon[language_ddragon[i]["CODE"]] = "default" #在CommunityDragon数据库上，美服正式服的数据资源代码是default，而不是小写的en_US（The code for English (US) data resources on CommunityDragon database is "default" instead of the lowercase of "en_US"）
        else:
            language_cdragon[language_ddragon[i]["CODE"]] = language_ddragon[i]["CODE"].lower()
    language_dict = {"No.": list(language_ddragon.keys()), "CODE": list(map(lambda x: x["CODE"], language_ddragon.values())), "LANGUAGE": list(map(lambda x: x["LANGUAGE (EN)"], language_ddragon.values())), "语言": list(map(lambda x: x["LANGUAGE (ZH)"], language_ddragon.values())), "Applicable CDragon Data Patches": list(map(lambda x: x["Applicable CDragon Data Patches"], language_ddragon.values()))}
    language_df = pandas.DataFrame(language_dict)
    print(format_df(language_df)[0])
    log.write(format_df(language_df, width_exceed_ask = False, direct_print = False)[0] + "\n")
    while True:
        language_option = input()
        if language_option == "" or language_option in [str(i) for i in range(1, 31)]:
            if language_option == "":
                language_option = "29"
            language_code = language_ddragon[int(language_option)]["CODE"]
            #下面声明一些数据资源的地址（The following code declare some data resources' URLs）
            URLPatch = "pbe" if platformId == "PBE1" or platformId == "PBE" else "latest"
            patches_url = "https://ddragon.leagueoflegends.com/api/versions.json"
            spell_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-spells.json" %(URLPatch, language_cdragon[language_code]) #CommunityDragon数据库只存储第7赛季及以后的数据（CommunityDragon database only stores data including and after Season 7）
            LoLChampion_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/champion-summary.json" %(URLPatch, language_cdragon[language_code])
            LoLItem_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/items.json" %(URLPatch, language_cdragon[language_code])
            summonerIcon_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-icons.json" %(URLPatch, language_cdragon[language_code])
            perk_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perks.json" %(URLPatch, language_cdragon[language_code])
            perkstyle_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perkstyles.json" %(URLPatch, language_cdragon[language_code])
            TFT_url = "https://raw.communitydragon.org/%s/cdragon/tft/%s.json" %(URLPatch, language_code.lower())
            TFTChampion_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tftchampions.json" %(URLPatch, language_cdragon[language_code])
            TFTItem_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tftitems.json" %(URLPatch, language_cdragon[language_code])
            TFTCompanion_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/companions.json" %(URLPatch, language_cdragon[language_code])
            TFTTrait_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tfttraits.json" %(URLPatch, language_cdragon[language_code])
            CherryAugment_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/cherry-augments.json" %(URLPatch, language_cdragon[language_code])
            #下面声明离线数据资源的默认地址（The following code declare the default paths of offline data resources）
            patches_local_default = "离线数据（Offline Data）\\versions.json"
            spell_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\summoner-spells.json" %(URLPatch, language_cdragon[language_code])
            LoLChampion_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\champion-summary.json" %(URLPatch, language_cdragon[language_code])
            LoLItem_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\items.json" %(URLPatch, language_cdragon[language_code])
            summonerIcon_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\summoner-icons.json" %(URLPatch, language_cdragon[language_code])
            perk_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\perks.json" %(URLPatch, language_cdragon[language_code])
            perkstyle_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\perkstyles.json" %(URLPatch, language_cdragon[language_code])
            TFT_local_default = "离线数据（Offline Data）\\cdragon\\%s\\cdragon\\tft\\%s.json" %(URLPatch, language_code.lower())
            TFTChampion_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\tftchampions.json" %(URLPatch, language_cdragon[language_code])
            TFTItem_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\tftitems.json" %(URLPatch, language_cdragon[language_code])
            TFTCompanion_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\companions.json" %(URLPatch, language_cdragon[language_code])
            TFTTrait_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\tfttraits.json" %(URLPatch, language_cdragon[language_code])
            CherryAugment_local_default = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\cherry-augments.json" %(URLPatch, language_code.lower())
            logPrint("请选择数据资源获取模式：\nPlease select the data resource capture mode:\n1\t在线模式（Online）\n2\t离线模式（Offline）")
            prepareMode = logInput()
            switch_language = False
            while True:
                if prepareMode != "" and prepareMode[0] == "1":
                    switch_prepare_mode = False
                    #下面获取版本信息（The following code get the patch data）
                    try:
                        patches_initial = requests.get(patches_url).json()
                    except requests.exceptions.RequestException:
                        logPrint('版本信息获取超时！正在尝试离线加载数据……\nPatch information capture timeout! Trying loading offline data ...\n请输入版本Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the patch Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(patches_local_default, patches_local_default))
                        while True:
                            patches_local = logInput()
                            if patches_local == "":
                                patches_local = patches_local_default
                            elif patches_local[0] == "0":
                                logPrint("版本信息获取失败！请检查系统网络状况和代理设置。\nPatch information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(patches_local, "r", encoding = "utf-8") as fp:
                                    patches_initial = json.load(fp)
                                if isinstance(patches_initial, list) and patches_initial[-1] == "lolpatch_3.7":
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合DataDragon数据库中记录的版本数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the patch data archived in DataDragon database (%s)!" %(patches_url, patches_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的版本Json数据文件路径！\nFile "%s" NOT found! Please input a correct patch Json data file path!' %(patches_local, patches_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有版本信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with patch information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合DataDragon数据库中记录的版本数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the patch data archived in DataDragon database (%s)!" %(patches_url, patches_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    latest_patch = patches_initial[0]
                    patches_dict = {}
                    smallPatches = []
                    bigPatches = []
                    for patch in patches_initial:
                        if not patch.startswith("lolpatch"):
                            patch_split = patch.split(".")
                            smallPatch = ".".join(patch_split[:3])
                            smallPatches.append(smallPatch)
                            bigPatch = ".".join(patch_split[:2])
                            bigPatches.append(bigPatch)
                            patches_dict[bigPatch] = []
                    for i in range(len(bigPatches)):
                        patches_dict[bigPatches[i]].append(smallPatches[i])
                    #下面获取召唤师技能数据（The following code get summoenr spell data）
                    try:
                        logPrint("正在加载召唤师技能信息……\nLoading summoner spell information from CommunityDragon...")
                        spell_initial = requests.get(spell_url) #spell存储召唤师技能信息（Variable `spell_initial` stores summoner spell information）
                        if spell_initial.ok:
                            spell_initial = spell_initial.json()
                        else:
                            logPrint(spell_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('召唤师技能信息获取超时！正在尝试离线加载数据……\nSummoner spell information capture timeout! Trying loading offline data ...\n请输入召唤师技能Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the summoner spell Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(spell_local_default, spell_local_default))
                        while True:
                            spell_local = logInput()
                            if spell_local == "":
                                spell_local = spell_local_default
                            elif spell_local[0] == "0":
                                logPrint("召唤师技能信息获取失败！请检查系统网络状况和代理设置。\nSummoner spell information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(spell_local, "r", encoding = "utf-8") as fp:
                                    spell_initial = json.load(fp)
                                if isinstance(spell_initial, list) and all(i in spell_initial[j] for i in ["id", "name", "description", "summonerLevel", "cooldown", "gameModes", "iconPath"] for j in range(len(spell_initial))):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师技能数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner spell data archived in CommunityDragon database (%s)!" %(spell_url, spell_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的召唤师技能Json数据文件路径！\nFile "%s" NOT found! Please input a correct summoner spell Json data file path!' %(spell_local, spell_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有召唤师技能信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with summoner spell information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师技能数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner spell data archived in CommunityDragon database (%s)!" %(spell_url, spell_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取英雄信息（The following code get LoL champion data）
                    try:
                        logPrint("正在加载英雄信息……\nLoading LoL champion information from CommunityDragon...")
                        LoLChampion_initial = requests.get(LoLChampion_url) #LoLItem存储英雄信息。（Variable `LoLChampion_initial` stores information of LoL champions）
                        if LoLChampion_initial.ok:
                            LoLChampion_initial = LoLChampion_initial.json()
                        else:
                            logPrint(LoLChampion_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('英雄信息获取超时！正在尝试离线加载数据……\nLoL champion information capture timeout! Trying loading offline data ...\n请输入英雄Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the LoL champion Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(LoLChampion_local_default, LoLChampion_local_default))
                        while True:
                            LoLChampion_local = logInput()
                            if LoLChampion_local == "":
                                LoLChampion_local = LoLChampion_local_default
                            elif LoLChampion_local[0] == "0":
                                logPrint("英雄信息获取失败！请检查系统网络状况和代理设置。\nLoL champion information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(LoLChampion_local, "r", encoding = "utf-8") as fp:
                                    LoLChampion_initial = json.load(fp)
                                if isinstance(LoLChampion_initial, list) and all(isinstance(LoLChampion_initial[i], dict) for i in range(len(LoLChampion_initial))) and all(j in LoLChampion_initial[i] for i in range(len(LoLChampion_initial)) for j in ["id", "name", "alias", "squarePortraitPath", "roles"]) and all(isinstance(LoLChampion_initial[i]["id"], int) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["name"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["alias"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["squarePortraitPath"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["roles"], list) for i in range(len(LoLChampion_initial))):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL champion data archived in CommunityDragon database (%s)!" %(LoLChampion_url, LoLChampion_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的英雄Json数据文件路径！\nFile "%s" NOT found! Please input a correct LoL champion Json data file path!' %(LoLChampion_local, LoLChampion_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with LoL champion information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL champion data archived in CommunityDragon database (%s)!" %(LoLChampion_url, LoLChampion_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取英雄联盟装备信息（The following code get LoL item data）
                    try:
                        logPrint("正在加载英雄联盟装备信息……\nLoading LoL item information from CommunityDragon...")
                        LoLItem_initial = requests.get(LoLItem_url) #LoLItem存储经典模式的装备信息。（Variable `LoLItem_initial` stores information of LoL items）
                        if LoLItem_initial.ok:
                            LoLItem_initial = LoLItem_initial.json()
                        else:
                            logPrint(LoLItem_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('英雄联盟装备信息获取超时！正在尝试离线加载数据……\nLoL item information capture timeout! Trying loading offline data ...\n请输入英雄联盟装备Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the LoL item Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(LoLItem_local_default, LoLItem_local_default))
                        while True:
                            LoLItem_local = logInput()
                            if LoLItem_local == "":
                                LoLItem_local = LoLItem_local_default
                            elif LoLItem_local[0] == "0":
                                logPrint("英雄联盟装备信息获取失败！请检查系统网络状况和代理设置。\nLoL item information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(LoLItem_local, "r", encoding = "utf-8") as fp:
                                    LoLItem_initial = json.load(fp)
                                if isinstance(LoLItem_initial, list) and all(i in LoLItem_initial[j] for i in ["id", "name", "description", "active", "inStore", "from", "to", "categories", "maxStacks", "requiredChampion", "requiredAlly", "requiredBuffCurrencyName", "requiredBuffCurrencyCost", "specialRecipe", "isEnchantment", "price", "priceTotal", "iconPath"] for j in range(len(LoLItem_initial))):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄联盟装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL item data archived in CommunityDragon database (%s)!" %(LoLItem_url, LoLItem_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的英雄联盟装备Json数据文件路径！\nFile "%s" NOT found! Please input a correct LoL item Json data file path!' %(LoLItem_local, LoLItem_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有英雄联盟装备信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with LoL item information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄联盟装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL item data archived in CommunityDragon database (%s)!" %(LoLItem_url, LoLItem_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取召唤师图标信息（The following code get summoner icon data）
                    try:
                        logPrint("正在加载召唤师图标信息……\nLoading summoner icon information from CommunityDragon...")
                        summonerIcon_initial = requests.get(summonerIcon_url) #LoLItem存储召唤师图标信息。（Variable `summonerIcon_initial` stores information of summoner icons）
                        if summonerIcon_initial.ok:
                            summonerIcon_initial = summonerIcon_initial.json()
                        else:
                            logPrint(summonerIcon_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('召唤师图标信息获取超时！正在尝试离线加载数据……\nSummoner icon information capture timeout! Trying loading offline data ...\n请输入召唤师图标Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the summoner icon Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(summonerIcon_local_default, summonerIcon_local_default))
                        while True:
                            summonerIcon_local = logInput()
                            if summonerIcon_local == "":
                                summonerIcon_local = summonerIcon_local_default
                            elif summonerIcon_local[0] == "0":
                                logPrint("召唤师图标信息获取失败！请检查系统网络状况和代理设置。\nSummoner icon information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(summonerIcon_local, "r", encoding = "utf-8") as fp:
                                    summonerIcon_initial = json.load(fp)
                                if isinstance(summonerIcon_initial, list) and all(map(lambda x: isinstance(x, dict), summonerIcon_initial)) and all(i in j for i in ["id", "title", "yearReleased", "isLegacy", "descriptions", "rarities", "disabledRegions"] for j in summonerIcon_initial):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师图标数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner icon data archived in CommunityDragon database (%s)!" %(summonerIcon_url, summonerIcon_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的召唤师图标Json数据文件路径！\nFile "%s" NOT found! Please input a correct summoner icon Json data file path!' %(summonerIcon_local, summonerIcon_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有召唤师图标信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with summoner icon information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师图标数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner icon data archived in CommunityDragon database (%s)!" %(summonerIcon_url, summonerIcon_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取基石符文信息（The following code get perk data）
                    try:
                        logPrint("正在加载基石符文信息……\nLoading perk information from CommunityDragon...")
                        perk_initial = requests.get(perk_url) #perk存储基石符文信息。（Variable `perk_initial` stores information of perks）
                        if perk_initial.ok:
                            perk_initial = perk_initial.json()
                        else:
                            logPrint(perk_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('基石符文信息获取超时！正在尝试离线加载数据……\nPerk information capture timeout! Trying loading offline data ...\n请输入基石符文Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the perk Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(perk_local_default, perk_local_default))
                        while True:
                            perk_local = logInput()
                            if perk_local == "":
                                perk_local = perk_local_default
                            elif perk_local[0] == "0":
                                logPrint("基石符文信息获取失败！请检查系统网络状况和代理设置。\nPerk information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(perk_local, "r", encoding = "utf-8") as fp:
                                    perk_initial = json.load(fp)
                                if isinstance(perk_initial, list) and all(i in perk_initial[j] for i in ["id", "name", "majorChangePatchVersion", "tooltip", "shortDesc", "longDesc", "recommendationDescriptor", "iconPath", "endOfGameStatDescs", "recommendationDescriptorAttributes"] for j in range(len(perk_initial))):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的基石符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perk data archived in CommunityDragon database (%s)!" %(perk_url, perk_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的基石符文Json数据文件路径！\nFile "%s" NOT found! Please input a correct perk Json data file path!' %(perk_local, perk_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有基石符文信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with perk information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的基石符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perk data archived in CommunityDragon database (%s)!" %(perk_url, perk_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取符文系信息（The following code get perkstyle data）
                    try:
                        logPrint("正在加载符文系信息……\nLoading perkstyle information from CommunityDragon...")
                        perkstyle_initial = requests.get(perkstyle_url) #perkstyle存储符文系信息。（Variable `perkstyle_initial` stores information of perkstyles）
                        if perkstyle_initial.ok:
                            perkstyle_initial = perkstyle_initial.json()
                        else:
                            logPrint(perkstyle_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('符文系信息获取超时！正在尝试离线加载数据……\nPerkstyle information capture timeout! Trying loading offline data ...\n请输入符文系Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the perkstyle Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(perkstyle_local_default, perkstyle_local_default))
                        while True:
                            perkstyle_local = logInput()
                            if perkstyle_local == "":
                                perkstyle_local = perkstyle_local_default
                            elif perkstyle_local[0] == "0":
                                logPrint("符文系信息获取失败！请检查系统网络状况和代理设置。\nperkstyle information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(perkstyle_local, "r", encoding = "utf-8") as fp:
                                    perkstyle_initial = json.load(fp)
                                if isinstance(perkstyle_initial, dict) and all(perkstyle_initial.get(i, 0) for i in ["schemaVersion", "styles"]) and isinstance(perkstyle_initial["styles"], list) and all(j in perkstyle_initial["styles"][i] for i in range(len(perkstyle_initial["styles"])) for j in ["id", "name", "tooltip", "iconPath", "assetMap", "isAdvanced", "allowedSubStyles", "subStyleBonus", "slots", "defaultPageName", "defaultSubStyle", "defaultPerks", "defaultPerksWhenSplashed", "defaultStatModsPerSubStyle"]):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的符文系数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perkstyle data archived in CommunityDragon database (%s)!" %(perkstyle_url, perkstyle_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的符文系Json数据文件路径！\nFile "%s" NOT found! Please input a correct perkstyle Json data file path!' %(perkstyle_local, perkstyle_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有符文系信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with perkstyle information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的符文系数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perkstyle data archived in CommunityDragon database (%s)!" %(perkstyle_url, perkstyle_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取云顶之弈强化符文数据（The following code get TFT augment data）
                    try:
                        logPrint("正在加载云顶之弈基础数据……\nLoading TFT basic data from CommunityDragon ...")
                        TFT_initial = requests.get(TFT_url) #TFT存储云顶之弈中至今为止所有的强化符文、英雄和羁绊信息和各赛季的英雄和羁绊信息（Variable `TFT_initial` stores information of all augments, champions and traits so far and information of champions and traits with respect to season）
                        if TFT_initial.ok:
                            TFT_initial = TFT_initial.json()
                        else:
                            logPrint(TFT_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('云顶之弈基础信息获取超时！正在尝试离线加载数据……\nTFT basic information capture timeout! Trying loading offline data ...\n请输入云顶之弈基础数据Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT basics Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFT_local_default, TFT_local_default))
                        while True:
                            TFT_local = logInput()
                            if TFT_local == "":
                                TFT_local = TFT_local_default
                            elif TFT_local[0] == "0":
                                logPrint("云顶之弈基础信息获取失败！请检查系统网络状况和代理设置。\nTFT basic information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(TFT_local, "r", encoding = "utf-8") as fp:
                                    TFT_initial = json.load(fp)
                                if isinstance(TFT_initial, dict) and all(i in TFT_initial for i in ["items", "setData", "sets"]) and all(isinstance(TFT_initial[i], list) for i in ["items", "setData"]) and all(isinstance(TFT_initial[i], dict) for i in ["sets"]) and all(j in TFT_initial["items"][i] for i in range(len(TFT_initial["items"])) for j in ["apiName", "associatedTraits", "composition", "desc", "effects", "from", "icon", "id", "incompatibleTraits", "name", "tags", "unique"]) and all(isinstance(TFT_initial["items"][i][j], str) or TFT_initial["items"][i][j] == None for i in range(len(TFT_initial["items"])) for j in ["apiName", "desc", "icon", "name"]) and all(isinstance(TFT_initial["items"][i][j], list) for i in range(len(TFT_initial["items"])) for j in ["associatedTraits", "composition", "tags"]) and all(isinstance(TFT_initial["items"][i][j], dict) for i in range(len(TFT_initial["items"])) for j in ["effects"]) and all(isinstance(TFT_initial["items"][i][j], bool) for i in range(len(TFT_initial["items"])) for j in ["unique"]) and all(j in TFT_initial["setData"][i] for i in range(len(TFT_initial["setData"])) for j in ["augments", "champions", "items", "mutator", "name", "number", "traits"]) and all(isinstance(TFT_initial["setData"][i][j], list) for i in range(len(TFT_initial["setData"])) for j in ["augments", "champions", "items", "traits"]) and all(isinstance(TFT_initial["setData"][i][j], str) for i in range(len(TFT_initial["setData"])) for j in ["mutator", "name"]) and all(isinstance(TFT_initial["setData"][i][j], int) for i in range(len(TFT_initial["setData"])) for j in ["number"]) and all(map(lambda x: isinstance(x, str), TFT_initial["setData"][i][j]) for i in range(len(TFT_initial["setData"])) for j in ["augments", "items"]) and all(map(lambda x: isinstance(x, dict), TFT_initial["setData"][i]["champions"]) for i in range(len(TFT_initial["setData"]))) and all(k in TFT_initial["setData"][i]["champions"][j] for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in ["ability", "apiName", "characterName", "cost", "icon", "name", "role", "squareIcon", "stats", "tileIcon", "traits"]) and all(isinstance(TFT_initial["setData"][i]["champions"][j][k], dict) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in ["ability", "stats"]) and all(isinstance(TFT_initial["setData"][i]["champions"][j][k], str) or TFT_initial["setData"][i]["champions"][j][k] == None for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in ["apiName", "characterName", "icon", "name", "squareIcon", "tileIcon"]) and all(isinstance(TFT_initial["setData"][i]["champions"][j][k], int) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in ["cost"]) and all(k in TFT_initial["setData"][i]["champions"][j]["ability"] for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in ["desc", "icon", "name", "variables"]) and all(isinstance(TFT_initial["setData"][i]["champions"][j]["ability"][k], str) or TFT_initial["setData"][i]["champions"][j]["ability"][k] == None for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in ["desc", "icon", "name"]) and all(isinstance(TFT_initial["setData"][i]["champions"][j]["ability"][k], list) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in ["variables"]) and all(map(lambda x: isinstance(x, dict), TFT_initial["setData"][i]["champions"][j]["ability"]["variables"]) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"]))) and all(l in TFT_initial["setData"][i]["champions"][j]["ability"]["variables"][k] for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in range(len(TFT_initial["setData"][i]["champions"][j]["ability"]["variables"])) for l in ["name", "value"]) and all(isinstance(TFT_initial["setData"][i]["champions"][j]["ability"]["variables"][k][l], str) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in range(len(TFT_initial["setData"][i]["champions"][j]["ability"]["variables"])) for l in ["name"]) and all(isinstance(TFT_initial["setData"][i]["champions"][j]["ability"]["variables"][k][l], list) or TFT_initial["setData"][i]["champions"][j]["ability"]["variables"][k][l] == None for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["champions"])) for k in range(len(TFT_initial["setData"][i]["champions"][j]["ability"]["variables"])) for l in ["value"]) and all(k in TFT_initial["setData"][i]["traits"][j] for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["traits"])) for k in ["apiName", "desc", "effects", "icon", "name"]) and all(isinstance(TFT_initial["setData"][i]["traits"][j][k], str) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["traits"])) for k in ["apiName", "desc", "icon", "name"]) and all(isinstance(TFT_initial["setData"][i]["traits"][j][k], list) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["traits"])) for k in ["effects"]) and all(map(lambda x: isinstance(x, dict), TFT_initial["setData"][i]["traits"][j]["effects"]) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["traits"]))) and all(l in TFT_initial["setData"][i]["traits"][j]["effects"][k] for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["traits"])) for k in range(len(TFT_initial["setData"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style", "variables"]) and all(isinstance(TFT_initial["setData"][i]["traits"][j]["effects"][k][l], int) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["traits"])) for k in range(len(TFT_initial["setData"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style"]) and all(isinstance(TFT_initial["setData"][i]["traits"][j]["effects"][k][l], dict) for i in range(len(TFT_initial["setData"])) for j in range(len(TFT_initial["setData"][i]["traits"])) for k in range(len(TFT_initial["setData"][i]["traits"][j]["effects"])) for l in ["variables"]) and all(j in TFT_initial["sets"][i] for i in TFT_initial["sets"] for j in ["champions", "name", "traits"]) and all(isinstance(TFT_initial["sets"][i][j], list) for i in TFT_initial["sets"] for j in ["champions", "traits"]) and all(isinstance(TFT_initial["sets"][i][j], str) for i in TFT_initial["sets"] for j in ["name"]) and all(k in TFT_initial["sets"][i]["champions"][j] for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in ["ability", "apiName", "characterName", "cost", "icon", "name", "role", "squareIcon", "stats", "tileIcon", "traits"]) and all(isinstance(TFT_initial["sets"][i]["champions"][j][k], dict) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in ["ability", "stats"]) and all(isinstance(TFT_initial["sets"][i]["champions"][j][k], str) or TFT_initial["sets"][i]["champions"][j][k] == None for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in ["apiName", "characterName", "icon", "name", "squareIcon", "tileIcon"]) and all(isinstance(TFT_initial["sets"][i]["champions"][j][k], int) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in ["cost"]) and all(k in TFT_initial["sets"][i]["champions"][j]["ability"] for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in ["desc", "icon", "name", "variables"]) and all(isinstance(TFT_initial["sets"][i]["champions"][j]["ability"][k], str) or TFT_initial["sets"][i]["champions"][j]["ability"][k] == None for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in ["desc", "icon", "name"]) and all(isinstance(TFT_initial["sets"][i]["champions"][j]["ability"][k], list) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in ["variables"]) and all(map(lambda x: isinstance(x, dict), TFT_initial["sets"][i]["champions"][j]["ability"]["variables"]) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"]))) and all(l in TFT_initial["sets"][i]["champions"][j]["ability"]["variables"][k] for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in range(len(TFT_initial["sets"][i]["champions"][j]["ability"]["variables"])) for l in ["name", "value"]) and all(isinstance(TFT_initial["sets"][i]["champions"][j]["ability"]["variables"][k][l], str) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in range(len(TFT_initial["sets"][i]["champions"][j]["ability"]["variables"])) for l in ["name"]) and all(isinstance(TFT_initial["sets"][i]["champions"][j]["ability"]["variables"][k][l], list) or TFT_initial["sets"][i]["champions"][j]["ability"]["variables"][k][l] == None for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["champions"])) for k in range(len(TFT_initial["sets"][i]["champions"][j]["ability"]["variables"])) for l in ["value"]) and all(k in TFT_initial["sets"][i]["traits"][j] for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["traits"])) for k in ["apiName", "desc", "effects", "icon", "name"]) and all(isinstance(TFT_initial["sets"][i]["traits"][j][k], str) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["traits"])) for k in ["apiName", "desc", "icon", "name"]) and all(isinstance(TFT_initial["sets"][i]["traits"][j][k], list) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["traits"])) for k in ["effects"]) and all(map(lambda x: isinstance(x, dict), TFT_initial["sets"][i]["traits"][j]["effects"]) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["traits"]))) and all(l in TFT_initial["sets"][i]["traits"][j]["effects"][k] for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["traits"])) for k in range(len(TFT_initial["sets"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style", "variables"]) and all(isinstance(TFT_initial["sets"][i]["traits"][j]["effects"][k][l], int) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["traits"])) for k in range(len(TFT_initial["sets"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style"]) and all(isinstance(TFT_initial["sets"][i]["traits"][j]["effects"][k][l], dict) for i in TFT_initial["sets"] for j in range(len(TFT_initial["sets"][i]["traits"])) for k in range(len(TFT_initial["sets"][i]["traits"][j]["effects"])) for l in ["variables"]):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈基础数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT basic data archived in CommunityDragon database (%s)!" %(TFT_url, TFT_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的云顶之弈基础信息Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT basics Json data file path!' %(TFT_local, TFT_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有云顶之弈基础信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT basic information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈基础数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT basic data archived in CommunityDragon database (%s)!" %(TFT_url, TFT_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取云顶之弈英雄数据（The following code get TFT champion data）
                    try:
                        logPrint("正在加载云顶之弈棋子信息……\nLoading TFT champion information from CommunityDragon ...")
                        TFTChampion_initial = requests.get(TFTChampion_url) #TFTChampion存储云顶之弈的棋子信息（Variable `TFTChampion_initial` stores information of TFT champions）
                        if TFTChampion_initial.ok:
                            TFTChampion_initial = TFTChampion_initial.json()
                        else:
                            logPrint(TFTChampion_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('云顶之弈英雄信息获取超时！正在尝试离线加载数据……\nTFT champion information capture timeout! Trying loading offline data ...\n请输入云顶之弈英雄Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT champion Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTChampion_local_default, TFTChampion_local_default))
                        while True:
                            TFTChampion_local = logInput()
                            if TFTChampion_local == "":
                                TFTChampion_local = TFTChampion_local_default
                            elif TFTChampion_local[0] == "0":
                                logPrint("云顶之弈英雄信息获取失败！请检查系统网络状况和代理设置。\nTFT champion information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(TFTChampion_local, "r", encoding = "utf-8") as fp:
                                    TFTChampion_initial = json.load(fp)
                                if isinstance(TFTChampion_initial, list) and all(isinstance(TFTChampion_initial[i], dict) for i in range(len(TFTChampion_initial))) and all(TFTChampion_initial[i].get(j, 0) for i in range(len(TFTChampion_initial)) for j in ["name", "character_record"]):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈棋子数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT champion data archived in CommunityDragon database (%s)!" %(TFTChampion_url, TFTChampion_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的云顶之弈棋子Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT champion Json data file path!' %(TFTChampion_local, TFTChampion_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有云顶之弈英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT champion information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈棋子数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT champion data archived in CommunityDragon database (%s)!" %(TFTChampion_url, TFTChampion_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取云顶之弈装备数据（The following code get TFT item information）
                    try:
                        logPrint("正在加载云顶之弈装备信息……\nLoading TFT item information from CommunityDragon ...")
                        TFTItem_initial = requests.get(TFTItem_url) #TFTItem存储云顶之弈的装备信息（Variable `TFTItem_initial` stores information of TFT items）
                        if TFTItem_initial.ok:
                            TFTItem_initial = TFTItem_initial.json()
                        else:
                            logPrint(TFTItem_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('云顶之弈装备信息获取超时！正在尝试离线加载数据……\nTFT item information capture timeout! Trying loading offline data ...\n请输入云顶之弈装备Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT item Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTItem_local_default, TFTItem_local_default))
                        while True:
                            TFTItem_local = logInput()
                            if TFTItem_local == "":
                                TFTItem_local = TFTItem_local_default
                            elif TFTItem_local[0] == "0":
                                logPrint("云顶之弈装备信息获取失败！请检查系统网络状况和代理设置。\nTFT item information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(TFTItem_local, "r", encoding = "utf-8") as fp:
                                    TFTItem_initial = json.load(fp)
                                if isinstance(TFTItem_initial, list) and all(isinstance(TFTItem_initial[i], dict) for i in range(len(TFTItem_initial))) and (all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "loadoutsIcon"]) or all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "squareIconPath"])):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT item data archived in CommunityDragon database (%s)!" %(TFTItem_url, TFTItem_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的云顶之弈装备Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT item Json data file path!' %(TFTItem_local, TFTItem_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有云顶之弈装备信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT companion information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT item data archived in CommunityDragon database (%s)!" %(TFTItem_url, TFTItem_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取云顶之弈小小英雄数据（The following code get TFT companion data）
                    try:
                        logPrint("正在加载云顶之弈小小英雄信息……\nLoading companion information from CommunityDragon ...")
                        TFTCompanion_initial = requests.get(TFTCompanion_url) #TFTChampion存储云顶之弈的小小英雄信息（Variable `TFTChampion_initial` stores information of companions）
                        if TFTCompanion_initial.ok:
                            TFTCompanion_initial = TFTCompanion_initial.json()
                        else:
                            logPrint(TFTCompanion_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('云顶之弈小小英雄信息获取超时！正在尝试离线加载数据……\nTFT companion information capture timeout! Trying loading offline data ...\n请输入云顶之弈小小英雄Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT companion Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTCompanion_local_default, TFTCompanion_local_default))
                        while True:
                            TFTCompanion_local = logInput()
                            if TFTCompanion_local == "":
                                TFTCompanion_local = TFTCompanion_local_default
                            elif TFTCompanion_local[0] == "0":
                                logPrint("云顶之弈小小英雄信息获取失败！请检查系统网络状况和代理设置。\nTFT companion information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(TFTCompanion_local, "r", encoding = "utf-8") as fp:
                                    TFTCompanion_initial = json.load(fp)
                                if isinstance(TFTCompanion_initial, list) and all(isinstance(TFTCompanion_initial[i], dict) for i in range(len(TFTCompanion_initial))) and all(j in TFTCompanion_initial[i] for i in range(len(TFTCompanion_initial)) for j in ["contentId", "itemId", "name", "loadoutsIcon", "description", "level", "speciesName", "speciesId", "rarity", "rarityValue", "isDefault", "upgrades", "TFTOnly"]):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈小小英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT companion data archived in CommunityDragon database (%s)!" %(TFTCompanion_url, TFTCompanion_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的云顶之弈小小英雄Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT companion Json data file path!' %(TFTCompanion_local, TFTCompanion_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有云顶之弈小小英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT companion information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈小小英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT companion data archived in CommunityDragon database (%s)!" %(TFTCompanion_url, TFTCompanion_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取云顶之弈羁绊数据（The following code get TFT trait data）
                    try:
                        logPrint("正在加载云顶之弈羁绊信息……\nLoading TFT trait information from CommunityDragon ...")
                        TFTTrait_initial = requests.get(TFTTrait_url) #TFTTrait存储云顶之弈的羁绊信息（Variable `TFTTrait_initial` stores information of TFT traits）
                        if TFTTrait_initial.ok:
                            TFTTrait_initial = TFTTrait_initial.json()
                        else:
                            logPrint(TFTTrait_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('云顶之弈羁绊信息获取超时！正在尝试离线加载数据……\nTFT trait information capture timeout! Trying loading offline data ...\n请输入云顶之弈羁绊Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT trait Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTTrait_local_default, TFTTrait_local_default))
                        while True:
                            TFTTrait_local = logInput()
                            if TFTTrait_local == "":
                                TFTTrait_local = TFTTrait_local_default
                            elif TFTTrait_local[0] == "0":
                                logPrint("云顶之弈羁绊信息获取失败！请检查系统网络状况和代理设置。\nTFT trait information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(TFTTrait_local, "r", encoding = "utf-8") as fp:
                                    TFTTrait_initial = json.load(fp)
                                if isinstance(TFTTrait_initial, list) and all(isinstance(TFTTrait_initial[i], dict) for i in range(len(TFTTrait_initial))) and all(j in TFTTrait_initial[i] for i in range(len(TFTTrait_initial)) for j in ["display_name", "trait_id", "set", "icon_path", "tooltip_text", "innate_trait_sets", "conditional_trait_sets"]):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈羁绊数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT trait data archived in CommunityDragon database (%s)!" %(TFTTrait_url, TFTTrait_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的云顶之弈羁绊Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT trait Json data file path!' %(TFTTrait_local, TFTTrait_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有云顶之弈小小英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT companion information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈羁绊数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT trait data archived in CommunityDragon database (%s)!" %(TFTTrait_url, TFTTrait_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    #下面获取斗魂竞技场强化符文数据（The following code get Arena augment data）
                    try:
                        logPrint("正在加载斗魂竞技场强化符文信息……\nLoading Arena augment information from CommunityDragon ...")
                        CherryAugment_initial = requests.get(CherryAugment_url) #Arena存储斗魂竞技场的强化符文信息（Variable `CherryAugment_initial` stores information of Arena augments）
                        if CherryAugment_initial.ok:
                            CherryAugment_initial = CherryAugment_initial.json()
                            break
                        else:
                            logPrint(CherryAugment_initial)
                            logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                            switch_language = True
                            break
                    except requests.exceptions.RequestException:
                        logPrint('斗魂竞技场强化符文信息获取超时！正在尝试离线加载数据……\nArena augment information capture timeout! Trying loading offline data ...\n请输入斗魂竞技场强化符文Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the Arena augment Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(CherryAugment_local_default, CherryAugment_local_default))
                        while True:
                            CherryAugment_local = logInput()
                            if CherryAugment_local == "":
                                CherryAugment_local = CherryAugment_local_default
                            elif CherryAugment_local[0] == "0":
                                logPrint("斗魂竞技场强化符文信息获取失败！请检查系统网络状况和代理设置。\nArena augment information capture failure! Please check the system network condition and agent configuration.")
                                time.sleep(5)
                                return 1
                            else:
                                switch_prepare_mode = True
                                break
                            try:
                                with open(CherryAugment_local, "r", encoding = "utf-8") as fp:
                                    CherryAugment_initial = json.load(fp)
                                if isinstance(CherryAugment_initial, list) and all(isinstance(CherryAugment_initial[i], dict) for i in range(len(CherryAugment_initial))) and all(j in CherryAugment_initial[i] for i in range(len(CherryAugment_initial)) for j in ["id", "nameTRA", "augmentSmallIconPath", "rarity"]) and all(isinstance(CherryAugment_initial[i]["id"], int) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["nameTRA"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["augmentSmallIconPath"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["rarity"], str) for i in range(len(CherryAugment_initial))):
                                    break
                                else:
                                    logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的斗魂竞技场强化符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the Arena augment data archived in CommunityDragon database (%s)!" %(CherryAugment_url, CherryAugment_url))
                                    continue
                            except FileNotFoundError:
                                logPrint('未找到文件“%s”！请输入正确的斗魂竞技场强化符文Json数据文件路径！\nFile "%s" NOT found! Please input a correct Arena augment Json data file path!' %(CherryAugment_local, CherryAugment_local))
                                continue
                            except OSError:
                                logPrint("数据文件名不合法！请输入含有斗魂竞技场强化符文信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with Arena augment information.")
                                continue
                            except json.decoder.JSONDecodeError:
                                logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的斗魂竞技场强化符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the Arena augment data archived in CommunityDragon database (%s)!" %(CherryAugment_url, CherryAugment_url))
                                continue
                    if switch_prepare_mode:
                        prepareMode = ""
                        continue
                    break
                else:
                    switch_prepare_mode = False
                    logPrint('请在浏览器中打开以下网页，待加载完成后按Ctrl + S保存网页json文件至同目录的“离线数据（Offline Data）”文件夹下，并根据括号内的提示放置和命名文件。\nPlease open the following URLs in a browser, then press Ctrl + S to save the online json files into the folder "离线数据（Offline Data）" under the same directory after the website finishes loading and organize and rename the downloaded files according to the hints in the circle brackets.\n版本信息（%s）： %s\n召唤师技能（%s）： %s\n英雄（%s）： %s\n英雄联盟装备（%s）： %s\n召唤师图标（%s）： %s\n基石符文（%s）： %s\n符文系（%s）： %s\n云顶之弈基础信息（%s）： %s\n云顶之弈棋子（%s）： %s\n云顶之弈装备（%s）： %s\n云顶之弈小小英雄（%s）： %s\n云顶之弈羁绊（%s）： %s\n斗魂竞技场强化符文（%s）： %s' %(patches_local_default[19:], patches_url, spell_local_default[19:], spell_url, LoLChampion_local_default[19:], LoLChampion_url, LoLItem_local_default[19:], LoLItem_url, summonerIcon_local_default[19:], summonerIcon_url, perk_local_default[19:], perk_url, perkstyle_local_default[19:], perkstyle_url, TFT_local_default[19:], TFT_url, TFTChampion_local_default[19:], TFTChampion_url, TFTItem_local_default[19:], TFTItem_url, TFTCompanion_local_default[19:], TFTCompanion_url, TFTTrait_local_default[19:], TFTTrait_url, CherryAugment_local_default[19:], CherryAugment_url))
                    offline_files_loaded = {"patch": False, "spell": False, "LoLChampion": False, "LoLItem": False, "summonerIcon": False, "perk": False, "perkstyle": False, "TFT": False, "TFTChampion": False, "TFTItem": False, "TFTCompanion": False, "TFTTrait": False, "CherryAugment": False}
                    offline_files = {"patch": {"file": patches_local_default, "URL": patches_url, "content": "版本信息"}, "spell": {"file": spell_local_default, "URL": spell_url, "content": "召唤师技能"}, "LoLChampion": {"file": LoLChampion_local_default, "URL": LoLChampion_url, "content": "英雄"}, "LoLItem": {"file": LoLItem_local_default, "URL": LoLItem_url, "content": "英雄联盟装备"}, "summonerIcon": {"file": summonerIcon_local_default, "URL": summonerIcon_url, "content": "召唤师图标"}, "perk": {"file": perk_local_default, "URL": perk_url, "content": "基石符文"}, "perkstyle": {"file": perkstyle_local_default, "URL": perkstyle_url, "content": "符文系"}, "TFT": {"file": TFT_local_default, "URL": TFT_url, "content": "云顶之弈基础信息"}, "TFTChampion": {"file": TFTChampion_local_default, "URL": TFTChampion_url, "content": "云顶之弈英雄"}, "TFTItem": {"file": TFTItem_local_default, "URL": TFTItem_url, "content": "云顶之弈装备"}, "TFTCompanion": {"file": TFTCompanion_local_default, "URL": TFTCompanion_url, "content": "云顶之弈小小英雄"}, "TFTTrait": {"file": TFTTrait_local_default, "URL": TFTTrait_url, "content": "云顶之弈羁绊"}, "CherryAugment": {"file": CherryAugment_local_default, "URL": CherryAugment_url, "content": "斗魂竞技场强化符文"}}
                    logPrint('请按任意键以加载离线数据。输入“1”以转为在线模式。输入“0”以退出程序。\nPlease input anything to load offline data. Input "1" to switch to online mode. Submit "0" to exit.')
                    while any(not i for i in offline_files_loaded.values()):
                        offline_files_notfound = {"patch": False, "spell": False, "LoLChampion": False, "LoLItem": False, "summonerIcon": False, "perk": False, "perkstyle": False, "TFT": False, "TFTChampion": False, "TFTItem": False, "TFTCompanion": False, "TFTTrait": False, "CherryAugment": False}
                        offline_files_formaterror = {"patch": False, "spell": False, "LoLChampion": False, "LoLItem": False, "summonerIcon": False, "perk": False, "perkstyle": False, "TFT": False, "TFTChampion": False, "TFTItem": False, "TFTCompanion": False, "TFTTrait": False, "CherryAugment": False}
                        prepareMode = logInput()
                        if prepareMode != "" and prepareMode[0] == "1":
                            switch_prepare_mode = True
                            break
                        if prepareMode != "" and prepareMode[0] == "0":
                            return 0
                        #下面获取版本信息（The following code get the patch data）
                        if not offline_files_loaded["patch"]:
                            try:
                                with open(patches_local_default, "r", encoding = "utf-8") as fp:
                                    patches_initial = json.load(fp)
                                if not (isinstance(patches_initial, list) and patches_initial[-1] == "lolpatch_3.7"): #之所以将patches的最后一个元素作为判断版本文件数据格式合法的依据，是因为按照这样的逻辑，代码在一般情况下就不需要频繁变动（The reason why I use the last element of the variable `patches_initial` as the judgment whether the patch file data format is legal is, that under this logic, the code won't need further adjustment as the update goes on）
                                    offline_files_formaterror["patch"] = True
                            except FileNotFoundError:
                                offline_files_notfound["patch"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["patch"] = True
                            else:
                                if not offline_files_formaterror["patch"]:
                                    offline_files_loaded["patch"] = True
                                    latest_patch = patches_initial[0]
                                    patches_dict = {}
                                    smallPatches = []
                                    bigPatches = []
                                    for patch in patches_initial:
                                        if not patch.startswith("lolpatch"):
                                            patch_split = patch.split(".")
                                            smallPatch = ".".join(patch_split[:3])
                                            smallPatches.append(smallPatch)
                                            bigPatch = ".".join(patch_split[:2])
                                            bigPatches.append(bigPatch)
                                            patches_dict[bigPatch] = []
                                    for i in range(len(bigPatches)):
                                        patches_dict[bigPatches[i]].append(smallPatches[i])
                        #下面获取召唤师技能数据（The following code get summoenr spell data）
                        if not offline_files_loaded["spell"]:
                            try:
                                with open(spell_local_default, "r", encoding = "utf-8") as fp:
                                    spell_initial = json.load(fp)
                                if not(isinstance(spell_initial, list) and all(i in spell_initial[j] for i in ["id", "name", "description", "summonerLevel", "cooldown", "gameModes", "iconPath"] for j in range(len(spell_initial)))):
                                    offline_files_formaterror["spell"] = True
                            except FileNotFoundError:
                                offline_files_notfound["spell"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["spell"] = True
                            else:
                                if not offline_files_formaterror["spell"]:
                                    offline_files_loaded["spell"] = True
                        #下面获取英雄信息（The following code get LoL champion data）
                        if not offline_files_loaded["LoLChampion"]:
                            try:
                                with open(LoLChampion_local_default, "r", encoding = "utf-8") as fp:
                                    LoLChampion_initial = json.load(fp)
                                if not(isinstance(LoLChampion_initial, list) and all(isinstance(LoLChampion_initial[i], dict) for i in range(len(LoLChampion_initial))) and all(j in LoLChampion_initial[i] for i in range(len(LoLChampion_initial)) for j in ["id", "name", "alias", "squarePortraitPath", "roles"]) and all(isinstance(LoLChampion_initial[i]["id"], int) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["name"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["alias"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["squarePortraitPath"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["roles"], list) for i in range(len(LoLChampion_initial)))):
                                    offline_files_formaterror["LoLChampion"] = True
                            except FileNotFoundError:
                                offline_files_notfound["LoLChampion"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["LoLChampion"] = True
                            else:
                                if not offline_files_formaterror["LoLChampion"]:
                                    offline_files_loaded["LoLChampion"] = True
                        #下面获取英雄联盟装备信息（The following code get LoL item data）
                        if not offline_files_loaded["LoLItem"]:
                            try:
                                with open(LoLItem_local_default, "r", encoding = "utf-8") as fp:
                                    LoLItem_initial = json.load(fp)
                                if not(isinstance(LoLItem_initial, list) and all(i in LoLItem_initial[j] for i in ["id", "name", "description", "active", "inStore", "from", "to", "categories", "maxStacks", "requiredChampion", "requiredAlly", "requiredBuffCurrencyName", "requiredBuffCurrencyCost", "specialRecipe", "isEnchantment", "price", "priceTotal", "iconPath"] for j in range(len(LoLItem_initial)))):
                                    offline_files_formaterror["LoLItem"] = True
                            except FileNotFoundError:
                                offline_files_notfound["LoLItem"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["LoLItem"] = True
                            else:
                                if not offline_files_formaterror["LoLItem"]:
                                    offline_files_loaded["LoLItem"] = True
                        #下面获取召唤师图标信息（The following code get summoner icon data）
                        if not offline_files_loaded["summonerIcon"]:
                            try:
                                with open(summonerIcon_local_default, "r", encoding = "utf-8") as fp:
                                    summonerIcon_initial = json.load(fp)
                                if not(isinstance(summonerIcon_initial, list) and all(map(lambda x: isinstance(x, dict), summonerIcon_initial)) and all(i in j for i in ["id", "title", "yearReleased", "isLegacy", "descriptions", "rarities", "disabledRegions"] for j in summonerIcon_initial)):
                                    offline_files_formaterror["summonerIcon"] = True
                            except FileNotFoundError:
                                offline_files_notfound["summonerIcon"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["summonerIcon"] = True
                            else:
                                if not offline_files_formaterror["summonerIcon"]:
                                    offline_files_loaded["summonerIcon"] = True
                        #下面获取基石符文信息（The following code get perk data）
                        if not offline_files_loaded["perk"]:
                            try:
                                with open(perk_local_default, "r", encoding = "utf-8") as fp:
                                    perk_initial = json.load(fp)
                                if not(isinstance(perk_initial, list) and all(i in perk_initial[j] for i in ["id", "name", "majorChangePatchVersion", "tooltip", "shortDesc", "longDesc", "recommendationDescriptor", "iconPath", "endOfGameStatDescs", "recommendationDescriptorAttributes"] for j in range(len(perk_initial)))):
                                    offline_files_formaterror["perk"] = True
                            except FileNotFoundError:
                                offline_files_notfound["perk"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["perk"] = True
                            else:
                                if not offline_files_formaterror["perk"]:
                                    offline_files_loaded["perk"] = True
                        #下面获取符文系信息（The following code get perkstyle data）
                        if not offline_files_loaded["perkstyle"]:
                            try:
                                with open(perkstyle_local_default, "r", encoding = "utf-8") as fp:
                                    perkstyle_initial = json.load(fp)
                                if not(isinstance(perkstyle_initial, dict) and all(perkstyle_initial.get(i, 0) for i in ["schemaVersion", "styles"]) and isinstance(perkstyle_initial["styles"], list) and all(j in perkstyle_initial["styles"][i] for i in range(len(perkstyle_initial["styles"])) for j in ["id", "name", "tooltip", "iconPath", "assetMap", "isAdvanced", "allowedSubStyles", "subStyleBonus", "slots", "defaultPageName", "defaultSubStyle", "defaultPerks", "defaultPerksWhenSplashed", "defaultStatModsPerSubStyle"])):
                                    offline_files_formaterror["perkstyle"] = True
                            except FileNotFoundError:
                                offline_files_notfound["perkstyle"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["perkstyle"] = True
                            else:
                                if not offline_files_formaterror["perkstyle"]:
                                    offline_files_loaded["perkstyle"] = True
                        #下面获取云顶之弈强化符文数据（The following code get TFT augment data）
                        if not offline_files_loaded["TFT"]:
                            try:
                                with open(TFT_local_default, "r", encoding = "utf-8") as fp:
                                    TFT_initial = json.load(fp)
                                if not(isinstance(TFT_initial, dict) and all(i in TFT_initial for i in ["items", "setData", "sets"])):
                                    offline_files_formaterror["TFT"] = True
                            except FileNotFoundError:
                                offline_files_notfound["TFT"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["TFT"] = True
                            else:
                                if not offline_files_formaterror["TFT"]:
                                    offline_files_loaded["TFT"] = True
                        #下面获取云顶之弈英雄数据（The following code get TFT champion data）
                        if not offline_files_loaded["TFTChampion"]:
                            try:
                                with open(TFTChampion_local_default, "r", encoding = "utf-8") as fp:
                                    TFTChampion_initial = json.load(fp)
                                if not(isinstance(TFTChampion_initial, list) and all(isinstance(TFTChampion_initial[i], dict) for i in range(len(TFTChampion_initial))) and all(TFTChampion_initial[i].get(j, 0) for i in range(len(TFTChampion_initial)) for j in ["name", "character_record"])):
                                    offline_files_formaterror["TFTChampion"] = True
                            except FileNotFoundError:
                                offline_files_notfound["TFTChampion"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["TFTChampion"] = True
                            else:
                                if not offline_files_formaterror["TFTChampion"]:
                                    offline_files_loaded["TFTChampion"] = True
                        #下面获取云顶之弈装备数据（The following code get TFT item information）
                        if not offline_files_loaded["TFTItem"]:
                            try:
                                with open(TFTItem_local_default, "r", encoding = "utf-8") as fp:
                                    TFTItem_initial = json.load(fp)
                                if not(isinstance(TFTItem_initial, list) and all(isinstance(TFTItem_initial[i], dict) for i in range(len(TFTItem_initial))) and (all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "loadoutsIcon"]) or all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "squareIconPath"]))):
                                    offline_files_formaterror["TFTItem"] = True
                            except FileNotFoundError:
                                offline_files_notfound["TFTItem"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["TFTItem"] = True
                            else:
                                if not offline_files_formaterror["TFTItem"]:
                                    offline_files_loaded["TFTItem"] = True
                        #下面获取云顶之弈小小英雄数据（The following code get TFT companion data）
                        if not offline_files_loaded["TFTCompanion"]:
                            try:
                                with open(TFTCompanion_local_default, "r", encoding = "utf-8") as fp:
                                    TFTCompanion_initial = json.load(fp)
                                if not(isinstance(TFTCompanion_initial, list) and all(isinstance(TFTCompanion_initial[i], dict) for i in range(len(TFTCompanion_initial))) and all(j in TFTCompanion_initial[i] for i in range(len(TFTCompanion_initial)) for j in ["contentId", "itemId", "name", "loadoutsIcon", "description", "level", "speciesName", "speciesId", "rarity", "rarityValue", "isDefault", "upgrades", "TFTOnly"])):
                                    offline_files_formaterror["TFTCompanion"] = True
                            except FileNotFoundError:
                                offline_files_notfound["TFTCompanion"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["TFTCompanion"] = True
                            else:
                                if not offline_files_formaterror["TFTCompanion"]:
                                    offline_files_loaded["TFTCompanion"] = True
                        #下面获取云顶之弈羁绊数据（The following code get TFT trait data）
                        if not offline_files_loaded["TFTTrait"]:
                            try:
                                with open(TFTTrait_local_default, "r", encoding = "utf-8") as fp:
                                    TFTTrait_initial = json.load(fp)
                                if not(isinstance(TFTTrait_initial, list) and all(isinstance(TFTTrait_initial[i], dict) for i in range(len(TFTTrait_initial))) and all(j in TFTTrait_initial[i] for i in range(len(TFTTrait_initial)) for j in ["display_name", "trait_id", "set", "icon_path", "tooltip_text", "innate_trait_sets", "conditional_trait_sets"])):
                                    offline_files_formaterror["TFTTrait"] = True
                            except FileNotFoundError:
                                offline_files_notfound["TFTTrait"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["TFTTrait"] = True
                            else:
                                if not offline_files_formaterror["TFTTrait"]:
                                    offline_files_loaded["TFTTrait"] = True
                        #下面获取斗魂竞技场强化符文数据（The following code get Arena augment data）
                        if not offline_files_loaded["CherryAugment"]:
                            try:
                                with open(CherryAugment_local_default, "r", encoding = "utf-8") as fp:
                                    CherryAugment_initial = json.load(fp)
                                if not(isinstance(CherryAugment_initial, list) and all(isinstance(CherryAugment_initial[i], dict) for i in range(len(CherryAugment_initial))) and all(j in CherryAugment_initial[i] for i in range(len(CherryAugment_initial)) for j in ["id", "nameTRA", "augmentSmallIconPath", "rarity"]) and all(isinstance(CherryAugment_initial[i]["id"], int) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["nameTRA"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["augmentSmallIconPath"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["rarity"], str) for i in range(len(CherryAugment_initial)))):
                                    offline_files_formaterror["CherryAugment"] = True
                            except FileNotFoundError:
                                offline_files_notfound["CherryAugment"] = True
                            except json.decoder.JSONDecodeError:
                                offline_files_formaterror["CherryAugment"] = True
                            else:
                                if not offline_files_formaterror["CherryAugment"]:
                                    offline_files_loaded["CherryAugment"] = True
                        #下面总结离线数据加载情况（The following code conclude the result of loading offline data）
                        unloaded_offline_files = []
                        notfound_offline_files = []
                        formaterror_offline_files = []
                        if any(offline_files_notfound.values()):
                            for i in offline_files_notfound:
                                if offline_files_notfound[i]:
                                    notfound_offline_files.append(i)
                                    unloaded_offline_files.append(i)
                            logPrint("以下信息文件不存在：\nNot existing file(s):")
                            for i in notfound_offline_files:
                                logPrint(offline_files[i]["file"] + "\t" + offline_files[i]["content"] + "\t" + offline_files[i]["URL"])
                        if any(offline_files_formaterror.values()):
                            for i in offline_files_formaterror:
                                if offline_files_formaterror[i]:
                                    formaterror_offline_files.append(i)
                                    unloaded_offline_files.append(i)
                            logPrint("以下信息文件格式错误：\nFormatError file(s):")
                            for i in formaterror_offline_files:
                                logPrint(offline_files[i]["file"] + "\t" + offline_files[i]["content"] + "\t" + offline_files[i]["URL"])
                        if any(not i for i in offline_files_loaded.values()):
                            logPrint('请按任意键以加载离线数据。输入“1”以转为在线模式。输入“0”以退出程序。\nPlease input anything to load offline data. Input "1" to switch to online mode. Submit "0" to exit.')
                    if switch_prepare_mode:
                        continue
                    else:
                        break
            if switch_language:
                continue
            break
        elif language_option[0] == "0":
            return 2
        else:
            logPrint("语言选项输入错误！请重新输入：\nERROR input of language option! Please try again:")
    #首先准备一些数据（First, prepare some data）
    #准备自己的召唤师数据（Prepare the information of the user himself/herself）
    current_info = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    current_infos = [current_info] #检测模式的小号模式中存在多个自己（There're many selves in Smurf Mode of Detect Mode）
    ##准备游戏模式数据（Prepare game mode data）
    gamemode = await (await connection.request("GET", "/lol-game-queues/v1/queues")).json()
    gamemodes = {-1: {"name": "自定义", "gameMode": "CUSTOM", "category": "CUSTOM", "description": "", "type": "CUSTOM"}, 0: {"name": "自定义", "gameMode": "CUSTOM", "category": "CUSTOM", "description": "", "type": "CUSTOM"}} #在对局记录中，自定义对局的队列序号是0；在邀请中，自定义房间的队列序号是-1（A custom game's queueId is 0 in the match history. A custom lobby's queueId is -1 in an invitation）
    for gamemode_iter in gamemode:
        gamemode_id = gamemode_iter["id"]
        gamemodes_iter = {}
        gamemodes_iter["name"] = gamemode_iter["name"]
        gamemodes_iter["gameMode"] = gamemode_iter["gameMode"]
        gamemodes_iter["category"] = gamemode_iter["category"] #用于生成模式对对局类别的判断（Used for judgments on the game category in Generate Mode）
        gamemodes_iter["description"] = gamemode_iter["description"]
        gamemodes_iter["type"] = gamemode_iter["type"] #用于检测邀请信息中的玩家时对对局类型的描述（Used for descriptions about the game type when the program is detecting the players in invitations）
        gamemodes[gamemode_id] = gamemodes_iter
    queues = {queue["id"]: queue for queue in gamemode}
    ##准备召唤师技能数据（Prepare summoner spell data）
    spells_initial = {} #spells为嵌套字典，键为召唤师技能序号，值为召唤师技能信息字典。一个键值对的示例如右：（Variable `spells` is a nested dictionary, whose keys are spellIds and values are spell information dictionaries. An example of the key-value pairs is shown as follows: ）{1: {"name": "净化", "description": "移除身上的所有限制效果（压制效果和击飞效果除外）和召唤师技能的减益效果，并且若在接下来的3秒里再次被施加限制效果时，新效果的持续时间会减少65%。", "summonerLevel": 9, "cooldown": 210, "gameModes": ["URF", "CLASSIC", "ARSR", "ARAM", "ULTBOOK", "WIPMODEWIP", "TUTORIAL", "DOOMBOTSTEEMO", "PRACTICETOOL", "FIRSTBLOOD", "NEXUSBLITZ", "PROJECT", "ONEFORALL"], "iconPath": "/lol-game-data/assets/DATA/Spells/Icons2D/Summoner_boost.png"}}
    for spell_iter in spell_initial:
        spell_id = spell_iter["id"]
        spells_initial[spell_id] = spell_iter
    ##准备英雄数据（Prepare champion data）
    LoLChampions_initial = {} #LoLChampions为嵌套字典，键为英雄序号，值为英雄信息字典。一个键值对的示例如右：（Variable `LoLItems` is a nested dictionary, whose keys are itemIds and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{1: {"name": "黑暗之女", "alias": "Annie", "squarePortraitPath": "/lol-game-data/assets/v1/champion-icons/1.png", "roles": ["mage", "support"]}}
    for LoLChampion_iter in LoLChampion_initial:
        LoLChampion_id = LoLChampion_iter["id"]
        LoLChampions_initial[LoLChampion_id] = LoLChampion_iter
    ##准备英雄联盟装备数据（Prapare LoL item data）
    LoLItems_initial = {} #LoLItems为嵌套字典，键为装备序号，值为装备信息字典。一个键值对的示例如右：（Variable `LoLItems` is a nested dictionary, whose keys are itemIds and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{1001: {"name": "鞋子", "description": "<mainText><stats><attention>25</attention>移动速度</stats></mainText><br>", "active": False, "inStore": True, "from": [], "to": [3111, 3006, 3005, 3009, 3020, 3047, 3117, 3158], "categories": ["Boots"], "maxStacks": 1, "requiredChampion": "", "requiredAlly": "", "requiredBuffCurrencyName": "", "requiredBuffCurrencyCost": 0, "specialRecipe": 0, "isEnchantment": False, "price": 300, "priceTotal": 300, "iconPath": "/lol-game-data/assets/ASSETS/Items/Icons2D/1001_Class_T1_BootsofSpeed.png"}}
    for LoLItem_iter in LoLItem_initial:
        LoLItem_id = int(LoLItem_iter["id"])
        LoLItems_initial[LoLItem_id] = LoLItem_iter
    ##准备召唤师图标数据（Prepare summoner icon data）
    summonerIcons_initial = {} #summonerIcons为嵌套字典，键为装备序号，值为装备信息字典。一个键值对的示例如右：（Variable `summonerIcons` is a nested dictionary, whose keys are itemIds and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{0: {"id":0,"title":"可爱凯尔 图标","yearReleased":2009,"isLegacy":false,"imagePath":"/lol-game-data/assets/v1/profile-icons/0.jpg","descriptions":[{"region":"riot","description":" "}],"rarities":[{"region":"riot","rarity":0}],"disabledRegions":[]},{"id":1000,"title":"2016 LCL Hard Random","yearReleased":2016,"isLegacy":false,"imagePath":"/lol-game-data/assets/v1/profile-icons/1000.jpg","esportsTeam":"Hard Random","esportsRegion":"RU","esportsEvent":"英雄联盟欧陆联赛 LCL","descriptions":[{"region":"riot","description":" "}],"rarities":[{"region":"riot","rarity":0}],"disabledRegions":[]}}
    for summonerIcon_iter in summonerIcon_initial:
        summonerIcon_id = int(summonerIcon_iter["id"])
        summonerIcons_initial[summonerIcon_id] = summonerIcon_iter
    ##准备符文数据（Prepare runes data）
    perks_initial = {} #perks为嵌套字典，键为符文序号，值为符文信息字典。一个键值对的示例如右：（Variable `perks` is a nested dictionary, whose keys are perkIds and values are perk information dictionaries. An example of the key-value pairs is shown as follows: ）{8369: {"name": "先攻", "majorChangePatchVersion": "11.23", "tooltip": "在进入与英雄战斗的@GraceWindow.2@秒内，对一名敌方英雄进行的攻击或技能将提供@GoldProcBonus@金币和<b>先攻</b>效果，持续@Duration@秒，来使你对英雄们造成<truedamage>@DamageAmp*100@%</truedamage>额外<truedamage>伤害</truedamage>，并提供<gold>{{ Item_Melee_Ranged_Split }}</gold>该额外伤害值的<gold>金币</gold>。<br><br>冷却时间：<scaleLevel>@Cooldown@</scaleLevel>秒<br><hr><br>已造成的伤害：@f1@<br>已提供的金币：@f2@", "shortDesc": "在你率先发起与英雄的战斗时，造成8%额外伤害，持续3秒，并基于该额外伤害提供金币。", "longDesc": "在进入与英雄战斗的0.25秒内，对一名敌方英雄进行的攻击或技能将提供5金币和<b>先攻</b>效果，持续3秒，来使你对英雄们造成<truedamage>8%</truedamage>额外<truedamage>伤害</truedamage>，并提供<gold>100% (远程英雄为70%)</gold>该额外伤害值的<gold>金币</gold>。<br><br>冷却时间：<scaleLevel>25 ~ 15</scaleLevel>秒", "recommendationDescriptor": "真实伤害，金币收入", "iconPath": "/lol-game-data/assets/v1/perk-images/Styles/Inspiration/FirstStrike/FirstStrike.png", "endOfGameStatDescs": ["已造成的伤害：@eogvar1@", "已提供的金币：@eogvar2@"], "recommendationDescriptorAttributes": {}}}
    for perk_iter in perk_initial:
        perk_id = perk_iter["id"]
        perks_initial[perk_id] = perk_iter
    ##准备符文系数据（Prepare perkstyle data）
    perkstyles_initial = {} #perkstyles为嵌套字典，键为符文系序号，值为符文系信息字典。一个键值对的示例如右：（Variable `perkstyles` is a nested dictionary, whose keys are perkstyle ids and values are perkstyle information dictionaries. An example of the key-value pairs is as follows: ）{8400: {"name": "坚决", "tooltip": "耐久和控制", "iconPath": "/lol-game-data/assets/v1/perk-images/Styles/7204_Resolve.png", "assetMap": {"p8400_s0_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k0.jpg", "p8400_s0_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k8437.jpg", "p8400_s0_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k8439.jpg", "p8400_s0_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k8465.jpg", "p8400_s8000_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k0.jpg", "p8400_s8000_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k8437.jpg", "p8400_s8000_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k8439.jpg", "p8400_s8000_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k8465.jpg", "p8400_s8100_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k0.jpg", "p8400_s8100_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k8437.jpg", "p8400_s8100_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k8439.jpg", "p8400_s8100_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k8465.jpg", "p8400_s8200_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k0.jpg", "p8400_s8200_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k8437.jpg", "p8400_s8200_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k8439.jpg", "p8400_s8200_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k8465.jpg", "p8400_s8300_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k0.jpg", "p8400_s8300_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k8437.jpg", "p8400_s8300_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k8439.jpg", "p8400_s8300_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k8465.jpg", "svg_icon": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/resolve_icon.svg", "svg_icon_16": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/resolve_icon_16.svg"}, "isAdvanced": False, "allowedSubStyles": [8000, 8100, 8200, 8300], "subStyleBonus": [{"styleId": 8000, "perkId": 8414}, {"styleId": 8100, "perkId": 8454}, {"styleId": 8200, "perkId": 8415}, {"styleId": 8300, "perkId": 8416}], "slots": [{"type": "kKeyStone", "slotLabel": "", "perks": [8437, 8439, 8465]}, {"type": "kMixedRegularSplashable", "slotLabel": "蛮力", "perks": [8446, 8463, 8401]}, {"type": "kMixedRegularSplashable", "slotLabel": "抵抗", "perks": [8429, 8444, 8473]}, {"type": "kMixedRegularSplashable", "slotLabel": "生机", "perks": [8451, 8453, 8242]}, {"type": "kStatMod", "slotLabel": "进攻", "perks": [5008, 5005, 5007]}, {"type": "kStatMod", "slotLabel": "灵活", "perks": [5008, 5002, 5003]}, {"type": "kStatMod", "slotLabel": "防御", "perks": [5001, 5002, 5003]}], "defaultPageName": "坚决：巨像", "defaultSubStyle": 8200, "defaultPerks": [8437, 8446, 8444, 8451, 8224, 8237, 5008, 5002, 5001], "defaultPerksWhenSplashed": [8444, 8446], "defaultStatModsPerSubStyle": [{"id": "8000", "perks": [5005, 5002, 5001]}, {"id": "8100", "perks": [5008, 5002, 5001]}, {"id": "8200", "perks": [5008, 5002, 5001]}, {"id": "8300", "perks": [5007, 5002, 5001]}]}}
    for perkstyle_iter in perkstyle_initial["styles"]:
        perkstyle_id = perkstyle_iter["id"]
        perkstyles_initial[perkstyle_id] = perkstyle_iter
    ##准备云顶之弈强化符文数据（Prepare TFT augment data）
    TFTAugments_initial = {} #TFTAugments为嵌套字典，键为物件在LCU API上的表达形式，值为物件信息字典。一个键值对的示例如右：（Variable `TFTAugments` is a nested dictionary, whose keys are LCU API representation of items and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{"TFT7_Consumable_NeekosHelpDragon": {"associatedTraits": [], "composition": [], "desc": "TFT7_Consumable_Description_Dragonling", "effects": {}, "from": None, "icon": "ASSETS/Maps/Particles/TFT/TFT7_Consumable_Dragonling.tex", "id": None, "incompatibleTraits": [], "name": "TFT7_Consumable_Name_Dragonling", "unique": False}}
    for item in TFT_initial["items"]:
        item_apiName = item["apiName"]
        TFTAugments_initial[item_apiName] = item
    ##准备云顶之弈英雄数据（Prepare TFT champion data）
    TFTChampions_initial = {} #TFTChampions为嵌套字典，键为棋子在LCU API上的表达形式，值为棋子信息字典。一个键值对的示例如右：（Variable `TFTChampions` is a nested dictionary, whose keys are LCU API representation of TFT Champions and values are TFT Champion information dictionaries. An example of the key-value pairs is shown as follows: ）{"TFT9_Aatrox": {"character_record": {"path": "Characters/TFT9_Aatrox/CharacterRecords/Root", "character_id": "TFT9_Aatrox", "rarity": 9, "display_name": "亚托克斯", "traits": [{"name": "暗裔", "id": "Set9_Darkin"}, {"name": "裁决战士", "id": "Set9_Slayer"}, {"name": "主宰", "id": "Set9_Armorclad"}], "squareIconPath": "/lol-game-data/assets/ASSETS/Characters/TFT9_Aatrox/HUD/TFT9_Aatrox_Square.TFT_Set9.png"}}}
    for TFTChampion_iter in TFTChampion_initial:
        champion_name = TFTChampion_iter["name"]
        TFTChampions_initial[champion_name] = TFTChampion_iter["character_record"]
    ##准备云顶之弈装备数据（Prepare TFT item data）
    TFTItems_initial = {} #TTItems为嵌套字典，键为云顶之弈装备名称序号，值为云顶之弈装备信息字典。一个键值对的示例如右：（Variable `TFTItems` is a nested dictionary, whose keys are TFT item nameIds and values are TFT item information dictionaries. An example of the key-value pairs is shown as follows: ）{"TFTTutorial_Item_BFSword": {"guid": "9f6e75bb-7ba2-49aa-8724-04c550279034", "name": "暴风大剑", "id": 0, "color": {"R": 73, "B": 54, "G": 68, "A": 255}, "loadoutsIcon": "/lol-game-data/assets/ASSETS/Maps/Particles/TFT/Item_Icons/Standard/BF_Sword.png"}}
    for TFTItem_iter in TFTItem_initial:
        item_nameId = TFTItem_iter["nameId"]
        TFTItems_initial[item_nameId] = TFTItem_iter
    ##准备云顶之弈小小英雄数据（Prepare TFT companion data）
    TFTCompanions_initial = {} #TFTCompanions为嵌套字典，键为小小英雄序号，值为小小英雄信息字典。一个键值对的示例如右：（Variable `TFTCompanions` is a nested dictionary, whose keys are companion contentIds and values are companion information dictionaries. An example of the key-value pairs is shown as follows: ）{"91f2e228-4e36-4dad-9a97-36036e3eca36": {"itemId": 13010, "name": "节奏大师 奥希雅", "loadoutsIcon": "/lol-game-data/assets/ASSETS/Loadouts/Companions/Tooltip_AkaliDragon_Beatmaker_Tier1.png", "description": "奥希雅是酷炫的具象化。它用毫不费力的语流，指挥着韵脚和节奏，甚至能让最出色的小小英雄们羡慕不休。", "level": 1, "speciesName": "奥希雅", "speciesId": 13, "rarity": "Epic", "rarityValue": 1, "isDefault": false, "upgrades": ["0e251d36-d86e-4c58-9b7f-bcee2376a408", "e3151dc2-c45c-4949-89e9-6afda3b2fd5f"], "TFTOnly": false}}
    for companion_iter in TFTCompanion_initial:
        contentId = companion_iter["contentId"]
        TFTCompanions_initial[contentId] = companion_iter
    ##准备云顶之弈羁绊数据（Prepare TFT trait data）
    TFTTraits_initial = {} #TFTTraits为嵌套字典，键为羁绊在LCU API上的表达形式，值为羁绊信息字典。一个键值对的示例如右：（Variable `TFTTraits` is a nested dictionary, whose keys are LCU API representation of traits and values are trait information dictionaries. An example of the key-value pairs is shown as follows: ）{"Assassin": {"display_name": "刺客", "set": "TFTSet1", "icon_path": "/lol-game-data/assets/ASSETS/UX/TraitIcons/Trait_Icon_Assassin.png", "tooltip_text": "固有：在战斗环节开始时，刺客们会跃至距离最远的敌人处。<br><br>刺客们会获得额外的暴击伤害和暴击几率。<br><br><expandRow>(@MinUnits@) +@CritAmpPercent@%暴击伤害和+@CritChanceAmpPercent@%暴击几率</expandRow><br>", "innate_trait_sets": [], "conditional_trait_sets": {2: {"effect_amounts": [{"name": "CritAmpPercent", "value": 75.0, "format_string": ""}, {"name": "CritChanceAmpPercent", "value": 5.0, "format_string": ""}], "min_units": 3, "max_units": 5, "style_name": "kBronze"}, 3: {"effect_amounts": [{"name": "CritAmpPercent", "value": 150.0, "format_string": ""}, {"name": "CritChanceAmpPercent", "value": 20.0, "format_string": ""}], "min_units": 6, "max_units": 8, "style_name": "kSilver"}, 4: {"effect_amounts": [{"name": "CritAmpPercent", "value": 225.0, "format_string": ""}, {"name": "CritChanceAmpPercent", "value": 30.0, "format_string": ""}], "min_units": 9, "max_units": 25000, "style_name": "kGold"}}}}
    for trait_iter in TFTTrait_initial:
        trait_id = trait_iter["trait_id"]
        conditional_trait_sets = {}
        for conditional_trait_set in trait_iter["conditional_trait_sets"]:
            style_idx = conditional_trait_set["style_idx"]
            conditional_trait_sets[style_idx] = conditional_trait_set
        trait_iter["conditional_trait_sets"] = conditional_trait_sets
        TFTTraits_initial[trait_id] = trait_iter
    ##准备斗魂竞技场强化符文数据（Prepare Arena augment data）
    CherryAugments_initial = {} #CherryAugments为嵌套字典，键为斗魂竞技场强化符文在LCU API上的表达形式，值为斗魂竞技场强化符文信息字典。一个键值对的实例如右：（Variable `CherryAugments` is a nested dictionary, whose keys are LCU API representation of Arena augments and values are Arena augment information dictionaries. An example of the key-value pairs is shown as follows: ）{205: {"nameTRA": "物理转魔法", "augmentSmallIconPath": "/lol-game-data/assets/ASSETS/UX/Cherry/Augments/Icons/ADAPt_small.png", "rarity": "kSilver"}}
    for CherryAugment in CherryAugment_initial:
        CherryAugment_id = CherryAugment["id"]
        CherryAugments_initial[CherryAugment_id] = CherryAugment
    #下面创建一个字典，用来存储程序正在使用的各数据资源的版本（The following code create a dictionary to store the versions of data resources that the program currently uses）
    current_versions = {"summonerIcon": URLPatch, "spell": URLPatch, "LoLChampion": URLPatch, "LoLItem": URLPatch, "summonerIcon": URLPatch, "perk": URLPatch, "perkstyle": URLPatch, "TFTAugment": URLPatch, "TFTChampion": URLPatch, "TFTItem": URLPatch, "TFTCompanion": URLPatch, "TFTTrait": URLPatch, "CherryAugment": URLPatch}
    #下面创建一个字典，用来存储程序正在使用的各数据资源的版本下发生错误的键。当某个数据资源更换版本时，其出错的键会被清空（The following code create a dictionary to store the keys that fail to map to the constant dictionaries under certain versions of each kind of data resource. Once the version of a data resource changes, its unmapped keys will be cleared）
    unmapped_keys = {"summonerIcon": set(), "spell": set(), "LoLChampion": set(), "LoLItem": set(), "summonerIcon": set(), "perk": set(), "perkstyle": set(), "TFTAugment": set(), "TFTChampion": set(), "TFTItem": set(), "TFTCompanion": set(), "TFTTrait": set(), "CherryAugment": set()}
    ##准备大区数据（Prepare server / platform data）
    platform_TENCENT = {"BGP1": "全网通区 男爵领域（Baron Zone）", "BGP2": "峡谷之巅（Super Zone）", "EDU1": "教育网专区（CRENET Server）", "HN1": "电信一区 艾欧尼亚（Ionia）", "HN2": "电信二区 祖安（Zaun）", "HN3": "电信三区 诺克萨斯（Noxus 1）", "HN4": "电信四区 班德尔城（Bandle City）", "HN4_NEW": "电信四区 班德尔城（Bandle City）", "HN5": "电信五区 皮尔特沃夫（Piltover）", "HN6": "电信六区 战争学院（the Institute of War）", "HN7": "电信七区 巨神峰（Mount Targon）", "HN8": "电信八区 雷瑟守备（Noxus 2）", "HN9": "电信九区 裁决之地（the Proving Grounds）", "HN10": "电信十区 黑色玫瑰（the Black Rose）", "HN11": "电信十一区 暗影岛（Shadow Isles）", "HN12": "电信十二区 钢铁烈阳（the Iron Solari）", "HN13": "电信十三区 水晶之痕（Crystal Scar）", "HN14": "电信十四区 均衡教派（the Kinkou Order）", "HN15": "电信十五区 影流（the Shadow Order）", "HN16": "电信十六区 守望之海（Guardian's Sea）", "HN17": "电信十七区 征服之海（Conqueror's Sea）", "HN18": "电信十八区 卡拉曼达（Kalamanda）", "HN19": "电信十九区 皮城警备（Piltover Wardens）", "PBE": "体验服 试炼之地（Chinese PBE）", "WT1": "网通一区 比尔吉沃特（Bilgewater）", "WT1_NEW": "网通一区 比尔吉沃特（Bilgewater）", "WT2": "网通二区 德玛西亚（Demacia）", "WT2_NEW": "网通二区 德玛西亚（Demacia）", "WT3": "网通三区 弗雷尔卓德（Freljord）", "WT3_NEW": "网通三区 弗雷尔卓德（Freljord）", "WT4": "网通四区 无畏先锋（House Crownguard）", "WT4_NEW": "网通四区 无畏先锋（House Crownguard）", "WT5": "网通五区 恕瑞玛（Shurima）", "WT6": "网通六区 扭曲丛林（Twisted Treeline）", "WT7": "网通七区 巨龙之巢（the Dragon Camp）", "FORCES": "比赛服 艾欧尼亚（Tournament - Ionia）", "NJ100": "联盟一区", "GZ100": "联盟二区", "CQ100": "联盟三区", "TJ100": "联盟四区", "TJ101": "联盟五区", "PREPBE": "试炼之地 临时过渡服务器（Chinese PBE Temporary）"}
    platform_RIOT = {"ME1": "中东服（Middle East）", "BR1": "巴西服（Brazil）", "EUN1": "北欧和东欧服（Europe Nordic & East）", "EUW1": "西欧服（Europe West）", "JP1": "日服（Japan）", "KR": "韩服（Republic of Korea）", "LA1": "北拉美服（Latin America North）", "LA2": "南拉美服（Latin America South）", "NA1": "北美服（North America）", "OC1": "大洋洲服（Oceania）", "TR1": "土耳其服（Turkey）", "RU": "俄罗斯服（Russia）", "PH2": "菲律宾服（Philippines）", "SG2": "新加坡服（Singapore）", "TH2": "泰服（Thailand）", "TW2": "台服（Taiwan, Hong Kong and Macau）", "VN2": "越南服（Vietnam）", "PBE1": "测试服（Public Beta Environment）"}
    platform_GARENA = {"PH1": "菲律宾服（Philippines）", "SG1": "新加坡服（Singapore, Malaysia and Indonesia）", "TW1": "台服（Taiwan, Hong Kong and Macau）", "VN1": "越南服（Vietnam）", "TH1": "泰服（Thailand）"}
    platform = {"TENCENT": "国服（TENCENT）", "RIOT": "外服（RIOT）", "GARENA": "竞舞（GARENA）"}
    platformIds = list((platform_TENCENT | platform_GARENA | platform_RIOT).keys())
    #定义常量字典（Define constant dictionaries）
    ##基础信息（Basic information）
    tiers = {"": "", "NONE": "没有段位", "IRON": "坚韧黑铁", "BRONZE": "英勇黄铜", "SILVER": "不屈白银", "GOLD": "荣耀黄金", "PLATINUM": "华贵铂金", "EMERALD": "流光翡翠", "DIAMOND": "璀璨钻石", "MASTER": "超凡大师", "GRANDMASTER": "傲世宗师", "CHALLENGER": "最强王者"}
    #tiers = {"": "", "NONE": "NONE", "IRON": "IRON", "BRONZE": "BRONZE", "SILVER": "SILVER", "GOLD": "GOLD", "PLATINUM": "PLATINUM", "EMERALD": "EMERALD", "DIAMOND": "DIAMOND", "MASTER": "MASTER", "GRANDMASTER": "GRANDMASTER", "CHALLENGER": "CHALLENGER"}
    ratedTiers = {"": "", "NONE": "没有段位", "GRAY": "灰白", "GREEN": "翠绿", "BLUE": "天蓝", "PURPLE": "绛紫", "ORANGE": "耀橙"}
    #ratedTiers = {"": "", "NONE": "NONE", "GRAY": "GRAY", "GREEN": "GREEN", "BLUE": "BLUE", "PURPLE": "PURPLE", "ORANGE": "ORANGE"}
    tiers_all = tiers | ratedTiers
    ##排位信息（Ranked）
    #queueTypes = {"ARAM_BOT": "极地大乱斗 人机对战", "ARAM_CLASH": "极地大乱斗 冠军杯赛", "ARAM_UNRANKED_1x1": "极地大乱斗1v1", "ARAM_UNRANKED_5x5": "极地大乱斗5v5", "BOT": "人机对战", "CHERRY": "斗魂竞技场", "CHERRY_UNRANKED": "斗魂竞技场 匹配模式", "CHONCC_TREASURE_TFT": "云顶之弈（恭喜发财）", "CLASH": "冠军杯赛", "FIVE_YEAR_ANNIVERSARY_TFT": "云顶之弈 5周年时光机", "LNY23_TFT": "云顶之弈（恭喜发财）", "LNY24_TFT": "云顶之弈 （第3.5赛季回归：再战星海）", "NEXUSBLITZ": "极限闪击", "NORMAL": "匹配模式", "NORMAL_TFT": "云顶之弈 匹配模式", "ONEFORALL": "克隆大作战", "RANKED_FLEX_SR": "灵活 5V5", "RANKED_SOLO_5x5": "单人/双人", "RANKED_TFT": "云顶之弈 排位赛", "RANKED_TFT_DOUBLE_UP": "双人作战", "RANKED_TFT_PAIRS": "2V0", "RANKED_TFT_TURBO": "狂暴模式", "RIOTSCRIPT_BOT": "人机对战", "SF_TFT": "云顶之弈（斗魂锦标赛）", "STRAWBERRY": "无尽狂潮", "TURBO_TFT": "云顶之弈 狂暴模式 自定义", "TUTORIAL_MODULE_1": "新手教程 第一部分", "TUTORIAL_MODULE_2": "新手教程 第二部分", "TUTORIAL_MODULE_3": "新手教程 第三部分", "TUTORIAL_TFT": "云顶之弈 新手教程", "ULTBOOK": "终极魔典", "URF": "无限火力"}
    queueTypes = {"RANKED_SOLO_5x5": "单人/双人", "RANKED_FLEX_SR": "灵活 5V5", "RANKED_TFT": "云顶之弈", "RANKED_TFT_PAIRS": "2V0", "RANKED_TFT_DOUBLE_UP": "双人作战", "RANKED_TFT_TURBO": "狂暴模式", "CHERRY": "斗魂竞技场"} #2V0模式仅美测服可用（RANKED_TFT_PAIRS is only available on PBE）
    #queueTypes = {"RANKED_SOLO_5x5": "Ranked Solo/Duo", "RANKED_FLEX_SR": "Ranked Flex", "RANKED_TFT": "Ranked TFT", "RANKED_TFT_PAIRS": "2V0", "RANKED_TFT_DOUBLE_UP": "Double Up", "RANKED_TFT_TURBO": "Hyper Roll", "CHERRY": "Arena"}
    ##英雄联盟对局记录（LoL match history）
    gameTypes = {"MATCHED_GAME": "匹配对局", "CUSTOM_GAME": "自定义对局", "TUTORIAL_GAME": "新手教程"}
    #gameTypes = {"MATCHED_GAME": "MATCHED_GAME", "CUSTOM_GAME": "CUSTOM_GAME", "TUTORIAL_GAME": "TUTORIAL_GAME"}
    team_color = {0: "", 100: "蓝方", 200: "红方"}
    #team_color = {0: "", 100: "Blue", 200: "Red"}
    endOfGameResults = {"": "", "GameComplete": "游戏结束", "Abort_Unexpected": "意外终止", "Abort_TooFewPlayers": "全员提前退出", "Abort_AntiCheatExit": "检测到作弊而终止"}
    #endOfGameResults = {"": "", "GameComplete": "GameComplete", "Abort_Unexpected": "Abort_Unexpected", "Abort_TooFewPlayers": "Abort_TooFewPlayers", "Abort_AntiCheatExit": "Abort_AntiCheatExit"}
    lanes = {"TOP": "上路", "JUNGLE": "打野", "MIDDLE": "中路", "BOTTOM": "下路", "NONE": ""}
    #lanes = {"TOP": "TOP", "JUNGLE": "JUNGLE", "MIDDLE": "MIDDLE", "BOTTOM": "BOTTOM", "NONE": ""}
    roles = {"CARRY": "C位", "DUO": "游走", "SOLO": "单人", "SUPPORT": "辅助", "NONE": ""}
    #roles = {"CARRY": "CARRY", "DUO": "DUO", "SOLO": "SOLO", "SUPPORT": "SUPPORT", "NONE": ""}
    ##英雄联盟对局信息（LoL match information）
    subteam_color = {0: "", 1: "魄罗", 2: "小兵", 3: "迅捷蟹", 4: "石甲虫", 5: "锋喙鸟", 6: "哨卫", 7: "狼", 8: "魔沼蛙"} #仅用于斗魂竞技场（Only for Arena mode）
    #subteam_color = {0: "", 1: "Poro", 2: "Minion", 3: "Scuttle", 4: "Krug", 5: "Raptor", 6: "Sentinel", 7: "Wolf", 8: "Gromp"}
    augment_rarity = {0: "白银", 1: "黄金", 2: "棱彩", 4: "黄金", 8: "棱彩", "kBronze": "青铜", "kSilver": "白银", "kGold": "黄金", "kPrismatic": "棱彩"}
    #augment_rarity = {0: "Silver", 1: "Gold", 2: "Prismatic", 4: "Gold", 8: "Prismatic", "KBronze": "Bronze", "KSilver": "Silver", "kGold": "Gold", "kPrismatic": "Prismatic"}
    win = {True: "胜利", False: "失败"}
    #win = {True: "V", False: "D"}
    ##英雄联盟事件（LoL events）
    eventTypes = {"CHAMPION_KILL": "英雄击杀", "ELITE_MONSTER_KILL": "史诗级野怪击杀", "BUILDING_KILL": "建筑物击杀"}
    #eventTypes = {"CHAMPION_KILL": "Champion Kills", "ELITE_MONSTER_KILL": "Elite Monster Kills", "BUILDING_KILL": "Building Kills"}
    buildingTypes = {"": "", "TOWER_BUILDING": "防御塔", "INHIBITOR_BUILDING": "召唤水晶"}
    #buildingTypes = {"": "", "TOWER_BUILDING": "Turret", "INHIBITOR_BUILDING": "Inhibitor"}
    laneTypes = {"": "", "TOP_LANE": "上路", "MID_LANE": "中路", "BOT_LANE": "下路"}
    #laneTypes = {"": "", "TOP_LANE": "Top", "MID_LANE": "Middle", "BOT_LANE": "Bottom"}
    monsterSubTypes = {"": "", "EARTH_DRAGON": "山脉亚龙", "CHEMTECH_DRAGON": "炼金科技亚龙", "WATER_DRAGON": "海洋亚龙", "HEXTECH_DRAGON": "海克斯科技亚龙", "AIR_DRAGON": "云霄亚龙", "FIRE_DRAGON": "炼狱亚龙", "ELDER_DRAGON": "远古巨龙", "RUINED_DRAGON": "破败巨龙", "UNKNOWN": "未知"}
    #monsterSubTypes = {"": "", "EARTH_DRAGON": "Mountain Drake", "CHEMTECH_DRAGON": "Chemtech Drake", "WATER_DRAGON": "Ocean Drake", "HEXTECH_DRAGON": "Hextech Drake", "AIR_DRAGON": "Cloud Drake", "FIRE_DRAGON": "Infernal Drake", "ELDER_DRAGON": "Elder Dragon", "RUINED_DRAGON": "Ruined Dragon", "UNKNOWN": "Unknown"}
    monsterTypes = {"": "", "RIFTHERALD": "峡谷先锋", "HORDE": "虚空巢虫", "BARON_NASHOR": "纳什男爵", "DRAGON": "巨龙", "ATAKHAN": "厄塔汗"}
    #monsterTypes = {"": "", "RIFTHERALD": "Rift Herald", "HORDE": "Voidgrub", "BARON_NASHOR": "Baron Nashor", "DRAGON": "Drake", "ATAKHAN": "Atakhan"}
    towerTypes = {"": "", "OUTER_TURRET": "外防御塔", "INNER_TURRET": "内防御塔", "BASE_TURRET": "水晶防御塔", "NEXUS_TURRET": "枢纽防御塔"}
    #towerTypes = {"": "", "OUTER_TURRET": "Outer Turret", "INNER_TURRET": "Inner Turret", "BASE_TURRET": "Inhibitor Turret", "NEXUS_TURRET": "Nexus Turret"}
    ##云顶之弈对局记录（TFT match history）
    #traitStyles = {"kThreat": "威慑", "kBronze": "青铜", "kSilver": "白银", "kGold": "黄金", "kChromatic": "炫金"}
    traitStyles = {0: "", 1: "青铜", 2: "白银", 3: "黄金", 4: "炫金", 5: "独特"}
    rarities = {"Default": "经典", "NoRarity": "其它", "Epic": "史诗", "Legendary": "传说", "Mythic": "神话", "Rare": "稀有", "Ultimate": "终极", "Exalted": "圣者至尊", "Transcendant": "超凡"}
    #控制只输出一遍的提示（Control the hint to be displayed only once）
    puuid_change_warning_printed = False
    Vanguard_warning_printed = False
    #设置电脑玩家的玩家通用唯一识别码（Set the puuid of a bot player）
    bot_puuid = "00000000-0000-0000-0000-000000000000"
    logPrint("请选择本脚本的使用模式：\nPlease select a mode for use:\n1\t生成模式（Generate Mode）\n2\t检测模式（Detect Mode）")
    detectMode = False
    mode = logInput()
    if mode == "" or mode[0] != "1":
        detectMode = True
        smurf_asked = False #在检测模式时，询问用户是否导入其它账号。当用户从检测模式切换到生成模式时，该变量会置为假（Under Detect Mode, the program asks if the user wants to import other accounts. When the user switch from Detect Mode to Generate Mode, this variable is set False）
    switch_mode = False #模式转换变量定义（Definition of the mode transfer variable）
    #然后获取历史记录（Next, fetch the history）
    #logPrint('''在腾讯代理的服务器上，如果查询某名玩家的对局记录，请尝试以下操作：\nTo search for the match history of a player on Tencent servers, try out the following operations:\n1. 在浏览器中打开本地主机网络协议：%s\n   Open the localhost IP in any browser: %s\n2. 尝试用以下用户名和密码登录：\n   Try logining in with the following username and password:\n   用户名（Username）：riot\n   密码（Password）：%s\n3. （如果可以立即知道一位玩家的玩家通用唯一识别码，则可以跳过第3和4步）在浏览器的地址栏中的地址最后，添加“lol-summoner/v1/summoners?name={name}”，其中{name}指的是召唤师名称编码后的字符串。当召唤师名称只包含英文字母和阿拉伯数字时，直接以召唤师名称去空格后的字符串代入{name}即可；当召唤师名称存在非美国标准信息交换代码时，以召唤师名称编码后的字符串代入{name}。\n(If a summoner's puuid can be immediately known, the user may skip Steps 3 and 4) Add to following the last character of the address in the browser's address bar "lol-summoner/v1/summoners?name={name}", where {name} refers to strings encoded from summonerName. When summonerName contains only English letters and Arabic numbers, simply substitute {name} with the strings with the spaces removed from summonerName. When a non-ASCII character exists in summonerName, substitute {name} by encoded summonerName.\n3.1 对于包含非美国标准信息交换代码的召唤师名称，如果可以得到该召唤师的精确名称（如通过复制到剪贴板），那么在Python中可以得知其编码后的字符串。在Python中使用from urllib.parse import quote命令引入quote函数，再使用quote(x)函数获取字符串x编码后的字符串。\nFor summonerNames that include non-ASCII characters, if the exact summonerName can be obtained (e. g. by copying to clipboard), then its encoded string can be returned in Python. In Python console, use "from urllib.parse import quote" to introduce the "quote" function. Then use quote(x) function to get the string encoded from the string x.\n4. 在lol-summoner/v1/summoners?name={name}返回的结果中找到puuid并复制。\n   Find "puuid" in the result returned by "lol-summoner/v1/summoners?name={name}" and copy it.\n5. 将地址栏中4位IP地址后的斜杠后的内容删除，再添加“lol-match-history/v1/products/lol/{puuid}/matches?begIndex=0&endIndex=20”，其中{puuid}是事先获知的玩家通用唯一识别码，或者是第4步复制到剪贴板的puuid。\nDelete the content following the slash after the 4-bit IP address in the address bar and then add to the end "lol-match-history/v1/products/lol/{puuid}/matches?begIndex=0&endIndex=20", where {puuid} refers to the puuid previously known, or copied to clipboard in Step 4.\n6. 尝试将上一步输入的地址中的“endIndex=”后的数字依次替换成21、199、200和500，观察每次替换后返回的网页结果有没有变多。\nTry changing the number following "endIndex=" in the last step into 21, 199, 200 and 500 one by one, and observe whether the returned webpage contains more information after each change.\n7. 教程完成，请继续执行本脚本……\n   Instruction finished. Please continue to run this program ...''' %(connection.address, connection.address, connection.auth_key))
    while True:
        infos = {} #存储程序运行过程中遇到的玩家信息，防止后续程序反复获取已经获取过的玩家信息（Store the summoner information fetched during the program execution, in case the program would keep capturing the summoner information already fetched before）
        #通过小号模式导入其它自己玩过的账号（Import other accounts that the user has played by Smurf Mode）
        detectMode = not detectMode if switch_mode else detectMode
        selfDetect = detectMode #标记检测模式是否检测自己（Marks whether Detect Mode detects the user itself）
        smurf_asked = False if not detectMode else smurf_asked
        switch_mode = False #模式转换变量初始化（Initialization of the mode transfer variable）
        if detectMode and not smurf_asked:
            smurf_asked = True
            logPrint("是否导入其它账号？（输入任意非空字符串以导入，否则不导入。）\nImport other accounts? (Submit any non-empty string to import, or null to refuse importing.)")
            smurfMode_str = logInput()
            smurfMode = bool(smurfMode_str)
            if smurfMode:
                smurfs = []
                smurf_header = {"displayName": "显示名", "gameName": "玩家昵称", "tagLine": "昵称编号", "summonerId": "召唤师序号", "puuid": "玩家通用唯一识别码"}
                smurf_df = pandas.DataFrame(data = smurf_header, index = [0])
                logPrint("请选择导入方式：\nPlease select an option to import:\n☆1\t读取文件（Read a file）\n2\t手动输入（Manually input）")
                smurf_option = logInput()
                if smurf_option != "" and smurf_option[0] == "2":
                    smurf_option = "2"
                else:
                    smurf_option = "1"
                #在下面的代码中，关键是列表`smurfs`中追加小号信息（The key point of the following code is to append smurf information into the list `smurfs`）
                smurf_file_read = False #标记程序是否成功读取到含有小号信息的数据文件（Marks whether the smurf data file is read successfully）
                smurf_file = "Smurf Accounts.json"
                smurf_file_rename = "Smurf Accounts (Invalid).json"
                if smurf_option == "1":
                    while os.path.exists(smurf_file_rename): #确保下面的重命名操作不会引发报错（Ensure the following renaming operation won't cause an error）
                        smurf_file_rename = "(1).".join(smurf_file_rename.split("."))
                    if os.path.exists(smurf_file):
                        try:
                            with open(smurf_file, "r", encoding = "utf-8") as fp:
                                smurf_local = json.load(fp)
                        except json.decoder.JSONDecodeError:
                            os.rename(smurf_file, smurf_file_rename) #上面的while循环保证这里重命名后的文件不可能存在（The above while-loop ensures the result file can't exist）
                            logPrint(f'''在同目录下发现了格式不正确的数据文件。该文件已重命名为“{smurf_file_rename}”。程序将转为手动输入。\nA smurf data file with invalid format is found under the same directory. This file has been renamed into "{smurf_file_rename}". You may need to input the smurfs' names manually.''')
                            smurf_option = "2"
                        else:
                            if isinstance(smurf_local, dict) and all(map(lambda x: x in platformIds, smurf_local.keys())) and all(map(lambda x: isinstance(x, dict), smurf_local.values())) and all(len(smurf_local_iter) == 0 or all(map(lambda x: isinstance(x, str) and verify_uuid(x), smurf_local_iter.keys())) and all(map(lambda x: isinstance(x, list) and all(map(lambda y: isinstance(y, str) and verify_uuid(y), x)), smurf_local_iter.values())) for smurf_local_iter in smurf_local.values()): #格式的严格校验（A serious verification of the format）
                                smurf_file_read = True
                                if platformId in smurf_local:
                                    if current_info["puuid"] in smurf_local[platformId]: #一定要注意，在本脚本中，`current_puuid`和`current_info["puuid"]`不是一回事（Pay attention that `current_puuid` and `current_info["puuid"]` aren't the same thing in this program）
                                        count = 0 #标识小号的序号（Number the smurfs）
                                        valid_puuid_count = 0 #记录能查询到玩家的玩家通用唯一识别码的数量（Record the number of puuids that can correspond to players）
                                        for smurf_puuid in smurf_local[platformId][current_info["puuid"]]:
                                            count += 1
                                            logPrint(f"{count}.\t{smurf_puuid}")
                                            info = await get_info(connection, smurf_puuid)
                                            if info["info_got"]:
                                                info_body = info["body"]
                                                if info_body["puuid"] in list(map(lambda x: x["puuid"], smurfs)):
                                                    logPrint("您已经输入过该玩家了。\nYou've entered this player.")
                                                else:
                                                    valid_puuid_count += 1
                                                    logPrint(info_body)
                                                    smurfs.append(info_body)
                                                    smurf_record = {key: info["body"][key] for key in smurf_header}
                                                    smurf_df = pandas.concat([smurf_df, pandas.DataFrame([smurf_record])], ignore_index = True)
                                                    print(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0], end = "\n\n")
                                                    log.write(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0] + "\n\n")
                                                    infos[info_body["puuid"]] = info_body
                                            else:
                                                logPrint(info["message"])
                                        logPrint(f"从离线文件中读取到了{valid_puuid_count}个小号信息。是否继续输入更多小号？（输入任意键以继续添加小号，否则不添加。）\nThe program has detected {valid_puuid_count} smurf account(s) from the local file. Do you want to continue with more smurf accounts? (Submit any non-empty string to continue, or null to refuse adding.)")
                                        input_more_smurf_str = logInput()
                                        smurf_option = "2" if bool(input_more_smurf_str) else "1"
                                    else:
                                        logPrint("在同目录下发现了含有小号信息的数据文件，但是没有找到您的小号信息。程序将转为手动输入。\nThe smurf data file is found under the same directory, but without yours. You may need to input the smurfs' names manually.")
                                        smurf_option = "2"
                                else:
                                    logPrint("在同目录下发现了含有小号信息的数据文件，但是没有找到您的大区信息。如果您确认您的本地文件没有问题，请向作者反馈该问题。\nThe smurf data file is found under the same directory, but without your server's. If you're sure that there's not any problem in your local data file, please file the feedback to the author.\n一个可用的反馈链接：\nAn available feedback link:\nhttps://github.com/WordlessMeteor/LoL-DIY-Programs/issues/new \n程序将转为手动输入。\nYou may need to input the smurfs' names manually.")
                                    smurf_option = "2"
                            else:
                                os.rename(smurf_file, smurf_file_rename)
                                logPrint(f'''在同目录下发现了格式不正确的数据文件。该文件已重命名为“{smurf_file_rename}”。程序将转为手动输入。\nA smurf data file with invalid format is found under the same directory. This file has been renamed into "{smurf_file_rename}". You may need to input the smurfs' names manually.''')
                                smurf_option = "2"
                    else:
                        logPrint("没有找到含有小号信息的数据文件。程序将转为手动输入。\nSmurf data file not found. You may need to input the smurfs' names manually.")
                        smurf_option = "2"
                if smurf_option == "2":
                    logPrint('请输入小号的召唤师名。输入“0”以清空已经输入的小号。输入-1以结束。\nPlease input the summoner names of the smurf accounts. Submit "0" to clear the entered smurfs. Submit "-1" to finish the importation.')
                    while True:
                        smurfName = logInput()
                        if smurfName == "-1":
                            break
                        elif smurfName == "0":
                            smurfs = []
                            smurf_df = pandas.DataFrame(data = smurf_header, index = [0])
                            logPrint("已清空小号。\nSmurfs cleared.")
                        elif smurfName == "":
                            continue
                        elif smurfName in {"current-summoner", get_info_name(current_info), current_info["puuid"], str(current_info["summonerId"])} and selfDetect:
                            logPrint("您不能把主账号作为小号！请添加其它账号。\nYou're not allowed to add your main account as a smurf account! Please try another account.")
                        else:
                            info = await get_info(connection, smurfName)
                            if info["info_got"]:
                                info_body = info["body"]
                                if info_body["puuid"] in list(map(lambda x: x["puuid"], smurfs)):
                                    logPrint("您已经输入过该玩家了。\nYou've entered this player.")
                                else:
                                    logPrint(info_body)
                                    smurfs.append(info_body)
                                    smurf_record = {key: info["body"][key] for key in smurf_header}
                                    smurf_df = pandas.concat([smurf_df, pandas.DataFrame([smurf_record])], ignore_index = True)
                                    print(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0])
                                    log.write(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0] + "\n")
                                    infos[info_body["puuid"]] = info_body
                            else:
                                logPrint(info["message"])
                logPrint("是否需要将小号信息保存到本地，以便日后直接读取文件来添加小号？（输入任意键以确认，否则不保存。）\nDo you want to save the smurf information into a local data file, so that you may read this file directly to load the smurfs? (Submit any non-empty string to confirm, or null to refuse saving.)\n注意：程序会覆盖之前的小号信息，因此如果您不想丢失以前的小号信息，请在不输入任何字符的情况下直接按回车键不保存，然后将原来的小号信息做好备份。\nNote: The old smurf information will be overwritten, so if you expect the previous smurf information not to be lost, please directly press Enter without any other characters entered, and then make a backup of the original smurf information.")
                save_smurf_str = logInput()
                save_smurf = bool(save_smurf_str)
                if save_smurf:
                    if smurf_file_read:
                        if platformId in smurf_local:
                            smurf_local[platformId][current_info["puuid"]] = list(map(lambda x: x["puuid"], smurfs)) #之所以考虑用玩家通用唯一识别码，而不用召唤师名或者召唤师序号作为小号信息存储介质的原因有两个方面的考量：从对人类友好的角度上，召唤师名的确更胜一筹，但是缺少唯一性。在调用get_info函数时，两个召唤师名如果只是差几个空格，就很有可能指向同一个召唤师。这样，上面和下面的代码在识别召唤师信息是否添加过时，就不太好实现；从存储格式的角度上来考虑，玩家通用唯一识别码服从通用唯一识别码的格式，相对比较统一，而且是全球统一的，这样在校验数据文件格式时比较方便。而召唤师序号只是整数，而且不同召唤师序号存在长短不一的情况，这样校验起来不够充分（The reason why I consider using puuid as the smurf data storing media, instead of the summoner name or summonerId, has two considerations. On the one hand, in terms of being human-friendly, a summoner name does far outweigh the puuid or summonerId. However, it lacks uniformity. When `get_info` function is called, if two parameters differ only in several spaces, the result might directs to a same summoner. In that case, it's not easy to implement the code to identify whether a summoner's information has been added to the list before, within the context. On the other hand, in terms of the format, a puuid obeys the format of uuids, so it's relatively general, let alone being "universally unqiue", which makes it convenient to verify the format of the smurf data file. Nevertheless, summonerId is just an integer, and different summonerIds may be of different lengths, so it's not sufficient to determine a summoner by summonerId）
                        else:
                            smurf_local[platformId] = {current_info["puuid"]: list(map(lambda x: x["puuid"], smurfs))}
                    else:
                        smurf_local = {platformId: {current_info["puuid"]: list(map(lambda x: x["puuid"], smurfs))}}
                    with open(smurf_file, "w", encoding = "utf-8") as fp:
                        json.dump(smurf_local, fp, indent = 4, ensure_ascii = False)
                    logPrint(f"小号信息已保存到“{smurf_file}”中。\nSmurf information has been saved into {smurf_file}.")
        #初始化所有数据资源（Initialize all data resources）
        logPrint("\n正在初始化所有数据资源……\nInitializing all data resources ...\n")
        patches = copy.deepcopy(patches_initial)
        spells = copy.deepcopy(spells_initial)
        LoLChampions = copy.deepcopy(LoLChampions_initial)
        LoLItems = copy.deepcopy(LoLItems_initial)
        summonerIcons = copy.deepcopy(summonerIcons_initial)
        perks = copy.deepcopy(perks_initial)
        perkstyles = copy.deepcopy(perkstyles_initial)
        TFTAugments = copy.deepcopy(TFTAugments_initial)
        TFTChampions = copy.deepcopy(TFTChampions_initial)
        TFTItems = copy.deepcopy(TFTItems_initial)
        TFTCompanions = copy.deepcopy(TFTCompanions_initial)
        TFTTraits = copy.deepcopy(TFTTraits_initial)
        CherryAugments = copy.deepcopy(CherryAugments_initial)
        current_versions = {"summonerIcon": URLPatch, "spell": URLPatch, "LoLChampion": URLPatch, "LoLItem": URLPatch, "summonerIcon": URLPatch, "perk": URLPatch, "perkstyle": URLPatch, "TFTAugment": URLPatch, "TFTChampion": URLPatch, "TFTItem": URLPatch, "TFTCompanion": URLPatch, "TFTTrait": URLPatch, "CherryAugment": URLPatch}
        unmapped_keys = {"summonerIcon": set(), "spell": set(), "LoLChampion": set(), "LoLItem": set(), "summonerIcon": set(), "perk": set(), "perkstyle": set(), "TFTAugment": set(), "TFTChampion": set(), "TFTItem": set(), "TFTCompanion": set(), "TFTTrait": set(), "CherryAugment": set()}
        if not detectMode:
            logPrint('请输入要查询的召唤师名称，退出请输入“0”，切换成检测模式请输入“3”：\nPlease input the summoner name to be searched. Submit "0" to exit. Submit "3" to switch to Detect Mode.')
            summoner_name = logInput()
        else: #检测模式一律把玩家通用唯一识别码传入summoner_name变量（In Detect Mode, puuid is always passed to the variable `summoner_name`）
            members_to_detect = [current_info]
            gameflow_phase = await (await connection.request("GET", "/lol-gameflow/v1/gameflow-phase")).json()
            if gameflow_phase in {"Lobby", "Matchmaking", "ReadyCheck", "ChampSelect", "InProgress", "Reconnect"}:
                if gameflow_phase in {"Lobby", "Matchmaking", "ReadyCheck"}:
                    lobby = await (await connection.request("GET", "/lol-lobby/v2/lobby")).json()
                    for member in lobby["members"]:
                        if not member["puuid"] in {current_info["puuid"], "", bot_puuid}:
                            member_info_recapture = 0
                            if member["puuid"] in infos:
                                member_info_body = infos[member["puuid"]]
                            else:
                                member_info = await get_info(connection, member["puuid"])
                                while not member_info["info_got"] and member_info["body"]["httpStatus"] != 404 and member_info_recapture < 3:
                                    logPrint(member_info["message"])
                                    member_info_recapture += 1
                                    logPrint("成员信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of a member (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(member["puuid"], member_info_recapture, member["puuid"], member_info_recapture))
                                    member_info = await get_info(connection, member["puuid"])
                                if member_info["info_got"]:
                                    member_info_body = member_info["body"]
                                    infos[member["puuid"]] = member_info_body
                                else:
                                    logPrint(member_info["message"])
                                    logPrint("成员信息（玩家通用唯一识别码：%s）获取失败！将忽略该名成员。\nInformation of a member (puuid: %s) capture failed! The program will ignore this member.")
                                    continue
                            members_to_detect.append(member_info_body)
                    if len(members_to_detect) > 1:
                        logPrint("检测到您正在房间内。是否检测其他玩家的近期一起玩过的玩家？（输入下方其他玩家对应的编号以查询其他玩家，或者直接按回车键以查询用户本人。）\nThe program detected that you're currently in a lobby. Do you want to detect recently played summoners of another player? (Submit the number corresponding to another player below to search for his/her recently player summoners, or press Enter directly to search for recently played summoners of the user itself.)")
                elif gameflow_phase == "ChampSelect":
                    champ_select_session = await (await connection.request("GET", "/lol-champ-select/v1/session")).json()
                    for ally in champ_select_session["myTeam"]:
                        if not ally["puuid"] in {current_info["puuid"], "", bot_puuid} and (ally["nameVisibilityType"] == "VISIBLE" or ally["nameVisibilityType"] == ""):
                            ally_info_recapture = 0
                            if ally["puuid"] in infos:
                                ally_info_body = infos[ally["puuid"]]
                            else:
                                ally_info = await get_info(connection, ally["puuid"])
                                while not ally_info["info_got"] and ally_info["body"]["httpStatus"] != 404 and ally_info_recapture < 3:
                                    logPrint(ally_info["message"])
                                    ally_info_recapture += 1
                                    logPrint("队友信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an ally (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(ally["puuid"], ally_info_recapture, ally["puuid"], ally_info_recapture))
                                    ally_info = await get_info(connection, ally["puuid"])
                                if ally_info["info_got"]:
                                    ally_info_body = ally_info["body"]
                                    infos[ally["puuid"]] = ally_info_body
                                else:
                                    logPrint(ally_info["message"])
                                    logPrint("队友信息（玩家通用唯一识别码：%s）获取失败！将忽略该名队友。\nInformation of an ally (puuid: %s) capture failed! The program will ignore this ally.")
                                    continue
                            members_to_detect.append(ally_info_body)
                    if champ_select_session["theirTeam"]:
                        for enemy in champ_select_session["theirTeam"]:
                            if not enemy["puuid"] in {current_info["puuid"], "", bot_puuid} and (enemy["nameVisibilityType"] == "VISIBLE" or enemy["nameVisibilityType"] == ""):
                                enemy_info_recapture = 0
                                if enemy["puuid"] in infos:
                                    enemy_info_body = infos[enemy["puuid"]]
                                else:
                                    enemy_info = await get_info(connection, enemy["puuid"])
                                    while not enemy_info["info_got"] and enemy_info["body"]["httpStatus"] != 404 and enemy_info_recapture < 3:
                                        logPrint(enemy_info["message"])
                                        enemy_info_recapture += 1
                                        logPrint("对手信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an enemy (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(enemy["puuid"], enemy_info_recapture, enemy["puuid"], enemy_info_recapture))
                                        enemy_info = await get_info(connection, enemy["puuid"])
                                    if enemy_info["info_got"]:
                                        enemy_info_body = enemy_info["body"]
                                        infos[enemy["puuid"]] = enemy_info_body
                                    else:
                                        logPrint(enemy_info["message"])
                                        logPrint("对手信息（玩家通用唯一识别码：%s）获取失败！将忽略该名对手。\nInformation of an enemy (puuid: %s) capture failed! The program will ignore this enemy.")
                                        continue
                                members_to_detect.append(enemy_info_body)
                    if len(members_to_detect) > 1:
                        logPrint("检测到您正在英雄选择阶段。是否检测其他玩家的近期一起玩过的玩家？（输入下方其他玩家对应的编号以查询其他玩家，或者直接按回车键以查询用户本人。）\nThe program detected that you're currently during champ select stage. Do you want to detect recently played summoners of another player? (Submit the number corresponding to another player below to search for his/her recently player summoners, or press Enter directly to search for recently played summoners of the user itself.)")
                else:
                    gameflow_session = await (await connection.request("GET", "/lol-gameflow/v1/session")).json()
                    gameData = gameflow_session["gameData"]
                    for player in gameData["teamOne"] + gameData["teamTwo"]:
                        if "puuid" in player and player["puuid"] != current_info["puuid"]: #电脑玩家没有玩家通用唯一识别码（Bot players don't have puuids）
                            player_info_recapture = 0
                            if player["puuid"] in infos:
                                player_info_body = infos[player["puuid"]]
                            else:
                                player_info = await get_info(connection, player["puuid"])
                                while not player_info["info_got"] and player_info["body"]["httpStatus"] != 404 and player_info_recapture < 3:
                                    logPrint(player_info["message"])
                                    player_info_recapture += 1
                                    logPrint("玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an player (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(player["puuid"], player_info_recapture, player["puuid"], player_info_recapture))
                                    player_info = await get_info(connection, player["puuid"])
                                if player_info["info_got"]:
                                    player_info_body = player_info["body"]
                                    infos[player_info_body["puuid"]] = player_info_body
                                else:
                                    logPrint(player_info["message"])
                                    logPrint("玩家信息（玩家通用唯一识别码：%s）获取失败！将忽略该名队友。\nInformation of an player (puuid: %s) capture failed! The program will ignore this player.")
                                    continue
                            members_to_detect.append(player_info_body)
                    if len(members_to_detect) > 1:
                        logPrint("检测到您正在游戏中。是否检测其他玩家的近期一起玩过的玩家？（输入下方其他玩家对应的编号以查询其他玩家，或者直接按回车键以查询用户本人。）\nThe program detected that you're currently in a game. Do you want to detect recently played summoners of another player? (Submit the number corresponding to another player below to search for his/her recently player summoners, or press Enter directly to search for recently played summoners of the user itself.)")
            if len(members_to_detect) > 1:
                for i in range(len(members_to_detect)):
                    member_info_body = members_to_detect[i]
                    logPrint("%d\t%s\t%s" %(i, member_info_body["puuid"], get_info_name(member_info_body)))
                memberId = logInput()
                if memberId != "" and memberId in list(map(str, range(1, len(members_to_detect)))):
                    selfDetect = False
                    summoner_name = members_to_detect[int(memberId)]["puuid"]
                elif memberId == "0":
                    continue
                else:
                    selfDetect = True
                    summoner_name = "current-summoner"
            else:
                selfDetect = True
                summoner_name = "current-summoner"
        if summoner_name == "0":
            break
        elif summoner_name == "3":
            switch_mode = True
            continue
        elif summoner_name == "":
            logPrint("请输入非空字符串！\nPlease input a string instead of null!")
            continue
        else:
            if summoner_name == "current-summoner":
                info = {"searchType": "current-summoner", "endpoint": "/lol-summoner/v1/current-summoner", "info_got": True, "network_error": False, "body": current_info.copy(), "message": "", "selfInfo": True}
            else:
                info = await get_info(connection, summoner_name)
            if not info["info_got"]:
                logPrint(info["message"])
                continue
            else:
                info_body = info["body"]
                displayName = get_info_name(info_body) #用于扫描模式定位到某召唤师（Determines the directory which contains the summoner's data）
                current_summonerId = info_body["summonerId"] #用于排除房间邀请信息中的自己（Defined to exclude the user itself from the lobby invitations）
                current_puuid = info_body["puuid"] #用于核验对局是否包含该召唤师。此外，还用于扫描模式从对局的所有玩家信息中定位到该玩家（For use of checking whether the searched matches include this summoner. In addition, it's used for localization of this player from all players in a match in "scan" mode）
                current_summonerName = "" if info_body["gameName"] == "" and info_body["tagLine"] == "" else info_body["gameName"] + "#" + info_body["tagLine"] #作用同上，用于模糊定位，主要应用于玩家通用唯一识别码发生变动的大区且在昵称编号引入后注册的主召唤师的对局记录扫描模式（Acts as the same role as the above variable for a rough localization. It's mainly designed for Scan Mode on players that signed up after tagLine was introduced on servers that changed the players' puuids）
                infos[current_puuid] = info_body
                #下面准备一些数据资源（The following code prepare data resources）
                tiers = {"": "", "NONE": "没有段位", "IRON": "坚韧黑铁", "BRONZE": "英勇黄铜", "SILVER": "不屈白银", "GOLD": "荣耀黄金", "PLATINUM": "华贵铂金", "EMERALD": "流光翡翠", "DIAMOND": "璀璨钻石", "MASTER": "超凡大师", "GRANDMASTER": "傲世宗师", "CHALLENGER": "最强王者"}
                #tiers = {"": "", "NONE": "NONE", "IRON": "IRON", "BRONZE": "BRONZE", "SILVER": "SILVER", "GOLD": "GOLD", "PLATINUM": "PLATINUM", "EMERALD": "EMERALD", "DIAMOND": "DIAMOND", "MASTER": "MASTER", "GRANDMASTER": "GRANDMASTER", "CHALLENGER": "CHALLENGER"}
                ratedTiers = {"": "", "NONE": "没有段位", "GRAY": "灰白", "GREEN": "翠绿", "BLUE": "天蓝", "PURPLE": "绛紫", "ORANGE": "耀橙"}
                #ratedTiers = {"": "", "NONE": "NONE", "GRAY": "GRAY", "GREEN": "GREEN", "BLUE": "BLUE", "PURPLE": "PURPLE", "ORANGE": "ORANGE"}
                tiers_all = tiers | ratedTiers
                #下面设置扫描模式的扫描目录（The following code determines the scanning directory for scan mode）
                riot_client_info = await (await connection.request("GET", "/riotclient/command-line-args")).json()
                client_info = {}
                for i in range(len(riot_client_info)):
                    try:
                        client_info[riot_client_info[i].split("=")[0]] = riot_client_info[i].split("=")[1]
                    except IndexError:
                        pass
                region = client_info["--region"]
                if region == "TENCENT":
                    platform_folder = "召唤师信息（Summoner Information）\\" + "国服（TENCENT）" + "\\" + platform_TENCENT[platformId]
                    folder = platform_folder + "\\" + get_info_name(info_body, 2)
                elif region == "GARENA":
                    platform_folder = "召唤师信息（Summoner Information）\\" + "竞舞（GARENA）" + "\\" + platform_GARENA[platformId]
                    folder = platform_folder + "\\" + get_info_name(info_body, 2)
                else: #拳头公司与竞舞娱乐公司的合同于2023年1月终止（In January 2023, Riot Games ended its contract with Garena）
                    platform_folder = "召唤师信息（Summoner Information）\\" + "外服（RIOT）" + "\\" + (platform_RIOT | platform_GARENA)[platformId]
                    folder = platform_folder + "\\" + get_info_name(info_body, 3)
                platform_config_filepath = platform_folder + "\\" + "platform_config_namespaces.json"
                while True:
                    try:
                        with open(platform_config_filepath, "w", encoding = "utf-8") as fp:
                            json.dump(platform_config, fp, indent = 4, ensure_ascii = False)
                    except FileNotFoundError: #这里需要注意是否具有创建文件夹的权限。下同（Pay attention to the authority to create the folder. So are the following）
                        os.makedirs(os.path.dirname(platform_config_filepath), exist_ok = True)
                    else:
                        break
                
                AllAccounts = [info_body] + smurfs if selfDetect and smurfMode else [info_body]
                current_puuid_list = list(map(lambda x: x["puuid"], AllAccounts))
                current_summonerName_list = list(map(lambda x: "" if x["gameName"] == "" and x["tagLine"] == "" else x["gameName"] + "#" + x["tagLine"], AllAccounts))
                #下面获取最近一起玩过的英雄联盟玩家的信息（The following code captures the recently played LoL players' information）
                logPrint("是否查询英雄联盟对局记录？（输入任意键查询，否则不查询）\nSearch LoL matches? (Input anything to search or null to skip searching LoL matches)")
                search_LoL_str = logInput()
                search_LoL = bool(search_LoL_str)
                LoLHistory_dfs = []
                if search_LoL:
                    for info_body in AllAccounts:
                        LoLChampions = copy.deepcopy(LoLChampions_initial) #切换召唤师时应回到最新版本的数据资源查询（When switching summoners, the program should use the latest version of data resources）
                        spells = copy.deepcopy(spells_initial)
                        LoLItems = copy.deepcopy(LoLItems_initial)
                        current_versions["LoLChampion"] = current_versions["spell"] = current_versions["LoLItem"] = URLPatch
                        unmapped_keys["LoLChampion"], unmapped_keys["spell"], unmapped_keys["LoLItem"] = set(), set(), set()
                        #logPrint("召唤师英雄联盟对局记录如下：\nLoL match history is as follows:")
                        LoLHistory_get = False
                        begIndex_get, endIndex_get = 0, 500
                        while True:
                            try:
                                LoLHistory = await (await connection.request("GET", "/lol-match-history/v1/products/lol/%s/matches?begIndex=%d&endIndex=%d" %(info_body["puuid"], begIndex_get, endIndex_get))).json()
                                #logPrint(LoLHistory)
                                error_occurred = False
                                count = 0 #存储内部服务器错误次数（Stores the times of internal server error）
                                if "errorCode" in LoLHistory:
                                    if "500 Internal Server Error" in LoLHistory["message"]:
                                        if not error_occurred:
                                            logPrint("您所在大区的对局记录服务异常。尝试重新获取数据……\nThe match history service provided on your server isn't in place. Trying to recapture the history data ...")
                                            occurred = True
                                        while "errorCode" in LoLHistory and "500 Internal Server Error" in LoLHistory["message"] and count <= 3: #在查询艾欧尼亚和黑色玫瑰大区的对局记录时，有时会产生如下报错：An error when looking up match history on HN1 and HN10 servers might occur as follows: {'errorCode': 'RPC_ERROR', 'httpStatus': 500, 'implementationDetails': {}, 'message': 'Failed due to Error deserializing json response for GET https: //hn1-cloud-acs.lol.qq.com/v1/stats/player_history/HN1/2936900903?begIndex=0&endIndex=500: Error: Invalid value. at offset 0. given body <html>\r\n<head><title>500 Internal Server Error</title></head>\r\n<body bgcolor="white">\r\n<center><h1>500 Internal Server Error</h1></center>\r\n<hr><center>nginx/1.10.0</center>\r\n</body>\r\n</html>\r\n'}
                                            count += 1
                                            logPrint("正在进行第%d次尝试……\nTimes trying: No. %d ..." %(count, count))
                                            LoLHistory = await (await connection.request("GET", "/lol-match-history/v1/products/lol/%s/matches?begIndex=%d&endIndex=%d" %(info_body["puuid"], begIndex_get, endIndex_get))).json()
                                    elif "body was empty" in LoLHistory["message"]:
                                        logPrint("这位召唤师从5月1日起就没有进行过英雄联盟任何对局。\nThis summoner hasn't played any LoL game yet since May 1st.")
                                        break
                                if count > 3:
                                    logPrint("对局记录获取失败！请等待官方修复对局记录服务！\nMatch history capture failure! Please wait for Tencent to fix the match history service!")
                                    break
                                logPrint("玩家%s共进行%d场英雄联盟对局。\nPlayer %s has played %d LoL matches.\n" %(get_info_name(info_body), LoLHistory["games"]["gameCount"], get_info_name(info_body), LoLHistory["games"]["gameCount"])) #在这里引发键异常（Here may trigger a KeyError）
                            except KeyError:
                                logPrint(LoLHistory)
                                LoLHistory_url = "%s/lol-match-history/v1/products/lol/%s/matches?begIndex=0&endIndex=200" %(connection.address, info_body["puuid"])
                                logPrint("请打开以下网址，输入如下所示的用户名和密码，打开后在命令行中按回车键继续（Please open the following website, type in the username and password accordingly and press Enter to continue）：\n网址（URL）：\t\t%s\n用户名（Username）：\triot\n密码（Password）：\t%s\n或者输入空格分隔的两个自然数以重新指定对局索引下限和上限。\nOr submit two nonnegative integers split by space to respecify the begIndex and endIndex." %(LoLHistory_url, connection.auth_key))
                                cont = logInput()
                                if cont == "":
                                    continue
                                else:
                                    try:
                                        begIndex_get, endIndex_get = map(int, cont.split())
                                    except ValueError:
                                        break
                                    else:
                                        continue
                            else:
                                LoLHistory_get = True
                                break
                        if not LoLHistory_get:
                            continue
                        LoLGamePlayed = True #标记该玩家是否进行过英雄联盟对局（Mark whether this summoner has played any LoL game）
                        LoLHistory_header = {"gameIndex": "游戏序号", "endOfGameResult": "对局终止情况", "gameCreation": "对局创建时间戳", "gameCreationDate": "对局创建日期", "gameDuration": "持续时长（秒）", "gameId": "对局序号", "gameMode": "游戏模式", "gameModeMutators": "游戏模式配置", "gameType": "游戏类型", "gameVersion": "对局版本", "mapId": "地图序号", "queueId": "队列序号", "seasonId": "赛季序号", "gameDuration_norm": "持续时长", "gameModeName": "游戏模式名称", "accountId": "帐户序号", "currentAccountId": "当前帐户序号", "currentPlatformId": "当前服务器代码", "gameName": "玩家昵称", "matchHistoryUri": "对局记录网址", "platformId": "服务器代码", "profileIcon": "召唤师图标序号", "puuid": "玩家通用唯一识别码", "summonerId": "召唤师序号", "summonerName": "召唤师名称", "tagLine": "昵称编号", "profileIcon_title": "召唤师图标名称", "profileIcon_imagePath": "召唤师图标路径", "championId": "英雄序号", "highestAchievedSeasonTier": "最高段位", "participantId": "玩家序号", "spell1Id": "召唤师技能1序号", "spell2Id": "召唤师技能2序号", "teamId": "阵营代号", "champion_name": "英雄", "champion_alias": "代号", "champion_squarePortraitPath": "方块头像路径", "spell1_name": "召唤师技能1", "spell2_name": "召唤师技能2", "spell1_iconPath": "召唤师技能1图标", "spell2_iconPath": "召唤师技能2图标", "team_color": "阵营", "assists": "助攻", "causedEarlySurrender": "发起提前投降", "champLevel": "英雄等级", "combatPlayerScore": "战斗得分", "damageDealtToObjectives": "对战略点的总伤害", "damageDealtToTurrets": "对防御塔的总伤害", "damageSelfMitigated": "自我缓和的伤害", "deaths": "死亡", "doubleKills": "双杀", "earlySurrenderAccomplice": "同意提前投降", "firstBloodAssist": "协助获得第一滴血", "firstBloodKill": "第一滴血", "firstInhibitorAssist": "协助摧毁第一座召唤水晶", "firstInhibitorKill": "摧毁第一座召唤水晶", "firstTowerAssist": "协助摧毁第一座塔", "firstTowerKill": "摧毁第一座塔", "gameEndedInEarlySurrender": "提前投降导致比赛结束", "gameEndedInSurrender": "投降导致比赛结束", "goldEarned": "金币", "goldSpent": "金币使用", "inhibitorKills": "摧毁召唤水晶", "item0": "装备1序号", "item1": "装备2序号", "item2": "装备3序号", "item3": "装备4序号", "item4": "装备5序号", "item5": "装备6序号", "item6": "饰品序号", "killingSprees": "大杀特杀", "kills": "击杀", "largestCriticalStrike": "最大暴击伤害", "largestKillingSpree": "最高连杀", "largestMultiKill": "最高多杀", "longestTimeSpentLiving": "最长生存时间", "magicDamageDealt": "造成的魔法伤害", "magicDamageDealtToChampions": "对英雄的魔法伤害", "magicalDamageTaken": "承受的魔法伤害", "neutralMinionsKilled": "击杀野怪", "neutralMinionsKilledEnemyJungle": "击杀敌方野区野怪", "neutralMinionsKilledTeamJungle": "击杀我方野区野怪", "objectivePlayerScore": "战略点玩家得分", "pentaKills": "五杀", "perk0": "符文1序号", "perk0Var1": "符文1：参数1", "perk0Var2": "符文1：参数2", "perk0Var3": "符文1：参数3", "perk1": "符文2序号", "perk1Var1": "符文2：参数1", "perk1Var2": "符文2：参数2", "perk1Var3": "符文2：参数3", "perk2": "符文3序号", "perk2Var1": "符文3：参数1", "perk2Var2": "符文3：参数2", "perk2Var3": "符文3：参数3", "perk3": "符文4序号", "perk3Var1": "符文4：参数1", "perk3Var2": "符文4：参数2", "perk3Var3": "符文4：参数3", "perk4": "符文5序号", "perk4Var1": "符文5：参数1", "perk4Var2": "符文5：参数2", "perk4Var3": "符文5：参数3", "perk5": "符文6序号", "perk5Var1": "符文6：参数1", "perk5Var2": "符文6：参数2", "perk5Var3": "符文6：参数3", "perkPrimaryStyle": "主系序号", "perkSubStyle": "副系序号", "physicalDamageDealt": "造成的物理伤害", "physicalDamageDealtToChampions": "对英雄的物理伤害", "physicalDamageTaken": "承受的物理伤害", "playerAugment1": "强化符文1", "playerAugment2": "强化符文2", "playerAugment3": "强化符文3", "playerAugment4": "强化符文4", "playerAugment5": "强化符文5", "playerAugment6": "强化符文6", "playerScore0": "玩家得分1", "playerScore1": "玩家得分2", "playerScore2": "玩家得分3", "playerScore3": "玩家得分4", "playerScore4": "玩家得分5", "playerScore5": "玩家得分6", "playerScore6": "玩家得分7", "playerScore7": "玩家得分8", "playerScore8": "玩家得分9", "playerScore9": "玩家得分10", "playerSubteamId": "子阵营代号", "quadraKills": "四杀", "sightWardsBoughtInGame": "购买洞察之石", "subteamPlacement": "队伍排名", "teamEarlySurrendered": "队伍提前投降", "timeCCingOthers": "控制得分", "totalDamageDealt": "造成的伤害总和", "totalDamageDealtToChampions": "对英雄的伤害总和", "totalDamageTaken": "承受伤害", "totalHeal": "输出治疗效果", "totalMinionsKilled": "击杀小兵", "totalPlayerScore": "玩家总得分", "totalScoreRank": "总得分排名", "totalTimeCrowdControlDealt": "控制时间", "totalUnitsHealed": "治疗单位数", "tripleKills": "三杀", "trueDamageDealt": "造成真实伤害", "trueDamageDealtToChampions": "对英雄的真实伤害", "trueDamageTaken": "承受的真实伤害", "turretKills": "摧毁防御塔", "unrealKills": "六杀及以上", "visionScore": "视野得分", "visionWardsBoughtInGame": "购买控制守卫", "wardsKilled": "摧毁守卫", "wardsPlaced": "放置守卫", "win": "胜利", "item0_name": "装备1", "item1_name": "装备2", "item2_name": "装备3", "item3_name": "装备4", "item4_name": "装备5", "item5_name": "装备6", "item6_name": "饰品", "item0_iconPath": "装备1图标路径", "item1_iconPath": "装备2图标路径", "item2_iconPath": "装备3图标路径", "item3_iconPath": "装备4图标路径", "item4_iconPath": "装备5图标路径", "item5_iconPath": "装备6图标路径", "item6_iconPath": "饰品图标路径", "perk0EndOfGameStatDescs": "符文1游戏结算数据", "perk1EndOfGameStatDescs": "符文2游戏结算数据", "perk2EndOfGameStatDescs": "符文3游戏结算数据", "perk3EndOfGameStatDescs": "符文4游戏结算数据", "perk4EndOfGameStatDescs": "符文5游戏结算数据", "perk5EndOfGameStatDescs": "符文6游戏结算数据", "perk0_name": "符文1名称", "perk1_name": "符文2名称", "perk2_name": "符文3名称", "perk3_name": "符文4名称", "perk4_name": "符文5名称", "perk5_name": "符文6名称", "perk0_iconPath": "符文1图标路径", "perk1_iconPath": "符文2图标路径", "perk2_iconPath": "符文3图标路径", "perk3_iconPath": "符文4图标路径", "perk4_iconPath": "符文5图标路径", "perk5_iconPath": "符文6图标路径", "perkPrimaryStyle_name": "主系名称", "perkPrimaryStyle_iconPath": "主系图标路径", "perkSubStyle_name": "副系名称", "perkSubStyle_iconPath": "副系图标路径", "playerAugment1_nameTRA": "强化符文1名称", "playerAugment2_nameTRA": "强化符文2名称", "playerAugment3_nameTRA": "强化符文3名称", "playerAugment4_nameTRA": "强化符文4名称", "playerAugment5_nameTRA": "强化符文5名称", "playerAugment6_nameTRA": "强化符文6名称", "playerAugment1_augmentIconPath": "强化符文1图标路径", "playerAugment2_augmentIconPath": "强化符文2图标路径", "playerAugment3_augmentIconPath": "强化符文3图标路径", "playerAugment4_augmentIconPath": "强化符文4图标路径", "playerAugment5_augmentIconPath": "强化符文5图标路径", "playerAugment6_augmentIconPath": "强化符文6图标路径", "playerAugment1_rarity": "强化符文1等级", "playerAugment2_rarity": "强化符文2等级", "playerAugment3_rarity": "强化符文3等级", "playerAugment4_rarity": "强化符文4等级", "playerAugment5_rarity": "强化符文5等级", "playerAugment6_rarity": "强化符文6等级", "playerSubteamColor": "子阵营", "K/D/A": "击杀/死亡/助攻", "KDA": "战损比", "CS": "补刀", "GPM": "分均经济", "GUE": "金币利用率", "CSPM": "分均补刀", "D/G": "伤害转化率", "result": "结果", "lane": "分路", "role": "角色定位"}
                        LoLHistory_header_keys = list(LoLHistory_header.keys())
                        LoLHistory_data = {}
                        games = LoLHistory["games"]["games"]
                        versions = [] #该变量并不是用来呈现在Excel中的，而是用来存储不同装备的合适版本的信息（This variable isn't intended to be displyed in the Excel Sheets. Instead, it stores information of appropriate patches of different patches）
                        if len(games) == 0:
                            LoLGamePlayed = False
                        for i in range(len(LoLHistory_header_keys)):
                            key = LoLHistory_header_keys[i]
                            LoLHistory_data[key] = []
                        for i in range(len(games)):
                            game = games[i]
                            version = game["gameVersion"]
                            version_digits = version.split(".")
                            bigVersion = ".".join(version_digits[:2])
                            stats = game["participants"][0]["stats"]
                            timeline = game["participants"][0]["timeline"]
                            try: #这一部分语句无关紧要（This piece of statements doesn't matter）
                                versions.append(patches_dict[bigVersion][0])
                            except KeyError: #有可能存在美测服的临时版本未收录到DataDragon数据库中。详见patch_compare函数的注释（Possibly an intermediate patch on PBE isn't archived in DataDragon database. More details in the annotation of `patch_compare` function）
                                if patch_compare(bigVersion, latest_patch):
                                    patches_dict[bigVersion] = [FindPostPatch(version, patches)]
                                else:
                                    patches_dict[bigVersion] = [latest_patch]
                                versions.append(patches_dict[bigVersion][0])
                            #下面针对每场对局建立总的数据资源异常处理机制（Builds the summarized data resource exceptional handling mechanism for each match）
                            ##召唤师图标（Summoner icon）
                            summonerIconIds_match_list = sorted(set(map(lambda x: x["player"]["profileIcon"], game["participantIdentities"])))
                            for j in summonerIconIds_match_list:
                                if not j in summonerIcons and current_versions["summonerIcon"] != bigVersion:
                                    summonerIconPatch_adopted = bigVersion
                                    summonerIcon_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）召唤师图标信息（%d）获取失败！正在第%d次尝试改用%s版本的召唤师图标信息……\nSummoner icon information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to summoner icons of Patch %s ... Times tried: %d." %(i + 1, len(games), game["gameId"], j, summonerIcon_recapture, summonerIconPatch_adopted, j, i + 1, len(games), game["gameId"], summonerIconPatch_adopted, summonerIcon_recapture))
                                    while True:
                                        try:
                                            summonerIcon = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-icons.json" %(summonerIconPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            summonerIconPatch_deserted = summonerIconPatch_adopted
                                            summonerIconPatch_adopted = FindPostPatch(summonerIconPatch_adopted, bigPatches)
                                            summonerIcon_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to LoL champions of Patch %s ... Times tried: %d." %(summonerIconPatch_deserted, summonerIcon_recapture, summonerIconPatch_adopted, summonerIconPatch_deserted, summonerIconPatch_adopted, summonerIcon_recapture))
                                        except requests.exceptions.RequestException:
                                            if summonerIcon_recapture < 3:
                                                summonerIcon_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的召唤师图标信息……\nYour network environment is abnormal! Changing to summoner icons of Patch %s ... Times tried: %d." %(summonerIcon_recapture, summonerIconPatch_adopted, summonerIconPatch_adopted, summonerIcon_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的召唤师图标信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the summoner icon (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(games), game["gameId"], j, j, i + 1, len(games), game["gameId"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的召唤师图标信息。\nSummoner icon information changed to Patch %s." %(summonerIconPatch_adopted, summonerIconPatch_adopted))
                                            summonerIcons = {}
                                            for summonerIcon_iter in summonerIcon:
                                                summonerIcon_id = summonerIcon_iter["id"]
                                                summonerIcons[summonerIcon_id] = summonerIcon_iter
                                            current_versions["summonerIcon"] = summonerIconPatch_adopted
                                            unmapped_keys["summonerIcon"].clear()
                                            break
                                    break
                            ##英雄：包含选用英雄和禁用英雄（LoL champions, which contain picked and banned ones）
                            LoLChampionIds_match_list = sorted(set(map(lambda x: x["championId"], game["participants"])))
                            for j in LoLChampionIds_match_list:
                                if not j in LoLChampions and current_versions["LoLChampion"] != bigVersion:
                                    LoLChampionPatch_adopted = bigVersion
                                    LoLChampion_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）英雄信息（%d）获取失败！正在第%d次尝试改用%s版本的英雄信息……\nLoL champion information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to LoL champions of Patch %s ... Times tried: %d." %(i + 1, len(games), game["gameId"], j, LoLChampion_recapture, LoLChampionPatch_adopted, j, i + 1, len(games), game["gameId"], LoLChampionPatch_adopted, LoLChampion_recapture))
                                    while True:
                                        try:
                                            LoLChampion = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/champion-summary.json" %(LoLChampionPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            LoLChampionPatch_deserted = LoLChampionPatch_adopted
                                            LoLChampionPatch_adopted = FindPostPatch(LoLChampionPatch_adopted, bigPatches)
                                            LoLChampion_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to LoL champions of Patch %s ... Times tried: %d." %(LoLChampionPatch_deserted, LoLChampion_recapture, LoLChampionPatch_adopted, LoLChampionPatch_deserted, LoLChampionPatch_adopted, LoLChampion_recapture))
                                        except requests.exceptions.RequestException:
                                            if LoLChampion_recapture < 3:
                                                LoLChampion_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的英雄信息……\nYour network environment is abnormal! Changing to LoL champions of Patch %s ... Times tried: %d." %(LoLChampion_recapture, LoLChampionPatch_adopted, LoLChampionPatch_adopted, LoLChampion_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的英雄信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the LoL champion (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(games), game["gameId"], j, j, i + 1, len(games), game["gameId"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的英雄信息。\nLoL champion information changed to Patch %s." %(LoLChampionPatch_adopted, LoLChampionPatch_adopted))
                                            LoLChampions = {}
                                            for LoLChampion_iter in LoLChampion:
                                                LoLChampion_id = LoLChampion_iter["id"]
                                                LoLChampions[LoLChampion_id] = LoLChampion_iter
                                            current_versions["LoLChampion"] = LoLChampionPatch_adopted
                                            unmapped_keys["LoLChampion"].clear() #切换版本时，未对应的键应当清空。下同（When the version is switched, the unmapped keys should be cleared. This applies to other data resources）
                                            break
                                    break #切换版本只需一次即可。如果对局版本还不对，那就不用再找下去了（The version of data resources only needs changing once. If data resources of the version of this match don't match all the game data, then there's no need of retrying）
                            ##召唤师技能（Summoner spells）
                            spellIds_match_list = sorted(set(map(lambda x: x["spell1Id"], game["participants"])) | set(map(lambda x: x["spell2Id"], game["participants"])))
                            for j in spellIds_match_list:
                                if not j in spells and current_versions["spell"] != bigVersion and j != 0: #需要注意电脑玩家的召唤师技能序号都是0（Note that Spell Ids of bot players are both 0s）
                                    spellPatch_adopted = bigVersion
                                    spell_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）召唤师技能信息（%d）获取失败！正在第%d次尝试改用%s版本的召唤师技能信息……\nSpell information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to spells of Patch %s ... Times tried: %d." %(i + 1, len(games), game["gameId"], j, spell_recapture, spellPatch_adopted, j, i + 1, len(games), game["gameId"], spellPatch_adopted, spell_recapture))
                                    while True:
                                        try:
                                            spell = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-spells.json" %(spellPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            spellPatch_deserted = spellPatch_adopted
                                            spellPatch_adopted = FindPostPatch(spellPatch_adopted, bigPatches)
                                            spell_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to spells of Patch %s ... Times tried: %d." %(spellPatch_deserted, spell_recapture, spellPatch_adopted, spellPatch_deserted, spellPatch_adopted, spell_recapture))
                                        except requests.exceptions.RequestException:
                                            if spell_recapture < 3:
                                                spell_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的召唤师技能信息……\nYour network environment is abnormal! Changing to spells of Patch %s ... Times tried: %d." %(spell_recapture, spellPatch_adopted, spellPatch_adopted, spell_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的召唤师技能信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the spell (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(games), game["gameId"], j, j, i + 1, len(games), game["gameId"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的召唤师技能信息。\nSpell information changed to Patch %s." %(spellPatch_adopted, spellPatch_adopted))
                                            spells = {}
                                            for spell_iter in spell:
                                                spell_id = spell_iter["id"]
                                                spells[spell_id] = spell_iter
                                            current_versions["spell"] = spellPatch_adopted
                                            unmapped_keys["spell"].clear()
                                            break
                                    break
                            ##英雄联盟装备（LoL items）
                            LoLItemIds_match_list = sorted(set(map(lambda x: x["stats"]["item0"], game["participants"])) | set(map(lambda x: x["stats"]["item1"], game["participants"])) | set(map(lambda x: x["stats"]["item2"], game["participants"])) | set(map(lambda x: x["stats"]["item3"], game["participants"])) | set(map(lambda x: x["stats"]["item4"], game["participants"])) | set(map(lambda x: x["stats"]["item5"], game["participants"])) | set(map(lambda x: x["stats"]["item6"], game["participants"])))
                            for j in LoLItemIds_match_list:
                                if not j in LoLItems and current_versions["LoLItem"] != bigVersion and j != 0: #空装备序号是0（The itemId of an empty item is 0）
                                    LoLItemPatch_adopted = bigVersion
                                    LoLItem_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）英雄联盟装备信息（%d）获取失败！正在第%d次尝试改用%s版本的英雄联盟装备信息……\nLoL item information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to LoL items of Patch %s ... Times tried: %d." %(i + 1, len(games), game["gameId"], j, LoLItem_recapture, LoLItemPatch_adopted, j, i + 1, len(games), game["gameId"], LoLItemPatch_adopted, LoLItem_recapture))
                                    while True:
                                        try:
                                            LoLItem = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/items.json" %(LoLItemPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            LoLItemPatch_deserted = LoLItemPatch_adopted
                                            LoLItemPatch_adopted = FindPostPatch(LoLItemPatch_adopted, bigPatches)
                                            LoLItem_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to LoL items of Patch %s ... Times tried: %d." %(LoLItemPatch_deserted, LoLItem_recapture, LoLItemPatch_adopted, LoLItemPatch_deserted, LoLItemPatch_adopted, LoLItem_recapture))
                                        except requests.exceptions.RequestException:
                                            if LoLItem_recapture < 3:
                                                LoLItem_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的英雄联盟装备信息……\nYour network environment is abnormal! Changing to LoL items of Patch %s ... Times tried: %d." %(LoLItem_recapture, LoLItemPatch_adopted, LoLItemPatch_adopted, LoLItem_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的英雄联盟装备信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the LoL item (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(games), game["gameId"], j, j, i + 1, len(games), game["gameId"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的英雄联盟装备信息。\nLoL item information changed to Patch %s." %(LoLItemPatch_adopted, LoLItemPatch_adopted))
                                            LoLItems = {}
                                            for LoLItem_iter in LoLItem:
                                                LoLItem_id = LoLItem_iter["id"]
                                                LoLItems[LoLItem_id] = LoLItem_iter
                                            current_versions["LoLItem"] = LoLItemPatch_adopted
                                            unmapped_keys["LoLItem"].clear()
                                            break
                                    break
                            ##符文（Perks）
                            perkIds_match_list = sorted(set(perk for s in [set(map(lambda x: x["stats"]["perk" + str(i)], game["participants"])) for i in range(6)] for perk in s))
                            for j in perkIds_match_list:
                                if not j in perks and current_versions["perk"] != bigVersion and j != 0: #在一些非常规模式（如新手训练）的对局中，玩家可能没有携带任何符文（In matches with unconventional game mode (e.g. TUTORIAL), maybe the player doesn't take any runes）
                                    perkPatch_adopted = bigVersion
                                    perk_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）基石符文信息（%d）获取失败！正在第%d次尝试改用%s版本的基石符文信息……\nPerk information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to perks of Patch %s ... Times tried: %d." %(i + 1, len(games), game["gameId"], j, perk_recapture, perkPatch_adopted, j, i + 1, len(games), game["gameId"], perkPatch_adopted, perk_recapture))
                                    while True:
                                        try:
                                            perk = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perks.json" %(perkPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            perkPatch_deserted = perkPatch_adopted
                                            perkPatch_adopted = FindPostPatch(perkPatch_adopted, bigPatches)
                                            perk_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to perks of Patch %s ... Times tried: %d." %(perkPatch_deserted, perk_recapture, perkPatch_adopted, perkPatch_deserted, perkPatch_adopted, perk_recapture))
                                        except requests.exceptions.RequestException:
                                            if perk_recapture < 3:
                                                perk_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的基石符文信息……\nYour network environment is abnormal! Changing to perks of Patch %s ... Times tried: %d." %(perk_recapture, perkPatch_adopted, perkPatch_adopted, perk_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的基石符文信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the perk (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(games), game["gameId"], j, j, i + 1, len(games), game["gameId"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的基石符文信息。\nPerk information changed to Patch %s." %(perkPatch_adopted, perkPatch_adopted))
                                            perks = {}
                                            for perk_iter in perk:
                                                perk_id = perk_iter["id"]
                                                perks[perk_id] = perk_iter
                                            current_versions["perk"] = perkPatch_adopted
                                            unmapped_keys["perk"].clear()
                                            break
                                    break
                            ##符文系（Perkstyles）
                            perkstyleIds_match_list = sorted(list(set(map(lambda x: x["stats"]["perkPrimaryStyle"], game["participants"])) | set(map(lambda x: x["stats"]["perkSubStyle"], game["participants"]))))
                            for j in perkstyleIds_match_list:
                                if not j in perkstyles and current_versions["perkstyle"] != bigVersion and j != 0: #在一些非常规模式（如新手训练）的对局中，玩家可能没有携带任何符文（In matches with unconventional game mode (e.g. TUTORIAL), maybe the player doesn't take any runes）
                                    perkstylePatch_adopted = bigVersion
                                    perkstyle_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）符文系信息（%d）获取失败！正在第%d次尝试改用%s版本的符文系信息……\nPerkstyle information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to perkstyles of Patch %s ... Times tried: %d." %(i + 1, len(games), game["gameId"], j, perkstyle_recapture, perkstylePatch_adopted, j, i + 1, len(games), game["gameId"], perkstylePatch_adopted, perkstyle_recapture))
                                    while True:
                                        try:
                                            perkstyle = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perkstyles.json" %(perkstylePatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            perkstylePatch_deserted = perkstylePatch_adopted
                                            perkstylePatch_adopted = FindPostPatch(perkstylePatch_adopted, bigPatches)
                                            perkstyle_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to perks of Patch %s ... Times tried: %d." %(perkstylePatch_deserted, perkstyle_recapture, perkstylePatch_adopted, perkstylePatch_deserted, perkstylePatch_adopted, perkstyle_recapture))
                                        except requests.exceptions.RequestException:
                                            if perkstyle_recapture < 3:
                                                perkstyle_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的符文系信息……\nYour network environment is abnormal! Changing to perkstyles of Patch %s ... Times tried: %d." %(perkstyle_recapture, perkstylePatch_adopted, perkstylePatch_adopted, perkstyle_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的符文系信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the perkstyle (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(games), game["gameId"], j, j, i + 1, len(games), game["gameId"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的符文系信息。\nPerkstyle information changed to Patch %s." %(perkstylePatch_adopted, perkstylePatch_adopted))
                                            perkstyles = {}
                                            for perkstyle_iter in perkstyle["styles"]:
                                                perkstyle_id = perkstyle_iter["id"]
                                                perkstyles[perkstyle_id] = perkstyle_iter
                                            current_versions["perkstyle"] = perkstylePatch_adopted
                                            unmapped_keys["perkstyle"].clear()
                                            break
                                    break
                            ##斗魂竞技场强化符文（Cherry augments）
                            CherryAugmentIds_match_list = sorted(set(augment for s in [set(map(lambda x: x["stats"]["playerAugment" + str(i)], game["participants"])) for i in range(1, 7)] for augment in s))
                            for j in CherryAugmentIds_match_list:
                                if not j in CherryAugments and current_versions["CherryAugment"] != bigVersion and j != 0:
                                    CherryAugmentPatch_adopted = bigVersion
                                    CherryAugment_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）强化符文信息（%d）获取失败！正在第%d次尝试改用%s版本的斗魂竞技场强化符文信息……\nAugment information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to Cherry augments of Patch %s ... Times tried: %d." %(i + 1, len(games), game["gameId"], j, CherryAugment_recapture, CherryAugmentPatch_adopted, j, i + 1, len(games), game["gameId"], CherryAugmentPatch_adopted, CherryAugment_recapture))
                                    while True:
                                        try:
                                            CherryAugment = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/cherry-augments.json" %(CherryAugmentPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            CherryAugmentPatch_deserted = CherryAugmentPatch_adopted
                                            CherryAugmentPatch_adopted = FindPostPatch(CherryAugmentPatch_adopted, bigPatches)
                                            CherryAugment_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to Cherry augments of Patch %s ... Times tried: %d." %(CherryAugmentPatch_deserted, CherryAugment_recapture, CherryAugmentPatch_adopted, CherryAugmentPatch_deserted, CherryAugmentPatch_adopted, CherryAugment_recapture))
                                        except requests.exceptions.RequestException:
                                            if CherryAugment_recapture < 3:
                                                CherryAugment_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的斗魂竞技场强化符文信息……\nYour network environment is abnormal! Changing to Cherry augments of Patch %s ... Times tried: %d." %(CherryAugment_recapture, CherryAugmentPatch_adopted, CherryAugmentPatch_adopted, CherryAugment_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的强化符文信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the Cherry augment (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(games), game["gameId"], j, j, i + 1, len(games), game["gameId"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的斗魂竞技场强化符文信息。\nCherry augment information changed to Patch %s." %(CherryAugmentPatch_adopted, CherryAugmentPatch_adopted))
                                            CherryAugments = {}
                                            for CherryAugment_iter in CherryAugment:
                                                CherryAugment_id = CherryAugment_iter["id"]
                                                CherryAugments[CherryAugment_id] = CherryAugment_iter
                                            current_versions["CherryAugment"] = CherryAugmentPatch_adopted
                                            unmapped_keys["CherryAugment"].clear()
                                            break
                                    break
                            #下面开始整理数据（Sorts out the data）
                            for j in range(len(LoLHistory_header_keys)):
                                key = LoLHistory_header_keys[j]
                                if j == 0:
                                    LoLHistory_data[key].append(i + 1)
                                elif j <= 14:
                                    if j == 1: #对局终止情况（`endOfGameResult`）
                                        LoLHistory_data[key].append(endOfGameResults[game["endOfGameResult"]])
                                    if j == 3: #对局创建日期（`gameCreationDate`）
                                        LoLHistory_data[key].append(game["gameCreationDate"][:10] + " " + game["gameCreationDate"][11:23])
                                    elif j == 8: #游戏类型（`gameType`）
                                        LoLHistory_data[key].append(gameTypes[game[key]])
                                    elif j == 13: #持续时长（`gameDuration_norm`）
                                        LoLHistory_data[key].append(str(game["gameDuration"] // 60) + ":" + "%02d" %(game["gameDuration"] % 60))
                                    elif j == 14: #游戏模式名称（`gameModeName`）
                                        LoLHistory_data[key].append("自定义" if game["queueId"] == 0 else gamemodes[game["queueId"]]["name"] if game["queueId"] in gamemodes else "")
                                    else:
                                        LoLHistory_data[key].append(game[key])
                                elif j <= 27:
                                    if j >= 26: #召唤师图标相关键（Summoner icon-related keys）
                                        profileIconId = game["participantIdentities"][0]["player"]["profileIcon"]
                                        if profileIconId in summonerIcons:
                                            try:
                                                LoLHistory_data[key].append(summonerIcons[profileIconId][key.split("_")[-1]])
                                            except KeyError:
                                                traceback_info = traceback.format_exc()
                                                logPrint(traceback_info)
                                                LoLHistory_data[key].append("")
                                        elif profileIconId in summonerIcons_initial:
                                            try:
                                                LoLHistory_data[key].append(summonerIcons_initial[profileIconId][key.split("_")[-1]])
                                            except KeyError:
                                                traceback_info = traceback.format_exc()
                                                logPrint(traceback_info)
                                                LoLHistory_data[key].append("")
                                        else:
                                            if not profileIconId in unmapped_keys["summonerIcon"]:
                                                unmapped_keys["summonerIcon"].add(profileIconId)
                                                logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）召唤师图标信息（%d）获取失败！将采用原始数据！\n[%d. %s] Summoner icon information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, profileIconId, j, key, profileIconId, i + 1, len(games), game["gameId"], version))
                                            LoLHistory_data[key].append(profileIconId if j == 26 else "")
                                    else:
                                        LoLHistory_data[key].append(game["participantIdentities"][0]["player"][key])
                                elif j <= 41:
                                    if j == 29: #最高段位（`highestAchievedSeasonTier`）
                                        LoLHistory_data[key].append(tiers[game["participants"][0]["highestAchievedSeasonTier"]])
                                    elif j >= 34 and j <= 36: #英雄相关键（Champion-related keys）
                                        championId = game["participants"][0][key.split("_")[0] + "Id"]
                                        if championId in LoLChampions:
                                            LoLHistory_data[key].append(LoLChampions[championId][key.split("_")[1]])
                                        elif championId in LoLChampions_initial: #一些旧版的键可能出现在最新的数据资源中，而随着程序的进行，程序所使用的数据资源可能并不是最新的，因此再和最新的数据资源做一次比较。这种现象在云顶之弈对局中尤其普遍（Some old-version keys might appear in the latest data resource, while as the program executes, the data resource it uses may not be the latest. Therefore, a lookup in the latest data resource is performed. This phenomenon is especially normal in TFT matches）
                                            LoLHistory_data[key].append(LoLChampions_initial[championId][key.split("_")[1]])
                                        else: #在国服体验服的对局序号为696083511的对局中，出现了英雄序号为37225015（In a match with matchId 696083511 on Chinese PBE, there's a champion with championId 37225015）
                                            if not championId in unmapped_keys["LoLChampion"]:
                                                unmapped_keys["LoLChampion"].add(championId)
                                                logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）英雄信息（%d）获取失败！将采用原始数据！\n[%d. %s] Champion information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, championId, j, key, championId, i + 1, len(games), game["gameId"], version))
                                            LoLHistory_data[key].append(championId if j == 34 else "")
                                    elif j >= 37 and j <= 40: #召唤师技能相关键（Summoner spell-related keys）
                                        spellId = game["participants"][0][key.split("_")[0] + "Id"]
                                        if spellId in spells:
                                            LoLHistory_data[key].append(spells[spellId][key.split("_")[1]])
                                        elif spellId in spells_initial:
                                            LoLHistory_data[key].append(spells_initial[spellId][key.split("_")[1]])
                                        else:
                                            if not spellId in unmapped_keys["spell"]:
                                                unmapped_keys["spell"].add(spellId)
                                                logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）召唤师技能信息（%d）获取失败！将采用原始数据！\n[%d. %s] Spell information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, spellId, j, key, spellId, i + 1, len(games), game["gameId"], version))
                                            LoLHistory_data[key].append(spellId if j <= 38 else "")
                                    elif j == 41: #阵营（`team_color`）
                                        LoLHistory_data[key].append(team_color[game["participants"][0]["teamId"]])
                                    else:
                                        LoLHistory_data[key].append(game["participants"][0][key])
                                elif j <= 217:
                                    if j >= 155 and j <= 168: #英雄联盟装备相关键（LoLItems-related keys）
                                        itemId = stats[key.split("_")[0]]
                                        if itemId == 0:
                                            LoLHistory_data[key].append("")
                                        elif itemId in LoLItems:
                                            LoLHistory_data[key].append(LoLItems[itemId][key.split("_")[-1]])
                                        elif itemId in LoLItems_initial:
                                            LoLHistory_data[key].append(LoLItems_initial[itemId][key.split("_")[-1]])
                                        else:
                                            if not itemId in unmapped_keys["LoLItem"]:
                                                unmapped_keys["LoLItem"].add(itemId)
                                                logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）装备信息（%d）获取失败！将采用原始数据！\n[%d. %s] LoL item information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, itemId, j, key, itemId, i + 1, len(games), game["gameId"], version))
                                            LoLHistory_data[key].append(itemId if j <= 161 else "")
                                    elif j >= 169 and j <= 186: #符文相关键（Perks-related keys）
                                        if j <= 174:
                                            perkId = stats[key[:5]]
                                            if perkId == 0:
                                                LoLHistory_data[key].append("")
                                            elif perkId in perks:
                                                perk_EndOfGameStatDescs = "".join(list(map(lambda x: x + "。", perks[perkId]["endOfGameStatDescs"])))
                                                perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar1@", str(stats[key[:5] + "Var1"]))
                                                perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar2@", str(stats[key[:5] + "Var2"]))
                                                perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar3@", str(stats[key[:5] + "Var3"]))
                                                LoLHistory_data[key].append(perk_EndOfGameStatDescs)
                                            elif perkId in perks_initial:
                                                perk_EndOfGameStatDescs = "".join(list(map(lambda x: x + "。", perks_initial[perkId]["endOfGameStatDescs"])))
                                                perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar1@", str(stats[key[:5] + "Var1"]))
                                                perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar2@", str(stats[key[:5] + "Var2"]))
                                                perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar3@", str(stats[key[:5] + "Var3"]))
                                                LoLHistory_data[key].append(perk_EndOfGameStatDescs)
                                            else:
                                                if not perkId in unmapped_keys["perk"]:
                                                    unmapped_keys["perk"].add(perkId)
                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）符文信息（%d）获取失败！将采用原始数据！\n[%d. %s] Runes information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, perkId, j, key, perkId, i + 1, len(games), game["gameId"], version))
                                                LoLHistory_data[key].append("")
                                        else:
                                            perkId = stats[key.split("_")[0]]
                                            if perkId == 0:
                                                LoLHistory_data[key].append("")
                                            elif perkId in perks:
                                                LoLHistory_data[key].append(perks[perkId][key.split("_")[-1]])
                                            elif perkId in perks_initial:
                                                LoLHistory_data[key].append(perks_initial[perkId][key.split("_")[-1]])
                                            else:
                                                if not perkId in unmapped_keys["perk"]:
                                                    unmapped_keys["perk"].add(perkId)
                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）符文信息（%d）获取失败！将采用原始数据！\n[%d. %s] Runes information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, perkId, j, key, perkId, i + 1, len(games), game["gameId"], version))
                                                LoLHistory_data[key].append(perkId if j <= 180 else "")
                                    elif j >= 187 and j <= 190: #符文系相关键（Perkstyles-related keys）
                                        perkstyleId = stats[key.split("_")[0]]
                                        if perkstyleId == 0:
                                            LoLHistory_data[key].append("")
                                        elif perkstyleId in perkstyles:
                                            LoLHistory_data[key].append(perkstyles[perkstyleId][key.split("_")[-1]])
                                        elif perkstyleId in perkstyles_initial:
                                            LoLHistory_data[key].append(perkstyles_initial[perkstyleId][key.split("_")[-1]])
                                        else:
                                            if not perkstyleId in unmapped_keys["perkstyle"]:
                                                unmapped_keys["perkstyle"].add(perkstyleId)
                                                logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）符文系信息（%d）获取失败！将采用原始数据！\n[%d. %s] Perkstyle information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, perkstyleId, j, key, perkstyleId, i + 1, len(games), game["gameId"], version))
                                            LoLHistory_data[key].append(perkstyleId if (j - 187) % 2 == 0 else "")
                                    elif j >= 191 and j <= 208: #强化符文相关键（Augment-related keys）
                                        CherryAugmentId = stats[key.split("_")[0]]
                                        if CherryAugmentId == 0:
                                            LoLHistory_data[key].append("")
                                        elif CherryAugmentId in CherryAugments:
                                            if j <= 196: #强化符文名称（`nameTRA`）
                                                LoLHistory_data[key].append(CherryAugments[CherryAugmentId][key.split("_")[-1]])
                                            elif j <= 202: #强化符文图标路径（`augmentIconPath`）
                                                LoLHistory_data[key].append(CherryAugments[CherryAugmentId]["augmentSmallIconPath"].replace("_small.png", "_large.png"))
                                            else: #强化符文等级（`rarity`）
                                                LoLHistory_data[key].append(augment_rarity[CherryAugments[CherryAugmentId][key.split("_")[-1]]])
                                        elif CherryAugmentId in CherryAugments_initial:
                                            if j <= 196: #强化符文名称（`nameTRA`）
                                                LoLHistory_data[key].append(CherryAugments_initial[CherryAugmentId][key.split("_")[-1]])
                                            elif j <= 202: #强化符文图标路径（`augmentIconPath`）
                                                LoLHistory_data[key].append(CherryAugments_initial[CherryAugmentId]["augmentSmallIconPath"].replace("_small.png", "_large.png"))
                                            else: #强化符文等级（`rarity`）
                                                LoLHistory_data[key].append(augment_rarity[CherryAugments_initial[CherryAugmentId][key.split("_")[-1]]])
                                        else:
                                            if not CherryAugmentId in unmapped_keys["CherryAugment"]:
                                                unmapped_keys["CherryAugment"].add(CherryAugmentId)
                                                logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）强化符文信息（%d）获取失败！将采用原始数据！\n[%d. %s] Cherry augment information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(games), game["gameId"], version, CherryAugmentId, j, key, CherryAugmentId, i + 1, len(games), game["gameId"], version))
                                            LoLHistory_data[key].append(CherryAugmentId if j <= 196 else "")
                                    elif j == 209: #子阵营（`playerSubteam_color`）
                                        LoLHistory_data[key].append(subteam_color[stats["playerSubteamId"]])
                                    elif j == 210: #击杀/死亡/助攻（`K/D/A`）
                                        LoLHistory_data[key].append("/".join([str(stats["kills"]), str(stats["deaths"]), str(stats["assists"])]))
                                    elif j == 211: #战损比（`KDA`）
                                        LoLHistory_data[key].append((stats["kills"] + stats["assists"]) / max(1, stats["deaths"]))
                                    elif j == 212: #补刀（`CS`）
                                        LoLHistory_data[key].append(stats["neutralMinionsKilled"] + stats["totalMinionsKilled"])
                                    elif j == 213: #分均经济（`GPM`）
                                        LoLHistory_data[key].append(0 if game["gameDuration"] == 0 else stats["goldEarned"] * 60 / game["gameDuration"])
                                    elif j == 214: #金币利用率（`GUE` - Gold Utilization Efficiency）
                                        LoLHistory_data[key].append(0 if stats["goldEarned"] == 0 else stats["goldSpent"] / stats["goldEarned"])
                                    elif j == 215: #分均补刀（`CSPM`）
                                        LoLHistory_data[key].append(0 if game["gameDuration"] == 0 else (stats["neutralMinionsKilled"] + stats["totalMinionsKilled"]) * 60 / game["gameDuration"])
                                    elif j == 216: #伤害转化率（`D/G`）
                                        LoLHistory_data[key].append(0 if stats["goldEarned"] == 0 else stats["totalDamageDealtToChampions"] / stats["goldEarned"])
                                    elif j == 217: #胜负（`result`）
                                        LoLHistory_data[key].append("胜利" if stats["win"] else "失败")
                                    else:
                                        LoLHistory_data[key].append(stats[key])
                                else: #时间轴相关键（Timeline-related keys）
                                    LoLHistory_data[key].append(lanes[timeline[key]] if j == 218 else roles[timeline[key]])
                            print("对局记录查询进度（Match history query process）：%d/%d\t对局序号（MatchId）：%d" %(i + 1, len(games), game["gameId"]), end = "\r")
                        LoLHistory_statistics_output_order = [0, 24, 5, 3, 13, 11, 6, 14, 10, 9, 34, 35, 44, 37, 38, 155, 156, 157, 158, 159, 160, 161, 210, 212, 60, 217, 132]
                        LoLHistory_data_organized = {}
                        for i in LoLHistory_statistics_output_order:
                            key = LoLHistory_header_keys[i]
                            LoLHistory_data_organized[key] = LoLHistory_data[key]
                        LoLHistory_df = pandas.DataFrame(data = LoLHistory_data_organized)
                        LoLHistory_df = pandas.concat([pandas.DataFrame([LoLHistory_header])[LoLHistory_df.columns], LoLHistory_df], ignore_index = True)
                        LoLHistory_dfs.append(LoLHistory_df)
                        #LoLHistory_df.apply(lambda x: pandas.Series([-3], index = ["K/D/A"]))
                        if LoLGamePlayed:
                            logPrint(LoLHistory_df[:min(21, len(games) + 1)], write_time = False)
                    LoLHistory_df_all = pandas.concat([LoLHistory_dfs[0].iloc[:1]] + list(map(lambda x: x.iloc[1:], LoLHistory_dfs)), ignore_index = True) #需要注意数据框的中文表头占用了一行（Note that the Chinese header takes up a record）
                    #按对局序号去重（Drop duplicates according to gameId）
                    gameIds_occurred = {LoLHistory_df_all.loc[0, "gameId"]}
                    lines_to_drop = []
                    for i in range(1, len(LoLHistory_df_all)):
                        if LoLHistory_df_all.loc[i, "gameId"] in gameIds_occurred:
                            lines_to_drop.append(i)
                        else:
                            gameIds_occurred.add(LoLHistory_df_all.loc[i, "gameId"])
                    LoLHistory_df_all.drop(lines_to_drop, inplace = True)
                    LoLHistory_df_all = LoLHistory_df_all.reset_index(drop = True)
                    LoLHistory_df_all = pandas.concat([LoLHistory_df_all.iloc[:1], LoLHistory_df_all.iloc[1:].sort_values(by = "gameCreationDate", ascending = False)], ignore_index = True) #这里弃用了根据对局序号排序（Here gameId isn't used to sort the values）
                    
                    #下面获取最近一起玩过的英雄联盟玩家的信息（The following code captures the recently played LoL players' information）
                    logPrint('请输入要查询的英雄联盟对局序号，批量查询对局请输入对局序号列表，批量查询全部对局请输入“3”，退出英雄联盟对局查询“0”：\nPlease enter the LoL match ID to check. Submit a list containing matchIDs to search in batches. Submit "3" to search the currently stored history in batches. Submit "0" to quit searching for LoL matches.')
                    gameIds = []
                    for gameId in LoLHistory_df_all.loc[1:, "gameId"]:
                        if not gameId in gameIds:
                            gameIds.append(gameId)
                    while True:
                        matchID = logInput()
                        if matchID == "":
                            continue
                        elif matchID == "0":
                            search_LoL = False
                            break
                        else:
                            if matchID == "3":
                                logPrint("请设置需要查询的对局索引下界和上界，以空格为分隔符（输入空字符以默认查询近20场对局）：\nPlease set the begIndex and endIndex of the matches to be searched, split by space (Enter an empty string to search for the recent 20 matches):") #在13.13版本以前，腾讯代理的服务器只支持近20场对局查询（Before Patch 13.13, Tencent servers only provide search of the latest 20 matches）
                                while True:
                                    gameIndex = logInput()
                                    if gameIndex == "":
                                        begIndex, endIndex = 0, 20 * len(AllAccounts)
                                    elif gameIndex == "0":
                                        break
                                    else:
                                        try:
                                            begIndex, endIndex = map(int, gameIndex.split())
                                        except ValueError:
                                            logPrint("请以空格为分隔符输入对局索引的自然数类型的下界和上界！\nPlease enter the two nonnegative integers as the begIndex and endIndex of the matches split by space!")
                                            continue
                                    break
                                if gameIndex == "0":
                                    search_LoL = False
                                    break
                                LoLMatchIDs = gameIds[begIndex:endIndex]
                            elif matchID == "scan":
                                LoLMatchIDs = gameIds
                                filenames = os.listdir(folder)
                                for filename in filenames:
                                    if filename.startswith("Match Information (LoL) - "):
                                        LoLMatchIDs.append(int(filename.split("-")[-1].split(".")[0]))
                                if LoLMatchIDs == list():
                                    logPrint("尚未保存过该玩家的数据！\nYou haven't saved this summoner's matches yet!\n")
                                    break
                                else:
                                    LoLMatchIDs = sorted(set(LoLMatchIDs), reverse = True)
                                    logPrint("检测到%d场对局。是否继续？（输入任意键以重新输入要查询的对局序号，否则重新获取这些对局的数据）\nDetected %d matches. Continue? (Input any nonempty string to return to the last step of inputting the matchID, or null to recapture those matches' data)" %(len(LoLMatchIDs), len(LoLMatchIDs)))
                                    recapture_str = logInput()
                                    recapture = bool(recapture_str)
                                    if recapture:
                                        LoLMatchIDs = [] #如果没有这句语句，那么当重新输入对局序号列表时，从本地文件中检测到的对局数量相比上次检测数的基础上会多出本地文件中包含的对局的数量（Without this assignment, when reinputting the matchID list, the number of matches detected from the local files will become more than that of the last time's check）
                                        logPrint('请输入要查询的英雄联盟对局序号，批量查询对局请输入对局序号列表，批量查询全部对局请输入“3”，退出英雄联盟对局查询“0”：\nPlease enter the LoL match ID to check. Submit a list containing matchIDs to search in batches. Submit "3" to search the currently stored history in batches. Submit "0" to quit searching for LoL matches.')
                                        continue
                                    #在沿用查战绩脚本时，后续对局记录重新生成的代码不再需要了。因为这只是查召唤师信息的脚本，不是查对局记录的脚本（When inheritting code from Customized Program 5, the following code to regenerate match history is no longer needed. That's because this program is just designed to search for recently played summoners, rather than sort out match history）
                            else:
                                try:
                                    matchID = eval(matchID)
                                    LoLMatchIDs = []
                                    if isinstance(matchID, int):
                                        LoLMatchIDs.append(matchID)
                                    elif isinstance(matchID, list):
                                        for match in matchID:
                                            if isinstance(match, int):
                                                LoLMatchIDs.append(match)
                                    if len(LoLMatchIDs) == 0:
                                        logPrint("您输入的对局序号集不合法！请重新输入。\nThe matchID set you've input is illegal! Please try again.")
                                        continue
                                except (SyntaxError, NameError):
                                    logPrint("您的输入存在语法错误。请重新输入！\nSyntax ERROR detected in this input! Please try again!")
                                    continue
                            #开始获取各对局内的玩家信息。数据结构参考/lol-match-history/v1/recently-played-summoners（Begin to capture the players' information in each match. The data structure can be referred to "/lol-match-history/v1/recently-played-summoners"）
                            ##首先定义存储玩家信息的数据框的数据结构（First, define the data structure of the dataframe that stores player information）
                            LoLGame_info_header = {"gameIndex": "游戏序号", "endOfGameResult": "对局终止情况", "gameCreation": "对局创建时间戳", "gameCreationDate": "创建日期", "gameDuration": "持续时长（秒）", "gameId": "对局序号", "gameMode": "游戏模式", "gameType": "游戏类型", "gameVersion": "对局版本", "mapId": "地图序号", "queueId": "队列序号", "gameDuration_norm": "持续时长", "gameModeName": "游戏模式名称", "participantId": "玩家序号", "accountId": "账户序号", "currentAccountId": "当前账户序号", "currentPlatformId": "当前大区", "gameName": "玩家昵称", "matchHistoryUri": "对局记录网址", "platformId": "原大区", "profileIcon": "召唤师图标序号", "puuid": "玩家通用唯一识别码", "summonerId": "召唤师序号", "summonerName": "召唤师名称", "tagLine": "昵称编号", "profileIcon_title": "召唤师图标名称", "profileIcon_imagePath": "召唤师图标路径", "championId": "选用英雄序号", "highestAchievedSeasonTier": "最高段位", "spell1Id": "召唤师技能1序号", "spell2Id": "召唤师技能2序号", "teamId": "阵营代号", "champion_name": "选用英雄", "champion_alias": "选用英雄代号", "champion_squarePortraitPath": "选用英雄方块头像路径", "spell1_name": "召唤师技能1", "spell2_name": "召唤师技能2", "spell1_iconPath": "召唤师技能1图标", "spell2_iconPath": "召唤师技能2图标", "team_color": "阵营", "assists": "助攻", "causedEarlySurrender": "发起提前投降", "champLevel": "英雄等级", "combatPlayerScore": "战斗得分", "damageDealtToObjectives": "对战略点的总伤害", "damageDealtToTurrets": "对防御塔的总伤害", "damageSelfMitigated": "自我缓和的伤害", "deaths": "死亡", "doubleKills": "双杀", "earlySurrenderAccomplice": "同意提前投降", "firstBloodAssist": "协助获得第一滴血", "firstBloodKill": "第一滴血", "firstInhibitorAssist": "协助摧毁第一座召唤水晶", "firstInhibitorKill": "摧毁第一座召唤水晶", "firstTowerAssist": "协助摧毁第一座塔", "firstTowerKill": "摧毁第一座塔", "gameEndedInEarlySurrender": "提前投降导致比赛结束", "gameEndedInSurrender": "投降导致比赛结束", "goldEarned": "金币获取", "goldSpent": "金币使用", "inhibitorKills": "摧毁召唤水晶", "item0": "装备1序号", "item1": "装备2序号", "item2": "装备3序号", "item3": "装备4序号", "item4": "装备5序号", "item5": "装备6序号", "item6": "饰品序号", "killingSprees": "大杀特杀", "kills": "击杀", "largestCriticalStrike": "最大暴击伤害", "largestKillingSpree": "最高连杀", "largestMultiKill": "最高多杀", "longestTimeSpentLiving": "最长生存时间", "magicDamageDealt": "造成的魔法伤害", "magicDamageDealtToChampions": "对英雄的魔法伤害", "magicalDamageTaken": "承受的魔法伤害", "neutralMinionsKilled": "击杀野怪", "neutralMinionsKilledEnemyJungle": "击杀敌方野区野怪", "neutralMinionsKilledTeamJungle": "击杀我方野区野怪", "objectivePlayerScore": "战略点玩家得分", "pentaKills": "五杀", "perk0": "符文1序号", "perk0Var1": "符文1：参数1", "perk0Var2": "符文1：参数2", "perk0Var3": "符文1：参数3", "perk1": "符文2序号", "perk1Var1": "符文2：参数1", "perk1Var2": "符文2：参数2", "perk1Var3": "符文2：参数3", "perk2": "符文3序号", "perk2Var1": "符文3：参数1", "perk2Var2": "符文3：参数2", "perk2Var3": "符文3：参数3", "perk3": "符文4序号", "perk3Var1": "符文4：参数1", "perk3Var2": "符文4：参数2", "perk3Var3": "符文4：参数3", "perk4": "符文5序号", "perk4Var1": "符文5：参数1", "perk4Var2": "符文5：参数2", "perk4Var3": "符文5：参数3", "perk5": "符文6序号", "perk5Var1": "符文6：参数1", "perk5Var2": "符文6：参数2", "perk5Var3": "符文6：参数3", "perkPrimaryStyle": "主系序号", "perkSubStyle": "副系序号", "physicalDamageDealt": "造成的物理伤害", "physicalDamageDealtToChampions": "对英雄的物理伤害", "physicalDamageTaken": "承受的物理伤害", "playerAugment1": "强化符文1", "playerAugment2": "强化符文2", "playerAugment3": "强化符文3", "playerAugment4": "强化符文4", "playerAugment5": "强化符文5", "playerAugment6": "强化符文6", "playerScore0": "玩家得分1", "playerScore1": "玩家得分2", "playerScore2": "玩家得分3", "playerScore3": "玩家得分4", "playerScore4": "玩家得分5", "playerScore5": "玩家得分6", "playerScore6": "玩家得分7", "playerScore7": "玩家得分8", "playerScore8": "玩家得分9", "playerScore9": "玩家得分10", "playerSubteamId": "子阵营代号", "quadraKills": "四杀", "sightWardsBoughtInGame": "购买洞察之石", "subteamPlacement": "队伍排名", "teamEarlySurrendered": "队伍提前投降", "timeCCingOthers": "控制得分", "totalDamageDealt": "造成的伤害总和", "totalDamageDealtToChampions": "对英雄的伤害总和", "totalDamageTaken": "承受伤害", "totalHeal": "输出治疗效果", "totalMinionsKilled": "击杀小兵", "totalPlayerScore": "玩家总得分", "totalScoreRank": "总得分排名", "totalTimeCrowdControlDealt": "控制时间", "totalUnitsHealed": "治疗单位数", "tripleKills": "三杀", "trueDamageDealt": "造成真实伤害", "trueDamageDealtToChampions": "对英雄的真实伤害", "trueDamageTaken": "承受的真实伤害", "turretKills": "摧毁防御塔", "unrealKills": "六杀及以上", "visionScore": "视野得分", "visionWardsBoughtInGame": "购买控制守卫", "wardsKilled": "摧毁守卫", "wardsPlaced": "放置守卫", "win": "胜利", "item0_name": "装备1", "item1_name": "装备2", "item2_name": "装备3", "item3_name": "装备4", "item4_name": "装备5", "item5_name": "装备6", "item6_name": "饰品", "item0_iconPath": "装备1图标路径", "item1_iconPath": "装备2图标路径", "item2_iconPath": "装备3图标路径", "item3_iconPath": "装备4图标路径", "item4_iconPath": "装备5图标路径", "item5_iconPath": "装备6图标路径", "item6_iconPath": "饰品图标路径", "perk0EndOfGameStatDescs": "符文1游戏结算数据", "perk1EndOfGameStatDescs": "符文2游戏结算数据", "perk2EndOfGameStatDescs": "符文3游戏结算数据", "perk3EndOfGameStatDescs": "符文4游戏结算数据", "perk4EndOfGameStatDescs": "符文5游戏结算数据", "perk5EndOfGameStatDescs": "符文6游戏结算数据", "perk0_name": "符文1名称", "perk1_name": "符文2名称", "perk2_name": "符文3名称", "perk3_name": "符文4名称", "perk4_name": "符文5名称", "perk5_name": "符文6名称", "perk0_iconPath": "符文1图标路径", "perk1_iconPath": "符文2图标路径", "perk2_iconPath": "符文3图标路径", "perk3_iconPath": "符文4图标路径", "perk4_iconPath": "符文5图标路径", "perk5_iconPath": "符文6图标路径", "perkPrimaryStyle_name": "主系名称", "perkPrimaryStyle_iconPath": "主系图标路径", "perkSubStyle_name": "副系名称", "perkSubStyle_iconPath": "副系图标路径", "playerAugment1_nameTRA": "强化符文1名称", "playerAugment2_nameTRA": "强化符文2名称", "playerAugment3_nameTRA": "强化符文3名称", "playerAugment4_nameTRA": "强化符文4名称", "playerAugment5_nameTRA": "强化符文5名称", "playerAugment6_nameTRA": "强化符文6名称", "playerAugment1_augmentIconPath": "强化符文1图标路径", "playerAugment2_augmentIconPath": "强化符文2图标路径", "playerAugment3_augmentIconPath": "强化符文3图标路径", "playerAugment4_augmentIconPath": "强化符文4图标路径", "playerAugment5_augmentIconPath": "强化符文5图标路径", "playerAugment6_augmentIconPath": "强化符文6图标路径", "playerAugment1_rarity": "强化符文1等级", "playerAugment2_rarity": "强化符文2等级", "playerAugment3_rarity": "强化符文3等级", "playerAugment4_rarity": "强化符文4等级", "playerAugment5_rarity": "强化符文5等级", "playerAugment6_rarity": "强化符文6等级", "playerSubteam_color": "子阵营", "K/D/A": "击杀/死亡/助攻", "KDA": "战损比", "CS": "补刀", "GPM": "分均经济", "GUE": "金币利用率", "CSPM": "分均补刀", "D/G": "伤害转化率", "win/lose": "胜负", "bannedChampionId": "禁用英雄序号", "bannedChampion_name": "禁用英雄", "bannedChampion_alias": "禁用英雄代号", "bannedChampion_squarePortraitPath": "禁用英雄方块头像路径", "lane": "分路", "role": "角色定位", "ally?": "是否队友？", "assists_percent": "助攻次数占比", "combatPlayerScore_percent": "战斗得分占比", "damageDealtToObjectives_percent": "对战略点的总伤害占比", "damageDealtToTurrets_percent": "对防御塔的总伤害占比", "damageSelfMitigated_percent": "自我缓和的伤害占比", "deaths_percent": "死亡次数占比", "doubleKills_percent": "双杀次数占比", "goldEarned_percent": "金币获取占比", "goldSpent_percent": "金币使用占比", "inhibitorKills_percent": "摧毁召唤水晶数量占比", "killingSprees_percent": "大杀特杀次数占比", "kills_percent": "击杀数量占比", "largestCriticalStrike_percent": "最大暴击伤害占比", "largestKillingSpree_percent": "最高连杀占比", "largestMultiKill_percent": "最高多杀占比", "longestTimeSpentLiving_percent": "最长生存时间占比", "magicDamageDealt_percent": "造成的魔法伤害占比", "magicDamageDealtToChampions_percent": "对英雄的魔法伤害占比", "magicalDamageTaken_percent": "承受的魔法伤害占比", "neutralMinionsKilled_percent": "击杀野怪数量占比", "neutralMinionsKilledEnemyJungle_percent": "击杀敌方野区野怪数量占比", "neutralMinionsKilledTeamJungle_percent": "击杀我方野区野怪数量占比", "objectivePlayerScore_percent": "战略点玩家得分占比", "pentaKills_percent": "五杀次数占比", "physicalDamageDealt_percent": "造成的物理伤害占比", "physicalDamageDealtToChampions_percent": "对英雄的物理伤害占比", "physicalDamageTaken_percent": "承受的物理伤害占比", "playerScore0_percent": "玩家得分1占比", "playerScore1_percent": "玩家得分2占比", "playerScore2_percent": "玩家得分3占比", "playerScore3_percent": "玩家得分4占比", "playerScore4_percent": "玩家得分5占比", "playerScore5_percent": "玩家得分6占比", "playerScore6_percent": "玩家得分7占比", "playerScore7_percent": "玩家得分8占比", "playerScore8_percent": "玩家得分9占比", "playerScore9_percent": "玩家得分10占比", "quadraKills_percent": "四杀次数占比", "sightWardsBoughtInGame_percent": "购买洞察之石数量占比", "timeCCingOthers_percent": "控制得分占比", "totalDamageDealt_percent": "造成的伤害总和占比", "totalDamageDealtToChampions_percent": "对英雄的伤害总和占比", "totalDamageTaken_percent": "承受伤害占比", "totalHeal_percent": "输出治疗效果占比", "totalMinionsKilled_percent": "击杀小兵数量占比", "totalPlayerScore_percent": "玩家总得分占比", "totalTimeCrowdControlDealt_percent": "控制时间占比", "totalUnitsHealed_percent": "治疗单位数占比", "tripleKills_percent": "三杀次数占比", "trueDamageDealt_percent": "造成真实伤害占比", "trueDamageDealtToChampions_percent": "对英雄的真实伤害占比", "trueDamageTaken_percent": "承受的真实伤害占比", "turretKills_percent": "摧毁防御塔数量占比", "unrealKills_percent": "六杀及以上连杀次数占比", "visionScore_percent": "视野得分占比", "visionWardsBoughtInGame_percent": "购买控制守卫数量占比", "wardsKilled_percent": "摧毁守卫数量占比", "wardsPlaced_percent": "放置守卫数量占比", "KP_percent": "参团率", "CS_percent": "补刀数占比", "assists_order": "助攻次数位次", "champLevel_order": "英雄等级位次", "combatPlayerScore_order": "战斗得分位次", "damageDealtToObjectives_order": "对战略点的总伤害位次", "damageDealtToTurrets_order": "对防御塔的总伤害位次", "damageSelfMitigated_order": "自我缓和的伤害位次", "deaths_order": "死亡次数位次", "doubleKills_order": "双杀次数位次", "goldEarned_order": "金币获取位次", "goldSpent_order": "金币使用位次", "inhibitorKills_order": "摧毁召唤水晶数量位次", "killingSprees_order": "大杀特杀次数位次", "kills_order": "击杀数量位次", "largestCriticalStrike_order": "最大暴击伤害位次", "largestKillingSpree_order": "最高连杀位次", "largestMultiKill_order": "最高多杀位次", "longestTimeSpentLiving_order": "最长生存时间位次", "magicDamageDealt_order": "造成的魔法伤害位次", "magicDamageDealtToChampions_order": "对英雄的魔法伤害位次", "magicalDamageTaken_order": "承受的魔法伤害位次", "neutralMinionsKilled_order": "击杀野怪数量位次", "neutralMinionsKilledEnemyJungle_order": "击杀敌方野区野怪数量位次", "neutralMinionsKilledTeamJungle_order": "击杀我方野区野怪数量位次", "objectivePlayerScore_order": "战略点玩家得分位次", "pentaKills_order": "五杀次数位次", "physicalDamageDealt_order": "造成的物理伤害位次", "physicalDamageDealtToChampions_order": "对英雄的物理伤害位次", "physicalDamageTaken_order": "承受的物理伤害位次", "playerScore0_order": "玩家得分1位次", "playerScore1_order": "玩家得分2位次", "playerScore2_order": "玩家得分3位次", "playerScore3_order": "玩家得分4位次", "playerScore4_order": "玩家得分5位次", "playerScore5_order": "玩家得分6位次", "playerScore6_order": "玩家得分7位次", "playerScore7_order": "玩家得分8位次", "playerScore8_order": "玩家得分9位次", "playerScore9_order": "玩家得分10位次", "quadraKills_order": "四杀次数位次", "sightWardsBoughtInGame_order": "购买洞察之石数量位次", "timeCCingOthers_order": "控制得分位次", "totalDamageDealt_order": "造成的伤害总和位次", "totalDamageDealtToChampions_order": "对英雄的伤害总和位次", "totalDamageTaken_order": "承受伤害位次", "totalHeal_order": "输出治疗效果位次", "totalMinionsKilled_order": "击杀小兵数量位次", "totalPlayerScore_order": "玩家总得分位次", "totalTimeCrowdControlDealt_order": "控制时间位次", "totalUnitsHealed_order": "治疗单位数位次", "tripleKills_order": "三杀次数位次", "trueDamageDealt_order": "造成真实伤害位次", "trueDamageDealtToChampions_order": "对英雄的真实伤害位次", "trueDamageTaken_order": "承受的真实伤害位次", "turretKills_order": "摧毁防御塔数量位次", "unrealKills_order": "六杀及以上连杀次数位次", "visionScore_order": "视野得分位次", "visionWardsBoughtInGame_order": "购买控制守卫数量位次", "wardsKilled_order": "摧毁守卫数量位次", "wardsPlaced_order": "放置守卫数量位次", "KDA_order": "战损比位次", "KP_order": "参团率位次", "CS_order": "补刀数位次", "D/G_order": "伤害转化率位次", "GUE_order": "金币利用率位次"}
                            LoLGame_info_header_keys = list(LoLGame_info_header.keys())
                            LoLGame_info_data = {}
                            fetched_info = False #用于控制程序走向，防止在没有获取到任何对局信息的情况下程序进入可视化部分（Used to control the running of the program, in case the program enters visualization part without fetching any match information）
                            error_LoLMatchIDs = [] #记录实际存在但未如期获取的对局序号（Records the LoL matchIDs that really exist but fail to be fetched）
                            matches_to_remove = [] #记录获取成功但不包含主玩家的对局序号（Records the matches that are fetched successfully but don't contain the main player）
                            LoLGameDuration_raw = [] #用于存储未转化成几分几秒格式的游戏持续时间。主要是为了方便可视化时呈现不同玩家的累计游戏时间的图表（Used to store the gameDuration that is not transformed into "(X)X:XX" form. Mainly for convenience of displaying the chart regarding the total time for which a player has accompanied the main player）
                            LoLChampions = copy.deepcopy(LoLChampions_initial) #接下来查询具体的对局信息和时间轴，使用的可能并不是历史记录中记载的对局序号形成的列表。考虑实际使用需求，这里对于装备的合适版本信息采取的思路是默认从最新版本开始获取，如果有装备不存在于最新版本的装备信息，则获取游戏信息中存储的版本对应的装备信息。该思路仍然有问题，详见后续关于美测服的装备获取的注释（The next step is to capture the information and timeline for each specific match, which may not originate from the matchIDs recorded in the match history. Considering the practical use, here the stream of thought for an appropriate version for items is to get items' information from the latest patch, and if some item doesn't exist in the items information of the latest patch, then get the items of the version corresponding to the game according to gameVersion recorded in the match information. There's a flaw of this idea. Please refer to the annotation regarding PBE data crawling for further solution）
                            spells = copy.deepcopy(spells_initial)
                            LoLItems = copy.deepcopy(LoLItems_initial)
                            current_versions["summonerIcon"] = current_versions["LoLChampion"] = current_versions["spell"] = current_versions["LoLItem"] = current_versions["perk"] = current_versions["perkstyle"] = current_versions["CherryAugment"] = URLPatch
                            unmapped_keys["summonerIcon"], unmapped_keys["LoLChampion"], unmapped_keys["spell"], unmapped_keys["LoLItem"], unmapped_keys["perk"], unmapped_keys["perkstyle"], unmapped_keys["CherryAugment"] = set(), set(), set(), set(), set(), set(), set()
                            for key in LoLGame_info_header_keys:
                                LoLGame_info_data[key] = []
                            for matchID in LoLMatchIDs:
                                LoLGame_info = await (await connection.request("GET", f"/lol-match-history/v1/games/{matchID}")).json()
                                #logPrint(LoLGame_info)
                                
                                #尝试修复错误（Try to fix the error）
                                if "errorCode" in LoLGame_info:
                                    count = 0
                                    if LoLGame_info["httpStatus"] == 404:
                                        logPrint(f"未找到序号为{matchID}的回放文件！将忽略该序号。\nMatch file with matchID {matchID} not found! The program will ignore this matchID.")
                                        continue
                                    if "500 Internal Server Error" in LoLGame_info["message"]:
                                        if not error_occurred:
                                            logPrint("您所在大区的对局记录服务异常。尝试重新获取数据……\nThe match history service provided on your server isn't in place. Trying to recapture the history data ...")
                                            error_occurred = True
                                        while "errorCode" in LoLGame_info and "500 Internal Server Error" in LoLGame_info["message"] and count <= 3:
                                            count += 1
                                            logPrint("正在第%d次尝试获取对局%d信息……\nTimes trying to capture Match %d: No. %d ..." %(count, matchID, matchID, count))
                                            LoLGame_info = await (await connection.request("GET", f"/lol-match-history/v1/games/{matchID}")).json()
                                    elif "Connection timed out after " in LoLGame_info["message"]:
                                        logPrint("对局信息保存超时！请检查网速状况！\nGame information saving operation timed out after 20000 milliseconds with 0 bytes received! Please check the netspeed!")
                                    elif "Service Unavailable - Connection retries limit exceeded. Response timed out" in LoLGame_info["message"]:
                                        if not error_occurred:
                                            logPrint("访问频繁。尝试重新获取数据……\nConnection retries limit exceeded! Trying to recapture the match data ...")
                                            error_occurred = True
                                        while "errorCode" in LoLGame_info and "Service Unavailable - Connection retries limit exceeded. Response timed out" in LoLGame_info["message"] and count <= 3:
                                            count += 1
                                            logPrint("正在第%d次尝试获取对局%d信息……\nTimes trying to capture Match %d: No. %d ..." %(count, matchID, matchID, count))
                                            LoLGame_info = await (await connection.request("GET", f"/lol-match-history/v1/games/{matchID}")).json()
                                    if count > 3:
                                        logPrint("对局%d信息获取失败！\nMatch %d information capture failure!" %(matchID, matchID))
                                
                                if "errorCode" in LoLGame_info:
                                    logPrint(LoLGame_info, end = "\n\n")
                                    error_LoLMatchIDs.append(matchID)
                                else:
                                    version = LoLGame_info["gameVersion"]
                                    bigVersion = ".".join(version.split(".")[:2])
                                    #整理对局禁用信息（Sort out the team ban information）
                                    bans_team100 = LoLGame_info["teams"][0]["bans"]
                                    try:
                                        bans_team200 = LoLGame_info["teams"][1]["bans"]
                                    except IndexError:
                                        bans = bans_team100 #空对局也会进入历史记录。空对局定义为完成选英雄但是无法正常进入游戏，而后游戏不存在的对局。而训练模式的空对局只有一方，因此LoLGame_info["teams"]中只有一个元素（Empty matches are included in the match history. An empty match is defined as the matches which can't be launched after the ChmpSlct period. Since an empty match of Practice Tool has only one team, there's only 1 element in LoLGame_info["teams"]）
                                    else:
                                        bans = bans_team100 + bans_team200
                                    if LoLGame_info["gameMode"] == "CHERRY" and patch_compare("14.8", version):
                                        bans_tmp = bans[:]
                                        bans = []
                                        emptyBan = {"championId": -1, "pickTurn": 0} #定义一个初始化禁用字典，用于后续数据框填充空值（Define an initialized banning dictionary so that empty values are appended to the dataframe at certain times subsequently）
                                        playerSubteam = {} #存储不同子阵营的玩家，键是子阵营序号，值是该子阵营中的玩家的API序号列表（Stores different subteams' players. Keys are playerSubteamIds, and values are index lists from API for players in the subteams）
                                        for i in range(len(LoLGame_info["participants"])):
                                            bans.append(emptyBan.copy())
                                            playerSubteamId = LoLGame_info["participants"][i]["stats"]["playerSubteamId"]
                                            if not playerSubteamId in playerSubteam:
                                                playerSubteam[playerSubteamId] = []
                                            playerSubteam[playerSubteamId].append(i)
                                        if patch_compare("14.12", version):
                                            participantBanIds = []
                                            for i in sorted(playerSubteam.keys()):
                                                participantBanIds += [playerSubteam[i][0], playerSubteam[i][1]] #这里默认采用某个子阵营在API中记录的第一名玩家作为该子阵营的先选者。这可能与实际选用顺序有出入（Here the first player of a subteam recorded in API is considered as the player that picks a champion first. This player may not be the real first player.）
                                        else:
                                            participantBanIds = [playerSubteam[i][0] for i in sorted(playerSubteam.keys())] #这里默认采用某个子阵营在API中记录的第一名玩家作为禁用英雄的玩家。这可能与实际禁用英雄的玩家有出入（Here the first player of a subteam recorded in API is considered as the player that banned some champion. This player may not be the real player that banned it）
                                        for i in range(len(participantBanIds)):
                                            bans[participantBanIds[i]] = bans_tmp[i]
                                    legacy_banData_team100_appended = legacy_banData_team200_appended = False #自定义对局中的征召模式是由每个阵营的1号选手禁用3个英雄，所以当禁用信息添加到一个阵营的第一名玩家后，后续玩家不需要再添加禁用信息。这两个逻辑变量就是用来判断这一点的（Draft mode in custom matches is performed by the first player of each team banning 3 champions, so if the ban information is added into the first player, the subsequent player in the same team doesn't need to add this information. That's what these two boolean variables are used for）
                                    legacy_banData_team100_last_i = legacy_banData_team200_last_i = -1 #上面两个逻辑变量需要在切换i时才转变为真。i从0开始遍历，如果这两者在程序进行到判断禁用信息是否已添加的阶段时仍然等于-1，说明还没添加过，将它们赋值为i；一旦i发生变化，则把上面两个逻辑变量转变为真（The above two boolean variables become True only when the loop traverses the next `i`. `i` traverses from 0. When the program is going to judge whether the ban information has been added, if these two variables are still -1, then the ban information hasn't been added, and they're assigned `i`. Once `i` changes, the above two boolean variables are assigned True）
                                    #判断对局序号列表中的对局是否包含主玩家（Judges whether the matches in the matchID list contain the main player）
                                    participant_puuids = []
                                    for i in LoLGame_info["participantIdentities"]:
                                        participant_puuids.append(i["player"]["puuid"])
                                    if any(puuid in participant_puuids for puuid in current_puuid_list) or args.reserve: #之所以使用玩家通用唯一识别码，而不是用召唤师名称来识别对局是否包含主玩家，是因为该玩家可能使用过改名卡。这里也没有选择帐户序号，这是因为保存在对局中的各玩家的帐户序号竟然是0！（The reason why the puuid instead of the displayName or summonerName is used to identify whether the matches contain the main player is that the player may have used name changing card. AccountId isn't chosen here, because all players' accountIds saved in the match fetched from 127 API is 0, to my surprise!）
                                        for currentParticipantId in range(len(LoLGame_info["participantIdentities"])): #定位主召唤师（Find the index of the main player in a match）
                                            if LoLGame_info["participantIdentities"][currentParticipantId]["player"]["puuid"] in current_puuid_list or LoLGame_info["participantIdentities"][currentParticipantId]["player"]["gameName"] + "#" + LoLGame_info["participantIdentities"][currentParticipantId]["player"]["tagLine"] in current_summonerName_list:
                                                break
                                        #下面针对每场对局建立总的数据资源异常处理机制（Builds the summarized data resource exceptional handling mechanism for each match）
                                        ##召唤师图标（Summoner icon）
                                        summonerIconIds_match_list = sorted(set(map(lambda x: x["player"]["profileIcon"], LoLGame_info["participantIdentities"])))
                                        for i in summonerIconIds_match_list:
                                            if not i in summonerIcons and current_versions["summonerIcon"] != bigVersion:
                                                summonerIconPatch_adopted = bigVersion
                                                summonerIcon_recapture = 1
                                                logPrint("第%d/%d场对局（对局序号：%d）召唤师图标信息（%d）获取失败！正在第%d次尝试改用%s版本的召唤师图标信息……\nSummoner icon information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to summoner icons of Patch %s ... Times tried: %d." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, summonerIcon_recapture, summonerIconPatch_adopted, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, summonerIconPatch_adopted, summonerIcon_recapture))
                                                while True:
                                                    try:
                                                        summonerIcon = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-icons.json" %(summonerIconPatch_adopted, language_cdragon[language_code])).json()
                                                    except requests.exceptions.JSONDecodeError:
                                                        summonerIconPatch_deserted = summonerIconPatch_adopted
                                                        summonerIconPatch_adopted = FindPostPatch(summonerIconPatch_adopted, bigPatches)
                                                        summonerIcon_recapture = 1
                                                        logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to LoL champions of Patch %s ... Times tried: %d." %(summonerIconPatch_deserted, summonerIcon_recapture, summonerIconPatch_adopted, summonerIconPatch_deserted, summonerIconPatch_adopted, summonerIcon_recapture))
                                                    except requests.exceptions.RequestException:
                                                        if summonerIcon_recapture < 3:
                                                            summonerIcon_recapture += 1
                                                            logPrint("网络环境异常！正在第%d次尝试改用%s版本的召唤师图标信息……\nYour network environment is abnormal! Changing to summoner icons of Patch %s ... Times tried: %d." %(summonerIcon_recapture, summonerIconPatch_adopted, summonerIconPatch_adopted, summonerIcon_recapture))
                                                        else:
                                                            logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的召唤师图标信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the summoner icon (%s) of Match %d / %d (matchID: %d)!" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID))
                                                            break
                                                    else:
                                                        logPrint("已改用%s版本的召唤师图标信息。\nSummoner icon information changed to Patch %s." %(summonerIconPatch_adopted, summonerIconPatch_adopted))
                                                        summonerIcons = {}
                                                        for summonerIcon_iter in summonerIcon:
                                                            summonerIcon_id = summonerIcon_iter["id"]
                                                            summonerIcons[summonerIcon_id] = summonerIcon_iter
                                                        current_versions["summonerIcon"] = summonerIconPatch_adopted
                                                        unmapped_keys["summonerIcon"].clear()
                                                        break
                                                break
                                        ##英雄：包含选用英雄和禁用英雄（LoL champions, which contain picked and banned ones）
                                        LoLChampionIds_match_list = sorted(set(map(lambda x: x["championId"], LoLGame_info["participants"])) | set(map(lambda x: x["championId"], bans)))
                                        for i in LoLChampionIds_match_list:
                                            if not i in LoLChampions and current_versions["LoLChampion"] != bigVersion:
                                                LoLChampionPatch_adopted = bigVersion
                                                LoLChampion_recapture = 1
                                                logPrint("第%d/%d场对局（对局序号：%d）英雄信息（%d）获取失败！正在第%d次尝试改用%s版本的英雄信息……\nLoL champion information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to LoL champions of Patch %s ... Times tried: %d." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, LoLChampion_recapture, LoLChampionPatch_adopted, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, LoLChampionPatch_adopted, LoLChampion_recapture))
                                                while True:
                                                    try:
                                                        LoLChampion = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/champion-summary.json" %(LoLChampionPatch_adopted, language_cdragon[language_code])).json()
                                                    except requests.exceptions.JSONDecodeError:
                                                        LoLChampionPatch_deserted = LoLChampionPatch_adopted
                                                        LoLChampionPatch_adopted = FindPostPatch(LoLChampionPatch_adopted, bigPatches)
                                                        LoLChampion_recapture = 1
                                                        logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to LoL champions of Patch %s ... Times tried: %d." %(LoLChampionPatch_deserted, LoLChampion_recapture, LoLChampionPatch_adopted, LoLChampionPatch_deserted, LoLChampionPatch_adopted, LoLChampion_recapture))
                                                    except requests.exceptions.RequestException:
                                                        if LoLChampion_recapture < 3:
                                                            LoLChampion_recapture += 1
                                                            logPrint("网络环境异常！正在第%d次尝试改用%s版本的英雄信息……\nYour network environment is abnormal! Changing to LoL champions of Patch %s ... Times tried: %d." %(LoLChampion_recapture, LoLChampionPatch_adopted, LoLChampionPatch_adopted, LoLChampion_recapture))
                                                        else:
                                                            logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的英雄信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the LoL champion (%s) of Match %d / %d (matchID: %d)!" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID))
                                                            break
                                                    else:
                                                        logPrint("已改用%s版本的英雄信息。\nLoL champion information changed to Patch %s." %(LoLChampionPatch_adopted, LoLChampionPatch_adopted))
                                                        LoLChampions = {}
                                                        for LoLChampion_iter in LoLChampion:
                                                            LoLChampion_id = LoLChampion_iter["id"]
                                                            LoLChampions[LoLChampion_id] = LoLChampion_iter
                                                        current_versions["LoLChampion"] = LoLChampionPatch_adopted
                                                        unmapped_keys["LoLChampion"].clear()
                                                        break
                                                break
                                        ##召唤师技能（Summoner spells）
                                        spellIds_match_list = sorted(set(map(lambda x: x["spell1Id"], LoLGame_info["participants"])) | set(map(lambda x: x["spell2Id"], LoLGame_info["participants"])))
                                        for i in spellIds_match_list:
                                            if not i in spells and current_versions["spell"] != bigVersion and i != 0: #需要注意电脑玩家的召唤师技能序号都是0（Note that Spell Ids of bot players are both 0s）
                                                spellPatch_adopted = bigVersion
                                                spell_recapture = 1
                                                logPrint("第%d/%d场对局（对局序号：%d）召唤师技能信息（%d）获取失败！正在第%d次尝试改用%s版本的召唤师技能信息……\nSpell information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to spells of Patch %s ... Times tried: %d." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, spell_recapture, spellPatch_adopted, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, spellPatch_adopted, spell_recapture))
                                                while True:
                                                    try:
                                                        spell = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-spells.json" %(spellPatch_adopted, language_cdragon[language_code])).json()
                                                    except requests.exceptions.JSONDecodeError:
                                                        spellPatch_deserted = spellPatch_adopted
                                                        spellPatch_adopted = FindPostPatch(spellPatch_adopted, bigPatches)
                                                        spell_recapture = 1
                                                        logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to spells of Patch %s ... Times tried: %d." %(spellPatch_deserted, spell_recapture, spellPatch_adopted, spellPatch_deserted, spellPatch_adopted, spell_recapture))
                                                    except requests.exceptions.RequestException:
                                                        if spell_recapture < 3:
                                                            spell_recapture += 1
                                                            logPrint("网络环境异常！正在第%d次尝试改用%s版本的召唤师技能信息……\nYour network environment is abnormal! Changing to spells of Patch %s ... Times tried: %d." %(spell_recapture, spellPatch_adopted, spellPatch_adopted, spell_recapture))
                                                        else:
                                                            logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的召唤师技能信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the spell (%s) of Match %d / %d (matchID: %d)!" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID))
                                                            break
                                                    else:
                                                        logPrint("已改用%s版本的召唤师技能信息。\nSpell information changed to Patch %s." %(spellPatch_adopted, spellPatch_adopted))
                                                        spells = {}
                                                        for spell_iter in spell:
                                                            spell_id = spell_iter["id"]
                                                            spells[spell_id] = spell_iter
                                                        current_versions["spell"] = spellPatch_adopted
                                                        unmapped_keys["spell"].clear()
                                                        break
                                                break
                                        ##英雄联盟装备（LoL items）
                                        LoLItemIds_match_list = sorted(set(item for s in [set(map(lambda x: x["stats"]["item" + str(i)], LoLGame_info["participants"])) for i in range(7)] for item in s)) #该表达式等价于以下表达式（This expression is equivalent to the following expression）：`LoLItemIds_match_list = sorted(set(map(lambda x: x["stats"]["item0"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["item1"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["item2"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["item3"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["item4"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["item5"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["item6"], LoLGame_info["participants"])))`
                                        for i in LoLItemIds_match_list:
                                            if not i in LoLItems and current_versions["LoLItem"] != bigVersion and i != 0: #空装备序号是0（The itemId of an empty item is 0）
                                                LoLItemPatch_adopted = bigVersion
                                                LoLItem_recapture = 1
                                                logPrint("第%d/%d场对局（对局序号：%d）英雄联盟装备信息（%d）获取失败！正在第%d次尝试改用%s版本的英雄联盟装备信息……\nLoL item information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to LoL items of Patch %s ... Times tried: %d." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, LoLItem_recapture, LoLItemPatch_adopted, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, LoLItemPatch_adopted, LoLItem_recapture))
                                                while True:
                                                    try:
                                                        LoLItem = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/items.json" %(LoLItemPatch_adopted, language_cdragon[language_code])).json()
                                                    except requests.exceptions.JSONDecodeError:
                                                        LoLItemPatch_deserted = LoLItemPatch_adopted
                                                        LoLItemPatch_adopted = FindPostPatch(LoLItemPatch_adopted, bigPatches)
                                                        LoLItem_recapture = 1
                                                        logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to LoL items of Patch %s ... Times tried: %d." %(LoLItemPatch_deserted, LoLItem_recapture, LoLItemPatch_adopted, LoLItemPatch_deserted, LoLItemPatch_adopted, LoLItem_recapture))
                                                    except requests.exceptions.RequestException:
                                                        if LoLItem_recapture < 3:
                                                            LoLItem_recapture += 1
                                                            logPrint("网络环境异常！正在第%d次尝试改用%s版本的英雄联盟装备信息……\nYour network environment is abnormal! Changing to LoL items of Patch %s ... Times tried: %d." %(LoLItem_recapture, LoLItemPatch_adopted, LoLItemPatch_adopted, LoLItem_recapture))
                                                        else:
                                                            logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的英雄联盟装备信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the LoL item (%s) of Match %d / %d (matchID: %d)!" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID))
                                                            break
                                                    else:
                                                        logPrint("已改用%s版本的英雄联盟装备信息。\nLoL item information changed to Patch %s." %(LoLItemPatch_adopted, LoLItemPatch_adopted))
                                                        LoLItems = {}
                                                        for LoLItem_iter in LoLItem:
                                                            LoLItem_id = LoLItem_iter["id"]
                                                            LoLItems[LoLItem_id] = LoLItem_iter
                                                        current_versions["LoLItem"] = LoLItemPatch_adopted
                                                        unmapped_keys["LoLItem"].clear()
                                                        break
                                                break
                                        ##符文（Perks）
                                        perkIds_match_list = sorted(set(perk for s in [set(map(lambda x: x["stats"]["perk" + str(i)], LoLGame_info["participants"])) for i in range(6)] for perk in s))
                                        for i in perkIds_match_list:
                                            if not i in perks and current_versions["perk"] != bigVersion and i != 0: #在一些非常规模式（如新手训练）的对局中，玩家可能没有携带任何符文（In matches with unconventional game mode (e.g. TUTORIAL), maybe the player doesn't take any runes）
                                                perkPatch_adopted = bigVersion
                                                perk_recapture = 1
                                                logPrint("第%d/%d场对局（对局序号：%d）基石符文信息（%d）获取失败！正在第%d次尝试改用%s版本的基石符文信息……\nPerk information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to perks of Patch %s ... Times tried: %d." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, perk_recapture, perkPatch_adopted, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, perkPatch_adopted, perk_recapture))
                                                while True:
                                                    try:
                                                        perk = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perks.json" %(perkPatch_adopted, language_cdragon[language_code])).json()
                                                    except requests.exceptions.JSONDecodeError:
                                                        perkPatch_deserted = perkPatch_adopted
                                                        perkPatch_adopted = FindPostPatch(perkPatch_adopted, bigPatches)
                                                        perk_recapture = 1
                                                        logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to perks of Patch %s ... Times tried: %d." %(perkPatch_deserted, perk_recapture, perkPatch_adopted, perkPatch_deserted, perkPatch_adopted, perk_recapture))
                                                    except requests.exceptions.RequestException:
                                                        if perk_recapture < 3:
                                                            perk_recapture += 1
                                                            logPrint("网络环境异常！正在第%d次尝试改用%s版本的基石符文信息……\nYour network environment is abnormal! Changing to perks of Patch %s ... Times tried: %d." %(perk_recapture, perkPatch_adopted, perkPatch_adopted, perk_recapture))
                                                        else:
                                                            logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的基石符文信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the perk (%s) of Match %d / %d (matchID: %d)!" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID))
                                                            break
                                                    else:
                                                        logPrint("已改用%s版本的基石符文信息。\nPerk information changed to Patch %s." %(perkPatch_adopted, perkPatch_adopted))
                                                        perks = {}
                                                        for perk_iter in perk:
                                                            perk_id = perk_iter["id"]
                                                            perks[perk_id] = perk_iter
                                                        current_versions["perk"] = perkPatch_adopted
                                                        unmapped_keys["perk"].clear()
                                                        break
                                                break
                                        ##符文系（Perkstyles）
                                        perkstyleIds_match_list = sorted(list(set(map(lambda x: x["stats"]["perkPrimaryStyle"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["perkSubStyle"], LoLGame_info["participants"]))))
                                        for i in perkstyleIds_match_list:
                                            if not i in perkstyles and current_versions["perkstyle"] != bigVersion and i != 0: #在一些非常规模式（如新手训练）的对局中，玩家可能没有携带任何符文（In matches with unconventional game mode (e.g. TUTORIAL), maybe the player doesn't take any runes）
                                                perkstylePatch_adopted = bigVersion
                                                perkstyle_recapture = 1
                                                logPrint("第%d/%d场对局（对局序号：%d）符文系信息（%d）获取失败！正在第%d次尝试改用%s版本的符文系信息……\nPerkstyle information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to perkstyles of Patch %s ... Times tried: %d." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, perkstyle_recapture, perkstylePatch_adopted, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, perkstylePatch_adopted, perkstyle_recapture))
                                                while True:
                                                    try:
                                                        perkstyle = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perkstyles.json" %(perkstylePatch_adopted, language_cdragon[language_code])).json()
                                                    except requests.exceptions.JSONDecodeError:
                                                        perkstylePatch_deserted = perkstylePatch_adopted
                                                        perkstylePatch_adopted = FindPostPatch(perkstylePatch_adopted, bigPatches)
                                                        perkstyle_recapture = 1
                                                        logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to perks of Patch %s ... Times tried: %d." %(perkstylePatch_deserted, perkstyle_recapture, perkstylePatch_adopted, perkstylePatch_deserted, perkstylePatch_adopted, perkstyle_recapture))
                                                    except requests.exceptions.RequestException:
                                                        if perkstyle_recapture < 3:
                                                            perkstyle_recapture += 1
                                                            logPrint("网络环境异常！正在第%d次尝试改用%s版本的符文系信息……\nYour network environment is abnormal! Changing to perkstyles of Patch %s ... Times tried: %d." %(perkstyle_recapture, perkstylePatch_adopted, perkstylePatch_adopted, perkstyle_recapture))
                                                        else:
                                                            logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的符文系信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the perkstyle (%s) of Match %d / %d (matchID: %d)!" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID))
                                                            break
                                                    else:
                                                        logPrint("已改用%s版本的符文系信息。\nPerkstyle information changed to Patch %s." %(perkstylePatch_adopted, perkstylePatch_adopted))
                                                        perkstyles = {}
                                                        for perkstyle_iter in perkstyle["styles"]:
                                                            perkstyle_id = perkstyle_iter["id"]
                                                            perkstyles[perkstyle_id] = perkstyle_iter
                                                        current_versions["perkstyle"] = perkstylePatch_adopted
                                                        unmapped_keys["perkstyle"].clear()
                                                        break
                                                break
                                        ##斗魂竞技场强化符文（Cherry augments）
                                        CherryAugmentIds_match_list = sorted(set(augment for s in [set(map(lambda x: x["stats"]["playerAugment" + str(i)], LoLGame_info["participants"])) for i in range(1, 7)] for augment in s)) #该表达式等价于以下表达式（This expression is equivalent to the following expression）：CherryAugmentIds_match_list = sorted(list(set(map(lambda x: x["stats"]["playerAugment1"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["playerAugment2"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["playerAugment3"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["playerAugment4"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["playerAugment5"], LoLGame_info["participants"])) | set(map(lambda x: x["stats"]["playerAugment6"], LoLGame_info["participants"]))))
                                        for i in CherryAugmentIds_match_list:
                                            if not i in CherryAugments and current_versions["CherryAugment"] != bigVersion and i != 0:
                                                CherryAugmentPatch_adopted = bigVersion
                                                CherryAugment_recapture = 1
                                                logPrint("第%d/%d场对局（对局序号：%d）强化符文信息（%d）获取失败！正在第%d次尝试改用%s版本的斗魂竞技场强化符文信息……\nAugment information (%d) of Match %d / %d (matchID: %d) capture failed! Changing to Cherry augments of Patch %s ... Times tried: %d." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, CherryAugment_recapture, CherryAugmentPatch_adopted, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, CherryAugmentPatch_adopted, CherryAugment_recapture))
                                                while True:
                                                    try:
                                                        CherryAugment = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/cherry-augments.json" %(CherryAugmentPatch_adopted, language_cdragon[language_code])).json()
                                                    except requests.exceptions.JSONDecodeError:
                                                        CherryAugmentPatch_deserted = CherryAugmentPatch_adopted
                                                        CherryAugmentPatch_adopted = FindPostPatch(CherryAugmentPatch_adopted, bigPatches)
                                                        CherryAugment_recapture = 1
                                                        logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to Cherry augments of Patch %s ... Times tried: %d." %(CherryAugmentPatch_deserted, CherryAugment_recapture, CherryAugmentPatch_adopted, CherryAugmentPatch_deserted, CherryAugmentPatch_adopted, CherryAugment_recapture))
                                                    except requests.exceptions.RequestException:
                                                        if CherryAugment_recapture < 3:
                                                            CherryAugment_recapture += 1
                                                            logPrint("网络环境异常！正在第%d次尝试改用%s版本的斗魂竞技场强化符文信息……\nYour network environment is abnormal! Changing to Cherry augments of Patch %s ... Times tried: %d." %(CherryAugment_recapture, CherryAugmentPatch_adopted, CherryAugmentPatch_adopted, CherryAugment_recapture))
                                                        else:
                                                            logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的强化符文信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the Cherry augment (%s) of Match %d / %d (matchID: %d)!" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, i, i, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID))
                                                            break
                                                    else:
                                                        logPrint("已改用%s版本的斗魂竞技场强化符文信息。\nCherry augment information changed to Patch %s." %(CherryAugmentPatch_adopted, CherryAugmentPatch_adopted))
                                                        CherryAugments = {}
                                                        for CherryAugment_iter in CherryAugment:
                                                            CherryAugment_id = CherryAugment_iter["id"]
                                                            CherryAugments[CherryAugment_id] = CherryAugment_iter
                                                        current_versions["CherryAugment"] = CherryAugmentPatch_adopted
                                                        unmapped_keys["CherryAugment"].clear()
                                                        break
                                                break
                                        #下面开始整理数据（Sorts out the data）
                                        for i in range(len(LoLGame_info["participants"])):
                                            if LoLGame_info["participantIdentities"][i]["player"]["puuid"] != "00000000-0000-0000-0000-000000000000" and not LoLGame_info["participantIdentities"][i]["player"]["puuid"] in current_puuid_list or args.save_self: #统计玩家，当然指的是不包括自己的人类玩家（Of course, the players counted are human players but not himself / herself）
                                                stats = LoLGame_info["participants"][i]["stats"]
                                                timeline = LoLGame_info["participants"][i]["timeline"]
                                                team_participants = [participant for participant in LoLGame_info["participants"] if LoLGame_info["gameMode"] == "CHERRY" and participant["stats"]["playerSubteamId"] == stats["playerSubteamId"] or LoLGame_info["gameMode"] != "CHERRY" and participant["teamId"] == LoLGame_info["participants"][i]["teamId"]] #存储对局信息中同一队伍的玩家。斗魂竞技场对局应该使用子阵营（Store the participants of the same team from the game information. Subteam should be used to evaluate a player）
                                                LoLGameDuration_raw.append(LoLGame_info["gameDuration"])
                                                for j in range(len(LoLGame_info_header)):
                                                    key = LoLGame_info_header_keys[j]
                                                    if j == 0: #游戏序号（`gameIndex`）
                                                        LoLGame_info_data[key].append(LoLMatchIDs.index(matchID) + 1)
                                                    elif j <= 12:
                                                        if j == 1: #对局终止情况（`endOfGameResults`）
                                                            LoLGame_info_data[key].append(endOfGameResults[LoLGame_info["endOfGameResult"]])
                                                        if j == 3: #创建日期（`gameCreationDate`）
                                                            LoLGame_info_data[key].append(LoLGame_info["gameCreationDate"][:10] + " " + LoLGame_info["gameCreationDate"][11:23])
                                                        elif j == 7: #游戏类型（`gameType`）
                                                            LoLGame_info_data[key].append(gameTypes[LoLGame_info[key]])
                                                        elif j == 11: #持续时长（`gameDuration_norm`）
                                                            LoLGame_info_data[key].append(str(LoLGame_info["gameDuration"] // 60) + ":" + "%02d" %(LoLGame_info["gameDuration"] % 60))
                                                        elif j == 12: #游戏模式名称（`gameModeName`）
                                                            LoLGame_info_data[key].append("自定义" if LoLGame_info["queueId"] == 0 else gamemodes[LoLGame_info["queueId"]]["name"] if LoLGame_info["queueId"] in gamemodes else "")
                                                        else:
                                                            LoLGame_info_data[key].append(LoLGame_info[key])
                                                    elif j == 13: #玩家序号（`participantId`）
                                                        LoLGame_info_data[key].append(LoLGame_info["participantIdentities"][i][key])
                                                    elif j <= 26:
                                                        if j >= 25: #召唤师图标相关键（Summoner icon-related keys）
                                                            profileIconId = LoLGame_info["participantIdentities"][i]["player"]["profileIcon"]
                                                            if profileIconId in summonerIcons:
                                                                try:
                                                                    LoLGame_info_data[key].append(summonerIcons[profileIconId][key.split("_")[-1]])
                                                                except KeyError:
                                                                    traceback_info = traceback.format_exc()
                                                                    logPrint(traceback_info)
                                                                    LoLGame_info_data[key].append("")
                                                            elif profileIconId in summonerIcons_initial:
                                                                try:
                                                                    LoLGame_info_data[key].append(summonerIcons_initial[profileIconId][key.split("_")[-1]])
                                                                except KeyError:
                                                                    traceback_info = traceback.format_exc()
                                                                    logPrint(traceback_info)
                                                                    LoLGame_info_data[key].append("")
                                                            else:
                                                                if not profileIconId in unmapped_keys["summonerIcon"]:
                                                                    unmapped_keys["summonerIcon"].add(profileIconId)
                                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）召唤师图标信息（%d）获取失败！将采用原始数据！\n[%d. %s] Summoner icon information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, profileIconId, j, key, profileIconId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                LoLGame_info_data[key].append(profileIconId if j == 25 else "")
                                                        else:
                                                            LoLGame_info_data[key].append(LoLGame_info["participantIdentities"][i]["player"][key])
                                                    elif j <= 39:
                                                        if j == 28: #最高段位（`highestAchievedSeasonTier`）
                                                            LoLGame_info_data[key].append(tiers[LoLGame_info["participants"][i]["highestAchievedSeasonTier"]])
                                                        elif j >= 32 and j <= 34: #选用英雄序号相关键（`championId`-related keys）
                                                            championId = LoLGame_info["participants"][i][key.split("_")[0] + "Id"]
                                                            if championId in LoLChampions:
                                                                LoLGame_info_data[key].append(LoLChampions[championId][key.split("_")[-1]])
                                                            elif championId in LoLChampions_initial:
                                                                LoLGame_info_data[key].append(LoLChampions_initial[championId][key.split("_")[-1]])
                                                            else:
                                                                if not championId in unmapped_keys["LoLChampion"]:
                                                                    unmapped_keys["LoLChampion"].add(championId)
                                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）英雄信息（%d）获取失败！将采用原始数据！\n[%d. %s] Champion information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, championId, j, key, championId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                LoLGame_info_data[key].append(championId if j == 32 else "")
                                                        elif j >= 35 and j <= 38: #召唤师技能相关键（Summoner spell-related keys）
                                                            spellId = LoLGame_info["participants"][i][key.split("_")[0] + "Id"]
                                                            if spellId in spells:
                                                                LoLGame_info_data[key].append(spells[spellId][key.split("_")[-1]])
                                                            elif spellId in spells_initial:
                                                                LoLGame_info_data[key].append(spells_initial[spellId][key.split("_")[-1]])
                                                            else:
                                                                if not spellId in unmapped_keys["spell"]:
                                                                    unmapped_keys["spell"].add(spellId)
                                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）召唤师技能信息（%d）获取失败！将采用原始数据！\n[%d. %s] Spell information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, spellId, j, key, spellId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                LoLGame_info_data[key].append(spellId if j <= 36 else "")
                                                        elif j == 39: #阵营（`team_color`）
                                                            LoLGame_info_data[key].append(team_color[LoLGame_info["participants"][i]["teamId"]])
                                                        else:
                                                            LoLGame_info_data[key].append(LoLGame_info["participants"][i][key])
                                                    elif j <= 215:
                                                        if j >= 153 and j <= 166: #英雄联盟装备相关键（LoLItems-related keys）
                                                            itemId = stats[key.split("_")[0]]
                                                            if itemId == 0:
                                                                LoLGame_info_data[key].append("")
                                                            elif itemId in LoLItems:
                                                                LoLGame_info_data[key].append(LoLItems[itemId][key.split("_")[-1]])
                                                            elif itemId in LoLItems_initial:
                                                                LoLGame_info_data[key].append(LoLItems_initial[itemId][key.split("_")[-1]])
                                                            else:
                                                                if not itemId in unmapped_keys["LoLItem"]:
                                                                    unmapped_keys["LoLItem"].add(itemId)
                                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）装备信息（%d）获取失败！将采用原始数据！\n[%d. %s] LoL item information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, itemId, j, key, itemId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                LoLGame_info_data[key].append(itemId if j <= 159 else "")
                                                        elif j >= 167 and j <= 184: #符文相关键（Perks-related keys）
                                                            if j <= 172:
                                                                perkId = stats[key[:5]]
                                                                if perkId == 0:
                                                                    LoLGame_info_data[key].append("")
                                                                elif perkId in perks:
                                                                    perk_EndOfGameStatDescs = "".join(list(map(lambda x: x + "。", perks[perkId]["endOfGameStatDescs"])))
                                                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar1@", str(stats[key[:5] + "Var1"]))
                                                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar2@", str(stats[key[:5] + "Var2"]))
                                                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar3@", str(stats[key[:5] + "Var3"]))
                                                                    LoLGame_info_data[key].append(perk_EndOfGameStatDescs)
                                                                elif perkId in perks_initial:
                                                                    perk_EndOfGameStatDescs = "".join(list(map(lambda x: x + "。", perks_initial[perkId]["endOfGameStatDescs"])))
                                                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar1@", str(stats[key[:5] + "Var1"]))
                                                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar2@", str(stats[key[:5] + "Var2"]))
                                                                    perk_EndOfGameStatDescs = perk_EndOfGameStatDescs.replace("@eogvar3@", str(stats[key[:5] + "Var3"]))
                                                                    LoLGame_info_data[key].append(perk_EndOfGameStatDescs)
                                                                else:
                                                                    if not perkId in unmapped_keys["perk"]:
                                                                        unmapped_keys["perk"].add(perkId)
                                                                        logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）符文信息（%d）获取失败！将采用原始数据！\n[%d. %s] Runes information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, perkId, j, key, perkId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                    LoLGame_info_data[key].append("")
                                                            else:
                                                                perkId = stats[key.split("_")[0]]
                                                                if perkId == 0:
                                                                    LoLGame_info_data[key].append("")
                                                                elif perkId in perks:
                                                                    LoLGame_info_data[key].append(perks[perkId][key.split("_")[-1]])
                                                                elif perkId in perks_initial:
                                                                    LoLGame_info_data[key].append(perks_initial[perkId][key.split("_")[-1]])
                                                                else:
                                                                    if not perkId in unmapped_keys["perk"]:
                                                                        unmapped_keys["perk"].add(perkId)
                                                                        logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）符文信息（%d）获取失败！将采用原始数据！\n[%d. %s] Runes information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, perkId, j, key, perkId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                    LoLGame_info_data[key].append(perkId if j <= 178 else "")
                                                        elif j >= 185 and j <= 188: #符文系相关键（Perkstyles-related keys）
                                                            perkstyleId = stats[key.split("_")[0]]
                                                            if perkstyleId == 0:
                                                                LoLGame_info_data[key].append("")
                                                            elif perkstyleId in perkstyles:
                                                                LoLGame_info_data[key].append(perkstyles[perkstyleId][key.split("_")[-1]])
                                                            elif perkstyleId in perkstyles_initial:
                                                                LoLGame_info_data[key].append(perkstyles_initial[perkstyleId][key.split("_")[-1]])
                                                            else:
                                                                if not perkstyleId in unmapped_keys["perkstyle"]:
                                                                    unmapped_keys["perkstyle"].add(perkstyleId)
                                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）符文系信息（%d）获取失败！将采用原始数据！\n[%d. %s] Perkstyle information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, perkstyleId, j, key, perkstyleId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                LoLGame_info_data[key].append(perkstyleId if (j - 185) % 2 == 0 else "")
                                                        elif j >= 189 and j <= 206: #强化符文相关键（Augment-related keys）
                                                            CherryAugmentId = stats[key.split("_")[0]]
                                                            if CherryAugmentId == 0:
                                                                LoLGame_info_data[key].append("")
                                                            elif CherryAugmentId in CherryAugments:
                                                                if j <= 194: #强化符文名称（`nameTRA`）
                                                                    LoLGame_info_data[key].append(CherryAugments[CherryAugmentId][key.split("_")[-1]])
                                                                elif j <= 200: #强化符文图标路径（`augmentIconPath`）
                                                                    LoLGame_info_data[key].append(CherryAugments[CherryAugmentId]["augmentSmallIconPath"].replace("_small.png", "_large.png"))
                                                                else: #强化符文等级（`rarity`）
                                                                    LoLGame_info_data[key].append(augment_rarity[CherryAugments[CherryAugmentId][key.split("_")[-1]]])
                                                            elif CherryAugmentId in CherryAugments_initial:
                                                                if j <= 194: #强化符文名称（`nameTRA`）
                                                                    LoLGame_info_data[key].append(CherryAugments_initial[CherryAugmentId][key.split("_")[-1]])
                                                                elif j <= 200: #强化符文图标路径（`augmentIconPath`）
                                                                    LoLGame_info_data[key].append(CherryAugments_initial[CherryAugmentId]["augmentSmallIconPath"].replace("_small.png", "_large.png"))
                                                                else: #强化符文等级（`rarity`）
                                                                    LoLGame_info_data[key].append(augment_rarity[CherryAugments_initial[CherryAugmentId][key.split("_")[-1]]])
                                                            else:
                                                                if not CherryAugmentId in unmapped_keys["CherryAugment"]:
                                                                    unmapped_keys["CherryAugment"].add(CherryAugmentId)
                                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）强化符文信息（%d）获取失败！将采用原始数据！\n[%d. %s] Cherry augment information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, CherryAugmentId, j, key, CherryAugmentId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                LoLGame_info_data[key].append(CherryAugmentId if j <= 194 else "")
                                                        elif j == 207: #子阵营（`playerSubteam_color`）
                                                            LoLGame_info_data[key].append(subteam_color[stats["playerSubteamId"]])
                                                        elif j == 208: #击杀/死亡/助攻（`K/D/A`）
                                                            LoLGame_info_data[key].append("/".join([str(stats["kills"]), str(stats["deaths"]), str(stats["assists"])]))
                                                        elif j == 209: #战损比（`KDA`）
                                                            LoLGame_info_data[key].append((stats["kills"] + stats["assists"]) / max(1, stats["deaths"]))
                                                        elif j == 210: #补刀（`CS`）
                                                            LoLGame_info_data[key].append(stats["neutralMinionsKilled"] + stats["totalMinionsKilled"])
                                                        elif j == 211: #分均经济（`GPM`）
                                                            LoLGame_info_data[key].append(0 if LoLGame_info["gameDuration"] == 0 else stats["goldEarned"] * 60 / LoLGame_info["gameDuration"])
                                                        elif j == 212: #金币利用率（`GUE` - Gold Utilization Efficiency）
                                                            LoLGame_info_data[key].append(0 if stats["goldEarned"] == 0 else stats["goldSpent"] / stats["goldEarned"])
                                                        elif j == 213: #分均补刀（`CSPM`）
                                                            LoLGame_info_data[key].append(0 if LoLGame_info["gameDuration"] == 0 else (stats["neutralMinionsKilled"] + stats["totalMinionsKilled"]) * 60 / LoLGame_info["gameDuration"])
                                                        elif j == 214: #伤害转化率（`D/G`）
                                                            LoLGame_info_data[key].append(0 if stats["goldEarned"] == 0 else stats["totalDamageDealtToChampions"] / stats["goldEarned"])
                                                        elif j == 215: #胜负（`win/lose`）
                                                            LoLGame_info_data[key].append("胜利" if stats["win"] else "失败")
                                                        else:
                                                            LoLGame_info_data[key].append(stats[key])
                                                    elif j <= 219:
                                                        if bans == []: #修改说明：以前判断禁用数据是否为空是通过禁用模式进行的，如果禁用模式是经典策略就记录禁用信息，否则直接追加空值到列表中。但是在终极魔典中，先前版本记录禁用信息，后来却不记录了。因此，这里判断禁用数据是否为空，直接通过判断bans是否为空【Modification note: To judge whether the ban information of a match is empty, banMode (teams\bans) is used: if banMode is StandardBanStrategy, record the ban information; otherwise, append empty values to the list (by player_count times). But in Ultbook, ban information is recorded in previous versions but not anymore recorded later. Therefore, to judge whether the ban information is empty, whether the variable bans is empty is directly checked】
                                                            LoLGame_info_data[key].append("")
                                                        else:
                                                            if LoLGame_info["queueId"] == 0:
                                                                if LoLGame_info["participants"][i]["teamId"] == 100:
                                                                    if legacy_banData_team100_last_i == -1:
                                                                        legacy_banData_team100_last_i = i
                                                                    elif legacy_banData_team100_last_i != i:
                                                                        legacy_banData_team100_appended = True
                                                                    if not legacy_banData_team100_appended:
                                                                        if j == 216:
                                                                            LoLGame_info_data[key].append(list(map(lambda x: x["championId"], bans_team100)))
                                                                        else:
                                                                            championIds = list(map(lambda x: x["championId"], bans_team100))
                                                                            to_append = []
                                                                            for championId in championIds:
                                                                                if championId in LoLChampions:
                                                                                    to_append.append(LoLChampions[championId][key.split("_")[-1]])
                                                                                elif championId in LoLChampions_initial:
                                                                                    to_append.append(LoLChampions_initial[championId][key.split("_")[-1]])
                                                                                else:
                                                                                    if not championId in unmapped_keys["LoLChampion"]:
                                                                                        unmapped_keys["LoLChampion"].add(championId)
                                                                                        logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）英雄信息（%d）获取失败！将采用原始数据！\n[%d. %s] Champion information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, championId, j, key, championId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                                    to_append.append(championId if j == 217 else "")
                                                                            LoLGame_info_data[key].append(to_append)
                                                                    else:
                                                                        LoLGame_info_data[key].append("")
                                                                if LoLGame_info["participants"][i]["teamId"] == 200:
                                                                    if legacy_banData_team200_last_i == -1:
                                                                        legacy_banData_team200_last_i = i
                                                                    elif legacy_banData_team200_last_i != i:
                                                                        legacy_banData_team200_appended = True
                                                                    if not legacy_banData_team200_appended:
                                                                        if j == 216:
                                                                            LoLGame_info_data[key].append(list(map(lambda x: x["championId"], bans_team200)))
                                                                        else:
                                                                            championIds = list(map(lambda x: x["championId"], bans_team200))
                                                                            to_append = []
                                                                            for championId in championIds:
                                                                                if championId in LoLChampions:
                                                                                    to_append.append(LoLChampions[championId][key.split("_")[-1]])
                                                                                elif championId in LoLChampions_initial:
                                                                                    to_append.append(LoLChampions_initial[championId][key.split("_")[-1]])
                                                                                else:
                                                                                    if not championId in unmapped_keys["LoLChampion"]:
                                                                                        unmapped_keys["LoLChampion"].add(championId)
                                                                                        logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）英雄信息（%d）获取失败！将采用原始数据！\n[%d. %s] Champion information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, championId, j, key, championId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                                    to_append.append(championId if j == 217 else "")
                                                                            LoLGame_info_data[key].append(to_append)
                                                                    else:
                                                                        LoLGame_info_data[key].append("")
                                                            else:
                                                                if bans[i]["championId"] == -1:
                                                                    LoLGame_info_data[key].append("")
                                                                else:
                                                                    if j == 216:
                                                                        LoLGame_info_data[key].append(bans[i]["championId"])
                                                                    else:
                                                                        championId = bans[i]["championId"]
                                                                        if championId in LoLChampions:
                                                                            LoLGame_info_data[key].append(LoLChampions[championId][key.split("_")[-1]])
                                                                        elif championId in LoLChampions_initial:
                                                                            LoLGame_info_data[key].append(LoLChampions_initial[championId][key.split("_")[-1]])
                                                                        else:
                                                                            if not championId in unmapped_keys["LoLChampion"]:
                                                                                unmapped_keys["LoLChampion"].add(championId)
                                                                                logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）英雄信息（%d）获取失败！将采用原始数据！\n[%d. %s] Champion information (%d) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version, championId, j, key, championId, LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, version))
                                                                            LoLGame_info_data[key].append(championId if j == 217 else "")
                                                    elif j <= 221: #时间轴相关键（Timeline-related keys）
                                                        LoLGame_info_data[key].append(lanes[timeline[key]] if j == 220 else roles[timeline[key]])
                                                    elif j == 222: #是否队友？（`ally?`）
                                                        if any(puuid in participant_puuids for puuid in current_puuid_list) and LoLGame_info["participants"][i]["teamId"] == LoLGame_info["participants"][currentParticipantId]["teamId"] and stats["playerSubteamId"] == LoLGame_info["participants"][currentParticipantId]["stats"]["playerSubteamId"]: #如果小号出现在大号对面的阵营，以大号为主要参考（If a smurf account is against the main account, the main account is referred in priority）
                                                            LoLGame_info_data[key].append(True)
                                                        else:
                                                            LoLGame_info_data[key].append(False)
                                                    else: #对局信息转换键（Keys transformed according to game information）
                                                        subkey = key.split("_")[0]
                                                        if key.endswith("_percent"): #团队占比键（Team percentage keys）
                                                            if j == 281: #参团率（`KP_percent`）
                                                                self_stat = stats["kills"] + stats["assists"]
                                                                total_stat = sum(map(lambda x: x["stats"]["kills"], team_participants))
                                                            elif j == 282: #补刀数占比（`CS_percent`）
                                                                self_stat = stats["totalMinionsKilled"] + stats["neutralMinionsKilled"]
                                                                total_stat = sum(map(lambda x: x["stats"]["totalMinionsKilled"] + x["stats"]["neutralMinionsKilled"], team_participants))
                                                            else:
                                                                self_stat = stats[subkey]
                                                                total_stat = sum(map(lambda x: x["stats"][subkey], team_participants))
                                                            value = 0 if total_stat == 0 else self_stat / total_stat
                                                            LoLGame_info_data[key].append(value)
                                                        else: #位次键（Order keys）
                                                            if j == 342: #战损比位次（`KDA_order`）
                                                                self_stat = (stats["kills"] + stats["assists"]) / max(1, stats["deaths"])
                                                                stat_list = sorted(map(lambda x: (x["stats"]["kills"] + x["stats"]["assists"]) / max(1, x["stats"]["deaths"]), team_participants), reverse = True)
                                                            elif j == 343: #参团率位次（`KP_order`）
                                                                self_stat = stats["kills"] + stats["assists"]
                                                                stat_list = sorted(map(lambda x: x["stats"]["kills"] + x["stats"]["assists"], team_participants), reverse = True)
                                                            elif j == 344: #补刀数位次（`CS_order`）
                                                                self_stat = stats["totalMinionsKilled"] + stats["neutralMinionsKilled"]
                                                                stat_list = sorted(map(lambda x: x["stats"]["totalMinionsKilled"] + x["stats"]["neutralMinionsKilled"], team_participants), reverse = True)
                                                            elif j == 345: #伤害转化率位次（`D/G_order`）
                                                                self_stat = 0 if stats["goldEarned"] == 0 else stats["totalDamageDealtToChampions"] / stats["goldEarned"]
                                                                stat_list = sorted(map(lambda x: 0 if x["stats"]["goldEarned"] == 0 else x["stats"]["totalDamageDealtToChampions"] / x["stats"]["goldEarned"], team_participants), reverse = True)
                                                            elif j == 346: #金币利用率位次（`GUE_order`）
                                                                self_stat = 0 if stats["goldEarned"] == 0 else stats["goldSpent"] / stats["goldEarned"]
                                                                stat_list = sorted(map(lambda x: 0 if x["stats"]["goldEarned"] == 0 else x["stats"]["goldSpent"] / x["stats"]["goldEarned"], team_participants), reverse = True)
                                                            else:
                                                                self_stat = stats[subkey]
                                                                stat_list = sorted(map(lambda x: x["stats"][subkey], team_participants), reverse = j != 289) #死亡次数越低，死亡位次越小（For deaths, the lower the number of deaths is, the smaller the death order is）
                                                            LoLGame_info_data[key].append(0 if len(set(stat_list)) == 1 else stat_list.index(self_stat) + 1) #当所有人的数据一样时，则不用比较位次（When some stat of every player is the same, there's no need to compare it）
                                        fetched_info = True
                                        if args.reserve:
                                            logPrint("[%d/%d]对局%d不包含主玩家。已保留该对局。\nMatch %d doesn't contain the main player but is reserved." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, matchID), print_time = True)
                                        else:
                                            logPrint("加载进度（Loading process）：%d/%d\t对局序号（MatchID）： %s" %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID), print_time = True)
                                    else:
                                        matches_to_remove.append(matchID)
                                        logPrint("[%d/%d]对局%d不包含主玩家。已舍弃该对局。\nMatch %d doesn't contain the main player and is deprecated." %(LoLMatchIDs.index(matchID) + 1, len(LoLMatchIDs), matchID, matchID), print_time = True)

                            if len(error_LoLMatchIDs) > 0:
                                logPrint("警告：以下%d场对局获取失败。\nWarning: The following %d match(es) fail to be fetched." %(len(error_LoLMatchIDs), len(error_LoLMatchIDs)))
                                logPrint(error_LoLMatchIDs)
                            if len(matches_to_remove) > 0:
                                logPrint("注意：以下%d场对局因不包含主玩家而被舍弃。\nAttention: The following %d match(es) are deprecated because they don't contain the main player." %(len(matches_to_remove), len(matches_to_remove)))
                                logPrint(matches_to_remove)
                            if not fetched_info:
                                logPrint("未获取到有效对局。请重新输入要查询的对局序号。\nThe program didn't fetch any valid match. Please reinput the match ID to check.")
                                continue
                            recent_LoLPlayers_statistics_output_order = [0, 13, 23, 17, 24, 22, 21, 28, 5, 3, 11, 10, 6, 12, 9, 8, 222, 32, 33, 217, 218, 220, 221, 42, 35, 36, 153, 154, 155, 156, 157, 158, 159, 189, 201, 190, 202, 191, 203, 192, 204, 193, 205, 194, 206, 208, 209, 210, 213, 214, 43, 138, 139, 71, 68, 72, 51, 50, 55, 54, 53, 52, 48, 142, 128, 81, 147, 132, 140, 134, 109, 75, 144, 133, 108, 74, 143, 70, 45, 44, 136, 141, 135, 110, 76, 145, 46, 148, 151, 150, 129, 149, 58, 211, 59, 212, 137, 77, 79, 78, 146, 60, 73, 185, 187, 173, 167, 174, 168, 175, 169, 176, 170, 177, 171, 178, 172, 41, 49, 131, 56, 57, 215, 130, 234, 228, 223, 281, 224, 268, 236, 233, 237, 229, 271, 260, 246, 276, 262, 269, 264, 248, 240, 273, 263, 247, 239, 272, 235, 226, 225, 266, 270, 265, 249, 241, 274, 227, 277, 280, 279, 261, 278, 230, 231, 267, 242, 244, 243, 282, 275, 232, 238, 284, 295, 289, 283, 342, 343, 345, 285, 329, 297, 294, 298, 290, 332, 321, 307, 337, 323, 330, 325, 309, 301, 334, 324, 308, 300, 333, 296, 287, 286, 327, 331, 326, 310, 302, 335, 288, 338, 341, 340, 322, 339, 291, 292, 346, 328, 303, 304, 305, 344, 336, 293, 299]
                            recent_LoLPlayers_data_organized = {}
                            for i in range(len(recent_LoLPlayers_statistics_output_order)):
                                key = LoLGame_info_header_keys[recent_LoLPlayers_statistics_output_order[i]]
                                recent_LoLPlayers_data_organized[key] = LoLGame_info_data[key]
                                #logPrint("近期一起玩过的英雄联盟玩家数据重排进度（Rearranging process of recently played summoner (LoL) data）：%d/%d" %(i + 1, len(recent_LoLPlayers_statistics_output_order)), end = "\r")
                            #logPrint("正在创建数据框……\nCreating the dataframe ...")
                            recent_LoLPlayers_df = pandas.DataFrame(data = recent_LoLPlayers_data_organized)
                            #logPrint("数据框创建完成！\nDataframe creation finished!")
                            logPrint("正在优化逻辑值显示……\nOptimizing the display of boolean values ...")
                            for column in recent_LoLPlayers_df:
                                if recent_LoLPlayers_df[column].dtype == "bool":
                                    recent_LoLPlayers_df[column] = recent_LoLPlayers_df[column].astype(str)
                                    for i in range(len(recent_LoLPlayers_df)):
                                        recent_LoLPlayers_df.loc[i, column] = "√" if recent_LoLPlayers_df[column][i] == "True" else ""
                            logPrint("逻辑值显示优化完成！\nBoolean value display optimization finished!")
                            recent_LoLPlayers_df = pandas.concat([pandas.DataFrame([LoLGame_info_header])[recent_LoLPlayers_df.columns], recent_LoLPlayers_df], ignore_index = True)
                            break
                if not (search_LoL and fetched_info):
                    LoLGame_info_header = {"gameIndex": "游戏序号", "endOfGameResult": "对局终止情况", "gameCreation": "对局创建时间戳", "gameCreationDate": "创建日期", "gameDuration": "持续时长（秒）", "gameId": "对局序号", "gameMode": "游戏模式", "gameType": "游戏类型", "gameVersion": "对局版本", "mapId": "地图序号", "queueId": "队列序号", "gameDuration_norm": "持续时长", "gameModeName": "游戏模式名称", "participantId": "玩家序号", "accountId": "账户序号", "currentAccountId": "当前账户序号", "currentPlatformId": "当前大区", "gameName": "玩家昵称", "matchHistoryUri": "对局记录网址", "platformId": "原大区", "profileIcon": "召唤师图标序号", "puuid": "玩家通用唯一识别码", "summonerId": "召唤师序号", "summonerName": "召唤师名称", "tagLine": "昵称编号", "profileIcon_title": "召唤师图标名称", "profileIcon_imagePath": "召唤师图标路径", "championId": "选用英雄序号", "highestAchievedSeasonTier": "最高段位", "spell1Id": "召唤师技能1序号", "spell2Id": "召唤师技能2序号", "teamId": "阵营代号", "champion_name": "选用英雄", "champion_alias": "选用英雄代号", "champion_squarePortraitPath": "选用英雄方块头像路径", "spell1_name": "召唤师技能1", "spell2_name": "召唤师技能2", "spell1_iconPath": "召唤师技能1图标", "spell2_iconPath": "召唤师技能2图标", "team_color": "阵营", "assists": "助攻", "causedEarlySurrender": "发起提前投降", "champLevel": "英雄等级", "combatPlayerScore": "战斗得分", "damageDealtToObjectives": "对战略点的总伤害", "damageDealtToTurrets": "对防御塔的总伤害", "damageSelfMitigated": "自我缓和的伤害", "deaths": "死亡", "doubleKills": "双杀", "earlySurrenderAccomplice": "同意提前投降", "firstBloodAssist": "协助获得第一滴血", "firstBloodKill": "第一滴血", "firstInhibitorAssist": "协助摧毁第一座召唤水晶", "firstInhibitorKill": "摧毁第一座召唤水晶", "firstTowerAssist": "协助摧毁第一座塔", "firstTowerKill": "摧毁第一座塔", "gameEndedInEarlySurrender": "提前投降导致比赛结束", "gameEndedInSurrender": "投降导致比赛结束", "goldEarned": "金币获取", "goldSpent": "金币使用", "inhibitorKills": "摧毁召唤水晶", "item0": "装备1序号", "item1": "装备2序号", "item2": "装备3序号", "item3": "装备4序号", "item4": "装备5序号", "item5": "装备6序号", "item6": "饰品序号", "killingSprees": "大杀特杀", "kills": "击杀", "largestCriticalStrike": "最大暴击伤害", "largestKillingSpree": "最高连杀", "largestMultiKill": "最高多杀", "longestTimeSpentLiving": "最长生存时间", "magicDamageDealt": "造成的魔法伤害", "magicDamageDealtToChampions": "对英雄的魔法伤害", "magicalDamageTaken": "承受的魔法伤害", "neutralMinionsKilled": "击杀野怪", "neutralMinionsKilledEnemyJungle": "击杀敌方野区野怪", "neutralMinionsKilledTeamJungle": "击杀我方野区野怪", "objectivePlayerScore": "战略点玩家得分", "pentaKills": "五杀", "perk0": "符文1序号", "perk0Var1": "符文1：参数1", "perk0Var2": "符文1：参数2", "perk0Var3": "符文1：参数3", "perk1": "符文2序号", "perk1Var1": "符文2：参数1", "perk1Var2": "符文2：参数2", "perk1Var3": "符文2：参数3", "perk2": "符文3序号", "perk2Var1": "符文3：参数1", "perk2Var2": "符文3：参数2", "perk2Var3": "符文3：参数3", "perk3": "符文4序号", "perk3Var1": "符文4：参数1", "perk3Var2": "符文4：参数2", "perk3Var3": "符文4：参数3", "perk4": "符文5序号", "perk4Var1": "符文5：参数1", "perk4Var2": "符文5：参数2", "perk4Var3": "符文5：参数3", "perk5": "符文6序号", "perk5Var1": "符文6：参数1", "perk5Var2": "符文6：参数2", "perk5Var3": "符文6：参数3", "perkPrimaryStyle": "主系序号", "perkSubStyle": "副系序号", "physicalDamageDealt": "造成的物理伤害", "physicalDamageDealtToChampions": "对英雄的物理伤害", "physicalDamageTaken": "承受的物理伤害", "playerAugment1": "强化符文1", "playerAugment2": "强化符文2", "playerAugment3": "强化符文3", "playerAugment4": "强化符文4", "playerAugment5": "强化符文5", "playerAugment6": "强化符文6", "playerScore0": "玩家得分1", "playerScore1": "玩家得分2", "playerScore2": "玩家得分3", "playerScore3": "玩家得分4", "playerScore4": "玩家得分5", "playerScore5": "玩家得分6", "playerScore6": "玩家得分7", "playerScore7": "玩家得分8", "playerScore8": "玩家得分9", "playerScore9": "玩家得分10", "playerSubteamId": "子阵营代号", "quadraKills": "四杀", "sightWardsBoughtInGame": "购买洞察之石", "subteamPlacement": "队伍排名", "teamEarlySurrendered": "队伍提前投降", "timeCCingOthers": "控制得分", "totalDamageDealt": "造成的伤害总和", "totalDamageDealtToChampions": "对英雄的伤害总和", "totalDamageTaken": "承受伤害", "totalHeal": "输出治疗效果", "totalMinionsKilled": "击杀小兵", "totalPlayerScore": "玩家总得分", "totalScoreRank": "总得分排名", "totalTimeCrowdControlDealt": "控制时间", "totalUnitsHealed": "治疗单位数", "tripleKills": "三杀", "trueDamageDealt": "造成真实伤害", "trueDamageDealtToChampions": "对英雄的真实伤害", "trueDamageTaken": "承受的真实伤害", "turretKills": "摧毁防御塔", "unrealKills": "六杀及以上", "visionScore": "视野得分", "visionWardsBoughtInGame": "购买控制守卫", "wardsKilled": "摧毁守卫", "wardsPlaced": "放置守卫", "win": "胜利", "item0_name": "装备1", "item1_name": "装备2", "item2_name": "装备3", "item3_name": "装备4", "item4_name": "装备5", "item5_name": "装备6", "item6_name": "饰品", "item0_iconPath": "装备1图标路径", "item1_iconPath": "装备2图标路径", "item2_iconPath": "装备3图标路径", "item3_iconPath": "装备4图标路径", "item4_iconPath": "装备5图标路径", "item5_iconPath": "装备6图标路径", "item6_iconPath": "饰品图标路径", "perk0EndOfGameStatDescs": "符文1游戏结算数据", "perk1EndOfGameStatDescs": "符文2游戏结算数据", "perk2EndOfGameStatDescs": "符文3游戏结算数据", "perk3EndOfGameStatDescs": "符文4游戏结算数据", "perk4EndOfGameStatDescs": "符文5游戏结算数据", "perk5EndOfGameStatDescs": "符文6游戏结算数据", "perk0_name": "符文1名称", "perk1_name": "符文2名称", "perk2_name": "符文3名称", "perk3_name": "符文4名称", "perk4_name": "符文5名称", "perk5_name": "符文6名称", "perk0_iconPath": "符文1图标路径", "perk1_iconPath": "符文2图标路径", "perk2_iconPath": "符文3图标路径", "perk3_iconPath": "符文4图标路径", "perk4_iconPath": "符文5图标路径", "perk5_iconPath": "符文6图标路径", "perkPrimaryStyle_name": "主系名称", "perkPrimaryStyle_iconPath": "主系图标路径", "perkSubStyle_name": "副系名称", "perkSubStyle_iconPath": "副系图标路径", "playerAugment1_nameTRA": "强化符文1名称", "playerAugment2_nameTRA": "强化符文2名称", "playerAugment3_nameTRA": "强化符文3名称", "playerAugment4_nameTRA": "强化符文4名称", "playerAugment5_nameTRA": "强化符文5名称", "playerAugment6_nameTRA": "强化符文6名称", "playerAugment1_augmentIconPath": "强化符文1图标路径", "playerAugment2_augmentIconPath": "强化符文2图标路径", "playerAugment3_augmentIconPath": "强化符文3图标路径", "playerAugment4_augmentIconPath": "强化符文4图标路径", "playerAugment5_augmentIconPath": "强化符文5图标路径", "playerAugment6_augmentIconPath": "强化符文6图标路径", "playerAugment1_rarity": "强化符文1等级", "playerAugment2_rarity": "强化符文2等级", "playerAugment3_rarity": "强化符文3等级", "playerAugment4_rarity": "强化符文4等级", "playerAugment5_rarity": "强化符文5等级", "playerAugment6_rarity": "强化符文6等级", "playerSubteam_color": "子阵营", "K/D/A": "击杀/死亡/助攻", "KDA": "战损比", "CS": "补刀", "GPM": "分均经济", "GUE": "金币利用率", "CSPM": "分均补刀", "D/G": "伤害转化率", "win/lose": "胜负", "bannedChampionId": "禁用英雄序号", "bannedChampion_name": "禁用英雄", "bannedChampion_alias": "禁用英雄代号", "bannedChampion_squarePortraitPath": "禁用英雄方块头像路径", "lane": "分路", "role": "角色定位", "ally?": "是否队友？", "assists_percent": "助攻次数占比", "combatPlayerScore_percent": "战斗得分占比", "damageDealtToObjectives_percent": "对战略点的总伤害占比", "damageDealtToTurrets_percent": "对防御塔的总伤害占比", "damageSelfMitigated_percent": "自我缓和的伤害占比", "deaths_percent": "死亡次数占比", "doubleKills_percent": "双杀次数占比", "goldEarned_percent": "金币获取占比", "goldSpent_percent": "金币使用占比", "inhibitorKills_percent": "摧毁召唤水晶数量占比", "killingSprees_percent": "大杀特杀次数占比", "kills_percent": "击杀数量占比", "largestCriticalStrike_percent": "最大暴击伤害占比", "largestKillingSpree_percent": "最高连杀占比", "largestMultiKill_percent": "最高多杀占比", "longestTimeSpentLiving_percent": "最长生存时间占比", "magicDamageDealt_percent": "造成的魔法伤害占比", "magicDamageDealtToChampions_percent": "对英雄的魔法伤害占比", "magicalDamageTaken_percent": "承受的魔法伤害占比", "neutralMinionsKilled_percent": "击杀野怪数量占比", "neutralMinionsKilledEnemyJungle_percent": "击杀敌方野区野怪数量占比", "neutralMinionsKilledTeamJungle_percent": "击杀我方野区野怪数量占比", "objectivePlayerScore_percent": "战略点玩家得分占比", "pentaKills_percent": "五杀次数占比", "physicalDamageDealt_percent": "造成的物理伤害占比", "physicalDamageDealtToChampions_percent": "对英雄的物理伤害占比", "physicalDamageTaken_percent": "承受的物理伤害占比", "playerScore0_percent": "玩家得分1占比", "playerScore1_percent": "玩家得分2占比", "playerScore2_percent": "玩家得分3占比", "playerScore3_percent": "玩家得分4占比", "playerScore4_percent": "玩家得分5占比", "playerScore5_percent": "玩家得分6占比", "playerScore6_percent": "玩家得分7占比", "playerScore7_percent": "玩家得分8占比", "playerScore8_percent": "玩家得分9占比", "playerScore9_percent": "玩家得分10占比", "quadraKills_percent": "四杀次数占比", "sightWardsBoughtInGame_percent": "购买洞察之石数量占比", "timeCCingOthers_percent": "控制得分占比", "totalDamageDealt_percent": "造成的伤害总和占比", "totalDamageDealtToChampions_percent": "对英雄的伤害总和占比", "totalDamageTaken_percent": "承受伤害占比", "totalHeal_percent": "输出治疗效果占比", "totalMinionsKilled_percent": "击杀小兵数量占比", "totalPlayerScore_percent": "玩家总得分占比", "totalTimeCrowdControlDealt_percent": "控制时间占比", "totalUnitsHealed_percent": "治疗单位数占比", "tripleKills_percent": "三杀次数占比", "trueDamageDealt_percent": "造成真实伤害占比", "trueDamageDealtToChampions_percent": "对英雄的真实伤害占比", "trueDamageTaken_percent": "承受的真实伤害占比", "turretKills_percent": "摧毁防御塔数量占比", "unrealKills_percent": "六杀及以上连杀次数占比", "visionScore_percent": "视野得分占比", "visionWardsBoughtInGame_percent": "购买控制守卫数量占比", "wardsKilled_percent": "摧毁守卫数量占比", "wardsPlaced_percent": "放置守卫数量占比", "KP_percent": "参团率", "CS_percent": "补刀数占比", "assists_order": "助攻次数位次", "champLevel_order": "英雄等级位次", "combatPlayerScore_order": "战斗得分位次", "damageDealtToObjectives_order": "对战略点的总伤害位次", "damageDealtToTurrets_order": "对防御塔的总伤害位次", "damageSelfMitigated_order": "自我缓和的伤害位次", "deaths_order": "死亡次数位次", "doubleKills_order": "双杀次数位次", "goldEarned_order": "金币获取位次", "goldSpent_order": "金币使用位次", "inhibitorKills_order": "摧毁召唤水晶数量位次", "killingSprees_order": "大杀特杀次数位次", "kills_order": "击杀数量位次", "largestCriticalStrike_order": "最大暴击伤害位次", "largestKillingSpree_order": "最高连杀位次", "largestMultiKill_order": "最高多杀位次", "longestTimeSpentLiving_order": "最长生存时间位次", "magicDamageDealt_order": "造成的魔法伤害位次", "magicDamageDealtToChampions_order": "对英雄的魔法伤害位次", "magicalDamageTaken_order": "承受的魔法伤害位次", "neutralMinionsKilled_order": "击杀野怪数量位次", "neutralMinionsKilledEnemyJungle_order": "击杀敌方野区野怪数量位次", "neutralMinionsKilledTeamJungle_order": "击杀我方野区野怪数量位次", "objectivePlayerScore_order": "战略点玩家得分位次", "pentaKills_order": "五杀次数位次", "physicalDamageDealt_order": "造成的物理伤害位次", "physicalDamageDealtToChampions_order": "对英雄的物理伤害位次", "physicalDamageTaken_order": "承受的物理伤害位次", "playerScore0_order": "玩家得分1位次", "playerScore1_order": "玩家得分2位次", "playerScore2_order": "玩家得分3位次", "playerScore3_order": "玩家得分4位次", "playerScore4_order": "玩家得分5位次", "playerScore5_order": "玩家得分6位次", "playerScore6_order": "玩家得分7位次", "playerScore7_order": "玩家得分8位次", "playerScore8_order": "玩家得分9位次", "playerScore9_order": "玩家得分10位次", "quadraKills_order": "四杀次数位次", "sightWardsBoughtInGame_order": "购买洞察之石数量位次", "timeCCingOthers_order": "控制得分位次", "totalDamageDealt_order": "造成的伤害总和位次", "totalDamageDealtToChampions_order": "对英雄的伤害总和位次", "totalDamageTaken_order": "承受伤害位次", "totalHeal_order": "输出治疗效果位次", "totalMinionsKilled_order": "击杀小兵数量位次", "totalPlayerScore_order": "玩家总得分位次", "totalTimeCrowdControlDealt_order": "控制时间位次", "totalUnitsHealed_order": "治疗单位数位次", "tripleKills_order": "三杀次数位次", "trueDamageDealt_order": "造成真实伤害位次", "trueDamageDealtToChampions_order": "对英雄的真实伤害位次", "trueDamageTaken_order": "承受的真实伤害位次", "turretKills_order": "摧毁防御塔数量位次", "unrealKills_order": "六杀及以上连杀次数位次", "visionScore_order": "视野得分位次", "visionWardsBoughtInGame_order": "购买控制守卫数量位次", "wardsKilled_order": "摧毁守卫数量位次", "wardsPlaced_order": "放置守卫数量位次", "KDA_order": "战损比位次", "KP_order": "参团率位次", "CS_order": "补刀数位次", "D/G_order": "伤害转化率位次", "GUE_order": "金币利用率位次"}
                    LoLGame_info_header_keys = list(LoLGame_info_header.keys())
                    recent_LoLPlayers_statistics_output_order = [0, 13, 23, 17, 24, 22, 21, 28, 5, 3, 11, 10, 6, 12, 9, 8, 222, 32, 33, 217, 218, 220, 221, 42, 35, 36, 153, 154, 155, 156, 157, 158, 159, 189, 201, 190, 202, 191, 203, 192, 204, 193, 205, 194, 206, 208, 209, 210, 213, 214, 43, 138, 139, 71, 68, 72, 51, 50, 55, 54, 53, 52, 48, 142, 128, 81, 147, 132, 140, 134, 109, 75, 144, 133, 108, 74, 143, 70, 45, 44, 136, 141, 135, 110, 76, 145, 46, 148, 151, 150, 129, 149, 58, 211, 59, 212, 137, 77, 79, 78, 146, 60, 73, 185, 187, 173, 167, 174, 168, 175, 169, 176, 170, 177, 171, 178, 172, 41, 49, 131, 56, 57, 215, 130, 234, 228, 223, 281, 224, 268, 236, 233, 237, 229, 271, 260, 246, 276, 262, 269, 264, 248, 240, 273, 263, 247, 239, 272, 235, 226, 225, 266, 270, 265, 249, 241, 274, 227, 277, 280, 279, 261, 278, 230, 231, 267, 242, 244, 243, 282, 275, 232, 238, 284, 295, 289, 283, 342, 343, 345, 285, 329, 297, 294, 298, 290, 332, 321, 307, 337, 323, 330, 325, 309, 301, 334, 324, 308, 300, 333, 296, 287, 286, 327, 331, 326, 310, 302, 335, 288, 338, 341, 340, 322, 339, 291, 292, 346, 328, 303, 304, 305, 344, 336, 293, 299]
                    recent_LoLPlayers_data_organized = {}
                    for i in range(len(recent_LoLPlayers_statistics_output_order)):
                        key = LoLGame_info_header_keys[recent_LoLPlayers_statistics_output_order[i]]
                        recent_LoLPlayers_data_organized[key] = [LoLGame_info_header[key]]
                    recent_LoLPlayers_df = pandas.DataFrame(data = recent_LoLPlayers_data_organized)

                #下面获取最近一起玩过的云顶之弈玩家的信息（The following code captures the recently played TFT players' information）
                logPrint("是否查询云顶之弈对局记录？（输入任意键查询，否则不查询）\nSearch TFT matches? (Input anything to search or null to export data or switch for another summoner)")
                search_TFT_str = logInput()
                search_TFT = bool(search_TFT_str)
                TFTHistory_dict = {}
                if search_TFT:
                    logPrint("请设置需要查询的对局索引下界和对局数，以空格为分隔符（输入空字符以默认查询近20场对局）：\nPlease set the begin and count of the matches to be searched, split by space (Enter an empty string to search for the recent 20 matches):")
                    while True:
                        gameIndex = logInput()
                        if gameIndex == "":
                            begin_get, count_get = 0, 20
                        elif gameIndex == "0":
                            search_TFT = False
                            break
                        else:
                            try:
                                begin_get, count_get = map(int, gameIndex.split())
                            except ValueError:
                                logPrint("请以空格为分隔符输入自然数类型的对局索引下界和对局数！\nPlease enter the two nonnegative integers as the begin and count of the matches split by space!")
                                continue
                        break
                    if gameIndex == "0":
                        search_TFT = False
                        continue
                    for info_body in AllAccounts:
                        logPrint("正在加载云顶之弈对局信息……\nLoading TFT match information ...")
                        TFTHistory_get = False
                        while True:
                            try:
                                TFTHistory = await (await connection.request("GET", "/lol-match-history/v1/products/tft/%s/matches?begin=%d&count=%d" %(info_body["puuid"], begin_get, count_get))).json()
                                #logPrint(TFTHistory)
                                count = 0 #存储内部服务器错误次数（Stores the times of internal server error）
                                if "errorCode" in TFTHistory:
                                    if "500 Internal Server Error" in TFTHistory["message"]:
                                        if not error_occurred:
                                            logPrint("您所在大区的对局记录服务异常。尝试重新获取数据……\nThe match history service provided on your server isn't in place. Trying to recapture the history data ...")
                                            occurred = True
                                        while "errorCode" in TFTHistory and "500 Internal Server Error" in TFTHistory["message"] and count <= 3:
                                            count += 1
                                            logPrint("正在进行第%d次尝试……\nTimes trying: No. %d ..." %(count, count))
                                            TFTHistory = await (await connection.request("GET", "/lol-match-history/v1/products/tft/%s/matches?begin=%d&count=%d" %(info_body["puuid"], begin_get, count_get))).json()
                                currentTime = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
                                pkl5name = "Intermediate Object - TFTHistory - %s (%s).pkl" %(displayName, currentTime)
                                #with open(os.path.join(folder, pkl5name), "wb") as IntObj5:
                                    #pickle.dump(TFTHistory, IntObj5)
                                if count > 3:
                                    logPrint("云顶之弈对局记录获取失败！请等待官方修复对局记录服务！\nTFT match history capture failure! Please wait for Tencent to fix the match history service!")
                                    break
                                logPrint("玩家%s共进行%d场云顶之弈对局。\nPlayer %s has played %d TFT matches.\n" %(get_info_name(info_body), len(TFTHistory["games"]), get_info_name(info_body), len(TFTHistory["games"]))) #在这里引发键异常（Here may trigger a KeyError）
                            except KeyError:
                                if "errorCode" in TFTHistory:
                                    logPrint(TFTHistory)
                                    TFTHistory_url = "%s/lol-match-history/v1/products/tft/%s/matches?begin=0&count=200" %(connection.address, info_body["puuid"])
                                    logPrint("请打开以下网址，输入如下所示的用户名和密码，打开后在命令行中按回车键继续，或输入任意字符以切换召唤师（Please open the following website, type in the username and password accordingly and press Enter to continue or input anything to switch to another summoner）：\n网址（URL）：\t\t%s\n用户名（Username）：\triot\n密码（Password）：\t%s\n或者输入空格分隔的两个自然数以重新指定对局索引下限和对局数。\nOr submit two nonnegative integers split by space to respecify the begin and count." %(TFTHistory_url, connection.auth_key))
                                    cont = logInput()
                                    if cont == "":
                                        continue
                                    else:
                                        try:
                                            begin_get, count_get = map(int, cont.split())
                                        except ValueError:
                                            break
                                        else:
                                            continue
                            else:
                                TFTHistory_get = True
                                break
                        if not TFTHistory_get:
                            continue
                        for game in TFTHistory["games"]:
                            match_id = int(game["metadata"]["match_id"].split("_")[-1])
                            if not match_id in TFTHistory_dict: #由于云顶之弈的对局记录包含所有玩家的信息，所以如果多个玩家的对局记录包含同一场对局，则这些对局的信息一定是相同的（Because TFT match history includes all players' information, if a match is included in multiple players' match histories, then information of the matches recorded in different players' match histories must be the same）
                                TFTHistory_dict[match_id] = game
                    #由于云顶之弈的对局记录包含所有玩家的信息，所以这里考虑先整合所有小号的对局记录，再对总对局记录进行整理。如果先整理再整合，后续排序时玩家顺序的信息会丢失，因为在这种情形下根据对局序号排序，而数据框中不包含玩家序号键，无法按照玩家序号进行升序排列（Because TFT match history includes all players' information, here the program first merges all smurf accounts' match history, and then sort out the aggregate match history. Otherwise, if the program first sort out the match history respectively and then merge the result dataframe, the participantId order may be lost during the subsequent ordering, for gameId is taken to arrange the aggregate dataframe, but the key `participantId` isn't in the dataframe, and therefore the dataframe can't be arranged in the ascending order of participantId）
                    TFTHistory = list(map(lambda x: TFTHistory_dict[x], sorted(TFTHistory_dict.keys(), reverse = True)))
                    TFTHistory_header = {"gameIndex": "游戏序号", "endOfGameResult": "对局终止情况", "gameCreation": "对局创建时间戳", "game_datetime": "对局结算时间戳", "game_id": "对局序号", "game_length": "持续时长（秒）", "game_version": "对局版本", "queue_id": "队列序号", "tft_game_type": "游戏类型", "tft_set_core_name": "数据版本名称", "tft_set_number": "赛季", "gameCreationDate": "对局创建时间", "gameDate": "对局结算时间", "gameLength": "持续时长", "participantId": "玩家序号", "augment1 apiName": "强化符文1接口名称", "augment2 apiName": "强化符文2接口名称", "augment3 apiName": "强化符文3接口名称", "augment1 name": "强化符文1名称", "augment2 name": "强化符文2名称", "augment3 name": "强化符文3名称", "augment1 icon": "强化符文1图标", "augment2 icon": "强化符文2图标", "augment3 icon": "强化符文3图标", "companion content_ID": "小小英雄商品编号", "companion item_ID": "小小英雄序号", "companion skin_ID": "小小英雄皮肤序号", "companion species": "小小英雄物种", "companion name": "小小英雄名称", "companion level": "小小英雄星级", "companion rarity": "小小英雄稀有度", "gold_left": "剩余金币", "last_round": "存活回合数", "level": "等级", "placement": "名次", "players_eliminated": "淘汰玩家数", "puuid": "玩家通用唯一识别码", "riotIdGameName": "玩家昵称", "riotIdTagLine": "昵称编号", "time_eliminated": "存活时长（秒）", "total_damage_to_players": "造成玩家伤害", "last_round_format": "存活回合", "time_eliminated_norm": "存活时长", "trait0 name": "羁绊1", "trait0 num_units": "羁绊1单位数", "trait0 style": "羁绊1羁绊框颜色", "trait0 tier_current": "羁绊1当前等级", "trait0 tier_total": "羁绊1最高等级", "trait0 display_name": "羁绊1显示名", "trait0 icon_path": "羁绊1图标路径", "trait1 name": "羁绊2", "trait1 num_units": "羁绊2单位数", "trait1 style": "羁绊2羁绊框颜色", "trait1 tier_current": "羁绊2当前等级", "trait1 tier_total": "羁绊2最高等级", "trait1 display_name": "羁绊2显示名", "trait1 icon_path": "羁绊2图标路径", "trait2 name": "羁绊3", "trait2 num_units": "羁绊3单位数", "trait2 style": "羁绊3羁绊框颜色", "trait2 tier_current": "羁绊3当前等级", "trait2 tier_total": "羁绊3最高等级", "trait2 display_name": "羁绊3显示名", "trait2 icon_path": "羁绊3图标路径", "trait3 name": "羁绊4", "trait3 num_units": "羁绊4单位数", "trait3 style": "羁绊4羁绊框颜色", "trait3 tier_current": "羁绊4当前等级", "trait3 tier_total": "羁绊4最高等级", "trait3 display_name": "羁绊4显示名", "trait3 icon_path": "羁绊4图标路径", "trait4 name": "羁绊5", "trait4 num_units": "羁绊5单位数", "trait4 style": "羁绊5羁绊框颜色", "trait4 tier_current": "羁绊5当前等级", "trait4 tier_total": "羁绊5最高等级", "trait4 display_name": "羁绊5显示名", "trait4 icon_path": "羁绊5图标路径", "trait5 name": "羁绊6", "trait5 num_units": "羁绊6单位数", "trait5 style": "羁绊6羁绊框颜色", "trait5 tier_current": "羁绊6当前等级", "trait5 tier_total": "羁绊6最高等级", "trait5 display_name": "羁绊6显示名", "trait5 icon_path": "羁绊6图标路径", "trait6 name": "羁绊7", "trait6 num_units": "羁绊7单位数", "trait6 style": "羁绊7羁绊框颜色", "trait6 tier_current": "羁绊7当前等级", "trait6 tier_total": "羁绊7最高等级", "trait6 display_name": "羁绊7显示名", "trait6 icon_path": "羁绊7图标路径", "trait7 name": "羁绊8", "trait7 num_units": "羁绊8单位数", "trait7 style": "羁绊8羁绊框颜色", "trait7 tier_current": "羁绊8当前等级", "trait7 tier_total": "羁绊8最高等级", "trait7 display_name": "羁绊8显示名", "trait7 icon_path": "羁绊8图标路径", "trait8 name": "羁绊9", "trait8 num_units": "羁绊9单位数", "trait8 style": "羁绊9羁绊框颜色", "trait8 tier_current": "羁绊9当前等级", "trait8 tier_total": "羁绊9最高等级", "trait8 display_name": "羁绊9显示名", "trait8 icon_path": "羁绊9图标路径", "trait9 name": "羁绊10", "trait9 num_units": "羁绊10单位数", "trait9 style": "羁绊10羁绊框颜色", "trait9 tier_current": "羁绊10当前等级", "trait9 tier_total": "羁绊10最高等级", "trait9 display_name": "羁绊10显示名", "trait9 icon_path": "羁绊10图标路径", "trait10 name": "羁绊11", "trait10 num_units": "羁绊11单位数", "trait10 style": "羁绊11羁绊框颜色", "trait10 tier_current": "羁绊11当前等级", "trait10 tier_total": "羁绊11最高等级", "trait10 display_name": "羁绊11显示名", "trait10 icon_path": "羁绊11图标路径", "trait11 name": "羁绊12", "trait11 num_units": "羁绊12单位数", "trait11 style": "羁绊12羁绊框颜色", "trait11 tier_current": "羁绊12当前等级", "trait11 tier_total": "羁绊12最高等级", "trait11 display_name": "羁绊12显示名", "trait11 icon_path": "羁绊12图标路径", "trait12 name": "羁绊13", "trait12 num_units": "羁绊13单位数", "trait12 style": "羁绊13羁绊框颜色", "trait12 tier_current": "羁绊13当前等级", "trait12 tier_total": "羁绊13最高等级", "trait12 display_name": "羁绊13显示名", "trait12 icon_path": "羁绊13图标路径", "unit0 character_id": "英雄1：角色编号", "unit0 rarity": "英雄1：卡费", "unit0 tier": "英雄1：星级", "unit0 display_name": "英雄1：显示名", "unit0 squareIconPath": "英雄1：方块图标路径", "unit1 character_id": "英雄2：角色编号", "unit1 rarity": "英雄2：卡费", "unit1 tier": "英雄2：星级", "unit1 display_name": "英雄2：显示名", "unit1 squareIconPath": "英雄2：方块图标路径", "unit2 character_id": "英雄3：角色编号", "unit2 rarity": "英雄3：卡费", "unit2 tier": "英雄3：星级", "unit2 display_name": "英雄3：显示名", "unit2 squareIconPath": "英雄3：方块图标路径", "unit3 character_id": "英雄4：角色编号", "unit3 rarity": "英雄4：卡费", "unit3 tier": "英雄4：星级", "unit3 display_name": "英雄4：显示名", "unit3 squareIconPath": "英雄4：方块图标路径", "unit4 character_id": "英雄5：角色编号", "unit4 rarity": "英雄5：卡费", "unit4 tier": "英雄5：星级", "unit4 display_name": "英雄5：显示名", "unit4 squareIconPath": "英雄5：方块图标路径", "unit5 character_id": "英雄6：角色编号", "unit5 rarity": "英雄6：卡费", "unit5 tier": "英雄6：星级", "unit5 display_name": "英雄6：显示名", "unit5 squareIconPath": "英雄6：方块图标路径", "unit6 character_id": "英雄7：角色编号", "unit6 rarity": "英雄7：卡费", "unit6 tier": "英雄7：星级", "unit6 display_name": "英雄7：显示名", "unit6 squareIconPath": "英雄7：方块图标路径", "unit7 character_id": "英雄8：角色编号", "unit7 rarity": "英雄8：卡费", "unit7 tier": "英雄8：星级", "unit7 display_name": "英雄8：显示名", "unit7 squareIconPath": "英雄8：方块图标路径", "unit8 character_id": "英雄9：角色编号", "unit8 rarity": "英雄9：卡费", "unit8 tier": "英雄9：星级", "unit8 display_name": "英雄9：显示名", "unit8 squareIconPath": "英雄9：方块图标路径", "unit9 character_id": "英雄10：角色编号", "unit9 rarity": "英雄10：卡费", "unit9 tier": "英雄10：星级", "unit9 display_name": "英雄10：显示名", "unit9 squareIconPath": "英雄10：方块图标路径", "unit10 character_id": "英雄11：角色编号", "unit10 rarity": "英雄11：卡费", "unit10 tier": "英雄11：星级", "unit10 display_name": "英雄11：显示名", "unit10 squareIconPath": "英雄11：方块图标路径", "unit0 item0 nameId": "英雄1：装备1序号", "unit0 item0 name": "英雄1：装备1名称", "unit0 item0 squareIconPath": "英雄1：装备1方块图像路径", "unit0 item1 nameId": "英雄1：装备2序号", "unit0 item1 name": "英雄1：装备2名称", "unit0 item1 squareIconPath": "英雄1：装备2方块图像路径", "unit0 item2 nameId": "英雄1：装备3序号", "unit0 item2 name": "英雄1：装备3名称", "unit0 item2 squareIconPath": "英雄1：装备3方块图像路径", "unit1 item0 nameId": "英雄2：装备1序号", "unit1 item0 name": "英雄2：装备1名称", "unit1 item0 squareIconPath": "英雄2：装备1方块图像路径", "unit1 item1 nameId": "英雄2：装备2序号", "unit1 item1 name": "英雄2：装备2名称", "unit1 item1 squareIconPath": "英雄2：装备2方块图像路径", "unit1 item2 nameId": "英雄2：装备3序号", "unit1 item2 name": "英雄2：装备3名称", "unit1 item2 squareIconPath": "英雄2：装备3方块图像路径", "unit2 item0 nameId": "英雄3：装备1序号", "unit2 item0 name": "英雄3：装备1名称", "unit2 item0 squareIconPath": "英雄3：装备1方块图像路径", "unit2 item1 nameId": "英雄3：装备2序号", "unit2 item1 name": "英雄3：装备2名称", "unit2 item1 squareIconPath": "英雄3：装备2方块图像路径", "unit2 item2 nameId": "英雄3：装备3序号", "unit2 item2 name": "英雄3：装备3名称", "unit2 item2 squareIconPath": "英雄3：装备3方块图像路径", "unit3 item0 nameId": "英雄4：装备1序号", "unit3 item0 name": "英雄4：装备1名称", "unit3 item0 squareIconPath": "英雄4：装备1方块图像路径", "unit3 item1 nameId": "英雄4：装备2序号", "unit3 item1 name": "英雄4：装备2名称", "unit3 item1 squareIconPath": "英雄4：装备2方块图像路径", "unit3 item2 nameId": "英雄4：装备3序号", "unit3 item2 name": "英雄4：装备3名称", "unit3 item2 squareIconPath": "英雄4：装备3方块图像路径", "unit4 item0 nameId": "英雄5：装备1序号", "unit4 item0 name": "英雄5：装备1名称", "unit4 item0 squareIconPath": "英雄5：装备1方块图像路径", "unit4 item1 nameId": "英雄5：装备2序号", "unit4 item1 name": "英雄5：装备2名称", "unit4 item1 squareIconPath": "英雄5：装备2方块图像路径", "unit4 item2 nameId": "英雄5：装备3序号", "unit4 item2 name": "英雄5：装备3名称", "unit4 item2 squareIconPath": "英雄5：装备3方块图像路径", "unit5 item0 nameId": "英雄6：装备1序号", "unit5 item0 name": "英雄6：装备1名称", "unit5 item0 squareIconPath": "英雄6：装备1方块图像路径", "unit5 item1 nameId": "英雄6：装备2序号", "unit5 item1 name": "英雄6：装备2名称", "unit5 item1 squareIconPath": "英雄6：装备2方块图像路径", "unit5 item2 nameId": "英雄6：装备3序号", "unit5 item2 name": "英雄6：装备3名称", "unit5 item2 squareIconPath": "英雄6：装备3方块图像路径", "unit6 item0 nameId": "英雄7：装备1序号", "unit6 item0 name": "英雄7：装备1名称", "unit6 item0 squareIconPath": "英雄7：装备1方块图像路径", "unit6 item1 nameId": "英雄7：装备2序号", "unit6 item1 name": "英雄7：装备2名称", "unit6 item1 squareIconPath": "英雄7：装备2方块图像路径", "unit6 item2 nameId": "英雄7：装备3序号", "unit6 item2 name": "英雄7：装备3名称", "unit6 item2 squareIconPath": "英雄7：装备3方块图像路径", "unit7 item0 nameId": "英雄8：装备1序号", "unit7 item0 name": "英雄8：装备1名称", "unit7 item0 squareIconPath": "英雄8：装备1方块图像路径", "unit7 item1 nameId": "英雄8：装备2序号", "unit7 item1 name": "英雄8：装备2名称", "unit7 item1 squareIconPath": "英雄8：装备2方块图像路径", "unit7 item2 nameId": "英雄8：装备3序号", "unit7 item2 name": "英雄8：装备3名称", "unit7 item2 squareIconPath": "英雄8：装备3方块图像路径", "unit8 item0 nameId": "英雄9：装备1序号", "unit8 item0 name": "英雄9：装备1名称", "unit8 item0 squareIconPath": "英雄9：装备1方块图像路径", "unit8 item1 nameId": "英雄9：装备2序号", "unit8 item1 name": "英雄9：装备2名称", "unit8 item1 squareIconPath": "英雄9：装备2方块图像路径", "unit8 item2 nameId": "英雄9：装备3序号", "unit8 item2 name": "英雄9：装备3名称", "unit8 item2 squareIconPath": "英雄9：装备3方块图像路径", "unit9 item0 nameId": "英雄10：装备1序号", "unit9 item0 name": "英雄10：装备1名称", "unit9 item0 squareIconPath": "英雄10：装备1方块图像路径", "unit9 item1 nameId": "英雄10：装备2序号", "unit9 item1 name": "英雄10：装备2名称", "unit9 item1 squareIconPath": "英雄10：装备2方块图像路径", "unit9 item2 nameId": "英雄10：装备3序号", "unit9 item2 name": "英雄10：装备3名称", "unit9 item2 squareIconPath": "英雄10：装备3方块图像路径", "unit10 item0 nameId": "英雄11：装备1序号", "unit10 item0 name": "英雄11：装备1名称", "unit10 item0 squareIconPath": "英雄11：装备1方块图像路径", "unit10 item1 nameId": "英雄11：装备2序号", "unit10 item1 name": "英雄11：装备2名称", "unit10 item1 squareIconPath": "英雄11：装备2方块图像路径", "unit10 item2 nameId": "英雄11：装备3序号", "unit10 item2 name": "英雄11：装备3名称", "unit10 item2 squareIconPath": "英雄11：装备3方块图像路径"}
                    TFTHistory_header_keys = list(TFTHistory_header.keys())
                    TFTHistory_data = {}
                    TFT_main_player_indices = [] #云顶之弈对局记录中记录了所有玩家的数据，但是在历史记录的工作表中只要显示主召唤师的数据，因此必须知道每场对局中主召唤师的索引（Each match in TFT history records all players' data, but only the main player's data are needed to display in the match history worksheet, so the index of the main player in each match is necessary）
                    version_re = re.compile(r"\d*\.\d*\.\d*\.\d*") #云顶之弈的对局版本信息是一串字符串，从中识别四位对局版本（TFT match version is a long string, from which the 4-number version is identified）
                    TFTGameDuration_raw = []
                    for game in TFTHistory:
                        TFT_main_player_found = False
                        try:
                            for i in range(len(game["json"]["participants"])):
                                if game["json"]["participants"][i]["puuid"] in current_puuid_list or args.save_self:
                                    TFT_main_player_found = True
                                    TFT_main_player_indices.append(i) #在用户选择始终保存自己的数据时，这里的i一定是0。因此，查战绩脚本中的一些等价替换在这里不适用（When the user chooses to always save the data of the main summoner, `i` must be equal to 0 here. Therefore, the subsequent equivalent substitutions reflected in the comments at the corresponding code line of Customized Program 05 are removed in this program）
                                    break
                            if not TFT_main_player_found: #在美测服的对局序号为4420772721的对局中，不存在Volibear  PBE6玩家。这是极少见的情况，如果没有此处的判断，一旦发生这种情况，就会引起下标越界的错误（Player "Volibear  PBE6" is absent from a PBE match with matchId 4420772721, which is quite rare. Nevertheless, once it happens, an IndexError that list index out of range will be definitely thrown）
                                TFT_main_player_indices.append(-1)
                        except TypeError: #在艾欧尼亚的对局序号为8346130449的对局中，不存在玩家。这可能是因为系统维护的原因，所有人未正常进入对局，但是对局确实创建了（There doesn't exist any player in an HN1 match with matchID 8346130499. This may be due to system mainteinance, which causes all players to fail to start the game, even if the match itself has been created）
                            TFT_main_player_indices.append(-1) #当主玩家索引为-1时，表示本场对局存在异常（Main player index being -1 represents an abnormal match）
                    TFTGamePlayed = len(TFTHistory) != 0 and any(TFT_main_player_indices[i] != -1 for i in range(len(TFTHistory))) #标记该玩家是否进行过云顶之弈对局（Mark whether this summoner has played any TFT game）
                    for i in range(len(TFTHistory_header)): #各项目初始化（Initialize every feature / column）
                        key = TFTHistory_header_keys[i]
                        TFTHistory_data[key] = []
                    for i in range(len(TFTHistory)): #由于不同对局意味着不同版本，不同版本的云顶之弈数据相差较大，所以为了使得一次获取的版本能够尽可能用到多个对局中，第一层迭代器应当是对局序号（Because different matches mean different patches, and TFT data differ greatly among different patches, to make a recently captured version of TFT data applicable in as more matches as possible, the first iterator should be the ID of the matches）
                        if TFT_main_player_indices[i] == -1: #对局数据记录存在异常时的处理（Exception of match data recording exception）
                            logPrint("加载进度（Loading process）：%d/%d\t对局序号（MatchID）： %s （Exceptional match neglected）" %(i + 1, len(TFTHistory), TFTHistory[i]["metadata"]["match_id"].split("_")[1]), print_time = True)
                        else:
                            TFTHistoryJson = TFTHistory[i]["json"]
                            TFTGameVersion = version_re.search(TFTHistoryJson["game_version"]).group()
                            TFTGamePatch = ".".join(TFTGameVersion.split(".")[:2])
                            #下面针对每场对局建立总的数据资源异常处理机制（Builds the summarized data resource exceptional handling mechanism for each match）
                            ##云顶之弈强化符文（TFT augments）
                            TFTAugmentIds_match_list = sorted(set(augment for lst in list(map(lambda x: x["augments"] if "augments" in x else [], TFTHistoryJson["participants"])) for augment in lst)) #`if "augments" in x`的作用是防止早期云顶之弈对局无强化符文导致程序报错（`if "augments" in x` is used here because some early TFT matches don't contain augments and result in KeyErrors consequently）
                            for j in TFTAugmentIds_match_list:
                                if not j in TFTAugments and current_versions["TFTAugment"] != TFTGamePatch:
                                    TFTAugmentPatch_adopted = TFTGamePatch
                                    TFTAugment_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）强化符文信息（%s）获取失败！正在第%d次尝试改用%s版本的云顶之弈强化符文信息……\nAugment information (%s) of Match %d / %d (matchID: %d) capture failed! Changing to TFT augments of Patch %s ... Times tried: %d." %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, TFTAugment_recapture, TFTAugmentPatch_adopted, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTAugmentPatch_adopted, TFTAugment_recapture))
                                    while True:
                                        try:
                                            TFT = requests.get("https://raw.communitydragon.org/%s/cdragon/tft/%s.json" %(TFTAugmentPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError: #存在版本合并更新的情况（Situation like merged update exists）
                                            TFTAugmentPatch_deserted = TFTAugmentPatch_adopted
                                            TFTAugmentPatch_adopted = FindPostPatch(TFTAugmentPatch_adopted, bigPatches)
                                            TFTAugment_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to TFT augments of Patch %s ... Times tried: %d." %(TFTAugmentPatch_deserted, TFTAugment_recapture, TFTAugmentPatch_adopted, TFTAugmentPatch_deserted, TFTAugmentPatch_adopted, TFTAugment_recapture))
                                        except requests.exceptions.RequestException: #如果重新获取数据的过程中出现网络异常，那么暂时先将原始数据导入工作表中（If a network error occurs when recapturing the data, then temporarily export the initial data into the worksheet）
                                            if TFTAugment_recapture < 3:
                                                TFTAugment_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的云顶之弈强化符文信息……\nYour network environment is abnormal! Changing to TFT augments of Patch %s ... Times tried: %d." %(TFTAugment_recapture, TFTAugmentPatch_adopted, TFTAugmentPatch_adopted, TFTAugment_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的强化符文信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the augment (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的云顶之弈强化符文信息。\nTFT augment information changed to Patch %s." %(TFTAugmentPatch_adopted, TFTAugmentPatch_adopted))
                                            TFTAugments = {}
                                            for item in TFT["items"]:
                                                item_apiName = item["apiName"]
                                                TFTAugments[item_apiName] = item
                                            current_versions["TFTAugment"] = TFTAugmentPatch_adopted
                                            unmapped_keys["TFTAugment"].clear()
                                            break
                                    break
                            ##云顶之弈小小英雄（TFT companions）
                            TFTCompanionIds_match_list = sorted(set(map(lambda x: x["companion"]["content_ID"], TFTHistoryJson["participants"])))
                            for j in TFTCompanionIds_match_list:
                                if not j in TFTCompanions and current_versions["TFTCompanion"] != TFTGamePatch:
                                    TFTCompanionPatch_adopted = TFTGamePatch
                                    TFTCompanion_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）小小英雄信息（%s）获取失败！正在第%d次尝试改用%s版本的小小英雄信息……\nTFT companion information (%s) of Match %d / %d (matchID: %d) capture failed! Changing to TFT companions of Patch %s ... Times tried: %d." %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, TFTCompanion_recapture, TFTCompanionPatch_adopted, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTCompanionPatch_adopted, TFTCompanion_recapture))
                                    while True:
                                        try:
                                            TFTCompanion = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/companions.json" %(TFTCompanionPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            TFTCompanionPatch_deserted = TFTCompanionPatch_adopted
                                            TFTCompanionPatch_adopted = FindPostPatch(TFTCompanionPatch_adopted, bigPatches)
                                            TFTCompanion_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to TFT traits of Patch %s ... Times tried: %d." %(TFTCompanionPatch_deserted, TFTCompanion_recapture, TFTCompanionPatch_adopted, TFTCompanionPatch_deserted, TFTCompanionPatch_adopted, TFTCompanion_recapture))
                                        except requests.exceptions.RequestException:
                                            if TFTCompanion_recapture < 3:
                                                TFTCompanion_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的小小英雄信息……\nYour network environment is abnormal! Changing to TFT companions of Patch %s ... Times tried: %d." %(TFTCompanion_recapture, TFTCompanionPatch_adopted, TFTCompanionPatch_adopted, TFTCompanion_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的小小英雄信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the companion (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的小小英雄信息。\nTFT companion information changed to Patch %s." %(TFTCompanionPatch_adopted, TFTCompanionPatch_adopted))
                                            TFTCompanions = {}
                                            for companion_iter in TFTCompanion:
                                                contentId = companion_iter["contentId"]
                                                TFTCompanions[contentId] = companion_iter
                                            current_versions["TFTCompanion"] = TFTCompanionPatch_adopted
                                            unmapped_keys["TFTCompanion"].clear()
                                            break
                                    break
                            ##云顶之弈羁绊（TFT Traits）
                            TFTTraitIds_match_list = sorted(set(trait for s in [set(map(lambda x: x["name"], participant["traits"])) for participant in TFTHistoryJson["participants"]] for trait in s))
                            for j in TFTTraitIds_match_list:
                                if not j in TFTTraits and current_versions["TFTTrait"] != TFTGamePatch:
                                    TFTTraitPatch_adopted = TFTGamePatch
                                    TFTTrait_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）羁绊信息（%s）获取失败！正在第%d次尝试改用%s版本的羁绊信息……\nTFT trait information (%s) of Match %d / %d (matchID: %d) capture failed! Changing to TFT traits of Patch %s ... Times tried: %d." %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, TFTTrait_recapture, TFTTraitPatch_adopted, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTTraitPatch_adopted, TFTTrait_recapture))
                                    while True:
                                        try:
                                            TFTTrait = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tfttraits.json" %(TFTTraitPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            TFTTraitPatch_deserted = TFTTraitPatch_adopted
                                            TFTTraitPatch_adopted = FindPostPatch(TFTTraitPatch_adopted, bigPatches)
                                            TFTTrait_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to TFT traits of Patch %s ... Times tried: %d." %(TFTTraitPatch_deserted, TFTTrait_recapture, TFTTraitPatch_adopted, TFTTraitPatch_deserted, TFTTraitPatch_adopted, TFTTrait_recapture))
                                        except requests.exceptions.RequestException:
                                            if TFTTrait_recapture < 3:
                                                TFTTrait_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的羁绊信息……\nYour network environment is abnormal! Changing to TFT traits of Patch %s ... Times tried: %d." %(TFTTrait_recapture, TFTTraitPatch_adopted, TFTTraitPatch_adopted, TFTTrait_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的羁绊信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the trait (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的羁绊信息。\nTFT trait information changed to Patch %s." %(TFTTraitPatch_adopted, TFTTraitPatch_adopted))
                                            TFTTraits = {}
                                            for trait_iter in TFTTrait:
                                                trait_id = trait_iter["trait_id"]
                                                conditional_trait_sets = {}
                                                if "conditional_trait_sets" in trait_iter: #在英雄联盟第13赛季之前，CommunityDragon数据库中记录的羁绊信息无conditional_trait_sets项（Before Season 13, `conditional_trait_sets` item is absent from tfttraits from CommunityDragon database）
                                                    for conditional_trait_set in trait_iter["conditional_trait_sets"]:
                                                        style_idx = conditional_trait_set["style_idx"]
                                                        conditional_trait_sets[style_idx] = conditional_trait_set
                                                trait_iter["conditional_trait_sets"] = conditional_trait_sets
                                                TFTTraits[trait_id] = trait_iter
                                            current_versions["TFTTrait"] = TFTTraitPatch_adopted
                                            unmapped_keys["TFTTrait"].clear()
                                            break
                                    break
                            ##云顶之弈英雄（TFT champions）
                            TFTChampionIds_match_list = sorted(set(champion for s in [set(map(lambda x: x["character_id"], participant["units"])) for participant in TFTHistoryJson["participants"]] for champion in s))
                            for j in TFTChampionIds_match_list:
                                if not j in TFTChampions and not j.lower() in map(lambda x: x.lower(), TFTChampions.keys()) and current_versions["TFTChampion"] != TFTGamePatch:
                                    TFTChampionPatch_adopted = TFTGamePatch
                                    TFTChampion_recapture = 1
                                    logPrint("第%d/%d场对局（对局序号：%d）英雄信息（%s）获取失败！正在第%d次尝试改用%s版本的棋子信息……\nTFT champion (%s) information of Match %d / %d (matchID: %d) capture failed! Changing to TFT champions of Patch %s ... Times tried: %d." %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, TFTChampion_recapture, TFTChampionPatch_adopted, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTChampionPatch_adopted, TFTChampion_recapture))
                                    while True:
                                        try:
                                            TFTChampion = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tftchampions.json" %(TFTChampionPatch_adopted, language_cdragon[language_code])).json()
                                        except requests.exceptions.JSONDecodeError:
                                            TFTChampionPatch_deserted = TFTChampionPatch_adopted
                                            TFTChampionPatch_adopted = FindPostPatch(TFTChampionPatch_adopted, bigPatches)
                                            TFTChampion_recapture = 1
                                            logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to TFT champions of Patch %s ... Times tried: %d." %(TFTChampionPatch_deserted, TFTChampion_recapture, TFTChampionPatch_adopted, TFTChampionPatch_deserted, TFTChampionPatch_adopted, TFTChampion_recapture))
                                        except requests.exceptions.RequestException:
                                            if TFTChampion_recapture < 3:
                                                TFTChampion_recapture += 1
                                                logPrint("网络环境异常！正在第%d次尝试改用%s版本的棋子信息……\nYour network environment is abnormal! Changing to TFT champions of Patch %s ... Times tried: %d." %(TFTChampion_recapture, TFTChampionPatch_adopted, TFTChampionPatch_adopted, TFTChampion_recapture))
                                            else:
                                                logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）将采用原始数据！\nNetwork error! The original data will be used for Match %d / %d (matchID: %d)!" %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], i + 1, len(TFTHistory), TFTHistoryJson["game_id"]))
                                                break
                                        else:
                                            logPrint("已改用%s版本的棋子信息。\nTFT champion information changed to Patch %s." %(TFTChampionPatch_adopted, TFTChampionPatch_adopted))
                                            TFTChampions = {}
                                            if patch_compare(TFTChampionPatch_adopted, "13.17"): #从13.17版本开始，CommunityDragon数据库中关于云顶之弈小小英雄的数据格式发生微调（Since Patch 13.17, the format of TFT Champion data in CommunityDragon database has been modified）
                                                for TFTChampion_iter in TFTChampion:
                                                    champion_name = TFTChampion_iter["character_id"]
                                                    TFTChampions[champion_name] = TFTChampion_iter
                                            else:
                                                for TFTChampion_iter in TFTChampion:
                                                    champion_name = TFTChampion_iter["name"]
                                                    TFTChampions[champion_name] = TFTChampion_iter["character_record"] #请注意该语句与4行之前的语句的差异，并看看一开始准备数据文件时使用的是哪一种——其实你应该猜的出来（Have you noticed the difference between this statement and the statement that is 4 lines above from this statement? Also, check which statement I chose for the beginning, when I prepared the data resources. Actually, you should be able to speculate it without referring to the code）
                                            current_versions["TFTChampion"] = TFTChampionPatch_adopted
                                            unmapped_keys["TFTChampion"].clear()
                                            break
                                    break
                            ##云顶之弈装备（TFT items）
                            s = set()
                            for participant in TFTHistoryJson["participants"]:
                                for unit in participant["units"]:
                                    if "itemNames" in unit:
                                        s |= set(unit["itemNames"])
                                    elif "items" in unit:
                                        s |= set(unit["items"])
                                    else:
                                        s |= set()
                            TFTItemIds_match_list = sorted(s)
                            for j in TFTItemIds_match_list:
                                if not j in TFTItems and not j in TFTAugments:
                                    if current_versions["TFTItem"] != TFTGamePatch:
                                        TFTItemPatch_adopted = TFTGamePatch
                                        TFTItem_recapture = 1
                                        logPrint("第%d/%d场对局（对局序号：%d）装备信息（%s）获取失败！正在第%d次尝试改用%s版本的云顶之弈装备信息……\nTFT item information (%s) of Match %d / %d (matchID: %d) capture failed! Changing to TFT items of Patch %s ... Times tried: %d." %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, TFTItem_recapture, TFTItemPatch_adopted, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTItemPatch_adopted, TFTItem_recapture))
                                        while True:
                                            try:
                                                TFTItem = requests.get("https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tftitems.json" %(TFTItemPatch_adopted, language_cdragon[language_code])).json()
                                            except requests.exceptions.JSONDecodeError:
                                                TFTItemPatch_deserted = TFTItemPatch_adopted
                                                TFTItemPatch_adopted = FindPostPatch(TFTItemPatch_adopted, bigPatches)
                                                TFTItemPatch_recapture = 1
                                                logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to TFT items of Patch %s ... Times tried: %d." %(TFTItemPatch_deserted, TFTItem_recapture, TFTItemPatch_adopted, TFTItemPatch_deserted, TFTItemPatch_adopted, TFTItem_recapture))
                                            except requests.exceptions.RequestException:
                                                if TFTItem_recapture < 3:
                                                    TFTItem_recapture += 1
                                                    logPrint("网络环境异常！正在第%d次尝试改用%s版本的云顶之弈装备信息……\nYour network environment is abnormal! Changing to TFT items of Patch %s ... Times tried: %d." %(TFTItem_recapture, TFTItemPatch_adopted, TFTItemPatch_adopted, TFTItem_recapture))
                                                else:
                                                    logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的装备信息（%d）将采用原始数据！\nNetwork error! The original data will be used for the item (%d) of Match %d / %d (matchID: %d)!" %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"]))
                                                    break
                                            else:
                                                logPrint("已改用%s版本的云顶之弈装备信息。\nTFT item information changed to Patch %s." %(TFTItemPatch_adopted, TFTItemPatch_adopted))
                                                TFTItems = {}
                                                for TFTItem_iter in TFTItem:
                                                    item_id = TFTItem_iter["id"]
                                                    TFTItems[item_id] = TFTItem_iter
                                                current_versions["TFTItem"] = TFTItemPatch_adopted
                                                unmapped_keys["TFTItem"].clear()
                                                break
                                    #由于云顶之弈基础数据中也包含装备信息，这里将重新获取对局版本的云顶之弈基础数据（Because TFT basic data contain item data, here the program recaptures TFT basic data of the match version）
                                    if current_versions["TFTAugment"] != TFTGamePatch:
                                        TFTAugmentPatch_adopted = TFTGamePatch
                                        TFTAugment_recapture = 1
                                        while True:
                                            try:
                                                TFT = requests.get("https://raw.communitydragon.org/%s/cdragon/tft/%s.json" %(TFTAugmentPatch_adopted, language_cdragon[language_code])).json()
                                            except requests.exceptions.JSONDecodeError:
                                                TFTAugmentPatch_deserted = TFTAugmentPatch_adopted
                                                TFTAugmentPatch_adopted = FindPostPatch(TFTAugmentPatch_adopted, bigPatches)
                                                TFTAugment_recapture = 1
                                                logPrint("%s版本文件不存在！正在第%s次尝试转至%s版本……\n%s patch file doesn't exist! Changing to TFT augments of Patch %s ... Times tried: %d." %(TFTAugmentPatch_deserted, TFTAugment_recapture, TFTAugmentPatch_adopted, TFTAugmentPatch_deserted, TFTAugmentPatch_adopted, TFTAugment_recapture))
                                            except requests.exceptions.RequestException: #如果重新获取数据的过程中出现网络异常，那么暂时先将原始数据导入工作表中（If a network error occurs when recapturing the data, then temporarily export the initial data into the worksheet）
                                                if TFTAugment_recapture < 3:
                                                    TFTAugment_recapture += 1
                                                    logPrint("网络环境异常！正在第%d次尝试改用%s版本的云顶之弈强化符文信息……\nYour network environment is abnormal! Changing to TFT augments of Patch %s ... Times tried: %d." %(TFTAugment_recapture, TFTAugmentPatch_adopted, TFTAugmentPatch_adopted, TFTAugment_recapture))
                                                else:
                                                    logPrint("网络环境异常！第%d/%d场对局（对局序号：%d）的强化符文信息（%s）将采用原始数据！\nNetwork error! The original data will be used for the augment (%s) of Match %d / %d (matchID: %d)!" %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], j, j, i + 1, len(TFTHistory), TFTHistoryJson["game_id"]))
                                                    break
                                            else:
                                                logPrint("已改用%s版本的云顶之弈强化符文信息。\nTFT augment information changed to Patch %s." %(TFTAugmentPatch_adopted, TFTAugmentPatch_adopted))
                                                TFTAugments = {}
                                                for item in TFT["items"]:
                                                    item_apiName = item["apiName"]
                                                    TFTAugments[item_apiName] = item
                                                current_versions["TFTAugment"] = TFTAugmentPatch_adopted
                                                unmapped_keys["TFTAugment"].clear()
                                                break
                                    break
                            #下面开始整理数据（Sorts out the data）
                            for j in range(len(TFTHistory_header)):
                                key = TFTHistory_header_keys[j]
                                if j == 0: #游戏序号（`gameIndex`）
                                    for k in range(len(TFTHistory[i]["metadata"]["participants"])): #这里选择遍历元数据子字典中的玩家，而不是json子字典中的玩家，是因为前者不会包含电脑玩家的玩家通用唯一识别码，而后者会。显然，统计最近一起玩过的玩家数据不应当包含电脑玩家（Here the for-loop traverses the participants saved in the "metadata" sub-dictionary instead of the "json" sub-dictionary. This is becasue puuid of bot players isn't included in the former dictionary, but included in the latter dictionary. Obviously, they shouldn't counted as a recently played summoner）
                                        if not TFTHistory[i]["json"]["participants"][k]["puuid"] in current_puuid_list or args.save_self:
                                            TFTHistory_data[key].append(i + 1)
                                elif j <= 13:
                                    for k in range(len(TFTHistory[i]["metadata"]["participants"])):
                                        if not TFTHistory[i]["json"]["participants"][k]["puuid"] in current_puuid_list or args.save_self:
                                            if j == 1: #对局终止情况（`endOfGameResult`）
                                                if "endOfGameResult" in TFTHistoryJson:
                                                    TFTHistory_data[key].append(endOfGameResults[TFTHistoryJson["endOfGameResult"]])
                                                else:
                                                    TFTHistory_data[key].append("")
                                            elif j == 2: #对局创建时间戳（`gameCreation`）
                                                TFTHistory_data[key].append(TFTHistoryJson.get("gameCreation", "")) #14.6版本之前的云顶之弈对局信息中没有`gameCreation`这个键（The key `gameCreation` doesn't exist in information of TFT matches before Patch 14.6）
                                            elif j == 6: #对局版本（`game_version`）
                                                TFTHistory_data[key].append(TFTGameVersion)
                                            elif j == 8: #游戏类型（`tft_game_type`）
                                                TFTHistory_data[key].append(gamemodes[TFTHistoryJson["queue_id"]]["description"] if TFTHistoryJson["queue_id"] in gamemodes else "")
                                            elif j == 9: #数据版本名称（`tft_set_core_name`）
                                                TFTHistory_data[key].append(TFTHistoryJson.get("tft_set_core_name", "")) #在云顶之弈第7赛季之前，TFTHistoryJson中无tft_set_core_name这一键（Before TFTSet7, tft_set_core_name isn't present as a key of `TFTHistoryJson`）
                                            elif j == 11: #对局创建时间（`gameCreationDate`）
                                                if "gameCreation" in TFTHistoryJson:
                                                    gameCreation = int(TFTHistoryJson["gameCreation"])
                                                    gameCreationDate = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(gameCreation // 1000))
                                                    gameCreationDate_fraction = gameCreation / 1000 - gameCreation // 1000
                                                    to_append = gameCreationDate + ("{0:.3}".format(gameCreationDate_fraction))[1:5]
                                                else:
                                                    to_append = ""
                                                TFTHistory_data[key].append(to_append)
                                            elif j == 12: #对局结算时间（`gameDate`）
                                                game_datetime = int(TFTHistoryJson["game_datetime"])
                                                game_date = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(game_datetime // 1000))
                                                game_date_fraction = game_datetime / 1000 - game_datetime // 1000
                                                to_append = game_date + ("{0:.3}".format(game_date_fraction))[1:5]
                                                TFTHistory_data[key].append(to_append)
                                            elif j == 13: #持续时长（`gameLength`）
                                                TFTHistory_data[key].append("%d:%02d" %(int(TFTHistoryJson["game_length"]) // 60, int(TFTHistoryJson["game_length"]) % 60))
                                            else:
                                                TFTHistory_data[key].append(TFTHistoryJson[key])
                                elif j <= 42: #对于一些容易产生争议和报错的情况，引入to_append变量以简化代码。下同（Variable `to_append` is introduced to simplify the code in case of some controversy that produces errors easily. So does the following）
                                    for k in range(len(TFTHistory[i]["metadata"]["participants"])): #这里没有遵循迭代器命名原则，因为云顶之弈对局记录的赋值代码中包含了云顶之弈对局信息的赋值代码（Here the iterator naming principle isn't followed, because assignment code of TFT game information are included in those of TFT match information）
                                        TFTPlayer = TFTHistoryJson["participants"][k]
                                        if j == 14: #玩家序号（`participantId`）
                                            if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(k + 1)
                                        elif j >= 15 and j <= 23: #强化符文相关键（Augment-related keys）
                                            if "augments" in TFTPlayer:
                                                augment_index = (j - 15) % 3
                                                subkey_index = (j - 15) // 3
                                                if augment_index < len(TFTPlayer["augments"]):
                                                    TFTAugmentId = TFTPlayer["augments"][augment_index]
                                                    if subkey_index == 0:
                                                        to_append = TFTAugmentId
                                                    elif TFTAugmentId in TFTAugments:
                                                        to_append = TFTAugments[TFTAugmentId][key.split()[-1]]
                                                    elif TFTAugmentId in TFTAugments_initial:
                                                        to_append = TFTAugments_initial[TFTAugmentId][key.split()[-1]]
                                                    else:
                                                        if not TFTAugmentId in unmapped_keys["TFTAugment"]:
                                                            unmapped_keys["TFTAugment"].add(TFTAugmentId)
                                                            logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）强化符文信息（%s）获取失败！将采用原始数据！\n[%d. %s] TFT augment information (%s) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion, TFTAugmentId, j, key, TFTAugmentId, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion))
                                                        to_append = TFTAugmentId if subkey_index == 1 else ""
                                                else:
                                                    to_append = ""
                                            else:
                                                to_append = "" #云顶之弈刚出的时候，没有强化符文的概念（The concept of "augment" didn't appear at the beginning of TFT）
                                            if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                                        elif j >= 24 and j <= 30: #小小英雄相关键（Companion-related keys）
                                            TFTCompanionId = TFTPlayer["companion"]["content_ID"]
                                            if j <= 27:
                                                to_append = TFTPlayer["companion"][key.split()[-1]]
                                            elif TFTCompanionId in TFTCompanions:
                                                to_append = TFTCompanions[TFTCompanionId][key.split()[-1]] if j <= 29 else rarities[TFTCompanions[TFTCompanionId][key.split()[-1]]]
                                            elif TFTCompanionId in TFTCompanions_initial:
                                                to_append = TFTCompanions_initial[TFTCompanionId][key.split()[-1]] if j <= 29 else rarities[TFTCompanions_initial[TFTCompanionId][key.split()[-1]]]
                                            else:
                                                if not TFTCompanionId in unmapped_keys["TFTCompanion"]:
                                                    unmapped_keys["TFTCompanion"].add(TFTCompanionId)
                                                    logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）小小英雄信息（%s）获取失败！将采用原始数据！\n[%d. %s] TFT companion information (%s) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion, TFTCompanionId, j, key, TFTCompanionId, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion))
                                                to_append = TFTCompanionId if j == 28 else ""
                                            if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                                        elif j == 37 or j == 38: #玩家昵称和昵称编号（`riotIdGameName` and `riotIdTagLine`）
                                            if key in TFTPlayer:
                                                to_append = TFTPlayer[key]
                                            else:
                                                if TFTPlayer["puuid"] in infos:
                                                    TFTPlayer_info_body = infos[TFTPlayer["puuid"]]
                                                    to_append = TFTPlayer_info_body["gameName"] if j == 37 else TFTPlayer_info_body["tagLine"]
                                                else:
                                                    if TFTPlayer["puuid"] == "00000000-0000-0000-0000-000000000000": #在云顶之弈（新手教程）中，无法通过电脑玩家的玩家通用唯一识别码（00000000-0000-0000-0000-000000000000）来查询其召唤师名称和序号（Summoner names and IDs of bot players in TFT (Tutorial) can't be searched for according to their puuid: 00000000-0000-0000-0000-000000000000）
                                                        to_append = ""
                                                    else:
                                                        TFTPlayer_info_recapture = 0
                                                        TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"])
                                                        while not TFTPlayer_info["info_got"] and TFTPlayer_info["body"]["httpStatus"] != 404 and TFTPlayer_info_recapture < 3:
                                                            logPrint(TFTPlayer_info["message"])
                                                            TFTPlayer_info_recapture += 1
                                                            logPrint("第%d/%d场对局（对局序号：%d）玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of Player (puuid: %s) in Match %d / %d (matchID: %d) capture failed! Recapturing this player's information ... Times tried: %d." %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTPlayer["puuid"], TFTPlayer_info_recapture, TFTPlayer["puuid"], i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTPlayer_info_recapture))
                                                            TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"])
                                                        if TFTPlayer_info["info_got"]:
                                                            TFTPlayer_info_body = TFTPlayer_info["body"]
                                                            infos[TFTPlayer["puuid"]] = TFTPlayer_info_body #虽然即使infos中已经存在该召唤师信息时也会执行这一步，但不会影响数据的准确性（Despite the this summoner's existence in `infos`, running this statement won't influence data accuracy）
                                                            to_append = TFTPlayer_info_body["gameName"] if j == 37 else TFTPlayer_info_body["tagLine"]
                                                        else:
                                                            logPrint(TFTPlayer_info["message"])
                                                            to_append = ""
                                            if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                                        elif j == 41: #存活回合（`last_round_format`）
                                            lastRound = TFTPlayer["last_round"]
                                            if lastRound <= 3:
                                                bigRound = 1
                                                smallRound = lastRound
                                            else:
                                                bigRound = (lastRound + 3) // 7 + 1
                                                smallRound = (lastRound + 3) % 7 + 1
                                            to_append = "%d-%d" %(bigRound, smallRound)
                                            if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                                        elif j == 42: #存活时长（`time_eliminated_norm`）
                                            to_append = "%d:%02d" %(int(TFTPlayer["time_eliminated"]) // 60, int(TFTPlayer["time_eliminated"]) % 60)
                                            if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                                                TFTGameDuration_raw.append(TFTPlayer["time_eliminated"])
                                        else:
                                            to_append = TFTPlayer[key]
                                            if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                                elif j <= 133: #云顶之弈羁绊相关键（TFT trait-related keys）
                                    trait_index = (j - 43) // 7
                                    subkey_index = (j - 43) % 7
                                    for k in range(len(TFTHistory[i]["metadata"]["participants"])):
                                        TFTPlayer = TFTHistoryJson["participants"][k]
                                        TFTPlayer_Traits = TFTPlayer["traits"]
                                        if TFTPlayer["puuid"] in infos:
                                            TFTPlayer_info_body = infos[TFTPlayer["puuid"]]
                                        elif TFTPlayer["puuid"] != "00000000-0000-0000-0000-000000000000":
                                            TFTPlayer_info_recapture = 0
                                            TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"]) #这里的玩家信息仅用于模板羁绊的提示（The summoner information here is only used for the prompt of TemplateTrait）
                                            while not TFTPlayer_info["info_got"] and TFTPlayer_info["body"]["httpStatus"] != 404 and TFTPlayer_info_recapture < 3:
                                                logPrint(TFTPlayer_info["message"])
                                                TFTPlayer_info_recapture += 1
                                                logPrint("第%d/%d场对局（对局序号：%d）玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of Player (puuid: %s) in Match %d / %d (matchID: %d) capture failed! Recapturing this player's information ... Times tried: %d." %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTPlayer["puuid"], TFTPlayer_info_recapture, TFTPlayer["puuid"], i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTPlayer_info_recapture))
                                                TFTPlayer_info = await get_info(connection, TFTPlayer["puuid"])
                                            if TFTPlayer_info["info_got"]:
                                                TFTPlayer_info_body = TFTPlayer_info["body"]
                                                infos[TFTPlayer["puuid"]] = TFTPlayer_info_body
                                            else:
                                                logPrint(TFTPlayer_info["message"])
                                                logPrint("第%d/%d场对局（对局序号：%d）玩家信息（玩家通用唯一识别码：%s）获取失败！\nInformation of Player (puuid: %s) in Match %d / %d (matchID: %d) capture failed!" %(i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTPlayer["puuid"], TFTPlayer["puuid"], i + 1, len(TFTHistory), TFTHistoryJson["game_id"], TFTPlayer_info_recapture))
                                        if trait_index < len(TFTPlayer_Traits): #在这个小于的问题上纠结了很久[敲打]——下标是从0开始的。假设API上记录了n个羁绊，那么当程序正在获取第n个羁绊时，就会引起下标越界的问题。所以这里不能使用小于等于号（I stuck at this less than sign for too long xD - note that the index begins from 0. Suppose there're totally n traits recorded in LCU API. Then, when the program is trying to capture the n-th trait, it'll throw an IndexError. That's why the "less than or equal to" sign can't be used here）
                                            TFTTrait_iter = TFTPlayer_Traits[trait_index]
                                            TFTTraitId = TFTTrait_iter["name"]
                                            if TFTTraitId == "TemplateTrait": #CommunityDragon数据库中没有收录模板羁绊的数据（Data about TemplateTrait aren't archived in CommunityDragon database）
                                                if subkey_index == 4 and TFTPlayer["puuid"] != "00000000-0000-0000-0000-000000000000": #在艾欧尼亚的对局序号为4959597974的对局中，存在一个模板羁绊，没有tier_total这个键（There exists a TemplateTrait without the key `tier_total` in an Ionia match with matchID 4959597974）
                                                    to_append = ""
                                                    logPrint("警告：对局%d中玩家%s（玩家通用唯一识别码：%s）的第%d个羁绊是模板羁绊！\nWarning: Trait No. %d of the player %s (puuid: %s) in the match %d is TemplateTrait." %(TFTHistoryJson["game_id"], get_info_name(TFTPlayer_info_body), TFTPlayer["puuid"], trait_index + 1, trait_index + 1, get_info_name(TFTPlayer_info_body), TFTPlayer["puuid"], TFTHistoryJson["game_id"]))
                                                else:
                                                    to_append = TFTTraitId if subkey_index == 5 else "" if subkey_index == 6 else TFTTrait_iter[key.split()[-1]]
                                            else:
                                                if subkey_index <= 4:
                                                    if subkey_index == 2:
                                                        to_append = traitStyles[TFTTrait_iter[key.split()[-1]]]
                                                    else:
                                                        to_append = TFTTrait_iter[key.split()[-1]]
                                                elif TFTTraitId in TFTTraits:
                                                    to_append = TFTTraits[TFTTraitId][key.split()[-1]]
                                                elif TFTTraitId in TFTTraits_initial:
                                                    to_append = TFTTraits_initial[TFTTraitId][key.split()[-1]]
                                                else:
                                                    if not TFTTraitId in unmapped_keys["TFTTrait"]:
                                                        unmapped_keys["TFTTrait"].add(TFTTraitId)
                                                        logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）羁绊信息（%s）获取失败！将采用原始数据！\n[%d. %s] TFT trait information (%s) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion, TFTTraitId, j, key, TFTTraitId, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion))
                                                    to_append = TFTTraitId if subkey_index == 5 else ""
                                        else:
                                            to_append = ""
                                        if not TFTPlayer["puuid"] in current_puuid_list or args.save_self:
                                            TFTHistory_data[key].append(to_append)
                                else:
                                    for k in range(len(TFTHistory[i]["metadata"]["participants"])):
                                        TFTPlayer_Units = TFTHistoryJson["participants"][k]["units"]
                                        if j <= 188: #云顶之弈英雄相关键（TFT champion-related keys）
                                            unit_index = (j - 134) // 5
                                            subkey_index = (j - 134) % 5
                                            if unit_index < len(TFTPlayer_Units):
                                                TFTChampion_iter = TFTPlayer_Units[unit_index]
                                                TFTChampionId = TFTChampion_iter["character_id"]
                                                if subkey_index >= 3:
                                                    #character_id_lower = TFTPlayer_Units[unit_index]["character_id"].lower()
                                                    #TFTChampion_keys_lower = list(map(lambda x: x.lower(), list(TFTChampions.keys())))
                                                    if TFTChampionId in TFTChampions:
                                                        to_append = TFTChampions[TFTChampionId][key.split()[-1]]
                                                    elif TFTChampionId in TFTChampions_initial:
                                                        to_append = TFTChampions_initial[TFTChampionId][key.split()[-1]]
                                                    elif TFTChampionId.lower() in map(lambda x: x.lower(), TFTChampions.keys()): #在获取艾欧尼亚对局序号为8390690410的英雄信息时，由于雷克塞的英雄序号大小写的原因，会引发键异常（KeyError is caused due to the case of "RekSai" string when the program is getting data from an Ionia match with matchID 8390690410）
                                                        TFTChampion_index = list(map(lambda x: x.lower(), TFTChampions.keys())).index(TFTChampionId.lower())
                                                        to_append = list(TFTChampions.values())[TFTChampion_index][key.split()[-1]]
                                                    elif TFTChampionId.lower() in map(lambda x: x.lower(), TFTChampions_initial.keys()):
                                                        TFTChampion_index = list(map(lambda x: x.lower(), TFTChampions_initial.keys())).index(TFTChampionId.lower())
                                                        to_append = list(TFTChampions_initial.values())[TFTChampion_index][key.split()[-1]]
                                                    else:
                                                        if not TFTChampionId in unmapped_keys["TFTCompanion"]:
                                                            unmapped_keys["TFTCompanion"].add(TFTChampionId)
                                                            logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）棋子信息（%s）获取失败！将采用原始数据！\n[%d. %s] TFT champion information (%s) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion, TFTChampionId, j, key, TFTChampionId, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion))
                                                        to_append = TFTChampionId if subkey_index == 3 else ""
                                                else:
                                                    to_append = TFTPlayer_Units[unit_index][key.split()[-1]]
                                            else:
                                                to_append = ""
                                            if not TFTHistoryJson["participants"][k]["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                                        else:
                                            unit_index = (j - 189) // 9
                                            item_index = (j - 189) // 3 % 3
                                            subkey_index = (j - 189) % 3
                                            if unit_index < len(TFTPlayer_Units): #很少有英雄单位可以有3个装备（Merely do champion units have full items）
                                                if "itemNames" in TFTPlayer_Units[unit_index] and item_index < len(TFTPlayer_Units[unit_index]["itemNames"]):
                                                    TFTItemId = TFTPlayer_Units[unit_index]["itemNames"][item_index]
                                                    if subkey_index == 0:
                                                        to_append = TFTItemId
                                                    elif TFTItemId in TFTItems:
                                                        to_append = TFTItems[TFTItemId][key.split()[-1]]
                                                    elif TFTItemId in TFTItems_initial:
                                                        to_append = TFTItems_initial[TFTItemId][key.split()[-1]]
                                                    elif TFTItemId in TFTAugments: #云顶之弈基础数据文件中存在部分云顶之弈装备数据文件中没有的装备（Some items are present in the TFT basic data file but absent from the TFT item data file）
                                                        item_basic_dict = {"nameId": "apiName", "name": "name", "squareIconPath": "icon"} #云顶之弈装备数据文件和云顶之弈基础数据文件的格式不一致（The formats between TFT basic data and TFT item data are different）
                                                        to_append = TFTAugments[TFTItemId][item_basic_dict[key.split()[-1]]]
                                                    elif TFTItemId in TFTAugments_initial:
                                                        item_basic_dict = {"nameId": "apiName", "name": "name", "squareIconPath": "icon"} #云顶之弈装备数据文件和云顶之弈基础数据文件的格式不一致（The formats between TFT basic data and TFT item data are different）
                                                        to_append = TFTAugments_initial[TFTItemId][item_basic_dict[key.split()[-1]]]
                                                    else:
                                                        if not TFTItemId in unmapped_keys["TFTItem"]:
                                                            unmapped_keys["TFTItem"].add(TFTItemId)
                                                            logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）装备信息（%s）获取失败！将采用原始数据！\n[%d. %s] TFT item information (%s) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion, TFTItemId, j, key, TFTItemId, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion))
                                                        to_append = TFTItemId if subkey_index == 1 else ""
                                                elif "items" in TFTPlayer_Units[unit_index] and item_index < len(TFTPlayer_Units[unit_index]["items"]): #在12.4版本之前，装备是通过序号而不是接口名称在LCU API中被存储的（Before Patch 12.4, items are stored via itemIDs instead of itemNames）
                                                    TFTItemId = TFTPlayer_Units[unit_index]["items"][item_index]
                                                    if subkey_index == 0:
                                                        to_append = TFTItemId
                                                    elif TFTItemId in TFTItems:
                                                        to_append = TFTItems[TFTItemId][key.split()[-1]]
                                                    elif TFTItemId in TFTItems_initial:
                                                        to_append = TFTItems_initial[TFTItemId][key.split()[-1]]
                                                    elif TFTItemId in TFTAugments:
                                                        item_basic_dict = {"nameId": "apiName", "name": "name", "squareIconPath": "icon"}
                                                        to_append = TFTAugments[TFTItemId][item_basic_dict[key.split()[-1]]]
                                                    elif TFTItemId in TFTAugments_initial:
                                                        item_basic_dict = {"nameId": "apiName", "name": "name", "squareIconPath": "icon"}
                                                        to_append = TFTAugments_initial[TFTItemId][item_basic_dict[key.split()[-1]]]
                                                    else:
                                                        if not TFTItemId in unmapped_keys["TFTItem"]:
                                                            unmapped_keys["TFTItem"].add(TFTItemId)
                                                            logPrint("【%d. %s】第%d/%d场对局（对局序号：%d，对局版本：%s）装备信息（%s）获取失败！将采用原始数据！\n[%d. %s] TFT item information (%s) of Match %d / %d (matchID: %d, gameVersion: %s) capture failed! The original data will be used for this match!" %(j, key, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion, TFTItemId, j, key, TFTItemId, i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"], TFTGameVersion))
                                                        to_append = TFTItemId if subkey_index == 1 else ""
                                                else:
                                                    to_append = ""
                                            else:
                                                to_append = ""
                                            if not TFTHistoryJson["participants"][k]["puuid"] in current_puuid_list or args.save_self:
                                                TFTHistory_data[key].append(to_append)
                            logPrint("加载进度（Loading process）：%d/%d\t对局序号（MatchID）： %d" %(i + 1, len(TFTHistory), TFTHistory[i]["json"]["game_id"]), print_time = True)
                    recent_TFTPlayers_statistics_output_order = [37, 38, 36, 4, 11, 12, 6, 13, 7, 8, 28, 29, 30, 33, 41, 42, 31, 40, 35, 34, 18, 19, 20, 137, 135, 136, 190, 193, 196, 142, 140, 141, 199, 202, 205, 147, 145, 146, 208, 211, 214, 152, 150, 151, 217, 220, 223, 157, 155, 156, 226, 229, 232, 162, 160, 161, 235, 238, 241, 167, 165, 166, 244, 247, 250, 172, 170, 171, 253, 256, 259, 177, 175, 176, 262, 265, 268, 182, 180, 181, 271, 274, 277, 187, 185, 186, 280, 283, 286, 48, 44, 45, 46, 47, 55, 51, 52, 53, 54, 62, 58, 59, 60, 61, 69, 65, 66, 67, 68, 76, 72, 73, 74, 75, 83, 79, 80, 81, 82, 90, 86, 87, 88, 89, 97, 93, 94, 95, 96, 104, 100, 101, 102, 103, 111, 107, 108, 109, 110, 118, 114, 115, 116, 117, 125, 121, 122, 123, 124, 132, 128, 129, 130, 131]
                    recent_TFTPlayers_data_organized = {}
                    for i in range(len(recent_TFTPlayers_statistics_output_order)):
                        key = TFTHistory_header_keys[recent_TFTPlayers_statistics_output_order[i]]
                        recent_TFTPlayers_data_organized[key] = TFTHistory_data[key]
                        #logPrint("近期一起玩过的云顶之弈玩家数据重排进度（Rearranging process of recently played summoner (TFT) data）：%d/%d" %(i + 1, len(recent_TFTPlayers_statistics_output_order)), end = "\r")
                    #logPrint("正在创建数据框……\nCreating the dataframe ...")
                    recent_TFTPlayers_df = pandas.DataFrame(data = recent_TFTPlayers_data_organized)
                    #logPrint("数据框创建完成！\nDataframe creation finished!")
                    if not TFTGamePlayed:
                        logPrint("这位召唤师从5月1日起就没有进行过任何云顶之弈对局。\nThis summoner hasn't played any TFT game yet since May 1st.")
                    recent_TFTPlayers_df = pandas.concat([pandas.DataFrame([TFTHistory_header])[recent_TFTPlayers_df.columns], recent_TFTPlayers_df], ignore_index = True)
                    # recent_TFTPlayers_df = pandas.concat([recent_TFTPlayers_dfs[0].iloc[:1]] + list(map(lambda x: x.iloc[1:], recent_TFTPlayers_dfs)), ignore_index = True)
                    # recent_TFTPlayers_df = pandas.concat([recent_TFTPlayers_df.iloc[:1], recent_TFTPlayers_df.iloc[1:].sort_values(by = "gameCreation")], ignore_index = True)
                if not search_TFT:
                    TFTHistory_header = {"gameIndex": "游戏序号", "endOfGameResult": "对局终止情况", "gameCreation": "对局创建时间戳", "game_datetime": "对局结算时间戳", "game_id": "对局序号", "game_length": "持续时长（秒）", "game_version": "对局版本", "queue_id": "队列序号", "tft_game_type": "游戏类型", "tft_set_core_name": "数据版本名称", "tft_set_number": "赛季", "gameCreationDate": "对局创建时间", "gameDate": "对局结算时间", "gameLength": "持续时长", "participantId": "玩家序号", "augment1 apiName": "强化符文1接口名称", "augment2 apiName": "强化符文2接口名称", "augment3 apiName": "强化符文3接口名称", "augment1 name": "强化符文1名称", "augment2 name": "强化符文2名称", "augment3 name": "强化符文3名称", "augment1 icon": "强化符文1图标", "augment2 icon": "强化符文2图标", "augment3 icon": "强化符文3图标", "companion content_ID": "小小英雄商品编号", "companion item_ID": "小小英雄序号", "companion skin_ID": "小小英雄皮肤序号", "companion species": "小小英雄物种", "companion name": "小小英雄名称", "companion level": "小小英雄星级", "companion rarity": "小小英雄稀有度", "gold_left": "剩余金币", "last_round": "存活回合数", "level": "等级", "placement": "名次", "players_eliminated": "淘汰玩家数", "puuid": "玩家通用唯一识别码", "riotIdGameName": "玩家昵称", "riotIdTagLine": "昵称编号", "time_eliminated": "存活时长（秒）", "total_damage_to_players": "造成玩家伤害", "last_round_format": "存活回合", "time_eliminated_norm": "存活时长", "trait0 name": "羁绊1", "trait0 num_units": "羁绊1单位数", "trait0 style": "羁绊1羁绊框颜色", "trait0 tier_current": "羁绊1当前等级", "trait0 tier_total": "羁绊1最高等级", "trait0 display_name": "羁绊1显示名", "trait0 icon_path": "羁绊1图标路径", "trait1 name": "羁绊2", "trait1 num_units": "羁绊2单位数", "trait1 style": "羁绊2羁绊框颜色", "trait1 tier_current": "羁绊2当前等级", "trait1 tier_total": "羁绊2最高等级", "trait1 display_name": "羁绊2显示名", "trait1 icon_path": "羁绊2图标路径", "trait2 name": "羁绊3", "trait2 num_units": "羁绊3单位数", "trait2 style": "羁绊3羁绊框颜色", "trait2 tier_current": "羁绊3当前等级", "trait2 tier_total": "羁绊3最高等级", "trait2 display_name": "羁绊3显示名", "trait2 icon_path": "羁绊3图标路径", "trait3 name": "羁绊4", "trait3 num_units": "羁绊4单位数", "trait3 style": "羁绊4羁绊框颜色", "trait3 tier_current": "羁绊4当前等级", "trait3 tier_total": "羁绊4最高等级", "trait3 display_name": "羁绊4显示名", "trait3 icon_path": "羁绊4图标路径", "trait4 name": "羁绊5", "trait4 num_units": "羁绊5单位数", "trait4 style": "羁绊5羁绊框颜色", "trait4 tier_current": "羁绊5当前等级", "trait4 tier_total": "羁绊5最高等级", "trait4 display_name": "羁绊5显示名", "trait4 icon_path": "羁绊5图标路径", "trait5 name": "羁绊6", "trait5 num_units": "羁绊6单位数", "trait5 style": "羁绊6羁绊框颜色", "trait5 tier_current": "羁绊6当前等级", "trait5 tier_total": "羁绊6最高等级", "trait5 display_name": "羁绊6显示名", "trait5 icon_path": "羁绊6图标路径", "trait6 name": "羁绊7", "trait6 num_units": "羁绊7单位数", "trait6 style": "羁绊7羁绊框颜色", "trait6 tier_current": "羁绊7当前等级", "trait6 tier_total": "羁绊7最高等级", "trait6 display_name": "羁绊7显示名", "trait6 icon_path": "羁绊7图标路径", "trait7 name": "羁绊8", "trait7 num_units": "羁绊8单位数", "trait7 style": "羁绊8羁绊框颜色", "trait7 tier_current": "羁绊8当前等级", "trait7 tier_total": "羁绊8最高等级", "trait7 display_name": "羁绊8显示名", "trait7 icon_path": "羁绊8图标路径", "trait8 name": "羁绊9", "trait8 num_units": "羁绊9单位数", "trait8 style": "羁绊9羁绊框颜色", "trait8 tier_current": "羁绊9当前等级", "trait8 tier_total": "羁绊9最高等级", "trait8 display_name": "羁绊9显示名", "trait8 icon_path": "羁绊9图标路径", "trait9 name": "羁绊10", "trait9 num_units": "羁绊10单位数", "trait9 style": "羁绊10羁绊框颜色", "trait9 tier_current": "羁绊10当前等级", "trait9 tier_total": "羁绊10最高等级", "trait9 display_name": "羁绊10显示名", "trait9 icon_path": "羁绊10图标路径", "trait10 name": "羁绊11", "trait10 num_units": "羁绊11单位数", "trait10 style": "羁绊11羁绊框颜色", "trait10 tier_current": "羁绊11当前等级", "trait10 tier_total": "羁绊11最高等级", "trait10 display_name": "羁绊11显示名", "trait10 icon_path": "羁绊11图标路径", "trait11 name": "羁绊12", "trait11 num_units": "羁绊12单位数", "trait11 style": "羁绊12羁绊框颜色", "trait11 tier_current": "羁绊12当前等级", "trait11 tier_total": "羁绊12最高等级", "trait11 display_name": "羁绊12显示名", "trait11 icon_path": "羁绊12图标路径", "trait12 name": "羁绊13", "trait12 num_units": "羁绊13单位数", "trait12 style": "羁绊13羁绊框颜色", "trait12 tier_current": "羁绊13当前等级", "trait12 tier_total": "羁绊13最高等级", "trait12 display_name": "羁绊13显示名", "trait12 icon_path": "羁绊13图标路径", "unit0 character_id": "英雄1：角色编号", "unit0 rarity": "英雄1：卡费", "unit0 tier": "英雄1：星级", "unit0 display_name": "英雄1：显示名", "unit0 squareIconPath": "英雄1：方块图标路径", "unit1 character_id": "英雄2：角色编号", "unit1 rarity": "英雄2：卡费", "unit1 tier": "英雄2：星级", "unit1 display_name": "英雄2：显示名", "unit1 squareIconPath": "英雄2：方块图标路径", "unit2 character_id": "英雄3：角色编号", "unit2 rarity": "英雄3：卡费", "unit2 tier": "英雄3：星级", "unit2 display_name": "英雄3：显示名", "unit2 squareIconPath": "英雄3：方块图标路径", "unit3 character_id": "英雄4：角色编号", "unit3 rarity": "英雄4：卡费", "unit3 tier": "英雄4：星级", "unit3 display_name": "英雄4：显示名", "unit3 squareIconPath": "英雄4：方块图标路径", "unit4 character_id": "英雄5：角色编号", "unit4 rarity": "英雄5：卡费", "unit4 tier": "英雄5：星级", "unit4 display_name": "英雄5：显示名", "unit4 squareIconPath": "英雄5：方块图标路径", "unit5 character_id": "英雄6：角色编号", "unit5 rarity": "英雄6：卡费", "unit5 tier": "英雄6：星级", "unit5 display_name": "英雄6：显示名", "unit5 squareIconPath": "英雄6：方块图标路径", "unit6 character_id": "英雄7：角色编号", "unit6 rarity": "英雄7：卡费", "unit6 tier": "英雄7：星级", "unit6 display_name": "英雄7：显示名", "unit6 squareIconPath": "英雄7：方块图标路径", "unit7 character_id": "英雄8：角色编号", "unit7 rarity": "英雄8：卡费", "unit7 tier": "英雄8：星级", "unit7 display_name": "英雄8：显示名", "unit7 squareIconPath": "英雄8：方块图标路径", "unit8 character_id": "英雄9：角色编号", "unit8 rarity": "英雄9：卡费", "unit8 tier": "英雄9：星级", "unit8 display_name": "英雄9：显示名", "unit8 squareIconPath": "英雄9：方块图标路径", "unit9 character_id": "英雄10：角色编号", "unit9 rarity": "英雄10：卡费", "unit9 tier": "英雄10：星级", "unit9 display_name": "英雄10：显示名", "unit9 squareIconPath": "英雄10：方块图标路径", "unit10 character_id": "英雄11：角色编号", "unit10 rarity": "英雄11：卡费", "unit10 tier": "英雄11：星级", "unit10 display_name": "英雄11：显示名", "unit10 squareIconPath": "英雄11：方块图标路径", "unit0 item0 nameId": "英雄1：装备1序号", "unit0 item0 name": "英雄1：装备1名称", "unit0 item0 squareIconPath": "英雄1：装备1方块图像路径", "unit0 item1 nameId": "英雄1：装备2序号", "unit0 item1 name": "英雄1：装备2名称", "unit0 item1 squareIconPath": "英雄1：装备2方块图像路径", "unit0 item2 nameId": "英雄1：装备3序号", "unit0 item2 name": "英雄1：装备3名称", "unit0 item2 squareIconPath": "英雄1：装备3方块图像路径", "unit1 item0 nameId": "英雄2：装备1序号", "unit1 item0 name": "英雄2：装备1名称", "unit1 item0 squareIconPath": "英雄2：装备1方块图像路径", "unit1 item1 nameId": "英雄2：装备2序号", "unit1 item1 name": "英雄2：装备2名称", "unit1 item1 squareIconPath": "英雄2：装备2方块图像路径", "unit1 item2 nameId": "英雄2：装备3序号", "unit1 item2 name": "英雄2：装备3名称", "unit1 item2 squareIconPath": "英雄2：装备3方块图像路径", "unit2 item0 nameId": "英雄3：装备1序号", "unit2 item0 name": "英雄3：装备1名称", "unit2 item0 squareIconPath": "英雄3：装备1方块图像路径", "unit2 item1 nameId": "英雄3：装备2序号", "unit2 item1 name": "英雄3：装备2名称", "unit2 item1 squareIconPath": "英雄3：装备2方块图像路径", "unit2 item2 nameId": "英雄3：装备3序号", "unit2 item2 name": "英雄3：装备3名称", "unit2 item2 squareIconPath": "英雄3：装备3方块图像路径", "unit3 item0 nameId": "英雄4：装备1序号", "unit3 item0 name": "英雄4：装备1名称", "unit3 item0 squareIconPath": "英雄4：装备1方块图像路径", "unit3 item1 nameId": "英雄4：装备2序号", "unit3 item1 name": "英雄4：装备2名称", "unit3 item1 squareIconPath": "英雄4：装备2方块图像路径", "unit3 item2 nameId": "英雄4：装备3序号", "unit3 item2 name": "英雄4：装备3名称", "unit3 item2 squareIconPath": "英雄4：装备3方块图像路径", "unit4 item0 nameId": "英雄5：装备1序号", "unit4 item0 name": "英雄5：装备1名称", "unit4 item0 squareIconPath": "英雄5：装备1方块图像路径", "unit4 item1 nameId": "英雄5：装备2序号", "unit4 item1 name": "英雄5：装备2名称", "unit4 item1 squareIconPath": "英雄5：装备2方块图像路径", "unit4 item2 nameId": "英雄5：装备3序号", "unit4 item2 name": "英雄5：装备3名称", "unit4 item2 squareIconPath": "英雄5：装备3方块图像路径", "unit5 item0 nameId": "英雄6：装备1序号", "unit5 item0 name": "英雄6：装备1名称", "unit5 item0 squareIconPath": "英雄6：装备1方块图像路径", "unit5 item1 nameId": "英雄6：装备2序号", "unit5 item1 name": "英雄6：装备2名称", "unit5 item1 squareIconPath": "英雄6：装备2方块图像路径", "unit5 item2 nameId": "英雄6：装备3序号", "unit5 item2 name": "英雄6：装备3名称", "unit5 item2 squareIconPath": "英雄6：装备3方块图像路径", "unit6 item0 nameId": "英雄7：装备1序号", "unit6 item0 name": "英雄7：装备1名称", "unit6 item0 squareIconPath": "英雄7：装备1方块图像路径", "unit6 item1 nameId": "英雄7：装备2序号", "unit6 item1 name": "英雄7：装备2名称", "unit6 item1 squareIconPath": "英雄7：装备2方块图像路径", "unit6 item2 nameId": "英雄7：装备3序号", "unit6 item2 name": "英雄7：装备3名称", "unit6 item2 squareIconPath": "英雄7：装备3方块图像路径", "unit7 item0 nameId": "英雄8：装备1序号", "unit7 item0 name": "英雄8：装备1名称", "unit7 item0 squareIconPath": "英雄8：装备1方块图像路径", "unit7 item1 nameId": "英雄8：装备2序号", "unit7 item1 name": "英雄8：装备2名称", "unit7 item1 squareIconPath": "英雄8：装备2方块图像路径", "unit7 item2 nameId": "英雄8：装备3序号", "unit7 item2 name": "英雄8：装备3名称", "unit7 item2 squareIconPath": "英雄8：装备3方块图像路径", "unit8 item0 nameId": "英雄9：装备1序号", "unit8 item0 name": "英雄9：装备1名称", "unit8 item0 squareIconPath": "英雄9：装备1方块图像路径", "unit8 item1 nameId": "英雄9：装备2序号", "unit8 item1 name": "英雄9：装备2名称", "unit8 item1 squareIconPath": "英雄9：装备2方块图像路径", "unit8 item2 nameId": "英雄9：装备3序号", "unit8 item2 name": "英雄9：装备3名称", "unit8 item2 squareIconPath": "英雄9：装备3方块图像路径", "unit9 item0 nameId": "英雄10：装备1序号", "unit9 item0 name": "英雄10：装备1名称", "unit9 item0 squareIconPath": "英雄10：装备1方块图像路径", "unit9 item1 nameId": "英雄10：装备2序号", "unit9 item1 name": "英雄10：装备2名称", "unit9 item1 squareIconPath": "英雄10：装备2方块图像路径", "unit9 item2 nameId": "英雄10：装备3序号", "unit9 item2 name": "英雄10：装备3名称", "unit9 item2 squareIconPath": "英雄10：装备3方块图像路径", "unit10 item0 nameId": "英雄11：装备1序号", "unit10 item0 name": "英雄11：装备1名称", "unit10 item0 squareIconPath": "英雄11：装备1方块图像路径", "unit10 item1 nameId": "英雄11：装备2序号", "unit10 item1 name": "英雄11：装备2名称", "unit10 item1 squareIconPath": "英雄11：装备2方块图像路径", "unit10 item2 nameId": "英雄11：装备3序号", "unit10 item2 name": "英雄11：装备3名称", "unit10 item2 squareIconPath": "英雄11：装备3方块图像路径"}
                    TFTHistory_header_keys = list(TFTHistory_header.keys())
                    recent_TFTPlayers_statistics_output_order = [37, 38, 36, 4, 11, 12, 6, 13, 7, 8, 28, 29, 30, 33, 41, 42, 31, 40, 35, 34, 18, 19, 20, 137, 135, 136, 190, 193, 196, 142, 140, 141, 199, 202, 205, 147, 145, 146, 208, 211, 214, 152, 150, 151, 217, 220, 223, 157, 155, 156, 226, 229, 232, 162, 160, 161, 235, 238, 241, 167, 165, 166, 244, 247, 250, 172, 170, 171, 253, 256, 259, 177, 175, 176, 262, 265, 268, 182, 180, 181, 271, 274, 277, 187, 185, 186, 280, 283, 286, 48, 44, 45, 46, 47, 55, 51, 52, 53, 54, 62, 58, 59, 60, 61, 69, 65, 66, 67, 68, 76, 72, 73, 74, 75, 83, 79, 80, 81, 82, 90, 86, 87, 88, 89, 97, 93, 94, 95, 96, 104, 100, 101, 102, 103, 111, 107, 108, 109, 110, 118, 114, 115, 116, 117, 125, 121, 122, 123, 124, 132, 128, 129, 130, 131]
                    recent_TFTPlayers_data_organized = {}
                    for i in range(len(recent_TFTPlayers_statistics_output_order)):
                        key = TFTHistory_header_keys[recent_TFTPlayers_statistics_output_order[i]]
                        recent_TFTPlayers_data_organized[key] = [TFTHistory_header[key]]
                    recent_TFTPlayers_df = pandas.DataFrame(data = recent_TFTPlayers_data_organized)
                
                if search_LoL and fetched_info or search_TFT and TFTGamePlayed:
                    if not detectMode:
                        recent_players_metadata = {} #这里另外设置元数据是为了整理出用于可视化的数据（Here the metadata is designed to sort out data for visualization）
                        if search_LoL:
                            for i in range(1, len(recent_LoLPlayers_df)): #第0行是中文表头，所以要从第1行开始（The 0th line contains the Chinese headers, so the iteration should start from the first line）
                                puuid_iter = recent_LoLPlayers_df.loc[i, "puuid"]
                                summonerName_iter = recent_LoLPlayers_df.loc[i, "gameName"] + "#" + recent_LoLPlayers_df.loc[i, "tagLine"]
                                matchID_iter = recent_LoLPlayers_df.loc[i, "gameId"]
                                LoLGameDuration_iter = LoLGameDuration_raw[i - 1] #由于列表变量LoLGameDuration_raw独立于recent_players_data之外单独存储信息，其中不包含中文表头，全是数据，因此在设置索引时应当减1，以对应其它与recent_LoLPlayers_df有关的列表变量（Since the list variable `LoLGameDuration_raw` stores data independently from the variable `recent_players_data`, it only contains data without a header. Therefore, its index is subtracted by 1 to correspond to other list variables with regard to `recent_LoLPlayers_df`）
                                isPvP_iter = True if recent_LoLPlayers_df["queueId"][i] in gamemodes and gamemodes[recent_LoLPlayers_df["queueId"][i]]["category"] == "PvP" else False #添加是否玩家对战的信息，以便单独统计一同进行玩家对战的总时间。下同（Added the information whether a match is PvP, so that the total time of only PvP matches can be calculated. So do the following two variables）
                                isPvE_iter = True if recent_LoLPlayers_df["queueId"][i] in gamemodes and gamemodes[recent_LoLPlayers_df["queueId"][i]]["category"] == "VersusAi" else False
                                isCustom_iter = True if recent_LoLPlayers_df["queueId"][i] in gamemodes and gamemodes[recent_LoLPlayers_df["queueId"][i]]["category"] == "CUSTOM" else False
                                if not puuid_iter in recent_players_metadata:
                                    recent_players_metadata[puuid_iter] = {}
                                    recent_players_metadata[puuid_iter]["name"] = summonerName_iter #该语句不会在else部分出现。这是考虑到如果召唤师改过名字，那么呈现在频数直方图上的横轴的召唤师名应当是最新的（This statement won't appear in the else-part, considering if a summoner has changed its name, then the summonerName near the horizontal axis of the frequency histogram should be latest）
                                    recent_players_metadata[puuid_iter]["puuid"] = puuid_iter
                                    recent_players_metadata[puuid_iter]["gameCount"] = 1
                                    recent_players_metadata[puuid_iter]["matches"] = [matchID_iter]
                                    recent_players_metadata[puuid_iter]["durations"] = [LoLGameDuration_iter]
                                    recent_players_metadata[puuid_iter]["isPvP"] = [isPvP_iter]
                                    recent_players_metadata[puuid_iter]["isPvE"] = [isPvE_iter]
                                    recent_players_metadata[puuid_iter]["isCustom"] = [isCustom_iter]
                                    recent_players_metadata[puuid_iter]["PvPCount"] = int(isPvP_iter)
                                    recent_players_metadata[puuid_iter]["PvECount"] = int(isPvE_iter)
                                    recent_players_metadata[puuid_iter]["CustomCount"] = int(isCustom_iter)
                                    recent_players_metadata[puuid_iter]["totalTime"] = LoLGameDuration_iter
                                    recent_players_metadata[puuid_iter]["totalPvPTime"] = LoLGameDuration_iter * isPvP_iter
                                    recent_players_metadata[puuid_iter]["totalPvETime"] = LoLGameDuration_iter * isPvE_iter
                                    recent_players_metadata[puuid_iter]["totalCustomTime"] = LoLGameDuration_iter * isCustom_iter
                                else:
                                    recent_players_metadata[puuid_iter]["gameCount"] += 1
                                    recent_players_metadata[puuid_iter]["matches"].append(matchID_iter)
                                    recent_players_metadata[puuid_iter]["durations"].append(LoLGameDuration_iter)
                                    recent_players_metadata[puuid_iter]["isPvP"].append(isPvP_iter)
                                    recent_players_metadata[puuid_iter]["isPvE"].append(isPvE_iter)
                                    recent_players_metadata[puuid_iter]["isCustom"].append(isCustom_iter)
                                    recent_players_metadata[puuid_iter]["PvPCount"] += isPvP_iter
                                    recent_players_metadata[puuid_iter]["PvECount"] += isPvE_iter
                                    recent_players_metadata[puuid_iter]["CustomCount"] += isCustom_iter
                                    recent_players_metadata[puuid_iter]["totalTime"] += LoLGameDuration_iter
                                    recent_players_metadata[puuid_iter]["totalPvPTime"] += LoLGameDuration_iter * isPvP_iter
                                    recent_players_metadata[puuid_iter]["totalPvETime"] += LoLGameDuration_iter * isPvE_iter
                                    recent_players_metadata[puuid_iter]["totalCustomTime"] += LoLGameDuration_iter * isCustom_iter
                                #logPrint("用于可视化的元数据创建进度（Creating process of metadata for visualization）：%d/%d" %(i, len(recent_LoLPlayers_df) - 1), end = "\r")
                        if search_TFT:
                            for i in range(1, len(recent_TFTPlayers_df)):
                                puuid_iter = recent_TFTPlayers_df.loc[i, "puuid"]
                                summonerName_iter = recent_TFTPlayers_df.loc[i, "riotIdGameName"] + "#" + recent_TFTPlayers_df.loc[i, "riotIdTagLine"]
                                matchID_iter = recent_TFTPlayers_df.loc[i, "game_id"]
                                TFTGameDuration_iter = TFTGameDuration_raw[i - 1]
                                isPvP_iter = True if recent_TFTPlayers_df["queue_id"][i] in gamemodes and gamemodes[recent_TFTPlayers_df["queue_id"][i]]["category"] == "PvP" else False
                                isPvE_iter = True if recent_TFTPlayers_df["queue_id"][i] in gamemodes and gamemodes[recent_TFTPlayers_df["queue_id"][i]]["category"] == "VersusAi" else False
                                isCustom_iter = True if recent_TFTPlayers_df["queue_id"][i] in gamemodes and gamemodes[recent_TFTPlayers_df["queue_id"][i]]["category"] == "CUSTOM" else False
                                if not puuid_iter in recent_players_metadata:
                                    recent_players_metadata[puuid_iter] = {}
                                    recent_players_metadata[puuid_iter]["name"] = summonerName_iter
                                    recent_players_metadata[puuid_iter]["puuid"] = puuid_iter
                                    recent_players_metadata[puuid_iter]["gameCount"] = 1
                                    recent_players_metadata[puuid_iter]["matches"] = [matchID_iter]
                                    recent_players_metadata[puuid_iter]["durations"] = [TFTGameDuration_iter]
                                    recent_players_metadata[puuid_iter]["isPvP"] = [isPvP_iter]
                                    recent_players_metadata[puuid_iter]["isPvE"] = [isPvE_iter]
                                    recent_players_metadata[puuid_iter]["isCustom"] = [isCustom_iter]
                                    recent_players_metadata[puuid_iter]["PvPCount"] = int(isPvP_iter)
                                    recent_players_metadata[puuid_iter]["PvECount"] = int(isPvE_iter)
                                    recent_players_metadata[puuid_iter]["CustomCount"] = int(isCustom_iter)
                                    recent_players_metadata[puuid_iter]["totalTime"] = TFTGameDuration_iter
                                    recent_players_metadata[puuid_iter]["totalPvPTime"] = TFTGameDuration_iter * isPvP_iter
                                    recent_players_metadata[puuid_iter]["totalPvETime"] = TFTGameDuration_iter * isPvE_iter
                                    recent_players_metadata[puuid_iter]["totalCustomTime"] = TFTGameDuration_iter * isCustom_iter
                                else:
                                    recent_players_metadata[puuid_iter]["gameCount"] += 1
                                    recent_players_metadata[puuid_iter]["matches"].append(matchID_iter)
                                    recent_players_metadata[puuid_iter]["durations"].append(TFTGameDuration_iter)
                                    recent_players_metadata[puuid_iter]["isPvP"].append(isPvP_iter)
                                    recent_players_metadata[puuid_iter]["isPvE"].append(isPvE_iter)
                                    recent_players_metadata[puuid_iter]["isCustom"].append(isCustom_iter)
                                    recent_players_metadata[puuid_iter]["PvPCount"] += isPvP_iter
                                    recent_players_metadata[puuid_iter]["PvECount"] += isPvE_iter
                                    recent_players_metadata[puuid_iter]["CustomCount"] += isCustom_iter
                                    recent_players_metadata[puuid_iter]["totalTime"] += TFTGameDuration_iter
                                    recent_players_metadata[puuid_iter]["totalPvPTime"] += TFTGameDuration_iter * isPvP_iter
                                    recent_players_metadata[puuid_iter]["totalPvETime"] += TFTGameDuration_iter * isPvE_iter
                                    recent_players_metadata[puuid_iter]["totalCustomTime"] += TFTGameDuration_iter * isCustom_iter
                                #logPrint("用于可视化的元数据创建进度（Creating process of metadata for visualization）：%d/%d" %(i, len(recent_TFTPlayers_df) - 1), end = "\r")
                        #pyperclip.copy(str(recent_players_metadata))
                        jsonname = "Recently Played Summoners - %s.json" %displayName
                        while True:
                            try:
                                jsonfile = open(os.path.join(folder, jsonname), "w", encoding = "utf-8")
                            except FileNotFoundError:
                                os.makedirs(folder, exist_ok = True)
                            else:
                                break
                        try:
                            jsonfile.write(str(json.dumps(recent_players_metadata, indent = 4, ensure_ascii = False)))
                        except UnicodeEncodeError:
                            logPrint("近期一起玩过的玩家元数据文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nRecently played summoner metadata text generation failure! Please check if the summoner name includes any abnormal characters!\n")
                        recent_players_metadata_list = sorted(recent_players_metadata.values(), key = lambda x: x["gameCount"], reverse = True)
                        recent_players_metadata_header = {"name": "召唤师名", "puuid": "玩家通用唯一识别码", "gameCount": "共同作战局数", "matches": "共同对局序号", "durations": "对局持续时间列表", "isPvP": "玩家对战逻辑值列表", "isPvE": "人机对战逻辑值列表", "isCustom": "自定义对战逻辑值列表", "PvPCount": "玩家对战局数", "PvECount": "人机对战局数", "CustomCount": "自定义对战局数", "totalTime": "共同作战时长（秒）", "totalPvPTime": "共同玩家对战时长（秒）", "totalPvETime": "共同人机对战时长（秒）", "totalCustomTime": "共同自定义对战时长（秒）"}
                        recent_players_metadata_header_keys = list(recent_players_metadata_header.keys())
                        recent_players_metadata_statistics_output_order = [0, 1, 3, 2, 8, 9, 10, 11, 12, 13, 14]
                        recent_players_metadata_organized = {}
                        for i in recent_players_metadata_statistics_output_order:
                            key = recent_players_metadata_header_keys[i]
                            recent_players_metadata_organized[key] = list(map(lambda x: x[key], recent_players_metadata_list))
                        recent_players_metaDf = pandas.concat([pandas.DataFrame(data = recent_players_metadata_organized)])
                        recent_players_metaDf = pandas.concat([pandas.DataFrame([recent_players_metadata_header])[recent_players_metaDf.columns], recent_players_metaDf], ignore_index = True)
                        #默认导出玩家对局数量统计表（Export recent played summoner count table by default）
                        while True:
                            try:
                                with pandas.ExcelWriter(path = os.path.join(folder, f"Recently Played Summoner Count - {displayName}.xlsx")) as writer:
                                    recent_players_metaDf.to_excel(excel_writer = writer)
                            except PermissionError:
                                logPrint("近期一起玩过的玩家对局数量统计表导出失败！请检查文件的权限以及是否被占用！按回车键重试，或者输入任意非空字符串以放弃导出。\nRecently played summoner count table export failure! Please check the permission and if the file is occupied! Press Enter to try again, or submit any non-empty string to give up exporting.")
                                gameCount_export_str = logInput()
                                gameCount_export = not bool(gameCount_export_str)
                                if not gameCount_export:
                                    break
                            else:
                                break
                        
                        #针对元数据中记录的每个玩家的累计游戏时长和游戏对局数输出条形图（Output the bar chart of each summoner's total time and game counts in the metadata）
                        totalTime = {}
                        PvPTime = {}
                        PvETime = {}
                        CustomTime = {}
                        totalCount = {}
                        PvPCount = {}
                        PvECount = {}
                        CustomCount = {}
                        for player in recent_players_metadata.values():
                            totalTime[player["name"]] = player["totalTime"]
                            PvPTime[player["name"]] = player["totalPvPTime"]
                            PvETime[player["name"]] = player["totalPvETime"]
                            CustomTime[player["name"]] = player["totalCustomTime"]
                            totalCount[player["name"]] = player["gameCount"]
                            PvPCount[player["name"]] = player["PvPCount"]
                            PvECount[player["name"]] = player["PvECount"]
                            CustomCount[player["name"]] = player["CustomCount"]
                        totalTime_sorted = sorted(totalTime.items(), key = lambda x: x[1], reverse = True)
                        PvPTime_sorted = sorted(PvPTime.items(), key = lambda x: x[1], reverse = True)
                        PvETime_sorted = sorted(PvETime.items(), key = lambda x: x[1], reverse = True)
                        CustomTime_sorted = sorted(CustomTime.items(), key = lambda x: x[1], reverse = True)
                        totalCount_sorted = sorted(totalCount.items(), key = lambda x: x[1], reverse = True)
                        PvPCount_sorted = sorted(PvPCount.items(), key = lambda x: x[1], reverse = True)
                        PvECount_sorted = sorted(PvECount.items(), key = lambda x: x[1], reverse = True)
                        CustomCount_sorted = sorted(CustomCount.items(), key = lambda x: x[1], reverse = True)
                        logPrint("您希望条形图中显示游戏时长最长的前几名玩家？（默认为前20名）\nHow many players of the longest game time do you want to display in the bar chart? (20 by default)")
                        while True:
                            try:
                                topN = logInput()
                                if topN == "":
                                    topN = 20
                                    break
                                else:
                                    topN = int(topN)
                            except ValueError:
                                logPrint("请输入整数！\nPlease input an integer!")
                            else:
                                if topN <= 0:
                                    logPrint("请输入正整数！\nPlease input a positive integer!")
                                else:
                                    break
                        topN = min(topN, len(recent_LoLPlayers_df) + len(recent_TFTPlayers_df) - 2)
                        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei'] #设置默认字体为微软雅黑（Set the default font Microsoft YaHei）
                        plt.figure(figsize = (topN / 2, 12)) #设置导出图象的大小（Set the size of the exported figure）
                        valuefont = {"family": "Times New Roman", "weight": "normal", "size": 9} #指定柱上显示的数据的字体格式（Determines the font of the values above the bars）
                        #绘制各玩家的总游戏时间柱状图（Plot the bar chart of the total time of certain number of players）
                        totalTimePic = plt.bar([totalTime_sorted[i][0] for i in range(topN)], [totalTime_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("游戏时长（秒）\ntotalGameTime (s)")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in totalTime_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("总游戏时间\ntotal game time")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Time_total.png" %displayName))
                        plt.clf()
                        #绘制各玩家的玩家对战时间柱状图（Plot the bar chart of the PvP time of certain number of players）
                        PvPTimePic = plt.bar([PvPTime_sorted[i][0] for i in range(topN)], [PvPTime_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("游戏时长（秒）\ntotalGameTime (s)")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in PvPTime_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("玩家对战时间\nPvP game time")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Time_PvP.png" %displayName))
                        plt.clf()
                        #绘制各玩家的人机对战时间柱状图（Plot the bar chart of the PvE time of certain number of players）
                        PvETimePic = plt.bar([PvETime_sorted[i][0] for i in range(topN)], [PvETime_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("游戏时长（秒）\ntotalGameTime (s)")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in PvETime_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("人机对战时间\nPvE game time")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Time_PvE.png" %displayName))
                        plt.clf()
                        #绘制各玩家的自定义对战时间柱状图（Plot the bar chart of the Custom time of certain number of players）
                        CustomTimePic = plt.bar([CustomTime_sorted[i][0] for i in range(topN)], [CustomTime_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("游戏时长（秒）\ntotalGameTime (s)")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in CustomTime_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("自定义对战时间\nCustom game time")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Time_Custom.png" %displayName))
                        plt.clf()
                        #绘制各玩家的总游戏对局数柱状图（Plot the bar chart of the total game count of certain number of players）
                        totalCountPic = plt.bar([totalCount_sorted[i][0] for i in range(topN)], [totalCount_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("对局数\ntotalGameCount")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in totalCount_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("总游戏对局数\ntotal game count")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Count_total.png" %displayName))
                        plt.clf()
                        #绘制各玩家的玩家对战局数柱状图（Plot the bar chart of the PvP game count of certain number of players）
                        PvPCountPic = plt.bar([PvPCount_sorted[i][0] for i in range(topN)], [PvPCount_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("对局数\ntotalGameCount")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in PvPCount_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("玩家对战局数\nPvP game count")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Count_PvP.png" %displayName))
                        plt.clf()
                        #绘制各玩家的人机对战局数柱状图（Plot the bar chart of the PvE game count of certain number of players）
                        PvECountPic = plt.bar([PvECount_sorted[i][0] for i in range(topN)], [PvECount_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("对局数\ntotalGameCount")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in PvECount_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("人机对战局数\nPvE game count")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Count_PvE.png" %displayName))
                        plt.clf()
                        #绘制各玩家的自定义对战局数柱状图（Plot the bar chart of the Custom game count of certain number of players）
                        CustomCountPic = plt.bar([CustomCount_sorted[i][0] for i in range(topN)], [CustomCount_sorted[i][1] for i in range(topN)])
                        plt.xticks(rotation = 45, ha = "right")
                        plt.ylabel("对局数\ntotalGameCount")
                        plt.yticks(fontproperties = "Calibri", size = 12)
                        for player, playtime in CustomCount_sorted[:topN]:
                            plt.text(player, playtime, playtime, ha = "center", va = "bottom", fontdict = valuefont)
                        plt.title("自定义对战局数\nCustom game count")
                        plt.savefig(os.path.join(folder, "Recently Played Summoners - %s - Count_Custom.png" %displayName))
                        plt.clf()
                        
                        #定义条件格式（Define the conditional formats）
                        twoDigitPercentage_columns_lol = [column for column in recent_LoLPlayers_df.columns if column.endswith("_percent") or column == "GUE"] #百分比（Percentage）
                        oneDigitFloat_columns_lol = ["KDA"] #一位小数（One-digit float）
                        threeDigitFloat_columns_lol = ["CSPM", "D/G", "GPM"] #三位小数（Three-digit float）
                        colorScale_columns_lol = [column for column in recent_LoLPlayers_df.columns if column.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
                        dataBar_columns_lol = [column for column in recent_LoLPlayers_df.columns if column.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
                        max_numPlayersPerTeam_lol = 5 if len(recent_LoLPlayers_df) <= 1 else max(map(lambda x: 5 if x == 0 else 2 if queues[x]["gameMode"] == "CHERRY" else queues[x]["numPlayersPerTeam"], recent_LoLPlayers_df.loc[1:, "queueId"])) #自定义对局的队伍规模视为5；斗魂竞技场的队伍规模虽然在API中记录为16，但这里应该考虑的是子阵营（The team size of any custom game is regarded as 5; although the team size of an Arena game is recorded as in LCU API, the subteam has more reference value）
                        order_colorScaleRule_lol = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = max_numPlayersPerTeam_lol, end_color = "FF6B6B") #跳过名次为0的单元格（Skip the order cells whose values are 0）
                        percent_dataBarRule_lol = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
                        
                        logPrint("是否导出以上近期一起玩过的玩家数据？（输入任意键导出，否则不导出）\nDo you want to export the above recently played summoner data? (Input anything to export or null to refuse exporting)")
                        export_str = logInput()
                        export = bool(export_str)
                        if export:
                            excel_name = "Summoner Profile - " + displayName + ".xlsx"
                            while True:
                                try:
                                    with pandas.ExcelWriter(path = os.path.join(folder, excel_name), engine = "openpyxl", mode = "a", if_sheet_exists = "replace") as writer:
                                        if search_LoL:
                                            recent_LoLPlayers_df.to_excel(excel_writer = writer, sheet_name = "Recently Played Summoners (LoL)")
                                            worksheet = writer.sheets["Recently Played Summoners (LoL)"]
                                            worksheet.conditional_formatting.rules = [] #读取时清空原规则（Clear original rules when reading）
                                            if len(recent_LoLPlayers_df) > 1:
                                                #套用保留两位小数的百分比格式（Two-digit percentage）
                                                for column in twoDigitPercentage_columns_lol:
                                                    col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    for row in range(3, len(recent_LoLPlayers_df) + 2):
                                                        worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
                                                #套用一位小数（One-digit float）
                                                for column in oneDigitFloat_columns_lol:
                                                    col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    for row in range(3, len(recent_LoLPlayers_df) + 2):
                                                        worksheet.cell(row = row, column = col_idx).number_format = "0.0"
                                                #套用三位小数（Three-digit float）
                                                for column in threeDigitFloat_columns_lol:
                                                    col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    for row in range(3, len(recent_LoLPlayers_df) + 2):
                                                        worksheet.cell(row = row, column = col_idx).number_format = "0.000"
                                                #胜负颜色（Win/Lose color）
                                                col_idx = recent_LoLPlayers_df.columns.get_loc("win/lose") + 2
                                                col_letter = get_column_letter(col_idx)
                                                rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(recent_LoLPlayers_df) + 1)
                                                win_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
                                                lose_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
                                                worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
                                                worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
                                                #百分比颜色（Percent color）
                                                rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
                                                for i in range(len(dataBar_columns_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
                                                    column = dataBar_columns_lol[i]
                                                    if i == 0:
                                                        startCol_idx = endCol_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    else:
                                                        col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                        if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                                                            endCol_idx = col_idx
                                                        else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                                                            startCol_letter = get_column_letter(startCol_idx)
                                                            endCol_letter = get_column_letter(endCol_idx)
                                                            rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                            rangeStrs.append(rangeStr)
                                                            startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
                                                else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
                                                    startCol_letter = get_column_letter(startCol_idx)
                                                    endCol_letter = get_column_letter(endCol_idx)
                                                    rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                    rangeStrs.append(rangeStr)
                                                for rangeStr in rangeStrs:
                                                    worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
                                                #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
                                                col_idx = recent_LoLPlayers_df.columns.get_loc("subteamPlacement") + 2
                                                col_letter = get_column_letter(col_idx)
                                                rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(recent_LoLPlayers_df) + 1)
                                                firstPlace_formulaRule_lol = FormulaRule(formula = ['$%s3=1' %(col_letter)], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
                                                worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)
                                                #位次颜色（Order color）
                                                rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
                                                rangeTuples = []
                                                for i in range(len(colorScale_columns_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
                                                    column = colorScale_columns_lol[i]
                                                    if i == 0:
                                                        startCol_idx = endCol_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    else:
                                                        col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                        if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                                                            endCol_idx = col_idx
                                                        else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                                                            startCol_letter = get_column_letter(startCol_idx)
                                                            endCol_letter = get_column_letter(endCol_idx)
                                                            rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                            rangeStrs.append(rangeStr)
                                                            rangeTuples.append((startCol_letter, endCol_letter))
                                                            startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
                                                else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
                                                    startCol_letter = get_column_letter(startCol_idx)
                                                    endCol_letter = get_column_letter(endCol_idx)
                                                    rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                    rangeStrs.append(rangeStr)
                                                    rangeTuples.append((startCol_letter, endCol_letter))
                                                for i in range(len(rangeStrs)):
                                                    rangeStr = rangeStrs[i]
                                                    rangeTuple = rangeTuples[i]
                                                    order_noFillRule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
                                                    worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
                                                    worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)
                                            logPrint("近期一起玩过的英雄联盟玩家数据导出完成！\nRecently played summoner data (LoL) exported!\n")
                                        if search_TFT:
                                            recent_TFTPlayers_df.to_excel(excel_writer = writer, sheet_name = "Recently Played Summoners (TFT)")
                                            logPrint("近期一起玩过的云顶之弈玩家数据导出完成！\nRecently played summoner data (TFT) exported!\n")
                                except PermissionError:
                                    logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                    logInput()
                                except FileNotFoundError:
                                    try:
                                        os.makedirs(folder)
                                    except FileExistsError:
                                        pass
                                    with pandas.ExcelWriter(path = os.path.join(folder, excel_name), engine = "openpyxl") as writer:
                                        if search_LoL:
                                            recent_LoLPlayers_df.to_excel(excel_writer = writer, sheet_name = "Recently Played Summoners (LoL)")
                                            worksheet = writer.sheets["Recently Played Summoners (LoL)"]
                                            worksheet.conditional_formatting.rules = [] #读取时清空原规则（Clear original rules when reading）
                                            if len(recent_LoLPlayers_df) > 1:
                                                #套用保留两位小数的百分比格式（Two-digit percentage）
                                                for column in twoDigitPercentage_columns_lol:
                                                    col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    for row in range(3, len(recent_LoLPlayers_df) + 2):
                                                        worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
                                                #套用一位小数（One-digit float）
                                                for column in oneDigitFloat_columns_lol:
                                                    col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    for row in range(3, len(recent_LoLPlayers_df) + 2):
                                                        worksheet.cell(row = row, column = col_idx).number_format = "0.0"
                                                #套用三位小数（Three-digit float）
                                                for column in threeDigitFloat_columns_lol:
                                                    col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    for row in range(3, len(recent_LoLPlayers_df) + 2):
                                                        worksheet.cell(row = row, column = col_idx).number_format = "0.000"
                                                #胜负颜色（Win/Lose color）
                                                col_idx = recent_LoLPlayers_df.columns.get_loc("win/lose") + 2
                                                col_letter = get_column_letter(col_idx)
                                                rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(recent_LoLPlayers_df) + 1)
                                                win_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
                                                lose_formulaRule_lol = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
                                                worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
                                                worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
                                                #百分比颜色（Percent color）
                                                rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
                                                for i in range(len(dataBar_columns_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
                                                    column = dataBar_columns_lol[i]
                                                    if i == 0:
                                                        startCol_idx = endCol_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    else:
                                                        col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                        if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                                                            endCol_idx = col_idx
                                                        else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                                                            startCol_letter = get_column_letter(startCol_idx)
                                                            endCol_letter = get_column_letter(endCol_idx)
                                                            rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                            rangeStrs.append(rangeStr)
                                                            startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
                                                else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
                                                    startCol_letter = get_column_letter(startCol_idx)
                                                    endCol_letter = get_column_letter(endCol_idx)
                                                    rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                    rangeStrs.append(rangeStr)
                                                for rangeStr in rangeStrs:
                                                    worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
                                                #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
                                                col_idx = recent_LoLPlayers_df.columns.get_loc("subteamPlacement") + 2
                                                col_letter = get_column_letter(col_idx)
                                                rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(recent_LoLPlayers_df) + 1)
                                                firstPlace_formulaRule_lol = FormulaRule(formula = ['$%s3=1' %(col_letter)], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
                                                worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)
                                                #位次颜色（Order color）
                                                rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
                                                rangeTuples = []
                                                for i in range(len(colorScale_columns_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
                                                    column = colorScale_columns_lol[i]
                                                    if i == 0:
                                                        startCol_idx = endCol_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                    else:
                                                        col_idx = recent_LoLPlayers_df.columns.get_loc(column) + 2
                                                        if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                                                            endCol_idx = col_idx
                                                        else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                                                            startCol_letter = get_column_letter(startCol_idx)
                                                            endCol_letter = get_column_letter(endCol_idx)
                                                            rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                            rangeStrs.append(rangeStr)
                                                            rangeTuples.append((startCol_letter, endCol_letter))
                                                            startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
                                                else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
                                                    startCol_letter = get_column_letter(startCol_idx)
                                                    endCol_letter = get_column_letter(endCol_idx)
                                                    rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(recent_LoLPlayers_df) + 1)
                                                    rangeStrs.append(rangeStr)
                                                    rangeTuples.append((startCol_letter, endCol_letter))
                                                for i in range(len(rangeStrs)):
                                                    rangeStr = rangeStrs[i]
                                                    rangeTuple = rangeTuples[i]
                                                    order_noFillRule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
                                                    worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
                                                    worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)
                                            logPrint("近期一起玩过的英雄联盟玩家数据导出完成！\nRecently played summoner data (LoL) exported!\n")
                                        if search_TFT:
                                            recent_TFTPlayers_df.to_excel(excel_writer = writer, sheet_name = "Recently Played Summoners (TFT)")
                                            logPrint("近期一起玩过的云顶之弈玩家数据导出完成！\nRecently played summoner data (TFT) exported!\n")
                                    break
                                else:
                                    break
                    else:
                        logPrint("近期一起玩过的玩家数据已加载完成！\nRecently played summoner data loaded successfully!")
                        update = False
                        while True:
                            recent_LoLPlayer_fields = ["gameName", "tagLine", "gameCreationDate", "gameModeName", "champion_name", "K/D/A"]
                            recent_TFTPlayer_fields = ["riotIdGameName", "riotIdTagLine", "gameDate", "tft_game_type", "last_round_format"]
                            recent_LoLPlayer_dict_to_print = {}
                            recent_TFTPlayer_dict_to_print = {}
                            for key in recent_LoLPlayer_fields:
                                recent_LoLPlayer_dict_to_print[key] = []
                            for key in recent_TFTPlayer_fields:
                                recent_TFTPlayer_dict_to_print[key] = []
                            logPrint("请选择检测场景：\nPlease select the situation to detect:\n1\t房间内/英雄选择阶段/游戏中（默认）【In-lobby/During champ select/In-game (Default)】\n2\t好友列表（Friend list）\n3\t好友请求（Friend requests）\n4\t组队邀请（Party invitations）\n5\t聊天黑名单（Block list）\n6\t自定义召唤师名称列表（List of any summoners' names）")
                            detect_scene = logInput()
                            if detect_scene == "":
                                detect_scene = "1"
                            elif detect_scene[0] == "0":
                                break
                            elif detect_scene[0] in set(map(str, range(1, 7))):
                                detect_scene = detect_scene[0]
                            else:
                                detect_scene = "6"
                            if detect_scene == "1":
                                member_count = 0
                                ally_count = 0
                                enemy_count = 0
                                player_count = 0
                                recent_friends = []
                                LoLMember_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                TFTMember_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                LoLAlly_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                LoLEnemy_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print) #在玩家对战的英雄选择阶段，所有敌方玩家的信息都是不可见的；在人机对战的英雄选择阶段，无敌方玩家。统计敌方信息只适用于自定义对局的英雄选择阶段和任意对局的游戏内（During champ select of PVP games, all enemies' information is hidden; during champ select of PVE games, there're no enemy players. Counting enemy stats only applys in the champ select stage of custom games and the in-game stage of any game）
                                LoLPlayer_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                TFTAlly_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                TFTEnemy_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                TFTPlayer_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                recent_LoLPlayer_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                recent_TFTPlayer_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                logPrint('''请确保您在房间内、英雄选择阶段或在游戏中，以便本脚本检测是否存在曾经遇到过的队友。按回车键开始检测，或者按“0”以返回上一步。\nPlease confirm you're in lobby, during champ select or in game, so that this script can detect whether there's an ally encountered before. Press Enter to start detection, or press "0" to return to the last step.''')
                                while True:
                                    detect = logInput()
                                    if detect != "" and detect[0] == "0":
                                        break
                                    gameflow_phase = await (await connection.request("GET", "/lol-gameflow/v1/gameflow-phase")).json()
                                    if gameflow_phase == "None":
                                        logPrint("您尚未创建任何房间！请创建房间后再按回车键开始检测。\nYou haven't created any lobby yet! Please create a lobby and then press Enter to start detection.")
                                        continue
                                    elif gameflow_phase in {"Lobby", "Matchmaking", "ReadyCheck", "ChampSelect", "InProgress", "Reconnect"}:
                                        # if gameflow_phase == "ChampSelect":
                                        #     if Vanguard_warning_printed:
                                        #         logPrint("您已进入英雄选择阶段！请在进入游戏后再按回车键开始检测。\nChamp select stage has started! Please press Enter to start detection after entering the game.")
                                        #     else:
                                        #         logPrint("鉴于拳头反作弊系统对于房间内队友信息访问行为的打击，本脚本已停用英雄选择阶段对曾经遇到过的队友的检测。请在进入游戏后再按回车键开始检测。\nIn view of Riot Vanguard's fight against Lobby Reveal behaviors, this program has banned the detection of recently played summoners during champ select stage. Please press Enter to start detection after entering the game.")
                                        #         Vanguard_warning_printed = True
                                        #     continue
                                        break
                                    elif gameflow_phase in {"WaitingForStats", "EndOfGame", "PreEndOfGame"}:
                                        logPrint("您已完成对局！请使用生成模式以查看最近一局比赛中遇到的玩家信息，或者开启下一局以查看下一局遇到的队友是否曾经遇到过。\nYou've finished the match! Please use [Generate Mode] to check the information of players encountered in the latest match, or start another game and use [Detect Mode] to check whether an ally has been met before.")
                                        continue
                                if detect != "" and detect[0] == "0":
                                    continue
                                friends = await (await connection.request("GET", "/lol-chat/v1/friends")).json()
                                friends = list(map(lambda x: x["puuid"], friends))
                                update = False
                                if gameflow_phase in {"Lobby", "Matchmaking", "ReadyCheck"}:
                                    lobby = await (await connection.request("GET", "/lol-lobby/v2/lobby")).json()
                                    logPrint(lobby)
                                    excel_name = "Recently Played Summoners in Lobby %s-%s.xlsx" %(platformId, lobby["partyId"])
                                    for member in lobby["members"]:
                                        if member["puuid"] != current_puuid: #这里不需要改成自己的玩家通用唯一识别码列表。有两个原因：一是一个会话仅属于一名英雄联盟玩家；二是前面整理玩家信息时，小号已经被排除，所以这里不可能会有成员为小号（Here the `current_puuid` doesn't need to be replaced by the self puuid list. Two reasons: first, a session only belongs to a single League of Legends player; second, while sorting out the player information before, smurf accounts have been excluded, so it's impossible for any member to correspond to a smurf）
                                            member_info_recapture = 0
                                            if member["puuid"] in infos:
                                                member_info_body = infos[member["puuid"]]
                                            else:
                                                member_info = await get_info(connection, member["puuid"])
                                                while not member_info["info_got"] and member_info["body"]["httpStatus"] != 404 and member_info_recapture < 3:
                                                    logPrint(member_info["message"])
                                                    member_info_recapture += 1
                                                    logPrint("成员信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of a member (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(member["puuid"], member_info_recapture, member["puuid"], member_info_recapture))
                                                    member_info = await get_info(connection, member["puuid"])
                                                if member_info["info_got"]:
                                                    member_info_body = member_info["body"]
                                                    infos[member["puuid"]] = member_info_body
                                                else:
                                                    logPrint(member_info["message"])
                                                    logPrint("成员信息（玩家通用唯一识别码：%s）获取失败！将忽略该名成员。\nInformation of a member (puuid: %s) capture failed! The program will ignore this member.")
                                                    continue
                                            LoLMember_index = [0]
                                            TFTMember_index = [0]
                                            if search_LoL:
                                                for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                    if recent_LoLPlayers_df.at[i, "puuid"] == member["puuid"]:
                                                        LoLMember_index.append(i)
                                            if search_TFT:
                                                for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                                    if recent_TFTPlayers_df.at[i, "puuid"] == member["puuid"]:
                                                        TFTMember_index.append(i)
                                            if len(LoLMember_index) + len(TFTMember_index) > 2: #这里不需要关于是否查询了云顶之弈对局记录分类讨论，因为不管有没有查询云顶之弈对局记录，TFTMember_index都存在，且长度至少为1（Here it's not necessary to discuss whether TFT match history has been searched before, because no matter whether it's searched, TFTMember_index is defined and its length is at least 1）
                                                member_count += 1
                                                LoLMember_df = recent_LoLPlayers_df.loc[LoLMember_index, :]
                                                LoLMember_df_to_print = pandas.concat([LoLMember_df_to_print, LoLMember_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                                TFTMember_df = recent_TFTPlayers_df.loc[TFTMember_index, :]
                                                TFTMember_df_to_print = pandas.concat([TFTMember_df_to_print, TFTMember_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                                if member["puuid"] in friends:
                                                    recent_friends.append(get_info_name(member_info_body))
                                                while True:
                                                    try:
                                                        with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                            if search_LoL and len(LoLMember_index) > 1:
                                                                LoLMember_df.to_excel(excel_writer = writer, sheet_name = get_info_name(member_info_body) + " (LoL)")
                                                            if search_TFT and len(TFTMember_index) > 1:
                                                                TFTMember_df.to_excel(excel_writer = writer, sheet_name = get_info_name(member_info_body) + " (TFT)")
                                                            logPrint("成员%s曾经与您一同战斗过%d次。\nMember %s has fought with you for %d time(s)." %(get_info_name(member_info_body), len(LoLMember_index) + len(TFTMember_index) - 2, get_info_name(member_info_body), len(LoLMember_index) + len(TFTMember_index) - 2))
                                                    except PermissionError:
                                                        logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                        logInput()
                                                    except FileNotFoundError:
                                                        with pandas.ExcelWriter(path = excel_name) as writer:
                                                            if search_LoL and len(LoLMember_index) > 1:
                                                                LoLMember_df.to_excel(excel_writer = writer, sheet_name = get_info_name(member_info_body) + " (LoL)")
                                                            if search_TFT and len(TFTMember_index) > 1:
                                                                TFTMember_df.to_excel(excel_writer = writer, sheet_name = get_info_name(member_info_body) + " (TFT)")
                                                            logPrint("成员%s曾经与您一同战斗过%d次。\nMember %s has fought with you for %d time(s)." %(get_info_name(member_info_body), len(LoLMember_index) + len(TFTMember_index) - 2, get_info_name(member_info_body), len(LoLMember_index) + len(TFTMember_index) - 2))
                                                        break
                                                    else:
                                                        break
                                    if len(lobby["members"]) == 1:
                                        if gameflow_phase == "Lobby":
                                            logPrint('''房间内无其它玩家。请单击寻找对局或开始游戏按钮，在进入英雄选择阶段后再按回车键开始检测。\nThere's not any other player in the lobby. Please click the "FIND MATCH" or "START GAME" button and press Enter to start detection after entering champ select stage.''')
                                        elif gameflow_phase == "Matchmaking":
                                            logPrint("房间内无其它玩家。请在接受对局进入英雄选择阶段后再按回车键开始检测。\nThere's not any other player in the lobby. Please press Enter to start detection after accepting a match and entering champ select stage.")
                                        elif gameflow_phase == "ReadyCheck":
                                            logPrint("房间内无其它玩家。请接受对局，并在进入英雄选择阶段后按回车键开始检测。\nThere's not any other player in the lobby. Please accept this match and press Enter to start detection after entering champ select stage.")
                                    elif member_count == 0:
                                        if gameflow_phase == "Lobby":
                                            logPrint('''您目前遇到的都是新的成员。请单击寻找对局或开始游戏按钮，在进入英雄选择阶段后再按回车键开始检测。\nThe members you've met now are all new. Please click the "FIND MATCH" or "START GAME" button and press Enter to start detection after entering champ select stage.''')
                                        elif gameflow_phase == "Matchmaking":
                                            logPrint("您目前遇到的都是新的成员。请在接受对局进入英雄选择阶段后再按回车键开始检测。\nThe members you've met now are all new. Please press Enter to start detection after accepting a match and entering champ select stage.")
                                        elif gameflow_phase == "ReadyCheck":
                                            logPrint("您目前遇到的都是新的成员。请接受对局，并在进入英雄选择阶段后按回车键开始检测。\nThe members you've met now are all new. Please accept this match and press Enter to start detection after entering champ select stage.")
                                    else:
                                        logPrint()
                                        if search_LoL:
                                            print(format_df(LoLMember_df_to_print, print_index = True, reserve_index = True)[0])
                                            log.write(format_df(LoLMember_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                        if search_LoL and search_TFT:
                                            logPrint()
                                        if search_TFT:
                                            print(format_df(TFTMember_df_to_print, print_index = True, reserve_index = True)[0])
                                            log.write(format_df(TFTMember_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                        if member_count == 1:
                                            logPrint('''一名成员曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's a member present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                        else:
                                            logPrint('''%d名成员曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d members present in your past matches. Please check the workbook "%s" in the main directory.''' %(member_count, excel_name, member_count, excel_name))
                                    if len(recent_friends) == 1:
                                        logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s is your friend." %(recent_friends[0], recent_friends[0]))
                                    elif len(recent_friends) > 1:
                                        logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s are your friends." %("、".join(recent_friends), ", ".join(recent_friends)))
                                elif gameflow_phase == "ChampSelect":
                                    champ_select_session = await (await connection.request("GET", "/lol-champ-select/v1/session")).json()
                                    logPrint(champ_select_session)
                                    if "errorCode" in champ_select_session:
                                        if champ_select_session["message"] == "No active delegate": #在没有英雄选择阶段的游戏模式中，有时gameflow_phase的结果是“ChampSelect”，但是实际上没有可用的英雄选择会话（In game modes without champ select stage, sometimes `gameflow_phase` is "ChampSelect", but there's actually no available champ select session）
                                            logPrint("英雄选择会话已过期。\nChamp select session has expired.")
                                        continue
                                    excel_name = "Recently Played Summoners in Match %s-%s.xlsx" %(platformId, champ_select_session["gameId"])
                                    lobby = await (await connection.request("GET", "/lol-lobby/v2/lobby")).json()
                                    skip_lobby_member = False
                                    if not "errorCode" in lobby and len(lobby["members"]) > 1:
                                        logPrint("检测时是否忽略小队成员？（输入任意键忽略，否则不忽略。）\nNeglect lobby members when detecting? (Submit any non-empty string to neglect, or null to refust neglecting.)")
                                        skip_lobby_member_str = logInput()
                                        skip_lobby_member = bool(skip_lobby_member_str)
                                        lobby_member_puuids = list(map(lambda x: x["puuid"], lobby["members"]))
                                    for ally in champ_select_session["myTeam"]:
                                        if not ally["puuid"] in {current_info["puuid"], "", bot_puuid} and (ally["nameVisibilityType"] == "VISIBLE" or ally["nameVisibilityType"] == ""):
                                            ally_info_recapture = 0
                                            if ally["puuid"] in infos:
                                                ally_info_body = infos[ally["puuid"]]
                                            else:
                                                ally_info = await get_info(connection, ally["puuid"])
                                                while not ally_info["info_got"] and ally_info["body"]["httpStatus"] != 404 and ally_info_recapture < 3:
                                                    logPrint(ally_info["message"])
                                                    ally_info_recapture += 1
                                                    logPrint("队友信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an ally (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(ally["puuid"], ally_info_recapture, ally["puuid"], ally_info_recapture))
                                                    ally_info = await get_info(connection, ally["puuid"])
                                                if ally_info["info_got"]:
                                                    ally_info_body = ally_info["body"]
                                                    infos[ally["puuid"]] = ally_info_body
                                                else:
                                                    logPrint(ally_info["message"])
                                                    logPrint("队友信息（玩家通用唯一识别码：%s）获取失败！将忽略该名队友。\nInformation of an ally (puuid: %s) capture failed! The program will ignore this ally.")
                                                    continue
                                            LoLAlly_index = [0] #第0行是中文表头，所以一开始要包含在内（The 0th line is Chinese header, so it should be contained in the beginning）
                                            TFTAlly_index = [0]
                                            if search_LoL:
                                                for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                    if recent_LoLPlayers_df.at[i, "puuid"] == ally["puuid"] and not (skip_lobby_member and recent_LoLPlayers_df.at[i, "puuid"] in lobby_member_puuids):
                                                        LoLAlly_index.append(i)
                                            if search_TFT:
                                                for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                                    if recent_TFTPlayers_df.at[i, "puuid"] == ally["puuid"] and not (skip_lobby_member and recent_TFTPlayers_df.at[i, "puuid"] in lobby_member_puuids):
                                                        TFTAlly_index.append(i)
                                            if len(LoLAlly_index) + len(TFTAlly_index) > 2: #这里不需要关于是否查询了云顶之弈对局记录分类讨论，因为不管有没有查询云顶之弈对局记录，TFTAlly_index都存在，且长度至少为1（Here it's not necessary to discuss whether TFT match history has been searched before, because no matter whether it's searched, TFTAlly_index is defined and its length is at least 1）
                                                ally_count += 1
                                                LoLAlly_df = recent_LoLPlayers_df.loc[LoLAlly_index, :]
                                                LoLAlly_df_to_print = pandas.concat([LoLAlly_df_to_print, LoLAlly_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                                TFTAlly_df = recent_TFTPlayers_df.loc[TFTAlly_index, :]
                                                TFTAlly_df_to_print = pandas.concat([TFTAlly_df_to_print, TFTAlly_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                                if ally["puuid"] in friends:
                                                    recent_friends.append(get_info_name(ally_info_body))
                                                while True:
                                                    try:
                                                        with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                            if search_LoL and len(LoLAlly_index) > 1:
                                                                LoLAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (LoL)")
                                                            if search_TFT and len(TFTAlly_index) > 1:
                                                                TFTAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (TFT)")
                                                            logPrint("队友%s曾经与您一同战斗过%d次。\nAlly %s has fought with you for %d time(s)." %(get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2, get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2))
                                                    except PermissionError:
                                                        logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                        logInput()
                                                    except FileNotFoundError:
                                                        with pandas.ExcelWriter(path = excel_name) as writer:
                                                            if search_LoL and len(LoLAlly_index) > 1:
                                                                LoLAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (LoL)")
                                                            if search_TFT and len(TFTAlly_index) > 1:
                                                                TFTAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (TFT)")
                                                            logPrint("队友%s曾经与您一同战斗过%d次。\nAlly %s has fought with you for %d time(s)." %(get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2, get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2))
                                                        break
                                                    else:
                                                        break
                                    if champ_select_session["theirTeam"]: #在人机对战、云顶之弈和斗魂竞技场中，无敌方玩家（There're no enemy players in bot games, TFT and Arena）
                                        for enemy in champ_select_session["theirTeam"]:
                                            if not enemy["puuid"] in {current_info["puuid"], "", bot_puuid} and (enemy["nameVisibilityType"] == "VISIBLE" or enemy["nameVisibilityType"] == ""):
                                                enemy_info_recapture = 0
                                                if enemy["puuid"] in infos:
                                                    enemy_info_body = infos[enemy["puuid"]]
                                                else:
                                                    enemy_info = await get_info(connection, enemy["puuid"])
                                                    while not enemy_info["info_got"] and enemy_info["body"]["httpStatus"] != 404 and enemy_info_recapture < 3:
                                                        logPrint(enemy_info["message"])
                                                        enemy_info_recapture += 1
                                                        logPrint("对手信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an enemy (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(enemy["puuid"], enemy_info_recapture, enemy["puuid"], enemy_info_recapture))
                                                        enemy_info = await get_info(connection, enemy["puuid"])
                                                    if enemy_info["info_got"]:
                                                        enemy_info_body = enemy_info["body"]
                                                        infos[enemy["puuid"]] = enemy_info_body
                                                    else:
                                                        logPrint(enemy_info["message"])
                                                        logPrint("对手信息（玩家通用唯一识别码：%s）获取失败！将忽略该名对手。\nInformation of an enemy (puuid: %s) capture failed! The program will ignore this enemy.")
                                                        continue
                                                LoLEnemy_index = [0]
                                                TFTEnemy_index = [0]
                                                if search_LoL:
                                                    for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                        if recent_LoLPlayers_df.at[i, "puuid"] == enemy["puuid"] and not (skip_lobby_member and recent_LoLPlayers_df.at[i, "puuid"] in lobby_member_puuids):
                                                            LoLEnemy_index.append(i)
                                                if search_TFT:
                                                    for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])) and not (skip_lobby_member and recent_TFTPlayers_df.at[i, "puuid"] in lobby_member_puuids):
                                                        if recent_TFTPlayers_df.at[i, "puuid"] == enemy["puuid"]:
                                                            TFTEnemy_index.append(i)
                                                if len(LoLEnemy_index) + len(TFTEnemy_index) > 2:
                                                    enemy_count += 1
                                                    LoLEnemy_df = recent_LoLPlayers_df.loc[LoLEnemy_index, :]
                                                    LoLEnemy_df_to_print = pandas.concat([LoLEnemy_df_to_print, LoLEnemy_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                                    TFTEnemy_df = recent_TFTPlayers_df.loc[TFTEnemy_index, :]
                                                    TFTEnemy_df_to_print = pandas.concat([TFTEnemy_df_to_print, TFTEnemy_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                                    if enemy["puuid"] in friends:
                                                        recent_friends.append((get_info_name(enemy_info_body)))
                                                    while True:
                                                        try:
                                                            with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                                if search_LoL and len(LoLEnemy_index) > 1:
                                                                    LoLEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTEnemy_index) > 1:
                                                                    TFTEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (TFT)")
                                                                logPrint("对手%s曾经与您一同战斗过%d次。\nEnemy %s has fought with you for %d time(s)." %(get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2, get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2))
                                                        except PermissionError:
                                                            logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                            logInput()
                                                        except FileNotFoundError:
                                                            with pandas.ExcelWriter(path = excel_name) as writer:
                                                                if search_LoL and len(LoLEnemy_index) > 1:
                                                                    LoLEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTEnemy_index) > 1:
                                                                    TFTEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (TFT)")
                                                                logPrint("对手%s曾经与您一同战斗过%d次。\nEnemy %s has fought with you for %d time(s)." %(get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2, get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2))
                                                            break
                                                        else:
                                                            break
                                    if ally_count == 0:
                                        logPrint("您目前遇到的都是新的队友。尝试拓展人缘吧！\nThe allies you've met now are all new. Try extending your friendship!")
                                    else:
                                        logPrint()
                                        if search_LoL:
                                            print(format_df(LoLAlly_df_to_print, print_index = True, reserve_index = True)[0])
                                            log.write(format_df(LoLAlly_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                        if search_LoL and search_TFT:
                                            logPrint()
                                        if search_TFT:
                                            print(format_df(TFTAlly_df_to_print, print_index = True, reserve_index = True)[0])
                                            log.write(format_df(TFTAlly_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                        if ally_count == 1:
                                            logPrint('''一名队友曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's an ally present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                        else:
                                            logPrint('''%d名队友曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d allies present in your past matches. Please check the workbook "%s" in the main directory.''' %(ally_count, excel_name, ally_count, excel_name))
                                    if any(map(lambda x: x["nameVisibilityType"] == "VISIBLE" or x["nameVisibilityType"] == "", champ_select_session["theirTeam"])):
                                        if enemy_count > 0:
                                            logPrint()
                                            if search_LoL:
                                                print(format_df(LoLEnemy_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(LoLEnemy_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                            if search_LoL and search_TFT:
                                                logPrint()
                                            if search_TFT:
                                                print(format_df(TFTEnemy_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(TFTEnemy_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                            if enemy_count == 1:
                                                logPrint('''一名对手曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's an enemy present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                            else:
                                                logPrint('''%d名对手曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d enemies present in your past matches. Please check the workbook "%s" in the main directory.''' %(enemy_count, excel_name, enemy_count, excel_name))
                                    if len(recent_friends) == 1:
                                        logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s is your friend." %(recent_friends[0], recent_friends[0]))
                                    elif len(recent_friends) > 1:
                                        logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s are your friends." %("、".join(recent_friends), ", ".join(recent_friends)))
                                    if not (all(map(lambda x: x["nameVisibilityType"] == "VISIBLE", champ_select_session["theirTeam"])) or all(map(lambda x: x["nameVisibilityType"] == "HIDDEN", champ_select_session["theirTeam"])) or all(map(lambda x: x["nameVisibilityType"] == "", champ_select_session["theirTeam"])) or all(map(lambda x: x["nameVisibilityType"] == "", champ_select_session["theirTeam"]))):
                                        logPrint("检测到敌方信息可见性异常！请检查之前输出的英雄选择阶段信息。\nDetected enemies' visibility abnormal! Please check the champ select session information printed before.")
                                elif gameflow_phase == "InProgress" or gameflow_phase == "Reconnect":
                                    gameflow_session = await (await connection.request("GET", "/lol-gameflow/v1/session")).json()
                                    logPrint(gameflow_session)
                                    gameData = gameflow_session["gameData"]
                                    excel_name = "Recently Played Summoners in Match %s-%s.xlsx" %(platformId, gameData["gameId"])
                                    if gameData["queue"]["mapId"] == "22" or gameData["queue"]["mapId"] == "30": #玩家在API上的阵营划分随对局模式而不同。云顶之弈和斗魂竞技场虽然有多个阵营，但是都是记录在gameData["teamOne"]中，这需要和其它模式区分开来。该条件语句与“if gameData["queue"]["gameMode"] == "TFT" or gameData["queue"]["gameMode"] == "CHERRY"”等价，但是因为召唤师峡谷还能分成CLASSIC、URF等模式，所以这里直接用地图序号作为判断依据（The team where a player belongs varies by the game mode. Although there're actually more than 2 teams in TFT and Arena, all players are recorded in `gameData["teamOne"]`, which needs ditinguishing from other game modes. This conditional statement is equivalent to `if gameData["queue"]["gameMode"] == "TFT" or gameData["queue"]["gameMode"] == "CHERRY"`, but since there're multiple modes based on one map, like CLASSIC and URF based on Summoner's Rift, the mapId is thus taken as the judgment criterium）
                                        for player in gameData["teamOne"]:
                                            if "puuid" in player and player["puuid"] != current_puuid: #电脑玩家没有玩家通用唯一识别码（Bot players don't have puuids）
                                                player_info_recapture = 0
                                                if player["puuid"] in infos:
                                                    player_info_body = infos[player["puuid"]]
                                                else:
                                                    player_info = await get_info(connection, player["puuid"])
                                                    while not player_info["info_got"] and player_info["body"]["httpStatus"] != 404 and player_info_recapture < 3:
                                                        logPrint(player_info["message"])
                                                        player_info_recapture += 1
                                                        logPrint("玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an player (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(player["puuid"], player_info_recapture, player["puuid"], player_info_recapture))
                                                        player_info = await get_info(connection, player["puuid"])
                                                    if player_info["info_got"]:
                                                        player_info_body = player_info["body"]
                                                        infos[player_info_body["puuid"]] = player_info_body
                                                    else:
                                                        logPrint(player_info["message"])
                                                        logPrint("玩家信息（玩家通用唯一识别码：%s）获取失败！将忽略该名队友。\nInformation of an player (puuid: %s) capture failed! The program will ignore this player.")
                                                        continue
                                                LoLPlayer_index = [0] #第0行是中文表头，所以一开始要包含在内（The 0th line is Chinese header, so it should be contained in the beginning）
                                                TFTPlayer_index = [0]
                                                if search_LoL:
                                                    for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                        if recent_LoLPlayers_df.at[i, "puuid"] == player["puuid"]:
                                                            LoLPlayer_index.append(i)
                                                if search_TFT:
                                                    for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                                        if recent_TFTPlayers_df.at[i, "puuid"] == player["puuid"]:
                                                            TFTPlayer_index.append(i)
                                                if len(LoLPlayer_index) + len(TFTPlayer_index) > 2: #这里不需要关于是否查询了云顶之弈对局记录分类讨论，因为不管有没有查询云顶之弈对局记录，TFTPlayer_index都存在，且长度至少为1（Here it's not necessary to discuss whether TFT match history has been searched before, because no matter whether it's searched, TFTPlayer_index is defined and its length is at least 1）
                                                    player_count += 1
                                                    LoLPlayer_df = recent_LoLPlayers_df.loc[LoLPlayer_index, :]
                                                    LoLPlayer_df_to_print = pandas.concat([LoLPlayer_df_to_print, LoLPlayer_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                                    TFTPlayer_df = recent_TFTPlayers_df.loc[TFTPlayer_index, :]
                                                    TFTPlayer_df_to_print = pandas.concat([TFTPlayer_df_to_print, TFTPlayer_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                                    if player["puuid"] in friends:
                                                        recent_friends.append(get_info_name(player_info_body))
                                                    while True:
                                                        try:
                                                            with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                                if search_LoL and len(LoLPlayer_index) > 1:
                                                                    LoLPlayer_df.to_excel(excel_writer = writer, sheet_name = get_info_name(player_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTPlayer_index) > 1:
                                                                    TFTPlayer_df.to_excel(excel_writer = writer, sheet_name = get_info_name(player_info_body) + " (TFT)")
                                                                logPrint("玩家%s曾经与您一同战斗过%d次。\nPlayer %s has fought with you for %d time(s)." %(get_info_name(player_info_body), len(LoLPlayer_index) + len(TFTPlayer_index) - 2, get_info_name(player_info_body), len(LoLPlayer_index) + len(TFTPlayer_index) - 2))
                                                        except PermissionError:
                                                            logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                            logInput()
                                                        except FileNotFoundError:
                                                            with pandas.ExcelWriter(path = excel_name) as writer:
                                                                if search_LoL and len(LoLPlayer_index) > 1:
                                                                    LoLPlayer_df.to_excel(excel_writer = writer, sheet_name = get_info_name(player_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTPlayer_index) > 1:
                                                                    TFTPlayer_df.to_excel(excel_writer = writer, sheet_name = get_info_name(player_info_body) + " (TFT)")
                                                                logPrint("玩家%s曾经与您一同战斗过%d次。\nPlayer %s has fought with you for %d time(s)." %(get_info_name(player_info_body), len(LoLPlayer_index) + len(TFTPlayer_index) - 2, get_info_name(player_info_body), len(LoLPlayer_index) + len(TFTPlayer_index) - 2))
                                                            break
                                                        else:
                                                            break
                                        if player_count == 0:
                                            logPrint("您目前遇到的都是新的玩家。尝试拓展人缘吧！\nThe players you've met now are all new. Try extending your friendship!")
                                        else:
                                            logPrint()
                                            if search_LoL:
                                                print(format_df(LoLPlayer_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(LoLPlayer_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                            if search_LoL and search_TFT:
                                                logPrint()
                                            if search_TFT:
                                                print(format_df(TFTPlayer_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(TFTPlayer_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                            if player_count == 1:
                                                logPrint('''一名玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's a player present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                            else:
                                                logPrint('''%d名玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d players present in your past matches. Please check the workbook "%s" in the main directory.''' %(player_count, excel_name, player_count, excel_name))
                                        if len(recent_friends) == 1:
                                            logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s is your friend." %(recent_friends[0], recent_friends[0]))
                                        elif len(recent_friends) > 1:
                                            logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s are your friends." %("、".join(recent_friends), ", ".join(recent_friends)))
                                    else:
                                        isSpectating = False #设置观战逻辑变量，确定游戏会话是不是观战的（This boolean variable is declared to tell whether the game session is spectating）
                                        teamOne_puuids = []
                                        for player in gameData["teamOne"]:
                                            if "puuid" in player:
                                                teamOne_puuids.append(player["puuid"])
                                        teamTwo_puuids = []
                                        for player in gameData["teamTwo"]:
                                            if "puuid" in player:
                                                teamTwo_puuids.append(player["puuid"])
                                        if current_puuid in teamOne_puuids: #API记录游戏中的玩家时，只会区分红蓝方，不会区分敌我。所以这里需要先判断那个阵营是我方（Players recorded in API only differentiate by blue or red team, instead of my or enemy team. So judging the own team or the enemy team is the first thing to do）
                                            myTeam = gameData["teamOne"]
                                            theirTeam = gameData["teamTwo"]
                                        elif current_puuid in teamTwo_puuids:
                                            myTeam = gameData["teamTwo"]
                                            theirTeam = gameData["teamOne"]
                                        else:
                                            myTeam = gameData["teamOne"] + gameData["teamTwo"]
                                            theirTeam = []
                                            isSpectating = True
                                        for ally in myTeam:
                                            if "puuid" in ally and ally["puuid"] != current_puuid:
                                                ally_info_recapture = 0
                                                if ally["puuid"] in infos:
                                                    ally_info_body = infos[ally["puuid"]]
                                                else:
                                                    ally_info = await get_info(connection, ally["puuid"])
                                                    while not ally_info["info_got"] and ally_info["body"]["httpStatus"] != 404 and ally_info_recapture < 3:
                                                        logPrint(ally_info["message"])
                                                        ally_info_recapture += 1
                                                        if isSpectating:
                                                            logPrint("玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of a player (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(ally["puuid"], ally_info_recapture, ally["puuid"], ally_info_recapture))
                                                        else:
                                                            logPrint("队友信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an ally (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(ally["puuid"], ally_info_recapture, ally["puuid"], ally_info_recapture))
                                                        ally_info = await get_info(connection, ally["puuid"])
                                                    if ally_info["info_got"]:
                                                        ally_info_body = ally_info["body"]
                                                        infos[ally_info_body["puuid"]] = ally_info_body
                                                    else:
                                                        logPrint(ally_info["message"])
                                                        if isSpectating:
                                                            logPrint("玩家信息（玩家通用唯一识别码：%s）获取失败！将忽略该名玩家。\nInformation of a player (puuid: %s) capture failed! The program will ignore this player.")
                                                        else:
                                                            logPrint("队友信息（玩家通用唯一识别码：%s）获取失败！将忽略该名队友。\nInformation of an ally (puuid: %s) capture failed! The program will ignore this ally.")
                                                        continue
                                                LoLAlly_index = [0] #第0行是中文表头，所以一开始要包含在内（The 0th line is Chinese header, so it should be contained in the beginning）
                                                TFTAlly_index = [0]
                                                if search_LoL:
                                                    for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                        if recent_LoLPlayers_df.at[i, "puuid"] == ally["puuid"]:
                                                            LoLAlly_index.append(i)
                                                if search_TFT:
                                                    for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                                        if recent_TFTPlayers_df.at[i, "puuid"] == ally["puuid"]:
                                                            TFTAlly_index.append(i)
                                                if len(LoLAlly_index) + len(TFTAlly_index) > 2: #这里不需要关于是否查询了云顶之弈对局记录分类讨论，因为不管有没有查询云顶之弈对局记录，TFTAlly_index都存在，且长度至少为1（Here it's not necessary to discuss whether TFT match history has been searched before, because no matter whether it's searched, TFTAlly_index is defined and its length is at least 1）
                                                    ally_count += 1
                                                    LoLAlly_df = recent_LoLPlayers_df.loc[LoLAlly_index, :]
                                                    LoLAlly_df_to_print = pandas.concat([LoLAlly_df_to_print, LoLAlly_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                                    TFTAlly_df = recent_TFTPlayers_df.loc[TFTAlly_index, :]
                                                    TFTAlly_df_to_print = pandas.concat([TFTAlly_df_to_print, TFTAlly_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                                    if ally["puuid"] in friends:
                                                        recent_friends.append(get_info_name(ally_info_body))
                                                    while True:
                                                        try:
                                                            with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                                if search_LoL and len(LoLAlly_index) > 1:
                                                                    LoLAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTAlly_index) > 1:
                                                                    TFTAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (TFT)")
                                                                if isSpectating:
                                                                    logPrint("玩家%s曾经与您一同战斗过%d次。\nPlayer %s has fought with you for %d time(s)." %(get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2, get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2))
                                                                else:
                                                                    logPrint("队友%s曾经与您一同战斗过%d次。\nAlly %s has fought with you for %d time(s)." %(get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2, get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2))
                                                        except PermissionError:
                                                            logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                            logInput()
                                                        except FileNotFoundError:
                                                            with pandas.ExcelWriter(path = excel_name) as writer:
                                                                if search_LoL and len(LoLAlly_index) > 1:
                                                                    LoLAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTAlly_index) > 1:
                                                                    TFTAlly_df.to_excel(excel_writer = writer, sheet_name = get_info_name(ally_info_body) + " (TFT)")
                                                                if isSpectating:
                                                                    logPrint("玩家%s曾经与您一同战斗过%d次。\nPlayer %s has fought with you for %d time(s)." %(get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2, get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2))
                                                                else:
                                                                    logPrint("队友%s曾经与您一同战斗过%d次。\nAlly %s has fought with you for %d time(s)." %(get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2, get_info_name(ally_info_body), len(LoLAlly_index) + len(TFTAlly_index) - 2))
                                                            break
                                                        else:
                                                            break
                                        for enemy in theirTeam:
                                            if "puuid" in enemy:
                                                if enemy["puuid"] in infos:
                                                    enemy_info_body = infos[enemy["puuid"]]
                                                else:
                                                    enemy_info_recapture = 0
                                                    enemy_info = await get_info(connection, enemy["puuid"])
                                                    while not enemy_info["info_got"] and enemy_info["body"]["httpStatus"] != 404 and enemy_info_recapture < 3:
                                                        logPrint(enemy_info["message"])
                                                        enemy_info_recapture += 1
                                                        logPrint("对手信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an enemy (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d." %(enemy["puuid"], enemy_info_recapture, enemy["puuid"], enemy_info_recapture))
                                                        enemy_info = await get_info(connection, enemy["puuid"])
                                                    if enemy_info["info_got"]:
                                                        enemy_info_body = enemy_info["body"]
                                                        infos[enemy["puuid"]] = enemy_info_body
                                                    else:
                                                        logPrint(enemy_info["message"])
                                                        logPrint("对手信息（玩家通用唯一识别码：%s）获取失败！将忽略该名对手。\nInformation of an enemy (puuid: %s) capture failed! The program will ignore this enemy.")
                                                        continue
                                                LoLEnemy_index = [0]
                                                TFTEnemy_index = [0]
                                                if search_LoL:
                                                    for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                        if recent_LoLPlayers_df.at[i, "puuid"] == enemy["puuid"]:
                                                            LoLEnemy_index.append(i)
                                                if search_TFT:
                                                    for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                                        if recent_TFTPlayers_df.at[i, "puuid"] == enemy["puuid"]:
                                                            TFTEnemy_index.append(i)
                                                if len(LoLEnemy_index) + len(TFTEnemy_index) > 2:
                                                    enemy_count += 1
                                                    LoLEnemy_df = recent_LoLPlayers_df.loc[LoLEnemy_index, :]
                                                    LoLEnemy_df_to_print = pandas.concat([LoLEnemy_df_to_print, LoLEnemy_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                                    TFTEnemy_df = recent_TFTPlayers_df.loc[TFTEnemy_index, :]
                                                    TFTEnemy_df_to_print = pandas.concat([TFTEnemy_df_to_print, TFTEnemy_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                                    if enemy["puuid"] in friends:
                                                        recent_friends.append((get_info_name(enemy_info_body)))
                                                    while True:
                                                        try:
                                                            with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                                if search_LoL and len(LoLEnemy_index) > 1:
                                                                    LoLEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTEnemy_index) > 1:
                                                                    TFTEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (TFT)")
                                                                logPrint("对手%s曾经与您一同战斗过%d次。\nEnemy %s has fought with you for %d time(s)." %(get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2, get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2))
                                                        except PermissionError:
                                                            logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                            logInput()
                                                        except FileNotFoundError:
                                                            with pandas.ExcelWriter(path = excel_name) as writer:
                                                                if search_LoL and len(LoLEnemy_index) > 1:
                                                                    LoLEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (LoL)")
                                                                if search_TFT and len(TFTEnemy_index) > 1:
                                                                    TFTEnemy_df.to_excel(excel_writer = writer, sheet_name = get_info_name(enemy_info_body) + " (TFT)")
                                                                logPrint("对手%s曾经与您一同战斗过%d次。\nEnemy %s has fought with you for %d time(s)." %(get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2, get_info_name(enemy_info_body), len(LoLEnemy_index) + len(TFTEnemy_index) - 2))
                                                            break
                                                        else:
                                                            break
                                        if isSpectating:
                                            if ally_count == 0:
                                                logPrint("您目前遇到的都是新的玩家。尝试拓展人缘吧！\nThe players you've met now are all new. Try extending your friendship!")
                                            else:
                                                logPrint()
                                                if search_LoL:
                                                    print(format_df(LoLAlly_df_to_print, print_index = True, reserve_index = True)[0])
                                                    log.write(format_df(LoLAlly_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                                if search_LoL and search_TFT:
                                                    logPrint()
                                                if search_TFT:
                                                    print(format_df(TFTAlly_df_to_print, print_index = True, reserve_index = True)[0])
                                                    log.write(format_df(TFTAlly_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                                if ally_count == 1:
                                                    logPrint('''一名玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's a player present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                                else:
                                                    logPrint('''%d名玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d players present in your past matches. Please check the workbook "%s" in the main directory.''' %(ally_count, excel_name, ally_count, excel_name))
                                        else:
                                            if ally_count == 0:
                                                logPrint("您目前遇到的都是新的玩家。尝试拓展人缘吧！\nThe players you've met now are all new. Try extending your friendship!")
                                            else:
                                                logPrint()
                                                if search_LoL:
                                                    print(format_df(LoLAlly_df_to_print, print_index = True, reserve_index = True)[0])
                                                    log.write(format_df(LoLAlly_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                                if search_LoL and search_TFT:
                                                    logPrint()
                                                if search_TFT:
                                                    print(format_df(TFTAlly_df_to_print, print_index = True, reserve_index = True)[0])
                                                    log.write(format_df(TFTAlly_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                                if ally_count == 1:
                                                    logPrint('''一名队友曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's an ally present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                                else:
                                                    logPrint('''%d名队友曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d allies present in your past matches. Please check the workbook "%s" in the main directory.''' %(ally_count, excel_name, ally_count, excel_name))
                                            if enemy_count > 0:
                                                logPrint()
                                                if search_LoL:
                                                    print(format_df(LoLEnemy_df_to_print, print_index = True, reserve_index = True)[0])
                                                    log.write(format_df(LoLEnemy_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                                if search_LoL and search_TFT:
                                                    logPrint()
                                                if search_TFT:
                                                    print(format_df(TFTEnemy_df_to_print, print_index = True, reserve_index = True)[0])
                                                    log.write(format_df(TFTEnemy_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                                if enemy_count == 1:
                                                    logPrint('''一名对手曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's an enemy present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                                else:
                                                    logPrint('''%d名对手曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d enemies present in your past matches. Please check the workbook "%s" in the main directory.''' %(enemy_count, excel_name, enemy_count, excel_name))
                                        if len(recent_friends) == 1:
                                            logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s is your friend." %(recent_friends[0], recent_friends[0]))
                                        elif len(recent_friends) > 1:
                                            logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s are your friends." %("、".join(recent_friends), ", ".join(recent_friends)))
                            elif detect_scene == "2":
                                recent_friend_count = 0
                                recent_LoLFriend_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                recent_TFTFriend_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                friends = await (await connection.request("GET", "/lol-chat/v1/friends")).json()
                                excel_name = "Recently Played Summoners in Friend List of %s - %s.xlsx" %(current_summonerName, platformId)
                                for friend in friends:
                                    friend_summonerName = friend["name"] if friend["gameName"] == "" and friend["gameTag"] == "" else friend["gameName"] + "#" + friend["gameTag"]
                                    LoLFriend_index = [0]
                                    TFTFriend_index = [0]
                                    if search_LoL:
                                        for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                            if recent_LoLPlayers_df.at[i, "puuid"] == friend["puuid"]:
                                                LoLFriend_index.append(i)
                                    if search_TFT:
                                        for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                            if recent_TFTPlayers_df.at[i, "puuid"] == friend["puuid"]:
                                                TFTFriend_index.append(i)
                                    if len(LoLFriend_index) + len(TFTFriend_index) > 2:
                                        recent_friend_count += 1
                                        recent_LoLFriend_df = recent_LoLPlayers_df.loc[LoLFriend_index, :]
                                        recent_LoLFriend_df.insert(1, "note", ["备注"] + [friend["note"]] * (len(LoLFriend_index) - 1))
                                        recent_LoLFriend_df_to_print = pandas.concat([recent_LoLFriend_df_to_print, recent_LoLFriend_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                        recent_TFTFriend_df = recent_TFTPlayers_df.loc[TFTFriend_index, :]
                                        recent_TFTFriend_df.insert(1, "note", ["备注"] + [friend["note"]] * (len(TFTFriend_index) - 1))
                                        recent_TFTFriend_df_to_print = pandas.concat([recent_TFTFriend_df_to_print, recent_TFTFriend_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                        while True:
                                            try:
                                                with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                    if search_LoL and len(LoLFriend_index) > 1:
                                                        recent_LoLFriend_df.to_excel(excel_writer = writer, sheet_name = friend_summonerName + " (LoL)")
                                                    if search_TFT and len(TFTFriend_index) > 1:
                                                        recent_TFTFriend_df.to_excel(excel_writer = writer, sheet_name = friend_summonerName + " (TFT)")
                                                    logPrint("好友%s曾经与您一同战斗过%d次。\nFriend %s has fought with you for %d time(s)." %(friend_summonerName, len(LoLFriend_index) + len(TFTFriend_index) - 2, friend_summonerName, len(LoLFriend_index) + len(TFTFriend_index) - 2))
                                            except PermissionError:
                                                logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                logInput()
                                            except FileNotFoundError:
                                                with pandas.ExcelWriter(path = excel_name) as writer:
                                                    if search_LoL and len(LoLFriend_index) > 1:
                                                        recent_LoLFriend_df.to_excel(excel_writer = writer, sheet_name = friend_summonerName + " (LoL)")
                                                    if search_TFT and len(TFTFriend_index) > 1:
                                                        recent_TFTFriend_df.to_excel(excel_writer = writer, sheet_name = friend_summonerName + " (TFT)")
                                                    logPrint("好友%s曾经与您一同战斗过%d次。\nFriend %s has fought with you for %d time(s)." %(friend_summonerName, len(LoLFriend_index) + len(TFTFriend_index) - 2, friend_summonerName, len(LoLFriend_index) + len(TFTFriend_index) - 2))
                                                break
                                            else:
                                                break
                                if len(friends) == 0:
                                    logPrint("您尚未添加任何好友。尝试拓展人缘吧！\nYou haven't added any friend. Try extending your friendship!")
                                elif recent_friend_count == 0:
                                    logPrint("您近期还没有和任何好友一起玩过。这不赶紧开个黑ヽ(*^ｰ^)人(^ｰ^*)ノ\nYou haven't played with any friend recently. Go for a game with one of your friends ...")
                                else:
                                    logPrint()
                                    if search_LoL:
                                        print(format_df(recent_LoLFriend_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_LoLFriend_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if search_LoL and search_TFT:
                                        logPrint()
                                    if search_TFT:
                                        print(format_df(recent_TFTFriend_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_TFTFriend_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if recent_friend_count == 1:
                                        logPrint('''一名好友曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's a friend present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                    else:
                                        logPrint('''%d名好友曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d friends present in your past matches. Please check the workbook "%s" in the main directory.''' %(recent_friend_count, excel_name, recent_friend_count, excel_name))
                            elif detect_scene == "3":
                                recent_prefriend_count = 0
                                recent_LoLPrefriend_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                recent_TFTPrefriend_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                friend_requests = await (await (connection.request("GET", "/lol-chat/v2/friend-requests"))).json()
                                excel_name = "Recently Played Summoners in Friend Requests of %s - %s.xlsx" %(current_summonerName, platformId)
                                for prefriend in friend_requests:
                                    prefriend_summonerName = prefriend["name"] if prefriend["gameName"] == "" and prefriend["tagLine"] == "" else prefriend["gameName"] + "#" + prefriend["tagLine"]
                                    LoLPrefriend_index = [0]
                                    TFTPrefriend_index = [0]
                                    if search_LoL:
                                        for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                            if recent_LoLPlayers_df.at[i, "puuid"] == prefriend["puuid"]:
                                                LoLPrefriend_index.append(i)
                                    if search_TFT:
                                        for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                            if recent_TFTPlayers_df.at[i, "puuid"] == prefriend["puuid"]:
                                                TFTPrefriend_index.append(i)
                                    if len(LoLPrefriend_index) + len(TFTPrefriend_index) > 2:
                                        recent_prefriend_count += 1
                                        recent_LoLPrefriend_df = recent_LoLPlayers_df.loc[LoLPrefriend_index, :]
                                        recent_LoLPrefriend_df_to_print = pandas.concat([recent_LoLPrefriend_df_to_print, recent_LoLPrefriend_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                        recent_TFTPrefriend_df = recent_TFTPlayers_df.loc[TFTPrefriend_index, :]
                                        recent_TFTPrefriend_df_to_print = pandas.concat([recent_TFTPrefriend_df_to_print, recent_TFTPrefriend_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                        while True:
                                            try:
                                                with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                    if search_LoL and len(LoLPrefriend_index) > 1:
                                                        recent_LoLPrefriend_df.to_excel(excel_writer = writer, sheet_name = prefriend_summonerName + " (" + prefriend["direction"] + ") (LoL)")
                                                    if search_TFT and len(TFTPrefriend_index) > 1:
                                                        recent_TFTPrefriend_df.to_excel(excel_writer = writer, sheet_name = prefriend_summonerName + " (" + prefriend["direction"] + ") (TFT)")
                                                    logPrint("好友请求列表中的%s曾经与您一同战斗过%d次。\nPlayer %s in friend request list has fought with you for %d time(s)." %(prefriend_summonerName, len(LoLPrefriend_index) + len(TFTPrefriend_index) - 2, prefriend_summonerName, len(LoLPrefriend_index) + len(TFTPrefriend_index) - 2))
                                            except PermissionError:
                                                logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                logInput()
                                            except FileNotFoundError:
                                                with pandas.ExcelWriter(path = excel_name) as writer:
                                                    if search_LoL and len(LoLPrefriend_index) > 1:
                                                        recent_LoLPrefriend_df.to_excel(excel_writer = writer, sheet_name = prefriend_summonerName + " (" + prefriend["direction"] + ") (LoL)")
                                                    if search_TFT and len(TFTPrefriend_index) > 1:
                                                        recent_TFTPrefriend_df.to_excel(excel_writer = writer, sheet_name = prefriend_summonerName + " (" + prefriend["direction"] + ") (TFT)")
                                                    logPrint("好友请求列表中的%s曾经与您一同战斗过%d次。\nPlayer %s in friend request list has fought with you for %d time(s)." %(prefriend_summonerName, len(LoLPrefriend_index) + len(TFTPrefriend_index) - 2, prefriend_summonerName, len(LoLPrefriend_index) + len(TFTPrefriend_index) - 2))
                                                break
                                            else:
                                                break
                                if len(friend_requests) == 0:
                                    logPrint("您尚未发送或收到任何好友请求。尝试拓展人缘吧！\nYou haven't sent or received any friend request. Try extending your friendship!")
                                elif recent_prefriend_count == 0:
                                    logPrint("您近期未曾和好友请求列表中的玩家一起战斗过。这可能是因为好友请求太久未审核，或者该请求源于朋友或视频推荐，或者该请求不正当。\nYou haven't fought with any player in the friend request list. This may be because this request is put aside for too long, this request results from the recommendation from a friend or a video, or this request isn't sent in a proper manner.")
                                else:
                                    logPrint()
                                    if search_LoL:
                                        print(format_df(recent_LoLPrefriend_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_LoLPrefriend_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if search_LoL and search_TFT:
                                        logPrint()
                                    if search_TFT:
                                        print(format_df(recent_TFTPrefriend_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_TFTPrefriend_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if recent_prefriend_count == 1:
                                        logPrint('''好友请求列表中的一名玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's a friend in the request list that is present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                    else:
                                        logPrint('''好友请求列表中的%d名好友曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d friends in the request that is present in your past matches. Please check the workbook "%s" in the main directory.''' %(recent_prefriend_count, excel_name, recent_prefriend_count, excel_name))
                            elif detect_scene == "4":
                                invitee_count = inviter_count = 0
                                recent_friends = []
                                LoLInvitee_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                TFTInvitee_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                LoLInviter_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                TFTInviter_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                excel_name = "Recently Played Summoners in Invitations to and from %s - %s.xlsx" %(current_summonerName, platformId)
                                lobby = await (await connection.request("GET", "/lol-lobby/v2/lobby")).json()
                                lobbyInvitations = await (await connection.request("GET", "/lol-lobby/v2/lobby/invitations")).json()
                                if not "errorCode" in lobbyInvitations:
                                    for invid in lobbyInvitations:
                                        if invid["toSummonerId"] != current_summonerId:
                                            invitee_info_recapture = 0
                                            invitee_info = await get_info(connection, invid["toSummonerId"])
                                            while not invitee_info["info_got"] and invitee_info["body"]["httpStatus"] != 404 and invitee_info_recapture < 3:
                                                logPrint(invitee_info["message"])
                                                invitee_info_recapture += 1
                                                logPrint("被邀请者信息（召唤师序号：%d）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an invitee (summonerId: %d) capture failed! Recapturing this player's information ... Times tried: %d." %(invid["toSummonerId"], invitee_info_recapture, invid["toSummonerId"], invitee_info_recapture))
                                                invitee_info = await get_info(connection, invid["toSummonerId"])
                                            if invitee_info["info_got"]:
                                                invitee_info_body = invitee_info["body"]
                                                infos[invitee_info_body["puuid"]] = invitee_info_body
                                            else:
                                                logPrint(invitee_info["message"])
                                                logPrint("被邀请者信息（召唤师序号：%d）获取失败！将忽略该被邀请者。\nInformation of an invitee (summonerId: %d) capture failed! The program will ignore this invitee.")
                                                continue
                                            LoLInvitee_index = [0]
                                            TFTInvitee_index = [0]
                                            if search_LoL:
                                                for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                    if recent_LoLPlayers_df.at[i, "puuid"] == invitee_info_body["puuid"]:
                                                        LoLInvitee_index.append(i)
                                            if search_TFT:
                                                for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                                    if recent_TFTPlayers_df.at[i, "puuid"] == invitee_info_body["puuid"]:
                                                        TFTInvitee_index.append(i)
                                            if len(LoLInvitee_index) + len(TFTInvitee_index) > 2: #这里不需要关于是否查询了云顶之弈对局记录分类讨论，因为不管有没有查询云顶之弈对局记录，TFTInvitee_index都存在，且长度至少为1（Here it's not necessary to discuss whether TFT match history has been searched before, because no matter whether it's searched, TFTInvitee_index is defined and its length is at least 1）
                                                invitee_count += 1
                                                LoLInvitee_df = recent_LoLPlayers_df.loc[LoLInvitee_index, :]
                                                LoLInvitee_df_to_print = pandas.concat([LoLInvitee_df_to_print, LoLInvitee_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                                TFTInvitee_df = recent_TFTPlayers_df.loc[TFTInvitee_index, :]
                                                TFTInvitee_df_to_print = pandas.concat([TFTInvitee_df_to_print, TFTInvitee_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                                if invitee_info_body["puuid"] in friends:
                                                    recent_friends.append(get_info_name(invitee_info_body))
                                                while True:
                                                    try:
                                                        with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                            if search_LoL and len(LoLInvitee_index) > 1:
                                                                LoLInvitee_df.to_excel(excel_writer = writer, sheet_name = get_info_name(invitee_info_body) + " (out) (LoL)")
                                                            if search_TFT and len(TFTInvitee_index) > 1:
                                                                TFTInvitee_df.to_excel(excel_writer = writer, sheet_name = get_info_name(invitee_info_body) + " (out) (TFT)")
                                                            logPrint("被邀请者%s曾经与您一同战斗过%d次。\nInvitee %s has fought with you for %d time(s)." %(get_info_name(invitee_info_body), len(LoLInvitee_index) + len(TFTInvitee_index) - 2, get_info_name(invitee_info_body), len(LoLInvitee_index) + len(TFTInvitee_index) - 2))
                                                    except PermissionError:
                                                        logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                        logInput()
                                                    except FileNotFoundError:
                                                        with pandas.ExcelWriter(path = excel_name) as writer:
                                                            if search_LoL and len(LoLInvitee_index) > 1:
                                                                LoLInvitee_df.to_excel(excel_writer = writer, sheet_name = get_info_name(invitee_info_body) + " (out) (LoL)")
                                                            if search_TFT and len(TFTInvitee_index) > 1:
                                                                TFTInvitee_df.to_excel(excel_writer = writer, sheet_name = get_info_name(invitee_info_body) + " (out) (TFT)")
                                                            logPrint("被邀请者%s曾经与您一同战斗过%d次。\nInvitee %s has fought with you for %d time(s)." %(get_info_name(invitee_info_body), len(LoLInvitee_index) + len(TFTInvitee_index) - 2, get_info_name(invitee_info_body), len(LoLInvitee_index) + len(TFTInvitee_index) - 2))
                                                        break
                                                    else:
                                                        break
                                receivedInvitations = await (await connection.request("GET", "/lol-lobby/v2/received-invitations")).json()
                                for invid in receivedInvitations:
                                    inviter_info_recapture = 0
                                    inviter_info = await get_info(connection, invid["fromSummonerId"])
                                    while not inviter_info["info_got"] and inviter_info["body"]["httpStatus"] != 404 and inviter_info_recapture < 3:
                                        logPrint(inviter_info["message"])
                                        inviter_info_recapture += 1
                                        logPrint("邀请者信息（召唤师序号：%d）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of an inviter (summonerId: %d) capture failed! Recapturing this player's information ... Times tried: %d." %(invid["fromSummonerId"], inviter_info_recapture, invid["fromSummonerId"], inviter_info_recapture))
                                        inviter_info = await get_info(connection, invid["fromSummonerId"])
                                    if inviter_info["info_got"]:
                                        inviter_info_body = inviter_info["body"]
                                        infos[inviter_info_body["puuid"]] = inviter_info_body
                                    else:
                                        logPrint(inviter_info["message"])
                                        logPrint("邀请者信息（召唤师序号：%d）获取失败！将忽略该邀请者。\nInformation of an inviter (summonerId: %d) capture failed! The program will ignore this inviter.")
                                        continue
                                    LoLInviter_index = [0]
                                    TFTInviter_index = [0]
                                    if search_LoL:
                                        for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                            if recent_LoLPlayers_df.at[i, "puuid"] == inviter_info_body["puuid"]:
                                                LoLInviter_index.append(i)
                                    if search_TFT:
                                        for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                            if recent_TFTPlayers_df.at[i, "puuid"] == inviter_info_body["puuid"]:
                                                TFTInviter_index.append(i)
                                    if len(LoLInviter_index) + len(TFTInviter_index) > 2: #这里不需要关于是否查询了云顶之弈对局记录分类讨论，因为不管有没有查询云顶之弈对局记录，TFTInviter_index都存在，且长度至少为1（Here it's not necessary to discuss whether TFT match history has been searched before, because no matter whether it's searched, TFTInviter_index is defined and its length is at least 1）
                                        inviter_count += 1
                                        LoLInviter_df = recent_LoLPlayers_df.loc[LoLInviter_index, :]
                                        LoLInviter_df_to_print = pandas.concat([LoLInviter_df_to_print, LoLInviter_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                        TFTInviter_df = recent_TFTPlayers_df.loc[TFTInviter_index, :]
                                        TFTInviter_df_to_print = pandas.concat([TFTInviter_df_to_print, TFTInviter_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                        if inviter_info_body["puuid"] in friends:
                                            recent_friends.append(get_info_name(inviter_info_body))
                                        while True:
                                            try:
                                                with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                    if search_LoL and len(LoLInviter_index) > 1:
                                                        LoLInviter_df.to_excel(excel_writer = writer, sheet_name = get_info_name(inviter_info_body) + " (in) (LoL)")
                                                    if search_TFT and len(TFTInviter_index) > 1:
                                                        TFTInviter_df.to_excel(excel_writer = writer, sheet_name = get_info_name(inviter_info_body) + " (in) (TFT)")
                                                    logPrint("邀请者%s曾经与您一同战斗过%d次。\nInviter %s has fought with you for %d time(s)." %(get_info_name(inviter_info_body), len(LoLInviter_index) + len(TFTInviter_index) - 2, get_info_name(inviter_info_body), len(LoLInviter_index) + len(TFTInviter_index) - 2))
                                            except PermissionError:
                                                logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                logInput()
                                            except FileNotFoundError:
                                                with pandas.ExcelWriter(path = excel_name) as writer:
                                                    if search_LoL and len(LoLInviter_index) > 1:
                                                        LoLInviter_df.to_excel(excel_writer = writer, sheet_name = get_info_name(inviter_info_body) + " (in) (LoL)")
                                                    if search_TFT and len(TFTInviter_index) > 1:
                                                        TFTInviter_df.to_excel(excel_writer = writer, sheet_name = get_info_name(inviter_info_body) + " (in) (TFT)")
                                                    logPrint("邀请者%s曾经与您一同战斗过%d次。\nInviter %s has fought with you for %d time(s)." %(get_info_name(inviter_info_body), len(LoLInviter_index) + len(TFTInviter_index) - 2, get_info_name(inviter_info_body), len(LoLInviter_index) + len(TFTInviter_index) - 2))
                                                break
                                            else:
                                                break
                                recent_friends = list(set(recent_friends)) #被邀请者和邀请者可能重复（Invitees may overlap with inviters）
                                if ("errorCode" in lobbyInvitations or lobby["gameConfig"]["isCustom"] and len(lobbyInvitations) == 0 or not lobby["gameConfig"]["isCustom"] and len(lobbyInvitations) == 1) and len(receivedInvitations) == 0:
                                    logPrint("您尚未发送邀请，也未被邀请。\nYou haven't sent any invitation or been invited by anyone.")
                                else:
                                    if invitee_count == 0 and inviter_count == 0:
                                        logPrint("您近期未曾和您邀请的玩家或者邀请您的玩家一起战斗过。尝试拓展人缘吧！\nYou haven't fought with any inviter or invitee. Try extending your friendship.")
                                    else:
                                        logPrint("\n近期一起玩过的被邀请者信息：\nRecently played invitee information:")
                                        if invitee_count > 0:
                                            if search_LoL:
                                                print(format_df(LoLInvitee_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(LoLInvitee_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                            if search_LoL and search_TFT:
                                                logPrint()
                                            if search_TFT:
                                                print(format_df(TFTInvitee_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(TFTInvitee_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                        logPrint("近期一起玩过的邀请者信息：\nRecently played inviter information:")
                                        if inviter_count > 0:
                                            if search_LoL:
                                                print(format_df(LoLInvitee_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(LoLInvitee_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                            if search_LoL and search_TFT:
                                                logPrint()
                                            if search_TFT:
                                                print(format_df(TFTInvitee_df_to_print, print_index = True, reserve_index = True)[0])
                                                log.write(format_df(TFTInvitee_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                        if invitee_count > 0:
                                            if invitee_count == 1:
                                                logPrint('''一名您邀请的玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's an invitee present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                            else:
                                                logPrint('''%d名您邀请的玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d invitees present in your past matches. Please check the workbook "%s" in the main directory.''' %(invitee_count, excel_name, invitee_count, excel_name))
                                        if inviter_count > 0:
                                            if inviter_count == 1:
                                                logPrint('''一名向您发起邀请的玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's an inviter present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                            else:
                                                logPrint('''%d名向您发起邀请的玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d inviters present in your past matches. Please check the workbook "%s" in the main directory.''' %(inviter_count, excel_name, inviter_count, excel_name))
                                        if len(recent_friends) == 1:
                                            logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s is your friend." %(recent_friends[0], recent_friends[0]))
                                        elif len(recent_friends) > 1:
                                            logPrint("以上玩家中，%s是您的好友。\nAmong the above players, %s are your friends." %("、".join(recent_friends), ", ".join(recent_friends)))
                            elif detect_scene == "5":
                                recent_blockedPlayer_count = 0
                                recent_LoLBlockedPlayer_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                recent_TFTBlockedPlayer_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                blockList = await (await connection.request("GET", "/lol-chat/v1/blocked-players")).json()
                                excel_name = "Recently Played Summoners in Block List of %s - %s.xlsx" %(current_summonerName, platformId)
                                for blockedPlayer in blockList:
                                    blockedPlayer_summonerName = blockedPlayer["name"] if blockedPlayer["gameName"] == "" and blockedPlayer["gameTag"] == "" else blockedPlayer["gameName"] + "#" + blockedPlayer["gameTag"]
                                    LoLBlockedPlayer_index = [0]
                                    TFTBlockedPlayer_index = [0]
                                    if search_LoL:
                                        for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                            if recent_LoLPlayers_df.at[i, "puuid"] == blockedPlayer["puuid"]:
                                                LoLBlockedPlayer_index.append(i)
                                    if search_TFT:
                                        for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                            if recent_TFTPlayers_df.at[i, "puuid"] == blockedPlayer["puuid"]:
                                                TFTBlockedPlayer_index.append(i)
                                    if len(LoLBlockedPlayer_index) + len(TFTBlockedPlayer_index) > 2:
                                        recent_blockedPlayer_count += 1
                                        recent_LoLBlockedPlayer_df = recent_LoLPlayers_df.loc[LoLBlockedPlayer_index, :]
                                        recent_LoLBlockedPlayer_df_to_print = pandas.concat([recent_LoLBlockedPlayer_df_to_print, recent_LoLBlockedPlayer_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                        recent_TFTBlockedPlayer_df = recent_TFTPlayers_df.loc[TFTBlockedPlayer_index, :]
                                        recent_TFTBlockedPlayer_df_to_print = pandas.concat([recent_TFTBlockedPlayer_df_to_print, recent_TFTBlockedPlayer_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                        while True:
                                            try:
                                                with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                    if search_LoL and len(LoLBlockedPlayer_index) > 1:
                                                        recent_LoLBlockedPlayer_df.to_excel(excel_writer = writer, sheet_name = blockedPlayer_summonerName + " (LoL)")
                                                    if search_TFT and len(TFTBlockedPlayer_index) > 1:
                                                        recent_TFTBlockedPlayer_df.to_excel(excel_writer = writer, sheet_name = blockedPlayer_summonerName + " (TFT)")
                                                    logPrint("黑名单玩家%s曾经与您一同战斗过%d次。\nThe blocked player %s has fought with you for %d time(s)." %(blockedPlayer_summonerName, len(LoLBlockedPlayer_index) + len(TFTBlockedPlayer_index) - 2, blockedPlayer_summonerName, len(LoLBlockedPlayer_index) + len(TFTBlockedPlayer_index) - 2))
                                            except PermissionError:
                                                logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                logInput()
                                            except FileNotFoundError:
                                                with pandas.ExcelWriter(path = excel_name) as writer:
                                                    if search_LoL and len(LoLBlockedPlayer_index) > 1:
                                                        recent_LoLBlockedPlayer_df.to_excel(excel_writer = writer, sheet_name = blockedPlayer_summonerName + " (LoL)")
                                                    if search_TFT and len(TFTBlockedPlayer_index) > 1:
                                                        recent_TFTBlockedPlayer_df.to_excel(excel_writer = writer, sheet_name = blockedPlayer_summonerName + " (TFT)")
                                                    logPrint("黑名单玩家%s曾经与您一同战斗过%d次。\nThe blocked player %s has fought with you for %d time(s)." %(blockedPlayer_summonerName, len(LoLBlockedPlayer_index) + len(TFTBlockedPlayer_index) - 2, blockedPlayer_summonerName, len(LoLBlockedPlayer_index) + len(TFTBlockedPlayer_index) - 2))
                                                break
                                            else:
                                                break
                                if len(blockList) == 0:
                                    logPrint("您尚未拉黑过人。恭喜！\nYou haven't blocked any friend. Congratulations!")
                                elif recent_blockedPlayer_count == 0:
                                    logPrint("您近期还没有和任何黑名单玩家一起玩过。\nYou haven't played with any blocked player recently.")
                                else:
                                    logPrint()
                                    if search_LoL:
                                        print(format_df(recent_LoLBlockedPlayer_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_LoLBlockedPlayer_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if search_LoL and search_TFT:
                                        logPrint()
                                    if search_TFT:
                                        print(format_df(recent_TFTBlockedPlayer_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_TFTBlockedPlayer_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if recent_blockedPlayer_count == 1:
                                        logPrint('''一名黑名单玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's a blocked player present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                    else:
                                        logPrint('''%d名黑名单玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d blocked players present in your past matches. Please check the workbook "%s" in the main directory.''' %(recent_blockedPlayer_count, excel_name, recent_blockedPlayer_count, excel_name))
                            elif detect_scene == "6":
                                logPrint('请输入一个由召唤师名称或玩家通用唯一识别码组成的列表。注意列表的每个元素都必须用半角引号括起来。示例：\nPlease input a list of summoner names or puuids. Note that each element of the list must be quoted with English quotation marks. Examples:\n["丿丶莫言丶丶丶", "WordlessMeteor", "沈黙の流れ星"]\n["d7669616-971c-53b1-a19e-570340d825dd", "671e9989-4165-59b6-8d3b-46c9090791a7", "60a6db11-8ff4-5eb4-b6fa-360e6e0eb8fc"]')
                                while True:
                                    try:
                                        summoners = logInput()
                                        if summoners == "":
                                            continue
                                        elif summoners[0] == "0":
                                            break
                                        else:
                                            summoners = eval(summoners)
                                    except SyntaxError:
                                        traceback_info = traceback.format_exc()
                                        logPrint(traceback_info)
                                        logPrint("语法错误！请重新输入。\nGrammar error! Please try again.")
                                    else:
                                        if not isinstance(summoners, list):
                                            logPrint("请输入一个列表！\nPlease input a list!")
                                        elif not all(map(lambda x: isinstance(x, str), summoners)):
                                            logPrint("请输入一个元素全为字符串的列表！\nPlease input a list consisting of only string elements.")
                                        else:
                                            break
                                if isinstance(summoners, str) and summoners[0] == "0":
                                    continue
                                recent_players_count = 0
                                recent_LoLPlayer_df_to_print = pandas.DataFrame(data = recent_LoLPlayer_dict_to_print)
                                recent_TFTPlayer_df_to_print = pandas.DataFrame(data = recent_TFTPlayer_dict_to_print)
                                excel_name = "Recently Played Summoners in Specified Player List - %s.xlsx" %platformId
                                logPrint("是否呈现非法召唤师名称警告？（输入任意键呈现，否则不呈现。）\nDo you want to display illegal summoner name warning? (Input anything to display the warnings, or null to stop displaying.)")
                                illegal_name_warning_str = logInput()
                                illegal_name_warning = bool(illegal_name_warning_str)
                                legal_summoners = {}
                                for summoner in summoners:
                                    info_check = await get_info(connection, summoner)
                                    if info_check["info_got"]:
                                        info_check_body = info_check["body"]
                                        infos[info_check_body["puuid"]] = info_check_body
                                        legal_summoners[info_check_body["puuid"]] = get_info_name(info_check_body)
                                        LoLPlayer_index = [0]
                                        TFTPlayer_index = [0]
                                        if search_LoL:
                                            for i in range(len(recent_LoLPlayers_df.loc[:, "puuid"])):
                                                if recent_LoLPlayers_df.at[i, "puuid"] == info_check_body["puuid"]:
                                                    LoLPlayer_index.append(i)
                                        if search_TFT:
                                            for i in range(len(recent_TFTPlayers_df.loc[:, "puuid"])):
                                                if recent_TFTPlayers_df.at[i, "puuid"] == info_check_body["puuid"]:
                                                    TFTPlayer_index.append(i)
                                        if len(LoLPlayer_index) + len(TFTPlayer_index) > 2:
                                            recent_players_count += 1
                                            recent_LoLPlayer_df = recent_LoLPlayers_df.loc[LoLPlayer_index, :]
                                            recent_LoLPlayer_df_to_print = pandas.concat([recent_LoLPlayer_df_to_print, recent_LoLPlayer_df.loc[1:, recent_LoLPlayer_fields]], axis = 0)
                                            recent_TFTPlayer_df = recent_TFTPlayers_df.loc[TFTPlayer_index, :]
                                            recent_TFTPlayer_df_to_print = pandas.concat([recent_TFTPlayer_df_to_print, recent_TFTPlayer_df.loc[1:, recent_TFTPlayer_fields]], axis = 0)
                                            while True:
                                                try:
                                                    with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                                                        if search_LoL and len(LoLPlayer_index) > 1:
                                                            recent_LoLPlayer_df.to_excel(excel_writer = writer, sheet_name = info_check_body["displayName"] + " (LoL)")
                                                        if search_TFT and len(TFTPlayer_index) > 1:
                                                            recent_TFTPlayer_df.to_excel(excel_writer = writer, sheet_name = info_check_body["displayName"] + " (TFT)")
                                                        logPrint("玩家%s曾经与您一同战斗过%d次。\nPlayer %s has fought with you for %d time(s)." %(info_check_body["displayName"], len(LoLPlayer_index) + len(TFTPlayer_index) - 2, info_check_body["displayName"], len(LoLPlayer_index) + len(TFTPlayer_index) - 2))
                                                except PermissionError:
                                                    logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                                                    logInput()
                                                except FileNotFoundError:
                                                    with pandas.ExcelWriter(path = excel_name) as writer:
                                                        if search_LoL and len(LoLPlayer_index) > 1:
                                                            recent_LoLPlayer_df.to_excel(excel_writer = writer, sheet_name = info_check_body["displayName"] + " (LoL)")
                                                        if search_TFT and len(TFTPlayer_index) > 1:
                                                            recent_TFTPlayer_df.to_excel(excel_writer = writer, sheet_name = info_check_body["displayName"] + " (TFT)")
                                                        logPrint("玩家%s曾经与您一同战斗过%d次。\nPlayer %s has fought with you for %d time(s)." %(info_check_body["displayName"], len(LoLPlayer_index) + len(TFTPlayer_index) - 2, info_check_body["displayName"], len(LoLPlayer_index) + len(TFTPlayer_index) - 2))
                                                    break
                                                else:
                                                    break
                                    elif illegal_name_warning:
                                        logPrint(info_check["message"])
                                logPrint("检测到%d名玩家：\nDetected %d players:" %(len(legal_summoners), len(legal_summoners)))
                                logPrint(pandas.DataFrame({"puuid": legal_summoners.keys(), "summonerName": legal_summoners.values()}), write_time = False)
                                if recent_players_count == 0:
                                    logPrint("未从以上玩家中检测到近期一起玩过的玩家。\nNo players detected in the above summoner list.")
                                else:
                                    logPrint()
                                    if search_LoL:
                                        print(format_df(recent_LoLPlayer_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_LoLPlayer_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if search_LoL and search_TFT:
                                        logPrint()
                                    if search_TFT:
                                        print(format_df(recent_TFTPlayer_df_to_print, print_index = True, reserve_index = True)[0])
                                        log.write(format_df(recent_TFTPlayer_df_to_print, width_exceed_ask = False, direct_print = False, print_index = True, reserve_index = True)[0] + "\n")
                                    if recent_players_count == 1:
                                        logPrint('''一名玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere's a player present in your past matches. Please check the workbook "%s" in the main directory.''' %(excel_name, excel_name))
                                    else:
                                        logPrint('''%d名玩家曾经出现在您的历史对局中。请查看主目录下的“%s”文件。\nThere're %d players present in your past matches. Please check the workbook "%s" in the main directory.''' %(recent_players_count, excel_name, recent_players_count, excel_name))
                            logPrint('是否更新数据？（输入任意键以返回上一层更新对局记录信息，否则在不更新对局信息的情况下再次查询近期一起玩过的玩家）\nUpdate data? (Submit any non-empty string to update match history information, otherwise check the recently played summoners again without updating match history)')
                            update_str = logInput()
                            update = bool(update_str)
                            if update:
                                break
            #with open("infos.json", "w", encoding = "utf-8") as fp:
                #json.dump(infos, fp, indent = 4, ensure_ascii = False)
            if detectMode:
                logPrint("是否从检测模式切换到生成模式？（输入任意键切换，否则不切换）\nDo you want to switch from Detect Mode to Generate Mode? (Submit anything to switch, or null to refuse switching)")
            else:
                logPrint("是否从生成模式切换到检测模式？（输入任意键切换，否则不切换）\nDo you want to switch from Detect Mode to Generate Mode? (Submit anything to switch, or null to refuse switching)")
            switch_mode_str = logInput()
            switch_mode = bool(switch_mode_str)

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection):
    await get_summoner_data(connection)
    await search_recent_players(connection)
    log.write("\n[Program terminated and returned status 0.]\n")
    log.close()

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

connector.start()
