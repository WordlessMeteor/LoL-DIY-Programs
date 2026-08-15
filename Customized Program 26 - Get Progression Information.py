from lcu_driver import Connector
from lcu_driver.connection import Connection
import argparse, os, pandas, re, time
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.core.config.headers import progression_header, progression_counter_header, progression_milestone_header, progression_repeatTrigger_header
from src.core.config.localization import counterDirections, milestone_triggerRequirements, milestoneSizes, milestoneTriggerTypes
from src.core.config.servers import save_platform_info, set_summonerInfo_folder
from src.utils.summoner import get_info_name, print_summoner_info
from src.utils.excel_workbook import create_workbook_win32, sort_worksheet
from src.utils.format import addDefaultStyle
from typing import Any

parser: argparse.ArgumentParser = argparse.ArgumentParser(formatter_class = argparse.RawTextHelpFormatter)
parser.add_argument("-f", "--fill", help = "激活父级元素填充。\nEnable parent element fill.", action = "store_true")
args: argparse.Namespace = parser.parse_args()

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2026/08/15
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

connector: Connector = Connector()

#-----------------------------------------------------------------------------
# 整理进度信息（Organize progression information）
#-----------------------------------------------------------------------------
async def organize_progression_information(connection: Connection) -> None:
    '''
    将进度信息整理成数据框并导出到Excel工作簿中。<br>Organize progression information into dataframes and export them into an Excel workbook.
    
    调用以下接口：<br>The following endpoint is called:
    - GET /lol-progression/v1/groups/configuration
    
    生成以下工作表：<br>The following worksheets are generated:
    - 基础信息（Basic info）
    - 计数器（Counter）
    - 里程碑（Milestone）
    - 重复触发（Repeat trigger）
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    '''
    current_party: dict[str, Any] = await (await connection.request("GET", "/lol-lobby/v1/parties/player")).json()
    platformId: str = current_party["platformId"]
    riot_client_info: list[str] = await (await connection.request("GET", "/riotclient/command-line-args")).json()
    client_info: dict[str, str] = {}
    for i in range(len(riot_client_info)):
        try:
            client_info[riot_client_info[i].split("=")[0]] = riot_client_info[i].split("=")[1]
        except IndexError:
            pass
    region: str = client_info["--region"]
    common_data: dict[str, Any] = await (await connection.request("GET", "/telemetry/v1/common-data")).json()
    version: str = common_data["common.application_version"]
    #设置输出路径（Set the output directory）
    current_info: dict[str, Any] = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    displayName: str = get_info_name(current_info)
    folder: str = set_summonerInfo_folder(region, platformId, current_info)
    #初始化数据结构（Initialize data structures）
    ##基础信息（Basic info）
    progression_header_keys: list[str] = list(progression_header.keys())
    progression_data: dict[str, list[Any]] = {key: [] for key in progression_header_keys}
    ##计数器（Counter）
    progression_counter_header_keys: list[str] = list(progression_counter_header.keys())
    progression_counter_data: dict[str, list[Any]] = {key: [] for key in progression_counter_header_keys}
    ##里程碑（Milestone）
    progression_milestone_header_keys: list[str] = list(progression_milestone_header.keys())
    progression_milestone_data: dict[str, list[Any]] = {key: [] for key in progression_milestone_header_keys}
    ##重复触发（Repeat trigger）
    progression_repeatTrigger_header_keys: list[str] = list(progression_repeatTrigger_header.keys())
    progression_repeatTrigger_data: dict[str, list[Any]] = {key: [] for key in progression_repeatTrigger_header_keys}
    #准备数据（Organize data）
    print("正在加载所有进度信息……\nLoading all progression information ...")
    progressions: list[dict[str, Any]] = await (await connection.request("GET", "/lol-progression/v1/groups/configuration")).json()
    print("正在加载所有进度实例信息……\nLoading all progression instance information ...")
    progression_instanceData: dict[str, dict[str, Any]] = {}
    for i in range(len(progressions)):
        groupId: str = progressions[i]["id"]
        print("[%d/%d]%s\t%s" %(i + 1, len(progressions), groupId, progressions[i]["name"]), end = "\r")
        instance: dict[str, Any] = await (await connection.request("GET", f"/lol-progression/v1/groups/{groupId}/instanceData")).json()
        if not "errorCode" in instance:
            progression_instanceData[groupId] = instance
    #准备一些常量字典（Prepare some constant dictionaries）
    counter_map: dict[str, str] = {counter["id"]: counter["name"] for progression in progressions for counter in progression["counters"]}
    #整理数据（Organize data）
    print("正在整理数据……\nOrganizing data ...")
    for progression_index in range(len(progressions)):
        progression: dict[str, Any] = progressions[progression_index]
        # print("[%d/%d]%s\t%s" %(progression_index + 1, len(progressions), progression["id"], progression["name"]), end = "\r")
        instance_got: bool = progression["id"] in progression_instanceData
        instance: dict[str, Any] = progression_instanceData.get(progression["id"], {})
        ##基础信息（Basic info）
        for i in range(len(progression_header_keys)):
            key: str = progression_header_keys[i]
            if i == 0: #序号（`index`）
                to_append: Any = progression_index + 1
            elif i <= 4:
                to_append = progression[key]
            else: #重复子键（`repeat`'s subkeys）
                to_append = progression["repeat"][key.split()[1]]
            progression_data[key].append(to_append)
        ##计数器（Counter）
        instance_counters: dict[str, dict[str, Any]] = {counter["counterId"]: counter for counter in instance["counters"]} if instance_got else {}
        for counter_index in range(len(progression["counters"])):
            counter: dict[str, Any] = progression["counters"][counter_index]
            instance_counter_got: bool = counter["id"] in instance_counters
            instance_counter: dict[str, Any] = instance_counters.get(counter["id"], {})
            for i in range(len(progression_counter_header_keys)):
                key: str = progression_counter_header_keys[i]
                if i <= 2: #进度相关键（Progression-related keys）
                    if counter_index == 0 or args.fill: #基础信息只在同一个进度中追加一次（Basic information is only appended once per progression）
                        if i == 0: #进度序号（`progression_index`）
                            to_append: Any = progression_index + 1
                        else:
                            to_append = progression[key.split("_")[1]]
                    else:
                        to_append = ""
                elif i == 3: #计数器序号（`counter_index`）
                    to_append = counter_index + 1
                elif i <= 8:
                    if i == 4: #方向（`direction`）
                        to_append = counterDirections[counter["direction"]]
                    else:
                        to_append = counter[key]
                else: #来自（From）：`/lol-progression/v1/groups/{groupId}/instanceData`
                    if instance_counter_got:
                        to_append = instance_counter[key]
                    else:
                        to_append = ""
                progression_counter_data[key].append(to_append)
        ##里程碑（Milestone）
        milestones: list[dict[str, Any]] = progression["milestones"] + progression["repeat"]["milestones"]
        instance_milestones: dict[str, dict[str, Any]] = {milestone["milestoneId"]: milestone for milestone in instance["milestones"]} if instance_got else {}
        for milestone_index in range(len(milestones)):
            milestone: dict[str, Any] = milestones[milestone_index]
            instance_milestone_got: bool = milestone["id"] in instance_milestones
            instance_milestone: dict[str, Any] = instance_milestones.get(milestone["id"], {})
            for trigger_index in range(max(1, len(milestone["triggers"]))): #即使复杂触发条件序列为空，里程碑信息如有，也要追加一次（Even if the trigger array is empty, if milestone information is there, it should be appended once）
                trigger: dict[str, Any] = {} if len(milestone["triggers"]) == 0 else milestone["triggers"][trigger_index]
                for i in range(len(progression_milestone_header_keys)):
                    key: str = progression_milestone_header_keys[i]
                    if i <= 2: #进度相关键（Progression-related keys）
                        if milestone_index == 0 and trigger_index == 0 or args.fill:
                            if i == 0: #进度序号（`progression_index`）
                                to_append: Any = progression_index + 1
                            else:
                                to_append = progression[key.split("_")[1]]
                        else:
                            to_append = ""
                    elif i <= 23:
                        if trigger_index == 0 or args.fill: #里程碑信息只在一个触发条件中追加一次（Milestone information should be appended only once per trigger）
                            if i <= 15:
                                if i == 3: #里程碑序号（`milestone_index`）
                                    to_append = milestone_index + 1 if milestone_index < len(progression["milestones"]) else milestone_index - len(progression["milestones"]) + 1
                                elif i == 4: #可重复（`isRepeat`）
                                    to_append = "" if milestone_index < len(progression["milestones"]) else "√"
                                elif i <= 10:
                                    if i == 9: #触发要求（`triggerRequirement`）
                                        to_append = milestone_triggerRequirements[milestone["triggerRequirement"]]
                                    elif i == 10: #计数器名称（`counter_name`）
                                        to_append = "" if milestone["counterId"] == "" else counter_map[milestone["counterId"]]
                                    else:
                                        to_append = milestone[key]
                                else: #里程碑属性子键（`properties`' subkeys）
                                    subkey: str = key.split()[1]
                                    if subkey in milestone["properties"]:
                                        if i == 14: #里程碑规格（`properties MILESTONE_SIZE`）
                                            to_append = milestoneSizes[milestone["properties"]["MILESTONE_SIZE"]]
                                        else:
                                            to_append = milestone["properties"][subkey]
                                    else:
                                        to_append = ""
                            else: #来自（From）：`/lol-progression/v1/groups/{groupId}/instanceData`
                                if instance_milestone_got:
                                    if i == 22: #已达成（`triggered`）
                                        to_append = "√" if instance_milestone["triggered"] else ""
                                    else:
                                        to_append = instance_milestone[key]
                                else:
                                    to_append = ""
                        else:
                            to_append = ""
                    else: #复杂触发条件子键（`triggers`' subkeys）
                        if len(milestone["triggers"]) == 0: #不存在触发条件的里程碑的这部分字段追加空值（Fields of milestones that don't have complex triggers are appended with empty values）
                            to_append = ""
                        else:
                            if i == 24: #复杂触发条件序号（`trigger_index`）
                                to_append = trigger_index + 1
                            elif i == 28: #复杂触发类型（`trigger type`）
                                to_append = milestoneTriggerTypes[trigger["type"]]
                            elif i == 29: #监控计数器名称（`trigger counter_name`）
                                to_append = "" if trigger["counterId"] == "" else counter_map[trigger["counterId"]]
                            else:
                                to_append = trigger[key.split()[1]]
                    progression_milestone_data[key].append(to_append)
        ##重复触发（Repeat trigger）
        for repeatTrigger_index in range(len(progression["repeat"]["repeatTriggers"])):
            repeatTrigger: dict[str, Any] = progression["repeat"]["repeatTriggers"][repeatTrigger_index]
            for i in range(len(progression_repeatTrigger_header_keys)):
                key: str = progression_repeatTrigger_header_keys[i]
                if i <= 2: #进度相关键（Progression-related keys）
                    if repeatTrigger_index == 0 or args.fill: #里程碑信息只在一个重复触发条件中追加一次（Milestone information should be appended only once per repeated trigger）
                        if i == 0: #进度序号（`progression_index`）
                            to_append: Any = progression_index + 1
                        else:
                            to_append = progression[key.split("_")[1]]
                    else:
                        to_append = ""
                elif i == 3: #重复触发条件序号（`repeatTrigger_index`）
                    to_append = repeatTrigger_index + 1
                else:
                    if i == 8: #触发类型（`type`）
                        to_append = milestoneTriggerTypes[repeatTrigger["type"]]
                    elif i == 9: #监控计数器名称（`counter_name`）
                        to_append = "" if repeatTrigger["counterId"] == "" else counter_map[repeatTrigger["counterId"]]
                    else:
                        to_append = repeatTrigger[key]
                progression_repeatTrigger_data[key].append(to_append)
    #构建数据框和排序（Build dataframes and sort the keys and values）
    ##基础信息（Basic info）
    progression_statistics_output_order: list[int] = [0, 1, 2, 5, 6, 7, 4]
    progression_data_organized: dict[str, list[Any]] = {progression_header_keys[i]: progression_data[progression_header_keys[i]] for i in progression_statistics_output_order}
    progression_df: pandas.DataFrame = pandas.DataFrame(data = progression_data_organized)
    progression_df = pandas.concat([pandas.DataFrame([progression_header])[progression_df.columns], progression_df], ignore_index = True)
    ##计数器（Counter）
    progression_counter_statistics_output_order: list[int] = [0, 2, 3, 7, 6, 12, 5, 4, 8, 10]
    progression_counter_data_organized: dict[str, list[Any]] = {progression_counter_header_keys[i]: progression_counter_data[progression_counter_header_keys[i]] for i in progression_counter_statistics_output_order}
    progression_counter_df: pandas.DataFrame = pandas.DataFrame(data = progression_counter_data_organized)
    progression_counter_df = pandas.concat([pandas.DataFrame([progression_counter_header])[progression_counter_df.columns], progression_counter_df], ignore_index = True)
    ##里程碑（Milestone）
    progression_milestone_statistics_output_order: list[int] = [0, 2, 3, 5, 10, 8, 4, 20, 7, 16, 19, 9, 21, 22, 23, 11, 12, 13, 14, 15, 24, 28, 25, 29, 26]
    progression_milestone_data_organized: dict[str, list[Any]] = {progression_milestone_header_keys[i]: progression_milestone_data[progression_milestone_header_keys[i]] for i in progression_milestone_statistics_output_order}
    progression_milestone_df: pandas.DataFrame = pandas.DataFrame(data = progression_milestone_data_organized)
    progression_milestone_df = pandas.concat([pandas.DataFrame([progression_milestone_header])[progression_milestone_df.columns], progression_milestone_df], ignore_index = True)
    ##重复触发（Repeat trigger）
    progression_repeatTrigger_statistics_output_order: list[int] = [0, 2, 3, 8, 4, 9, 7, 5, 6]
    progression_repeatTrigger_data_organized: dict[str, list[Any]] = {progression_repeatTrigger_header_keys[i]: progression_repeatTrigger_data[progression_repeatTrigger_header_keys[i]] for i in progression_repeatTrigger_statistics_output_order}
    progression_repeatTrigger_df: pandas.DataFrame = pandas.DataFrame(data = progression_repeatTrigger_data_organized)
    progression_repeatTrigger_df = pandas.concat([pandas.DataFrame([progression_repeatTrigger_header])[progression_repeatTrigger_df.columns], progression_repeatTrigger_df], ignore_index = True)
    #保存文件（Save file）
    print("开始导出到工作簿。\nBegin to export to the workbook.\n")
    excel_name: str = "Progression - %s.xlsx" %displayName
    excel_name_sorted: str = "Progression - %s (sorted).xlsx" %displayName
    currentTime: str = time.strftime("%Y-%m-%d %H-%M", time.localtime(time.time()))
    wbPath: str = os.path.join(folder, excel_name).replace("\\", "/")
    os.makedirs(folder, exist_ok = True)
    if not os.path.exists(wbPath):
        wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath), sheet1_name = f"Info - {currentTime}")
    workbook_exist: bool = os.path.exists(wbPath)
    while True:
        try:
            with (pandas.ExcelWriter(path = wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = wbPath)) as writer:
                addDefaultStyle(progression_df).to_excel(excel_writer = writer, sheet_name = f"Info - {currentTime}")
                addDefaultStyle(progression_counter_df).to_excel(excel_writer = writer, sheet_name = f"Counter - {currentTime}")
                addDefaultStyle(progression_milestone_df).to_excel(excel_writer = writer, sheet_name = f"Milestone - {currentTime}")
                addDefaultStyle(progression_repeatTrigger_df).to_excel(excel_writer = writer, sheet_name = f"RepTrigger - {currentTime}") #全名（Full name）： RepeatTrigger
                for sheet_name in [f"Info - {currentTime}", f"Counter - {currentTime}", f"Milestone - {currentTime}", f"RepTrigger - {currentTime}"]:
                    if sheet_name in writer.sheets:
                        worksheet: Worksheet = writer.sheets[sheet_name]
                        if worksheet.calculate_dimension() != "A1:A1":
                            worksheet.cell(row = 1, column = 1, value = version) #在A1单元格填充数据所在版本（Fill in A0 cell with the data version）
        except PermissionError:
            print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
            input()
        else:
            print('事件通行证信息已保存为“%s”！\nEvent pass information is saved as "%s"!' %(wbPath, wbPath))
            break
    if workbook_exist:
        print("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致工作表顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of worksheet order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
        sort: bool = bool(input())
        if sort:
            print("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
            while True:
                try:
                    wb: Workbook = load_workbook(wbPath)
                except FileNotFoundError:
                    print('进度工作簿读取失败！请确保“%s”文件夹内含有名为“%s”的工作簿。如果需要退出程序，请输入“0”。\nERROR reading the Progression workbook! Please make sure the workbook "%s" is in the folder "%s". If you want to exit the program, please submit "0".' %(folder, excel_name, excel_name, folder))
                    store_reload: str = input()
                    if store_reload == "0":
                        break
                else:
                    sheetnames: list[str] = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                    print("请选择排序方式：\nPlease select an ordering pattern:\n☆1\t时间优先（Time in priority）\n2\t类别优先（Type in priority）")
                    op: str = input()
                    print("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                    date_re: re.Pattern[str] = re.compile(r"\d{4}-\d{2}-\d{2} \d{2}-\d{2}") #设置正则表达式识别日期（Define a regular expression to identify a date pattern）
                    dOrder: list[str] = ["Info", "Counter", "Milestone", "RepTrigger"] #存储数据类型的排列顺序（Store the order of data types）
                    dOrder_type_map: dict[str, int] = {_: dOrder.index(_) for _ in dOrder} #定义数据类型权重字典，用于排序数据类型（Define a data type weight dictionary to order the data types）
                    sheetname_date_list: list[str] = list(map(lambda x: date_re.search(x).group(), sheetnames)) #从工作表名称提取日期信息形成列表（Extract the dates from the sheetnames to form a list）
                    sheetname_type_list: list[str] = list(map(lambda x: x.split()[0], sheetnames)) #从工作表名称提取数据类型信息形成列表（Extract the data types from the sheetnames to form a list）
                    sheetname_type_weight_list: list[int] = list(map(lambda x: dOrder_type_map.get(x, len(dOrder) + 1), sheetname_type_list)) #将数据类型列表转换为数据类型权重列表（Transform the data type list into the data type weight list）
                    sheetname_tmpDf: pandas.DataFrame = pandas.DataFrame(data = [sheetnames, sheetname_date_list, sheetname_type_list, sheetname_type_weight_list]).transpose() #创建一个四列数据框，各列分别是完整工作表名、日期信息、数据类型信息和大区信息（Create a 4-column dataframe whose columns are the complete sheetname, date, data type and platformId）
                    if op == "" or op[0] != "2": #按照时间优先的原则对工作表进行排序，时间相同则商品工作表在前，藏品工作表在后（Sort the sheets by time in priority. If the times are the same, then the store sheet is arranged in front of the collection sheet）
                        sheetnames_sorted: list[str] = sheetname_tmpDf.sort_values(by = [1, 3], ascending = True).iloc[:, 0].tolist() #将工作表名按照第一关键字——日期信息正序排列，第二关键字——数据类型权重正序排列（Order the sheetnames according to the ascending order of the first keyword - date and the ascending order of the second keyword - data type weight）
                    else:
                        sheetnames_sorted: list[str] = sheetname_tmpDf.sort_values(by = [3, 1], ascending = True).iloc[:, 0].tolist() #将工作表名按照第一关键字——数据类型权重正序排列，第二关键字——日期信息正序排列（Order the sheetnames according to the ascending order of the first keyword - data type weight and the ascending order of the second keyword - date）
                    #下面排列所有工作表（The following code arrange all sheets）
                    print("正在排序……\nOrdering ...")
                    sort_worksheet(wb, sheetnames_sorted)
                    print('正在保存中……\nSaving the ordered workbook ...')
                    wb.save(os.path.join(folder, excel_name_sorted))
                    print('排序完成！排好序的工作簿已保存为“%s”。请按任意键退出。\nOrdering finished! The ordered workbook is saved as "%s". Press any key to exit ...\n' %(excel_name_sorted, excel_name_sorted))
                    wb.close()
                    input()
                    break

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection: Connection) -> None:
    await print_summoner_info(connection)
    await save_platform_info(connection)
    await organize_progression_information(connection)

@connector.close
async def disconnect(connection: Connection) -> None:
    print("已从英雄联盟客户端断开连接。\nDisconnected from the League Client.")

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

connector.start()
