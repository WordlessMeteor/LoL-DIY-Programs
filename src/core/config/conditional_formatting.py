import pandas
from openpyxl.styles import Color, numbers, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any

def get_continuous_area(targetFields: list[str], df_fields: list[str], record_count: int, skiprecords: int = 0, transpose: bool = False) -> tuple[list[str], list[tuple[Any, Any]]]:
    '''
    求一个字段列表在一个数据框中的连续区域。<br>Calculate the continuous regions of a list of fields in a dataframe.
    
    :param targetFields: 目标字段。<br>Target fields.
    :type targetFields: list[str]
    :param df_fields: 数据框的所有字段。必须保证导出到工作簿后，第一个字段位于B列，最后一个字段位于最后一列。
    :type df_fields: list[str]
    :param record_count: 总记录数量。<br>Number of records in total.
    :type record_count: int
    :param skiprecords: 要跳过的前几行记录的数量。默认不跳过。<br>Number of the first several records to skip. 0 by default.
    :type skiprecords: int
    :param transpose: 数据框是否转置。默认为假。<br>Whether the dataframe is transposed. False by default.
    :type transpose: bool
    :return: 一个二元组。<br>A two-tuple.

        - 第一个元素是一些连续区域字符串。这样的字符串可以直接输入到微软Excel应用的粘贴按钮下方的名称框中，从而直接选中对应区域。<br>The first element is some continuous regions. Such string can be pasted into the name box under the paste button in Microsoft Excel app, so that the corresponding area can be selected.
        - 第二个元素是一个列表，每个元素是由每个连续区域的起始字段索引和终止字段索引组成的二元组。在转置时，这些索引为行，否则为列。<br>The second element is a list whose elements are two-tuples composed of starting and ending field indices of continuous regions. When the dataframe is transposed, these indices are lines, otherwise rows.
    :rtype: tuple[list[str], list[tuple[str, str]] | list[tuple[int, in]]]
    '''
    #参数校验（Parameter checking）
    missing_elements: list[str] = [field for field in targetFields if not field in df_fields]
    if len(missing_elements) > 0:
        raise ValueError("Fields %s not in dataframe." %(missing_elements))
    if record_count <= 0:
        raise ValueError(f"Record count must be positive, received {record_count}.")
    if skiprecords <= 0:
        skiprecords = 0
    if skiprecords >= record_count:
        raise ValueError("At least one record should be retained.")
    #初始化临时变量（Initialize temporary variables）
    startIndex: int = 0 #一段连续区域的起始索引（The starting index of a continuous region）
    endIndex: int = 0 #一段连续区域的终止索引（The ending index of a continuous region）
    fieldIndex: int = 0 #中间变量，用于起始索引和终止索引的赋值（An intermediate variable used for assignment of `startIndex` and `endIndex`）
    #函数主体（Function body）
    rangeStrs: list[str] = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
    rangeTuples: list[tuple[Any, Any]] = [] #存储连续区域的字段索引边界（Store the border field index of continuous regions）
    for i in range(len(targetFields)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
        field: str = targetFields[i]
        if i == 0: #遍历起始赋值（Assignment at the start of the traversal）
            startIndex = endIndex = df_fields.index(field) + 2
        else:
            fieldIndex = df_fields.index(field) + 2
            if fieldIndex == endIndex + 1: #如果下一个要添加条件格式的字段索引与上一个要添加条件格式的字段索引差1，那么这两个字段是相邻的，即连贯的（If the index of the current field to add conditional format is greater than the index of the predecessive field to add conditional format by 1, then these two fields are continuous）
                endIndex = fieldIndex
            else: #如果两个字段不相邻，则提取得到上一个连贯的区域（If these two fields aren't continuous, then get the previous continuous area）
                if transpose:
                    startCol_letter: str = get_column_letter(skiprecords + 2)
                    endCol_letter: str = get_column_letter(record_count + 1)
                    rangeStr: str = "%s%d:%s%d" %(startCol_letter, startIndex, endCol_letter, endIndex)
                    rangeTuples.append((startIndex, endIndex))
                else:
                    startCol_letter: str = get_column_letter(startIndex)
                    endCol_letter: str = get_column_letter(endIndex)
                    rangeStr = "%s%d:%s%d" %(startCol_letter, skiprecords + 2, endCol_letter, record_count + 1)
                    rangeTuples.append((startCol_letter, endCol_letter))
                rangeStrs.append(rangeStr)
                startIndex = endIndex = fieldIndex #将区域的起始索引和终止索引设置为当前索引（Set the starting and ending index as the current index）
    if len(targetFields) > 0: #执行完成后，把最后一个连贯区域也加上。此条件判断为异常处理：当目标字段为空列表时，上面的for循环不执行，起始和终止索引都是0，在转置时会形成“X0:Y0”的工作表区域，而这样的区域是不存在的，因为Excel工作表没有第0行。而且如果上面的循环不执行，本来也不应该执行下面的部分（After the for-loop finishes, add the last continuous area. This condition is designed for exception handling: when `targetFields` is an empty list, the above for-loop won't be executed, so that `startIndex` and `endIndex` are both 0. When `transpose` is True, an "X0:Y0" area will be formed, but an Excel worksheet doesn't have a 0th line. Besides, if the above loop isn't executed, this branch shouldn't be executed, either）
        if transpose:
            startCol_letter: str = get_column_letter(skiprecords + 2)
            endCol_letter: str = get_column_letter(record_count + 1)
            rangeStr: str = "%s%d:%s%d" %(startCol_letter, startIndex, endCol_letter, endIndex)
            rangeTuples.append((startIndex, endIndex))
        else:
            startCol_letter: str = get_column_letter(startIndex)
            endCol_letter: str = get_column_letter(endIndex)
            rangeStr = "%s%d:%s%d" %(startCol_letter, skiprecords + 2, endCol_letter, record_count + 1)
            rangeTuples.append((startCol_letter, endCol_letter))
        rangeStrs.append(rangeStr)
    return (rangeStrs, rangeTuples)

#声明：每个函数的命名与对应的表头一一对应（Declaration: Each function's naming obeys the one-to-one correspondence with the dataframe header）
def addFormat_LoLHistory_wb(worksheet: Worksheet, LoLHistory_df: pandas.DataFrame) -> None:
    '''
    为英雄联盟对局记录工作表添加条件格式。包括以下格式：<br>Add conditional formats to a LoL match history sheet. The following formats are involved:
    
    1. “结果”列：公式规则。<br>Column "Result": FormatRule.
    - 胜利/V(ictory): 绿色（Green）
    - 失败/D(efeat): 蓝色（Blue）
    - 被终止/T(erminated): 灰色（Grey）
    2. “名次”和“队伍排名”列：公式规则。<br>Columns "placement" and "subteamPlacement": FormatRule.
    - 1: 橙色（Orange）
    
    :param worksheet: 工作表对象，通过对一个`pandas.ExcelWriter`对象对工作表取下标得到，如`writer["Sheet1"]`。<br>A Worksheet object obtained by subscripting a `pandas.ExcelWriter` object, e.g. `writer["Sheet1"]`.
    :type worksheet: Worksheet.
    :param LoLHistory_df: 英雄联盟对局记录数据框。<br>LoL match history dataframe.
    :type LoLHistory_df: pandas.DataFrame
    '''
    #胜负颜色（Win/Lose color）
    col_idx: int = LoLHistory_df.columns.to_list().index("result") + 2
    col_letter: str = get_column_letter(col_idx)
    rangeStr: str = "%s3:%s%d" %(col_letter, col_letter, len(LoLHistory_df) + 1)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
    col_idx = LoLHistory_df.columns.to_list().index("subteamPlacement") + 2
    col_letter = get_column_letter(col_idx)
    rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(LoLHistory_df) + 1)
    firstPlace_formulaRule_lol: Rule = FormulaRule(formula = ["$%s3=1" %(col_letter)], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)

def addFormat_LoLGame_summary_wb(worksheet: Worksheet, LoLGame_summary_df: pandas.DataFrame, numColorScale_order: int = 5) -> None:
    '''
    为英雄联盟对局概要工作表添加条件格式。包括以下格式：<br>Add conditional formats to a LoL match summary sheet. The following formats are involved:
    
    1. 百分比列：<br>Percentage columns:
    - 保留两位小数。<br>Keep two decimal places.
    - 数据条，从0到1。<br>Data bar, from 0 to 1.
    2. “战损比”列：保留一位小数。<br>Column "KDA": Keep one decimal place.
    3. “分均补刀”“伤害转化率”和“分均经济”列：保留三位小数。<br>Columns "CSPM" "D/G" and "GPM": Keep three decimal places.
    4. 位次列：三色刻度。从1到该列最大值，颜色由红色逐渐变为绿色。<br>Order columns: Three-color scale. From 1 to the maximum value of this column, the color gradually changes from red to green.
    5. “胜负”列：公式规则。<br>Column "win/lose": FormatRule.
    - 胜利/V(ictory): 绿色（Green）
    - 失败/D(efeat): 蓝色（Blue）
    - 被终止/T(erminated): 灰色（Grey）
    6. “名次”和“队伍排名”列：公式规则。<br>Columns "placement" and "subteamPlacement": FormatRule.
    - 1: 橙色（Orange）
    
    :param worksheet: 工作表对象，通过对一个`pandas.ExcelWriter`对象对工作表取下标得到，如`writer["Sheet1"]`。<br>A Worksheet object obtained by subscripting a `pandas.ExcelWriter` object, e.g. `writer["Sheet1"]`.
    :type worksheet: Worksheet
    :param LoLGame_summary_df: 英雄联盟对局概要数据框。<br>LoL match summary dataframe.
    :type LoLGame_summary_df: pandas.DataFrame
    :param numColorScale_order: 三色刻度的颜色位阶数量。等于单队最大玩家数量。默认为5。<br>The number of color levels for three-color scale. Equal to the maximum number of players per team. 5 by default.
    :type numColorScale_order: int
    '''
    #准备一些常量（Prepare some constants）
    df_columns: list[str] = LoLGame_summary_df.columns.to_list()
    #定义条件格式（Define the conditional formats）
    twoDigitPercentage_columns_lol: list[str] = [column for column in df_columns if column.endswith("_percent") or column == "GUE"] #百分比（Percentage）
    oneDigitFloat_columns_lol: list[str] = [column for column in df_columns if column == "KDA"] #一位小数（One-digit float）
    threeDigitFloat_columns_lol: list[str] = [column for column in df_columns if column in {"CSPM", "D/G", "GPM"}] #三位小数（Three-digit float）
    colorScale_columns_lol: list[str] = [column for column in df_columns if column.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
    dataBar_columns_lol: list[str] = [column for column in df_columns if column.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
    firstPlace_columns_lol: list[str] = [column for column in df_columns if column in {"placement", "subteamPlacement"}] #条件格式——公式（Conditional formatting - formula）
    order_colorScaleRule_lol: Rule = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = numColorScale_order, end_color = "FF6B6B") #跳过名次为0的单元格（Skip the order cells whose values are 0）
    percent_dataBarRule_lol: Rule = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
    #套用保留两位小数的百分比格式（Two-digit percentage）
    for column in twoDigitPercentage_columns_lol:
        col_idx: int = df_columns.index(column) + 2
        for row in range(3, len(LoLGame_summary_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
    #套用一位小数（One-digit float）
    for column in oneDigitFloat_columns_lol:
        col_idx = df_columns.index(column) + 2
        for row in range(3, len(LoLGame_summary_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.0"
    #套用三位小数（Three-digit float）
    for column in threeDigitFloat_columns_lol:
        col_idx = df_columns.index(column) + 2
        for row in range(3, len(LoLGame_summary_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.000"
    #胜负颜色（Win/Lose color）
    col_idx = df_columns.index("win/lose") + 2
    col_letter: str = get_column_letter(col_idx)
    rangeStr: str = "%s3:%s%d" %(col_letter, col_letter, len(LoLGame_summary_df) + 1)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #百分比颜色（Percent color）
    rangeStrs, rangeTuples = get_continuous_area(dataBar_columns_lol, df_columns, len(LoLGame_summary_df), skiprecords = 1)
    for rangeStr in rangeStrs:
        worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
    #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
    rangeStrs, rangeTuples = get_continuous_area(firstPlace_columns_lol, df_columns, len(LoLGame_summary_df), skiprecords = 1)
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple: tuple[str, str] = rangeTuples[i]
        firstPlace_formulaRule_lol: Rule = FormulaRule(formula = ["$%s3=1" %(rangeTuple[0])], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
        worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)
    #位次颜色（Order color）
    rangeStrs, rangeTuples = get_continuous_area(colorScale_columns_lol, df_columns, len(LoLGame_summary_df), skiprecords = 1)
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple = rangeTuples[i]
        order_noFillRule: Rule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
        worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
        worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)

def addFormat_LoLGame_summary_wb_transpose(worksheet: Worksheet, LoLGame_summary_df: pandas.DataFrame, numColorScale_order: int = 5) -> None:
    '''
    为转置后的英雄联盟对局概要工作表添加条件格式。包括以下格式：<br>Add conditional formats to a transposed LoL match summary sheet. The following formats are involved:
    
    1. 百分比列：<br>Percentage columns:
    - 保留两位小数。<br>Keep two decimal places.
    - 数据条，从0到1。<br>Data bar, from 0 to 1.
    2. “战损比”列：保留一位小数。<br>Column "KDA": Keep one decimal place.
    3. “分均补刀”“伤害转化率”和“分均经济”列：保留三位小数。<br>Columns "CSPM" "D/G" and "GPM": Keep three decimal places.
    4. 位次列：三色刻度。从1到该列最大值，颜色由红色逐渐变为绿色。<br>Order columns: Three-color scale. From 1 to the maximum value of this column, the color gradually changes from red to green.
    5. “胜负”列：公式规则。<br>Column "win/lose": FormatRule.
    - 胜利/V(ictory): 绿色（Green）
    - 失败/D(efeat): 蓝色（Blue）
    - 被终止/T(erminated): 灰色（Grey）
    6. “名次”和“队伍排名”列：公式规则。<br>Columns "placement" and "subteamPlacement": FormatRule.
    - 1: 橙色（Orange）
    
    警告：在一次性导出多场对局概要时，添加条件格式可能显著降低导出速度，增大工作簿体积。<br>Warning: When a lot of match summary sheets are going to be exported, adding conditional formats may result in a significant slow of export speed and a significant increase of the workbook size.
    
    :param worksheet: 工作表对象，通过对一个`pandas.ExcelWriter`对象对工作表取下标得到，如`writer["Sheet1"]`。<br>A Worksheet object obtained by subscripting a `pandas.ExcelWriter` object, e.g. `writer["Sheet1"]`.
    :type worksheet: Worksheet
    :param LoLGame_summary_df: 英雄联盟对局概要数据框。<br>LoL match summary dataframe.
    :type LoLGame_summary_df: pandas.DataFrame
    :param numColorScale_order: 三色刻度的颜色位阶数量。等于单队最大玩家数量。默认为5。<br>The number of color levels for three-color scale. Equal to the maximum number of players per team. 5 by default.
    :type numColorScale_order: int
    '''
    #准备一些常量（Prepare some constants）
    df_indices: list[str] = LoLGame_summary_df.index.to_list()
    #定义条件格式（Define the conditional formats）
    twoDigitPercentage_rows_lol: list[str] = [row for row in df_indices if row.endswith("_percent") or row == "GUE"] #百分比（Percentage）
    oneDigitFloat_rows_lol: list[str] = [row for row in df_indices if row == "KDA"] #一位小数（One-digit float）
    threeDigitFloat_rows_lol: list[str] = [row for row in df_indices if row in {"CSPM", "D/G", "GPM"}] #三位小数（Three-digit float）
    colorScale_rows_lol: list[str] = [row for row in df_indices if row.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
    dataBar_rows_lol: list[str] = [row for row in df_indices if row.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
    firstPlace_rows_lol: list[str] = [row for row in df_indices if row in {"placement", "subteamPlacement"}] #条件格式——公式（Conditional formatting - formula）
    order_colorScaleRule_lol: Rule = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = numColorScale_order, end_color = "FF6B6B") #跳过名次为0的单元格（Skip the order cells whose values are 0）
    percent_dataBarRule_lol: Rule = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
    #套用保留两位小数的百分比格式（Two-digit percentage）
    for row in twoDigitPercentage_rows_lol:
        row_idx: int = df_indices.index(row) + 2
        for column in range(3, len(LoLGame_summary_df) + 2):
            worksheet.cell(column = column, row = row_idx).number_format = numbers.FORMAT_PERCENTAGE_00
    #套用一位小数（One-digit float）
    for row in oneDigitFloat_rows_lol:
        row_idx = df_indices.index(row) + 2
        for column in range(3, len(LoLGame_summary_df) + 2):
            worksheet.cell(column = column, row = row_idx).number_format = "0.0"
    #套用三位小数（Three-digit float）
    for row in threeDigitFloat_rows_lol:
        row_idx = df_indices.index(row) + 2
        for column in range(3, len(LoLGame_summary_df) + 2):
            worksheet.cell(column = column, row = row_idx).number_format = "0.000"
    #胜负颜色（Win/Lose color）
    row_idx = df_indices.index("win/lose") + 2
    col_letter: str = get_column_letter(len(LoLGame_summary_df) + 1)
    rangeStr: str = "C%d:%s%d" %(row_idx, col_letter, row_idx)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['C$%d="%s"' %(row_idx, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['C$%d="%s"' %(row_idx, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #百分比颜色（Percent color）
    rangeStrs, rangeTuples = get_continuous_area(dataBar_rows_lol, df_indices, LoLGame_summary_df.shape[1], skiprecords = 1, transpose = True)
    for rangeStr in rangeStrs:
        worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
    #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
    rangeStrs, rangeTuples = get_continuous_area(firstPlace_rows_lol, df_indices, LoLGame_summary_df.shape[1], skiprecords = 1, transpose = True)
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple: tuple[int, int] = rangeTuples[i]
        firstPlace_formulaRule_lol: Rule = FormulaRule(formula = ["C$%d=1" %(rangeTuple[0])], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
        worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)
    #位次颜色（Order color）
    rangeStrs, rangeTuples = get_continuous_area(colorScale_rows_lol, df_indices, LoLGame_summary_df.shape[1], skiprecords = 1, transpose = True)
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple = rangeTuples[i]
        order_noFillRule: Rule = FormulaRule(formula = ["C%d=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
        worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
        worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)

def addFormat_LoLPlayer_summary_wb(worksheet: Worksheet, LoLPlayer_summary_df: pandas.DataFrame, numColorScale_order: int = 5) -> None:
    '''
    `addFormat_LoLGame_summary_wb`函数的简化版本，仅用于战绩汇总脚本。包括以下格式：<br>A simplified version of `addFormat_LoLGame_summary_wb` function, only for Customized Program 20. Other schemes are same as `addFormat_LoLGame_summary_wb` function. The following formats are involved:
    
    1. 百分比列：<br>Percentage columns:
    - 保留两位小数。<br>Keep two decimal places.
    - 数据条，从0到1。<br>Data bar, from 0 to 1.
    2. “战损比”列：保留一位小数。<br>Column "KDA": Keep one decimal place.
    3. 位次列：三色刻度。从1到该列最大值，颜色由红色逐渐变为绿色。<br>Order columns: Three-color scale. From 1 to the maximum value of this column, the color gradually changes from red to green.
    4. “胜负”列：公式规则。<br>Column "win/lose": FormatRule.
    - 胜利/V(ictory): 绿色（Green）
    - 失败/D(efeat): 蓝色（Blue）
    - 被终止/T(erminated): 灰色（Grey）
    
    :param worksheet: 工作表对象，通过对一个`pandas.ExcelWriter`对象对工作表取下标得到，如`writer["Sheet1"]`。<br>A Worksheet object obtained by subscripting a `pandas.ExcelWriter` object, e.g. `writer["Sheet1"]`.
    :type worksheet: Worksheet
    :param LoLGame_summary_df: 英雄联盟玩家概要数据框。<br>LoL player summary dataframe.
    :type LoLGame_summary_df: pandas.DataFrame
    :param numColorScale_order: 三色刻度的颜色位阶数量。等于单队最大玩家数量。默认为5。<br>The number of color levels for three-color scale. Equal to the maximum number of players per team. 5 by default.
    :type numColorScale_order: int
    '''
    #准备一些常量（Prepare some constants）
    df_columns: list[str] = LoLPlayer_summary_df.columns.to_list()
    #定义条件格式（Define the conditional formats）
    twoDigitPercentage_columns_lol_summary: list[str] = ["KP_percent"] #百分比（Percentage）
    oneDigitFloat_columns_lol_summary: list[str] = ["KDA"] #一位小数（One-digit float）
    colorScale_columns_lol_summary: list[str] = [column for column in df_columns if column.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
    dataBar_columns_lol_summary: list[str] = [column for column in df_columns if column.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
    firstPlace_columns_lol_summary: list[str] = [column for column in df_columns if column in {"placement", "subteamPlacement"}] #条件格式——公式（Conditional formatting - formula）
    order_colorScaleRule_lol: Rule = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = numColorScale_order, end_color = "FF6B6B") #跳过名次为0的单元格。这里`end_value`的选取可以讨论一下，可以选取所有对局的队列信息中记录的队伍规模的最大值（Skip the order cells whose values are 0. Here the value of `end_value` is worth discussion: it may take the maximum of `numPlayersPerTeam` recorded in the queue data of the corresponding queueIds）
    percent_dataBarRule_lol: Rule = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
    #套用保留两位小数的百分比格式（Two-digit percentage）
    for column in twoDigitPercentage_columns_lol_summary:
        col_idx: int = df_columns.index(column) + 2 #Excel中的第一列（A列）的索引是1，且又是数据框的索引列【The index of the first column (Column A) in Excel is 1, and this column is the index of column of the dataframe）
        for row in range(3, len(LoLPlayer_summary_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
    #套用一位小数（One-digit float）
    for column in oneDigitFloat_columns_lol_summary:
        col_idx = df_columns.index(column) + 2
        for row in range(3, len(LoLPlayer_summary_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.0"
    #胜负颜色（Win/Lose color）
    col_idx = df_columns.index("win/lose") + 2
    col_letter: str = get_column_letter(col_idx)
    rangeStr: str = "%s3:%s%d" %(col_letter, col_letter, len(LoLPlayer_summary_df) + 1)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #百分比颜色（Percent color）
    rangeStrs, rangeTuples = get_continuous_area(dataBar_columns_lol_summary, df_columns, len(LoLPlayer_summary_df), skiprecords = 1)
    for rangeStr in rangeStrs:
        worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
    #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
    rangeStrs, rangeTuples = get_continuous_area(firstPlace_columns_lol_summary, df_columns, len(LoLPlayer_summary_df), skiprecords = 1)
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple: tuple[str, str] = rangeTuples[i]
        firstPlace_formulaRule_lol: Rule = FormulaRule(formula = ["$%s3=1" %(rangeTuple[0])], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
        worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)
    #位次颜色（Order color）
    rangeStrs, rangeTuples = get_continuous_area(colorScale_columns_lol_summary, df_columns, len(LoLPlayer_summary_df), skiprecords = 1)
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple = rangeTuples[i]
        order_noFillRule: Rule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
        worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
        worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)

def addFormat_inGame_allPlayer_wb(worksheet: Worksheet, inGame_allPlayer_df: pandas.DataFrame) -> None:
    '''
    为通过游戏客户端接口整理得到的游戏内玩家信息工作表设置单元格格式。包括以下格式：<br>Set cell format for in-game player information sheet organized from Game Client API. The following formats are involved:
    
    1. “战损比”列：保留一位小数。<br>Column "KDA": Keep one decimal place.
    2. “分均补刀”列：保留三位小数。<br>Column "CSPM": Keep three decimal places.
    
    :param worksheet: 工作表对象，通过对一个`pandas.ExcelWriter`对象对工作表取下标得到，如`writer["Sheet1"]`。<br>A Worksheet object obtained by subscripting a `pandas.ExcelWriter` object, e.g. `writer["Sheet1"]`.
    :type worksheet: Worksheet
    :param LoLGame_summary_df: 游戏内玩家数据框。<br>In-game player dataframe.
    :type LoLGame_summary_df: pandas.DataFrame
    '''
    #定义条件格式（Define the conditional formats）
    oneDigitFloat_columns_lol: list[str] = ["KDA"] #一位小数（One-digit float）
    threeDigitFloat_columns_lol: list[str] = ["CSPM"] #三位小数（Three-digit float）
    #套用一位小数（One-digit float）
    for column in oneDigitFloat_columns_lol:
        col_idx: int = inGame_allPlayer_df.columns.to_list().index(column) + 2
        for row in range(3, len(inGame_allPlayer_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.0"
    #套用三位小数（Three-digit float）
    for column in threeDigitFloat_columns_lol:
        col_idx = inGame_allPlayer_df.columns.to_list().index(column) + 2
        for row in range(3, len(inGame_allPlayer_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.000"
