import os, pywintypes, sys
import win32com.client as win32
from openpyxl import Workbook
from typing import Optional
wd: str = os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")).replace("\\", "/")
os.chdir(wd)
if not wd in sys.path:
    sys.path.append(wd) #确保在“src”文件夹的父级目录运行此代码（Make sure this program is run under the parent folder of the "src" folder）
from src.utils.logger import LogManager

def create_workbook_win32(wbPath: str, excel: Optional[win32.CDispatch] = None, sheet1_name: str = "Sheet1", log: Optional[LogManager] = None) -> bool:
    '''
    使用系统自带的Excel应用来新建工作簿，以保持工作簿的默认字体为西文字体——等线。<br>Use the built-in Excel application to create a workbook, so that the default font of the workbook is of English style: SimHei.
    
    :param wbPath: 新工作簿的**绝对**路径。<br>**Absolute** path of the new workbook.
    :type wbPath: str
    :param excel: Excel应用程序对象。如果未指定，则将在函数内创建一个临时的应用程序对象，并在函数返回时关闭该应用程序。<br>An Excel application object. If unspecified, the function will create a temporary application object and delete it when the function returns.
    :type excel: win32com.client.CDisplatch
    :param sheet1_name: 新工作簿时的第一张工作表的名称。默认为“Sheet1”。<br>The name of the first sheet of the new workbook. "Sheet1" by default.
    :type sheet1_name: str
    :param log: 日志管理对象。如果未指定，则使用传统的输入和打印函数。<br>A LogManager object. If unspecified, traditional `input` and `print` functions will be used instead.
    :type log: LogManager
    :return: 是否创建成功。<br>Whether the workbook is created successfully.
    :rtype: bool
    '''
    if log == None:
        log = LogManager()
    logPrint = log.logPrint
    wbPath = wbPath
    if os.path.exists(wbPath):
        logPrint("目标工作簿已存在。\nTarget workbook already exists.")
        return False
    logPrint("正在通过微软Excel应用新建工作簿……\nCreating a workbook using Microsoft Excel ...\n路径（Path）： %s\n工作表名（Sheet name）： %s" %(wbPath.replace("\\", "/"), sheet1_name))
    useTmpAppObject: bool = excel is None
    if useTmpAppObject:
        try:
            excel = win32.DispatchEx("Excel.Application") #创建一个新的Excel进程（Create a new Excel process）
        except pywintypes.com_error as e:
            logPrint(e)
            if e.args[0] == -2147221005:
                logPrint("未找到Excel应用程序。\nExcel application not found.")
            elif e.args[0] == -2147352567:
                logPrint("路径访问失败。请检查文件夹是否已创建，或者是否有权限访问。\nPath access failure. Please check if the folder has been created, or if access is denied.")
            else:
                logPrint("未知错误。\nUnknown error.")
            return False
        except Exception as e:
            logPrint(e)
            logPrint("创建Excel应用程序对象时出现了一个问题。\nAn error occurred when the program was trying to create an Excel application object.")
            return False
    workbook: win32.CDispatch = excel.Workbooks.Add() #新建工作簿（Create a workbook）
    worksheet: win32.CDispatch = workbook.Worksheets(1) #取第一个工作表（Get the first worksheet）
    try:
        worksheet.name = sheet1_name #重命名第一个工作表（Rename the first worksheet）
    except AttributeError as e:
        logPrint(e)
        logPrint("工作表名称设置失败。请检查名称是否过或者包含非法字符。将使用默认名称。\nSheet name failed to be set. Please check if the name is too long or contains anyu illegal character. The default name will be used.")
    del worksheet #消除工作表变量以解除其与Python运行环境的关联（Eliminate the worksheet variable to unbind it from Python runtime environment）
    workbook.SaveAs(wbPath) #保存工作簿。注意这个方法的参数中的路径分隔符必须是反斜杠，不能是正斜杠（Save the workbook. Note that the path separator in the parameter of this method must be a backslash instead of a slash）
    workbook.Close() #关闭工作簿对象（Close the workbook object）
    del workbook #消除工作簿变量以解除其与Python运行环境的关联（Eliminate the workbook variable to unbind it from Python runtime environment）
    logPrint("已创建工作簿。\nWorkbook created.")
    if useTmpAppObject:
        excel.Quit() #关闭Excel对象（Close the Excel application）
        del excel
    return True

def sort_worksheet(wb: Workbook, sheetnames_sorted: list[str]) -> None:
    '''
    按照一个给定的工作表顺序，整理一个工作簿中的工作表。<br>According to a given sheet order, sort the worksheets in a workbook.
    
    :param wb: 通过`openpyxl.load_workbook`函数打开的一个工作簿对象。<br>A Workbook object opened through `openpyxl.load_workbook` function.
    :type wb: openpyxl.Workbook
    :param sheetnames_sorted: 期望的工作表顺序。<br>Expected sheet order.
    
        只会将该参数包含的工作表放到工作簿的前部。其它工作表按照原来的顺序放在排好序的工作表的后面。<br>Only the sheets contained in this parameter will be put in the front of the workbook. Other workbooks not included will be put behind the sorted sheets following the original order.
    :type sheetnames_sorted: list[str]
    '''
    for i in range(len(sheetnames_sorted)): #排序的思路是每次将一个工作表根据其在原工作表列表中的索引和在顺序工作表列表中的索引的差值进行移动（The main idea of sheets' sorting is to move each sheet according to the difference of the indices between in the original sheet list and in the ordered sheet list）
        sheetnames = wb.sheetnames #因为一次移动可能导致很多其它工作表的位置发生变化，所以必须每次都重新获取工作表列表（Because a moving event may result in location change of many other sheets, the sheet list must be obtained each time）
        sheetname_iter: str = sheetnames_sorted[i] #这里以顺序工作表为迭代器进行遍历，因为顺序工作表是固定不变的（Here the ordered sheet list acts as the iterator to be traversed, for the ordered sheet list is fixed）
        if sheetnames[i] != sheetname_iter:
            preIndex: int = sheetnames.index(sheetname_iter)
            wb.move_sheet(sheetname_iter, i - preIndex) #注意移动距离数应当是排序后的索引减去排序前的索引（Note that the moving offset should be the index in the ordered list subtracted by that in the original list）
        #print("排序进度（Ordering process）：%d/%d\t工作表名称（Sheet name）： %s" %(i + 1, len(sheetnames_sorted), sheetname_iter))
