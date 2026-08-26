from lcu_driver.connection import Connection
import json, os, pandas, requests, time
from urllib.parse import urljoin
from typing import Any
from openpyxl.worksheet.worksheet import Worksheet
from src.utils.patch import get_cdragon_patchList
from src.utils.webRequest import requestUrl
from src.utils.format import format_df, addDefaultStyle
from src.utils.summoner import print_summoner_info
from src.utils.repeatConnect import LCUConnect
from src.utils.excel_workbook import create_workbook_win32
from src.utils.keyControl import isKeyPressed
from src.core.dataframes.champions import test_bot, sort_ddragon_champions, sort_inventory_champions, sort_plugin_champions
from src.localization.general import language_ddragon, language_dict, language_cdragon

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2026/08/27
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

session: requests.Session = requests.Session()

#-----------------------------------------------------------------------------
# 统计英雄数量（Count champions）
#-----------------------------------------------------------------------------
def get_ddragon_champions(locale: str = "zh_CN") -> tuple[dict[int, dict[str, Any]], str]:
    '''
    获取DataDragon数据库上的所有英雄数据，并将其整理成一个字典。<br>Get all champion information from DataDragon database and organize them into a dictionary.
    
    :param locale: 语言文化代码。默认使用简体中文。<br>Language code. Chinese Simplified by default.
    :type locale: str
    :return: 整理后的英雄数据以及所使用的版本。<br>Organize champion data and the patch used.
    
        整理后的英雄数据中，键是英雄序号，值是英雄信息字典。<br>In the organized champion data, each key is a championId, and each value is the champion information dictionary.
    :rtype: tuple[dict[int, dict[str, Any]], str]
    '''
    global session
    patches_url: str = "https://ddragon.leagueoflegends.com/api/versions.json"
    patches_local_default: str = "离线数据（Offline Data）/versions.json"
    print("请输入您想要获取的版本。输入空字符串以获取最新版本英雄信息。\nPlease input the patch you want to search from. Submit an empty string to get the latest champion data.")
    source, status, session = requestUrl("GET", patches_url, session = session)
    if status == 200:
        patches: list[str] = source.json()
    else:
        if status == -1:
            print('版本信息获取超时！正在尝试离线加载数据……\nPatch information capture timeout! Trying loading offline data ...\n请输入版本Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“0”以返回上一层。\nPlease enter the patch Json data file path. Enter an empty string to use the default relative path: "%s". Submit "0" to return to the last step.' %(patches_local_default, patches_local_default))
            while True:
                patches_local: str = input()
                if patches_local == "":
                    patches_local = patches_local_default
                elif patches_local[0] == "0":
                    print("版本信息获取失败！请检查系统网络状况和代理设置。\nPatch information capture failure! Please check the system network condition and proxy configuration.")
                    time.sleep(3)
                    return ({}, "")
                try:
                    with open(patches_local, "r", encoding = "utf-8") as fp:
                        patches = json.load(fp)
                    if isinstance(patches, list) and patches[-1] == "lolpatch_3.7":
                        break
                    else:
                        print("数据格式错误！请选择一个符合DataDragon数据库中记录的版本数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the patch data archived in DataDragon database (%s)!" %(patches_url, patches_url))
                        continue
                except FileNotFoundError:
                    print("未找到文件%s！请输入正确的版本Json数据文件路径！\nFile %s NOT found! Please input a correct patch Json data file path!" %(patches_local, patches_local))
                    continue
                except OSError:
                    print("数据文件名不合法！请输入含有版本信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with patch information.")
                    continue
                except json.decoder.JSONDecodeError:
                    print("数据格式错误！请选择一个符合DataDragon数据库中记录的版本数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the patch data archived in DataDragon database (%s)!" %(patches_url, patches_url))
                    continue
        elif status == 404:
            print("版本信息文件不存在！请联系作者修复程序。\nPatch information resource file not found! Please contact the author and ask for a repair.")
            time.sleep(3)
            return ({}, "")
        else:
            print("版本列表获取异常。\nAn error occurred when the program was trying to fetch the patch list.")
            time.sleep(3)
            return ({}, "")
    print(json.dumps(patches[:-98], ensure_ascii = False))
    while True:
        patch_in_url: str = input()
        if patch_in_url == "":
            patch_in_url = patches[0]
        elif patch_in_url[0] == "0":
            return ({}, "")
        if patch_in_url in patches[:-98]:
            champion_url = "http://ddragon.leagueoflegends.com/cdn/%s/data/%s/championFull.json" %(patch_in_url, locale)
            break
        else:
            print("版本输入有误！请重新输入。\nERROR input of patch! Please try again!")
    champion_local_default = "离线数据（Offline Data）/ddragon/%s/champion.json" %locale
    source, status, session = requestUrl("GET", champion_url, session = session)
    if status == 200:
        LoLChampion: dict[str, Any] = source.json()
    else:
        if status == -1:
            print('英雄数据获取超时！正在尝试离线加载数据……\nChampion data capture timeout! Trying loading offline data ...\n请输入英雄Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“0”以返回上一层。\nPlease enter the champion Json data file path. Enter an empty string to use the default relative path: "%s". Submit "0" to return to the last steo.' %(champion_local_default, champion_local_default))
            while True:
                champion_local: str = input()
                if champion_local == "":
                    champion_local = champion_local_default
                elif champion_local[0] == "0":
                    print("英雄数据获取失败！请检查系统网络状况和代理设置。\nChampion data capture failure! Please check the system network condition and proxy configuration.")
                    time.sleep(3)
                    return ({}, patch_in_url)
                try:
                    with open(champion_local, "r", encoding = "utf-8") as fp:
                        LoLChampion = json.load(fp)
                    if isinstance(LoLChampion, dict) and all(i in LoLChampion for i in ["type", "format", "version", "data"]) and LoLChampion["type"] == "champion" and all(j in LoLChampion["data"][i] for i in LoLChampion["data"] for j in ["id", "key", "name", "title", "image", "skins", "lore", "blurb", "allytips", "enemytips", "tags", "partype", "info", "stats", "speklls", "passive", "recommended"]):
                        break
                    else:
                        print("数据格式错误！请选择一个符合DataDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the champion data archived in DataDragon database (%s)!" %(champion_url, champion_url))
                        continue
                except FileNotFoundError:
                    print("未找到文件%s！请输入正确的英雄Json数据文件路径！\nFile %s NOT found! Please input a correct champion Json data file path!" %(champion_local, champion_local))
                    continue
                except OSError:
                    print("数据文件名不合法！请输入含有英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with champion information.")
                    continue
                except json.decoder.JSONDecodeError:
                    print("数据格式错误！请选择一个符合DataDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the champion data archived in DataDragon database (%s)!" %(champion_url, champion_url))
                    continue
        elif status == 404:
            print("英雄数据文件不存在！请联系作者修复程序。\nChampion data resource file not found! Please contact the author and ask for a repair.")
            time.sleep(3)
            return ({}, patch_in_url)
        else:
            print("英雄列表获取异常。\nAn error occurred when the program was trying to fetch the champion list.")
            time.sleep(3)
            return ({}, patch_in_url)
    LoLChampions: dict[int, dict[str, Any]] = {int(champion["key"]): champion for champion in LoLChampion["data"].values()}
    return (LoLChampions, patch_in_url)

def get_cdragon_champions(locale: str = "zh_CN") -> tuple[dict[int, dict[str, Any]], str]:
    '''
    获取CommunityDragon数据库上的所有英雄数据，并将其整理成一个字典。<br>Get all champion information from CommunityDragon database and organize them into a dictionary.
    
    :param locale: 语言文化代码。默认使用简体中文。<br>Language code. Chinese Simplified by default.
    :type locale: str
    :return: 整理后的英雄数据以及所使用的版本。<br>Organize champion data and the patch used.
    
        整理后的英雄数据中，键是英雄序号，值是英雄信息字典。<br>In the organized champion data, each key is a championId, and each value is the champion information dictionary.
    :rtype: tuple[dict[int, dict[str, Any]], str]
    '''
    global session
    print("请输入您想要获取的版本。输入空字符串以获取最新版本英雄信息。\nPlease input the patch you want to search from. Submit an empty string to get the latest champion data.")
    patches_cdragon, patchList_fetched = get_cdragon_patchList(session = session) #对应于DataDragon数据库的版本，从CommunityDragons数据库主页获取可用版本（Corresponding to getting patches DataDragon database, get the available patches in CommunityDragon database through its homepage）
    print(json.dumps(patches_cdragon, ensure_ascii = False))
    if patchList_fetched:
        while True:
            patch_in_url: str = input()
            if patch_in_url == "":
                patch_in_url = patches_cdragon[1]
            elif patch_in_url[0] == "0":
                return ({}, "")
            if patch_in_url in patches_cdragon:
                champion_folder_url: str = "https://raw.communitydragon.org/json/%s/plugins/rcp-be-lol-game-data/global/%s/v1/champions/" %(patch_in_url, locale.lower())
                patch_url: str = "https://raw.communitydragon.org/json/%s/compat-version-metadata.json" %patch_in_url
                break
            else:
                print("版本输入有误！请重新输入。\nERROR input of patch! Please try again!")
    else:
        time.sleep(3)
        return ({}, "")
    #下面获取版本信息（The following code obtain the patch information）
    source, status, session = requestUrl("GET", patch_url, session = session)
    if status == 200:
        version_dict: dict[str, str] = source.json()
        version: str = version_dict["version"]
    else:
        if status == -1:
            print("版本文件访问失败！\nPatch file access failed!")
        elif status == 404:
            print("版本文件不存在！请联系作者修复程序。\nPatch file not found! Please contact the author and ask for a repair.")
        else:
            print("版本信息获取异常。\nAn error occurred when the program was trying to fetch the patch information.")
        time.sleep(3)
        return ({}, "")
    #下面获取每个英雄的数据资源链接（The following code obtain the data resource url of each champion）
    source, status, session = requestUrl("GET", champion_folder_url, session = session)
    if status == 200:
        champion_folder_json: list[dict[str, Any]] = source.json()
        champion_urls: list[str] = []
        champion_files: dict[int, str] = {}
        for record in champion_folder_json:
            if record["type"] == "file":
                championId: int = int(os.path.splitext(record["name"])[0])
                champion_files[championId] = record["name"]
        for championId in sorted(champion_files.keys()):
            champion_urls.append(urljoin(champion_folder_url, champion_files[championId]))
    else:
        if status == -1:
            print("英雄文件夹访问失败！\nChampion folder access failed!")
        elif status == 404:
            print("英雄文件夹不存在！请联系作者修复程序。\nChampion folder not found! Please contact the author and ask for a repair.")
        else:
            print("英雄文件夹信息获取异常。\nAn error occurred when the program was trying to fetch the champion folder.")
        time.sleep(3)
        return ({}, version)
    champion_local_default: str = "离线数据（Offline Data）/cdragon/pbe/plugins/rcp-be-lol-game-data/global/%s/v1/champions/" %language_cdragon[locale]
    champion_files_ready: bool = False
    LoLChampions_source: list[dict[str, Any]] = []
    #注释以下代码以直接离线加载数据资源（Comment out the following code to load offline data resources directly）
    print("获取进度（Capturing process）：")
    for i in range(len(champion_urls)):
        if isKeyPressed(b"\x1b"):
            print("您已中断此过程。\nYou've interrupted this process.")
            return ({}, version)
        champion_url: str = champion_urls[i]
        source, status, session = requestUrl("GET", champion_url, session = session)
        if status != 200:
            print("英雄信息获取失败！请检查系统网络状况和代理设置。\nChampion information capture failure! Please check the system network condition and proxy configuration.")
            break
        champion: dict[str, Any] = source.json()
        LoLChampions_source.append(champion)
        print("[%s]" %(time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())), end = "")
        print("[%d/%d]%s %s" %(i + 1, len(champion_urls), champion["name"], champion["title"]))
    else:
        champion_files_ready = True #任何一个文件获取失败都会导致程序进入离线加载模式（Any file that failed to be loaded will cause to program to load all data again offline）
        champion_url: str = "" #占位变量，用于处理变量可能不存在的类型检查提示（A dummy variable to handle the "Variable is possibly unbound" type check hint）
    #注释以上代码以直接离线加载数据资源（Comment out the above code to load offline data resources directly）
    if not champion_files_ready:
        print('英雄信息获取失败！正在尝试离线加载数据……\nChampion information capture failure! Trying loading offline data ...\n请输入英雄Json数据文件夹路径。输入空字符以使用默认相对引用路径“%s”。输入“0”以返回上一层。\nPlease enter the champion Json data folder path. Enter an empty string to use the default relative path: "%s". Submit "0" to return to the last step.' %(champion_local_default, champion_local_default))
        while True:
            LoLChampions_source = []
            champion_local: str = input()
            if champion_local == "":
                champion_local = champion_local_default
            elif champion_local[0] == "0":
                print("英雄数据获取失败！请检查系统网络状况和代理设置。\nChampion data capture failure! Please check the system network condition and proxy configuration.")
                time.sleep(3)
                return ({}, version)
            try:
                for championId in sorted(champion_files.keys()):
                    with open(os.path.join(champion_local, champion_files[championId]), "r", encoding = "utf-8") as fp:
                        champion = json.load(fp)
                    if isinstance(champion, dict) and all([i in champion for i in ["id", "name", "alias", "title", "shortBio", "tacticalInfo", "playstyleInfo", "squarePortraitPath", "stingerSfxPath", "chooseVoPath", "banVoPath", "roles", "recommendedItemDefaults", "skins", "passive", "spells"]]) and all(isinstance(i, dict) for i in [champion["tacticalInfo"], champion["playstyleInfo"], champion["passive"]]) and all(i in champion["tacticalInfo"] for i in ["style", "difficulty", "damageType"]) and all(i in champion["playstyleInfo"] for i in ["damage", "durability", "crowdControl", "mobility", "utility"]) and all(i in champion["passive"] for i in ["name", "abilityIconPath", "abilityVideoPath", "abilityVideoImagePath", "description"]) and all(isinstance(i, int) for i in [champion["id"], champion["tacticalInfo"]["style"], champion["tacticalInfo"]["difficulty"], champion["playstyleInfo"]["damage"], champion["playstyleInfo"]["durability"], champion["playstyleInfo"]["crowdControl"], champion["playstyleInfo"]["mobility"], champion["playstyleInfo"]["utility"]]) and all(isinstance(i, str) for i in [champion["name"], champion["alias"], champion["title"], champion["shortBio"], champion["squarePortraitPath"], champion["stingerSfxPath"], champion["chooseVoPath"], champion["banVoPath"], champion["tacticalInfo"]["damageType"], champion["passive"]["name"], champion["passive"]["abilityIconPath"], champion["passive"]["abilityVideoPath"], champion["passive"]["abilityVideoImagePath"], champion["passive"]["description"]]) and all(isinstance(i, list) for i in [champion["roles"], champion["recommendedItemDefaults"], champion["skins"], champion["spells"]]):
                        LoLChampions_source.append(champion)
                    else:
                        print("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the champion data archived in CommunityDragon database (%s)!" %(champion_url, champion_url))
                        break
            except FileNotFoundError:
                print("未找到文件%s！请输入正确的英雄Json数据文件夹路径！\nFile %s NOT found! Please input a correct champion Json data folder path!" %(champion_local, champion_local))
                continue
            except OSError:
                print("数据文件名不合法！请输入含有英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with champion information.")
                continue
            except json.decoder.JSONDecodeError:
                print("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the champion data archived in CommunityDragon database (%s)!" %(champion_url, champion_url))
                continue
            else:
                break
    LoLChampions: dict[int, dict[str, Any]] = {champion["id"]: champion for champion in LoLChampions_source}
    return (LoLChampions, version)

async def get_plugin_champions(connection: Connection) -> list[dict[str, Any]]: #和整理静态英雄数据资源的函数不同，这里返回的是一个列表（What's different from the functions that organize static champion data resources is that this function returns a list）
    '''
    通过LCU API读取插件中的英雄数据，并将其整理成一个列表。<br>Read champion data in the plugin through LCU API and organize them into a list.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    :return: 汇总后的英雄原始数据。<br>Merged raw champion data.
    :rtype: list[dict[str, Any]]
    '''
    champion_summary: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/champion-summary.json")).json()
    championIds: list[int] = list(map(lambda x: x["id"], champion_summary))
    LoLChampions_source: list[dict[str, Any]] = []
    print("获取进度（Capturing process）：")
    for i in range(len(championIds)):
        championId: int = championIds[i]
        champion: dict[str, Any] = await (await connection.request("GET", f"/lol-game-data/assets/v1/champions/{championId}.json")).json() #插件从本地读取，因此一般不需要设置异常处理（Plugins are read locally, so exception handling isn't needed here）
        LoLChampions_source.append(champion)
        print("[%s]" %(time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())), end = "")
        print("[%d/%d]%s %s" %(i + 1, len(championIds), champion["name"], champion["title"]))
    return LoLChampions_source

async def count_champions(connection: Connection) -> None:
    current_summoner: dict[str, Any] = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    recommended_position_for_champion: dict[str, dict[str, Any]] = await (await connection.request("GET", "/lol-perks/v1/recommended-champion-positions")).json()
    common_data: dict[str, Any] = await (await connection.request("GET", "/telemetry/v1/common-data")).json()
    version: str = common_data["common.application_version"]
    print("请选择英雄数据类型：\nPlease a champion data type:\n1\t个人所有（Personal inventory）\n2\t插件（Plugins）")
    while True:
        data_type: str = input()
        if data_type == "":
            continue
        elif data_type[0] == "0":
            return
        elif data_type[0] == "1":
            LoLChampions_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-champions/v1/inventories/%s/champions" %current_summoner["summonerId"])).json()
            break
        elif data_type[0] == "2":
            LoLChampions_source = await get_plugin_champions(connection)
            break
        else:
            print("您的输入有误！请重新输入。\nERROR input! Please try again.")
    print("请选择统计类型：\nPlease select which type of champions to count:\n1\t所有英雄（All champions）\n2\t所有电脑英雄（All bot champions）\n3\t当前房间可用电脑英雄（Available bot champions in this lobby）")
    while True: #分类讨论确定`LoLChampions`（Discuss and determine `LoLChampions`）
        champion_got: bool = False #标记是否获取到英雄数据（Marks whether champion data have been fetched）
        LoLChampions: dict[int, dict[str, Any]] = {} #初始化英雄数据（Initialize champion data）
        mode: str = input()
        if mode == "":
            continue
        elif mode[0] == "0":
            return
        elif mode[0] == "1":
            sheet_name: str = "Sheet3"
            for champion in LoLChampions_source:
                LoLChampions[champion["id"]] = champion
            champion_got = True
        elif mode[0] == "2":
            sheet_name = "Sheet2"
            for champion in LoLChampions_source:
                LoLChampions[champion["id"]] = champion
            LoLChampions, count = await test_bot(connection, LoLChampions, verbose = True)
            champion_got = True
        elif mode[0] == "3":
            sheet_name = "Sheet1"
            lobby_information: dict[str, Any] = await (await connection.request("GET", "/lol-lobby/v2/lobby")).json()
            if "errorCode" in lobby_information and lobby_information["message"] == "LOBBY_NOT_FOUND":
                print("请确保您正在房间内！\nPlease make sure you're in a lobby!")
            else:
                bots_enabled: bool = await (await connection.request("GET", "/lol-lobby/v2/lobby/custom/bots-enabled")).json()
                if not bots_enabled:
                    print("该房间无可用电脑玩家。\nThere're no available bot champions in this lobby.")
                else:
                    available_bots: dict[str, Any] = await (await connection.request("GET", "/lol-lobby/v2/lobby/custom/available-bots")).json()
                    available_botIds: list[int] = list(map(lambda x: x["id"], available_bots))
                    for champion in LoLChampions_source:
                        if champion["id"] in available_botIds:
                            LoLChampions[champion["id"]] = champion
                    champion_got = True
        else:
            print("您的输入有误！请重新输入。\nERROR input! Please try again.")
            continue
        #下面按照程序需求对数据资源进行一定的整理（The following code organize the data resource according to the program's need）
        if champion_got:
            count: int = 0
            if mode[0] == "2":
                print("正在整理数据……\nOrganizing data ...")
            if data_type[0] == "1":
                LoLChampion_df, count = sort_inventory_champions(LoLChampions, recommended_position_for_champion, verbose = mode[0] != "2")
            elif data_type[0] == "2":
                LoLChampion_df, count = sort_plugin_champions(LoLChampions, verbose = mode[0] != "2")
            wbPath: str = "available-bots.xlsx"
            if not os.path.exists(wbPath):
                wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath))
            workbook_exist: bool = os.path.exists(wbPath)
            while True:
                try:
                    with (pandas.ExcelWriter(path = wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = wbPath)) as writer:
                        addDefaultStyle(LoLChampion_df).to_excel(excel_writer = writer, sheet_name = sheet_name)
                        worksheet: Worksheet = writer.sheets[sheet_name]
                        worksheet.cell(row = 1, column = 1, value = version)
                except PermissionError:
                    print("无写入权限！请确保文件未被打开且非只读状态！按回车键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press Enter to try again.")
                    input()
                else:
                    if mode[0] == "1" or mode[0] == "3":
                        print("\n统计完毕，共%d名英雄。按回车键继续。\nCount finished! There're %d champions in total. Press Enter to continue." %(count, count))
                    else:
                        print("英雄数据导出完成！按回车键继续。\nChampion data exported! Press Enter to continue.")
                    break
            input()
        print("请选择统计类型：\nPlease select which type of champions to count:\n1\t所有英雄（All champions）\n2\t所有电脑英雄（All bot champions）\n3\t当前房间可用电脑英雄（Available bot champions in this lobby）")

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
async def connect(connection: Connection) -> None: #注意到这里没有加装饰器（Note that here's not a decorator）
    await print_summoner_info(connection)
    await count_champions(connection)

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------
def main():
    print("请选择离线数据资源的语言【默认为中文（中国）】：\nPlease select a language for offline data resources (the default option is zh_CN):")
    language_df: pandas.DataFrame = pandas.DataFrame(language_dict)
    print(format_df(language_df)[0])
    while True:
        language_specified: bool = False
        language_option: str = input()
        if language_option == "" or language_option in [str(i) for i in range(1, 31)]:
            if language_option == "":
                language_option = "29"
            language_specified = True
            break
        elif language_option[0] == "0":
            break
        else:
            print("语言选项输入错误！\nERROR input of language option!")
    # loop_closed: bool = False #表明事件循环是否已经关闭。如果已经关闭，则同一线程中不能再执行此循环（Represent whether the event loop is closed. If it is, this loop shouldn't be accessed within the same thread）
    if language_specified:
        language_code: str = list(language_ddragon.keys())[int(language_option) - 1]
        print('请选择英雄数据来源：（输入“0”以退出程序。）\nPlease select the champion data source: (Submit "0" to exit.)\n1\tLCU API\n2\tDataDragon\n3\tCommunityDragon')
        while True:
            source: str = input()
            if source == "":
                continue
            elif source[0] == "0":
                break
            elif source[0] == "1":
                LCUConnect(connect)
            elif source[0] == "2":
                LoLChampions, version = get_ddragon_champions(locale = language_code)
                if len(LoLChampions) > 0:
                    LoLChampion_df, count = sort_ddragon_champions(LoLChampions, verbose = True)
                    version_df: pandas.DataFrame = pandas.DataFrame({"Patch": [version]})
                    wbPath: str = "available-bots.xlsx"
                    if not os.path.exists(wbPath):
                        wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath))
                    workbook_exist: bool = os.path.exists(wbPath)
                    while True:
                        try:
                            with (pandas.ExcelWriter(path = "available-bots.xlsx", mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = "available-bots.xlsx")) as writer:
                                addDefaultStyle(LoLChampion_df).to_excel(excel_writer = writer, sheet_name = "Sheet3")
                                worksheet: Worksheet = writer.sheets["Sheet3"]
                                worksheet.cell(row = 1, column = 1, value = version)
                        except PermissionError:
                            print("无写入权限！请确保文件未被打开且非只读状态！按回车键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press Enter to try again.")
                            input()
                        else:
                            print("\n统计完毕，共%d名英雄。按回车键继续。\nCount finished! There're %d champions in total. Press Enter to continue." %(count, count))
                            break
                    input()
            elif source[0] == "3":
                LoLChampions, version = get_cdragon_champions(locale = language_code)
                if len(LoLChampions) > 0:
                    LoLChampion_df, count = sort_plugin_champions(LoLChampions)
                    version_df = pandas.DataFrame({"Patch": [version]})
                    wbPath = "available-bots.xlsx"
                    if not os.path.exists(wbPath):
                        wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath))
                    workbook_exist = os.path.exists(wbPath)
                    while True:
                        try:
                            with (pandas.ExcelWriter(path = "available-bots.xlsx", mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = "available-bots.xlsx")) as writer:
                                addDefaultStyle(LoLChampion_df).to_excel(excel_writer = writer, sheet_name = "Sheet3")
                                worksheet: Worksheet = writer.sheets["Sheet3"]
                                worksheet.cell(row = 1, column = 1, value = version)
                        except PermissionError:
                            print("无写入权限！请确保文件未被打开且非只读状态！按回车键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press Enter to try again.")
                            input()
                        else:
                            print("\n统计完毕，共%d名英雄。按回车键继续。\nCount finished! There're %d champions in total. Press Enter to continue." %(count, count))
                            break
                    input()
            else:
                continue
            print('请选择英雄数据来源：（输入“0”以退出程序。）\nPlease select the champion data source: (Submit "0" to exit.)\n1\tLCU API\n2\tDataDragon\n3\tCommunityDragon')
    return 0

status = main()
