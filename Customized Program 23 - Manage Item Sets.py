from lcu_driver.connection import Connection
import copy, json, os, pandas, re, requests, time, traceback, uuid
from openpyxl import load_workbook, Workbook
from typing import Any, Optional
from src.utils.summoner import print_summoner_info
from src.utils.logger import LogManager
from src.utils.repeatConnect import LCUConnect
from src.utils.format import optimize_bool_display, format_df, addDefaultStyle, format_runtime, verify_uuid, pyobj2json
from src.utils.patch import Patch, get_ddragon_versionList, get_cdragon_patchList
from src.utils.runtimeDebug import subscope
from src.utils.webRequest import requestUrl
from src.utils.excel_workbook import create_workbook_win32, sort_worksheet
from src.core.config.localization import language_ddragon, language_dict, language_cdragon
from src.core.extractor import LoLDataExtractor
from src.core.config.localization import gamemaps
from src.core.dataframes.champions import sort_champion_summary

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2026/07/17
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

spells: dict[int, dict[str, Any]] = {}
LoLChampions: dict[int, dict[str, Any]] = {}
LoLItems: dict[int, dict[str, Any]] = {}

log_folder: str = "日志（Logs）/Customized Program 23 - Manage Item Sets"
os.makedirs(log_folder, exist_ok = True)
currentTime: str = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
log = LogManager(path = os.path.join(log_folder, currentTime + ".log"), mode = "a+", encoding = "utf-8")
logInput = log.logInput
logPrint = log.logPrint

#-----------------------------------------------------------------------------
# 配置配装方案（Configure item sets）
#-----------------------------------------------------------------------------
async def prepare_data_resources(connection: Connection) -> None:
    '''
    准备全局数据资源。<br>Prepare global data resources.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    '''
    logPrint("正在准备数据资源……\nPreparing data resources ...")
    global spells, LoLChampions, LoLItems
    ##召唤师技能（Summoner spell）
    logPrint("正在加载召唤师技能信息……\nLoading summoner spell information ...")
    spells_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/summoner-spells.json")).json()
    spells = {int(spell_iter["id"]): spell_iter for spell_iter in spells_source}
    ##英雄（Champion）
    logPrint("正在加载英雄信息……\nLoading champion information ...")
    LoLChampions_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/champion-summary.json")).json()
    LoLChampions = {int(LoLChampion_iter["id"]): LoLChampion_iter for LoLChampion_iter in LoLChampions_source}
    ##英雄联盟装备（LoL item）
    logPrint("正在加载英雄联盟装备信息……\nLoading LoL item information ...")
    LoLItems_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/items.json")).json()
    LoLItems = {int(LoLItem_iter["id"]): LoLItem_iter for LoLItem_iter in LoLItems_source}

def calculate_totalPrice(item_key: str, items_bin: dict[str, list[str] | dict[str, Any]]) -> int:
    '''
    在给定一个装备主键的情况下，迭代地计算其总价格。<br>While an item key is given, iteratively calculate its total price.
    
    :param item_key: 装备主键。<br>Item key.
    :type item_key: str
    :param items_bin: 装备二进制数据。<br>Item binary data.
    :type items_bin: dict[str, list[str] | dict[str, Any]]
    :return: 装备总价格。<br>The item's total price.
    :rtype: int
    '''
    priceTotal: int = 0
    #下面通过一个栈实现装备总价格的计算（Calculate the total price using a stack）
    recipeItem_key_stack: list[str] = [item_key]
    while len(recipeItem_key_stack) > 0:
        item_key_tmp: str = recipeItem_key_stack.pop()
        itemJson_tmp: dict[str, Any] = items_bin[item_key_tmp]
        priceTotal += itemJson_tmp.get("price", 0) #部分装备没有价格键，例如防御塔装备（Some items don't have the "price" key, e.g. turret items）
        if "recipeItemLinks" in itemJson_tmp:
            recipeItem_key_stack += itemJson_tmp["recipeItemLinks"][::-1] #保证金币的计算遵循正确的装备构件顺序，虽然这其实无关紧要（Make sure the calculation order of total price follows the correct in-game item component order, although it doesn't matter actually）
    return priceTotal

def create_test_itemPage(isZH: bool = True, bilingual: bool = False) -> tuple[dict[str, Any], bool, str]:
    '''
    创建一个包含所有装备的配装方案，并根据游戏模式进行分类。<br>Create an item set that contains all items and classify them according to game modes.
    
    程序首先获取在召唤师峡谷经典模式中的所有可用装备。其它游戏模式的装备区块中的装备对召唤师峡谷经典模式中的装备取差集得到。<br>The program first sorts out all available items in Summoner's Rift Classic mode. For other game modes, each item list is obtained by taking the difference set against that of Summoner's Rift Classic.
    
    :param isZH: 是否使用简体中文配装方案名称和装备区块名称。默认为真。<br>Whether to use item set name and item block name in Chinese Simplified. True by default.
    :type isZH: bool
    :param bilingual: 是否同时使用简体中文和美式英语配装方案名称和装备区块名称。默认为假。<br>Whether to use item set name and item block name in both Chinese Simplified and English US. False by default.<br>当该参数为真时，isZH参数不再生效。<br>When this parameter is True, `isZH` no longer takes effect.
    :return: 配装方案、获取状态和配装方案文件目录。<br>Item set, fetch status and the item set file path.
    :rtype: tuple[dict[str, Any], bool, str]
    '''
    session = requests.Session() #初始化网络请求会话（Initialize the web request session）
    #session.trust_env = False #忽略系统代理设置（Bypass system proxy）
    #从CommunityDragon读取装备相关二进制数据（Load relevant binary data from CommunityDragon）
    logPrint("正在加载地图和装备二进制数据……\nLoading binary data for maps and items from CommunityDragon ...")
    ##在线读取（Online loading）
    map11_bin_url: str = "https://raw.communitydragon.org/pbe/game/data/maps/shipping/map11/map11.bin.json"
    map12_bin_url: str = "https://raw.communitydragon.org/pbe/game/data/maps/shipping/map12/map12.bin.json"
    map21_bin_url: str = "https://raw.communitydragon.org/pbe/game/data/maps/shipping/map21/map21.bin.json"
    map22_bin_url: str = "https://raw.communitydragon.org/pbe/game/data/maps/shipping/map22/map22.bin.json"
    map30_bin_url: str = "https://raw.communitydragon.org/pbe/game/data/maps/shipping/map30/map30.bin.json"
    map33_bin_url: str = "https://raw.communitydragon.org/pbe/game/data/maps/shipping/map33/map33.bin.json"
    map35_bin_url: str = "https://raw.communitydragon.org/pbe/game/data/maps/shipping/map35/map35.bin.json"
    map453_bin_url: str = "https://raw.communitydragon.org/pbe/game/unknown/579b4182be3270f4.bin.json"
    items_bin_url: str = "https://raw.communitydragon.org/pbe/game/items.cdtb.bin.json"
    source, status, session = requestUrl("GET", map11_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("召唤师峡谷地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nSummoner's Rift map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map11_bin_url))
            map11_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("召唤师峡谷地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nSummoner's Rift map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map11_bin = source.json()
    source, status, session = requestUrl("GET", map12_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("嚎哭深渊地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nHowling Abyss map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map12_bin_url))
            map12_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("嚎哭深渊地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nHowling Abyss map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map12_bin = source.json()
    source, status, session = requestUrl("GET", map21_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("百合与莲花的神庙地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nTemple of Lily and Lotus map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map21_bin_url))
            map21_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("百合与莲花的神庙地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nTemple of Lily and Lotus map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map21_bin = source.json()
    source, status, session = requestUrl("GET", map22_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("聚点危机地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nConvergence map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map22_bin_url))
            map22_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("聚点危机地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nConvergence map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map22_bin = source.json()
    source, status, session = requestUrl("GET", map30_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("怒火角斗场地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nRings of Wrath map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map30_bin_url))
            map30_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("怒火角斗场地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nRings of Wrath map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map30_bin = source.json()
    source, status, session = requestUrl("GET", map33_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("最终都市地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nFinal City map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map33_bin_url))
            map33_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("最终都市地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nFinal City map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map33_bin = source.json()
    source, status, session = requestUrl("GET", map35_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("班德尔之森地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nThe Bandlewood map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map35_bin_url))
            map35_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("班德尔之森地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nThe Bandlewood map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map35_bin = source.json()
    source, status, session = requestUrl("GET", map453_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("经典召唤师峡谷地图信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nClassic Rift map data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(map453_bin_url))
            map453_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("经典召唤师峡谷地图信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nClassic Rift map data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        map453_bin = source.json()
    source, status, session = requestUrl("GET", items_bin_url, session = session, log = log)
    if status != 200:
        if status == 404:
            logPrint("装备信息获取失败！请检查以下链接的可用性。程序将返回上一层。\nItem data capture failure! Please check the URL availability. The program will return to the last step.\n%s" %(items_bin_url))
            items_bin: dict[str, list[str] | dict[str, Any]] = {}
        else:
            logPrint("装备信息获取失败！请检查系统网络状况和代理设置。程序即将返回上一层。\nItem data capture failure! Please check the system network condition and proxy configuration. The program will return to the last step soon.")
            time.sleep(3)
        return ({}, False, "")
    else:
        items_bin = source.json()
    ##离线读取（Offline reading）
    ###存储库（Repository）
    # cdragon_folder: str = "C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/"
    # map11_bin_path: str = os.path.join(cdragon_folder, "pbe/game/data/maps/shipping/map11/map11.bin.json").replace("\\", "/")
    # map12_bin_path: str = os.path.join(cdragon_folder, "pbe/game/data/maps/shipping/map12/map12.bin.json").replace("\\", "/")
    # map21_bin_path: str = os.path.join(cdragon_folder, "pbe/game/data/maps/shipping/map21/map21.bin.json").replace("\\", "/")
    # map22_bin_path: str = os.path.join(cdragon_folder, "pbe/game/data/maps/shipping/map22/map22.bin.json").replace("\\", "/")
    # map30_bin_path: str = os.path.join(cdragon_folder, "pbe/game/data/maps/shipping/map30/map30.bin.json").replace("\\", "/")
    # map33_bin_path: str = os.path.join(cdragon_folder, "pbe/game/data/maps/shipping/map33/map33.bin.json").replace("\\", "/")
    # map35_bin_path: str = os.path.join(cdragon_folder, "pbe/game/data/maps/shipping/map35/map35.bin.json").replace("\\", "/")
    # map453_bin_path: str = os.path.join(cdragon_folder, "pbe/game/unknown/579b4182be3270f4.bin.json").replace("\\", "/")
    # items_bin_path: str = os.path.join(cdragon_folder, "pbe/game/items.cdtb.bin.json")
    ###提取目录（Extracted directory）
    # extract_folder: str = "D:/Workspace/LoL-Wad-Extract-Riot/pbe-text1/Game/DATA/FINAL/"
    # map11_bin_path: str = os.path.join(extract_folder, "data/maps/shipping/map11/map11.bin.json").replace("\\", "/")
    # map12_bin_path: str = os.path.join(extract_folder, "data/maps/shipping/map12/map12.bin.json").replace("\\", "/")
    # map21_bin_path: str = os.path.join(extract_folder, "data/maps/shipping/map21/map21.bin.json").replace("\\", "/")
    # map22_bin_path: str = os.path.join(extract_folder, "data/maps/shipping/map22/map22.bin.json").replace("\\", "/")
    # map30_bin_path: str = os.path.join(extract_folder, "data/maps/shipping/map30/map30.bin.json").replace("\\", "/")
    # map33_bin_path: str = os.path.join(extract_folder, "data/maps/shipping/map33/map33.bin.json").replace("\\", "/")
    # map35_bin_path: str = os.path.join(extract_folder, "data/maps/shipping/map35/map35.bin.json").replace("\\", "/")
    # map453_bin_path: str = os.path.join(extract_folder, "unknown/579b4182be3270f4.bin.json").replace("\\", "/")
    # items_bin_path: str = os.path.join(extract_folder, "items.cdtb.bin.json").replace("\\", "/")
    # with open(map11_bin_path, "r", encoding = "utf-8") as fp:
    #     map11_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(map12_bin_path, "r", encoding = "utf-8") as fp:
    #     map12_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(map21_bin_path, "r", encoding = "utf-8") as fp:
    #     map21_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(map22_bin_path, "r", encoding = "utf-8") as fp:
    #     map22_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(map30_bin_path, "r", encoding = "utf-8") as fp:
    #     map30_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(map33_bin_path, "r", encoding = "utf-8") as fp:
    #     map33_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(map35_bin_path, "r", encoding = "utf-8") as fp:
    #     map35_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(map453_bin_path, "r", encoding = "utf-8") as fp:
    #     map453_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    # with open(items_bin_path, "r", encoding = "utf-8") as fp:
    #     items_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
    #定义一个获取装备序号的函数（Define a function that gets itemId）
    def get_itemId(items_bin: dict[str, list[str] | dict[str, Any]], key: str) -> int:
        '''
        根据给定的装备二进制描述数据和一个装备主键，返回这个装备的序号。<br>According to given item binary description data and an item key, return this item's ID.

        :param items_bin: 装备二进制描述数据。通过以下链接得到：<br>Item binary description data, which can be obtained through the following link:

            - https://raw.communitydragon.org/pbe/game/items.cdtb.bin.json
        :type items_bin: dict[str, list[str] | dict[str, Any]]
        :param key: 要查询的装备主键。<br>The item key to query.
        :type key: str
        :return: 装备序号。如果未找到，则返回0。<br>ItemId. If no item is found, then 0 is returned.
        :rtype: int
        '''
        binhash_re: re.Pattern[str] = re.compile(r"\{\w{8}\}")
        if key in items_bin: #装备主键可直接在装备二进制描述数据中找到的情形（The case where the item key can be directly found in the item binary description data）
            return items_bin[key]["itemID"]
        elif binhash_re.fullmatch(key) and key in item_key_hash_map: #装备主键在地图二进制描述数据中是散列值，在装备二进制描述数据中是原始字符串的情形。此处`item_key_hash_map`是全局变量，只需构建一次。如果放在函数内，反复构建成为决速步骤（The case the item key is a hash value in the map binary description data but an original string in the item binary description data. Here `item_key_hash_map` is a global variable, which only needs to be constructed once. If it's put into this function, the repeated constructions of it becomes a rate-determining step）
            return items_bin[item_key_hash_map[key]]["itemID"]
        elif not binhash_re.fullmatch(key) and (key_hash := LoLDataExtractor.compute_binhash(key)) in items_bin: #装备主键在地图二进制描述数据中是原始字符串，在装备二进制描述数据中是散列值的情形（The case the item key is an original string in the map binary description data but a hash value in the item binary description data）
            return items_bin[key_hash]["itemID"]
        else:
            return 0
    #构建不同模式的全装备列表（Build the full item list for different modes）
    logPrint("正在构建各模式的全装备列表……\nBuilding full item list for different modes ...", print_time = True)
    maps_bin: list[dict[str, list[str] | dict[str, Any]]] = [map11_bin, map12_bin, map21_bin, map22_bin, map30_bin, map33_bin, map35_bin, map453_bin]
    itemKey_itemId_map: dict[int, str] = {} #构建从装备序号到装备数据的映射（Build a map from the itemId to the corresponding item data）
    exclusive_itemIds: set[int] = set() #构建一个装备序号集合，用来存储没有在任何地图的任何模式中出现的装备序号（Create a set of itemIds to store items that don't appear in any mode of any map）
    for (key, value) in items_bin.items():
        if key != "__linked" and value["__type"] == "ItemData":
            itemKey_itemId_map[value["itemID"]] = key
            exclusive_itemIds.add(value["itemID"])
    binhash_re: re.Pattern[str] = re.compile(r"\{\w{8}\}")
    items_bin_keys: list[str] = list(items_bin.keys())
    if "__linked" in items_bin_keys: #这个条件一般情况下都是真（Basically this condition is True）
        items_bin_keys.remove("__linked")
    item_key_hash_map: dict[str, str] = {LoLDataExtractor.compute_binhash(key): key for key in items_bin.keys() if not binhash_re.fullmatch(key)} #构建从装备主键的散列值到原始字符串的映射（Build a map from an item key's hash value to its original string）
    gameModeNames: dict[str, dict[str, str]] = {
        "Maps/Shipping/Map11/Modes/ARSR": {
            "zh_CN": "峡谷大乱斗",
            "en_US": "ARSR",
            "bilingual": "峡谷大乱斗（ARSR）"
        },
        "Maps/Shipping/Map11/Modes/ASSASSINATE": {
            "zh_CN": "红月决",
            "en_US": "Blood Moon",
            "bilingual": "红月决（Blood Moon）"
        },
        "Maps/Shipping/Map11/Modes/CLASSIC": {
            "zh_CN": "召唤师峡谷经典模式",
            "en_US": "Summoner's Rift Normal",
            "bilingual": "召唤师峡谷经典模式（Summoner's Rift Normal）"
        },
        "Maps/Shipping/Map11/Modes/DOOMBOTSTEEMO": {
            "zh_CN": "末日人工智能",
            "en_US": "Doom Bots of Doom",
            "bilingual": "末日人工智能（Doom Bots of Doom）"
        },
        "Maps/Shipping/Map11/Modes/ONEFORALL": {
            "zh_CN": "克隆大作战",
            "en_US": "One For All",
            "bilingual": "克隆大作战（One For All）"
        },
        "Maps/Shipping/Map11/Modes/PRACTICETOOL": {
            "zh_CN": "训练模式",
            "en_US": "Practice Tool",
            "bilingual": "训练模式（Practice Tool）"
        },
        "Maps/Shipping/Map11/Modes/RUBY": {
            "zh_CN": "末日人工智能",
            "en_US": "Doom Bots",
            "bilingual": "末日人工智能（Doom Bots）"
        },
        "Maps/Shipping/Map11/Modes/RUBY_TRIAL_1": {
            "zh_CN": "末日人工智能：维迦的诅咒！",
            "en_US": "Doom Bots: Veigar's Curse!",
            "bilingual": "末日人工智能：维迦的诅咒！（Doom Bots: Veigar's Curse!）"
        },
        "Maps/Shipping/Map11/Modes/RUBY_TRIAL_2": {
            "zh_CN": "末日人工智能：维迦的邪咒！",
            "en_US": "Doom Bots: Veigar's Evil",
            "bilingual": "末日人工智能：维迦的邪咒！（Doom Bots: Veigar's Evil）"
        },
        "Maps/Shipping/Map11/Modes/RUBY_TRIAL_3": {
            "zh_CN": "末日人工智能：维迦的末日厄咒！",
            "en_US": "Doom Bots: Veigar's Doom!",
            "bilingual": "末日人工智能：维迦的末日厄咒！（Doom Bots: Veigar's Doom!）"
        },
        "Maps/Shipping/Map11/Modes/SNOWURF": {
            "zh_CN": "冰雪无限火力",
            "en_US": "Snow ARURF",
            "bilingual": "冰雪无限火力（Snow ARURF）"
        },
        "Maps/Shipping/Map11/Modes/SWIFTPLAY": {
            "zh_CN": "快速模式",
            "en_US": "Swiftplay",
            "bilingual": "快速模式（Swiftplay）"
        },
        "Maps/Shipping/Map11/Modes/TUTORIAL": {
            "zh_CN": "新手教程 召唤师峡谷",
            "en_US": "Tutorial: Summoner's Rift",
            "bilingual": "新手教程 召唤师峡谷（Tutorial: Summoner's Rift）"
        },
        "Maps/Shipping/Map11/Modes/TUTORIAL_MODULE_1": {
            "zh_CN": "新手教程 第一部分",
            "en_US": "Tutorial Part 1",
            "bilingual": "新手教程 第一部分（Tutorial Part 1）"
        },
        "Maps/Shipping/Map11/Modes/TUTORIAL_MODULE_2": {
            "zh_CN": "新手教程 第二部分",
            "en_US": "Tutorial Part 2",
            "bilingual": "新手教程 第二部分（Tutorial Part 2）"
        },
        "Maps/Shipping/Map11/Modes/TUTORIAL_MODULE_3": {
            "zh_CN": "新手教程 第三部分",
            "en_US": "Tutorial Part 3",
            "bilingual": "新手教程 第三部分（Tutorial Part 3）"
        },
        "Maps/Shipping/Map11/Modes/ULTBOOK": {
            "zh_CN": "终极魔典",
            "en_US": "Ultimate Spellbook",
            "bilingual": "终极魔典（Ultimate Spellbook）"
        },
        "Maps/Shipping/Map11/Modes/URF": {
            "zh_CN": "无限火力",
            "en_US": "URF",
            "bilingual": "无限火力（URF）"
        },
        "Maps/Shipping/Map11/Modes/WASD": {
            "zh_CN": "WASD",
            "en_US": "WASD",
            "bilingual": "WASD"
        },
        "Maps/Shipping/Map12/Modes/ARAM": {
            "zh_CN": "极地大乱斗",
            "en_US": "ARAM",
            "bilingual": "极地大乱斗（ARAM）"
        },
        "Maps/Shipping/Map12/Modes/FIRSTBLOOD": {
            "zh_CN": "大对决",
            "en_US": "Showdown",
            "bilingual": "大对决（Showdown）"
        },
        "Maps/Shipping/Map12/Modes/KINGPORO": {
            "zh_CN": "魄罗大乱斗",
            "en_US": "Legend of the Poro King",
            "bilingual": "魄罗大乱斗（Legend of the Poro King）"
        },
        "Maps/Shipping/Map12/Modes/KIWI": {
            "zh_CN": "海克斯大乱斗",
            "en_US": "ARAM: Mayhem",
            "bilingual": "海克斯大乱斗（ARAM: Mayhem）"
        },
        "Maps/Shipping/Map12/Modes/KIWI_JADE": {
            "zh_CN": "海克斯大乱斗 经典模式版",
            "en_US": "ARAM: Mayhem Classic-ish",
            "bilingual": "海克斯大乱斗 经典模式版（ARAM: Mayhem Classic-ish）"
        },
        "Maps/Shipping/Map12/Modes/TUTORIAL": {
            "zh_CN": "新手教程 嚎哭深渊",
            "en_US": "Tutorial: Howling Abyss",
            "bilingual": "新手教程 嚎哭深渊（Tutorial: Howling Abyss）"
        },
        "Maps/Shipping/Map21/Modes/NEXUSBLITZ": {
            "zh_CN": "极限闪击",
            "en_US": "Nexus Blitz",
            "bilingual": "极限闪击（Nexus Blitz）"
        },
        "Maps/Shipping/Map22/Modes/TFT": {
            "zh_CN": "云顶之弈",
            "en_US": "TFT",
            "bilingual": "云顶之弈（TFT）"
        },
        "Maps/Shipping/Map30/Modes/CHERRY": {
            "zh_CN": "斗魂竞技场",
            "en_US": "Arena",
            "bilingual": "斗魂竞技场（Arena）"
        },
        "Maps/Shipping/Map33/Modes/STRAWBERRY": {
            "zh_CN": "无尽狂潮",
            "en_US": "Swarm",
            "bilingual": "无尽狂潮（Swarm）"
        },
        "Maps/Shipping/Map35/Modes/BRAWL": {
            "zh_CN": "神木之门",
            "en_US": "Brawl",
            "bilingual": "神木之门（Brawl）"
        },
        "{8d691c1c}": {
            "zh_CN": "{44334b59}",
            "en_US": "{44334b59}",
            "bilingual": "{44334b59}"
        },
        "{c706490e}": {
            "zh_CN": "{6462680f}",
            "en_US": "{6462680f}",
            "bilingual": "{6462680f}"
        },
        "Maps/Shipping/Map453/Modes/BASELINESR": {
            "zh_CN": "英雄联盟经典模式",
            "en_US": "League Classic",
            "bilingual": "英雄联盟经典模式（League Classic）"
        },
        "Maps/Shipping/Map453/Modes/JADE": {
            "zh_CN": "英雄联盟经典模式",
            "en_US": "League Classic",
            "bilingual": "英雄联盟经典模式（League Classic）"
        }
    } #之所以使用地图二进制数据的主键而不是其值字典的mModeName值，是因为召唤师峡谷和嚎哭深渊的地图二进制数据的新手教程的主键不一致，但是其值字典的mModeName值一致（The reason why the key of the binary data of maps is used instead of the `mModeName` value of this key's value dictionary is that the keys of TUTORIAL of the binary data of the maps 11 and 12 are different, but the `mModeName` values are the same）
    gameModes_ordered: list[str] = [
        "Maps/Shipping/Map11/Modes/CLASSIC",
        "Maps/Shipping/Map11/Modes/SWIFTPLAY",
        "Maps/Shipping/Map11/Modes/ARSR",
        "Maps/Shipping/Map11/Modes/URF",
        "Maps/Shipping/Map11/Modes/SNOWURF",
        "Maps/Shipping/Map11/Modes/ONEFORALL",
        "Maps/Shipping/Map11/Modes/ASSASSINATE",
        "Maps/Shipping/Map11/Modes/ULTBOOK",
        "Maps/Shipping/Map11/Modes/PRACTICETOOL",
        "Maps/Shipping/Map11/Modes/RUBY",
        "Maps/Shipping/Map11/Modes/RUBY_TRIAL_1",
        "Maps/Shipping/Map11/Modes/RUBY_TRIAL_2",
        "Maps/Shipping/Map11/Modes/RUBY_TRIAL_3",
        "Maps/Shipping/Map11/Modes/DOOMBOTSTEEMO",
        "Maps/Shipping/Map11/Modes/TUTORIAL",
        "Maps/Shipping/Map11/Modes/TUTORIAL_MODULE_1",
        "Maps/Shipping/Map11/Modes/TUTORIAL_MODULE_2",
        "Maps/Shipping/Map11/Modes/TUTORIAL_MODULE_3",
        "Maps/Shipping/Map11/Modes/WASD",
        "Maps/Shipping/Map12/Modes/ARAM",
        "Maps/Shipping/Map12/Modes/KIWI",
        "Maps/Shipping/Map12/Modes/KIWI_JADE",
        "Maps/Shipping/Map12/Modes/FIRSTBLOOD",
        "Maps/Shipping/Map12/Modes/KINGPORO",
        "Maps/Shipping/Map12/Modes/TUTORIAL",
        "Maps/Shipping/Map21/Modes/NEXUSBLITZ",
        "Maps/Shipping/Map22/Modes/TFT",
        "Maps/Shipping/Map30/Modes/CHERRY",
        "Maps/Shipping/Map33/Modes/STRAWBERRY",
        "Maps/Shipping/Map35/Modes/BRAWL",
        "Maps/Shipping/Map453/Modes/JADE",
        "Maps/Shipping/Map453/Modes/BASELINESR",
        "{8d691c1c}",
        "{c706490e}"
    ] #综合考虑发布时间和受欢迎程度排序（Ordered according to release date and popularity）
    itemPage: dict[str, Any] = {
        "associatedChampions": [],
        "associatedMaps": [11, 12, 21, 22, 30, 33, 35, 453],
        "blocks": [],
        "map": "any",
        "mode": "any",
        "preferredItemSlots": [],
        "sortrank": 0,
        "startedFrom": "blank",
        "title": "测试配装 | Test Item Set" if bilingual else "测试配装" if isZH else "Test Item Set",
        "type": "custom",
        "uid": str(uuid.uuid4())
    }
    itemBlocks: dict[str, dict[str, Any]] = {} #注意，这并不是最终要形成的配装方案的装备数据。因为这个字典里存在大量的重复装备（Note that this isn't the block data of the final item set, because there're a lot of redundant items in this dictionary）
    itemBlock_exclusive: dict[str, Any] = {"items": [], "type": "其它（Others）" if bilingual else "其它" if isZH else "Others"} #存储那些没有在任何地图的任何模式中出现的装备（Stores items that don't appear in any mode of any map）
    for map_bin in maps_bin:
        for (key, value) in map_bin.items():
            if key != "__linked" and value["__type"] == "GameModeMapData":
                if not key in itemBlocks: #这个条件总是真（This condition is always True）
                    itemBlocks[key] = {"items": [], "type": gameModeNames[key]["bilingual" if bilingual else "zh_CN" if isZH else "en_US"]}
                    gameModeItemIds: list[int] = []
                    if "itemLists" in value:
                        for itemList_key in value["itemLists"]:
                            for item_key in map_bin[itemList_key]["mItems"]:
                                if (itemId := get_itemId(items_bin, item_key)) != 0:
                                    gameModeItemIds.append(itemId)
                    if "Configs" in value:
                        for config_key in value["Configs"]:
                            if config_key in map_bin and map_bin[config_key]["__type"] == "{3d900309}":
                                for itemList_key in map_bin[config_key]["itemLists"].values():
                                    for item_key in map_bin[itemList_key]["mItems"]:
                                        if (itemId := get_itemId(items_bin, item_key)) != 0:
                                            gameModeItemIds.append(itemId)
                    gameModeItemIds = sorted(set(gameModeItemIds)) #依据装备序号去重（Remove redundancy according to itemId）
                    exclusive_itemIds -= set(gameModeItemIds)
                    for i in range(len(gameModeItemIds)):
                        itemId: int = gameModeItemIds[i]
                        itemJson: dict[str, str | int] = {"id": str(itemId), "count": 1, "priceTotal": 0}
                        item_key: str = itemKey_itemId_map[itemId]
                        priceTotal: int = calculate_totalPrice(item_key, items_bin)
                        itemJson["priceTotal"] = priceTotal
                        itemBlocks[key]["items"].append(itemJson)
    #排序装备区块（Order the item blocks）
    itemBlocks_ordered: dict[str, dict[str, Any]] = {}
    for key in gameModes_ordered: #先将已经确定顺序的装备区块放在前面（Put the item blocks whose order are known in the front）
        if key in itemBlocks:
            itemBlocks_ordered[key] = itemBlocks.pop(key)
    for key in itemBlocks: #再将剩余部分按照原顺序添加到字典中（Add the rest of blocks into the ordered dictionary, reserving their original order）
        itemBlocks_ordered[key] = itemBlocks.pop(key)
    itemBlocks = itemBlocks_ordered
    del itemBlocks_ordered
    #下面开始制作配装方案（Build the item set）
    logPrint("正在制作配装方案……\nCreating the item set ...", print_time = True)
    baseBlock_key = "Maps/Shipping/Map11/Modes/CLASSIC"
    ##首先提取所有召唤师峡谷经典模式的装备，作为配装方案的第一块（Extract all Summoner's Rift Normal items as the first block of the item set）
    itemPage["blocks"].append(copy.deepcopy(itemBlocks[baseBlock_key]))
    ##将召唤师峡谷经典模式的装备中的防御塔类型的装备提取出来，单独作为一块（Extract all turret items from Summoner's Rift Normal item set as a single block）
    itemBlock_turret: dict[str, Any] = {"items": [], "type": "防御塔（Turret）" if bilingual else "防御塔" if isZH else "Turret"}
    for itemJson in itemBlocks[baseBlock_key]["items"]:
        if int(itemJson["id"]) >= 1500 and int(itemJson["id"]) < 1600:
            itemBlock_turret["items"].append(copy.deepcopy(itemJson))
            itemPage["blocks"][0]["items"].remove(itemJson) #将防御塔的装备从召唤师峡谷经典模式装备区块中移除（Remove turret items from Summoner's Rift Normal item block）
    itemPage["blocks"].append(itemBlock_turret)
    ##将召唤师峡谷经典模式的装备中的英雄特定装备提取出来，单独作为一块（Extract all champion-specific items from Summoner's Rift Normal item set as a single block）
    itemBlock_championSpecific: dict[str, Any] = {"items": [], "type": "英雄特定（Champion Specific）" if bilingual else "英雄特定" if isZH else "Champion Specific"}
    for itemJson in itemBlocks[baseBlock_key]["items"]:
        item_key: str = itemKey_itemId_map[int(itemJson["id"])]
        if "mRequiredChampion" in items_bin[item_key]:
            itemBlock_championSpecific["items"].append(copy.deepcopy(itemJson))
            itemPage["blocks"][0]["items"].remove(itemJson) #将英雄特定装备从召唤师峡谷经典模式装备区块中移除（Remove champion-specific items from Summoner's Rift Normal item block）
    itemPage["blocks"].append(itemBlock_championSpecific)
    ##依次追加其它模式的装备（Append other mode's item set block by block）
    for key in itemBlocks:
        if key != baseBlock_key:
            itemBlock_otherMode: dict[str, Any] = copy.deepcopy(itemBlocks[key]) #深复制，防止后续处理装备区块时影响到itemBlocks变量中存储的原始数据（Deep copy prevents the original stored in `itemBlocks` being influenced when the item block is further processed）
            for itemJson in itemBlocks[baseBlock_key]["items"]: #移除召唤师峡谷经典模式的所有装备，包括防御塔装备。这就是为什么前面一直强调itemBlocks中的内容不能被修改。如果被修改，那这里就要写成两个for循环：一个for循环去除召唤师峡谷经典模式的基础装备，另一个for循环去除召唤师峡谷经典模式的防御塔装备（Remvoe all Summoner's Rift Normal items, including turret items. That's why I keep emphasizing the content in `itemBlocks` can't be changed. If it's changed, this for-loop would have been split into two parts: one loop removes the basic items in Summoner's Rift Normal, and the other loop removes the turret items in Summoner's Rift Normal）
                if itemJson in itemBlock_otherMode["items"]:
                    itemBlock_otherMode["items"].remove(itemJson)
            if len(itemBlock_otherMode["items"]) > 0: #没有模式特殊装备的游戏模式不需要添加到配装方案中（Game modes without mode-specific items don't need to be added into the item set）
                itemPage["blocks"].append(itemBlock_otherMode)
    ##最后追加排除在任何地图的任何模式中的装备（Finally, add items excluded from any mode of any map）
    for itemId in sorted(exclusive_itemIds):
        itemJson: dict[str, str | int] = {"id": str(itemId), "count": 1, "priceTotal": 0}
        item_key: str = itemKey_itemId_map[itemId]
        priceTotal: int = calculate_totalPrice(item_key, items_bin)
        itemJson["priceTotal"] = priceTotal
        itemBlock_exclusive["items"].append(itemJson)
    itemPage["blocks"].append(itemBlock_exclusive)
    #将配装方案的每一块按照金币价格正序排列。无须担心第二关键字，因为前面的sorted(set())方法已经表明所有装备起码是按照装备序号正序排列的，并且后续处理过程没有涉及装备列表的错序（Arrange each block in the item set in the ascending order of total price. Don't worry about the second keyword, because the previous `sorted(set())` method already declares that all items must be arranged in the ascending order of itemId, plus the later process doesn't involve shuffling the item order）
    for i in range(len(itemPage["blocks"])):
        if len(itemPage["blocks"][i]["items"]) > 0:
            itemBlock_sorted: dict[str, Any] = {"items": sorted(itemPage["blocks"][i]["items"], key = lambda x: x["priceTotal"], reverse = False), "type": itemPage["blocks"][i]["type"]}
            itemPage["blocks"][i] = itemBlock_sorted
    #保存数据（Save data）
    logPrint("正在导出数据……\nExporting data ...", print_time = True)
    documents_dir: str = os.path.expanduser("~/Documents")
    filepath: str = os.path.join(documents_dir, "测试装备.json").replace("\\", "/")
    with open(filepath, "w", encoding = "utf-8") as fp:
        json.dump(itemPage, fp, indent = 4, ensure_ascii = False)
    logPrint("测试装备配装方案已保存到%s。\nTest item set has been saved to %s." %(filepath, filepath), print_time = True)
    return (itemPage, True, filepath)

class ItemSet:
    def __init__(self, itemSet: Any, spells: dict[int, dict[str, Any]], LoLChampions: dict[int, dict[str, Any]], LoLItems: dict[int, dict[str, Any]], isZH: bool = False) -> None:
        '''
        配装方案类构造函数。<br>Constructor of `ItemSet` class.
        
        :param itemSet: 配装方案集合。通过`GET /lol-item-sets/v1/item-sets/{summonerId}/sets`接口得到，而不是从客户端导出。<br>Item scheme set. Obtained through the endpoint `GET /lol-item-sets/v1/item-sets/{summonerId}/sets`, rather than exported through League Client.
        :type itemSet: dict[str, Any]
        :param LoLChampions: 整理后的英雄概要数据。键为英雄序号，值为英雄对象。<br>Organized champion data, whose keys are championIds and values are champion objects.
        :type LoLChampions: dict[str, Any]
        :param LoLItems: 整理后的装备数据。键为装备序号，值为装备对象。<br>Organized item data, whose keys are itemIds and values are item objects.
        :type LoLItems: dict[str, Any]
        :param isZH: 是否以简体中文输出配装方案内容。默认为假。<br>Whether to output the item set content in Chinese Simplified. False by default.
        :type isZH: bool
        
        注：在本类中，itemSet指配装方案集合。ItemPage指列表`self.itemPages`中的每个对象。<br>Note: In this class, `itemSet` denodes an item scheme set, while `ItemPage` denodes each object in the list `self.itemPages`.
        '''
        #配装相关属性（Item set related variables）
        if isinstance(itemSet, dict) and "accountId" in itemSet and isinstance(itemSet["accountId"], int) and "timestamp" in itemSet and isinstance(itemSet["timestamp"], int) and "itemSets" in itemSet and isinstance(itemSet["itemSets"], list) and all(map(self.isItemPage, itemSet["itemSets"])):
            self.itemSet: dict[str, Any] = itemSet
        else:
            logPrint("您输入的配装方案格式有误！配装方案已初始化为空。\nItem set format error! Item sets have been initialized as empty.")
            self.itemSet = {"accountId": 0, "itemSets": [], "timestamp": int(time.time() * 1000)}
        self.itemPages: list[dict[str, Any]] = self.itemSet["itemSets"]
        self.itemPage_uids: list[str] = list(map(lambda x: x["uid"], self.itemPages))
        self._itemPage_dict: dict[str, dict[str, Any]] = {page["uid"]: page for page in self.itemPages}
        #下面是数据资源部分（The following part is the data resource part）
        if self.isSpellData(spells):
            self.spells: dict[int, dict[str, Any]] = spells
        else:
            raise TypeError("Invalid format of summoner spell data.")
        if self.isChampionData(LoLChampions):
            self.LoLChampions: dict[int, dict[str, Any]] = LoLChampions
        else:
            raise TypeError("Invalid format of champion data.")
        if self.isItemData(LoLItems):
            self.LoLItems: dict[int, dict[str, Any]] = LoLItems
        else:
            raise TypeError("Invalid format of item data.")
        self.isZH: bool = isZH
    
    def update_itemSet_timestamp(self) -> None:
        '''
        更新配装方案的时间戳。<br>Update the timestamp of the item set.
        '''
        self.itemSet["timestamp"] = int(time.time() * 1000)
    
    def sync(self) -> None:
        '''
        将配装方案相关属性与最新配装方案同步。<br>Synchronize the attributes related to the item set with the latest item set.
        '''
        self.itemPages = self.itemSet["itemSets"]
        self.itemPage_uids = list(map(lambda x: x["uid"], self.itemPages))
        self._itemPage_dict = {page["uid"]: page for page in self.itemPages}
    
    def update(self, itemSet: Any) -> None:
        '''
        更新配装方案。<br>Update the item set.
        
        召唤师技能、英雄和装备等辅助数据在单次客户端会话内只需要获取一次。<br>Auxillary data like summoner spells, champions and items only need to be fetched once during a single client session.
        
        :param itemSet: 新的配装方案数据。<br>New item set data.
        :type itemSet: Any
        '''
        if self.isItemSet(itemSet):
            self.itemSet = itemSet
        else:
            logPrint("您输入的配装方案格式有误！配装方案已初始化为空。\nItem set format error! Item sets have been initialized as empty.")
            self.itemSet = {"accountId": 0, "itemSets": [], "timestamp": int(time.time() * 1000)}
        self.sync()
        self.itemSet["timestamp"] = int(time.time() * 1000)
    
    @staticmethod
    def isItemSet(itemSet: Any) -> bool:
        '''
        判断一个配装集合是否符合API中记录的格式。<br>Judge whether an item set obeys the format documented by API.
        
        :param itemSet: 配装集合。通过`GET /lol-item-sets/v1/item-sets/{summonerId}/sets`接口得到。<br>An item set. Obtained through the endpoint `GET /lol-item-sets/v1/item-sets/{summonerId}/sets`.
        :type itemSet: Any
        :return: 配装集合是否符合预期格式。<br>Whether `itemSet` obeys the expected format.
        :rtype: bool
        '''
        return isinstance(itemSet, dict) and "accountId" in itemSet and isinstance(itemSet["accountId"], int) and "timestamp" in itemSet and isinstance(itemSet["timestamp"], int) and "itemSets" in itemSet and isinstance(itemSet["itemSets"], list) and all(map(ItemSet.isItemPage, itemSet["itemSets"]))
    
    @staticmethod
    def isItemPage(itemPage: Any) -> bool:
        '''
        判断一个配装页是否符合API中记录的格式。<br>Judge whether an item set page obeys the format documented by API.
        
        :param itemPage: 配装页，配装方案的“itemSets”键的值中的一个元素。<br>An item set page, which is an element in the value of "itemSets" key of the item set.
        :type itemPage: Any
        :return: 配装页是否符合预期格式。<br>Whether `itemPage` obeys the expected format.
        :rtype: bool
        '''
        return isinstance(itemPage, dict) and all(key in itemPage for key in ["associatedChampions", "associatedMaps", "blocks", "map", "mode", "preferredItemSlots", "sortrank", "startedFrom", "title", "type", "uid"]) and all(isinstance(itemPage[key], list) and all(map(lambda x: isinstance(x, int), itemPage[key])) for key in ["associatedChampions", "associatedMaps"]) and all(isinstance(itemPage[key], list) for key in ["blocks", "preferredItemSlots"]) and all(isinstance(itemPage[key], int) for key in ["sortrank"]) and all(isinstance(itemPage[key], str) for key in ["map", "mode", "startedFrom", "title", "type"]) and verify_uuid(itemPage["uid"])
    
    @staticmethod
    def isSpellData(spells: Any) -> bool:
        '''
        判断召唤师技能数据是否符合插件数据的格式。<br>Judge whether the spell data obeys the plugin format.
        
        :param itemPage: 召唤师技能数据。由原始数据提取序号构建字典后得到。<br>Summoner spell data, which are obtained by extracting the ids and creating a dictionary.
        :type itemPage: Any
        :return: 召唤师技能数据是否符合预期格式。<br>Whether `spells` obeys the expected format.
        :rtype: bool
        '''
        return isinstance(spells, dict) and all(map(lambda x: isinstance(x, int), spells.keys())) and all(map(lambda x: isinstance(x, dict) and all(key in x for key in ["id", "name", "description", "summonerLevel", "cooldown", "gameModes", "iconPath"]) and all(isinstance(x[key], int) for key in ["id", "summonerLevel", "cooldown"]) and all(isinstance(x[key], str) for key in ["name", "description", "iconPath"]) and all(isinstance(x[key], list) for key in ["gameModes"]) and all(map(lambda y: isinstance(y, str), x["gameModes"])), spells.values()))
    
    @staticmethod
    def isChampionData(LoLChampions: Any) -> bool:
        '''
        判断英雄数据是否符合插件数据的格式。<br>Judge whether the LoL champion data obeys the plugin format.
        
        :param itemPage: 英雄数据。由原始数据提取序号构建字典后得到。<br>LoL champion data, which are obtained by extracting the ids and creating a dictionary.
        :type itemPage: Any
        :return: 英雄数据是否符合预期格式。<br>Whether `LoLChampions` obeys the expected format.
        :rtype: bool
        '''
        return isinstance(LoLChampions, dict) and all(map(lambda x: isinstance(x, int), LoLChampions.keys())) and all(map(lambda x: isinstance(x, dict) and all(key in x for key in ["id", "name", "description", "alias", "contentId", "squarePortraitPath", "roles"]) and all(isinstance(x[key], int) for key in ["id"]) and all(isinstance(x[key], str) for key in ["name", "description", "alias", "contentId", "squarePortraitPath"]) and all(isinstance(x[key], list) for key in ["roles"]) and all(map(lambda y: isinstance(y, str), x["roles"])), LoLChampions.values()))
    
    @staticmethod
    def isItemData(LoLItems: Any) -> bool:
        '''
        判断装备数据是否符合插件数据的格式。<br>Judge whether the LoL item data obeys the plugin format.
        
        :param itemPage: 装备数据。由原始数据提取序号构建字典后得到。<br>LoL item data, which are obtained by extracting the ids and creating a dictionary.
        :type itemPage: Any
        :return: 装备数据是否符合预期格式。<br>Whether `LoLItems` obeys the expected format.
        :rtype: bool
        '''
        return isinstance(LoLItems, dict) and all(map(lambda x: isinstance(x, int), LoLItems.keys())) and all(map(lambda x: isinstance(x, dict) and all(key in x for key in ["id", "name", "description", "active", "inStore", "from", "to", "categories", "maxStacks", "requiredChampion", "requiredAlly", "requiredBuffCurrencyName", "requiredBuffCurrencyCost", "specialRecipe", "isEnchantment", "price", "priceTotal", "displayInItemSets", "iconPath"]) and all(isinstance(x[key], int) for key in ["id", "maxStacks", "requiredBuffCurrencyCost", "specialRecipe", "price", "priceTotal"]) and all(isinstance(x[key], str) for key in ["name", "description", "requiredChampion", "requiredAlly", "requiredBuffCurrencyName", "iconPath"]) and all(isinstance(x[key], bool) for key in ["active", "inStore", "isEnchantment", "displayInItemSets"]) and all(isinstance(x[key], list) for key in ["from", "to", "categories"]) and all(map(lambda y: isinstance(y, int), x["from"] + x["to"])) and all(map(lambda y: isinstance(y, str), x["categories"])), LoLItems.values()))
    
    def __repr__(self) -> str: #自我描述（Self description）
        if len(self.itemPages) == 0:
            result: str = "该配装方案为空。" if self.isZH else "This item set is empty."
        else:
            result = ""
            for i in range(len(self.itemPages)):
                itemPage: dict[str, Any] = self.itemPages[i]
                if self.isZH:
                    result += "配装页%d：%s - %s\n" %(i + 1, itemPage["title"], itemPage["uid"])
                else:
                    result += "Item Page %d: %s - %s\n" %(i + 1, itemPage["title"], itemPage["uid"])
                itemPage: dict[str, Any] = self.itemPages[i]
                itemPage_str: str = self.format_page(itemPage["uid"])
                result += itemPage_str
        return result
    
    def format_block(self, page_uid: str, pos: int) -> str:
        '''
        格式化输出某个配装页中的装备区块内容。<br>Format print the content in an item block of an item set page.
        
        :param page_uid: 配装页的唯一识别码。用于定位到该区块。<br>Unique identifier of the item set page. Used to locate this block.
        :type page_uid: str
        :param pos: 配装页中的区块下标。用于定位到该区块。<br>The index of the block in the item set page. Used to locate this block.
        :type pos: int
        :return: 格式化字符串。<br>A formatted string.
        :rtype: str
        '''
        if page_uid in self._itemPage_dict:
            itemPage: dict[str, Any] = self._itemPage_dict[page_uid]
            if pos >= -len(itemPage["blocks"]) and pos < len(itemPage["blocks"]):
                result: str = ""
                if self.isZH:
                    block: dict[str, Any] = itemPage["blocks"][pos]
                    result += "区块%d——%s：\n" %(pos + 1, block["type"])
                    if block["showIfSummonerSpell"] != "":
                        result += "在选用召唤师技能%s时显示。\n" %(block["showIfSummonerSpell"])
                    if block["hideIfSummonerSpell"] != "":
                        result += "在选用召唤师技能%s时隐藏。\n" %(block["showIfSummonerSpell"])
                    item_grid_df: pandas.DataFrame = self.sort_item_grid(block["items"])
                    result += format_df(item_grid_df, width_exceed_ask = False, direct_print = False)[0] + "\n\n"
                else:
                    block: dict[str, Any] = itemPage["blocks"][pos]
                    result += "Block %d - %s:\n" %(pos + 1, block["type"])
                    if block["showIfSummonerSpell"] != "":
                        result += "Show if the following summoner spells are chosen: %s.\n" %(block["showIfSummonerSpell"])
                    if block["hideIfSummonerSpell"] != "":
                        result += "Hide if the following summoner spells are chosen: %s.\n" %(block["showIfSummonerSpell"])
                    item_grid_df: pandas.DataFrame = self.sort_item_grid(block["items"])
                    result += format_df(item_grid_df, width_exceed_ask = False, direct_print = False)[0] + "\n\n"
            else:
                result = "区块下标越界。" if self.isZH else "Block index out of range."
        else:
            result = "未找到该配装页。" if self.isZH else "Item set page not found."
        return result
    
    def format_page(self, page_uid: str) -> str:
        '''
        格式化输出某个配装页。<br>Format print the content in an item set page.
        
        :param page_uid: 配装页的唯一识别码。用于定位到该区块。<br>Unique identifier of the item set page. Used to locate this block.
        :type page_uid: str
        :return: 格式化字符串。<br>A formatted string.
        :rtype: str
        '''
        if page_uid in self._itemPage_dict:
            itemPage: dict[str, Any] = self._itemPage_dict[page_uid]
            result: str = ""
            if self.isZH:
                if len(itemPage["associatedChampions"]) > 0:
                    result += "适用英雄：\n"
                    for championId in itemPage["associatedChampions"]:
                        result += "%s\n" %(self.LoLChampions[championId]["name"])
                if len(itemPage["associatedMaps"]) > 0:
                    result += "适用地图：" + "、".join(list(map(lambda x: gamemaps[x]["zh_CN"], itemPage["associatedMaps"]))) + "\n"
                if len(itemPage["blocks"]) > 0:
                    for j in range(len(itemPage["blocks"])):
                        result += self.format_block(itemPage["uid"], j)
                result += "\n"
            else:
                if len(itemPage["associatedChampions"]) > 0:
                    result += "Associated champions:\n"
                    for championId in itemPage["associatedChampions"]:
                        result += "%s\n" %(self.LoLChampions[championId]["name"])
                if len(itemPage["associatedMaps"]) > 0:
                    result += "Associated maps: " + ", ".join(list(map(lambda x: gamemaps[x]["zh_CN"], itemPage["associatedMaps"]))) + "\n"
                if len(itemPage["blocks"]) > 0:
                    for j in range(len(itemPage["blocks"])):
                        result += self.format_block(itemPage["uid"], j)
                result += "\n"
        else:
            result = "未找到该配装页。" if self.isZH else "Item set page not found."
        return result
    
    def get_item_page(self, page_uid: str) -> dict[str, Any]:
        '''
        根据给定的唯一识别码获取配装方案中的某个配装页数据。如果唯一识别码不存在，则返回空字典。<br>Get the data of an item set page according to a specified uid. If that uid doesn't exist, return an empty dictionary instead.
        
        :param page_uid: 配装页的唯一识别码。<br>Unique identifier of the item set page.
        :type page_uid: str
        :return: 配装页数据。如果不存在，则为空字典。<br>Item set page data. If no match is found, then return an empty dictionary instead.
        :rtype: dict[str, Any]
        '''
        return self._itemPage_dict.get(page_uid, {})
    
    def export_item_page(self, page_uid: str) -> dict[str, Any]:
        '''
        和get_item_page方法不同的是，该方法返回的是更加符合客户端格式的数据。<br>What's different of this method from `get_item_page` is that this method returns data of a format that resembles what the League Client returns more than `get_item_page` method.
        
        :param page_uid: 配装页的唯一识别码。<br>Unique identifier of the item set page.
        :type page_uid: str
        :return: 配装页数据。如果不存在，则为空字典。<br>Item set page data. If no match is found, then return an empty dictionary instead.
        :rtype: dict[str, Any]
        '''
        if page_uid in self._itemPage_dict:
            blocks_simplified: list[dict[str, Any]] = [{"items": [{"id": item["id"], "count": item["count"]} for item in block["items"]], "type": block["type"]} for block in self._itemPage_dict[page_uid]["blocks"]]
            itemPage: dict[str, Any] = {
                "title": self._itemPage_dict[page_uid]["title"],
                "associatedMaps": self._itemPage_dict[page_uid]["associatedMaps"],
                "associatedChampions": self._itemPage_dict[page_uid]["associatedChampions"],
                "blocks": blocks_simplified
            }
            return itemPage
        else:
            return {}
    
    def sort_item_grid(self, itemList: list[dict[str, int | str]]) -> pandas.DataFrame:
        '''
        用于将每个区块的每一格中的装备列表整理成一个数据框。<br>Designed to sort out items in each grid of each block into a dataframe.
        
        :param itemList: 每格中的装备列表。每个元素形如`{"count": 0, "id": "0"}`。<br>A list of items in a grid. Each element looks like `{"count": 0, "id": "0"}`.
        :type itemList: list[dict[str, int]]
        :param header_isZH: （已弃用。）表头是否以中文显示。默认为假。<br>(Deprecated) Whether to display the headers in Chinese. False by default.
        :type header_isZH: bool
        :return: 装备数据框，包含装备序号、装备名称、最大数量和总价。<br>An item dataframe containing itemId, name, count and total price.
        :rtype: pandas.DataFrame
        '''
        item_grid_header: dict[str, str] = {"count": "数量", "id": "序号", "name": "名称", "priceTotal": "总价"}
        item_grid_header_keys: list[str] = list(item_grid_header.keys())
        item_grid_header_values: list[str] = list(item_grid_header.values())
        item_grid_data: dict[str, list[Any]] = {key: [] for key in item_grid_header_keys}
        for item in itemList:
            itemId: int = int(item["id"])
            if itemId in self.LoLItems:
                itemObj: dict[str, Any] = self.LoLItems[itemId]
                for i in range(len(item_grid_header_keys)):
                    key: str = item_grid_header_keys[i]
                    if i == 0: #数量（`count`）
                        to_append: Any = "" if item["count"] == 1 and itemObj["maxStacks"] <= 1 else item["count"]
                    elif i == 1: #序号（`id`）
                        to_append = itemId
                    else:
                        to_append = itemObj[key]
                    item_grid_data[key].append(to_append)
            else:
                for i in range(len(item_grid_header_keys)):
                    key: str = item_grid_header_keys[i]
                    if i == 0: #数量（`count`）
                        to_append = "" if item["count"] == 1 else item["count"]
                    elif i == 1: #序号（`id`）
                        to_append = itemId
                    else:
                        to_append = ""
                    item_grid_data[key].append(to_append)
        item_grid_statistics_output_order: list[int] = [1, 2, 3, 0]
        item_grid_data_organized: dict[str, list[Any]] = {(item_grid_header_values if self.isZH else item_grid_header_keys)[i]: item_grid_data[item_grid_header_keys[i]] for i in item_grid_statistics_output_order} #由于下文涉及对区块内装备列表的修改，要求用户输入列表的索引，而列表的索引从0开始，因此没有设置双语表头同时出现（Because the item list in a block may be changed downstream, and users are asked to input the indices of a list, whose index starts with 0, here only monolingual headers instead of bilingual headers are used）
        item_grid_df: pandas.DataFrame = pandas.DataFrame(item_grid_data_organized)
        return item_grid_df
    
    def sort_item_block(self, item_blocks: list[dict[str, Any]]) -> pandas.DataFrame:
        '''
        用于将每个配装页中的装备区块信息整理成一个数据框。<br>Designed to sort out item blocks in each item set page into a dataframe.
        
        :param item_blocks: 每个配装页中的所有装备区块。<br>All item blocks in an item set page.
        :type item_blocks: list[dict[str, Any]]
        :return: 装备区块数据框，包含区块名称、显示召唤师技能和隐藏召唤师技能。<br>An item block dataframe containing the block name, the only summoner spells to make it show and the only summoner spells to make it hide.
        :rtype: pandas.DataFrame
        '''
        item_block_header: dict[str, str] = {"type": "名称", "showIfSummonerSpell": "显示召唤师技能", "hideIfSummonerSpell": "隐藏召唤师技能"}
        item_block_header_keys: list[str] = list(item_block_header.keys())
        item_block_header_values: list[str] = list(item_block_header.values())
        item_block_data: dict[str, list[Any]] = {key: [] for key in item_block_header_keys}
        for item_block in item_blocks:
            for i in range(len(item_block_header_keys)):
                key: str = item_block_header_keys[i]
                if i == 0: #名称（`type`）
                    to_append: Any = item_block["type"]
                else:
                    if item_block[key] == "":
                        to_append = ""
                    else:
                        summonerSpellNames: list[str] = list(map(lambda x: self.spells[int(x)]["name"] if int(x) in self.spells else int(x), item_block[key].split(",")))
                        to_append = summonerSpellNames
                item_block_data[key].append(to_append)
        item_block_statistics_output_order: list[int] = list(range(len(item_block_header_keys)))
        item_block_data_organized: dict[str, list[Any]] = {(item_block_header_values if self.isZH else item_block_header_keys)[i]: item_block_data[item_block_header_keys[i]] for i in item_block_statistics_output_order}
        item_block_df: pandas.DataFrame = pandas.DataFrame(item_block_data_organized)
        return item_block_df
    
    def sort_item_page(self, item_pages: list[dict[str, Any]]) -> pandas.DataFrame:
        '''
        用于整理所有配装页。<br>Designed to sort out all item set pages.
        
        :param item_pages: 所有配装页。通过对从接口直接获取到的配装信息取“itemSets”键的值得到。<br>All item set pages. Obtained by taking the value of "itemSets" key from the item set information obtained through LCU API.
        :type item_pages: list[dict[str, Any]]
        :return: 配装页数据框。<br>An item page dataframe.
        :rtype: pandas.DataFrame
        '''
        item_page_header: dict[str, str] = {"associatedChampions": "相关英雄序号列表", "associatedMaps": "相关地图序号列表", "blocks": "装备区块", "map": "适用地图", "mode": "适用游戏模式", "preferredItemSlots": "偏好栏", "sortrank": "排序优先级", "startedFrom": "来自", "title": "标题", "type": "类型", "uid": "唯一识别码", "associatedChampions_name": "相关英雄头衔", "associatedChampions_description": "相关英雄名称", "associatedMaps_name_zh": "相关地图名称（中文）", "associatedMaps_name_en": "相关地图名称（英文）", "blocks type": "装备区块名称列表"}
        item_page_header_keys: list[str] = list(item_page_header.keys())
        item_page_data: dict[str, list[Any]] = {key: [] for key in item_page_header_keys}
        item_page_data_json = copy.deepcopy(item_page_data)
        for item_page in item_pages:
            for i in range(len(item_page_header_keys)):
                key: str = item_page_header_keys[i]
                if i <= 10:
                    to_append = item_page[key]
                elif i == 11 or i == 12: #相关英雄序号列表子键（`associatedChampions`' subkeys）
                    subkey: str = key.split("_")[1]
                    to_append = list(map(lambda x: self.LoLChampions[x][subkey] if x in self.LoLChampions else "", item_page["associatedChampions"]))
                    if to_append == []:
                        to_append = ""
                elif i == 13 or i == 14: #相关地图序号列表（`associatedMaps`）
                    to_append = list(map(lambda x: gamemaps[x]["zh_CN" if i == 13 else "en_US"] if x in gamemaps else "", item_page["associatedMaps"]))
                    if to_append == []:
                        to_append = ""
                else: #装备区块名称列表（`blocks type`）
                    to_append = list(map(lambda x: x["type"], item_page["blocks"]))
                item_page_data[key].append(to_append)
                item_page_data_json[key].append(pyobj2json(to_append))
        item_page_statistics_output_order: list[int] = [8, 9, 10, 3, 4, 0, 11, 12, 1, 13, 14, 2, 15, 5, 6, 7]
        item_page_data_organized: dict[str, list[Any]] = {item_page_header_keys[i]: item_page_data[item_page_header_keys[i]] for i in item_page_statistics_output_order}
        item_page_df: pandas.DataFrame = pandas.DataFrame(item_page_data_organized)
        item_page_df = pandas.concat([pandas.DataFrame([item_page_header])[item_page_df.columns], item_page_df], ignore_index = True)
        return item_page_df
    
    def sort_preferredItemSlot(self, preferredItemSlot_dict: dict[int, int]) -> pandas.DataFrame:
        '''
        用于将配装页中的装备偏好栏信息整理成一个数据框。<br>Designed to sort out preferred item slot information into a dataframe.
        
        :param preferredItemSlot_dict: 每个配装页中的所有装备偏好栏信息。键是装备序号，值是偏好栏。<br>All preferred item slot information in an item set page, whose keys are itemIds and values are preferred slots.
        :type preferredItemSlot_dict: list[dict[str, Any]]
        :return: 装备偏好栏数据框，包含装备序号、装备名称和偏好栏。<br>A preferred item slot dataframe containing the itemId, item name and preferred slot.
        :rtype: pandas.DataFrame
        '''
        preferredItemSlot_header: dict[str, str] = {"id": "序号", "preferredItemSlot": "偏好栏", "name": "名称"}
        preferredItemSlot_header_keys: list[str] = list(preferredItemSlot_header.keys())
        preferredItemSlot_header_values: list[str] = list(preferredItemSlot_header.values())
        preferredItemSlot_data: dict[str, list[Any]] = {key: [] for key in preferredItemSlot_header_keys}
        for (key1, value) in preferredItemSlot_dict.items():
            for i in range(len(preferredItemSlot_header_keys)):
                key: str = preferredItemSlot_header_keys[i]
                if i == 0: #序号（`id`）
                    to_append: Any = key1
                elif i == 1: #偏好栏（`preferredItemSlot`）
                    to_append = value
                else: #名称（`name`）
                    to_append = self.LoLItems[key1]["name"] if key1 in self.LoLItems else ""
                preferredItemSlot_data[key].append(to_append)
        preferredItemSlot_statistics_output_order: list[int] = list(range(len(preferredItemSlot_header_keys)))
        preferredItemSlot_data_organized: dict[str, list[Any]] = {(preferredItemSlot_header_values if self.isZH else preferredItemSlot_header_keys)[i]: preferredItemSlot_data[preferredItemSlot_header_keys[i]] for i in preferredItemSlot_statistics_output_order}
        preferredItemSlot_df: pandas.DataFrame = pandas.DataFrame(preferredItemSlot_data_organized)
        return preferredItemSlot_df
    
    def edit_block(self, page_uid: str, pos: int, log: Optional[LogManager] = None) -> tuple[dict[str, Any], bool]:
        '''
        编辑一个装备区块。<br>Edit an item block.
        
        :param page_uid: 配装页的唯一识别码。用于定位到该区块。<br>Unique identifier of the item set page. Used to locate this block.
        :type page_uid: str
        :param pos: 配装页中的区块下标。用于定位到该区块。<br>The index of the block in the item set page. Used to locate this block.
        :type pos: int
        :param log: 日志管理类对象。如果未指定，则创建一个新的对象，用于内部使用。默认为None。<br>A LogManager object. If unspecified, a new object will be created and used internally. None by default.
        :type log: LogManager | None
        :return: 新装备区块信息，以及是否做出修改。<br>New item block information, plus whether a change has been made.
        :rtype: tuple[dict[str, Any], bool]
        '''
        if log == None:
            log = LogManager()
        if not verify_uuid(page_uid):
            raise TypeError("Page id isn't in UUID format.")
        logInput = log.logInput
        logPrint = log.logPrint
        #初始化装备区块相关变量（Initialize item block related variables）
        create: bool = False #标记是否新建区块（Marks whether to create a block）
        if page_uid in self._itemPage_dict:
            item_blocks: list[dict[str, Any]] = self._itemPage_dict[page_uid]["blocks"]
            if pos < -len(item_blocks) or pos >= len(item_blocks):
                create = True
                original_block: dict[str, Any] = {}
            else:
                original_block = item_blocks[pos]
        else: #创建新的配装页，即创建新的区块（Create a new item page, which means to create a new block）
            create = True
            original_block = {}
        spell_df_str: str = "id\tname"
        for spellId in spells:
            spell_df_str += "\n%d\t%s" %(spellId, self.spells[spellId]["name"])
        if create:
            title: str = "新的配装方案"
            showIfSummonerSpell: str = ""
            hideIfSummonerSpell: str = ""
            itemList: list[dict[str, str | int]] = []
        else:
            title = original_block["type"]
            showIfSummonerSpell = original_block["showIfSummonerSpell"]
            hideIfSummonerSpell = original_block["hideIfSummonerSpell"]
            itemList = original_block["items"]
        changed: bool = False #标记是否做出变化。当且仅当完成以下所有步骤后，该变量才会置为真（Marks whether a change has been made. When and only when all of the following steps are finished will this variable be set as True）
        #分步执行（Stepwise execution）
        step = 1
        while True:
            if step == 0:
                break
            elif step == 1:
                logPrint("第一步：请输入装备区块名称。\nStep 1: Please input the title of the item block.")
                title2: str = logInput()
                if title2 != "":
                    if title2[0] == chr(4):
                        step -= 2
                    else:
                        title = title2
            elif step == 2:
                logPrint("第二步：您希望该装备区块仅在选择什么召唤师技能时显示？\nStep 2: Which summoner spells do you wish this item block to show only when they're selected?\n%s\n示例（Examples）：\n\t#原设置（Original setting）\n11\t#惩戒（Smite）\n[4, 11]\t#闪现（Flash）、惩戒（Smite）" %(spell_df_str))
                while True:
                    spell_str: str = logInput()
                    if spell_str == "":
                        break
                    elif spell_str[0] == "0":
                        step -= 2
                        break
                    else:
                        try:
                            tmp = eval(spell_str)
                        except:
                            traceback_info = traceback.format_exc()
                            logPrint(traceback_info)
                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                        else:
                            if isinstance(tmp, int) and tmp in self.spells:
                                showIfSummonerSpell = str(tmp)
                                break
                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int) and x in self.spells)):
                                showIfSummonerSpell = ",".join(sorted(set(tmp)))
                                break
                            else:
                                logPrint("请输入一个正整数列表。\nPlease input a list of positive integers.")
            elif step == 3:
                logPrint("第三步：您希望该装备区块仅在选择什么召唤师技能时隐藏？\nStep 3: Which summoner spells do you wish this item block to hide only when they're selected?\n%s\n示例（Examples）：\n\t#原设置（Original setting）\n11\t#惩戒（Smite）\n[4, 11]\t#闪现（Flash）、惩戒（Smite）" %(spell_df_str))
                while True:
                    spell_str: str = logInput()
                    if spell_str == "":
                        break
                    elif spell_str[0] == "0":
                        step -= 2
                        break
                    else:
                        try:
                            tmp = eval(spell_str)
                        except:
                            traceback_info = traceback.format_exc()
                            logPrint(traceback_info)
                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                        else:
                            if isinstance(tmp, int) and tmp in self.spells:
                                hideIfSummonerSpell = str(tmp)
                                break
                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int) and x in self.spells)):
                                hideIfSummonerSpell = ",".join(sorted(set(tmp)))
                                break
                            else:
                                logPrint("请输入一个正整数列表。\nPlease input a list of positive integers.")
            elif step == 4: #目前装备列表仅支持完全覆盖写（Currently the item list only supports overwriting completely）
                logPrint("第四步：管理装备列表。\nStep 4: Manage the item list.")
                itemList_tmp: list[dict[str, str | int]] = copy.deepcopy(itemList)
                create_only: bool = len(itemList_tmp) == 0
                item_grid_df_tmp: pandas.DataFrame = self.sort_item_grid(itemList_tmp)
                if not create_only:
                    logPrint("当前装备列表如下：\nCurrent item list is as follows:")
                    logPrint(format_df(item_grid_df_tmp, print_index = True)[0], write_time = False)
                logPrint("请选择一个操作：\nPlease select an operation:\n-1\t返回上一步（Return to the last step）\n0\t完成（Finish）\n%s1\t添加装备（Add items）\n%s2\t移除装备（Remove items）\n%s3\t排序装备（Order items）\n%s4\t重写装备（Overwrite items）" %("☆" if create_only else "", "!" if create_only else "", "!" if create_only else "", "☆" if create_only else ""))
                while True:
                    option: str = logInput()
                    if option == "":
                        continue
                    elif option == "-1":
                        step -= 2
                        break
                    elif option[0] == "0":
                        itemList = itemList_tmp
                        break
                    elif option[0] == "1":
                        logPrint('请依次输入插入索引、装备序号和数量。输入“0”以结束输入。\nPlease input the insert index, item id and count. Submit "0" to cancel.\n示例（Examples）：\n(3, 2003, 5)\t#在3号位插入5瓶生命药水（Insert 5 Health Potions before the third place）\n(0, 4403)\t#在列表首位插入1个金铲铲（Insert The Golden Spatula at the start of the list）') #由于名称不唯一对应装备序号，所以本程序不可能声明通过输入装备名称来编辑装备列表的机制（Because names don't uniquely map to itemIds, it's impossible to declare a mechanism to edit the item list given a name）
                        while True:
                            tuple_got: bool = False
                            item_element_insertion: tuple[int, int, int] = (0, 0, 0) #初始化插入元组（Initialize the insertion tuple）
                            item_element_insertion_str: str = logInput()
                            if item_element_insertion_str == "":
                                continue
                            elif item_element_insertion_str[0] == "0":
                                break
                            else:
                                try:
                                    tmp = eval(item_element_insertion_str)
                                except:
                                    traceback_info = traceback.format_exc()
                                    logPrint(traceback_info)
                                    logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                else:
                                    if isinstance(tmp, tuple):
                                        if len(tmp) < 2:
                                            logPrint("请至少输入索引和装备序号。\nPlease input an index and an itemId at least.")
                                        elif len(tmp) <= 3:
                                            if len(tmp) == 2:
                                                item_element_insertion = tmp + (1,)
                                            else:
                                                item_element_insertion = tmp
                                            if item_element_insertion[1] in self.LoLItems:
                                                tuple_got = True
                                            else:
                                                logPrint("未找到装备序号为%d的装备。请切换一个装备后重试。\nItem with id %d not found. Please change an item and try again." %(item_element_insertion[1], item_element_insertion[1]))
                                        else:
                                            logPrint("元组元素个数应为2或者3个。\nThere should be 2 or 3 elements in the given tuple.")
                                    else:
                                        logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                            if tuple_got:
                                insertion_index: int = item_element_insertion[0]
                                itemId: int = item_element_insertion[1]
                                count: int = item_element_insertion[2]
                                itemList_tmp.insert(insertion_index, {"id": str(itemId), "count": count})
                                item_grid_df_tmp: pandas.DataFrame = self.sort_item_grid(itemList_tmp)
                                logPrint("当前装备列表如下：\nCurrent item list is as follows:")
                                logPrint(format_df(item_grid_df_tmp, print_index = True)[0], write_time = False)
                    elif option[0] == "2" or option[0] == "3":
                        if create_only:
                            logPrint("该操作目前不可用。\nThis operation isn't available currently.")
                        else:
                            logPrint("当前装备列表如下：\nCurrent item list is as follows:")
                            logPrint(format_df(item_grid_df_tmp, print_index = True)[0], write_time = False)
                            if option[0] == "2":
                                logPrint("请依次输入要删除的装备的索引。输入空字符串或Ctrl-D字符以结束输入。\nPlease input the indices of items to remove. Submit an empty string or the Ctrl-D character to cancel.\n示例（Examples）：\n0\t#删除第一个装备（Remove the first item）\n-1\t#删除最后一个装备（Remove the last item）\n[0, 1, -2]\t#删除前两个装备和倒数第二个装备（Remove the first two items and the second item to last）")
                                while True:
                                    index_got: bool = False
                                    remove_indices: list[int] = []
                                    remove_index_str: str = logInput()
                                    if remove_index_str == "" or remove_index_str[0] == chr(4):
                                        break
                                    else:
                                        try:
                                            tmp = eval(remove_index_str)
                                        except:
                                            traceback_info = traceback.format_exc()
                                            logPrint(traceback_info)
                                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                        else:
                                            if isinstance(tmp, int):
                                                remove_indices = [tmp]
                                                index_got = True
                                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int), tmp)):
                                                remove_indices = tmp[:]
                                                index_got = True
                                            else:
                                                logPrint("请输入由整数组成的索引列表。\nPlease input an index list of integers.")
                                    if index_got:
                                        #负索引转换成正索引（Transform negative indices into positive ones）
                                        for i in range(len(remove_indices)):
                                            if remove_indices[i] < 0:
                                                remove_indices[i] += len(itemList_tmp)
                                        #去重并排序（Remove duplication and sort）
                                        remove_indices = sorted(set(remove_indices))
                                        #删除元素（Remove elements）
                                        for i in range(len(remove_indices) - 1, -1, -1): #移除元素时倒序遍历（Reverse-order traversal to remove elements）
                                            if remove_indices[i] < len(itemList_tmp): #跳过越界的下标（Skip indices out of range）
                                                itemList_tmp.pop(remove_indices[i])
                                        create_only = len(itemList_tmp) == 0
                                        item_grid_df_tmp: pandas.DataFrame = self.sort_item_grid(itemList_tmp)
                                        if create_only:
                                            break
                                        else:
                                            logPrint("当前装备列表如下：\nCurrent item list is as follows:")
                                            logPrint(format_df(item_grid_df_tmp, print_index = True)[0], write_time = False)
                            elif option[0] == "3":
                                logPrint("是否对装备索引取子集？（输入任意非空字符串以开始打草稿，否则直接开始输入新索引列表。）\nDo you want to get a subset of the current item indices? (Submit any non-empty string to make a draft, or null to input the new index list directly.)")
                                draft_str: str = logInput()
                                draft: bool = bool(draft_str)
                                if draft:
                                    logPrint("请选择草稿选项：\nPlease select a draft option:\n0\t退出草稿（Quit drafting）\n1\t迭代式取子集（Take subsets iteratively）")
                                    while True:
                                        draft_option: str = logInput()
                                        if draft_option == "":
                                            continue
                                        elif draft_option[0] == "0":
                                            break
                                        elif draft_option[0] == "1":
                                            scope: dict[str, Any] = {"format_df": format_df, "sort_item_grid": self.sort_item_grid, "itemList_tmp": itemList_tmp, "LoLItems": self.LoLItems}
                                            logPrint('示例（Examples）：\nprint(dir())\nprint(format_df(itemList_tmp, print_index = True)[0])\nprint([i for (i, v) in sorted(enumerate(itemList_tmp), key = lambda x: LoLItems[x[1]]["priceTotal"], reverse = True)])\n输入“-1”以退出取子集。\nSubmit "-1" to quit taking subsets.')
                                            subscope(scope, log = log)
                                        else:
                                            logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                                        logPrint("请选择草稿选项：\nPlease select a draft option:\n0\t退出草稿（Quit drafting）\n1\t迭代式取子集（Take subsets iteratively）")
                                logPrint('请输入符合期望顺序的装备索引列表。\nPlease input a list of item indices that follows the expected order.\n示例（Examples）：\n[2, 1, 0]\t#将前三个装备反过来排列，并保持其后续元素的顺序（Reserve the order of the first three items while reserve the order of subsequent elements）\nlist(range(len(itemList_tmp) - 1, -1, -1))\t#取当前所有装备顺序的倒序（Take the reversed order of the current order of all elements）\n[i for (i, v) in sorted(enumerate(itemList_tmp), key = lambda x: LoLItems[x[1]]["priceTotal"], reverse = True)]\t#取所有元素按装备总价格倒序排列后的索引（Take the indices to arrange all elements in the descending order of total price）')
                                while True:
                                    itemList_ordered: list[dict[str, str | int]] = []
                                    new_index_list: list[int] = []
                                    index_got: bool = False
                                    new_index_str: str = logInput()
                                    if new_index_str == "":
                                        continue
                                    elif new_index_str[0] == "0": #当用户只输入“0”时，即将第一个装备放在第一位，相当于不做任何更改（When the user submits "0", the first item will put at the starting location of the original list, which means no change is made）
                                        break
                                    else:
                                        try:
                                            tmp = eval(new_index_str)
                                        except:
                                            traceback_info = traceback.format_exc()
                                            logPrint(traceback_info)
                                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                        else:
                                            if isinstance(tmp, int):
                                                if tmp < -len(itemList_tmp) or tmp >= len(itemList_tmp):
                                                    logPrint("下标越界。请重新输入。\nIndex out of range. Please try again.")
                                                else:
                                                    new_index_list = [tmp]
                                                    index_got = True
                                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int), tmp)):
                                                if any(map(lambda x: x < -len(itemList_tmp) or x >= len(itemList_tmp), tmp)):
                                                    logPrint("下标越界。请重新输入。\nIndex out of range. Please try again.")
                                                else:
                                                    new_index_list = tmp[:]
                                                    index_got = True
                                            else:
                                                logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                                    if index_got:
                                        #负索引转换成正索引（Transform negative indices into positive ones）
                                        for i in range(len(new_index_list)):
                                            if new_index_list[i] < 0:
                                                new_index_list[i] += len(itemList_tmp)
                                        #在新列表中添加按照顺序排列的装备（Add ordered items into the new list）
                                        for i in range(len(new_index_list)):
                                            itemList_ordered.append(itemList_tmp[new_index_list[i]])
                                        #将已添加的装备从旧列表中移除（Remove added items from the old list）
                                        for i in range(len(set(new_index_list))):
                                            itemList_tmp.pop(sorted(set(new_index_list), reverse = True)[i]) #倒序抛出元素（Pop elements in the reversed order）
                                        #将排列好的装备放到旧列表的前部（Add ordered items to the front of the old list）
                                        itemList_tmp = itemList_ordered + itemList_tmp
                                        break
                    elif option[0] == "4":
                        logPrint('请依次输入装备序号和数量。输入“0”以完成更改。输入“-1”以放弃更改。\nPlease input the item names or ids and counts one by one. Submit "0" to finish the change. Submit "-1" to give up the change.\n示例（Examples）：\n(2003, 5)\t#生命药水（Health Potion）：5\n(4403)\t#金铲铲（The Golden Spatula）')
                        itemList_new: list[dict[str, str | int]] = []
                        while True:
                            tuple_got: bool = False
                            item_element_appendix: tuple[int, int] = (0, 0) #初始化追加元组（Initialize the appendix tuple）
                            item_element_appendix_str: str = logInput()
                            if item_element_appendix_str == "":
                                continue
                            elif item_element_appendix_str == "-1":
                                break
                            elif item_element_appendix_str[0] == "0":
                                itemList_tmp = itemList_new
                                break
                            else:
                                try:
                                    tmp = eval(item_element_appendix_str)
                                except:
                                    traceback_info = traceback.format_exc()
                                    logPrint(traceback_info)
                                    logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                else:
                                    if isinstance(tmp, tuple):
                                        if len(tmp) == 0:
                                            logPrint("请至少输入装备序号。\nPlease input an itemId at least.")
                                        elif len(tmp) <= 2:
                                            if len(tmp) == 1:
                                                item_element_appendix = tmp + (1,)
                                            else:
                                                item_element_appendix = tmp
                                            if item_element_appendix[0] in self.LoLItems:
                                                tuple_got = True
                                            else:
                                                logPrint("未找到装备序号为%d的装备。请切换一个装备后重试。\nItem with id %d not found. Please change an item and try again." %(item_element_appendix[0], item_element_appendix[0]))
                                        else:
                                            logPrint("元组元素个数应为1或者2个。\nThere should be 1 or 2 elements in the given tuple.")
                                    else:
                                        logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                            if tuple_got:
                                itemId: int = item_element_appendix[0]
                                count: int = item_element_appendix[1]
                                itemList_new.append({"id": str(itemId), "count": count})
                                logPrint("新装备列表如下：\nNew item list is as follows:")
                                item_grid_df_new: pandas.DataFrame = self.sort_item_grid(itemList_new)
                                logPrint(format_df(item_grid_df_new, print_index = True)[0], write_time = False)
                    else:
                        logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                        continue
                    create_only: bool = len(itemList_tmp) == 0
                    item_grid_df_tmp: pandas.DataFrame = self.sort_item_grid(itemList_tmp)
                    if not create_only:
                        logPrint("当前装备列表如下：\nCurrent item list is as follows:")
                        logPrint(format_df(item_grid_df_tmp, print_index = True)[0], write_time = False)
                    logPrint("请选择一个操作：\nPlease select an operation:\n-1\t返回上一步（Return to the last step）\n0\t完成（Finish）\n%s1\t添加装备（Add items）\n%s2\t移除装备（Remove items）\n%s3\t排序装备（Order items）\n%s4\t重写装备（Overwrite items）" %("☆" if create_only else "", "!" if create_only else "", "!" if create_only else "", "☆" if create_only else ""))
            elif step == 5:
                changed = True
                break
            else:
                logPrint("步骤异常。请联系开发人员修复程序。\nStep error. Please contact the developer to fix the program.")
            step += 1
        block: dict[str, Any] = {"type": title, "hideIfSummonerSpell": hideIfSummonerSpell, "showIfSummonerSpell": showIfSummonerSpell, "items": itemList}
        return (block, changed)
    
    def edit_page(self, page_uid: str, log: Optional[LogManager] = None) -> tuple[dict[str, Any], bool]:
        '''
        编辑一个配装页。<br>Edit an item set page.
        
        :param page_uid: 配装页的唯一识别码。<br>Unique identifier of the item set page.
        :type page_uid: str
        :param log: 日志管理类对象。如果未指定，则创建一个新的对象，用于内部使用。默认为None。<br>A LogManager object. If unspecified, a new object will be created and used internally. None by default.
        :type log: LogManager | None
        :return: 新配装页信息，以及是否做出修改。<br>New item set page information, plus whether a change has been made.
        :rtype: tuple[dict[str, Any], bool]
        '''
        if log == None:
            log = LogManager()
        if not verify_uuid(page_uid):
            raise TypeError("Page id isn't in UUID format.")
        logInput = log.logInput
        logPrint = log.logPrint
        #初始化配装页相关变量（Initialize item set page related variables）
        if page_uid in self._itemPage_dict:
            original_page: dict[str, Any] = self._itemPage_dict[page_uid]
            title = original_page["title"]
            candidate_champions = original_page["associatedChampions"]
            candidate_maps = original_page["associatedMaps"]
            item_blocks = original_page["blocks"]
            preferredItemSlots = original_page["preferredItemSlots"]
        else:
            title: str = "新的配装方案" if self.isZH else "New Item Set"
            candidate_champions: list[int] = []
            candidate_maps: list[int] = []
            item_blocks: list[dict[str, Any]] = []
            preferredItemSlots: list[dict[str, Any]] = []
        preferredItemSlot_dict: dict[int, int] = {int(_["id"]): _["preferredItemSlot"] for _ in preferredItemSlots}
        changed: bool = False #标记是否做出变化。当且仅当完成以下所有步骤后，该变量才会置为真（Marks whether a change has been made. When and only when all of the following steps are finished will this variable be set as True）
        #分步执行（Stepwise execution）
        step: int = 1
        while True:
            if step == 0:
                break
            elif step == 1:
                logPrint("第一步：请输入配装方案的标题。\nStep 1: Please input the title of the item set.")
                title2: str = logInput()
                if title2 != "":
                    if title2[0] == chr(4):
                        step -= 2
                    else:
                        title = title2
            elif step == 2:
                logPrint("第二步：请选择相关英雄。\nStep 2: Please select the associated champions.")
                champion_df: pandas.DataFrame = sort_champion_summary(self.LoLChampions)
                champion_fields_to_print: list[str] = ["id", "name", "description", "alias"]
                logPrint(format_df(champion_df.loc[:, champion_fields_to_print])[0], write_time = False)
                logPrint("示例（Examples）：\n\t#无英雄（No champion）\n11\t#无极剑圣 易（Master Yi）\n[1, 2, 3, 5]\t#黑暗之女 安妮（Annie）、狂战士 奥拉夫（Olaf）、正义巨像 加里奥（Galio）、德邦总管 赵信（Xin Zhao）", write_time = False)
                while True:
                    candidate_champions_str: str = logInput()
                    if candidate_champions_str == "":
                        candidate_champions = []
                        break
                    elif candidate_champions_str[0] == "0":
                        candidate_champions = []
                        step -= 2
                        break
                    else:
                        try:
                            tmp = eval(candidate_champions_str)
                        except:
                            traceback_info = traceback.format_exc()
                            logPrint(traceback_info)
                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                        else:
                            if isinstance(tmp, int) and tmp in self.LoLChampions:
                                candidate_champions = [tmp]
                                break
                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int) and x in self.LoLChampions, tmp)):
                                candidate_champions = tmp
                                break
                            else:
                                logPrint("请输入由正整数组成的列表。\nPlease input a list of positive integers.")
            elif step == 3:
                logPrint("第三步：请选择相关地图。\nStep 3: Please select the associated maps.")
                gamemap_header: dict[str, str] = {"id": "序号", "name_zh": "名称", "name_en": "英文"}
                gamemap_header_keys: list[str] = list(gamemap_header.keys())
                gamemap_data: dict[str, list[Any]] = {key: [] for key in gamemap_header_keys}
                for mapId in gamemaps:
                    gamemap_data["id"].append(mapId)
                    gamemap_data["name_zh"].append(gamemaps[mapId]["zh_CN"])
                    gamemap_data["name_en"].append(gamemaps[mapId]["en_US"])
                gamemap_df: pandas.DataFrame = pandas.DataFrame(gamemap_data)
                logPrint(format_df(gamemap_df)[0], write_time = False)
                logPrint("示例（Example）：\n\t#不选择地图（Empty maps）\n11\t#召唤师峡谷（Summoner's Rift）\n[11, 12]\t#召唤师峡谷（Summoner's Rift）、随机地图（Random Map）", write_time = False)
                while True:
                    mapIds_str: str = logInput()
                    if mapIds_str == "":
                        candidate_maps: list[int] = []
                        break
                    elif mapIds_str[0] == "0":
                        candidate_maps = []
                        step -= 2
                        break
                    else:
                        try:
                            tmp = eval(mapIds_str)
                        except:
                            traceback_info = traceback.format_exc()
                            logPrint(traceback_info)
                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                        else:
                            if isinstance(tmp, int) and tmp in self.LoLChampions:
                                candidate_maps = [tmp]
                                break
                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int) and x in self.LoLChampions, tmp)):
                                candidate_maps = tmp
                                break
                            else:
                                logPrint("请输入由正整数组成的列表。\nPlease input a list of positive integers.")
            elif step == 4:
                logPrint("第四步：管理装备分区。\nStep 4: Manage the item blocks.")
                item_blocks_tmp: list[dict[str, Any]] = copy.deepcopy(item_blocks)
                create_only: bool = len(item_blocks_tmp) == 0
                item_block_df_tmp: pandas.DataFrame = self.sort_item_block(item_blocks_tmp)
                if not create_only:
                    logPrint("当前装备分区如下：\nCurrent item blocks are as follows:")
                    logPrint(format_df(item_block_df_tmp, print_index = True)[0], write_time = False)
                logPrint("请选择一个操作：\nPlease select an operation:\n-1\t返回上一步（Return to the last step）\n0\t完成（Finish）\n%s1\t添加区块（Add a block）\n%s2\t编辑区块（Edit a block）\n%s3\t删除区块（Delete blocks）\n%s4\t排序区块（Order blocks）" %("☆" if create_only else "", "!" if create_only else "", "!" if create_only else "", "!" if create_only else ""))
                while True:
                    option: str = logInput()
                    if option == "" and create_only:
                        option = "1"
                    if option == "":
                        continue
                    elif option == "-1":
                        step -= 2
                        break
                    elif option[0] == "0":
                        break
                    elif option[0] == "1":
                        item_block, created = self.edit_block(page_uid, len(item_blocks), log = log)
                        if created:
                            logPrint('您想要让这个新区块显示在第几位？（默认为末位。）\nWhich place do you want this new block to display at? (In the end by default.)\n注：第一位从“0”开始。\nNote: The index starts from "0".')
                            while True:
                                index_got: bool = False
                                item_block_index: int = 0
                                item_block_index_str: str = logInput()
                                if item_block_index_str == "":
                                    item_block_index = len(item_blocks_tmp)
                                    index_got = True
                                    break
                                elif item_block_index_str[0] == chr(4):
                                    index_got = False
                                    break
                                else:
                                    try:
                                        tmp = eval(item_block_index_str)
                                    except:
                                        traceback_info = traceback.format_exc()
                                        logPrint(traceback_info)
                                        logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                    else:
                                        if isinstance(tmp, int):
                                            item_block_index = tmp
                                            index_got = True
                                            break
                                        else:
                                            logPrint("请输入一个整数下标。\nPlease input an index of integer type.")
                            if index_got:
                                item_blocks_tmp.insert(item_block_index, item_block)
                    elif option[0] in {"2", "3", "4"}:
                        if create_only:
                            logPrint("该操作目前不可用。\nThis operation isn't available currently.")
                        else:
                            if option[0] == "2":
                                logPrint("请选择一个区块：\nPlease select a block:")
                                logPrint(format_df(item_block_df_tmp, print_index = True)[0], write_time = False)
                                while True:
                                    index_got: bool = False
                                    item_block_index: int = 0
                                    item_block_index_str: str = logInput()
                                    if item_block_index_str == "":
                                        continue
                                    elif item_block_index_str == chr(4):
                                        index_got = False
                                        break
                                    else:
                                        try:
                                            tmp = eval(item_block_index_str)
                                        except:
                                            traceback_info = traceback.format_exc()
                                            logPrint(traceback_info)
                                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                        else:
                                            if isinstance(tmp, int):
                                                if tmp < -len(item_blocks_tmp) or tmp >= len(item_blocks_tmp):
                                                    logPrint("区块下标越界。\nBlock index out of range.")
                                                else:
                                                    item_block_index = tmp
                                                    index_got = True
                                                    break
                                            else:
                                                logPrint("请输入一个整数下标。\nPlease input an index of integer type.")
                                if index_got:
                                    item_block, changed = self.edit_block(page_uid, item_block_index, log = log)
                                    if changed:
                                        item_blocks_tmp[item_block_index] = item_block
                            elif option[0] == "3":
                                logPrint("请依次输入要删除的装备区块的索引。输入空字符串或Ctrl-D字符以结束输入。\nPlease input the indices of item blocks to remove. Submit an empty string or the Ctrl-D character to cancel.\n示例（Examples）：\n0\t#删除第一个装备区块（Remove the first item block）\n-1\t#删除最后一个装备区块（Remove the last item block）\n[0, 1, -2]\t#删除前两个装备区块和倒数第二个装备区块（Remove the first two item blocks and the second item block to last）")
                                while True:
                                    index_got: bool = False
                                    remove_indices: list[int] = []
                                    remove_index_str: str = logInput()
                                    if remove_index_str == "" or remove_index_str[0] == chr(4):
                                        break
                                    else:
                                        try:
                                            tmp = eval(remove_index_str)
                                        except:
                                            traceback_info = traceback.format_exc()
                                            logPrint(traceback_info)
                                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                        else:
                                            if isinstance(tmp, int):
                                                remove_indices = [tmp]
                                                index_got = True
                                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int), tmp)):
                                                remove_indices = tmp[:]
                                                index_got = True
                                            else:
                                                logPrint("请输入由整数组成的索引列表。\nPlease input an index list of integers.")
                                    if index_got:
                                        #负索引转换成正索引（Transform negative indices into positive ones）
                                        for i in range(len(remove_indices)):
                                            if remove_indices[i] < 0:
                                                remove_indices[i] += len(item_blocks_tmp)
                                        #去重并排序（Remove duplication and sort）
                                        remove_indices = sorted(set(remove_indices))
                                        #删除元素（Remove elements）
                                        for i in range(len(remove_indices) - 1, -1, -1): #移除元素时倒序遍历（Reverse-order traversal to remove elements）
                                            if remove_indices[i] < len(item_blocks_tmp): #跳过越界的下标（Skip indices out of range）
                                                item_blocks_tmp.pop(remove_indices[i])
                                        create_only = len(item_blocks_tmp) == 0
                                        item_block_df_tmp: pandas.DataFrame = self.sort_item_block(item_blocks_tmp)
                                        if create_only:
                                            break
                                        else:
                                            logPrint("当前装备分区如下：\nCurrent item blocks are as follows:")
                                            logPrint(format_df(item_block_df_tmp, print_index = True)[0], write_time = False)
                            else:
                                logPrint('请输入符合期望顺序的装备区块索引列表。\nPlease input a list of item indices that follows the expected order.\n示例（Examples）：\n[2, 1, 0]\t#将前三个装备区块反过来排列，并保持其后续元素的顺序（Reserve the order of the first three item blocks while reserve the order of subsequent elements）\nlist(range(len(itemList_tmp) - 1, -1, -1))\t#取当前所有装备区块顺序的倒序（Take the reversed order of the current order of all elements）')
                                while True:
                                    item_blocks_ordered: list[dict[str, str | int]] = []
                                    new_index_list: list[int] = []
                                    index_got: bool = False
                                    new_index_str: str = logInput()
                                    if new_index_str == "":
                                        continue
                                    elif new_index_str[0] == "0": #当用户只输入“0”时，即将第一个装备放在第一位，相当于不做任何更改（When the user submits "0", the first item will put at the starting location of the original list, which means no change is made）
                                        break
                                    else:
                                        try:
                                            tmp = eval(new_index_str)
                                        except:
                                            traceback_info = traceback.format_exc()
                                            logPrint(traceback_info)
                                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                        else:
                                            if isinstance(tmp, int):
                                                if tmp < -len(item_blocks_tmp) or tmp >= len(item_blocks_tmp):
                                                    logPrint("下标越界。请重新输入。\nIndex out of range. Please try again.")
                                                else:
                                                    new_index_list = [tmp]
                                                    index_got = True
                                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int), tmp)):
                                                if any(map(lambda x: x < -len(item_blocks_tmp) or x >= len(item_blocks_tmp), tmp)):
                                                    logPrint("下标越界。请重新输入。\nIndex out of range. Please try again.")
                                                else:
                                                    new_index_list = tmp[:]
                                                    index_got = True
                                            else:
                                                logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                                    if index_got:
                                        #负索引转换成正索引（Transform negative indices into positive ones）
                                        for i in range(len(new_index_list)):
                                            if new_index_list[i] < 0:
                                                new_index_list[i] += len(item_blocks_tmp)
                                        #在新列表中添加按照顺序排列的装备（Add ordered items into the new list）
                                        for i in range(len(new_index_list)):
                                            item_blocks_ordered.append(item_blocks_tmp[new_index_list[i]])
                                        #将已添加的装备从旧列表中移除（Remove added items from the old list）
                                        for i in range(len(set(new_index_list))):
                                            item_blocks_tmp.pop(sorted(set(new_index_list), reverse = True)[i]) #倒序抛出元素（Pop elements in the reversed order）
                                        #将排列好的装备放到旧列表的前部（Add ordered items to the front of the old list）
                                        item_blocks_tmp = item_blocks_ordered + item_blocks_tmp
                                        break
                    else:
                        logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                        continue
                    item_blocks = copy.deepcopy(item_blocks_tmp)
                    create_only: bool = len(item_blocks_tmp) == 0
                    item_block_df_tmp: pandas.DataFrame = self.sort_item_block(item_blocks_tmp)
                    if not create_only:
                        logPrint("当前装备分区如下：\nCurrent item blocks are as follows:")
                        logPrint(format_df(item_block_df_tmp, print_index = True)[0], write_time = False)
                    logPrint("请选择一个操作：\nPlease select an operation:\n-1\t返回上一步（Return to the last step）\n0\t完成（Finish）\n1\t添加区块（Add a block）\n2\t编辑区块（Edit a block）\n3\t删除区块（Delete blocks）\n4\t排序区块（Order blocks）")
            elif step == 5:
                logPrint("第五步：设置装备偏好栏。\nStep 5: Set item slots.")
                preferredItemSlot_dict_tmp: dict[int, int] = preferredItemSlot_dict.copy()
                preferredItemSlot_df_tmp: pandas.DataFrame = self.sort_preferredItemSlot(preferredItemSlot_dict_tmp)
                if len(preferredItemSlot_dict_tmp) > 0:
                    logPrint("当前装备偏好栏如下：\nCurrent preferred item slots are as follows:")
                    logPrint(format_df(preferredItemSlot_df_tmp, print_index = True)[0], write_time = False)
                logPrint('''请依次输入装备序号及其偏好栏。输入“0”以完成更改。输入“-1”以返回上一步。\nPlease input the item names and preferred slots one by one. Submit "0" to finish the change. Submit "-1" to return to the last step.\n示例（Examples）：\n(1055, 0)\t#多兰之刃——1号位（Doran's Blade - First grid）\n(2055, 3)\t#控制守卫——5号位（Control Wards - Fifth grid）\n2003\t#移除生命药水的偏好栏（Remove Health Potion's preferred slot）''')
                while True:
                    tuple_got: bool = False
                    preferredItemSlot_item: tuple[int, int] = (0, 0) #初始化键值对元组（Initialize the key-value pair tuple）
                    preferredItemSlot_item_str: str = logInput()
                    if preferredItemSlot_item_str == "":
                        continue
                    elif preferredItemSlot_item_str == "-1":
                        step -= 2
                        break
                    elif preferredItemSlot_item_str[0] == "0":
                        preferredItemSlot_dict = preferredItemSlot_dict_tmp
                        break
                    else:
                        try:
                            tmp = eval(preferredItemSlot_item_str)
                        except:
                            traceback_info = traceback.format_exc()
                            logPrint(traceback_info)
                            logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                        else:
                            if isinstance(tmp, int):
                                if tmp in self.LoLItems:
                                    preferredItemSlot_item = (tmp, -1)
                                    tuple_got = True
                                else:
                                    logPrint("未找到装备序号为%d的装备。请切换一个装备后重试。\nItem with id %d not found. Please change an item and try again." %(tmp, tmp))
                            elif isinstance(tmp, tuple):
                                if len(tmp) == 0:
                                    logPrint("请至少输入装备序号。\nPlease input an itemId at least.")
                                elif len(tmp) <= 2:
                                    if len(tmp) == 1:
                                        preferredItemSlot_item = tmp + (-1,)
                                    else:
                                        preferredItemSlot_item = tmp
                                    if preferredItemSlot_item[0] in self.LoLItems:
                                        tuple_got = True
                                    else:
                                        logPrint("未找到装备序号为%d的装备。请切换一个装备后重试。\nItem with id %d not found. Please change an item and try again." %(preferredItemSlot_item[0], preferredItemSlot_item[0]))
                                else:
                                    logPrint("元组元素个数应为1或者2个。\nThere should be 1 or 2 elements in the given tuple.")
                            else:
                                logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                    if tuple_got:
                        itemId: int = preferredItemSlot_item[0]
                        preferredSlot: int = preferredItemSlot_item[1]
                        if preferredSlot >= 0 and preferredSlot < 6:
                            preferredItemSlot_dict_tmp[itemId] = preferredSlot
                        else:
                            if itemId in preferredItemSlot_dict_tmp:
                                del preferredItemSlot_dict_tmp[itemId]
                        preferredItemSlot_df_tmp: pandas.DataFrame = self.sort_preferredItemSlot(preferredItemSlot_dict_tmp)
                        if len(preferredItemSlot_dict_tmp) > 0:
                            logPrint("当前装备偏好栏如下：\nCurrent preferred item slots are as follows:")
                            logPrint(format_df(preferredItemSlot_df_tmp, print_index = True)[0], write_time = False)
                preferredItemSlots = [{"id": str(key), "preferredItemSlot": value} for (key, value) in preferredItemSlot_dict.items()]
            elif step == 6:
                changed = True
                break
            else:
                logPrint("步骤异常。请联系开发人员修复程序。\nStep error. Please contact the developer to fix the program.")
            step += 1
        itemPage: dict[str, Any] = {
            "uid": page_uid,
            "title": title,
            "mode": "any",
            "map": "any",
            "type": "custom",
            "sortrank": 0,
            "startedFrom": "blank",
            "associatedChampions": candidate_champions,
            "associatedMaps": candidate_maps,
            "blocks": item_blocks,
            "preferredItemSlots": preferredItemSlots
        }
        logPrint(json.dumps(itemPage, ensure_ascii = False))
        return (itemPage, changed)
    
    def delete_page(self, page_uid: str) -> bool:
        '''
        删除一个配装页。<br>Delete an item set page.
        
        在执行批量删除时，应倒序删除，即page_uid对应的索引应逐渐减小。<br>Batch removal should be performed in the reverse order. That is, the index of `page_uid` in `self.itemPage_uids` should get smaller gradually.
        
        :param page_uid: 配装页的唯一识别码。<br>Unique identifier of the item set page.
        :type page_uid: str
        :return: 删除状态。删除成功时返回真，否则返回假。<br>Deletion status. If deletion is success, return True; otherwise, return False.
        :rtype: bool
        '''
        if page_uid in self._itemPage_dict:
            delete_index: int = list(self._itemPage_dict.keys()).index(page_uid)
            self.itemSet["itemSets"].pop(delete_index)
            self.itemPage_uids.pop(delete_index)
            del self._itemPage_dict[page_uid]
            self.itemSet["timestamp"] = int(time.time() * 1000)
            return True
        else:
            return False
    
    def clear_page(self) -> None:
        '''
        清除所有配装页。<br>Delete all item set pages.
        '''
        self.itemSet["itemSets"] = []
        self.itemPages = []
        self.itemPage_uids = []
        self._itemPage_dict = {}
        self.itemSet["timestamp"] = int(time.time() * 1000)
    
    def replace_page(self, page_uid: str, item_page_new: dict[str, Any]) -> bool:
        '''
        将某个配装页整个直接替换为某个新的配装页。<br>Replace some whole item set page with a new one.
        
        :param page_uid: 配装页的唯一识别码。<br>Unique identifier of the item set page.
        :type page_uid: str
        :param item_page_new: 新配装页数据。必须符合配装页格式。<br>New item set page data. It must follow the format of an item set page.
        :type item_page_new: dict[str, Any]
        :return: 是否替换成功。<br>Whether the replace is successful.
        :rtype: bool
        '''
        if self.isItemPage(item_page_new):
            if page_uid in self._itemPage_dict:
                item_page_index: int = self.itemPage_uids.index(page_uid)
                self.itemSet["itemSets"][item_page_index] = copy.deepcopy(item_page_new)
                self.sync()
                self.itemSet["timestamp"] = int(time.time() * 1000)
                return True
            else:
                return False
        else:
            raise TypeError("Invalid format of item set page data.")
    
    def import_page(self, itemPage: Any) -> bool:
        '''
        导入一个配装页以快速添加配装方案。新增的配装页总是被追加到现有的配装页列表的最后。<br>Import an item set page to quickly supplement the item set. New item set pages are always appended to the end of the current list of item set pages.
        
        :param itemPage: 配装页。<br>An item set page.<br>该配装页的条件相较于构建类对象时没有那么严格。<br>The condition for this `itemPage` isn't so rigorous as when an object of this class is constructed.
        :type itemPage: Any
        :return: 导入是否成功。\nWhether the data has been imported successfully.
        :rtype: bool
        '''
        if isinstance(itemPage, dict) and all(key in itemPage for key in ["title", "associatedMaps", "associatedChampions", "blocks"]) and all(isinstance(itemPage[key], list) and all(map(lambda x: isinstance(x, int), itemPage[key])) for key in ["associatedChampions", "associatedMaps"]) and all(isinstance(itemPage[key], list) for key in ["blocks"]) and all(isinstance(itemPage[key], str) for key in ["title"]):
            new_itemPage: dict[str, Any] = {
                "uid": str(uuid.uuid4()),
                "title": itemPage["title"],
                "mode": "any",
                "map": "any",
                "type": "custom",
                "sortrank": 0,
                "startedFrom": "blank",
                "associatedChampions": itemPage["associatedChampions"],
                "associatedMaps": itemPage["associatedMaps"],
                "blocks": itemPage["blocks"],
                "preferredItemSlots": []
            }
            self.itemSet["itemSets"].append(new_itemPage)
            self.sync()
            self.itemSet["timestamp"] = int(time.time() * 1000)
            return True
        else:
            return False

async def manage_item_set(connection: Connection) -> None:
    '''
    管理配装方案。由此进入各个选项。<br>Manage the item set. Entry to each option.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    '''
    current_summoner: dict[str, Any] = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    current_summonerId: int = current_summoner["summonerId"]
    region_locale: dict[str, str] = await (await connection.request("GET", "/riotclient/region-locale")).json()
    isZH: bool = region_locale["locale"] in {"zh_CN", "zh_MY", "zh_TW"}
    logPrint("请选择一个选项：\nPlease select an option:\n0\t退出功能（Exit this function）\n1\t查看所有配装方案（Check the whole item set）\n2\t管理配装页（Manage item set pages）\n!3\t清空配装方案（Clear all item set）")
    while True:
        option = logInput()
        if option == "":
            continue
        elif option[0] == "0":
            break
        elif option[0] == "1":
            itemSet: dict[str, Any] = await (await connection.request("GET", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets")).json()
            itemSet_obj: ItemSet = ItemSet(itemSet, spells, LoLChampions, LoLItems, isZH = isZH)
            logPrint("请选择输出流：\nPlease select a stream to output:\n0\t返回上一层（Return to the last step）\n1\t终端（Terminal）\n2\t文件（File）")
            while True:
                stream: str = logInput()
                if stream == "":
                    continue
                elif stream[0] == "0":
                    break
                elif stream[0] == "1":
                    logPrint(itemSet_obj)
                    break
                elif stream[0] == "2":
                    with open("ItemSet.json", "w", encoding = "utf-8") as fp:
                        json.dump(itemSet, fp, indent = 4, ensure_ascii = False)
                    with open("ItemSet.txt", "w", encoding = "utf-8") as fp:
                        fp.write(str(itemSet_obj))
                    logPrint('配装方案已保存到工作目录下的“ItemSet.json”和“ItemSet.txt”中。\nItem set has been saved into "ItemSet.json" and "ItemSet.txt" under the working directory.')
                    break
                else:
                    logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
        elif option[0] == "2":
            item_page_fields_to_print: list[str] = ["title", "uid"]
            itemSet: dict[str, Any] = await (await connection.request("GET", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets")).json()
            itemSet_tmp: dict[str, Any] = copy.deepcopy(itemSet)
            itemSet_obj: ItemSet = ItemSet(itemSet_tmp, spells, LoLChampions, LoLItems, isZH = isZH)
            create_only: bool = len(itemSet_tmp["itemSets"]) == 0
            item_page_df: pandas.DataFrame = itemSet_obj.sort_item_page(itemSet_tmp["itemSets"])
            if create_only:
                logPrint("配装页信息如下：\nItem set pages are as follows:")
                logPrint(format_df(item_page_df.loc[:, item_page_fields_to_print], print_index = True)[0], write_time = False)
            logPrint("请选择一个操作：\nPlease select an operation:\n0\t返回上一层（Return to the last step）\n%s1\t查看配装页（Check a page）\n%s2\t导出配装页（Export pages）\n%s3\t添加配装页（Add a page）\n%s4\t导入配装页（Import a page）\n%s5\t编辑配装页（Edit a page）\n%s6\t删除配装页（Delete pages）" %("!" if create_only else "", "!" if create_only else "", "☆" if create_only else "", "☆" if create_only else "", "!" if create_only else "", "!" if create_only else ""))
            while True:
                suboption: str = logInput()
                if suboption == "":
                    continue
                elif suboption[0] == "0":
                    break
                elif suboption[0] == "1" or suboption[0] == "5":
                    if create_only:
                        logPrint("该操作目前不可用。\nThis operation isn't available currently.")
                    else:
                        logPrint("请选择一个配装页：\nPlease select an item set page:")
                        logPrint(format_df(item_page_df.loc[:, item_page_fields_to_print], print_index = True)[0], write_time = False)
                        while True:
                            index_got: bool = False
                            item_page_index: int = 0 #这里第零行是中文表头（Here the 0th line is Chinese headers）
                            item_page_index_str: str = logInput()
                            if item_page_index_str == "":
                                continue
                            elif item_page_index_str[0] == "0":
                                break
                            else:
                                try:
                                    tmp = eval(item_page_index_str)
                                except:
                                    traceback_info = traceback.format_exc()
                                    logPrint(traceback_info)
                                    logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                else:
                                    if isinstance(tmp, int):
                                        if tmp > 0 and tmp < len(item_page_df):
                                            item_page_index = tmp
                                            index_got = True
                                        else:
                                            logPrint("请输入一个正整数。\nPlease input a positive integer.")
                                    else:
                                        logPrint("请输入一个整数。\nPlease input an integer.")
                            if index_got:
                                page_uid: str = item_page_df["uid"][item_page_index]
                                if suboption[0] == "1":
                                    item_page_desc: str = itemSet_obj.format_page(page_uid)
                                    item_page: dict[str, Any] = itemSet_obj.get_item_page(page_uid)
                                    logPrint("请选择输出流：\nPlease select a stream to output:\n0\t返回上一层（Return to the last step）\n1\t终端（Terminal）\n2\t文件（File）")
                                    while True:
                                        stream: str = logInput()
                                        if stream == "":
                                            continue
                                        elif stream[0] == "0":
                                            break
                                        elif stream[0] == "1":
                                            logPrint(item_page_desc)
                                            break
                                        elif stream[0] == "2":
                                            item_page_title: str = item_page["title"]
                                            with open(f"ItemPage - {item_page_title}.json", "w", encoding = "utf-8") as fp:
                                                json.dump(item_page, fp, indent = 4, ensure_ascii = False)
                                            with open(f"ItemPage - {item_page_title}.txt", "w", encoding = "utf-8") as fp:
                                                fp.write(item_page_desc)
                                            logPrint(f'配装页{item_page_title}已保存到工作目录下的“ItemPage - {item_page_title}.json”和“ItemPage - {item_page_title}.txt”中。\nItem set page {item_page_title} has been saved into "ItemPage - {item_page_title}.json" and "ItemPage - {item_page_title}.txt" under the working directory.')
                                            break
                                        else:
                                            logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                                else:
                                    item_page_new, changed = itemSet_obj.edit_page(page_uid, log = log)
                                    if changed:
                                        item_page_index: int = itemSet_obj.itemPage_uids.index(page_uid)
                                        itemSet_tmp["itemSets"][item_page_index] = item_page_new
                                        itemSet_obj.update(itemSet_tmp)
                                        response: Optional[dict[str, Any]] = await (await connection.request("PUT", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets", data = itemSet_obj.itemSet)).json()
                                        logPrint(response)
                                break
                elif suboption[0] == "2" or suboption[0] == "6":
                    if create_only:
                        logPrint("该操作目前不可用。\nThis operation isn't available currently.")
                    else:
                        logPrint("请选择配装页：\nPlease select item set pages:")
                        logPrint(format_df(item_page_df.loc[:, item_page_fields_to_print], print_index = True)[0], write_time = False)
                        while True:
                            index_got: bool = False
                            item_page_indices: list[int] = []
                            item_page_indices_str: str = logInput()
                            if item_page_indices_str == "":
                                continue
                            elif item_page_indices_str[0] == "0":
                                break
                            elif item_page_indices_str == "all":
                                item_page_indices = list(range(1, len(item_page_df)))
                                index_got = True
                            else:
                                try:
                                    tmp = eval(item_page_indices_str)
                                except:
                                    traceback_info = traceback.format_exc()
                                    logPrint(traceback_info)
                                    logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                else:
                                    if isinstance(tmp, int):
                                        item_page_indices = [tmp]
                                        index_got = True
                                    elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int), tmp)):
                                        if len(tmp) != len(set(tmp)):
                                            logPrint("请输入不重复的索引。\nPlease input different indices.")
                                        elif any(map(lambda x: x <= 0, tmp)):
                                            logPrint("请输入正整数。\nPlease input positive integers.")
                                        elif any(map(lambda x: x >= len(item_page_df))):
                                            logPrint("请输入%d以内的正整数。\nPlease input positive integers less than %d." %(len(item_page_df), len(item_page_df)))
                                        else:
                                            item_page_indices = tmp
                                            index_got = True
                                    else:
                                        logPrint("请输入整数。\nPlease input integers.")
                            if index_got:
                                if suboption[0] == "2":
                                    logPrint('请输入导出路径。默认为“文档/配装方案.json”。\nPlease input the export path, "Documents/Item Set.json" by default.')
                                    export_path: str = logInput()
                                    if export_path == "":
                                        if isZH:
                                            export_path = os.path.expanduser("~/Documents/配装方案.json").replace("\\", "/")
                                        else:
                                            export_path = os.path.expanduser("~/Documents/Item Set.json").replace("\\", "/")
                                        break
                                    elif not export_path.endswith(".json"):
                                        export_path += ".json"
                                    os.makedirs(os.path.dirname(export_path), exist_ok = True)
                                    if len(item_page_indices) == 1:
                                        page_uid = item_page_df["uid"][item_page_indices[0]]
                                        item_page: dict[str, Any] = itemSet_obj.export_item_page(page_uid)
                                        try:
                                            with open(export_path, "w", encoding = "utf-8") as fp:
                                                json.dump(item_page, fp, indent = 4, ensure_ascii = False) #在客户端中，只导出一个配装页时，数据类型为字典而不是列表（In the League Client, when the user only exports one item set page, the exported data type is a dictionary instead of a list）
                                        except PermissionError:
                                            logPrint("拒绝访问。请检查路径的权限。\nPermission denied. Please check the permission of the path.")
                                        else:
                                            logPrint(f"配装方案已导出到{export_path}。\nItem set has been exported to {export_path}.")
                                    elif len(item_page_indices) > 1:
                                        item_pages: list[dict[str, Any]] = []
                                        for item_page_index in item_page_indices:
                                            page_uid = item_page_df["uid"][item_page_index]
                                            item_page = itemSet_obj.export_item_page(page_uid)
                                            item_pages.append(item_page)
                                        try:
                                            with open(export_path, "w", encoding = "utf-8") as fp:
                                                json.dump(item_pages, fp, indent = 4, ensure_ascii = False)
                                        except PermissionError:
                                            logPrint("拒绝访问。请检查路径的权限。\nPermission denied. Please check the permission of the path.")
                                        else:
                                            logPrint(f"配装方案已导出到{export_path}。\nItem set has been exported to {export_path}.")
                                else:
                                    for item_page_index in sorted(item_page_indices, reverse = True):
                                        page_uid = item_page_df["uid"][item_page_index]
                                        itemSet_obj.delete_page(page_uid)
                                    itemSet_tmp = itemSet_obj.itemSet
                                    response: Optional[dict[str, Any]] = await (await connection.request("PUT", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets", data = itemSet_tmp)).json()
                                    logPrint(response)
                                break
                elif suboption[0] == "3":
                    page_uid: str = str(uuid.uuid4())
                    while page_uid in itemSet_obj.itemPage_uids: #确保新建页使用的唯一识别码和已有页的唯一识别码不同（Make sure the new page's uid is different from the existing ones'）
                        page_uid = str(uuid.uuid4())
                    item_page_new, created = itemSet_obj.edit_page(page_uid, log = log)
                    if created:
                        if create_only:
                            item_page_index: int = 0
                            index_got: bool = True
                        else:
                            logPrint('您想要让这个新页面显示在第几位？（默认为末位。）\nWhich place do you want this new page to display at? (In the end by default.)\n注：第一位从“1”开始。\nNote: The index starts from "1".')
                            while True:
                                index_got = False
                                item_page_index = 0
                                item_page_index_str: str = logInput()
                                if item_page_index_str == "":
                                    item_page_index = len(itemSet_tmp["itemSets"])
                                    index_got = True
                                    break
                                elif item_page_index_str[0] == chr(4):
                                    index_got = False
                                    break
                                else:
                                    try:
                                        tmp = eval(item_page_index_str)
                                    except:
                                        traceback_info = traceback.format_exc()
                                        logPrint(traceback_info)
                                        logPrint("出现了一个异常信息。请检查您的输入。\nAn exception is thrown. Please check your input.")
                                    else:
                                        if isinstance(tmp, int):
                                            item_page_index = tmp - 1
                                            index_got = True
                                            break
                                        else:
                                            logPrint("请输入一个整数下标。\nPlease input an index of integer type.")
                        if index_got:
                            itemSet_tmp["itemSets"].insert(item_page_index, item_page_new)
                            itemSet_obj.update(itemSet_tmp)
                            response: Optional[dict[str, Any]] = await (await connection.request("PUT", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets", data = itemSet_obj.itemSet)).json()
                            logPrint(response)
                            break
                elif suboption[0] == "4":
                    item_pages_new: list[dict[str, Any]] = []
                    item_page_fromFile: Any = {}
                    logPrint("请输入配装方案文件路径。输入空字符串以返回上一层。\nPlease input the path of the item set file. Submit an empty string to return to the last step.")
                    while True:
                        file_read: bool = False #标记是否成功读取配装方案文件（Marks whether the program succeeds in reading the item set file）
                        import_path: str = logInput()
                        if import_path == "":
                            break
                        elif os.path.exists(import_path):
                            try:
                                with open(import_path, "r", encoding = "utf-8") as fp:
                                    item_page_fromFile = json.load(fp)
                            except json.decoder.JSONDecodeError:
                                logPrint("文件格式有误！请检查。\nFormat error! Please check it.")
                            else:
                                file_read = True
                                break
                        else:
                            logPrint("文件不存在！请重新输入。\nFile not found! Please try again.")
                    if file_read:
                        if isinstance(item_page_fromFile, dict):
                            item_pages_new = [item_page_fromFile]
                        elif isinstance(item_page_fromFile, list):
                            item_pages_new = item_page_fromFile
                        for item_page_new in item_pages_new:
                            imported: bool = itemSet_obj.import_page(item_page_new)
                            if imported:
                                itemSet_tmp = itemSet_obj.itemSet
                                response: Optional[dict[str, Any]] = await (await connection.request("PUT", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets", data = itemSet_obj.itemSet)).json()
                                logPrint(response)
                                logPrint("配装页%s导入成功。\nItem set page %s import success." %(item_page_new["title"], item_page_new["title"]))
                            else:
                                logPrint("配装页%s导入失败。请检查文件格式。\nItem set page %s import failure. Please check the file format." %(item_page_new["title"], item_page_new["title"]))
                else:
                    logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                    continue
                itemSet = await (await connection.request("GET", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets")).json()
                itemSet_tmp = copy.deepcopy(itemSet)
                itemSet_obj = ItemSet(itemSet_tmp, spells, LoLChampions, LoLItems, isZH = isZH)
                create_only = len(itemSet_tmp["itemSets"]) == 0
                item_page_df = itemSet_obj.sort_item_page(itemSet_tmp["itemSets"])
                if create_only:
                    logPrint("配装页信息如下：\nItem set pages are as follows:")
                    logPrint(format_df(item_page_df.loc[:, item_page_fields_to_print], print_index = True)[0], write_time = False)
                logPrint("请选择一个操作：\nPlease select an operation:\n0\t返回上一层（Return to the last step）\n%s1\t查看配装页（Check a page）\n%s2\t导出配装页（Export pages）\n%s3\t添加配装页（Add a page）\n%s4\t导入配装页（Import a page）\n%s5\t编辑配装页（Edit a page）\n%s6\t删除配装页（Delete pages）" %("!" if create_only else "", "!" if create_only else "", "☆" if create_only else "", "☆" if create_only else "", "!" if create_only else "", "!" if create_only else ""))
        elif option[0] == "3":
            itemSet: dict[str, Any] = await (await connection.request("GET", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets")).json()
            itemSet_tmp: dict[str, Any] = copy.deepcopy(itemSet)
            itemSet_obj: ItemSet = ItemSet(itemSet_tmp, spells, LoLChampions, LoLItems, isZH = isZH)
            if len(itemSet_tmp["itemSets"]) == 0:
                logPrint("你似乎没有任何自定义配装方案。点击下方的按钮来创建一套新的配装方案，从剪贴板或文件中导入，或者选择一个英雄和地图来获得一套推荐方案。\nYou don't seem to have any custom item sets. Click the buttons below to create a new one, import from clipboard or file, or choose a champion and map to get a recommended set.")
            else:
                logPrint("警告：清空配装方案是不可逆的。请提前做好备份。您确定要继续吗？（输入任意非空字符串以继续，否则取消。）\nWarning: Clearing the whole item set is irreversible. Please make backups in advance. Do you really want to continue? (Submit any non-empty string to continue, or null to cancel.)")
                continue_str: str = logInput()
                if bool(continue_str):
                    itemSet_obj.clear_page()
                    itemSet_tmp = itemSet_obj.itemSet
                    response: Optional[dict[str, Any]] = await (await connection.request("PUT", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets", data = itemSet_obj.itemSet)).json()
                    logPrint(response)
                    logPrint("配装方案已清空。\nItem set has been cleared.")
        elif option == "test": #隐藏功能（Hidden function）
            itemSet: dict[str, Any] = await (await connection.request("GET", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets")).json()
            itemSet_tmp: dict[str, Any] = copy.deepcopy(itemSet)
            itemSet_obj: ItemSet = ItemSet(itemSet_tmp, spells, LoLChampions, LoLItems, isZH = isZH)
            item_page_new, created, itemSet_filePath = create_test_itemPage(isZH = isZH)
            if created:
                item_page_index: int = len(itemSet_tmp["itemSets"])
                itemSet_tmp["itemSets"].insert(item_page_index, item_page_new)
                itemSet_obj.update(itemSet_tmp)
                response: Optional[dict[str, Any]] = await (await connection.request("PUT", f"/lol-item-sets/v1/item-sets/{current_summonerId}/sets", data = itemSet_obj.itemSet)).json()
                logPrint(response)
                if response == None:
                    logPrint("已添加全装备的配装页。\nAdded an item set page of all items.")
                else:
                    if response == {"errorCode": "BAD_REQUEST_HEADERS", "httpStatus": 413, "message": "Content length is too large"}:
                        logPrint(f"内容异常。请尝试手动从客户端导入以下配装方案文件：\nUnexpected content. Please try importing the following item set file into the client:\n{itemSet_filePath}")
                    else:
                        logPrint("在导入全装备配装方案时出现了一个问题。\nAn error occurred when the program was trying to export an item set of all items.")
            else:
                logPrint("在创建全装备配装方案时出现了一个问题。\nAn error occurred when the program was trying to create an item set of all items.")
        else:
            continue
        logPrint("请选择一个选项：\nPlease select an option:\n0\t退出功能（Exit this function）\n1\t查看所有配装方案（Check the whole item set）\n2\t管理配装页（Manage item set pages）\n!3\t清空配装方案（Clear all item set）")

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
async def connect(connection: Connection) -> None:
    await print_summoner_info(connection)
    await prepare_data_resources(connection)
    await manage_item_set(connection)

async def disconnect(connection: Connection) -> None:
    logPrint("已从英雄联盟客户端断开连接。\nDisconnected from the League Client.")

#-----------------------------------------------------------------------------
# 导出装备数据（Export item data）
#-----------------------------------------------------------------------------
def sort_item_ddragon(locale: str = "zh_CN") -> tuple[dict[str, pandas.DataFrame], bool]:
    '''
    整理DataDragon数据库的装备数据。可同时整理多个版本。<br>Organize item data in DataDragon database. Multiple versions can be organized at the same time.
    
    :param locale: 装备的语言文化代码。默认使用简体中文。<br>Langauge code of items. Chinese Simplified by default.
    :type locale: str
    :return: 二元组，包括以下内容：<br>A two-tuple that contains the following content:
    
        - 装备数据框字典。键是版本，值是装备数据框。<br>An item dataframe dictionary. Each key is a version, and each value is an item dataframe.
        - 是否形成了装备数据框，决定了是否导出到工作簿。<br>Whether an item dataframe is formed, which determines whether to open a workbook and export data into it.
    :rtype: tuple[dict[str, pandas.DataFrame], bool]
    '''
    item_df_formed: bool = False
    session: requests.Session = requests.Session()
    # session.trust_env = False
    LoLItem_dfs: dict[str, pandas.DataFrame] = {}
    while True:
        back: bool = False
        logPrint("请在以下版本号中选择并输入完整的版本号：\nPlease select a version and then enter it entirely:")
        versions, versionList_fetched = get_ddragon_versionList(session = session, log = log)
        if not versionList_fetched:
            break
        logPrint(json.dumps(versions, ensure_ascii = False))
        #下面的代码块生成英雄联盟专业术语中英文对照的Sheet1的内容。数据来源是DataDragon数据库（The following code block generates the Sheet1 in LoL Term Translation - zh_CN & en_US. Data resources are from DataDrabon database）
        while True:
            version: str = logInput()
            if version == "":
                versions_sort: list[str] = [versions[0]]
                break
            elif version == "0":
                back = True
                break
            elif version == "all":
                versions_sort = versions
                break
            elif version in versions:
                versions_sort = [version]
                break
            else:
                try:
                    tmp = eval(version)
                except:
                    logPrint("您的输入有误，请重新输入！\nERROR input! Please try again.")
                else:
                    if isinstance(tmp, list) and all(map(lambda x: x in versions, tmp)):
                        versions_sort = tmp
                        break
                    else:
                        logPrint("您的输入有误，请重新输入！\nERROR input! Please try again.")
        if back:
            break
        runTimes: list[float] = [] #记录整理一个版本的装备数据所花费的时间（Records the time spent in organizing item data of one version）
        total_used: float = 0
        failed_count: int = 0 #记录数据获取失败的版本的个数（Count the number of versions whose data fail to be fetched）
        for version_index in range(len(versions_sort)):
            start: float = time.time()
            logPrint("整理进度（Organization process）：%d/%d" %(version_index + 1, len(versions_sort)), print_time = True)
            version: str = versions_sort[version_index]
            LoLItems_locale_url: str = "https://ddragon.leagueoflegends.com/cdn/%s/data/%s/item.json" %(version, locale)
            LoLItems_default_url: str = "https://ddragon.leagueoflegends.com/cdn/%s/data/en_US/item.json" %version
            champions_locale_url: str = "https://ddragon.leagueoflegends.com/cdn/%s/data/%s/champion.json" %(version, locale)
            logPrint(f"正在获取{version}版本的目标语言装备信息……\nFetching LoL item information of version {version} in target language ...")
            source, status, session = requestUrl("GET", LoLItems_locale_url, session = session, log = log)
            if status != 200:
                if status == -1:
                    logPrint("目标语言装备信息获取失败！\nLoL item information in target language capture failed!")
                elif status == 403:
                    logPrint("目标语言装备信息文件不存在！\nLoL item file in target language not found!")
                failed_count += 1
                continue
            LoLItems_locale: dict[str, Any] = source.json()
            logPrint(f"正在获取{version}版本的英文装备信息……\nFetching LoL item information of version {version} in English ...")
            source, status, session = requestUrl("GET", LoLItems_default_url, session = session, log = log)
            if status != 200:
                if status == -1:
                    logPrint("英文装备信息获取失败！\nLoL item information in English capture failed!")
                elif status == 403:
                    logPrint("英文装备信息文件不存在！\nLoL item file in English not found!")
                failed_count += 1
                continue
            LoLItems_default: dict[str, Any] = source.json()
            logPrint(f"正在获取{version}版本的目标语言英雄信息……\nFetching champion information of version {version} in target language ...")
            source, status, session = requestUrl("GET", champions_locale_url, session = session, log = log)
            if status != 200:
                if status == -1:
                    logPrint("目标语言英雄信息获取失败！\nChampion information in target language capture failed!")
                elif status == 403:
                    logPrint("目标语言英雄信息文件不存在！\nChampion file in target language not found!")
                failed_count += 1
                continue
            champions_locale: dict[str, Any] = source.json()

            #下面设置装备表头的元数据部分（Set the metadata part of the item headers）
            item_base_header: dict[str, str] = {"id": "装备序号", "group": "分组", "description": "详细信息", "colloq": "检索关键字", "plaintext": "简述", "consumed": "消耗品", "stacks": "最大持有数量", "depth": "深度", "consumeOnFull": "满装备时自动消耗", "from": "合成材料序号", "into": "合成装备序号", "specialRecipe": "特殊合成材料", "inStore": "商店可见性", "hideFromAll": "不可见性", "requiredChampion": "装备持有者", "requiredAlly": "所需队友", "localizedName": "装备名称", "name": "英文名称", "fromName": "合成材料名称", "intoName": "合成装备名称", "requiredChampionName": "装备持有者名称", "requiredAllyName": "所需队友名称", "specialRecipeName": "特殊合成材料名称", "baseGold": "合成费用", "purchasable": "可以购买", "totalGold": "总费用", "sellGold": "售价"}
            item_base_header_keys: list[str] = list(item_base_header.keys())
            item_base_header_values: list[str] = list(item_base_header.values())
            #下面设置装备表头的分类（标签）部分（Set the category / tag part of the item headers）
            tags_initial: set[str] = set()
            for item in LoLItems_locale["data"].values():
                tags_initial |= set(map(lambda x: x.lower(), item["tags"])) #之所以要加lower（upper也可以），是因为在3.12.26版本以前，所有的标签/分类信息都是大写的（The reason why "lower" (or "upper") is needed is that all tags / categories before v3.12.26 are in upper case）
            tags_ordered: list[str] = sorted(tags_initial)
            tags_organized: list[str] = ["Lane", "Jungle", "GoldPer", "Boots", "Consumable", "Damage", "CriticalStrike", "AttackSpeed", "OnHit", "ArmorPenetration", "SpellDamage", "Mana", "ManaRegen", "MagicPenetration", "Health", "HealthRegen", "MagicResist", "AbilityHaste", "CooldownReduction", "Movement", "NonbootsMovement", "LifeSteal", "SpellVamp", "Active", "Armor", "Aura", "Slow", "SpellBlock", "Stealth", "Tenacity", "Trinket", "Vision", "Bilgewater"] #设置分类表头的顺序（Set the order of category headers）
            tags: list[str] = []
            for tag in tags_organized:
                if tag.lower() in tags_ordered:
                    tags_ordered.remove(tag.lower())
                tags.append(tag)
            tags += tags_ordered
            tags_dict: dict[str, str] = {"AbilityHaste": "技能急速", "Active": "主动", "Armor": "护甲", "ArmorPenetration": "护甲穿透", "AttackSpeed": "攻击速度", "Aura": "光环", "Bilgewater": "比尔吉沃特", "Boots": "鞋子", "Consumable": "消耗品", "CooldownReduction": "冷却缩减", "CriticalStrike": "暴击", "Damage": "攻击力", "GoldPer": "工资装", "Health": "生命值", "HealthRegen": "生命回复", "Jungle": "打野-起始", "Lane": "对线-起始", "LifeSteal": "生命偷取", "MagicPenetration": "法术穿透", "MagicResist": "魔法抗性", "Mana": "法力值", "ManaRegen": "法力回复", "Movement": "移动速度", "NonbootsMovement": "其它移动速度物品", "OnHit": "攻击特效", "Slow": "减速", "SpellBlock": "魔法抗性", "SpellDamage": "法术强度", "SpellVamp": "法术吸血", "Stealth": "潜行/隐身", "Tenacity": "韧性", "Trinket": "饰品", "Vision": "视野"}
            #下面设置装备表头的地图部分（Set the map part of the item headers）
            maps: dict[str, dict[str, str]] = {"8": {"zh_CN": "水晶之痕", "en_US": "Crystal Scar"}, "11": {"zh_CN": "召唤师峡谷", "en_US": "Summoner's Rift"}, "12": {"zh_CN": "嚎哭深渊", "en_US": "Howling Abyss"}, "14": {"zh_CN": "屠夫之桥", "en_US": "Butcher's Bridge"}, "16": {"zh_CN": "星界废墟", "en_US": "Cosmic Ruins"}, "18": {"zh_CN": "瓦洛兰城市公园", "en_US": "Valoran City Park"}, "19": {"zh_CN": "第43区", "en_US": "Substructure 43"}, "21": {"zh_CN": "百合与莲花的神庙", "en_US": "Temple of Lily and Lotus"}, "22": {"zh_CN": "聚点危机", "en_US": "Convergence"}, "30": {"zh_CN": "怒火角斗场", "en_US": "Rings of Wrath"}, "33": {"zh_CN": "最终都市", "en_US": "Final City"}, "35": {"zh_CN": "班德尔之森", "en_US": "The Bandlewood"}}
            mapIds: list[str] = list(map(str, sorted(map(int, maps.keys()))))
            #下面设置装备表头的基础属性部分。这一部分需要按照实际情况随时更新。只需要增添新的，不需要删除旧的（Set the stat part of the item headers. This part needs update with the latest knowledge. Only need to add new keys, but not delete old keys）
            attributes: dict[str, str] = {"Health": "生命值", "Bonus Health": "额外生命值", "Mana": "法力值", "Attack Damage": "攻击力", "Ability Power": "法术强度", "Adaptive Force": "适应之力", "Armor": "护甲", "Magic Resist": "魔法抗性", "Attack Speed": "攻击速度", "Ability Haste": "技能急速", "Cooldown Reduction": "冷却缩减", "Critical Strike Chance": "暴击几率", "Critical Strike Damage": "暴击伤害", "Move Speed": "移动速度", "Base Health Regen": "基础生命回复", "Base Mana Regen": "基础法力回复", "Heal and Shield Power": "治疗和护盾强度", "Increased Healing from Potions": "来自药水的治疗效果", "Mana per level": "每级法力", "Mana regen per 5 seconds": "法力回复/5秒", "Lethality": "穿甲", "Armor Penetration": "护甲穿透", "Magic Penetration": "法术穿透", "Life Steal": "生命偷取", "Omnivamp": "全能吸血", "Life Steal vs. Monsters": "对野怪的生命偷取", "Life on Hit": "攻击时回复生命值", "Tenacity": "韧性", "Gold Per 10 Seconds": "金币/10秒", "Ability Power per level": "每级法术强度"}
            attribute_correct_map: dict[str, str] = {"Base Health Regeneration": "Base Health Regen", "Mana per 5 seconds": "Mana regen per 5 seconds", "Movement Speed": "Move Speed"} #早期的装备数据中存在一些不规范的数值属性称呼，这里将其规范成以上字典中包含的属性（The early item data contain some irregular calling of attributes, and this dictionary is designed to standardize them to be included in the above `attributes` dictionary）
            #下面设置装备表头（Set the item headers）
            LoLItem_header_en: list[str] = item_base_header_keys + ["Map Availability: " + maps[mapId]["en_US"] for mapId in maps] + ["Class: " + tag for tag in tags] + list(attributes.keys())
            LoLItem_header_zh: list[str] = item_base_header_values + ["地图可用性：" + maps[mapId]["zh_CN"] for mapId in maps] + ["类别：" + tags_dict[tag] for tag in tags] + list(attributes.values())
            LoLItem_header: dict[str, str] = {LoLItem_header_en[i]: LoLItem_header_zh[i] for i in range(len(LoLItem_header_en))}
            LoLItem_header_keys: list[str] = list(LoLItem_header.keys())
            #print(LoLItem_header_keys)
            #定义常量字典（Define the constant dictionaries）
            LoLItem_name_map: dict[str, str] = {key: value["name"] for (key, value) in LoLItems_locale["data"].items()}
            logPrint("开始整理数据……\nOrganizing data ...")
            pStats: re.Pattern[str] = re.compile(r"<stats>.*</stats>")
            pFormat: re.Pattern[str] = re.compile(r"<[/\sA-Za-z0-9=#\'_@]*>")
            champions: dict[str, str] = {}
            for champion in champions_locale["data"]:
                champions[champion.lower()] = champions_locale["data"][champion]["name"] + " " + champions_locale["data"][champion]["title"] #装备数据中记录的英雄代号和英雄数据中的英雄代号有大小写上的差异（Case difference exists in the alias between item and champion data）
            LoLItem_data: dict[str, list[Any]] = {key: [] for key in LoLItem_header_keys}
            for i in LoLItems_locale["data"]:
                item: dict[str, Any] = LoLItems_locale["data"][i]
                item_default = LoLItems_default["data"][i]
                #首先处理共有部分（First, deal with the common part）
                ##下面填充装备的基本数据。这里参考的是英语描述（The following code fills the items' basic stats. Here the code refer to English descriptions）
                statDict: dict[str, str] = {}
                if "description" in item_default:
                    if pStats.search(item_default["description"]):
                        statStr = pStats.search(item_default["description"]).group().replace("<stats>", "").replace("</stats>", "").replace("<br>", "\n")
                        statList: list[str] = statStr.split("\n")
                    else: #在0.152.55版本以前，装备详细信息的数值部分没有被<stats>和</stats>标签包起来。其数值部分总是出现在第一行，并且不同的数值中间由空格分隔（无效信息），每个数值前都有加号【Before v0.152.22, the stat part of an item's description isn't enclosed by <stats> and </stats> tags. In the description, the stat part is always the first line, different stats are delimited by a space (useless information) and a plus sign is always in the front of the stat value】
                        pNonStat: re.Pattern[str] = re.compile(r"[^\s\+A-Za-z0-9]")
                        statStr = item_default["description"].replace("<br>", "\n").split("\n")[0]
                        if pNonStat.search(statStr) != None: #部分描述不规范，直接将其它非数值文字放在与数值同一行的位置。这里的处理方式是将所有非数值字符都当成分隔符，然后取第一个元素（Some descriptions don't obey the standard, because the nonstat descriptions are put in the same line as the stats. Here the strategy is to regard any nonstat character as a delimiter, and then get the first element of the string split by the delimiter）
                            statStr = statStr.split(pNonStat.search(statStr).group(0))[0]
                        statList_tmp = statStr.split("+")
                        statList = []
                        for stat_iter in statList_tmp:
                            if stat_iter != "":
                                statList.append(stat_iter.strip())
                    for stat_iter in statList:
                        if stat_iter != "": #有的装备没有基本属性，或者其字符串中存在几个连续的换行符（Some items don't have basic stats, or the string contains several continuous line feed characters）
                            # 下面注释起来的代码只适用于10.22版本后的装备数据（The following commented code only apply to item data after v10.22）
                            # figureType = pFormat.search(stat_iter).group().replace("<", "").replace(">", "")
                            # preFigure = "<" + figureType + ">"
                            # postFigure = "</" + figureType + ">"
                            # pFigure: re.Pattern[str] = re.compile(preFigure + ".*" + postFigure)
                            # figure = pFigure.search(stat_iter).group().replace(preFigure, "").replace(postFigure, "")
                            # figure_attr = stat_iter.replace(pFigure.search(stat_iter).group(), "").strip() #英文中，数值和属性之间有空格（In English, there's a space between the stat and the attribute）
                            # statDict[figure_attr] = figure
                            while (matchObj := pFormat.search(stat_iter)):
                                stat_iter = stat_iter.replace(matchObj.group(), "")
                            pFigure: re.Pattern[str] = re.compile(r"(\+|\-)?[0-9]+%?")
                            try:
                                figure: str = pFigure.search(stat_iter).group()
                                figure_attr: str = stat_iter.replace(figure, "").strip() #英文中，数值和属性之间有空格（In English, there's a space between the stat and the attribute）
                                if figure_attr in attribute_correct_map:
                                    figure_attr = attribute_correct_map[figure_attr]
                                statDict[figure_attr] = figure.replace("+", "")
                            except AttributeError:
                                pass
                #然后分类讨论（Then discuss about `j`)
                for j in range(len(LoLItem_header_keys)):
                    key: str = LoLItem_header_keys[j]
                    if j < len(item_base_header_keys): #基本表头部分（Base part）
                        if j == 0: #键（Key）
                            try:
                                to_append: Any = int(i)
                            except ValueError: #在12.21.1版本出现了装备序号为TalentReaperItem的装备（An item with itemId "TalentReaperItem" appears in the item data of v12.21.1）
                                traceback_info = traceback.format_exc()
                                logPrint(traceback_info)
                                to_append = i
                        elif j <= 22:
                            if j == 2: #详细信息（`description`）
                                desc: str = item["description"].replace("<br>", "\n")
                                while (matchObj := pFormat.search(desc)):
                                    desc = desc.replace(matchObj.group(), "")
                                to_append = desc
                            elif j == 16: #装备名称（`localizedName`）
                                to_append = item["name"]
                            elif j == 17: #英文名称（`name`）
                                to_append = item_default["name"]
                            elif j == 18: #合成材料名称（`fromName`）
                                to_append = list(map(lambda x: LoLItem_name_map.get(x, x), item["from"])) if "from" in item else ""
                            elif j == 19: #合成装备名称（`intoName`）
                                to_append = list(map(lambda x: LoLItem_name_map.get(x, x), item["into"])) if "into" in item else "" #吞噬者仅出现在合成装备中（Devourer only occurs as an item to upgrade into）
                            elif j == 20: #装备持有者名称（`requiredChampionName`）
                                to_append = champions[item["requiredChampion"].lower()] if "requiredChampion" in item and item["requiredChampion"] != "" else ""
                            elif j == 21: #所需队友名称（`requiredAllyName`）
                                to_append = champions[item["requiredAlly"].lower()] if "requiredAlly" in item and item["requiredAlly"] != "" else ""
                            elif j == 22: #特殊合成材料名称（`specialRecipeName`）
                                to_append = LoLItem_name_map.get(item["specialRecipe"], "") if "specialRecipe" in item and item["specialRecipe"] != 0 else ""
                            elif j in {5, 8, 12, 13, 24}: #逻辑值（Values of boolean type）
                                to_append = item.get(key, False)
                            else:
                                to_append = item.get(key, "")
                        else:
                            if j == 23: #合成费用（`baseGold`）
                                to_append = item["gold"]["base"]
                            elif j == 24: #可以购买（`purchasable`）
                                to_append = item["gold"]["purchasable"]
                            elif j == 25: #总费用（`totalGold`）
                                to_append = item["gold"]["total"]
                            else: #售价（`sellGold`）
                                to_append = item["gold"]["sell"]
                    elif j < len(item_base_header_keys) + len(mapIds): #地图部分（Map part）
                        to_append = item["maps"].get(mapIds[j - len(item_base_header_keys)], False) if "maps" in item else False
                    elif j < len(item_base_header_keys) + len(mapIds) + len(tags): #分类部分（Category part）
                        to_append = tags[j - len(item_base_header_keys) - len(mapIds)] in item["tags"]
                    else: #基础属性部分（Stat part）
                        if "description" in item_default:
                            key_default: str = list(attributes.keys())[j - len(item_base_header_keys) - len(mapIds) - len(tags)]
                            to_append = statDict.get(key_default, "")
                        else:
                            to_append = ""
                    LoLItem_data[key].append(to_append)
            base_statistics_display_order: list[int] = [0, 16, 17, 4, 5, 8, 12, 13, 24, 23, 25, 26, 6, 7, 20, 21, 18, 19, 22, 3, 2]
            LoLItem_statistics_display_order: list[int] = base_statistics_display_order + list(range(len(base_statistics_display_order), len(LoLItem_header)))
            LoLItem_data_organized: dict[str, list[Any]] = {LoLItem_header_keys[i]: LoLItem_data[LoLItem_header_keys[i]] for i in LoLItem_statistics_display_order}
            LoLItem_df: pandas.DataFrame = pandas.DataFrame(data = LoLItem_data_organized)
            optimize_bool_display(LoLItem_df)
            LoLItem_df.index = list(range(1, len(LoLItem_df) + 1))
            try:
                LoLItem_df = LoLItem_df.sort_values(by = "id", ascending = True)
            except TypeError: #在12.21.1版本出现了装备序号为TalentReaperItem的装备，导致无法正常按照装备序号排序（An item with itemId "TalentReaperItem" appears in the item data of v12.21.1, which makes it disallowed to sort the values by the itemIds）
                traceback_info = traceback.format_exc()
                logPrint(traceback_info)
            LoLItem_df = pandas.concat([pandas.DataFrame([LoLItem_header])[LoLItem_df.columns], LoLItem_df])
            LoLItem_dfs[version] = LoLItem_df.copy(deep = True)
            item_df_formed = True
            end: float = time.time()
            unit: float = end - start
            total_used += unit
            runTimes.append(unit)
            total_remaining: float = sum(runTimes) / (version_index + 1 - failed_count) * (len(versions_sort) - version_index - 1)
            logPrint("整理该版本数据所花费的时间（Time spent in organizing this version）：", format_runtime(unit))
            logPrint("已花费的总时间（Total time used）                                  ：", format_runtime(total_used))
            logPrint("剩余时间（Time remaining）                                         ：", format_runtime(total_remaining))
            logPrint("预计总时间（Expected total time）                                  ：", format_runtime(total_used + total_remaining), end = "\n\n")
    return (LoLItem_dfs, item_df_formed)

def sort_item_cdragon(locale: str = "zh_CN") -> tuple[dict[str, pandas.DataFrame], bool]:
    '''
    整理CommunityDragon数据库的装备数据。可同时整理多个版本。<br>Organize item data in CommunityDragon database. Multiple versions can be organized at the same time.
    
    :param locale: 装备的语言文化代码。默认使用简体中文。<br>Langauge code of items. Chinese Simplified by default.
    :type locale: str
    :return: 二元组，包括以下内容：<br>A two-tuple that contains the following content:
    
        - 装备数据框字典。键是版本，值是装备数据框。<br>An item dataframe dictionary. Each key is a version, and each value is an item dataframe.
        - 是否形成了装备数据框，决定了是否导出到工作簿。<br>Whether an item dataframe is formed, which determines whether to open a workbook and export data into it.
    :rtype: tuple[dict[str, pandas.DataFrame], bool]
    '''
    item_df_formed: bool = False
    session: requests.Session = requests.Session()
    # session.trust_env = False
    LoLItem_dfs: dict[str, pandas.DataFrame] = {}
    while True:
        back: bool = False
        logPrint("请在以下版本号中选择并输入完整的版本号：\nPlease select a version and then enter it entirely:")
        patches_cdragon, patchList_fetched = get_cdragon_patchList(session = session, log = log)
        if not patchList_fetched:
            break
        logPrint(json.dumps(patches_cdragon, ensure_ascii = False))
        #下面的代码块生成英雄联盟专业术语中英文对照的Sheet1的内容。数据来源是CommunityDragon数据库（The following code block generates the Sheet1 in LoL Term Translation - zh_CN & en_US. Data resources are from CommunityDragon database）
        while True:
            version: str = logInput()
            if version == "":
                versions_sort: list[str] = ["pbe"]
                break
            elif version == "both":
                versions_sort = ["latest", "pbe"]
                break
            elif version == "all":
                versions_sort = patches_cdragon
                break
            elif version[0] == "0":
                versions_sort = []
                back = True
                break
            elif version in patches_cdragon:
                versions_sort = [version]
                break
            else:
                try:
                    tmp = eval(version)
                except:
                    logPrint("您的输入有误，请重新输入！\nERROR input! Please try again.")
                else:
                    if isinstance(tmp, list) and all(map(lambda x: x in patches_cdragon, tmp)):
                        versions_sort = tmp
                        break
                    else:
                        logPrint("您的输入有误，请重新输入！\nERROR input! Please try again.")
        if back:
            break
        runTimes: list[float] = [] #记录整理一个版本的装备数据所花费的时间（Records the time spent in organizing item data of one version）
        total_used: float = 0
        failed_count: int = 0 #记录数据获取失败的版本的个数（Count the number of versions whose data fail to be fetched）
        for version_index in range(len(versions_sort)):
            start: float = time.time()
            logPrint("整理进度（Organization process）：%d/%d" %(version_index + 1, len(versions_sort)), print_time = True)
            version: str = versions_sort[version_index]
            LoLItems_locale_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/items.json" %(version, language_cdragon[locale])
            LoLItems_default_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/default/v1/items.json" %version
            champions_locale_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/champion-summary.json" %(version, language_cdragon[locale])
            fontconfig_locale_url: str = "https://raw.communitydragon.org/%s/game/data/menu/fontconfig_%s.txt.json" %(version, locale.lower())
            strtable_locale_url1: str = "https://raw.communitydragon.org/%s/game/data/menu/main_%s.stringtable.json" %(version, locale.lower())
            strtable_locale_url2: str = "https://raw.communitydragon.org/%s/game/%s/data/menu/en_us/main.stringtable.json" %(version, locale.lower())
            strtable_locale_url3: str = "https://raw.communitydragon.org/%s/game/%s/data/menu/en_us/lol.stringtable.json" %(version, locale.lower())
            LoLItems_binary_url1: str = "https://raw.communitydragon.org/%s/game/global/items/items.bin.json" %(version)
            LoLItems_binary_url2: str = "https://raw.communitydragon.org/%s/game/items.cdtb.bin.json" %(version)
            logPrint(f"正在获取{version}版本的目标语言装备信息……\nFetching LoL item information of version {version} in target language ...")
            source, status, session = requestUrl("GET", LoLItems_locale_url, session = session, log = log)
            if status != 200:
                if status == -1:
                    logPrint("目标语言装备信息获取失败！\nLoL item information in target language capture failed!")
                elif status == 404:
                    logPrint("目标语言装备信息文件不存在！\nLoL item file in target language not found!")
                failed_count += 1
                continue
            LoLItems_locale: list[dict[str, Any]] = source.json()
            logPrint(f"正在获取{version}版本的英文装备信息……\nFetching LoL item information of version {version} in English ...")
            source, status, session = requestUrl("GET", LoLItems_default_url, session = session, log = log)
            if status != 200:
                if status == -1:
                    logPrint("英文装备信息获取失败！\nLoL item information in English capture failed!")
                elif status == 404:
                    logPrint("英文装备信息文件不存在！\nLoL item file in English not found!")
                failed_count += 1
                continue
            LoLItems_default: list[dict[str, Any]] = source.json()
            logPrint(f"正在获取{version}版本的目标语言英雄信息……\nFetching champion information of version {version} in target language ...")
            source, status, session = requestUrl("GET", champions_locale_url, session = session, log = log)
            if status != 200:
                if status == -1:
                    logPrint("目标语言英雄信息获取失败！\nChampion information in target language capture failed!")
                elif status == 404:
                    logPrint("目标语言英雄信息文件不存在！\nChampion file in target language not found!")
                failed_count += 1
                continue
            champions_locale: list[dict[str, Any]] = source.json()
            logPrint(f"正在获取{version}版本的目标语言字符串常量池（英雄联盟）……\nFetching stringtable (LoL) of version {version} in target language ...")
            source, status, session = requestUrl("GET", fontconfig_locale_url if Patch(version) < Patch("12.23") else strtable_locale_url1 if Patch(version) < Patch("14.4") else strtable_locale_url2 if Patch(version) < Patch("14.15") else strtable_locale_url3, session = session, log = log) #翻译数据在12.23、14.4和14.15版本发生了路径迁移（Path transfer occurred to the localization data in Patches 12.23, 14.4 and 14.15）
            status_strtable_locale: int = status
            if status != 200:
                if status == -1:
                    logPrint("目标语言字符串常量池（英雄联盟）获取失败！\nStringtable (LoL) in target language capture failed!")
                elif status == 404:
                    logPrint("目标语言字符串常量池（英雄联盟）文件不存在！\nStringtable (LoL) file in target language not found!")
            if status_strtable_locale == 200:
                strtable_locale: dict[str, int | dict[str, str]] = source.json()
            elif status_strtable_locale == 404:
                strtable_locale = {}
            else:
                failed_count += 1
                continue
            logPrint(f"正在获取{version}版本的二进制装备信息……\nFetching binary LoL item of version {version} ...")
            source, status, session = requestUrl("GET", LoLItems_binary_url1 if Patch(version) < Patch("13.15") else LoLItems_binary_url2, session = session, log = log) #二进制装备信息在13.15版本发生了路径迁移（Path transfer occurred to the binary item information in Patch 13.15）
            status_item_binary: int = status
            if status != 200:
                if status == -1:
                    logPrint("目标语言二进制装备信息获取失败！\nBinary LoL item information capture failed!")
                elif status == 404:
                    logPrint("目标语言二进制装备信息文件不存在！\nBinary LoL item file not found!")
            if status_item_binary == 200:
                LoLItems_binary: dict[str, list[str] | dict[str, Any]] = source.json()
            elif status_item_binary == 404:
                LoLItems_binary = {}
            else:
                failed_count += 1
                continue
            
            #下面设置装备表头的元数据部分（Set the metadata part of the item headers）
            item_base_header: dict[str, str] = {"id": "装备序号", "active": "主动使用", "description": "描述", "inStore": "游戏内可见性", "from": "合成材料序号", "to": "合成装备序号", "maxStacks": "最大持有数量", "requiredChampion": "装备持有者", "requiredAlly": "所需队友", "requiredBuffCurrencyName": "其它货币类型", "requiredBuffCurrencyCost": "其它费用", "specialRecipe": "特殊合成材料", "isEnchantment": "附魔装备", "price": "合成费用", "priceTotal": "总费用", "displayInItemSets": "装备图册可见性", "iconPath": "缩略图路径", "localizedName": "装备名称", "name": "英文名称", "fromName": "合成材料名称", "toName": "合成装备名称", "requiredChampionName": "装备持有者名称", "requiredAllyName": "所需队友名称", "specialRecipeName": "特殊合成材料名称"}
            item_base_header_keys: list[str] = list(item_base_header.keys())
            item_base_header_values: list[str] = list(item_base_header.values())
            #下面设置装备表头的详细信息部分（Set the tooltip part of the item headers）
            tooltip_header: dict[str, str] = {"tooltip": "游戏内详细信息", "tooltip_burn": "游戏内详细信息（数值转换）"}
            tooltip_header_keys: list[str] = list(tooltip_header.keys())
            tooltip_header_values: list[str] = list(tooltip_header.values())
            #下面设置装备表头的分类（标签）部分（Set the category / tag part of the item headers）
            categories_initial_set: set[str] = set()
            for item in LoLItems_locale:
                categories_initial_set |= set(map(lambda x: x.lower(), item.get("categories", []))) #在7.8版本以前，装备数据中无“categories”键（Before Patch 7.8, "categories" key isn't present in item data）
            categories_initial: list[str] = sorted(categories_initial_set)
            categories_organized: list[str] = ["Lane", "Jungle", "GoldPer", "Boots", "Consumable", "Damage", "CriticalStrike", "AttackSpeed", "OnHit", "ArmorPenetration", "SpellDamage", "Mana", "ManaRegen", "MagicPenetration", "Health", "HealthRegen", "MagicResist", "AbilityHaste", "CooldownReduction", "Movement", "NonbootsMovement", "LifeSteal", "SpellVamp", "Active", "Armor", "Aura", "Slow", "SpellBlock", "Stealth", "Tenacity", "Trinket", "Vision", "Bilgewater"] #设置分类表头的顺序（Set the order of category headers）
            categories: list[str] = []
            for category in categories_organized:
                if category.lower() in categories_initial:
                    categories_initial.remove(category.lower())
                categories.append(category)
            categories += categories_initial
            categories_dict: dict[str, str] = {"AbilityHaste": "技能急速", "Active": "主动", "Armor": "护甲", "ArmorPenetration": "护甲穿透", "AttackSpeed": "攻击速度", "Aura": "光环", "Bilgewater": "比尔吉沃特", "Boots": "鞋子", "Consumable": "消耗品", "CooldownReduction": "冷却缩减", "CriticalStrike": "暴击", "Damage": "攻击力", "GoldPer": "工资装", "Health": "生命值", "HealthRegen": "生命回复", "Jungle": "打野-起始", "Lane": "对线-起始", "LifeSteal": "生命偷取", "MagicPenetration": "法术穿透", "MagicResist": "魔法抗性", "Mana": "法力值", "ManaRegen": "法力回复", "Movement": "移动速度", "NonbootsMovement": "其它移动速度物品", "OnHit": "攻击特效", "Slow": "减速", "SpellBlock": "魔法抗性", "SpellDamage": "法术强度", "SpellVamp": "法术吸血", "Stealth": "潜行/隐身", "Tenacity": "韧性", "Trinket": "饰品", "Vision": "视野"}
            #下面设置装备表头的基础属性部分。这一部分需要按照实际情况随时更新。只需要增添新的，不需要删除旧的（Set the basic stat part of the item headers. This part needs update with the latest knowledge. Only need to add new keys, but not delete old keys）
            attributes: dict[str, str] = {"Health": "生命值", "Bonus Health": "额外生命值", "Mana": "法力值", "Attack Damage": "攻击力", "Ability Power": "法术强度", "Adaptive Force": "适应之力", "Armor": "护甲", "Magic Resist": "魔法抗性", "Attack Speed": "攻击速度", "Ability Haste": "技能急速", "Cooldown Reduction": "冷却缩减", "Critical Strike Chance": "暴击几率", "Critical Strike Damage": "暴击伤害", "Move Speed": "移动速度", "Base Health Regen": "基础生命回复", "Base Mana Regen": "基础法力回复", "Heal and Shield Power": "治疗和护盾强度", "Increased Healing from Potions": "来自药水的治疗效果", "Mana per level": "每级法力", "Mana regen per 5 seconds": "法力回复/5秒", "Lethality": "穿甲", "Armor Penetration": "护甲穿透", "Magic Penetration": "法术穿透", "Life Steal": "生命偷取", "Omnivamp": "全能吸血", "Life Steal vs. Monsters": "对野怪的生命偷取", "Life on Hit": "攻击时回复生命值", "Tenacity": "韧性", "Gold Per 10 Seconds": "金币/10秒", "Ability Power per level": "每级法术强度"}
            attribute_correct_map: dict[str, str] = {"Base Health Regeneration": "Base Health Regen", "Mana per 5 seconds": "Mana regen per 5 seconds", "Movement Speed": "Move Speed"} #早期的装备数据中存在一些不规范的数值属性称呼，这里将其规范成以上字典中包含的属性（The early item data contain some irregular calling of attributes, and this dictionary is designed to standardize them to be included in the above `attributes` dictionary）
            #下面设置装备表头的所有数值部分。这一部分需要按照实际情况随时更新。只需要增添新的，不需要删除旧的（Set the detailed stat part of the item headers. This part needs update with the latest knowledge. Only need to add new keys, but not delete old keys）
            allStats_header: dict[str, str] = {}
            allStats_header_keys: list[str] = list(allStats_header.keys())
            allStats_header_values: list[str] = list(allStats_header.values())
            #下面设置装备表头（Set the item headers）
            LoLItem_header_en: list[str] = item_base_header_keys + tooltip_header_keys + ["Class: " + category for category in categories] + list(attributes.keys()) + (allStats_header_keys if status_item_binary == 200 else [])
            LoLItem_header_zh: list[str] = item_base_header_values + tooltip_header_values + ["类别：" + categories_dict[category] for category in categories] + list(attributes.values()) + (allStats_header_values if status_item_binary == 200 else [])
            LoLItem_header: dict[str, str] = {LoLItem_header_en[i]: LoLItem_header_zh[i] for i in range(len(LoLItem_header_en))}
            LoLItem_header_keys: list[str] = list(LoLItem_header.keys())
            #print(LoLItem_header_keys)
            #定义常量字典（Define the constant dictionaries）
            LoLItem_name_map: dict[int, str] = {item["id"]: item["name"] for item in LoLItems_locale}
            itemKey_itemId_map: dict[int, str] = {} #构建从装备序号到装备数据的映射（Build a map from the itemId to the corresponding item data）
            for (key, value) in LoLItems_binary.items():
                if key != "__linked" and value["__type"] == "ItemData":
                    itemKey_itemId_map[value["itemID"]] = key
            logPrint("开始整理数据……\nOrganizing data ...")
            pStats: re.Pattern[str] = re.compile(r"<stats>.*</stats>")
            pFormat: re.Pattern[str] = re.compile(r"<[/\sA-Za-z0-9=#\'_@]*>")
            pSection: re.Pattern[str] = re.compile(r"<section>.*?</section>") #在星号后添加问号以启用贪婪模式（Enable greedy match by adding a question mark after the asterisk）
            champions: dict[str, str] = {}
            for champion in champions_locale:
                champions[champion["alias"]] = champion["name"] + " " + (champion["description"] if "description" in champion else champion["alias"]) #15.9版本以前，“champion-summary.json”中没有“description”键（Before Patch 15.9, "description" key isn't present in "champion-summary.json"）
            LoLItem_data: dict[str, list[Any]] = {key: [] for key in LoLItem_header_keys}
            for i in range(len(LoLItems_locale)):
                item: dict[str, Any] = LoLItems_locale[i]
                item_default: dict[str, Any] = LoLItems_default[i]
                item_binary: dict[str, Any] = LoLItems_binary.get(itemKey_itemId_map[item["id"]], {}) #确定该装备在二进制json文件中的数据。在15.12.685.0388版本，神木之门引入的三个饮品没有出现在二进制json文件中（Determine item data in the binary json file. In Patch 15.12.685.0388, 3 juices introduced into Brawl don't exist in the binary json file）
                #首先处理共有部分（First, deal with the common part）
                ##下面填充装备的基本数据。这里参考的是英语描述（The following code fills the items' basic stats. Here the code refer to English descriptions）
                statDict: dict[str, str] = {}
                if "description" in item_default and pStats.search(item_default["description"]):
                    statStr: str = pStats.search(item_default["description"]).group().replace("<stats>", "").replace("</stats>", "").replace("<br>", "\n")
                    statList: list[str] = statStr.split("\n")
                    for stat_iter in statList:
                        if stat_iter != "": #有的装备没有基本属性，或者其字符串中存在几个连续的换行符（Some items don't have basic stats, or the string contains several continuous line feed characters）
                            while (matchObj := pFormat.search(stat_iter)):
                                stat_iter: str = stat_iter.replace(matchObj.group(), "")
                            pFigure: re.Pattern[str] = re.compile(r"(\+|\-)?[0-9]+%?")
                            if (matchObj := pFigure.search(stat_iter)):
                                figure: str = matchObj.group()
                                figure_attr = stat_iter.replace(figure, "").strip() #英文中，数值和属性之间有空格（In English, there's a space between the stat and the attribute）
                                if figure_attr in attribute_correct_map:
                                    figure_attr = attribute_correct_map[figure_attr]
                                statDict[figure_attr] = figure.replace("+", "")
                ##下面确定该装备在二进制json文件中存储的属性及其值（The following code determine attributes and corresponding values of this item stored in the binary json file）
                mDataValues: dict[str, float] = {}
                if "mDataValues" in item_binary:
                    for itemDataValue_iter in item_binary["mDataValues"]:
                        if all(j in itemDataValue_iter for j in ["mName", "mValue", "__type"]):
                            mDataValues[itemDataValue_iter["mName"]] = itemDataValue_iter["mValue"]
                #然后分类讨论（Then discuss about `j`)
                for j in range(len(LoLItem_header_keys)):
                    key: str = LoLItem_header_keys[j]
                    if j < len(item_base_header_keys): #基本表头部分（Base part）
                        if j == 2: #详细信息（`description`）
                            desc: str = item["description"].replace("<br>", "\n")
                            while (matchObj := pFormat.search(desc)):
                                desc = desc.replace(matchObj.group(), "")
                            to_append: Any = desc
                        elif j == 17: #装备名称（`localizedName`）
                            to_append = item["name"]
                        elif j == 18: #英文名称（`name`）
                            to_append = item_default["name"]
                        elif j == 19: #合成材料名称（`fromName`）
                            to_append = list(map(lambda x: LoLItem_name_map.get(x, x), item["from"])) if "from" in item else "" #在7.8版本以前，装备数据中无“from”键（Before Patch 7.8, "from" key isn't present in item data）
                        elif j == 20: #合成装备名称（`toName`）
                            to_append = list(map(lambda x: LoLItem_name_map.get(x, x), item["to"])) if "to" in item else "" #在7.8版本以前，装备数据中无“to”键（Before Patch 7.8, "to" key isn't present in item data）
                        elif j == 21: #装备持有者名称（`requiredChampionName`）
                            to_append = champions[item["requiredChampion"]] if "requiredChampion" in item and item["requiredChampion"] != "" else champions[item["requiredchampion"]] if "requiredchampion" in item and item["requiredchampion"] != "" else "" #在7.10版本以前，部分键是小写形式（Before Patch 7.10, some keys are in lower case）
                        elif j == 22: #所需队友名称（`requiredAllyName`）
                            to_append = champions[item["requiredAlly"]] if "requiredAlly" in item and item["requiredAlly"] != "" else champions[item["requiredally"]] if "requiredally" in item and item["requiredally"] != "" else "" #在7.16版本以前，装备数据中无“requiredAlly”键（Before Patch 7.16, "requiredAlly" key isn't present in item data）
                        elif j == 23: #特殊合成材料名称（`specialRecipeName`）
                            to_append = LoLItem_name_map.get(item["specialRecipe"], "") if "specialRecipe" in item and item["specialRecipe"] != 0 else LoLItem_name_map.get(item["specialrecipe"], "") if "specialrecipe" in item and item["specialrecipe"] != 0 else "" #在7.10版本以前，部分键是小写形式（Before Patch 7.10, some keys are in lower case）
                        elif j in {1, 3, 12, 15}: #逻辑类键（Keys of boolean type）
                            to_append = item.get(key, item.get(key.lower(), False)) #在14.15版本以前，装备数据中无“displayInItemSets”键（Before Patch 14.15, "displayInItemSets" key isn't in item data）
                        else:
                            to_append = item.get(key, item.get(key.lower(), "")) #在7.10版本以前，部分键是小写形式；在7.8版本以前，装备数据中无大多数键（Before Patch 7.10, some keys are in lower case. Before Patch 7.8, most keys aren't present in item data）
                    elif j < len(item_base_header_keys) + len(tooltip_header_keys): #游戏内详细信息部分（In-game tooltip part）
                        if status_strtable_locale == 200:
                            entry_key: str = "generatedtip_item_%d_tooltipinventoryextended" %(item["id"])
                            entry_key_hash: str = LoLDataExtractor.compute_rsthash(entry_key, strtable_locale["version"])
                            if entry_key in strtable_locale["entries"]:
                                tooltip: str = strtable_locale["entries"][entry_key]
                            elif entry_key_hash in strtable_locale["entries"]:
                                tooltip: str = strtable_locale["entries"][entry_key_hash]
                            else: #早期版本中没有按Shift查看详细信息的说法（In early versions, pressing Shift won't provide the detailed description）
                                entry_key = "generatedtip_item_%d_tooltipinventory" %(item["id"])
                                entry_key_hash = LoLDataExtractor.compute_rsthash(entry_key, strtable_locale["version"])
                                if entry_key in strtable_locale["entries"]:
                                    tooltip = strtable_locale["entries"][entry_key]
                                elif entry_key_hash in strtable_locale["entries"]:
                                    tooltip = strtable_locale["entries"][entry_key_hash]
                                else:
                                    tooltip = "" #空字符串仍然适用于下面的格式替换（The following format transformation applies to an empty string）
                        else:
                            tooltip = ""
                        if j == len(item_base_header_keys): #游戏内详细信息（`tooltip`）
                            to_append = tooltip
                        else: #游戏内详细信息（数值转换）（`tooltip_burn`）
                            LoLDataExtractor.calculatedVariables = {} #全局变量，存储已经计算过的变量的值。每次切换装备时重置（A global variable that calculates the value of variables that have been calculated. Resets when transforming another tooltip）
                            tooltip_text: str = LoLDataExtractor.tooltipTransform(tooltip, strtable_locale, item_binary, locale, enableModeOverride = False)
                            to_append = tooltip_text
                    elif j < len(item_base_header_keys) + len(tooltip_header_keys) + len(categories): #分类部分（Category part）
                        to_append = categories[j - len(item_base_header_keys) - len(tooltip_header_keys)] in item["categories"] if "categories" in item else False #在7.8版本以前，装备数据中无“categories”键（Before Patch 7.8, "categories" key isn't present in item data）
                    elif j < len(item_base_header_keys) + len(tooltip_header_keys) + len(categories) + len(attributes): #基础属性部分（Stat part）
                        if "description" in item_default:
                            key_default: str = list(attributes.keys())[j - len(item_base_header_keys) - len(tooltip_header_keys) - len(categories)]
                            to_append = statDict.get(key_default, "")
                        else:
                            to_append = ""
                    else:
                        key_decapitalized: str = key[0].lower() + key[1:]
                        if key in mDataValues:
                            attributeValue: str = mDataValues[key]
                        elif key in item_binary:
                            attributeValue = item_binary[key]
                        elif key_decapitalized in item_binary:
                            attributeValue = item_binary[key_decapitalized]
                        elif f"m{key}" in item_binary:
                            attributeValue = item_binary[f"m{key}"]
                        else:
                            attributeValue = ""
                        to_append = attributeValue
                    LoLItem_data[key].append(to_append)
            base_statistics_display_order: list[int] = [0, 17, 18, 1, 3, 19, 20, 6, 13, 14, 21, 22, 9, 10, 23, 12, 15, 16, 2]
            LoLItem_statistics_display_order: list[int] = base_statistics_display_order + list(range(len(item_base_header_keys), len(LoLItem_header)))
            LoLItem_data_organized: dict[str, list[Any]] = {LoLItem_header_keys[i]: LoLItem_data[LoLItem_header_keys[i]] for i in LoLItem_statistics_display_order}
            LoLItem_df: pandas.DataFrame = pandas.DataFrame(data = LoLItem_data_organized)
            optimize_bool_display(LoLItem_df)
            LoLItem_df.index = list(range(1, len(LoLItem_df) + 1))
            LoLItem_df = LoLItem_df.sort_values(by = "id", ascending = True)
            LoLItem_df = pandas.concat([pandas.DataFrame([LoLItem_header])[LoLItem_df.columns], LoLItem_df])
            LoLItem_dfs[version] = LoLItem_df.copy(deep = True)
            item_df_formed = True
            end: float = time.time()
            unit: float = end - start
            total_used += unit
            runTimes.append(unit)
            total_remaining = sum(runTimes) / (version_index + 1 - failed_count) * (len(versions_sort) - version_index - 1)
            logPrint("整理该版本数据所花费的时间（Time spent in organizing this version）：", format_runtime(unit))
            logPrint("已花费的总时间（Total time used）                                  ：", format_runtime(total_used))
            logPrint("剩余时间（Time remaining）                                         ：", format_runtime(total_remaining))
            logPrint("预计总时间（Expected total time）                                  ：", format_runtime(total_used + total_remaining), end = "\n\n")
    return (LoLItem_dfs, item_df_formed)

def export_item_data() -> None:
    '''
    导出装备数据的主函数。<br>The main function to export item data.
    
    选择一个语言，选择一个数据来源，然后导出数据。<br>Select a language, then select a data source, and finally export data.
    '''
    logPrint("请选择装备的输出语言【默认为中文（中国）】：\nPlease select a language to output the items (the default option is zh_CN):")
    language_df: pandas.DataFrame = pandas.DataFrame(language_dict)
    logPrint(format_df(language_df)[0], write_time = False)
    while True:
        language_option: str = logInput()
        if language_option == "" or language_option in [str(i) for i in range(1, 30)]:
            if language_option == "":
                language_option = "29"
            language_code: str = list(language_ddragon.keys())[int(language_option) - 1]
            break
        elif language_option[0] == "0":
            return
        else:
            logPrint("语言选项输入错误！请重新输入：\nERROR input of language option! Please try again:")

    workbook_exist: bool = False
    excel_name: str = "英雄联盟装备信息.xlsx"
    excel_name_sorted: str = "英雄联盟装备信息(sorted).xlsx"
    logPrint('请选择数据来源：（输入“0”以返回上一层。）\nPlease select a data source: (Submit "0" to return to the last step.)\n1\tDataDragon\n2\tCommunityDragon')
    while True:
        item_df_formed: bool = False #记录程序是否形成了装备数据框（Records whether the program has formed an item dataframe）
        source: str = logInput()
        if source == "":
            continue
        elif source[0] == "0":
            break
        elif source[0] == "1":
            LoLItem_dfs, item_df_formed_tmp = sort_item_ddragon(locale = language_code)
            item_df_formed = item_df_formed or item_df_formed_tmp
        elif source[0] == "2":
            LoLItem_dfs, item_df_formed_tmp = sort_item_cdragon(locale = language_code)
            item_df_formed = item_df_formed or item_df_formed_tmp
        else:
            continue
        if item_df_formed:
            logPrint("是否导出以上装备数据至Excel中？（输入任意键导出，否则不导出。）\nDo you want to export the above data into Excel? (Submit any non-empty string to export, or null to refuse exporting.)")
            export_str: str = logInput()
            export: bool = bool(export_str)
            if export:
                versions_sort: list[str] = list(LoLItem_dfs.keys())
                if not os.path.exists(excel_name):
                    wbCreateFlag: bool = create_workbook_win32(os.path.abspath(excel_name), sheet1_name = "latest")
                workbook_exist = os.path.exists(excel_name)
                logPrint("正在保存中……\nSaving the data ...")
                while True:
                    try:
                        with (pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = excel_name)) as writer:
                            runTimes: list[float] = [] #记录保存一个版本的装备数据所花费的时间（Records the time spent in saving item data of one version）
                            total_used: float = 0
                            for i in range(len(versions_sort)):
                                start: float = time.time()
                                version: str = versions_sort[i]
                                logPrint("装备信息导出进度（Item data export process）：%d/%d\t版本（Version）：%s" %(i + 1, len(versions_sort), version))
                                if version == "latest" or version == "pbe":
                                    addDefaultStyle(LoLItem_dfs[version]).to_excel(excel_writer = writer, sheet_name = version)
                                elif source != "" and source[0] == "1":
                                    addDefaultStyle(LoLItem_dfs[version]).to_excel(excel_writer = writer, sheet_name = version + " (ddragon)")
                                else:
                                    addDefaultStyle(LoLItem_dfs[version]).to_excel(excel_writer = writer, sheet_name = version + " (cdragon)")
                                end: float = time.time()
                                unit: float = end - start
                                total_used += unit
                                runTimes.append(unit)
                                total_remaining = sum(runTimes) / (i + 1) * (len(versions_sort) - i - 1)
                                logPrint("保存该版本数据所花费的时间（Time spent in saving this version）：", format_runtime(unit))
                                logPrint("已花费的总时间（Total time used）                              ：", format_runtime(total_used))
                                logPrint("剩余时间（Time remaining）                                     ：", format_runtime(total_remaining))
                                logPrint("预计总时间（Expected total time）                              ：", format_runtime(total_used + total_remaining), end = "\n\n")
                    except PermissionError:
                        logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                        logInput()
                    else:
                        break
        logPrint('请选择数据来源：（输入“0”以返回上一层。）\nPlease select a data source: (Submit "0" to return to the last step.)\n1\tDataDragon\n2\tCommunityDragon')
    if workbook_exist:
        logPrint("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致版本号顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of version order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
        sort_str: str = logInput()
        if sort_str != "": #所有工作表按顺序依次分为固定版本号、cdragon版本号和ddragon版本号（All sheets are divided into the fixed version class, cdragon version class and ddragon version class）
            logPrint("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
            while True:
                try:
                    wb: Workbook = load_workbook(excel_name)
                except FileNotFoundError:
                    logPrint('装备工作簿读取失败！请确保当前文件夹内含有名为“%s”的工作簿。如果需要重新生成该召唤师的工作簿，请输入“0”。\nERROR reading the summoner profile workbook! Please make sure the workbook "%s" is in the current folder". If you want to regenerate this summoner\'s workbook, please submit "0".' %(excel_name, excel_name))
                    items_reload: str = logInput()
                    if items_reload == "0":
                        break
                else:
                    sheetnames: list[str] = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                    #下面锁定工作表顺序（The following code determine the sheet order）
                    print("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                    fixed_version_list: list[str] = ["pbe", "latest"]
                    cdragon_version_list: list[Patch] = []
                    ddragon_version_list: list[Patch] = []
                    for version in sheetnames:
                        if version.endswith("(cdragon)"):
                            cdragon_version_list.append(Patch(version))
                        elif version.endswith("(ddragon)"):
                            ddragon_version_list.append(Patch(version))
                    sheetnames_sorted: list[str] = []
                    for sheet_iter in fixed_version_list:
                        if sheet_iter in sheetnames:
                            sheetnames.remove(sheet_iter)
                            sheetnames_sorted.append(sheet_iter)
                    cdragon_version_sorted: list[str] = list(map(lambda x: str(x) + " (cdragon)", Patch.sort(cdragon_version_list)))
                    cdragon_version_sorted.reverse()
                    ddragon_version_sorted: list[str] = list(map(lambda x: str(x) + " (ddragon)", Patch.sort(ddragon_version_list)))
                    ddragon_version_sorted.reverse()
                    sheetnames_sorted += cdragon_version_sorted + ddragon_version_sorted #所有工作表的期望顺序存储在sheetnames_sorted变量中（The ordered result of all sheets is stored in the variable `sheetnames_sorted`）
                    #下面排列所有工作表（The following code arrange all sheets）
                    print("正在排序……\nOrdering ...")
                    sort_worksheet(wb, sheetnames_sorted)
                    print('正在保存中……\nSaving the ordered workbook ...')
                    wb.save(excel_name_sorted)
                    print('排序完成！排好序的工作簿已保存为“%s”。\nOrdering finished! The ordered workbook is saved as "%s".\n' %(excel_name_sorted, excel_name_sorted))
                    wb.close()
                    break

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------
def main() -> int:
    logPrint("请选择一个功能：\nPlease select a function:\n1\t配置配装方案（Configure item sets）\n2\t导出装备数据（Export item data）")
    while True:
        option: str = logInput()
        if option == "":
            continue
        elif option[0] == "0":
            break
        elif option[0] == "1":
            LCUConnect(connect, on_close = disconnect) #这里要使用本脚本中disconnect函数中的logPrint函数，因此需要重载关闭函数（Here we need to use the `logPrint` inside the `disconnect` function of this program, so we need to override the internal close function）
        elif option[0] == "2":
            export_item_data()
        else:
            continue
        logPrint("请选择一个功能：\nPlease select a function:\n1\t配置配装方案（Configure item sets）\n2\t导出装备数据（Export item data）")
    return 0

status = main()
log.write("\n[Program terminated and returned status 0.]\n")
log.close()
