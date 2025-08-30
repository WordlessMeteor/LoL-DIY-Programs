#准备必要的库（Import necessary libraries）
import argparse, pandas, re, requests, traceback, shutil, time, unicodedata
from openpyxl import load_workbook
from wcwidth import wcswidth

parser = argparse.ArgumentParser()
parser.add_argument("-t", "--transform", help = "将从CommunityDragon获取的装备详细信息中的变量转换成二进制装备json数据中的实际值（Transform variables in tooltips of items obtained from CommunityDragon database into the representing values）", action = "store_true")
args = parser.parse_args()

def count_nonASCII(s: str): #统计一个字符串中占用命令行2个宽度单位的字符个数（Count the number of characters that take up 2 width unit in CMD）
    return sum([unicodedata.east_asian_width(character) in ("F", "W") for character in list(str(s))])

def rm_ctrl_char(s: str): #移除一个字符串中的所有C0和C1字符（Remove all C0 and C1 characters from a string）
    return "".join(ch for ch in s if unicodedata.category(ch) != "Cc")

def format_df(df: pandas.DataFrame, width_exceed_ask: bool = True, direct_print: bool = False, print_header: bool = True, print_index: bool = False, reserve_index = False, start_index = 0, header_align: str = "^", align: str = "^", align_replicate_rule: str = "all"): #按照每列最长字符串的命令行宽度加上2，再根据每个数据的中文字符数量决定最终格式化输出的字符串宽度（Get the width of the longest string of each column, add it by 2, and substract it by the number of each cell string's Chinese characters to get the final width for each cell to print using `format` function）
    df = df.copy(deep = True)
    old_index = df.index
    df.index = range(start_index, len(df) + start_index)
    maxLens = {}
    maxWidth = shutil.get_terminal_size()[0]
    fields = df.columns.tolist()
    for field in fields:
        maxLens[field] = max(0 if len(df) == 0 else max(map(lambda x: wcswidth(rm_ctrl_char(str(x))), df[field])), wcswidth(rm_ctrl_char(field))) + 2
    index_len = 0 if len(df) == 0 else max(map(lambda x: len(str(x)), old_index)) if reserve_index else max(len(str(start_index)), len(str(start_index + len(df) - 1)))
    if sum(maxLens.values()) + 2 * (len(fields) - 1) > maxWidth or print_index and index_len + sum(maxLens.values()) + 2 * len(fields) > maxWidth:
        if width_exceed_ask:
            print("单行数据字符串输出宽度超过当前终端窗口宽度！是否继续？（输入任意键继续，否则直接打印该数据框。）\nThe output width of each record string exceeds the current width of the terminal window! Continue? (Input anything to continue, or null to directly print this dataframe.)")
            if input() == "":
                #print(df)
                result = str(df)
                return (result, maxLens)
        elif direct_print:
            #print("单行数据字符串输出宽度超过当前终端窗口宽度！将直接打印该数据框！\nThe output width of each record string exceeds the current width of the terminal window! The program is going to directly print this dataframe!")
            result = str(df)
            return (result, maxLens)
        # else:
        #     print("单行数据字符串输出宽度超过当前终端窗口宽度！将继续格式化输出！\nThe output width of each record string exceeds the current width of the terminal window! The program is going on formatted printing!")
    result = ""
    #确定各列的排列方向（Determine the alignments of all columns）
    if isinstance(header_align, str) and isinstance(align, str):
        if not all(map(lambda x: x in {"<", "^", ">"}, header_align)) or not all(map(lambda x: x in {"<", "^", ">"}, align)):
            print('排列方式字符串参数错误！排列方式必须是“<”“^”或者“>”中的一个。请修改排列方式字符串参数。\nParameter ERROR of the alignment string! The alignment value must be one of {"<", "^", ">"}. Please change the alignment string parameter.')
        if len(header_align) == 0: #指定为空字符串，即默认居中输出（Specifying it as a null string means output centered by default）
            header_alignments = ["^"] * df.shape[1]
        elif len(header_align) == 1:
            header_alignments = [header_align] * df.shape[1]
        else:
            header_alignments_tmp = list(header_align)
            if len(header_align) < df.shape[1]:
                if align_replicate_rule == "last":
                    header_alignments = header_alignments_tmp + [header_alignments_tmp[-1]] * len(df.shape[1] - len(header_align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    header_alignments = header_alignments_tmp * (df.shape[1] // len(header_align)) + header_alignments_tmp[:df.shape[1] % len(header_align)]
            else:
                header_alignments = header_alignments_tmp[:df.shape[1]]
        if len(align) == 0:
            alignments = ["^"] * df.shape[1]
        elif len(align) == 1:
            alignments = [align] * df.shape[1]
        else:
            alignments_tmp = list(align)
            if len(align) < df.shape[1]:
                if align_replicate_rule == "last":
                    alignments = alignments_tmp + [alignments_tmp[-1]] * len(df.shape[1] - len(align))
                else:
                    if align_replicate_rule != "all":
                        print("排列方式列表补充规则不合法！将默认采用全部填充。\nAlignment list supplement rule illegal! The whole alignment string will be replicated.")
                    alignments = alignments_tmp * (df.shape[1] // len(align)) + alignments_tmp[:df.shape[1] % len(align)]
            else:
                alignments = alignments_tmp[:df.shape[1]]
        if print_header:
            if print_index:
                result += " " * (index_len + 2)
            for i in range(df.shape[1]):
                field = fields[i]
                tmp = "{0:{align}{w}}".format(rm_ctrl_char(field), align = header_alignments[i], w = maxLens[field] - count_nonASCII(field))
                result += tmp
                #print(tmp, end = "")
                if i != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            result += "\n"
            #print()
        index = start_index
        for i in range(df.shape[0]):
            if print_index:
                result += "{0:>{w}}".format(old_index[index - start_index] if reserve_index else index, w = index_len) + "  "
            for j in range(df.shape[1]):
                field = fields[j]
                cell = str(list(df[field])[i])
                tmp = "{0:{align}{w}}".format(rm_ctrl_char(cell), align = alignments[j], w = maxLens[field] - count_nonASCII(cell))
                result += tmp
                #print(tmp, end = "")
                if j != df.shape[1] - 1:
                    result += "  "
                    #print("  ", end = "")
            if i != df.shape[0] - 1:
                result += "\n"
            #print() #注意这里的缩进和上一行不同（Note that here the indentation is different from the last line）
            index += 1
    else:
        print("排列方式参数错误！请传入字符串。\nAlignment parameter ERROR! Please pass a string instead.")
    return (result, maxLens)

def patch_compare(patch1, patch2): #比较两个版本号的先后顺序。当patch1 < patch2时，返回True，否则返回False。用于比较DataDragon数据库中未收录的版本和收录的最新版本的关系。如果未收录的版本小于收录的最新版本，那么该版本是美测服的临时版本，后来被合并更新了，如正式服将13.2和13.3合并更新了，因此DataDragon数据库中未收录13.2版本的数据；如果未收录的版本大于收录的最新版本，那么该版本是美测服的当前版本，但是仍处于开发状态，尚未完全确定，所以DataDragon数据库尚未收录，将以最新版本代替该版本；二者不可能相等，因为如果相等的话，就不会引发报错而调用此函数（Compare the time order of two patches. When patch1 < patch2, return True and vice versa. Designed to compare a patch not archived in DataDragon database with the latest patch archived in DataDragon database. If the unarchived patch is less than the latest archived patch, then this patch must be the intermediate patch and be merged into the update of its successive patch, such as Patch 13.2 merged into the update of Patch 13.3, so that DataDragon database doesn't archive the data of Patch 13.2; If the unarchived patch is greater than the latest archived patch, then this patch must be the current patch on PBE but is under development and improvement, so that DataDragon database doesn't archive this patch, either, in which case the latest patch will be used to substitute this unarchived patch; The two patches can't be the same, for suppose they're same, then the error to cause the call of this function won't be triggered）
    if not isinstance(patch1, str):
        patch1 = str(patch1)
    if not isinstance(patch2, str):
        patch2 = str(patch2)
    lst1, lst2 = patch1.split("."), patch2.split(".")
    try:
        lst1 = list(map(int, lst1))
    except ValueError:
        if lst1[0] != "pbe":
            print("第1个版本字符串不合法！请输入用半角句号连接的正整数，如13.15.1、10.10.3216176。\nThe first patch variable is illegal! Please pass the integers concatenated by dot, such as 13.15.1 and 10.10.3216176.")
        return False
    try:
        lst2 = list(map(int, lst2))
    except ValueError:
        if lst1[0] != "pbe":
            print("第2个版本字符串不合法！请输入用半角句号连接的正整数，如13.15.1、10.10.3216176。\nThe second patch variable is illegal! Please pass the integers concatenated by dot, such as 13.15.1 and 10.10.3216176.")
            return False
        else:
            return True
    for i in range(min(len(lst1), len(lst2))):
        if lst1[i] < lst2[i]:
            return True
        elif lst1[i] > lst2[i]:
            return False
        else:
            continue
    if len(lst1) < len(lst2):
        return True
    else:
        return False #这里将两个版本相同视为假，暗示了在本程序用得到的地方，两个版本不可能相同（Here the case where the two patches are the same is regarded as False, which indicates that the two patches can't be same within its use in this program）

def patch_sort(patchList: list): #利用插入排序算法，根据patch_compare函数对版本列表进行升序排列（Sorts a patch list according to the principle of `patch_compare` function through the insertion sort algorithm）
    patch_re = re.compile("([0-9]+.[0-9]+.[0-9]+)|([0-9]+.[0-9]+)")
    if all(map(lambda x: isinstance(x, str), patchList)) and all(map(lambda x: patch_re.search(x), patchList)): #此处放宽了参数的格式限制：只要列表的每个元素都是包含版本字符串的字符串即可（Here the function relaxes the limit for the format of the parameter: any list whose elements are all strings that contain a patch string is OK）
        patchList = list(map(lambda x: patch_re.search(x).group(), patchList))
        for i in range(1, len(patchList)):
            tmp = patchList[i] #将第i个元素临时存储（Temporarily stores the i-th element of `patchList`）
            j = i - 1
            while j >= 0 and patch_compare(tmp, patchList[j]): #如果检测到第i个元素比第(j = i - 1)个元素小，就要逐渐减小j，直到找到一个j，使得第j个元素小于第i个元素，此时第j + 1个元素仍然大于第i个铁元素。把j + 1及以后的元素右移，空出的位置再插入第i个元素（1f an i-th element is detected to be less than the j-th element, namely the (i - 1)th element, then the program decrements j until it finds a j such that the j-th element is less than the i-th element, while the (j + 1)-th element is still greater than the i-the element. Then, shift all elements between the current j-th and i-th elements and insert the i-th elements into the empty space）
                patchList[j + 1] = patchList[j]
                j -= 1
            patchList[j + 1] = tmp
    else:
        print("您的版本列表格式有误！\nYour patch list is not correctly formatted!")
    return patchList

def getUrl(url: str):
    retry = 0
    while True:
        try:
            retry += 1
            source = requests.get(url)
            source.raise_for_status()
        except requests.exceptions.HTTPError as http_err:
            if retry > 5:
                break
            if http_err.response.status_code in {403, 404}:
                return (source, http_err.response.status_code)
        except requests.exceptions.SSLError as ssl_error:
            if retry > 5:
                break
            if "[SSL: UNEXPECTED_EOF_WHILE_READING] EOF occurred in violation of protocol" in str(ssl_error):
                print("违反协议导致读取中断！正在尝试第%d次重新获取数据！\nEOF occurred in violation of protocol! Trying to recapture the data with url: %s. Time(s) tried: %d" %(retry, url, retry))
            elif 'certificate verify failed' in str(ssl_error):
                print("SSL证书验证失败！正在尝试第%d次重新获取数据！\nSSL certificate verify failed! Trying to recapture the data with url: %s. Time(s) tried: %d" %(retry, url, retry))
            elif 'Max retries exceeded with url' in str(ssl_error):
                print("请求数量超过限制！正在尝试第%d次重新获取数据！\nMax retries exceed with url! Trying to recapture the data with url: %s. Time(s) tried: %d" %(retry, url, retry))
        except requests.exceptions.ProxyError:
            if retry > 5:
                break
            print("无法连接到代理！正在尝试第%d次重新获取数据！\nCannot connect to proxy! Trying to recapture the data with url: %s. Time(s) tried: %d" %(retry, url, retry))
        except requests.exceptions.ChunkedEncodingError:
            if retry > 5:
                break
            print("接收数据块长度不正确导致连接中断！正在尝试第%d次重新获取数据！\nConnection broken: InvalidChunkLength. Trying to recapture the data with url: %s. Time(s) tried: %d" %(retry, url, retry))
        except requests.exceptions.ConnectionError:
            if retry > 5:
                break
            print("由于远程服务器端无响应，连接已关闭！正在尝试第%d次重新获取数据！\nRemote end closed connection without response. Trying to recapture the data with url: %s. Time(s) tried: %d" %(retry, url, retry))
        else:
            return (source, 0)
    if retry > 5:
        return (None, 1)

def format_runtime(seconds: int):
    units = [(" d", 86400), (" h", 3600), (" m", 60), (" s", 1)]
    result = []
    for unit_name, unit_seconds in units:
        if seconds >= unit_seconds:
            unit_value = round(seconds // unit_seconds)
            seconds %= unit_seconds
            result.append(f"{unit_value}{unit_name}")
    
    return " ".join(result) if result else "0"

print("请选择装备的输出语言【默认为中文（中国）】：\nPlease select a language to output the items (the default option is zh_CN):") #本来考虑把可用CDragon数据版本放在第三列，但是后来发现表头名字太长了，索性放在最后了（I had considered putting "Applicable CDragon Data Patches" at the third column, but then found the header was too long. So I put it at the last column）
language_ddragon = {1: {"CODE": "ar_AE", "LANGUAGE (EN)": "Arabic (United Arab Emirates)", "LANGUAGE (ZH)": "阿拉伯语（阿拉伯联合酋长国）", "Applicable CDragon Data Patches": "9.20～10.1, 13.20+"}, 2: {"CODE": "cs_CZ", "LANGUAGE (EN)": "Czech (Czech Republic)", "LANGUAGE (ZH)": "捷克语（捷克共和国）", "Applicable CDragon Data Patches": "7.1+"}, 3: {"CODE": "el_GR", "LANGUAGE (EN)": "Greek (Greece)", "LANGUAGE (ZH)": "希腊语（希腊）", "Applicable CDragon Data Patches": "9.1+"}, 4: {"CODE": "pl_PL", "LANGUAGE (EN)": "Polish (Poland)", "LANGUAGE (ZH)": "波兰语（波兰）", "Applicable CDragon Data Patches": "9.1+"}, 5: {"CODE": "ro_RO", "LANGUAGE (EN)": "Romanian (Romania)", "LANGUAGE (ZH)": "罗马尼亚语（罗马尼亚）", "Applicable CDragon Data Patches": "9.1+"}, 6: {"CODE": "hu_HU", "LANGUAGE (EN)": "Hungarian (Hungary)", "LANGUAGE (ZH)": "匈牙利语（匈牙利）", "Applicable CDragon Data Patches": "9.1+"}, 7: {"CODE": "en_GB", "LANGUAGE (EN)": "English (United Kingdom)", "LANGUAGE (ZH)": "英语（英国）", "Applicable CDragon Data Patches": "9.1+"}, 8: {"CODE": "de_DE", "LANGUAGE (EN)": "German (Germany)", "LANGUAGE (ZH)": "德语（德国）", "Applicable CDragon Data Patches": "7.1+"}, 9: {"CODE": "es_ES", "LANGUAGE (EN)": "Spanish (Spain)", "LANGUAGE (ZH)": "西班牙语（西班牙）", "Applicable CDragon Data Patches": "9.1+"}, 10: {"CODE": "it_IT", "LANGUAGE (EN)": "Italian (Italy)", "LANGUAGE (ZH)": "意大利语（意大利）", "Applicable CDragon Data Patches": "9.1+"}, 11: {"CODE": "fr_FR", "LANGUAGE (EN)": "French (France)", "LANGUAGE (ZH)": "法语（法国）", "Applicable CDragon Data Patches": "9.1+"}, 12: {"CODE": "ja_JP", "LANGUAGE (EN)": "Japanese (Japan)", "LANGUAGE (ZH)": "日语（日本）", "Applicable CDragon Data Patches": "9.1+"}, 13: {"CODE": "ko_KR", "LANGUAGE (EN)": "Korean (Korea)", "LANGUAGE (ZH)": "朝鲜语（韩国）", "Applicable CDragon Data Patches": "9.7+"}, 14: {"CODE": "es_MX", "LANGUAGE (EN)": "Spanish (Mexico)", "LANGUAGE (ZH)": "西班牙语（墨西哥）", "Applicable CDragon Data Patches": "9.1+"}, 15: {"CODE": "es_AR", "LANGUAGE (EN)": "Spanish (Argentina)", "LANGUAGE (ZH)": "西班牙语（阿根廷）", "Applicable CDragon Data Patches": "9.7+"}, 16: {"CODE": "pt_BR", "LANGUAGE (EN)": "Portuguese (Brazil)", "LANGUAGE (ZH)": "葡萄牙语（巴西）", "Applicable CDragon Data Patches": "9.1+"}, 17: {"CODE": "en_US", "LANGUAGE (EN)": "English (United States)", "LANGUAGE (ZH)": "英语（美国）", "Applicable CDragon Data Patches": "9.1+"}, 18: {"CODE": "en_AU", "LANGUAGE (EN)": "English (Australia)", "LANGUAGE (ZH)": "英语（澳大利亚）", "Applicable CDragon Data Patches": "9.1+"}, 19: {"CODE": "ru_RU", "LANGUAGE (EN)": "Russian (Russia)", "LANGUAGE (ZH)": "俄语（俄罗斯）", "Applicable CDragon Data Patches": "9.1+"}, 20: {"CODE": "tr_TR", "LANGUAGE (EN)": "Turkish (Turkey)", "LANGUAGE (ZH)": "土耳其语（土耳其）", "Applicable CDragon Data Patches": "9.1+"}, 21: {"CODE": "ms_MY", "LANGUAGE (EN)": "Malay (Malaysia)", "LANGUAGE (ZH)": "马来语（马来西亚）", "Applicable CDragon Data Patches": ""}, 22: {"CODE": "en_PH", "LANGUAGE (EN)": "English (Republic of the Philippines)", "LANGUAGE (ZH)": "英语（菲律宾共和国）", "Applicable CDragon Data Patches": "10.5+"}, 23: {"CODE": "en_SG", "LANGUAGE (EN)": "English (Singapore)", "LANGUAGE (ZH)": "英语（新加坡）", "Applicable CDragon Data Patches": "10.5+"}, 24: {"CODE": "th_TH", "LANGUAGE (EN)": "Thai (Thailand)", "LANGUAGE (ZH)": "泰语（泰国）", "Applicable CDragon Data Patches": "9.7+"}, 25: {"CODE": "vn_VN", "LANGUAGE (EN)": "Vietnamese (Viet Nam)", "LANGUAGE (ZH)": "越南语（越南）", "Applicable CDragon Data Patches": "9.7～13.9"}, 26: {"CODE": "vi_VN", "LANGUAGE (EN)": "Vietnamese (Viet Nam)", "LANGUAGE (ZH)": "越南语（越南）", "Applicable CDragon Data Patches": "12.17+"}, 27: {"CODE": "id_ID", "LANGUAGE (EN)": "Indonesian (Indonesia)", "LANGUAGE (ZH)": "印度尼西亚语（印度尼西亚）", "Applicable CDragon Data Patches": ""}, 28: {"CODE": "zh_MY", "LANGUAGE (EN)": "Chinese (Malaysia)", "LANGUAGE (ZH)": "中文（马来西亚）", "Applicable CDragon Data Patches": "10.5+"}, 29: {"CODE": "zh_CN", "LANGUAGE (EN)": "Chinese (China)", "LANGUAGE (ZH)": "中文（中国）", "Applicable CDragon Data Patches": "9.7+"}, 30: {"CODE": "zh_TW", "LANGUAGE (EN)": "Chinese (Taiwan)", "LANGUAGE (ZH)": "中文（台湾）", "Applicable CDragon Data Patches": "9.7+"}}
language_cdragon = {}
for i in language_ddragon:
    if language_ddragon[i]["CODE"] == "en_US":
        language_cdragon[language_ddragon[i]["CODE"]] = "default" #在CommunityDragon数据库上，美服正式服的数据资源代码是default，而不是小写的en_US（The code for English (US) data resources on CommunityDragon database is "default" instead of the lowercase of "en_US"）
    else:
        language_cdragon[language_ddragon[i]["CODE"]] = language_ddragon[i]["CODE"].lower()
language_dict = {"No.": list(language_ddragon.keys()), "CODE": list(map(lambda x: x["CODE"], language_ddragon.values())), "LANGUAGE": list(map(lambda x: x["LANGUAGE (EN)"], language_ddragon.values())), "语言": list(map(lambda x: x["LANGUAGE (ZH)"], language_ddragon.values())), "Applicable CDragon Data Patches": list(map(lambda x: x["Applicable CDragon Data Patches"], language_ddragon.values()))}
language_df = pandas.DataFrame(language_dict)
print(format_df(language_df)[0])
while True:
    language_option = input()
    if language_option == "" or language_option in [str(i) for i in range(1, 30)]:
        if language_option == "":
            language_option = "29"
        language_code = language_ddragon[int(language_option)]["CODE"]
        break
    elif language_option[0] == "0":
        exit()
    else:
        print("语言选项输入错误！请重新输入：\nERROR input of language option! Please try again:")

workbook_exist = False
while True:
    item_df_formed = False #记录程序是否形成了装备数据框（Records whether the program has formed an item dataframe）
    print("请选择数据来源：\nPlease select a data source:\n1\tDataDragon\n2\tCommunityDragon")
    source = input()
    if source != "" and source[0] == "0":
        break
    if source != "" and source[0] == "1":
        version_url = "https://ddragon.leagueoflegends.com/api/versions.json"
        LoLItem_dfs = {}
        while True:
            back = False
            print("请在以下版本号中选择并输入完整的版本号：\nPlease select a version and then enter it entirely:")
            src, status = getUrl(version_url)
            if status != 0:
                if status == 1:
                    print("版本信息获取失败！\nVersion information capture failed!")
                elif status in {403, 404}:
                    print("版本信息文件不存在！\nVersion file not found!")
                break
            versions = src.json()
            print(versions)
            #下面的代码块生成英雄联盟专业术语中英文对照的Sheet1的内容。数据来源是DataDragon数据库（The following code block generates the Sheet1 in LoL Term Translation - zh_CN & en_US. Data resources are from DataDrabon database）
            while True:
                version = input()
                if version == "":
                    versions_sort = [versions[0]]
                    break
                elif version == "0":
                    back = True
                    break
                elif version == "all":
                    versions_sort = versions
                    break
                elif version in versions:
                    versions_sort = [version]
                    break
                else:
                    try:
                        versions_sort = eval(version)
                    except:
                        print("您的输入有误，请重新输入！\nERROR input! Please try again.")
                    else:
                        if isinstance(versions_sort, list) and all(map(lambda x: x in versions, versions_sort)):
                            break
                        else:
                            print("您的输入有误，请重新输入！\nERROR input! Please try again.")
            if back:
                break
            runTimes = [] #记录整理一个版本的装备数据所花费的时间（Records the time spent in sorting out item data of one version）
            total_used = 0
            failed_count = 0 #记录数据获取失败的版本的个数（Count the number of versions whose data fail to be fetched）
            for version_index in range(len(versions_sort)):
                start = time.time()
                print("[%s]" %(time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())), end = "")
                print("整理进度（Sorting process）：%d/%d" %(version_index + 1, len(versions_sort)))
                version = versions_sort[version_index]
                LoLItems_locale_url = "https://ddragon.leagueoflegends.com/cdn/%s/data/%s/item.json" %(version, language_code)
                LoLItems_default_url = "https://ddragon.leagueoflegends.com/cdn/%s/data/en_US/item.json" %version
                champions_locale_url = "https://ddragon.leagueoflegends.com/cdn/%s/data/%s/champion.json" %(version, language_code)
                print(f"正在获取{version}版本的目标语言装备信息……\nFetching LoL item information of version {version} in target language ...")
                src, status = getUrl(LoLItems_locale_url)
                if status != 0:
                    if status == 1:
                        print("目标语言装备信息获取失败！\nLoL item information in target language capture failed!")
                    elif status == 403:
                        print("目标语言装备信息文件不存在！\nLoL item file in target language not found!")
                    failed_count += 1
                    continue
                LoLItems_locale = src.json()
                print(f"正在获取{version}版本的英文装备信息……\nFetching LoL item information of version {version} in English ...")
                src, status = getUrl(LoLItems_default_url)
                if status != 0:
                    if status == 1:
                        print("英文装备信息获取失败！\nLoL item information in English capture failed!")
                    elif status == 403:
                        print("英文装备信息文件不存在！\nLoL item file in English not found!")
                    failed_count += 1
                    continue
                LoLItems_default = src.json()
                print(f"正在获取{version}版本的目标语言英雄信息……\nFetching champion information of version {version} in target language ...")
                src, status = getUrl(champions_locale_url)
                if status != 0:
                    if status == 1:
                        print("目标语言英雄信息获取失败！\nChampion information in target language capture failed!")
                    elif status == 403:
                        print("目标语言英雄信息文件不存在！\nChampion file in target language not found!")
                    failed_count += 1
                    continue
                champions_locale = src.json()

                #下面设置装备表头的元数据部分（Set the metadata part of the item headers）
                base_header = {"id": "装备序号", "group": "分组", "description": "详细信息", "colloq": "检索关键字", "plaintext": "简述", "consumed": "消耗品", "stacks": "最大持有数量", "depth": "深度", "consumeOnFull": "满装备时自动消耗", "from": "合成材料序号", "into": "合成装备序号", "specialRecipe": "特殊合成材料", "inStore": "商店可见性", "hideFromAll": "不可见性", "requiredChampion": "装备持有者", "requiredAlly": "所需队友", "localizedName": "装备名称", "name": "英文名称", "fromName": "合成材料名称", "intoName": "合成装备名称", "requiredChampionName": "装备持有者名称", "requiredAllyName": "所需队友名称", "specialRecipeName": "特殊合成材料名称", "baseGold": "合成费用", "purchasable": "可以购买", "totalGold": "总费用", "sellGold": "售价"}
                base_header_keys = list(base_header.keys())
                base_header_values = list(base_header.values())
                #下面设置装备表头的分类（标签）部分（Set the category / tag part of the item headers）
                tags_initial = set()
                for item in LoLItems_locale["data"].values():
                    tags_initial |= set(map(lambda x: x.lower(), item["tags"])) #之所以要加lower（upper也可以），是因为在3.12.26版本以前，所有的标签/分类信息都是大写的（The reason why "lower" (or "upper") is needed is that all tags / categories before v3.12.26 are in upper case）
                tags_initial = sorted(tags_initial)
                tags_organized = ["Lane", "Jungle", "GoldPer", "Boots", "Consumable", "Damage", "CriticalStrike", "AttackSpeed", "OnHit", "ArmorPenetration", "SpellDamage", "Mana", "ManaRegen", "MagicPenetration", "Health", "HealthRegen", "MagicResist", "AbilityHaste", "CooldownReduction", "Movement", "NonbootsMovement", "LifeSteal", "SpellVamp", "Active", "Armor", "Aura", "Slow", "SpellBlock", "Stealth", "Tenacity", "Trinket", "Vision", "Bilgewater"] #设置分类表头的顺序（Set the order of category headers）
                tags = []
                for tag in tags_organized:
                    if tag.lower() in tags_initial:
                        tags_initial.remove(tag.lower())
                    tags.append(tag)
                tags += tags_initial
                tags_dict = {"AbilityHaste": "技能急速", "Active": "主动", "Armor": "护甲", "ArmorPenetration": "护甲穿透", "AttackSpeed": "攻击速度", "Aura": "光环", "Bilgewater": "比尔吉沃特", "Boots": "鞋子", "Consumable": "消耗品", "CooldownReduction": "冷却缩减", "CriticalStrike": "暴击", "Damage": "攻击力", "GoldPer": "工资装", "Health": "生命值", "HealthRegen": "生命回复", "Jungle": "打野-起始", "Lane": "对线-起始", "LifeSteal": "生命偷取", "MagicPenetration": "法术穿透", "MagicResist": "魔法抗性", "Mana": "法力值", "ManaRegen": "法力回复", "Movement": "移动速度", "NonbootsMovement": "其它移动速度物品", "OnHit": "攻击特效", "Slow": "减速", "SpellBlock": "魔法抗性", "SpellDamage": "法术强度", "SpellVamp": "法术吸血", "Stealth": "潜行/隐身", "Tenacity": "韧性", "Trinket": "饰品", "Vision": "视野"}
                #下面设置装备表头的地图部分（Set the map part of the item headers）
                maps = {"8": {"zh_CN": "水晶之痕", "en_US": "Crystal Scar"}, "11": {"zh_CN": "召唤师峡谷", "en_US": "Summoner's Rift"}, "12": {"zh_CN": "嚎哭深渊", "en_US": "Howling Abyss"}, "14": {"zh_CN": "屠夫之桥", "en_US": "Butcher's Bridge"}, "16": {"zh_CN": "星界废墟", "en_US": "Cosmic Ruins"}, "18": {"zh_CN": "瓦洛兰城市公园", "en_US": "Valoran City Park"}, "19": {"zh_CN": "第43区", "en_US": "Substructure 43"}, "21": {"zh_CN": "百合与莲花的神庙", "en_US": "Temple of Lily and Lotus"}, "22": {"zh_CN": "聚点危机", "en_US": "Convergence"}, "30": {"zh_CN": "怒火角斗场", "en_US": "Rings of Wrath"}, "33": {"zh_CN": "最终都市", "en_US": "Final City"}, "35": {"zh_CN": "班德尔之森", "en_US": "The Bandlewood"}}
                mapIds = list(map(str, sorted(map(int, maps.keys()))))
                #下面设置装备表头的基础属性部分。这一部分需要按照实际情况随时更新。只需要增添新的，不需要删除旧的（Set the stat part of the item headers. This part needs update with the latest knowledge. Only need to add new keys, but not delete old keys）
                attributes = {"Health": "生命值", "Bonus Health": "额外生命值", "Mana": "法力值", "Attack Damage": "攻击力", "Ability Power": "法术强度", "Adaptive Force": "适应之力", "Armor": "护甲", "Magic Resist": "魔法抗性", "Attack Speed": "攻击速度", "Ability Haste": "技能急速", "Cooldown Reduction": "冷却缩减", "Critical Strike Chance": "暴击几率", "Critical Strike Damage": "暴击伤害", "Move Speed": "移动速度", "Base Health Regen": "基础生命回复", "Base Mana Regen": "基础法力回复", "Heal and Shield Power": "治疗和护盾强度", "Increased Healing from Potions": "来自药水的治疗效果", "Mana per level": "每级法力", "Mana regen per 5 seconds": "法力回复/5秒", "Lethality": "穿甲", "Armor Penetration": "护甲穿透", "Magic Penetration": "法术穿透", "Life Steal": "生命偷取", "Omnivamp": "全能吸血", "Life Steal vs. Monsters": "对野怪的生命偷取", "Life on Hit": "攻击时回复生命值", "Tenacity": "韧性", "Gold Per 10 Seconds": "金币/10秒", "Ability Power per level": "每级法术强度"}
                attribute_correct_map = {"Base Health Regeneration": "Base Health Regen", "Mana per 5 seconds": "Mana regen per 5 seconds", "Movement Speed": "Move Speed"} #早期的装备数据中存在一些不规范的数值属性称呼，这里将其规范成以上字典中包含的属性（The early item data contain some irregular calling of attributes, and this dictionary is designed to standardize them to be included in the above `attributes` dictionary）
                #下面设置装备表头（Set the item headers）
                LoLItem_header_en = base_header_keys + ["Map Availability: " + maps[mapId]["en_US"] for mapId in maps] + ["Class: " + tag for tag in tags] + list(attributes.keys())
                LoLItem_header_zh = base_header_values + ["地图可用性：" + maps[mapId]["zh_CN"] for mapId in maps] + ["类别：" + tags_dict[tag] for tag in tags] + list(attributes.values())
                LoLItem_header = {LoLItem_header_en[i]: LoLItem_header_zh[i] for i in range(len(LoLItem_header_en))}
                LoLItem_header_keys = list(LoLItem_header.keys())
                #print(LoLItem_header_keys)
                #定义常量字典（Define the constant dictionaries）
                LoLItem_name_map = {key: value["name"] for (key, value) in LoLItems_locale["data"].items()}
                print("开始整理数据……\nSorting data ...")
                LoLItem_data = {}
                pStats = re.compile(r"<stats>.*</stats>")
                pFormat = re.compile(r"<[/\sA-Za-z0-9=#\'_@]*>")
                champions = {}
                for champion in champions_locale["data"]:
                    champions[champion.lower()] = champions_locale["data"][champion]["name"] + " " + champions_locale["data"][champion]["title"] #装备数据中记录的英雄代号和英雄数据中的英雄代号有大小写上的差异（Case difference exists in the alias between item and champion data）
                for i in range(len(LoLItem_header_keys)):
                    key = LoLItem_header_keys[i]
                    LoLItem_data[key] = []
                for i in LoLItems_locale["data"]:
                    item = LoLItems_locale["data"][i]
                    item_default = LoLItems_default["data"][i]
                    #首先处理共有部分（First, deal with the common part）
                    ##下面填充装备的基本数据。这里参考的是英语描述（The following code fills the items' basic stats. Here the code refer to English descriptions）
                    statDict = {}
                    if "description" in item_default:
                        if pStats.search(item_default["description"]):
                            statStr = pStats.search(item_default["description"]).group().replace("<stats>", "").replace("</stats>", "").replace("<br>", "\n")
                            statList = statStr.split("\n")
                        else: #在0.152.55版本以前，装备详细信息的数值部分没有被<stats>和</stats>标签包起来。其数值部分总是出现在第一行，并且不同的数值中间由空格分隔（无效信息），每个数值前都有加号【Before v0.152.22, the stat part of an item's description isn't enclosed by <stats> and </stats> tags. In the description, the stat part is always the first line, different stats are delimited by a space (useless information) and a plus sign is always in the front of the stat value】
                            pNonStat = re.compile(r"[^\s\+A-Za-z0-9]")
                            statStr = item_default["description"].replace("<br>", "\n").split("\n")[0]
                            if pNonStat.search(statStr) != None: #部分描述不规范，直接将其它非数值文字放在与数值同一行的位置。这里的处理方式是将所有非数值字符都当成分隔符，然后取第一个元素（Some descriptions don't obey the standard, because the nonstat descriptions are put in the same line as the stats. Here the strategy is to regard any nonstat character as a delimiter, and then get the first element of the string split by the delimiter）
                                statStr = statStr.split(pNonStat.search(statStr).group(0))[0]
                            statList_tmp = statStr.split("+")
                            statList = []
                            for stat_iter in statList_tmp:
                                if stat_iter != "":
                                    statList.append(stat_iter.strip())
                        for stat_iter in statList:
                            if stat_iter != "": #有的装备没有基本属性，或者其字符串中存在几个连续的换行符（Some items don't have basic stats, or the string contains several continuous line feed characters）
                                # 下面注释起来的代码只适用于10.22版本后的装备数据（The following commented code only apply to item data after v10.22）
                                # figureType = pFormat.search(stat_iter).group().replace("<", "").replace(">", "")
                                # preFigure = "<" + figureType + ">"
                                # postFigure = "</" + figureType + ">"
                                # pFigure = re.compile(preFigure + ".*" + postFigure)
                                # figure = pFigure.search(stat_iter).group().replace(preFigure, "").replace(postFigure, "")
                                # figure_attr = stat_iter.replace(pFigure.search(stat_iter).group(), "").strip() #英文中，数值和属性之间有空格（In English, there's a space between the stat and the attribute）
                                # statDict[figure_attr] = figure
                                while pFormat.search(stat_iter):
                                    stat_iter = stat_iter.replace(pFormat.search(stat_iter).group(), "")
                                pFigure = re.compile(r"(\+|\-)?[0-9]+%?")
                                try:
                                    figure = pFigure.search(stat_iter).group()
                                    figure_attr = stat_iter.replace(figure, "").strip() #英文中，数值和属性之间有空格（In English, there's a space between the stat and the attribute）
                                    if figure_attr in attribute_correct_map:
                                        figure_attr = attribute_correct_map[figure_attr]
                                    statDict[figure_attr] = figure.replace("+", "")
                                except AttributeError:
                                    pass
                    #然后分类讨论（Then discuss about `j`)
                    for j in range(len(LoLItem_header_keys)):
                        key = LoLItem_header_keys[j]
                        if j < len(base_header_keys): #基本表头部分（Base part）
                            if j == 0: #键（Key）
                                try:
                                    LoLItem_data[key].append(int(i))
                                except ValueError: #在12.21.1版本出现了装备序号为TalentReaperItem的装备（An item with itemId "TalentReaperItem" appears in the item data of v12.21.1）
                                    traceback_info = traceback.format_exc()
                                    print(traceback_info)
                                    LoLItem_data[key].append(i)
                            elif j <= 22:
                                if j == 2: #详细信息（`description`）
                                    desc = item["description"].replace("<br>", "\n")
                                    while pFormat.search(desc):
                                        desc = desc.replace(pFormat.search(desc).group(), "")
                                    LoLItem_data[key].append(desc)
                                elif j == 16: #装备名称（`localizedName`）
                                    LoLItem_data[key].append(item["name"])
                                elif j == 17: #英文名称（`name`）
                                    LoLItem_data[key].append(item_default["name"])
                                elif j == 18: #合成材料名称（`fromName`）
                                    LoLItem_data[key].append(list(map(lambda x: LoLItem_name_map.get(x, x), item["from"])) if "from" in item else "")
                                elif j == 19: #合成装备名称（`intoName`）
                                    LoLItem_data[key].append(list(map(lambda x: LoLItem_name_map.get(x, x), item["into"])) if "into" in item else "") #吞噬者仅出现在合成装备中（Devourer only occurs as an item to upgrade into）
                                elif j == 20: #装备持有者名称（`requiredChampionName`）
                                    LoLItem_data[key].append(champions[item["requiredChampion"].lower()] if "requiredChampion" in item and item["requiredChampion"] != "" else "")
                                elif j == 21: #所需队友名称（`requiredAllyName`）
                                    LoLItem_data[key].append(champions[item["requiredAlly"].lower()] if "requiredAlly" in item and item["requiredAlly"] != "" else "")
                                elif j == 22: #特殊合成材料名称（`specialRecipeName`）
                                    LoLItem_data[key].append(LoLItem_name_map.get(item["specialRecipe"], "") if "specialRecipe" in item and item["specialRecipe"] != 0 else "")
                                elif j in {5, 8, 12, 13, 24}: #逻辑值（Values of boolean type）
                                    LoLItem_data[key].append(item.get(key, False))
                                else:
                                    LoLItem_data[key].append(item.get(key, ""))
                            else:
                                if j == 23: #合成费用（`baseGold`）
                                    LoLItem_data[key].append(item["gold"]["base"])
                                elif j == 24: #可以购买（`purchasable`）
                                    LoLItem_data[key].append(item["gold"]["purchasable"])
                                elif j == 25: #总费用（`totalGold`）
                                    LoLItem_data[key].append(item["gold"]["total"])
                                else: #售价（`sellGold`）
                                    LoLItem_data[key].append(item["gold"]["sell"])
                        elif j < len(base_header_keys) + len(mapIds): #地图部分（Map part）
                            LoLItem_data[key].append(item["maps"].get(mapIds[j - len(base_header_keys)], False) if "maps" in item else False)
                        elif j < len(base_header_keys) + len(mapIds) + len(tags): #分类部分（Category part）
                            LoLItem_data[key].append(tags[j - len(base_header_keys) - len(mapIds)] in item["tags"])
                        else: #基础属性部分（Stat part）
                            if "description" in item_default:
                                key_default = list(attributes.keys())[j - len(base_header_keys) - len(mapIds) - len(tags)]
                                LoLItem_data[key].append(statDict.get(key_default, ""))
                            else:
                                LoLItem_data[key].append("")
                base_statistics_display_order = [0, 16, 17, 4, 5, 8, 12, 13, 24, 23, 25, 26, 6, 7, 20, 21, 18, 19, 22, 3, 2]
                LoLItem_statistics_display_order = base_statistics_display_order + list(range(len(base_statistics_display_order), len(LoLItem_header)))
                LoLItem_data_organized = {}
                for i in LoLItem_statistics_display_order:
                    key = LoLItem_header_keys[i]
                    LoLItem_data_organized[key] = LoLItem_data[key]
                LoLItem_df = pandas.DataFrame(data = LoLItem_data_organized)
                for column in LoLItem_df:
                    if LoLItem_df[column].dtype == "bool":
                        LoLItem_df[column] = LoLItem_df[column].astype(str)
                        for i in range(len(LoLItem_df)):
                            LoLItem_df.loc[i, column] = "√" if LoLItem_df[column][i] == "True" else ""
                LoLItem_df.index = list(range(1, len(LoLItem_df) + 1))
                try:
                    LoLItem_df = LoLItem_df.sort_values(by = "id", ascending = True)
                except TypeError: #在12.21.1版本出现了装备序号为TalentReaperItem的装备，导致无法正常按照装备序号排序（An item with itemId "TalentReaperItem" appears in the item data of v12.21.1, which makes it disallowed to sort the values by the itemIds）
                    traceback_info = traceback.format_exc()
                    print(traceback_info)
                LoLItem_df = pandas.concat([pandas.DataFrame([LoLItem_header])[LoLItem_df.columns], LoLItem_df])
                LoLItem_dfs[version] = LoLItem_df.copy(deep = True)
                item_df_formed = True
                end = time.time()
                unit = end - start
                total_used += unit
                runTimes.append(unit)
                total_remaining = sum(runTimes) / (version_index + 1 - failed_count) * (len(versions_sort) - version_index - 1)
                print("整理该版本数据所花费的时间（Time spent in sorting out this version）：", format_runtime(unit))
                print("已花费的总时间（Total time used）                                   ：", format_runtime(total_used))
                print("剩余时间（Time remaining）                                          ：", format_runtime(total_remaining))
                print("预计总时间（Expected total time）                                   ：", format_runtime(total_used + total_remaining), end = "\n\n")
    else:
        cdragon_home_url = "https://raw.communitydragon.org/"
        LoLItem_dfs = {}
        while True:
            back = False
            #下面的代码块生成英雄联盟专业术语中英文对照的Sheet1的内容。数据来源是CommunityDragon数据库（The following code block generates the Sheet1 in LoL Term Translation - zh_CN & en_US. Data resources are from CommunityDragon database）
            print("请在以下版本号中选择并输入完整的版本号：\nPlease select a version and then enter it entirely:")
            #下面从CommunityDragons数据库主页的源代码获取可用版本（The following code crawl the available patches in CommunityDragon database through its homepage）
            src, status = getUrl(cdragon_home_url)
            if status != 0:
                if status == 1:
                    print("CommunityDragon数据库主页访问失败！\nCommunityDragon database homepage access failed!")
                elif status == 404:
                    print("CommunityDragon数据库主页不存在！可能它已经变更了。\nCommunityDragon database homepage not found! Maybe it's changed.")
                break
            cdragon_homepage = src
            source = cdragon_homepage.content.decode()
            source_list = list(map(lambda x: x.strip(), source.split("\n")))
            line_re = re.compile(r'<tr><td class="link"><a href="[0-9]*\.[0-9]*/" title="[0-9]*\.[0-9]*">[0-9]*\.[0-9]*/</a></td><td class="size">-</td><td class="date">[0-9]*-[a-zA-Z]*-[0-9]* [0-9]*:[0-9]*</td></tr>')
            patch_re = re.compile(r'[0-9]*\.[0-9]*')
            patches_cdragon = []
            for line in source_list:
                matchedLine = line_re.search(line) #先通过一个比较长的正则表达式筛选包含版本信息的CSS代码行（First filter the CSS code lines that contain patch information through a long regular expression）
                if matchedLine:
                    matchedPatch = patch_re.search(line).group() #在包含版本信息的CSS代码中再获取版本字符串（Then obtains patch string from the CSS code that contain it）
                    patches_cdragon.append(matchedPatch)
            patches_cdragon = patch_sort(patches_cdragon)
            patches_cdragon.reverse()
            patches_cdragon.insert(0, "pbe")
            patches_cdragon.insert(0, "latest")
            print(patches_cdragon)
            
            while True:
                version = input()
                if version == "":
                    versions_sort = ["pbe"]
                    break
                elif version == "all":
                    versions_sort = patches_cdragon
                    break
                elif version[0] == "0":
                    back = True
                    break
                elif version in patches_cdragon:
                    versions_sort = [version]
                    break
                else:
                    try:
                        versions_sort = eval(version)
                    except:
                        print("您的输入有误，请重新输入！\nERROR input! Please try again.")
                    else:
                        if isinstance(versions_sort, list) and all(map(lambda x: x in patches_cdragon, versions_sort)):
                            break
                        else:
                            print("您的输入有误，请重新输入！\nERROR input! Please try again.")
            if back:
                break
            runTimes = [] #记录整理一个版本的装备数据所花费的时间（Records the time spent in sorting out item data of one version）
            total_used = 0
            failed_count = 0 #记录数据获取失败的版本的个数（Count the number of versions whose data fail to be fetched）
            for version_index in range(len(versions_sort)):
                start = time.time()
                print("[%s]" %(time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())), end = "")
                print("整理进度（Sorting process）：%d/%d" %(version_index + 1, len(versions_sort)))
                version = versions_sort[version_index]
                LoLItems_locale_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/items.json" %(version, language_cdragon[language_code])
                LoLItems_default_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/default/v1/items.json" %version
                champions_locale_url = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/champion-summary.json" %(version, language_cdragon[language_code])
                fontconfig_locale_url = "https://raw.communitydragon.org/%s/game/data/menu/fontconfig_%s.txt.json" %(version, language_code.lower())
                strtable_locale_url1 = "https://raw.communitydragon.org/%s/game/data/menu/main_%s.stringtable.json" %(version, language_code.lower())
                strtable_locale_url2 = "https://raw.communitydragon.org/%s/game/%s/data/menu/en_us/main.stringtable.json" %(version, language_code.lower())
                strtable_locale_url3 = "https://raw.communitydragon.org/%s/game/%s/data/menu/en_us/lol.stringtable.json" %(version, language_code.lower())
                LoLItems_binary_url1 = "https://raw.communitydragon.org/%s/game/global/items/items.bin.json" %(version)
                LoLItems_binary_url2 = "https://raw.communitydragon.org/%s/game/items.cdtb.bin.json" %(version)
                print(f"正在获取{version}版本的目标语言装备信息……\nFetching LoL item information of version {version} in target language ...")
                src, status = getUrl(LoLItems_locale_url)
                if status != 0:
                    if status == 1:
                        print("目标语言装备信息获取失败！\nLoL item information in target language capture failed!")
                    elif status == 404:
                        print("目标语言装备信息文件不存在！\nLoL item file in target language not found!")
                    failed_count += 1
                    continue
                LoLItems_locale = src.json()
                print(f"正在获取{version}版本的英文装备信息……\nFetching LoL item information of version {version} in English ...")
                src, status = getUrl(LoLItems_default_url)
                if status != 0:
                    if status == 1:
                        print("英文装备信息获取失败！\nLoL item information in English capture failed!")
                    elif status == 404:
                        print("英文装备信息文件不存在！\nLoL item file in English not found!")
                    failed_count += 1
                    continue
                LoLItems_default = src.json()
                print(f"正在获取{version}版本的目标语言英雄信息……\nFetching champion information of version {version} in target language ...")
                src, status = getUrl(champions_locale_url)
                if status != 0:
                    if status == 1:
                        print("目标语言英雄信息获取失败！\nChampion information in target language capture failed!")
                    elif status == 404:
                        print("目标语言英雄信息文件不存在！\nChampion file in target language not found!")
                    failed_count += 1
                    continue
                champions_locale = src.json()
                print(f"正在获取{version}版本的目标语言字符串常量池（英雄联盟）……\nFetching stringtable (LoL) of version {version} in target language ...")
                src, status = getUrl(fontconfig_locale_url if patch_compare(version, "12.23") else strtable_locale_url1 if patch_compare(version, "14.4") else strtable_locale_url2 if patch_compare(version, "14.15") else strtable_locale_url3) #翻译数据在12.23、14.4和14.15版本发生了路径迁移（Path transfer occurred to the localization data in Patches 12.23, 14.4 and 14.15）
                status_strtable_locale = status
                if status != 0:
                    if status == 1:
                        print("目标语言字符串常量池（英雄联盟）获取失败！\nStringtable (LoL) in target language capture failed!")
                    elif status == 404:
                        print("目标语言字符串常量池（英雄联盟）文件不存在！\nStringtable (LoL) file in target language not found!")
                if status_strtable_locale == 0:
                    strtable_locale = src.json()
                elif status_strtable_locale == 404:
                    strtable_locale = {}
                else:
                    failed_count += 1
                    continue
                print(f"正在获取{version}版本的二进制装备信息……\nFetching binary LoL item of version {version} ...")
                src, status = getUrl(LoLItems_binary_url1 if patch_compare(version, "13.15") else LoLItems_binary_url2) #二进制装备信息在13.15版本发生了路径迁移（Path transfer occurred to the binary item information in Patch 13.15）
                status_item_binary = status
                if status != 0:
                    if status == 1:
                        print("目标语言二进制装备信息获取失败！\nBinary LoL item information capture failed!")
                    elif status == 404:
                        print("目标语言二进制装备信息文件不存在！\nBinary LoL item file not found!")
                if status_item_binary == 0:
                    LoLItems_binary = src.json()
                elif status_item_binary == 404:
                    LoLItems_binary = {}
                else:
                    failed_count += 1
                    continue
                
                #下面设置装备表头的元数据部分（Set the metadata part of the item headers）
                base_header = {"id": "装备序号", "active": "主动使用", "description": "描述", "inStore": "游戏内可见性", "from": "合成材料序号", "to": "合成装备序号", "maxStacks": "最大持有数量", "requiredChampion": "装备持有者", "requiredAlly": "所需队友", "requiredBuffCurrencyName": "其它货币类型", "requiredBuffCurrencyCost": "其它费用", "specialRecipe": "特殊合成材料", "isEnchantment": "附魔装备", "price": "合成费用", "priceTotal": "总费用", "displayInItemSets": "装备图册可见性", "iconPath": "缩略图路径", "localizedName": "装备名称", "name": "英文名称", "fromName": "合成材料名称", "toName": "合成装备名称", "requiredChampionName": "装备持有者名称", "requiredAllyName": "所需队友名称", "specialRecipeName": "特殊合成材料名称"}
                base_header_keys = list(base_header.keys())
                base_header_values = list(base_header.values())
                #下面设置装备表头的详细信息部分（Set the tooltip part of the item headers）
                tooltip_header = {"tooltip": "游戏内详细信息"}
                tooltip_header_keys = list(tooltip_header.keys())
                tooltip_header_values = list(tooltip_header.values())
                #下面设置装备表头的分类（标签）部分（Set the category / tag part of the item headers）
                categories_initial = set()
                for item in LoLItems_locale:
                    categories_initial |= set(map(lambda x: x.lower(), item.get("categories", []))) #在7.8版本以前，装备数据中无“categories”键（Before Patch 7.8, "categories" key isn't present in item data）
                categories_initial = sorted(categories_initial)
                categories_organized = ["Lane", "Jungle", "GoldPer", "Boots", "Consumable", "Damage", "CriticalStrike", "AttackSpeed", "OnHit", "ArmorPenetration", "SpellDamage", "Mana", "ManaRegen", "MagicPenetration", "Health", "HealthRegen", "MagicResist", "AbilityHaste", "CooldownReduction", "Movement", "NonbootsMovement", "LifeSteal", "SpellVamp", "Active", "Armor", "Aura", "Slow", "SpellBlock", "Stealth", "Tenacity", "Trinket", "Vision", "Bilgewater"] #设置分类表头的顺序（Set the order of category headers）
                categories = []
                for category in categories_organized:
                    if category.lower() in categories_initial:
                        categories_initial.remove(category.lower())
                    categories.append(category)
                categories += categories_initial
                categories_dict = {"AbilityHaste": "技能急速", "Active": "主动", "Armor": "护甲", "ArmorPenetration": "护甲穿透", "AttackSpeed": "攻击速度", "Aura": "光环", "Bilgewater": "比尔吉沃特", "Boots": "鞋子", "Consumable": "消耗品", "CooldownReduction": "冷却缩减", "CriticalStrike": "暴击", "Damage": "攻击力", "GoldPer": "工资装", "Health": "生命值", "HealthRegen": "生命回复", "Jungle": "打野-起始", "Lane": "对线-起始", "LifeSteal": "生命偷取", "MagicPenetration": "法术穿透", "MagicResist": "魔法抗性", "Mana": "法力值", "ManaRegen": "法力回复", "Movement": "移动速度", "NonbootsMovement": "其它移动速度物品", "OnHit": "攻击特效", "Slow": "减速", "SpellBlock": "魔法抗性", "SpellDamage": "法术强度", "SpellVamp": "法术吸血", "Stealth": "潜行/隐身", "Tenacity": "韧性", "Trinket": "饰品", "Vision": "视野"}
                #下面设置装备表头的基础属性部分。这一部分需要按照实际情况随时更新。只需要增添新的，不需要删除旧的（Set the basic stat part of the item headers. This part needs update with the latest knowledge. Only need to add new keys, but not delete old keys）
                attributes = {"Health": "生命值", "Bonus Health": "额外生命值", "Mana": "法力值", "Attack Damage": "攻击力", "Ability Power": "法术强度", "Adaptive Force": "适应之力", "Armor": "护甲", "Magic Resist": "魔法抗性", "Attack Speed": "攻击速度", "Ability Haste": "技能急速", "Cooldown Reduction": "冷却缩减", "Critical Strike Chance": "暴击几率", "Critical Strike Damage": "暴击伤害", "Move Speed": "移动速度", "Base Health Regen": "基础生命回复", "Base Mana Regen": "基础法力回复", "Heal and Shield Power": "治疗和护盾强度", "Increased Healing from Potions": "来自药水的治疗效果", "Mana per level": "每级法力", "Mana regen per 5 seconds": "法力回复/5秒", "Lethality": "穿甲", "Armor Penetration": "护甲穿透", "Magic Penetration": "法术穿透", "Life Steal": "生命偷取", "Omnivamp": "全能吸血", "Life Steal vs. Monsters": "对野怪的生命偷取", "Life on Hit": "攻击时回复生命值", "Tenacity": "韧性", "Gold Per 10 Seconds": "金币/10秒", "Ability Power per level": "每级法术强度"}
                attribute_correct_map = {"Base Health Regeneration": "Base Health Regen", "Mana per 5 seconds": "Mana regen per 5 seconds", "Movement Speed": "Move Speed"} #早期的装备数据中存在一些不规范的数值属性称呼，这里将其规范成以上字典中包含的属性（The early item data contain some irregular calling of attributes, and this dictionary is designed to standardize them to be included in the above `attributes` dictionary）
                #下面设置装备表头的所有数值部分。这一部分需要按照实际情况随时更新。只需要增添新的，不需要删除旧的（Set the detailed stat part of the item headers. This part needs update with the latest knowledge. Only need to add new keys, but not delete old keys）
                allStats_header = {}
                allStats_header_keys = list(allStats_header.keys())
                allStats_header_values = list(allStats_header.values())
                #下面设置装备表头（Set the item headers）
                LoLItem_header_en = base_header_keys + tooltip_header_keys + ["Class: " + category for category in categories] + list(attributes.keys()) + (allStats_header_keys if status_item_binary == 0 else [])
                LoLItem_header_zh = base_header_values + tooltip_header_values + ["类别：" + categories_dict[category] for category in categories] + list(attributes.values()) + (allStats_header_values if status_item_binary == 0 else [])
                LoLItem_header = {LoLItem_header_en[i]: LoLItem_header_zh[i] for i in range(len(LoLItem_header_en))}
                LoLItem_header_keys = list(LoLItem_header.keys())
                #print(LoLItem_header_keys)
                #定义常量字典（Define the constant dictionaries）
                LoLItem_name_map = {item["id"]: item["name"] for item in LoLItems_locale}
                print("开始整理数据……\nSorting data ...")
                LoLItem_data = {}
                pStats = re.compile(r"<stats>.*</stats>")
                pFormat = re.compile(r"<[/\sA-Za-z0-9=#\'_@]*>")
                pIcon = re.compile(r"{{[/\sA-Za-z0-9=#\'_@]*}}|%i:\w*%")
                pSection = re.compile(r"<section>.*?</section>") #在星号后添加问号以启用贪婪模式（Enable greedy match by adding a question mark after the asterisk）
                pAttributes = re.compile(r"@.*?@")
                pAlphaNumeric = re.compile(r"\w+")
                pAlphabetic = re.compile(r"[A-Za-z_]+")
                champions = {}
                for champion in champions_locale:
                    champions[champion["alias"]] = champion["name"] + " " + (champion["description"] if "description" in champion else champion["alias"]) #15.9版本以前，“champion-summary.json”中没有“description”键（Before Patch 15.9, "description" key isn't present in "champion-summary.json"）
                for i in range(len(LoLItem_header_keys)):
                    key = LoLItem_header_keys[i]
                    LoLItem_data[key] = []
                for i in range(len(LoLItems_locale)):
                    item = LoLItems_locale[i]
                    item_default = LoLItems_default[i]
                    item_binary = LoLItems_binary.get("Items/%d" %(item["id"]), {}) #确定该装备在二进制json文件中的数据。在15.12.685.0388版本，神木之门引入的三个饮品没有出现在二进制json文件中（Determine item data in the binary json file. In Patch 15.12.685.0388, 3 juices introduced into Brawl don't exist in the binary json file）
                    #首先处理共有部分（First, deal with the common part）
                    ##下面填充装备的基本数据。这里参考的是英语描述（The following code fills the items' basic stats. Here the code refer to English descriptions）
                    statDict = {}
                    if "description" in item_default and pStats.search(item_default["description"]):
                        statStr = pStats.search(item_default["description"]).group().replace("<stats>", "").replace("</stats>", "").replace("<br>", "\n")
                        statList = statStr.split("\n")
                        for stat_iter in statList:
                            if stat_iter != "": #有的装备没有基本属性，或者其字符串中存在几个连续的换行符（Some items don't have basic stats, or the string contains several continuous line feed characters）
                                while pFormat.search(stat_iter):
                                    stat_iter = stat_iter.replace(pFormat.search(stat_iter).group(), "")
                                pFigure = re.compile(r"(\+|\-)?[0-9]+%?")
                                try:
                                    figure = pFigure.search(stat_iter).group()
                                    figure_attr = stat_iter.replace(figure, "").strip() #英文中，数值和属性之间有空格（In English, there's a space between the stat and the attribute）
                                    if figure_attr in attribute_correct_map:
                                        figure_attr = attribute_correct_map[figure_attr]
                                    statDict[figure_attr] = figure.replace("+", "")
                                except AttributeError:
                                    pass
                    ##下面确定该装备在二进制json文件中存储的属性及其值（The following code determine attributes and corresponding values of this item stored in the binary json file）
                    mDataValues = {}
                    if "mDataValues" in item_binary:
                        for itemDataValue_iter in item_binary["mDataValues"]:
                            if all(j in itemDataValue_iter for j in ["mName", "mValue", "__type"]):
                                mDataValues[itemDataValue_iter["mName"]] = itemDataValue_iter["mValue"]
                    #然后分类讨论（Then discuss about `j`)
                    for j in range(len(LoLItem_header_keys)):
                        key = LoLItem_header_keys[j]
                        if j < len(base_header_keys): #基本表头部分（Base part）
                            if j == 2: #详细信息（`description`）
                                desc = item["description"].replace("<br>", "\n")
                                while pFormat.search(desc):
                                    desc = desc.replace(pFormat.search(desc).group(), "")
                                LoLItem_data[key].append(desc)
                            elif j == 17: #装备名称（`localizedName`）
                                LoLItem_data[key].append(item["name"])
                            elif j == 18: #英文名称（`name`）
                                LoLItem_data[key].append(item_default["name"])
                            elif j == 19: #合成材料名称（`fromName`）
                                LoLItem_data[key].append(list(map(lambda x: LoLItem_name_map.get(x, x), item["from"])) if "from" in item else "") #在7.8版本以前，装备数据中无“from”键（Before Patch 7.8, "from" key isn't present in item data）
                            elif j == 20: #合成装备名称（`toName`）
                                LoLItem_data[key].append(list(map(lambda x: LoLItem_name_map.get(x, x), item["to"])) if "to" in item else "") #在7.8版本以前，装备数据中无“to”键（Before Patch 7.8, "to" key isn't present in item data）
                            elif j == 21: #装备持有者名称（`requiredChampionName`）
                                LoLItem_data[key].append(champions[item["requiredChampion"]] if "requiredChampion" in item and item["requiredChampion"] != "" else champions[item["requiredchampion"]] if "requiredchampion" in item and item["requiredchampion"] != "" else "") #在7.10版本以前，部分键是小写形式（Before Patch 7.10, some keys are in lower case）
                            elif j == 22: #所需队友名称（`requiredAllyName`）
                                LoLItem_data[key].append(champions[item["requiredAlly"]] if "requiredAlly" in item and item["requiredAlly"] != "" else champions[item["requiredally"]] if "requiredally" in item and item["requiredally"] != "" else "") #在7.16版本以前，装备数据中无“requiredAlly”键（Before Patch 7.16, "requiredAlly" key isn't present in item data）
                            elif j == 23: #特殊合成材料名称（`specialRecipeName`）
                                LoLItem_data[key].append(LoLItem_name_map.get(item["specialRecipe"], "") if "specialRecipe" in item and item["specialRecipe"] != 0 else LoLItem_name_map.get(item["specialrecipe"], "") if "specialrecipe" in item and item["specialrecipe"] != 0 else "")  #在7.10版本以前，部分键是小写形式（Before Patch 7.10, some keys are in lower case）
                            elif j in {1, 3, 12, 15}: #逻辑类键（Keys of boolean type）
                                LoLItem_data[key].append(item.get(key, item.get(key.lower(), False))) #在14.15版本以前，装备数据中无“displayInItemSets”键（Before Patch 14.15, "displayInItemSets" key isn't in item data）
                            else:
                                LoLItem_data[key].append(item.get(key, item.get(key.lower(), ""))) #在7.10版本以前，部分键是小写形式；在7.8版本以前，装备数据中无大多数键（Before Patch 7.10, some keys are in lower case. Before Patch 7.8, most keys aren't present in item data）
                        elif j < len(base_header_keys) + len(tooltip_header_keys): #游戏内详细信息部分（In-game tooltip part）
                            if status_strtable_locale == 0:
                                entry_key = "generatedtip_item_%d_tooltipinventoryextended" %(item["id"])
                                if entry_key in strtable_locale["entries"]:
                                    tooltip = strtable_locale["entries"][entry_key]
                                else: #早期版本中没有按Shift查看详细信息的说法（In early versions, pressing Shift won't provide the detailed description）
                                    entry_key = "generatedtip_item_%d_tooltipinventory" %(item["id"])
                                    if entry_key in strtable_locale["entries"]:
                                        tooltip = strtable_locale["entries"][entry_key]
                                    else:
                                        tooltip = "" #空字符串仍然适用于下面的格式替换（The following format transformation applies to an empty string）
                            else:
                                tooltip = ""
                            if args.transform:
                                tooltip_tmp = tooltip
                                tooltip_layers = {} #将详细信息按照第一层级分为几个部分。一般包括titleLeft、titleRight、subtitleLeft、subtitleRight、mainText和postScriptTitle等几个部分（Divide the details into several parts according to the first layer, basically including titleLeft, titleRight, subtitleLeft, subtitleRight, mainText, postScriptTitle, etc.）
                                while len(tooltip_tmp) > 0:
                                    first_layer_tag_start = pFormat.search(tooltip_tmp).group()
                                    first_layer_tag_end = first_layer_tag_start[0] + "/" + first_layer_tag_start[1:]
                                    first_layer_tag_start_indices = []
                                    first_layer_tag_end_indices = []
                                    for match in re.finditer(first_layer_tag_start, tooltip_tmp):
                                        first_layer_tag_start_indices.append(match.start())
                                    for match in re.finditer(first_layer_tag_end, tooltip_tmp):
                                        first_layer_tag_end_indices.append(match.start())
                                    tag_index_dict = {}
                                    for k in first_layer_tag_start_indices:
                                        tag_index_dict[k] = 1 #1代表新一层级的开始（1 represents the start of a new layer）
                                    for k in first_layer_tag_end_indices:
                                        tag_index_dict[k] = -1 #-1代表当前层级的结束（-1 represents the end of the current layer）
                                    layer_tag_stack = 1 #通过堆栈来判断是否达到第一层级的结束开关（Judge by a stack whether the closing tag of the first layer is reached）
                                    for k in sorted(tag_index_dict.keys())[1:]:
                                        layer_tag_stack += tag_index_dict[k]
                                        if layer_tag_stack == 0:
                                            break
                                    tooltip_layer = tooltip_tmp[:k + len(first_layer_tag_end)]
                                    tooltip_layers[first_layer_tag_start.replace("<", "").replace(">", "")] = tooltip_layer
                                    tooltip_tmp = tooltip_tmp[k + len(first_layer_tag_end):]
                                tooltip_layers_text = {}
                                for tag in tooltip_layers:
                                    tooltip_layer = tooltip_layers[tag]
                                    if pSection.search(tooltip_layer) == None:
                                        sections = [tooltip_layer] #神话版本的装备数据中没有<section>和</section>标签。这里的处理方法是将整个层视为一节。由于列表长度是1，所以后续在合并成字符串时也不会出现节与节之间的分隔符（In mythic item versions, <section> and </section> tags weren't present. In that case, that whole layer is regarded as a section. Since the list size is 1, no delimiters will be added when this list is going to concatenate into a string）
                                    else:
                                        sections = pSection.findall(tooltip_layer)
                                    for i in range(len(sections)):
                                        section = sections[i].replace("<br>", "\n").replace("<li>", "\n-\n")
                                        while pFormat.search(section):
                                            section = section.replace(pFormat.search(section).group(), "")
                                        while pIcon.search(section):
                                            section = section.replace(pIcon.search(section).group(), "")
                                        section = section.strip()
                                        while section.startswith("<br>"):
                                            section = section.lstrip("<br")
                                        while section.endswith("<br>"):
                                            section = section.rstrip("<br")
                                        mNames = pAttributes.findall(section)
                                        for mName in mNames:
                                            mName_literal = mName.replace("@", "") #这里默认不存在“@@”（Suppose "@@" doesn't exist here）
                                            mName_literal_decapitalize = mName_literal[0].lower() + mName_literal[1:]
                                            if mName_literal in mDataValues:
                                                section = section.replace(mName, str(round(mDataValues[mName_literal])))
                                            elif mName_literal in item_binary:
                                                section = section.replace(mName, str(round(item_binary[mName_literal])))
                                            elif mName_literal_decapitalize in item_binary: #二进制装备数据中，部分键的首字母被设为小写（In the binary item data, the capital letter of some keys is in lower case）
                                                section = section.replace(mName, str(round(item_binary[mName_literal_decapitalize])))
                                            elif f"m{mName_literal}" in item_binary:
                                                section = section.replace(mName, str(round(item_binary[f"m{mName_literal}"])))
                                            else: #这部分处理被双@围起来的公式（This part handles the formula surrounded by two @s）
                                                mName_literal_names = pAlphaNumeric.findall(mName)
                                                for mName_literal_name in mName_literal_names:
                                                    mName_literal_name_decapitalize = mName_literal_name[0].lower() + mName_literal_name[1:]
                                                    if mName_literal_name in mDataValues:
                                                        mName_literal = mName_literal.replace(mName_literal_name, str(mDataValues[mName_literal_name]))
                                                    elif mName_literal_name in item_binary:
                                                        mName_literal = mName_literal.replace(mName_literal_name, str(item_binary[mName_literal_name]))
                                                    elif mName_literal_name_decapitalize in item_binary: #二进制装备数据中，部分键的首字母被设为小写（In the binary item data, the capital letter of some keys is in lower case）
                                                        mName_literal = mName_literal.replace(mName_literal_name, str(item_binary[mName_literal_name_decapitalize]))
                                                    elif f"m{mName_literal_name}" in item_binary:
                                                        mName_literal = mName_literal.replace(mName_literal_name, str(item_binary[f"m{mName_literal_name}"]))
                                                if pAlphabetic.search(mName_literal) == None:
                                                    try:
                                                        section = section.replace(mName, str(round(eval(mName_literal)))) #所有装备的数值都显示为整数（All item stats are displayed as integers）
                                                    except SyntaxError: #在13.14版本中，月石再生器和降星者的描述出现格式错误，导致一个百分号被放在了双@内（In Patch 13.14, there was a format mistake in Moonstone Renewer's and Starcaster's tooltips, where a percent sign is put within two @s）
                                                        traceback_info = traceback.format_exc()
                                                        print(traceback_info)
                                                        pass
                                                else:
                                                    section = section.replace(mName, f"@{mName_literal}@")
                                        sections[i] = section
                                    while "" in sections:
                                        sections.remove("")
                                    if tag in {"titleLeft", "titleRight", "subtitleLeft", "subtitleRight"}:
                                        tooltip_layer_text = " - ".join(sections)
                                    else:
                                        tooltip_layer_text = "\n----\n".join(sections)
                                    tooltip_layers_text[tag] = tooltip_layer_text
                                tooltip_text = ""
                                if len(tooltip_layers_text) > 2:
                                    for i in range(len(tooltip_layers_text) - 1):
                                        tag = list(tooltip_layers_text.keys())[i]
                                        tag_next = list(tooltip_layers_text.keys())[i + 1]
                                        tooltip_layer_text = tooltip_layers_text[tag]
                                        if tag == "titleLeft" and tag_next == "titleRight" or tag == "subtitleLeft" and tag_next == "subtitleRight":
                                            tooltip_text += tooltip_layer_text + " | "
                                        elif tag == "subtitleRight" and tag_next == "mainText" or tag == "mainText" and tag_next in {"postScriptTitle", "postScriptLeft"}:
                                            tooltip_text += tooltip_layer_text + "\n--------\n"
                                        else:
                                            tooltip_text += tooltip_layer_text + "\n"
                                    tooltip_text += tooltip_layers_text[tag_next]
                                LoLItem_data[key].append(tooltip_text)
                            else:
                                LoLItem_data[key].append(tooltip)
                        elif j < len(base_header_keys) + len(tooltip_header_keys) + len(categories): #分类部分（Category part）
                            LoLItem_data[key].append(categories[j - len(base_header_keys) - len(tooltip_header_keys)] in item["categories"] if "categories" in item else False) #在7.8版本以前，装备数据中无“categories”键（Before Patch 7.8, "categories" key isn't present in item data）
                        elif j < len(base_header_keys) + len(tooltip_header_keys) + len(categories) + len(attributes): #基础属性部分（Stat part）
                            if "description" in item_default:
                                key_default = list(attributes.keys())[j - len(base_header_keys) - len(tooltip_header_keys) - len(categories)]
                                LoLItem_data[key].append(statDict.get(key_default, ""))
                            else:
                                LoLItem_data[key].append("")
                        else:
                            key_decapitalized = key[0].lower() + key[1:]
                            if key in mDataValues:
                                attributeValue = mDataValues[key]
                            elif key in item_binary:
                                attributeValue = item_binary[key]
                            elif key_decapitalized in item_binary:
                                attributeValue = item_binary[key_decapitalized]
                            elif f"m{key}" in item_binary:
                                attributeValue = item_binary[f"m{key}"]
                            else:
                                attributeValue = ""
                            LoLItem_data[key].append(attributeValue)
                base_statistics_display_order = [0, 17, 18, 1, 3, 19, 20, 6, 13, 14, 21, 22, 9, 10, 23, 12, 15, 16, 2]
                LoLItem_statistics_display_order = base_statistics_display_order + list(range(len(base_header_keys), len(LoLItem_header)))
                LoLItem_data_organized = {}
                for i in LoLItem_statistics_display_order:
                    key = LoLItem_header_keys[i]
                    LoLItem_data_organized[key] = LoLItem_data[key]
                LoLItem_df = pandas.DataFrame(data = LoLItem_data_organized)
                for column in LoLItem_df:
                    if LoLItem_df[column].dtype == "bool":
                        LoLItem_df[column] = LoLItem_df[column].astype(str)
                        for i in range(len(LoLItem_df)):
                            LoLItem_df.loc[i, column] = "√" if LoLItem_df[column][i] == "True" else ""
                LoLItem_df.index = list(range(1, len(LoLItem_df) + 1))
                LoLItem_df = LoLItem_df.sort_values(by = "id", ascending = True)
                LoLItem_df = pandas.concat([pandas.DataFrame([LoLItem_header])[LoLItem_df.columns], LoLItem_df])
                LoLItem_dfs[version] = LoLItem_df.copy(deep = True)
                item_df_formed = True
                end = time.time()
                unit = end - start
                total_used += unit
                runTimes.append(unit)
                total_remaining = sum(runTimes) / (version_index + 1 - failed_count) * (len(versions_sort) - version_index - 1)
                print("整理该版本数据所花费的时间（Time spent in sorting out this version）：", format_runtime(unit))
                print("已花费的总时间（Total time used）                                   ：", format_runtime(total_used))
                print("剩余时间（Time remaining）                                          ：", format_runtime(total_remaining))
                print("预计总时间（Expected total time）                                   ：", format_runtime(total_used + total_remaining), end = "\n\n")
    if item_df_formed:
        print("是否导出以上装备数据至Excel中？（输入任意键导出，否则不导出。）\nDo you want to export the above data into Excel? (Submit any non-empty string to export, or null to refuse exporting.)")
        export = bool(input())
        if export:
            versions_sort = list(LoLItem_dfs.keys())
            print("正在保存中……\nSaving the data ...")
            excel_name = "英雄联盟装备信息.xlsx"
            excel_name_sorted = "英雄联盟装备信息(sorted).xlsx"
            while True:
                try:
                    with pandas.ExcelWriter(path = excel_name, mode = "a", if_sheet_exists = "replace") as writer:
                        runTimes = [] #记录保存一个版本的装备数据所花费的时间（Records the time spent in saving item data of one version）
                        total_used = 0
                        for i in range(len(versions_sort)):
                            start = time.time()
                            version = versions_sort[i]
                            print("装备信息导出进度（Item data export process）：%d/%d\t版本（Version）：%s" %(i + 1, len(versions_sort), version))
                            if version == "latest" or version == "pbe":
                                LoLItem_dfs[version].to_excel(excel_writer = writer, sheet_name = version)
                            elif source != "" and source[0] == "1":
                                LoLItem_dfs[version].to_excel(excel_writer = writer, sheet_name = version + " (ddragon)")
                            else:
                                LoLItem_dfs[version].to_excel(excel_writer = writer, sheet_name = version + " (cdragon)")
                            end = time.time()
                            unit = end - start
                            total_used += unit
                            runTimes.append(unit)
                            total_remaining = sum(runTimes) / (i + 1) * (len(versions_sort) - i - 1)
                            print("保存该版本数据所花费的时间（Time spent in saving this version）：", format_runtime(unit))
                            print("已花费的总时间（Total time used）                              ：", format_runtime(total_used))
                            print("剩余时间（Time remaining）                                     ：", format_runtime(total_remaining))
                            print("预计总时间（Expected total time）                              ：", format_runtime(total_used + total_remaining), end = "\n\n")
                    workbook_exist = True
                except FileNotFoundError:
                    with pandas.ExcelWriter(path = excel_name) as writer:
                        runTimes = [] #记录保存一个版本的装备数据所花费的时间（Records the time spent in saving item data of one version）
                        total_used = 0
                        for i in range(len(versions_sort)):
                            start = time.time()
                            version = versions_sort[i]
                            print("装备信息导出进度（Item data export process）：%d/%d\t版本（Version）：%s" %(i + 1, len(versions_sort), version))
                            if version == "latest" or version == "pbe":
                                LoLItem_dfs[version].to_excel(excel_writer = writer, sheet_name = version)
                            elif source != "" and source[0] == "1":
                                LoLItem_dfs[version].to_excel(excel_writer = writer, sheet_name = version + " (ddragon)")
                            else:
                                LoLItem_dfs[version].to_excel(excel_writer = writer, sheet_name = version + " (cdragon)")
                            end = time.time()
                            unit = end - start
                            total_used += unit
                            runTimes.append(unit)
                            total_remaining = sum(runTimes) / (i + 1) * (len(versions_sort) - i - 1)
                            print("保存该版本数据所花费的时间（Time spent in saving this match）：", format_runtime(unit))
                            print("已花费的总时间（Total time used）                            ：", format_runtime(total_used))
                            print("剩余时间（Time remaining）                                   ：", format_runtime(total_remaining))
                            print("预计总时间（Expected total time）                            ：", format_runtime(total_used + total_remaining), end = "\n\n")
                    break
                except PermissionError:
                    print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                    input()
                else:
                    break
                
if workbook_exist:
    print("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致版本号顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of version order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
    sort = input()
    if sort != "": #所有工作表按顺序依次分为固定版本号、cdragon版本号和ddragon版本号（All sheets are divided into the fixed version class, cdragon version class and ddragon version class）
        items_loaded = True
        print("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
        while True:
            try:
                wb = load_workbook(excel_name)
            except FileNotFoundError:
                print('装备工作簿读取失败！请确保当前文件夹内含有名为“%s”的工作簿。如果需要重新生成该召唤师的工作簿，请输入“0”。\nERROR reading the summoner profile workbook! Please make sure the workbook "%s" is in the current folder". If you want to regenerate this summoner\'s workbook, please submit "0".' %(excel_name, excel_name))
                items_reload = input()
                if items_reload == "0":
                    items_loaded = False
                    break
            else:
                break
        if items_loaded:
            sheetnames = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
            #下面锁定工作表顺序（The following code determine the sheet order）
            print("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
            fixed_version_list = ["pbe", "latest"]
            cdragon_version_list = []
            ddragon_version_list = []
            for version in sheetnames:
                if version.endswith("(cdragon)"):
                    cdragon_version_list.append(version)
                elif version.endswith("(ddragon)"):
                    ddragon_version_list.append(version)
            sheetnames_sorted = []
            for sheet_iter in fixed_version_list:
                if sheet_iter in sheetnames:
                    sheetnames.remove(sheet_iter)
                    sheetnames_sorted.append(sheet_iter)
            cdragon_version_sorted = list(map(lambda x: x + " (cdragon)", patch_sort(cdragon_version_list)))
            cdragon_version_sorted.reverse()
            ddragon_version_sorted = list(map(lambda x: x + " (ddragon)", patch_sort(ddragon_version_list)))
            ddragon_version_sorted.reverse()
            sheetnames_sorted += cdragon_version_sorted + ddragon_version_sorted #所有工作表的期望顺序存储在sheetnames_sorted变量中（The ordered result of all sheets is stored in the variable `sheetnames_sorted`）
            #下面排列所有工作表（The following code arrange all sheets）
            print("正在排序……\nOrdering ...")
            for i in range(len(sheetnames_sorted)): #排序的思路是每次将一个工作表根据其在原工作表列表中的索引和在顺序工作表列表中的索引的差值进行移动（The main idea of sheets' sorting is to move each sheet according to the difference of the indices between in the original sheet list and in the ordered sheet list）
                sheetnames = wb.sheetnames #因为一次移动可能导致很多其它工作表的位置发生变化，所以必须每次都重新获取工作表列表（Because a moving event may result in location change of many other sheets, the sheet list must be obtained each time）
                sheetname_iter = sheetnames_sorted[i] #这里以顺序工作表为迭代器进行遍历，因为顺序工作表是固定不变的（Here the ordered sheet list acts as the iterator to be traversed, for the ordered sheet list is fixed）
                if sheetnames[i] != sheetname_iter:
                    preIndex = sheetnames.index(sheetname_iter)
                    wb.move_sheet(sheetname_iter, i - preIndex) #注意移动距离数应当是排序后的索引减去排序前的索引（Note that the moving offset should be the index in the ordered list subtracted by that in the original list）
                #print("排序进度（Ordering process）：%d/%d\t工作表名称（Sheet name）： %s" %(i + 1, len(sheetnames_sorted), sheetname_iter))
            print('正在保存中……\nSaving the ordered workbook ...')
            wb.save(excel_name_sorted)
            print('排序完成！排好序的工作簿已保存为“%s”。\nOrdering finished! The ordered workbook is saved as "%s".\n' %(excel_name_sorted, excel_name_sorted))
