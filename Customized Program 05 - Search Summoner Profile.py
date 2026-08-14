from lcu_driver import Connector
from lcu_driver.connection import Connection
import argparse, datetime, gc, os, pandas, requests, time, json
from urllib.parse import urljoin
from openpyxl import load_workbook, Workbook
from openpyxl.workbook.workbook import Worksheet
import matplotlib.pyplot as plt
from typing import Any, Optional
from src.utils.summoner import print_summoner_info, get_info, get_infos, get_info_name
from src.utils.logger import LogManager
from src.utils.format import getISOTime, optimize_bool_display, format_df, eliminate_empty_fields, addDefaultStyle, format_runtime, verify_uuid
from src.utils.patch import Patch
from src.utils.webRequest import requestUrl, SGPSession
from src.utils.excel_workbook import create_workbook_win32, sort_worksheet
from src.core.config.headers import profile_header, mastery_header, ranked_header, ladder_header
from src.core.config.headers import TFTGame_summary_header as TFTGame_stat_header
from src.core.config.localization import language_ddragon, language_dict, language_cdragon, tiers, tiers_all, ratedTiers, challengeCategories, challengeCrystalLevels, titleAcquisitionTypes, queueTypes_ranked
from src.core.config.servers import valid_platformIds, set_platform_folder, set_summonerInfo_folder, save_platform_info
from src.core.config.conditional_formatting import addFormat_LoLHistory_wb, addFormat_LoLGame_summary_wb, addFormat_LoLGame_summary_wb_transpose
from src.core.dataframes.ranked import sort_game_leaderboard
from src.core.dataframes.matchHistory import get_LoLHistory, get_matchSummary_sgp, get_matchDetails_sgp, sort_LoLHistory, sort_LoLHistory_sgp, reconstruct_LoLHistory, reconstruct_LoLHistory_sgp, reconstruct_TFTHistory, get_LoLGame_summary, get_game_summary_sgp, get_LoLGame_timeline, get_game_timeline_sgp, sort_LoLGame_summary, sort_LoLGame_summary_sgp, sort_LoLGame_timeline, sort_LoLGame_timeline_sgp, get_TFTHistory, sort_TFTHistory, sort_TFTGame_summary

parser = argparse.ArgumentParser()
parser.add_argument("-a", "--lol-api", help = "指定通过什么接口获取英雄联盟对局概要和时间轴（Specify the interface used to fetch LoL game summary and timeline）", action = "store", type = str, choices = ["lcu", "sgp"], default = "sgp")
# parser.add_argument("-l", "--lol-api-legacy", help = "指定是否使用传统LCU API接口获取英雄联盟对局概要和时间轴（Specify whether to use the traditional LCU API to fetch LoL game summary and timeline）", action = "store_true") #这个变量和上面的作用相同（This argument works the same way as the above one）
parser.add_argument("-ic", "--info-color", help = "为对局概要工作表施加条件格式（Add conditional formatting to match summary sheets）", action = "store_true") #这会对性能和工作簿大小有较大影响（This seriously affects the program's performance and the workbook size）
parser.add_argument("-lb", "--export-leaderboard", help = "导出每场对局的社交排行榜工作表。时间开销大（Export the social leaderboard data of each match. Time consuming）", action = "store_true")
parser.add_argument("-n", "--deny-empty-sheet-creation", help = "在创建工作簿的情况下不创建空白工作表（Deny creating empty sheets if a new workbook is created）", action = "store_true") #主要应用于从小工作簿移动工作表到大工作簿的情形（Mainly used in the case where sheets are moved from a small workbook into a bigger workbook）
parser.add_argument("-r", "--reserve", help = "在对局不包含主玩家的情况下仍然加载该对局（Load a match even if it doesn't contain the main player）", action = "store_true")
parser.add_argument("-rt", "--reserve-text", help = "在对局不包含主玩家的情况下仍然保存该对局（Save a match even if it doesn't contain the main player）", action = "store_true")
args = parser.parse_args()
use_sgp: bool = args.lol_api == "sgp"
if use_sgp:
    from src.core.config.headers import LoLGame_summary_sgp_header as LoLGame_stat_header
else:
    from src.core.config.headers import LoLGame_summary_header as LoLGame_stat_header

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN, Awesome丶ABC
# 更新（Last update）：     2026/08/12
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

session: requests.Session = requests.Session()
sgpSession: SGPSession = SGPSession()
URLPatch: str = ""
patches_initial: list[str] = []
bigPatches: list[Patch] = []
queues_initial: dict[int, dict[str, Any]] = {}
spells_initial: dict[int, dict[str, Any]] = {}
LoLChampions_initial: dict[int, dict[str, Any]] = {}
LoLItems_initial: dict[int, dict[str, Any]] = {}
summonerIcons_initial: dict[int, dict[str, Any]] = {}
perks_initial: dict[int, dict[str, Any]] = {}
perkstyles_initial: dict[int, dict[str, Any]] = {}
TFTAugments_initial: dict[str, dict[str, Any]] = {}
TFTChampions_initial: dict[str, dict[str, Any]] = {}
TFTItems_initial: dict[str, dict[str, Any]] = {}
TFTCompanions_initial: dict[str, dict[str, Any]] = {}
TFTTraits_initial: dict[str, dict[str, Any]] = {}
CherryAugments_initial: dict[int, dict[str, Any]] = {}
log: LogManager = LogManager()
platformId: str = ""

error_header: dict[str, str] = {"errorCode": "异常代码", "httpStatus": "HTTP状态码", "implementationDetails": "细节", "message": "消息"}
error_header_keys: list[str] = list(error_header.keys())
connector: Connector = Connector()

#-----------------------------------------------------------------------------
# 搜索召唤师生涯（Search summoner profile）
#-----------------------------------------------------------------------------
def prepare_data_resources(platformId: str, locale: str) -> tuple[bool, bool]:
    '''
    准备全局数据资源。<br>Prepare global data resources.
    
    :param platformId: 服务器代号。决定使用正式服还是测试服的数据资源。<br>PlatformId, which determines one of Live and PBE data resources will be used.
    
        服务器代号可以通过以下LCU接口获取：<br>PlatformId can be obtained through the following LCU endpoint:
        - `GET /lol-lobby/v1/parties/player`
    :type platformId: str
    :param locale: 语言文化代码。决定了数据资源的语言。<br>Language code, which determines the language of the data resources.
    :type locale
    :return: 是否切换语言，以及是否退出程序。<br>Whether to switch language and whether to exit the program.
    :rtype: tuple[bool, bool]
    '''
    global session, URLPatch, patches_initial, bigPatches, queues_initial, spells_initial, LoLChampions_initial, LoLItems_initial, summonerIcons_initial, perks_initial, perkstyles_initial, TFTAugments_initial, TFTChampions_initial, TFTItems_initial, TFTCompanions_initial, TFTTraits_initial, CherryAugments_initial
    #下面声明一些数据资源的地址（The following code declare some data resources' URLs）
    URLPatch = "pbe" if platformId == "PBE1" or platformId == "PBE" else "latest"
    patches_url: str = "https://ddragon.leagueoflegends.com/api/versions.json"
    queue_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/queues.json" %(URLPatch, language_cdragon[locale]) #CommunityDragon数据库只存储第7赛季及以后的数据（CommunityDragon database only stores data including and after Season 7）
    spell_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-spells.json" %(URLPatch, language_cdragon[locale])
    LoLChampion_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/champion-summary.json" %(URLPatch, language_cdragon[locale])
    LoLItem_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/items.json" %(URLPatch, language_cdragon[locale])
    summonerIcon_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/summoner-icons.json" %(URLPatch, language_cdragon[locale])
    perk_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perks.json" %(URLPatch, language_cdragon[locale])
    perkstyle_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/perkstyles.json" %(URLPatch, language_cdragon[locale])
    TFTBasic_url: str = "https://raw.communitydragon.org/%s/cdragon/tft/%s.json" %(URLPatch, locale.lower())
    TFTChampion_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tftchampions.json" %(URLPatch, language_cdragon[locale])
    TFTItem_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tftitems.json" %(URLPatch, language_cdragon[locale])
    TFTCompanion_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/companions.json" %(URLPatch, language_cdragon[locale])
    TFTTrait_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/tfttraits.json" %(URLPatch, language_cdragon[locale])
    CherryAugment_url: str = "https://raw.communitydragon.org/%s/plugins/rcp-be-lol-game-data/global/%s/v1/cherry-augments.json" %(URLPatch, language_cdragon[locale])
    #下面声明离线数据资源的默认地址（The following code declare the default paths of offline data resources）
    patches_local_default: str = "离线数据（Offline Data）\\versions.json"
    queue_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\queues.json" %(URLPatch, language_cdragon[locale])
    spell_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\summoner-spells.json" %(URLPatch, language_cdragon[locale])
    LoLChampion_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\champion-summary.json" %(URLPatch, language_cdragon[locale])
    LoLItem_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\items.json" %(URLPatch, language_cdragon[locale])
    summonerIcon_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\summoner-icons.json" %(URLPatch, language_cdragon[locale])
    perk_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\perks.json" %(URLPatch, language_cdragon[locale])
    perkstyle_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\perkstyles.json" %(URLPatch, language_cdragon[locale])
    TFTBasic_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\cdragon\\tft\\%s.json" %(URLPatch, locale.lower())
    TFTChampion_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\tftchampions.json" %(URLPatch, language_cdragon[locale])
    TFTItem_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\tftitems.json" %(URLPatch, language_cdragon[locale])
    TFTCompanion_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\companions.json" %(URLPatch, language_cdragon[locale])
    TFTTrait_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\tfttraits.json" %(URLPatch, language_cdragon[locale])
    CherryAugment_local_default: str = "离线数据（Offline Data）\\cdragon\\%s\\plugins\\rcp-be-lol-game-data\\global\\%s\\v1\\cherry-augments.json" %(URLPatch, locale.lower())
    #加载数据（Load data）
    logPrint("请选择数据资源获取模式：\nPlease select the data resource capture mode:\n1\t在线模式（Online）\n2\t离线模式（Offline）")
    prepareMode: str = logInput()
    switch_language: bool = False
    while True:
        if prepareMode != "" and prepareMode[0] == "1":
            switch_prepare_mode: bool = False
            #下面获取版本信息（The following code get the patch data）
            try:
                source, status, session = requestUrl("GET", patches_url, session = session, log = log)
                patches_initial = source.json()
            except requests.exceptions.RequestException:
                logPrint('版本信息获取超时！正在尝试离线加载数据……\nPatch information capture timeout! Trying loading offline data ...\n请输入版本Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the patch Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(patches_local_default, patches_local_default))
                while True:
                    patches_local: str = logInput()
                    if patches_local == "":
                        patches_local = patches_local_default
                    elif patches_local[0] == "0":
                        logPrint("版本信息获取失败！请检查系统网络状况和代理设置。\nPatch information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(patches_local, "r", encoding = "utf-8") as fp:
                            patches_initial = json.load(fp)
                        if isinstance(patches_initial, list) and patches_initial[-1] == "lolpatch_3.7":
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合DataDragon数据库中记录的版本数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the patch data archived in DataDragon database (%s)!" %(patches_url, patches_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的版本Json数据文件路径！\nFile "%s" NOT found! Please input a correct patch Json data file path!' %(patches_local, patches_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有版本信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with patch information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合DataDragon数据库中记录的版本数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the patch data archived in DataDragon database (%s)!" %(patches_url, patches_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            latest_patch: str = patches_initial[0]
            patches_dict: dict[str, list[str]] = {}
            smallPatches: list[str] = []
            bigPatches = []
            for patch in patches_initial:
                if not patch.startswith("lolpatch"):
                    patch_split: str = patch.split(".")
                    smallPatch: str = ".".join(patch_split[:3])
                    smallPatches.append(Patch(smallPatch))
                    bigPatch: str = ".".join(patch_split[:2])
                    bigPatches.append(Patch(bigPatch))
                    patches_dict[bigPatch] = []
            for i in range(len(bigPatches)):
                patches_dict[str(bigPatches[i])].append(str(smallPatches[i]))
            #下面获取游戏模式数据（The following code get game mode data）
            try:
                logPrint("正在加载游戏模式信息……\nLoading game mode information from CommunityDragon...")
                source, status, session = requestUrl("GET", queue_url, session = session, log = log)
                if source.ok:
                    queue_initial: list[dict[str, Any]] = source.json() #queue存储游戏模式信息（Variable `queue_initial` stores game mode information）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('游戏模式信息获取超时！正在尝试离线加载数据……\nQueue information capture timeout! Trying loading offline data ...\n请输入游戏模式Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the game mode Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(queue_local_default, queue_local_default))
                while True:
                    queue_local: str = logInput()
                    if queue_local == "":
                        queue_local = queue_local_default
                    elif queue_local[0] == "0":
                        logPrint("游戏模式信息获取失败！请检查系统网络状况和代理设置。\nQueue information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(queue_local, "r", encoding = "utf-8") as fp:
                            queue_initial = json.load(fp)
                        if isinstance(queue_initial, list) and all(map(lambda x: all(i in x for i in ["id", "name", "shortName", "description", "detailedDescription", "gameSelectModeGroup", "gameSelectCategory", "gameSelectPriority", "isSkillTreeQueue", "isLimitedTimeQueue", "isBotHonoringAllowed", "hidePlayerPosition", "viableChampionRoster"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], int) for i in ["id", "gameSelectPriority"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], str) for i in ["name", "shortName", "description", "detailedDescription", "gameSelectModeGroup", "gameSelectCategory"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], bool) for i in ["isSkillTreeQueue", "isLimitedTimeQueue", "isBotHonoringAllowed", "hidePlayerPosition"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], list) or x[i] is None for i in ["viableChampionRoster"]), queue_initial)):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的游戏模式数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the game mode data archived in CommunityDragon database (%s)!" %(queue_url, queue_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的游戏模式Json数据文件路径！\nFile "%s" NOT found! Please input a correct game mode Json data file path!' %(queue_local, queue_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有游戏模式信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with game mode information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的游戏模式数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the game mode data archived in CommunityDragon database (%s)!" %(queue_url, queue_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取召唤师技能数据（The following code get summoner spell data）
            try:
                logPrint("正在加载召唤师技能信息……\nLoading summoner spell information from CommunityDragon...")
                source, status, session = requestUrl("GET", spell_url, session = session, log = log)
                if source.ok:
                    spell_initial: list[dict[str, Any]] = source.json() #spell存储召唤师技能信息（Variable `spell_initial` stores summoner spell information）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('召唤师技能信息获取超时！正在尝试离线加载数据……\nSummoner spell information capture timeout! Trying loading offline data ...\n请输入召唤师技能Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the summoner spell Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(spell_local_default, spell_local_default))
                while True:
                    spell_local: str = logInput()
                    if spell_local == "":
                        spell_local = spell_local_default
                    elif spell_local[0] == "0":
                        logPrint("召唤师技能信息获取失败！请检查系统网络状况和代理设置。\nSummoner spell information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(spell_local, "r", encoding = "utf-8") as fp:
                            spell_initial = json.load(fp)
                        if isinstance(spell_initial, list) and all(i in spell_initial[j] for i in ["id", "name", "description", "summonerLevel", "cooldown", "gameModes", "iconPath"] for j in range(len(spell_initial))):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师技能数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner spell data archived in CommunityDragon database (%s)!" %(spell_url, spell_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的召唤师技能Json数据文件路径！\nFile "%s" NOT found! Please input a correct summoner spell Json data file path!' %(spell_local, spell_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有召唤师技能信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with summoner spell information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师技能数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner spell data archived in CommunityDragon database (%s)!" %(spell_url, spell_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取英雄信息（The following code get LoL champion data）
            try:
                logPrint("正在加载英雄信息……\nLoading LoL champion information from CommunityDragon...")
                source, status, session = requestUrl("GET", LoLChampion_url, session = session, log = log)
                if source.ok:
                    LoLChampion_initial: list[dict[str, Any]] = source.json() #LoLItem存储英雄信息。（Variable `LoLChampion_initial` stores information of LoL champions）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('英雄信息获取超时！正在尝试离线加载数据……\nLoL champion information capture timeout! Trying loading offline data ...\n请输入英雄Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the LoL champion Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(LoLChampion_local_default, LoLChampion_local_default))
                while True:
                    LoLChampion_local: str = logInput()
                    if LoLChampion_local == "":
                        LoLChampion_local = LoLChampion_local_default
                    elif LoLChampion_local[0] == "0":
                        logPrint("英雄信息获取失败！请检查系统网络状况和代理设置。\nLoL champion information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(LoLChampion_local, "r", encoding = "utf-8") as fp:
                            LoLChampion_initial = json.load(fp)
                        if isinstance(LoLChampion_initial, list) and all(isinstance(LoLChampion_initial[i], dict) for i in range(len(LoLChampion_initial))) and all(j in LoLChampion_initial[i] for i in range(len(LoLChampion_initial)) for j in ["id", "name", "alias", "squarePortraitPath", "roles"]) and all(isinstance(LoLChampion_initial[i]["id"], int) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["name"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["alias"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["squarePortraitPath"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["roles"], list) for i in range(len(LoLChampion_initial))):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL champion data archived in CommunityDragon database (%s)!" %(LoLChampion_url, LoLChampion_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的英雄Json数据文件路径！\nFile "%s" NOT found! Please input a correct LoL champion Json data file path!' %(LoLChampion_local, LoLChampion_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with LoL champion information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL champion data archived in CommunityDragon database (%s)!" %(LoLChampion_url, LoLChampion_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取英雄联盟装备信息（The following code get LoL item data）
            try:
                logPrint("正在加载英雄联盟装备信息……\nLoading LoL item information from CommunityDragon...")
                source, status, session = requestUrl("GET", LoLItem_url, session = session, log = log)
                if source.ok:
                    LoLItem_initial: list[dict[str, Any]] = source.json() #LoLItem存储经典模式的装备信息。（Variable `LoLItem_initial` stores information of LoL items）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('英雄联盟装备信息获取超时！正在尝试离线加载数据……\nLoL item information capture timeout! Trying loading offline data ...\n请输入英雄联盟装备Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the LoL item Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(LoLItem_local_default, LoLItem_local_default))
                while True:
                    LoLItem_local: str = logInput()
                    if LoLItem_local == "":
                        LoLItem_local = LoLItem_local_default
                    elif LoLItem_local[0] == "0":
                        logPrint("英雄联盟装备信息获取失败！请检查系统网络状况和代理设置。\nLoL item information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(LoLItem_local, "r", encoding = "utf-8") as fp:
                            LoLItem_initial = json.load(fp)
                        if isinstance(LoLItem_initial, list) and all(i in LoLItem_initial[j] for i in ["id", "name", "description", "active", "inStore", "from", "to", "categories", "maxStacks", "requiredChampion", "requiredAlly", "requiredBuffCurrencyName", "requiredBuffCurrencyCost", "specialRecipe", "isEnchantment", "price", "priceTotal", "iconPath"] for j in range(len(LoLItem_initial))):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄联盟装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL item data archived in CommunityDragon database (%s)!" %(LoLItem_url, LoLItem_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的英雄联盟装备Json数据文件路径！\nFile "%s" NOT found! Please input a correct LoL item Json data file path!' %(LoLItem_local, LoLItem_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有英雄联盟装备信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with LoL item information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的英雄联盟装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the LoL item data archived in CommunityDragon database (%s)!" %(LoLItem_url, LoLItem_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取召唤师图标信息（The following code get summoner icon data）
            try:
                logPrint("正在加载召唤师图标信息……\nLoading summoner icon information from CommunityDragon...")
                source, status, session = requestUrl("GET", summonerIcon_url, session = session, log = log)
                if source.ok:
                    summonerIcon_initial: list[dict[str, Any]] = source.json() #LoLItem存储召唤师图标信息。（Variable `summonerIcon_initial` stores information of summoner icons）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('召唤师图标信息获取超时！正在尝试离线加载数据……\nSummoner icon information capture timeout! Trying loading offline data ...\n请输入召唤师图标Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the summoner icon Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(summonerIcon_local_default, summonerIcon_local_default))
                while True:
                    summonerIcon_local: str = logInput()
                    if summonerIcon_local == "":
                        summonerIcon_local = summonerIcon_local_default
                    elif summonerIcon_local[0] == "0":
                        logPrint("召唤师图标信息获取失败！请检查系统网络状况和代理设置。\nSummoner icon information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(summonerIcon_local, "r", encoding = "utf-8") as fp:
                            summonerIcon_initial = json.load(fp)
                        if isinstance(summonerIcon_initial, list) and all(map(lambda x: isinstance(x, dict), summonerIcon_initial)) and all(i in j for i in ["id", "title", "yearReleased", "isLegacy", "descriptions", "rarities", "disabledRegions"] for j in summonerIcon_initial):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师图标数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner icon data archived in CommunityDragon database (%s)!" %(summonerIcon_url, summonerIcon_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的召唤师图标Json数据文件路径！\nFile "%s" NOT found! Please input a correct summoner icon Json data file path!' %(summonerIcon_local, summonerIcon_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有召唤师图标信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with summoner icon information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的召唤师图标数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the summoner icon data archived in CommunityDragon database (%s)!" %(summonerIcon_url, summonerIcon_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取基石符文信息（The following code get perk data）
            try:
                logPrint("正在加载基石符文信息……\nLoading perk information from CommunityDragon...")
                source, status, session = requestUrl("GET", perk_url, session = session, log = log)
                if source.ok:
                    perk_initial: list[dict[str, Any]] = source.json() #perk存储基石符文信息。（Variable `perk_initial` stores information of perks）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('基石符文信息获取超时！正在尝试离线加载数据……\nPerk information capture timeout! Trying loading offline data ...\n请输入基石符文Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the perk Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(perk_local_default, perk_local_default))
                while True:
                    perk_local: str = logInput()
                    if perk_local == "":
                        perk_local = perk_local_default
                    elif perk_local[0] == "0":
                        logPrint("基石符文信息获取失败！请检查系统网络状况和代理设置。\nPerk information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(perk_local, "r", encoding = "utf-8") as fp:
                            perk_initial = json.load(fp)
                        if isinstance(perk_initial, list) and all(i in perk_initial[j] for i in ["id", "name", "majorChangePatchVersion", "tooltip", "shortDesc", "longDesc", "recommendationDescriptor", "iconPath", "endOfGameStatDescs", "recommendationDescriptorAttributes"] for j in range(len(perk_initial))):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的基石符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perk data archived in CommunityDragon database (%s)!" %(perk_url, perk_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的基石符文Json数据文件路径！\nFile "%s" NOT found! Please input a correct perk Json data file path!' %(perk_local, perk_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有基石符文信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with perk information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的基石符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perk data archived in CommunityDragon database (%s)!" %(perk_url, perk_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取符文系信息（The following code get perkstyle data）
            try:
                logPrint("正在加载符文系信息……\nLoading perkstyle information from CommunityDragon...")
                source, status, session = requestUrl("GET", perkstyle_url, session = session, log = log)
                if source.ok:
                    perkstyle_initial: dict[str, Any] = source.json() #perkstyle存储符文系信息。（Variable `perkstyle_initial` stores information of perkstyles）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('符文系信息获取超时！正在尝试离线加载数据……\nPerkstyle information capture timeout! Trying loading offline data ...\n请输入符文系Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the perkstyle Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(perkstyle_local_default, perkstyle_local_default))
                while True:
                    perkstyle_local: str = logInput()
                    if perkstyle_local == "":
                        perkstyle_local = perkstyle_local_default
                    elif perkstyle_local[0] == "0":
                        logPrint("符文系信息获取失败！请检查系统网络状况和代理设置。\nperkstyle information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(perkstyle_local, "r", encoding = "utf-8") as fp:
                            perkstyle_initial = json.load(fp)
                        if isinstance(perkstyle_initial, dict) and all(perkstyle_initial.get(i, 0) for i in ["schemaVersion", "styles"]) and isinstance(perkstyle_initial["styles"], list) and all(j in perkstyle_initial["styles"][i] for i in range(len(perkstyle_initial["styles"])) for j in ["id", "name", "tooltip", "iconPath", "assetMap", "isAdvanced", "allowedSubStyles", "subStyleBonus", "slots", "defaultPageName", "defaultSubStyle", "defaultPerks", "defaultPerksWhenSplashed", "defaultStatModsPerSubStyle"]):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的符文系数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perkstyle data archived in CommunityDragon database (%s)!" %(perkstyle_url, perkstyle_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的符文系Json数据文件路径！\nFile "%s" NOT found! Please input a correct perkstyle Json data file path!' %(perkstyle_local, perkstyle_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有符文系信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with perkstyle information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的符文系数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the perkstyle data archived in CommunityDragon database (%s)!" %(perkstyle_url, perkstyle_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取云顶之弈强化符文数据（The following code get TFT augment data）
            try:
                logPrint("正在加载云顶之弈基础数据……\nLoading TFT basic data from CommunityDragon ...")
                source, status, session = requestUrl("GET", TFTBasic_url, session = session, log = log)
                if source.ok:
                    TFTBasic_initial: dict[str, Any] = source.json() #TFT存储云顶之弈中至今为止所有的强化符文、英雄和羁绊信息和各赛季的英雄和羁绊信息（Variable `TFTBasic_initial` stores information of all augments, champions and traits so far and information of champions and traits with respect to season）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('云顶之弈基础信息获取超时！正在尝试离线加载数据……\nTFT basic information capture timeout! Trying loading offline data ...\n请输入云顶之弈基础数据Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT basics Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTBasic_local_default, TFTBasic_local_default))
                while True:
                    TFTBasic_local: str = logInput()
                    if TFTBasic_local == "":
                        TFTBasic_local = TFTBasic_local_default
                    elif TFTBasic_local[0] == "0":
                        logPrint("云顶之弈基础信息获取失败！请检查系统网络状况和代理设置。\nTFT basic information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(TFTBasic_local, "r", encoding = "utf-8") as fp:
                            TFTBasic_initial = json.load(fp)
                        if isinstance(TFTBasic_initial, dict) and all(i in TFTBasic_initial for i in ["items", "setData", "sets"]) and all(isinstance(TFTBasic_initial[i], list) for i in ["items", "setData"]) and all(isinstance(TFTBasic_initial[i], dict) for i in ["sets"]) and all(j in TFTBasic_initial["items"][i] for i in range(len(TFTBasic_initial["items"])) for j in ["apiName", "associatedTraits", "composition", "desc", "effects", "from", "icon", "id", "incompatibleTraits", "name", "tags", "unique"]) and all(isinstance(TFTBasic_initial["items"][i][j], str) or TFTBasic_initial["items"][i][j] == None for i in range(len(TFTBasic_initial["items"])) for j in ["apiName", "desc", "icon", "name"]) and all(isinstance(TFTBasic_initial["items"][i][j], list) for i in range(len(TFTBasic_initial["items"])) for j in ["associatedTraits", "composition", "tags"]) and all(isinstance(TFTBasic_initial["items"][i][j], dict) for i in range(len(TFTBasic_initial["items"])) for j in ["effects"]) and all(isinstance(TFTBasic_initial["items"][i][j], bool) for i in range(len(TFTBasic_initial["items"])) for j in ["unique"]) and all(j in TFTBasic_initial["setData"][i] for i in range(len(TFTBasic_initial["setData"])) for j in ["augments", "champions", "items", "mutator", "name", "number", "traits"]) and all(isinstance(TFTBasic_initial["setData"][i][j], list) for i in range(len(TFTBasic_initial["setData"])) for j in ["augments", "champions", "items", "traits"]) and all(isinstance(TFTBasic_initial["setData"][i][j], str) for i in range(len(TFTBasic_initial["setData"])) for j in ["mutator", "name"]) and all(isinstance(TFTBasic_initial["setData"][i][j], int) for i in range(len(TFTBasic_initial["setData"])) for j in ["number"]) and all(map(lambda x: isinstance(x, str), TFTBasic_initial["setData"][i][j]) for i in range(len(TFTBasic_initial["setData"])) for j in ["augments", "items"]) and all(map(lambda x: isinstance(x, dict), TFTBasic_initial["setData"][i]["champions"]) for i in range(len(TFTBasic_initial["setData"]))) and all(k in TFTBasic_initial["setData"][i]["champions"][j] for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in ["ability", "apiName", "characterName", "cost", "icon", "name", "role", "squareIcon", "stats", "tileIcon", "traits"]) and all(isinstance(TFTBasic_initial["setData"][i]["champions"][j][k], dict) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in ["ability", "stats"]) and all(isinstance(TFTBasic_initial["setData"][i]["champions"][j][k], str) or TFTBasic_initial["setData"][i]["champions"][j][k] == None for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in ["apiName", "characterName", "icon", "name", "squareIcon", "tileIcon"]) and all(isinstance(TFTBasic_initial["setData"][i]["champions"][j][k], int) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in ["cost"]) and all(k in TFTBasic_initial["setData"][i]["champions"][j]["ability"] for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in ["desc", "icon", "name", "variables"]) and all(isinstance(TFTBasic_initial["setData"][i]["champions"][j]["ability"][k], str) or TFTBasic_initial["setData"][i]["champions"][j]["ability"][k] == None for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in ["desc", "icon", "name"]) and all(isinstance(TFTBasic_initial["setData"][i]["champions"][j]["ability"][k], list) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in ["variables"]) and all(map(lambda x: isinstance(x, dict), TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"]) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"]))) and all(l in TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"][k] for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in range(len(TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"])) for l in ["name", "value"]) and all(isinstance(TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"][k][l], str) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in range(len(TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"])) for l in ["name"]) and all(isinstance(TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"][k][l], list) or TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"][k][l] == None for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["champions"])) for k in range(len(TFTBasic_initial["setData"][i]["champions"][j]["ability"]["variables"])) for l in ["value"]) and all(k in TFTBasic_initial["setData"][i]["traits"][j] for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["traits"])) for k in ["apiName", "desc", "effects", "icon", "name"]) and all(isinstance(TFTBasic_initial["setData"][i]["traits"][j][k], str) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["traits"])) for k in ["apiName", "desc", "icon", "name"]) and all(isinstance(TFTBasic_initial["setData"][i]["traits"][j][k], list) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["traits"])) for k in ["effects"]) and all(map(lambda x: isinstance(x, dict), TFTBasic_initial["setData"][i]["traits"][j]["effects"]) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["traits"]))) and all(l in TFTBasic_initial["setData"][i]["traits"][j]["effects"][k] for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["traits"])) for k in range(len(TFTBasic_initial["setData"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style", "variables"]) and all(isinstance(TFTBasic_initial["setData"][i]["traits"][j]["effects"][k][l], int) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["traits"])) for k in range(len(TFTBasic_initial["setData"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style"]) and all(isinstance(TFTBasic_initial["setData"][i]["traits"][j]["effects"][k][l], dict) for i in range(len(TFTBasic_initial["setData"])) for j in range(len(TFTBasic_initial["setData"][i]["traits"])) for k in range(len(TFTBasic_initial["setData"][i]["traits"][j]["effects"])) for l in ["variables"]) and all(j in TFTBasic_initial["sets"][i] for i in TFTBasic_initial["sets"] for j in ["champions", "name", "traits"]) and all(isinstance(TFTBasic_initial["sets"][i][j], list) for i in TFTBasic_initial["sets"] for j in ["champions", "traits"]) and all(isinstance(TFTBasic_initial["sets"][i][j], str) for i in TFTBasic_initial["sets"] for j in ["name"]) and all(k in TFTBasic_initial["sets"][i]["champions"][j] for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in ["ability", "apiName", "characterName", "cost", "icon", "name", "role", "squareIcon", "stats", "tileIcon", "traits"]) and all(isinstance(TFTBasic_initial["sets"][i]["champions"][j][k], dict) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in ["ability", "stats"]) and all(isinstance(TFTBasic_initial["sets"][i]["champions"][j][k], str) or TFTBasic_initial["sets"][i]["champions"][j][k] == None for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in ["apiName", "characterName", "icon", "name", "squareIcon", "tileIcon"]) and all(isinstance(TFTBasic_initial["sets"][i]["champions"][j][k], int) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in ["cost"]) and all(k in TFTBasic_initial["sets"][i]["champions"][j]["ability"] for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in ["desc", "icon", "name", "variables"]) and all(isinstance(TFTBasic_initial["sets"][i]["champions"][j]["ability"][k], str) or TFTBasic_initial["sets"][i]["champions"][j]["ability"][k] == None for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in ["desc", "icon", "name"]) and all(isinstance(TFTBasic_initial["sets"][i]["champions"][j]["ability"][k], list) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in ["variables"]) and all(map(lambda x: isinstance(x, dict), TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"]) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"]))) and all(l in TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"][k] for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in range(len(TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"])) for l in ["name", "value"]) and all(isinstance(TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"][k][l], str) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in range(len(TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"])) for l in ["name"]) and all(isinstance(TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"][k][l], list) or TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"][k][l] == None for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["champions"])) for k in range(len(TFTBasic_initial["sets"][i]["champions"][j]["ability"]["variables"])) for l in ["value"]) and all(k in TFTBasic_initial["sets"][i]["traits"][j] for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["traits"])) for k in ["apiName", "desc", "effects", "icon", "name"]) and all(isinstance(TFTBasic_initial["sets"][i]["traits"][j][k], str) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["traits"])) for k in ["apiName", "desc", "icon", "name"]) and all(isinstance(TFTBasic_initial["sets"][i]["traits"][j][k], list) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["traits"])) for k in ["effects"]) and all(map(lambda x: isinstance(x, dict), TFTBasic_initial["sets"][i]["traits"][j]["effects"]) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["traits"]))) and all(l in TFTBasic_initial["sets"][i]["traits"][j]["effects"][k] for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["traits"])) for k in range(len(TFTBasic_initial["sets"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style", "variables"]) and all(isinstance(TFTBasic_initial["sets"][i]["traits"][j]["effects"][k][l], int) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["traits"])) for k in range(len(TFTBasic_initial["sets"][i]["traits"][j]["effects"])) for l in ["maxUnits", "minUnits", "style"]) and all(isinstance(TFTBasic_initial["sets"][i]["traits"][j]["effects"][k][l], dict) for i in TFTBasic_initial["sets"] for j in range(len(TFTBasic_initial["sets"][i]["traits"])) for k in range(len(TFTBasic_initial["sets"][i]["traits"][j]["effects"])) for l in ["variables"]):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈基础数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT basic data archived in CommunityDragon database (%s)!" %(TFTBasic_url, TFTBasic_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的云顶之弈基础信息Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT basics Json data file path!' %(TFTBasic_local, TFTBasic_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有云顶之弈基础信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT basic information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈基础数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT basic data archived in CommunityDragon database (%s)!" %(TFTBasic_url, TFTBasic_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取云顶之弈英雄数据（The following code get TFT champion data）
            try:
                logPrint("正在加载云顶之弈棋子信息……\nLoading TFT champion information from CommunityDragon ...")
                source, status, session = requestUrl("GET", TFTChampion_url, session = session, log = log)
                if source.ok:
                    TFTChampion_initial: list[dict[str, Any]] = source.json() #TFTChampion存储云顶之弈的棋子信息（Variable `TFTChampion_initial` stores information of TFT champions）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('云顶之弈英雄信息获取超时！正在尝试离线加载数据……\nTFT champion information capture timeout! Trying loading offline data ...\n请输入云顶之弈英雄Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT champion Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTChampion_local_default, TFTChampion_local_default))
                while True:
                    TFTChampion_local: str = logInput()
                    if TFTChampion_local == "":
                        TFTChampion_local = TFTChampion_local_default
                    elif TFTChampion_local[0] == "0":
                        logPrint("云顶之弈英雄信息获取失败！请检查系统网络状况和代理设置。\nTFT champion information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(TFTChampion_local, "r", encoding = "utf-8") as fp:
                            TFTChampion_initial = json.load(fp)
                        if isinstance(TFTChampion_initial, list) and all(isinstance(TFTChampion_initial[i], dict) for i in range(len(TFTChampion_initial))) and all(TFTChampion_initial[i].get(j, 0) for i in range(len(TFTChampion_initial)) for j in ["name", "character_record"]):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈棋子数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT champion data archived in CommunityDragon database (%s)!" %(TFTChampion_url, TFTChampion_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的云顶之弈棋子Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT champion Json data file path!' %(TFTChampion_local, TFTChampion_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有云顶之弈英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT champion information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈棋子数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT champion data archived in CommunityDragon database (%s)!" %(TFTChampion_url, TFTChampion_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取云顶之弈装备数据（The following code get TFT item information）
            try:
                logPrint("正在加载云顶之弈装备信息……\nLoading TFT item information from CommunityDragon ...")
                source, status, session = requestUrl("GET", TFTItem_url, session = session, log = log)
                if source.ok:
                    TFTItem_initial: list[dict[str, Any]] = source.json() #TFTItem存储云顶之弈的装备信息（Variable `TFTItem_initial` stores information of TFT items）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('云顶之弈装备信息获取超时！正在尝试离线加载数据……\nTFT item information capture timeout! Trying loading offline data ...\n请输入云顶之弈装备Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT item Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTItem_local_default, TFTItem_local_default))
                while True:
                    TFTItem_local: str = logInput()
                    if TFTItem_local == "":
                        TFTItem_local = TFTItem_local_default
                    elif TFTItem_local[0] == "0":
                        logPrint("云顶之弈装备信息获取失败！请检查系统网络状况和代理设置。\nTFT item information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(TFTItem_local, "r", encoding = "utf-8") as fp:
                            TFTItem_initial = json.load(fp)
                        if isinstance(TFTItem_initial, list) and all(isinstance(TFTItem_initial[i], dict) for i in range(len(TFTItem_initial))) and (all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "loadoutsIcon"]) or all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "squareIconPath"])):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT item data archived in CommunityDragon database (%s)!" %(TFTItem_url, TFTItem_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的云顶之弈装备Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT item Json data file path!' %(TFTItem_local, TFTItem_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有云顶之弈装备信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT companion information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈装备数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT item data archived in CommunityDragon database (%s)!" %(TFTItem_url, TFTItem_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取云顶之弈小小英雄数据（The following code get TFT companion data）
            try:
                logPrint("正在加载云顶之弈小小英雄信息……\nLoading companion information from CommunityDragon ...")
                source, status, session = requestUrl("GET", TFTCompanion_url, session = session, log = log)
                if source.ok:
                    TFTCompanion_initial: list[dict[str, Any]] = source.json() #TFTChampion存储云顶之弈的小小英雄信息（Variable `TFTChampion_initial` stores information of companions）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('云顶之弈小小英雄信息获取超时！正在尝试离线加载数据……\nTFT companion information capture timeout! Trying loading offline data ...\n请输入云顶之弈小小英雄Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT companion Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTCompanion_local_default, TFTCompanion_local_default))
                while True:
                    TFTCompanion_local: str = logInput()
                    if TFTCompanion_local == "":
                        TFTCompanion_local = TFTCompanion_local_default
                    elif TFTCompanion_local[0] == "0":
                        logPrint("云顶之弈小小英雄信息获取失败！请检查系统网络状况和代理设置。\nTFT companion information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(TFTCompanion_local, "r", encoding = "utf-8") as fp:
                            TFTCompanion_initial = json.load(fp)
                        if isinstance(TFTCompanion_initial, list) and all(isinstance(TFTCompanion_initial[i], dict) for i in range(len(TFTCompanion_initial))) and all(j in TFTCompanion_initial[i] for i in range(len(TFTCompanion_initial)) for j in ["contentId", "itemId", "name", "loadoutsIcon", "description", "level", "speciesName", "speciesId", "rarity", "rarityValue", "isDefault", "upgrades", "TFTOnly"]):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈小小英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT companion data archived in CommunityDragon database (%s)!" %(TFTCompanion_url, TFTCompanion_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的云顶之弈小小英雄Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT companion Json data file path!' %(TFTCompanion_local, TFTCompanion_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有云顶之弈小小英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT companion information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈小小英雄数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT companion data archived in CommunityDragon database (%s)!" %(TFTCompanion_url, TFTCompanion_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取云顶之弈羁绊数据（The following code get TFT trait data）
            try:
                logPrint("正在加载云顶之弈羁绊信息……\nLoading TFT trait information from CommunityDragon ...")
                source, status, session = requestUrl("GET", TFTTrait_url, session = session, log = log)
                if source.ok:
                    TFTTrait_initial: list[dict[str, Any]] = source.json() #TFTTrait存储云顶之弈的羁绊信息（Variable `TFTTrait_initial` stores information of TFT traits）
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('云顶之弈羁绊信息获取超时！正在尝试离线加载数据……\nTFT trait information capture timeout! Trying loading offline data ...\n请输入云顶之弈羁绊Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the TFT trait Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(TFTTrait_local_default, TFTTrait_local_default))
                while True:
                    TFTTrait_local: str = logInput()
                    if TFTTrait_local == "":
                        TFTTrait_local = TFTTrait_local_default
                    elif TFTTrait_local[0] == "0":
                        logPrint("云顶之弈羁绊信息获取失败！请检查系统网络状况和代理设置。\nTFT trait information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(TFTTrait_local, "r", encoding = "utf-8") as fp:
                            TFTTrait_initial = json.load(fp)
                        if isinstance(TFTTrait_initial, list) and all(isinstance(TFTTrait_initial[i], dict) for i in range(len(TFTTrait_initial))) and all(j in TFTTrait_initial[i] for i in range(len(TFTTrait_initial)) for j in ["display_name", "trait_id", "set", "icon_path", "tooltip_text", "innate_trait_sets", "conditional_trait_sets"]):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈羁绊数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT trait data archived in CommunityDragon database (%s)!" %(TFTTrait_url, TFTTrait_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的云顶之弈羁绊Json数据文件路径！\nFile "%s" NOT found! Please input a correct TFT trait Json data file path!' %(TFTTrait_local, TFTTrait_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有云顶之弈小小英雄信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with TFT companion information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的云顶之弈羁绊数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the TFT trait data archived in CommunityDragon database (%s)!" %(TFTTrait_url, TFTTrait_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            #下面获取斗魂竞技场强化符文数据（The following code get Arena augment data）
            try:
                logPrint("正在加载斗魂竞技场强化符文信息……\nLoading Arena augment information from CommunityDragon ...")
                source, status, session = requestUrl("GET", CherryAugment_url, session = session, log = log)
                if source.ok:
                    CherryAugment_initial: list[dict[str, Any]] = source.json() #Arena存储斗魂竞技场的强化符文信息（Variable `CherryAugment_initial` stores information of Arena augments）
                    break
                else:
                    logPrint(source)
                    logPrint("当前语言不可用！请切换语言或检查源代码中的链接。\nCurrent language isn't available! Please change another language or check the requests link in the source code.")
                    switch_language = True
                    break
            except requests.exceptions.RequestException:
                logPrint('斗魂竞技场强化符文信息获取超时！正在尝试离线加载数据……\nArena augment information capture timeout! Trying loading offline data ...\n请输入斗魂竞技场强化符文Json数据文件路径。输入空字符以使用默认相对引用路径“%s”。输入“2”以转为离线模式。输入“0”以退出程序。\nPlease enter the Arena augment Json data file path. Enter an empty string to use the default relative path: "%s". Submit "2" to switch to offline mode. Submit "0" to exit.' %(CherryAugment_local_default, CherryAugment_local_default))
                while True:
                    CherryAugment_local: str = logInput()
                    if CherryAugment_local == "":
                        CherryAugment_local = CherryAugment_local_default
                    elif CherryAugment_local[0] == "0":
                        logPrint("斗魂竞技场强化符文信息获取失败！请检查系统网络状况和代理设置。\nArena augment information capture failure! Please check the system network condition and proxy configuration.")
                        time.sleep(5)
                        return (switch_language, True)
                    else:
                        switch_prepare_mode = True
                        break
                    try:
                        with open(CherryAugment_local, "r", encoding = "utf-8") as fp:
                            CherryAugment_initial = json.load(fp)
                        if isinstance(CherryAugment_initial, list) and all(isinstance(CherryAugment_initial[i], dict) for i in range(len(CherryAugment_initial))) and all(j in CherryAugment_initial[i] for i in range(len(CherryAugment_initial)) for j in ["id", "nameTRA", "augmentSmallIconPath", "rarity"]) and all(isinstance(CherryAugment_initial[i]["id"], int) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["nameTRA"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["augmentSmallIconPath"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["rarity"], str) for i in range(len(CherryAugment_initial))):
                            break
                        else:
                            logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的斗魂竞技场强化符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the Arena augment data archived in CommunityDragon database (%s)!" %(CherryAugment_url, CherryAugment_url))
                    except FileNotFoundError:
                        logPrint('未找到文件“%s”！请输入正确的斗魂竞技场强化符文Json数据文件路径！\nFile "%s" NOT found! Please input a correct Arena augment Json data file path!' %(CherryAugment_local, CherryAugment_local))
                    except OSError:
                        logPrint("数据文件名不合法！请输入含有斗魂竞技场强化符文信息的本地文件的路径！\nIllegal data filename! Please input the path of a local file with Arena augment information.")
                    except json.decoder.JSONDecodeError:
                        logPrint("数据格式错误！请选择一个符合CommunityDragon数据库中记录的斗魂竞技场强化符文数据格式（%s）的数据文件！\nData format mismatched! Please select a data file that corresponds to the format of the Arena augment data archived in CommunityDragon database (%s)!" %(CherryAugment_url, CherryAugment_url))
            if switch_prepare_mode:
                prepareMode = ""
                continue
            break
        else:
            switch_prepare_mode = False
            logPrint('请在浏览器中打开以下网页，待加载完成后按Ctrl + S保存网页json文件至同目录的“离线数据（Offline Data）”文件夹下，并根据括号内的提示放置和命名文件。\nPlease open the following URLs in a browser, then press Ctrl + S to save the online json files into the folder "离线数据（Offline Data）" under the same directory after the website finishes loading and organize and rename the downloaded files according to the hints in the circle brackets.\n版本信息（%s）： %s\n游戏模式（%s）： %s\n召唤师技能（%s）： %s\n英雄（%s）： %s\n英雄联盟装备（%s）： %s\n召唤师图标（%s）： %s\n基石符文（%s）： %s\n符文系（%s）： %s\n云顶之弈基础信息（%s）： %s\n云顶之弈棋子（%s）： %s\n云顶之弈装备（%s）： %s\n云顶之弈小小英雄（%s）： %s\n云顶之弈羁绊（%s）： %s\n斗魂竞技场强化符文（%s）： %s' %(patches_local_default[19:], patches_url, queue_local_default[19:], queue_url, spell_local_default[19:], spell_url, LoLChampion_local_default[19:], LoLChampion_url, LoLItem_local_default[19:], LoLItem_url, summonerIcon_local_default[19:], summonerIcon_url, perk_local_default[19:], perk_url, perkstyle_local_default[19:], perkstyle_url, TFTBasic_local_default[19:], TFTBasic_url, TFTChampion_local_default[19:], TFTChampion_url, TFTItem_local_default[19:], TFTItem_url, TFTCompanion_local_default[19:], TFTCompanion_url, TFTTrait_local_default[19:], TFTTrait_url, CherryAugment_local_default[19:], CherryAugment_url))
            offline_files_loaded: dict[str, bool] = {"patch": False, "queue": False, "spell": False, "LoLChampion": False, "LoLItem": False, "summonerIcon": False, "perk": False, "perkstyle": False, "TFT": False, "TFTChampion": False, "TFTItem": False, "TFTCompanion": False, "TFTTrait": False, "CherryAugment": False}
            offline_files: dict[str, dict[str, str]] = {"patch": {"file": patches_local_default, "URL": patches_url, "content": "版本信息"}, "queue": {"file": queue_local_default, "URL": queue_url, "content": "游戏模式"}, "spell": {"file": spell_local_default, "URL": spell_url, "content": "召唤师技能"}, "LoLChampion": {"file": LoLChampion_local_default, "URL": LoLChampion_url, "content": "英雄"}, "LoLItem": {"file": LoLItem_local_default, "URL": LoLItem_url, "content": "英雄联盟装备"}, "summonerIcon": {"file": summonerIcon_local_default, "URL": summonerIcon_url, "content": "召唤师图标"}, "perk": {"file": perk_local_default, "URL": perk_url, "content": "基石符文"}, "perkstyle": {"file": perkstyle_local_default, "URL": perkstyle_url, "content": "符文系"}, "TFT": {"file": TFTBasic_local_default, "URL": TFTBasic_url, "content": "云顶之弈基础信息"}, "TFTChampion": {"file": TFTChampion_local_default, "URL": TFTChampion_url, "content": "云顶之弈英雄"}, "TFTItem": {"file": TFTItem_local_default, "URL": TFTItem_url, "content": "云顶之弈装备"}, "TFTCompanion": {"file": TFTCompanion_local_default, "URL": TFTCompanion_url, "content": "云顶之弈小小英雄"}, "TFTTrait": {"file": TFTTrait_local_default, "URL": TFTTrait_url, "content": "云顶之弈羁绊"}, "CherryAugment": {"file": CherryAugment_local_default, "URL": CherryAugment_url, "content": "斗魂竞技场强化符文"}}
            logPrint('按回车键以加载离线数据。输入“1”以转为在线模式。输入“0”以退出程序。\nPress Enter to load offline data. Input "1" to switch to online mode. Submit "0" to exit.')
            while any(not _ for _ in offline_files_loaded.values()):
                offline_files_notfound: dict[str, bool] = {"patch": False, "queue": False, "spell": False, "LoLChampion": False, "LoLItem": False, "summonerIcon": False, "perk": False, "perkstyle": False, "TFT": False, "TFTChampion": False, "TFTItem": False, "TFTCompanion": False, "TFTTrait": False, "CherryAugment": False}
                offline_files_formaterror: dict[str, bool] = {"patch": False, "queue": False, "spell": False, "LoLChampion": False, "LoLItem": False, "summonerIcon": False, "perk": False, "perkstyle": False, "TFT": False, "TFTChampion": False, "TFTItem": False, "TFTCompanion": False, "TFTTrait": False, "CherryAugment": False}
                prepareMode = logInput()
                if prepareMode != "" and prepareMode[0] == "1":
                    switch_prepare_mode = True
                    break
                if prepareMode != "" and prepareMode[0] == "0":
                    return (switch_language, True)
                #下面获取版本信息（The following code get the patch data）
                if not offline_files_loaded["patch"]:
                    try:
                        with open(patches_local_default, "r", encoding = "utf-8") as fp:
                            patches_initial = json.load(fp)
                        if not (isinstance(patches_initial, list) and patches_initial[-1] == "lolpatch_3.7"): #之所以将patches的最后一个元素作为判断版本文件数据格式合法的依据，是因为按照这样的逻辑，代码在一般情况下就不需要频繁变动（The reason why I use the last element of the variable `patches_initial` as the judgment whether the patch file data format is legal is, that under this logic, the code won't need further adjustment as the update goes on）
                            offline_files_formaterror["patch"] = True
                    except FileNotFoundError:
                        offline_files_notfound["patch"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["patch"] = True
                    else:
                        if not offline_files_formaterror["patch"]:
                            offline_files_loaded["patch"] = True
                            latest_patch = patches_initial[0]
                            patches_dict = {}
                            smallPatches = []
                            bigPatches = []
                            for patch in patches_initial:
                                if not patch.startswith("lolpatch"):
                                    patch_split = patch.split(".")
                                    smallPatch = ".".join(patch_split[:3])
                                    smallPatches.append(Patch(smallPatch))
                                    bigPatch = ".".join(patch_split[:2])
                                    bigPatches.append(Patch(bigPatch))
                                    patches_dict[bigPatch] = []
                            for i in range(len(bigPatches)):
                                patches_dict[str(bigPatches[i])].append(str(smallPatches[i]))
                #下面获取游戏模式数据（The following code get game mode data）
                if not offline_files_loaded["queue"]:
                    try:
                        with open(queue_local_default, "r", encoding = "utf-8") as fp:
                            queue_initial = json.load(fp)
                        if not(isinstance(queue_initial, list) and all(map(lambda x: all(i in x for i in ["id", "name", "shortName", "description", "detailedDescription", "gameSelectModeGroup", "gameSelectCategory", "gameSelectPriority", "isSkillTreeQueue", "isLimitedTimeQueue", "isBotHonoringAllowed", "hidePlayerPosition", "viableChampionRoster"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], int) for i in ["id", "gameSelectPriority"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], str) for i in ["name", "shortName", "description", "detailedDescription", "gameSelectModeGroup", "gameSelectCategory"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], bool) for i in ["isSkillTreeQueue", "isLimitedTimeQueue", "isBotHonoringAllowed", "hidePlayerPosition"]), queue_initial)) and all(map(lambda x: all(isinstance(x[i], list) or x[i] is None for i in ["viableChampionRoster"]), queue_initial))):
                            offline_files_formaterror["queue"] = True
                    except FileNotFoundError:
                        offline_files_notfound["queue"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["queue"] = True
                    else:
                        if not offline_files_formaterror["queue"]:
                            offline_files_loaded["queue"] = True
                #下面获取召唤师技能数据（The following code get summoner spell data）
                if not offline_files_loaded["spell"]:
                    try:
                        with open(spell_local_default, "r", encoding = "utf-8") as fp:
                            spell_initial = json.load(fp)
                        if not(isinstance(spell_initial, list) and all(i in spell_initial[j] for i in ["id", "name", "description", "summonerLevel", "cooldown", "gameModes", "iconPath"] for j in range(len(spell_initial)))):
                            offline_files_formaterror["spell"] = True
                    except FileNotFoundError:
                        offline_files_notfound["spell"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["spell"] = True
                    else:
                        if not offline_files_formaterror["spell"]:
                            offline_files_loaded["spell"] = True
                #下面获取英雄信息（The following code get LoL champion data）
                if not offline_files_loaded["LoLChampion"]:
                    try:
                        with open(LoLChampion_local_default, "r", encoding = "utf-8") as fp:
                            LoLChampion_initial = json.load(fp)
                        if not(isinstance(LoLChampion_initial, list) and all(isinstance(LoLChampion_initial[i], dict) for i in range(len(LoLChampion_initial))) and all(j in LoLChampion_initial[i] for i in range(len(LoLChampion_initial)) for j in ["id", "name", "alias", "squarePortraitPath", "roles"]) and all(isinstance(LoLChampion_initial[i]["id"], int) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["name"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["alias"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["squarePortraitPath"], str) for i in range(len(LoLChampion_initial))) and all(isinstance(LoLChampion_initial[i]["roles"], list) for i in range(len(LoLChampion_initial)))):
                            offline_files_formaterror["LoLChampion"] = True
                    except FileNotFoundError:
                        offline_files_notfound["LoLChampion"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["LoLChampion"] = True
                    else:
                        if not offline_files_formaterror["LoLChampion"]:
                            offline_files_loaded["LoLChampion"] = True
                #下面获取英雄联盟装备信息（The following code get LoL item data）
                if not offline_files_loaded["LoLItem"]:
                    try:
                        with open(LoLItem_local_default, "r", encoding = "utf-8") as fp:
                            LoLItem_initial = json.load(fp)
                        if not(isinstance(LoLItem_initial, list) and all(i in LoLItem_initial[j] for i in ["id", "name", "description", "active", "inStore", "from", "to", "categories", "maxStacks", "requiredChampion", "requiredAlly", "requiredBuffCurrencyName", "requiredBuffCurrencyCost", "specialRecipe", "isEnchantment", "price", "priceTotal", "iconPath"] for j in range(len(LoLItem_initial)))):
                            offline_files_formaterror["LoLItem"] = True
                    except FileNotFoundError:
                        offline_files_notfound["LoLItem"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["LoLItem"] = True
                    else:
                        if not offline_files_formaterror["LoLItem"]:
                            offline_files_loaded["LoLItem"] = True
                #下面获取召唤师图标信息（The following code get summoner icon data）
                if not offline_files_loaded["summonerIcon"]:
                    try:
                        with open(summonerIcon_local_default, "r", encoding = "utf-8") as fp:
                            summonerIcon_initial = json.load(fp)
                        if not(isinstance(summonerIcon_initial, list) and all(map(lambda x: isinstance(x, dict), summonerIcon_initial)) and all(i in j for i in ["id", "title", "yearReleased", "isLegacy", "descriptions", "rarities", "disabledRegions"] for j in summonerIcon_initial)):
                            offline_files_formaterror["summonerIcon"] = True
                    except FileNotFoundError:
                        offline_files_notfound["summonerIcon"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["summonerIcon"] = True
                    else:
                        if not offline_files_formaterror["summonerIcon"]:
                            offline_files_loaded["summonerIcon"] = True
                #下面获取基石符文信息（The following code get perk data）
                if not offline_files_loaded["perk"]:
                    try:
                        with open(perk_local_default, "r", encoding = "utf-8") as fp:
                            perk_initial = json.load(fp)
                        if not(isinstance(perk_initial, list) and all(i in perk_initial[j] for i in ["id", "name", "majorChangePatchVersion", "tooltip", "shortDesc", "longDesc", "recommendationDescriptor", "iconPath", "endOfGameStatDescs", "recommendationDescriptorAttributes"] for j in range(len(perk_initial)))):
                            offline_files_formaterror["perk"] = True
                    except FileNotFoundError:
                        offline_files_notfound["perk"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["perk"] = True
                    else:
                        if not offline_files_formaterror["perk"]:
                            offline_files_loaded["perk"] = True
                #下面获取符文系信息（The following code get perkstyle data）
                if not offline_files_loaded["perkstyle"]:
                    try:
                        with open(perkstyle_local_default, "r", encoding = "utf-8") as fp:
                            perkstyle_initial = json.load(fp)
                        if not(isinstance(perkstyle_initial, dict) and all(perkstyle_initial.get(i, 0) for i in ["schemaVersion", "styles"]) and isinstance(perkstyle_initial["styles"], list) and all(j in perkstyle_initial["styles"][i] for i in range(len(perkstyle_initial["styles"])) for j in ["id", "name", "tooltip", "iconPath", "assetMap", "isAdvanced", "allowedSubStyles", "subStyleBonus", "slots", "defaultPageName", "defaultSubStyle", "defaultPerks", "defaultPerksWhenSplashed", "defaultStatModsPerSubStyle"])):
                            offline_files_formaterror["perkstyle"] = True
                    except FileNotFoundError:
                        offline_files_notfound["perkstyle"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["perkstyle"] = True
                    else:
                        if not offline_files_formaterror["perkstyle"]:
                            offline_files_loaded["perkstyle"] = True
                #下面获取云顶之弈强化符文数据（The following code get TFT augment data）
                if not offline_files_loaded["TFT"]:
                    try:
                        with open(TFTBasic_local_default, "r", encoding = "utf-8") as fp:
                            TFTBasic_initial = json.load(fp)
                        if not(isinstance(TFTBasic_initial, dict) and all(i in TFTBasic_initial for i in ["items", "setData", "sets"])):
                            offline_files_formaterror["TFT"] = True
                    except FileNotFoundError:
                        offline_files_notfound["TFT"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["TFT"] = True
                    else:
                        if not offline_files_formaterror["TFT"]:
                            offline_files_loaded["TFT"] = True
                #下面获取云顶之弈英雄数据（The following code get TFT champion data）
                if not offline_files_loaded["TFTChampion"]:
                    try:
                        with open(TFTChampion_local_default, "r", encoding = "utf-8") as fp:
                            TFTChampion_initial = json.load(fp)
                        if not(isinstance(TFTChampion_initial, list) and all(isinstance(TFTChampion_initial[i], dict) for i in range(len(TFTChampion_initial))) and all(TFTChampion_initial[i].get(j, 0) for i in range(len(TFTChampion_initial)) for j in ["name", "character_record"])):
                            offline_files_formaterror["TFTChampion"] = True
                    except FileNotFoundError:
                        offline_files_notfound["TFTChampion"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["TFTChampion"] = True
                    else:
                        if not offline_files_formaterror["TFTChampion"]:
                            offline_files_loaded["TFTChampion"] = True
                #下面获取云顶之弈装备数据（The following code get TFT item information）
                if not offline_files_loaded["TFTItem"]:
                    try:
                        with open(TFTItem_local_default, "r", encoding = "utf-8") as fp:
                            TFTItem_initial = json.load(fp)
                        if not(isinstance(TFTItem_initial, list) and all(isinstance(TFTItem_initial[i], dict) for i in range(len(TFTItem_initial))) and (all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "loadoutsIcon"]) or all(j in TFTItem_initial[i] for i in range(len(TFTItem_initial)) for j in ["guid", "name", "nameId", "id", "color", "squareIconPath"]))):
                            offline_files_formaterror["TFTItem"] = True
                    except FileNotFoundError:
                        offline_files_notfound["TFTItem"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["TFTItem"] = True
                    else:
                        if not offline_files_formaterror["TFTItem"]:
                            offline_files_loaded["TFTItem"] = True
                #下面获取云顶之弈小小英雄数据（The following code get TFT companion data）
                if not offline_files_loaded["TFTCompanion"]:
                    try:
                        with open(TFTCompanion_local_default, "r", encoding = "utf-8") as fp:
                            TFTCompanion_initial = json.load(fp)
                        if not(isinstance(TFTCompanion_initial, list) and all(isinstance(TFTCompanion_initial[i], dict) for i in range(len(TFTCompanion_initial))) and all(j in TFTCompanion_initial[i] for i in range(len(TFTCompanion_initial)) for j in ["contentId", "itemId", "name", "loadoutsIcon", "description", "level", "speciesName", "speciesId", "rarity", "rarityValue", "isDefault", "upgrades", "TFTOnly"])):
                            offline_files_formaterror["TFTCompanion"] = True
                    except FileNotFoundError:
                        offline_files_notfound["TFTCompanion"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["TFTCompanion"] = True
                    else:
                        if not offline_files_formaterror["TFTCompanion"]:
                            offline_files_loaded["TFTCompanion"] = True
                #下面获取云顶之弈羁绊数据（The following code get TFT trait data）
                if not offline_files_loaded["TFTTrait"]:
                    try:
                        with open(TFTTrait_local_default, "r", encoding = "utf-8") as fp:
                            TFTTrait_initial = json.load(fp)
                        if not(isinstance(TFTTrait_initial, list) and all(isinstance(TFTTrait_initial[i], dict) for i in range(len(TFTTrait_initial))) and all(j in TFTTrait_initial[i] for i in range(len(TFTTrait_initial)) for j in ["display_name", "trait_id", "set", "icon_path", "tooltip_text", "innate_trait_sets", "conditional_trait_sets"])):
                            offline_files_formaterror["TFTTrait"] = True
                    except FileNotFoundError:
                        offline_files_notfound["TFTTrait"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["TFTTrait"] = True
                    else:
                        if not offline_files_formaterror["TFTTrait"]:
                            offline_files_loaded["TFTTrait"] = True
                #下面获取斗魂竞技场强化符文数据（The following code get Arena augment data）
                if not offline_files_loaded["CherryAugment"]:
                    try:
                        with open(CherryAugment_local_default, "r", encoding = "utf-8") as fp:
                            CherryAugment_initial = json.load(fp)
                        if not(isinstance(CherryAugment_initial, list) and all(isinstance(CherryAugment_initial[i], dict) for i in range(len(CherryAugment_initial))) and all(j in CherryAugment_initial[i] for i in range(len(CherryAugment_initial)) for j in ["id", "nameTRA", "augmentSmallIconPath", "rarity"]) and all(isinstance(CherryAugment_initial[i]["id"], int) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["nameTRA"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["augmentSmallIconPath"], str) for i in range(len(CherryAugment_initial))) and all(isinstance(CherryAugment_initial[i]["rarity"], str) for i in range(len(CherryAugment_initial)))):
                            offline_files_formaterror["CherryAugment"] = True
                    except FileNotFoundError:
                        offline_files_notfound["CherryAugment"] = True
                    except json.decoder.JSONDecodeError:
                        offline_files_formaterror["CherryAugment"] = True
                    else:
                        if not offline_files_formaterror["CherryAugment"]:
                            offline_files_loaded["CherryAugment"] = True
                #下面总结离线数据加载情况（The following code conclude the result of loading offline data）
                unloaded_offline_files: list[str] = []
                notfound_offline_files: list[str] = []
                formaterror_offline_files: list[str] = []
                if any(offline_files_notfound.values()):
                    for i in offline_files_notfound:
                        if offline_files_notfound[i]:
                            notfound_offline_files.append(i)
                            unloaded_offline_files.append(i)
                    logPrint("以下信息文件不存在：\nNot existing file(s):")
                    for i in notfound_offline_files:
                        logPrint(offline_files[i]["file"] + "\t" + offline_files[i]["content"] + "\t" + offline_files[i]["URL"])
                if any(offline_files_formaterror.values()):
                    for i in offline_files_formaterror:
                        if offline_files_formaterror[i]:
                            formaterror_offline_files.append(i)
                            unloaded_offline_files.append(i)
                    logPrint("以下信息文件格式错误：\nFormatError file(s):")
                    for i in formaterror_offline_files:
                        logPrint(offline_files[i]["file"] + "\t" + offline_files[i]["content"] + "\t" + offline_files[i]["URL"])
                if any(not i for i in offline_files_loaded.values()):
                    logPrint('按回车键以加载离线数据。输入“1”以转为在线模式。输入“0”以退出程序。\nPress Enter to load offline data. Input "1" to switch to online mode. Submit "0" to exit.')
            if not switch_prepare_mode:
                break
    if not switch_language:
        #下面按照程序需求对数据资源进行一定的整理。需要注意，从Json中读取到的整数键会被转换为字符串（The following code organize the data resource according to the program's need. Note that integer keys read from local json files will transform into strings）
        queues_initial = {int(queue_iter["id"]): queue_iter for queue_iter in queue_initial} #queues为嵌套字典，键为队列序号，值为游戏模式信息字典。一个键值对的示例如右：（Variable `queues` is a nested dictionary, whose keys are queueIds and values are game mode information dictionaries. An example of the key-value pairs is shown as follows: ）{"id": 0, "name": "", "shortName": "", "description": "", "detailedDescription": "", "gameSelectModeGroup": "kARAM", "gameSelectCategory": "kPvP", "gameSelectPriority": 0, "isSkillTreeQueue": false, "isLimitedTimeQueue": false, "isBotHonoringAllowed": false, "hidePlayerPosition": false, "viableChampionRoster": null}
        spells_initial = {int(spell_iter["id"]): spell_iter for spell_iter in spell_initial} #spells为嵌套字典，键为召唤师技能序号，值为召唤师技能信息字典。一个键值对的示例如右：（Variable `spells` is a nested dictionary, whose keys are spellIds and values are spell information dictionaries. An example of the key-value pairs is shown as follows: ）{1: {"name": "净化", "description": "移除身上的所有限制效果（压制效果和击飞效果除外）和召唤师技能的减益效果，并且若在接下来的3秒里再次被施加限制效果时，新效果的持续时间会减少65%。", "summonerLevel": 9, "cooldown": 210, "gameModes": ["URF", "CLASSIC", "ARSR", "ARAM", "ULTBOOK", "WIPMODEWIP", "TUTORIAL", "DOOMBOTSTEEMO", "PRACTICETOOL", "FIRSTBLOOD", "NEXUSBLITZ", "PROJECT", "ONEFORALL"], "iconPath": "/lol-game-data/assets/DATA/Spells/Icons2D/Summoner_boost.png"}}
        LoLChampions_initial = {int(LoLChampion_iter["id"]): LoLChampion_iter for LoLChampion_iter in LoLChampion_initial} #LoLChampions为嵌套字典，键为英雄序号，值为英雄信息字典。一个键值对的示例如右：（Variable `LoLItems` is a nested dictionary, whose keys are itemIds and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{1: {"name": "黑暗之女", "alias": "Annie", "squarePortraitPath": "/lol-game-data/assets/v1/champion-icons/1.png", "roles": ["mage", "support"]}}
        LoLItems_initial = {int(LoLItem_iter["id"]): LoLItem_iter for LoLItem_iter in LoLItem_initial} #LoLItems为嵌套字典，键为装备序号，值为装备信息字典。一个键值对的示例如右：（Variable `LoLItems` is a nested dictionary, whose keys are itemIds and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{1001: {"name": "鞋子", "description": "<mainText><stats><attention>25</attention>移动速度</stats></mainText><br>", "active": False, "inStore": True, "from": [], "to": [3111, 3006, 3005, 3009, 3020, 3047, 3117, 3158], "categories": ["Boots"], "maxStacks": 1, "requiredChampion": "", "requiredAlly": "", "requiredBuffCurrencyName": "", "requiredBuffCurrencyCost": 0, "specialRecipe": 0, "isEnchantment": False, "price": 300, "priceTotal": 300, "iconPath": "/lol-game-data/assets/ASSETS/Items/Icons2D/1001_Class_T1_BootsofSpeed.png"}}
        summonerIcons_initial = {int(summonerIcon_iter["id"]): summonerIcon_iter for summonerIcon_iter in summonerIcon_initial} #summonerIcons为嵌套字典，键为装备序号，值为装备信息字典。一个键值对的示例如右：（Variable `summonerIcons` is a nested dictionary, whose keys are itemIds and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{0: {"id":0,"title":"可爱凯尔 图标","yearReleased":2009,"isLegacy":false,"imagePath":"/lol-game-data/assets/v1/profile-icons/0.jpg","descriptions":[{"region":"riot","description":" "}],"rarities":[{"region":"riot","rarity":0}],"disabledRegions":[]},{"id":1000,"title":"2016 LCL Hard Random","yearReleased":2016,"isLegacy":false,"imagePath":"/lol-game-data/assets/v1/profile-icons/1000.jpg","esportsTeam":"Hard Random","esportsRegion":"RU","esportsEvent":"英雄联盟欧陆联赛 LCL","descriptions":[{"region":"riot","description":" "}],"rarities":[{"region":"riot","rarity":0}],"disabledRegions":[]}}
        perks_initial = {int(perk_iter["id"]): perk_iter for perk_iter in perk_initial} #perks为嵌套字典，键为符文序号，值为符文信息字典。一个键值对的示例如右：（Variable `perks` is a nested dictionary, whose keys are perkIds and values are perk information dictionaries. An example of the key-value pairs is shown as follows: ）{8369: {"name": "先攻", "majorChangePatchVersion": "11.23", "tooltip": "在进入与英雄战斗的@GraceWindow.2@秒内，对一名敌方英雄进行的攻击或技能将提供@GoldProcBonus@金币和<b>先攻</b>效果，持续@Duration@秒，来使你对英雄们造成<truedamage>@DamageAmp*100@%</truedamage>额外<truedamage>伤害</truedamage>，并提供<gold>{{ Item_Melee_Ranged_Split }}</gold>该额外伤害值的<gold>金币</gold>。<br><br>冷却时间：<scaleLevel>@Cooldown@</scaleLevel>秒<br><hr><br>已造成的伤害：@f1@<br>已提供的金币：@f2@", "shortDesc": "在你率先发起与英雄的战斗时，造成8%额外伤害，持续3秒，并基于该额外伤害提供金币。", "longDesc": "在进入与英雄战斗的0.25秒内，对一名敌方英雄进行的攻击或技能将提供5金币和<b>先攻</b>效果，持续3秒，来使你对英雄们造成<truedamage>8%</truedamage>额外<truedamage>伤害</truedamage>，并提供<gold>100% (远程英雄为70%)</gold>该额外伤害值的<gold>金币</gold>。<br><br>冷却时间：<scaleLevel>25 ~ 15</scaleLevel>秒", "recommendationDescriptor": "真实伤害，金币收入", "iconPath": "/lol-game-data/assets/v1/perk-images/Styles/Inspiration/FirstStrike/FirstStrike.png", "endOfGameStatDescs": ["已造成的伤害：@eogvar1@", "已提供的金币：@eogvar2@"], "recommendationDescriptorAttributes": {}}}
        perkstyles_initial = {int(perkstyle_iter["id"]): perkstyle_iter for perkstyle_iter in perkstyle_initial["styles"]} #perkstyles为嵌套字典，键为符文系序号，值为符文系信息字典。一个键值对的示例如右：（Variable `perkstyles` is a nested dictionary, whose keys are perkstyle ids and values are perkstyle information dictionaries. An example of the key-value pairs is as follows: ）{8400: {"name": "坚决", "tooltip": "耐久和控制", "iconPath": "/lol-game-data/assets/v1/perk-images/Styles/7204_Resolve.png", "assetMap": {"p8400_s0_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k0.jpg", "p8400_s0_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k8437.jpg", "p8400_s0_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k8439.jpg", "p8400_s0_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s0_k8465.jpg", "p8400_s8000_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k0.jpg", "p8400_s8000_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k8437.jpg", "p8400_s8000_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k8439.jpg", "p8400_s8000_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8000_k8465.jpg", "p8400_s8100_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k0.jpg", "p8400_s8100_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k8437.jpg", "p8400_s8100_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k8439.jpg", "p8400_s8100_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8100_k8465.jpg", "p8400_s8200_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k0.jpg", "p8400_s8200_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k8437.jpg", "p8400_s8200_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k8439.jpg", "p8400_s8200_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8200_k8465.jpg", "p8400_s8300_k0": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k0.jpg", "p8400_s8300_k8437": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k8437.jpg", "p8400_s8300_k8439": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k8439.jpg", "p8400_s8300_k8465": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/p8400_s8300_k8465.jpg", "svg_icon": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/resolve_icon.svg", "svg_icon_16": "/lol-game-data/assets/v1/perk-images/Styles/Resolve/resolve_icon_16.svg"}, "isAdvanced": False, "allowedSubStyles": [8000, 8100, 8200, 8300], "subStyleBonus": [{"styleId": 8000, "perkId": 8414}, {"styleId": 8100, "perkId": 8454}, {"styleId": 8200, "perkId": 8415}, {"styleId": 8300, "perkId": 8416}], "slots": [{"type": "kKeyStone", "slotLabel": "", "perks": [8437, 8439, 8465]}, {"type": "kMixedRegularSplashable", "slotLabel": "蛮力", "perks": [8446, 8463, 8401]}, {"type": "kMixedRegularSplashable", "slotLabel": "抵抗", "perks": [8429, 8444, 8473]}, {"type": "kMixedRegularSplashable", "slotLabel": "生机", "perks": [8451, 8453, 8242]}, {"type": "kStatMod", "slotLabel": "进攻", "perks": [5008, 5005, 5007]}, {"type": "kStatMod", "slotLabel": "灵活", "perks": [5008, 5002, 5003]}, {"type": "kStatMod", "slotLabel": "防御", "perks": [5001, 5002, 5003]}], "defaultPageName": "坚决：巨像", "defaultSubStyle": 8200, "defaultPerks": [8437, 8446, 8444, 8451, 8224, 8237, 5008, 5002, 5001], "defaultPerksWhenSplashed": [8444, 8446], "defaultStatModsPerSubStyle": [{"id": "8000", "perks": [5005, 5002, 5001]}, {"id": "8100", "perks": [5008, 5002, 5001]}, {"id": "8200", "perks": [5008, 5002, 5001]}, {"id": "8300", "perks": [5007, 5002, 5001]}]}}
        TFTAugments_initial = {item["apiName"]: item for item in TFTBasic_initial["items"]} #TFTAugments为嵌套字典，键为物件在LCU API上的表达形式，值为物件信息字典。一个键值对的示例如右：（Variable `TFTAugments` is a nested dictionary, whose keys are LCU API representation of items and values are item information dictionaries. An example of the key-value pairs is shown as follows: ）{"TFT7_Consumable_NeekosHelpDragon": {"associatedTraits": [], "composition": [], "desc": "TFT7_Consumable_Description_Dragonling", "effects": {}, "from": None, "icon": "ASSETS/Maps/Particles/TFT/TFT7_Consumable_Dragonling.tex", "id": None, "incompatibleTraits": [], "name": "TFT7_Consumable_Name_Dragonling", "unique": False}}
        TFTChampions_initial = {TFTChampion_iter["name"]: TFTChampion_iter["character_record"] for TFTChampion_iter in TFTChampion_initial} #TFTChampions为嵌套字典，键为棋子在LCU API上的表达形式，值为棋子信息字典。一个键值对的示例如右：（Variable `TFTChampions` is a nested dictionary, whose keys are LCU API representation of TFT Champions and values are TFT Champion information dictionaries. An example of the key-value pairs is shown as follows: ）{"TFT9_Aatrox": {"character_record": {"path": "Characters/TFT9_Aatrox/CharacterRecords/Root", "character_id": "TFT9_Aatrox", "rarity": 9, "display_name": "亚托克斯", "traits": [{"name": "暗裔", "id": "Set9_Darkin"}, {"name": "裁决战士", "id": "Set9_Slayer"}, {"name": "主宰", "id": "Set9_Armorclad"}], "squareIconPath": "/lol-game-data/assets/ASSETS/Characters/TFT9_Aatrox/HUD/TFT9_Aatrox_Square.TFT_Set9.png"}}}
        TFTItems_initial = {TFTItem_iter["nameId"]: TFTItem_iter for TFTItem_iter in TFTItem_initial} #TTItems为嵌套字典，键为云顶之弈装备名称序号，值为云顶之弈装备信息字典。一个键值对的示例如右：（Variable `TFTItems` is a nested dictionary, whose keys are TFT item nameIds and values are TFT item information dictionaries. An example of the key-value pairs is shown as follows: ）{"TFTTutorial_Item_BFSword": {"guid": "9f6e75bb-7ba2-49aa-8724-04c550279034", "name": "暴风大剑", "id": 0, "color": {"R": 73, "B": 54, "G": 68, "A": 255}, "loadoutsIcon": "/lol-game-data/assets/ASSETS/Maps/Particles/TFT/Item_Icons/Standard/BF_Sword.png"}}
        TFTCompanions_initial = {companion_iter["contentId"]: companion_iter for companion_iter in TFTCompanion_initial} #TFTCompanions为嵌套字典，键为小小英雄序号，值为小小英雄信息字典。一个键值对的示例如右：（Variable `TFTCompanions` is a nested dictionary, whose keys are companion contentIds and values are companion information dictionaries. An example of the key-value pairs is shown as follows: ）{"91f2e228-4e36-4dad-9a97-36036e3eca36": {"itemId": 13010, "name": "节奏大师 奥希雅", "loadoutsIcon": "/lol-game-data/assets/ASSETS/Loadouts/Companions/Tooltip_AkaliDragon_Beatmaker_Tier1.png", "description": "奥希雅是酷炫的具象化。它用毫不费力的语流，指挥着韵脚和节奏，甚至能让最出色的小小英雄们羡慕不休。", "level": 1, "speciesName": "奥希雅", "speciesId": 13, "rarity": "Epic", "rarityValue": 1, "isDefault": false, "upgrades": ["0e251d36-d86e-4c58-9b7f-bcee2376a408", "e3151dc2-c45c-4949-89e9-6afda3b2fd5f"], "TFTOnly": false}}
        TFTTraits_initial = {} #TFTTraits为嵌套字典，键为羁绊在LCU API上的表达形式，值为羁绊信息字典。一个键值对的示例如右：（Variable `TFTTraits` is a nested dictionary, whose keys are LCU API representation of traits and values are trait information dictionaries. An example of the key-value pairs is shown as follows: ）{"Assassin": {"display_name": "刺客", "set": "TFTSet1", "icon_path": "/lol-game-data/assets/ASSETS/UX/TraitIcons/Trait_Icon_Assassin.png", "tooltip_text": "固有：在战斗环节开始时，刺客们会跃至距离最远的敌人处。<br><br>刺客们会获得额外的暴击伤害和暴击几率。<br><br><expandRow>(@MinUnits@) +@CritAmpPercent@%暴击伤害和+@CritChanceAmpPercent@%暴击几率</expandRow><br>", "innate_trait_sets": [], "conditional_trait_sets": {2: {"effect_amounts": [{"name": "CritAmpPercent", "value": 75.0, "format_string": ""}, {"name": "CritChanceAmpPercent", "value": 5.0, "format_string": ""}], "min_units": 3, "max_units": 5, "style_name": "kBronze"}, 3: {"effect_amounts": [{"name": "CritAmpPercent", "value": 150.0, "format_string": ""}, {"name": "CritChanceAmpPercent", "value": 20.0, "format_string": ""}], "min_units": 6, "max_units": 8, "style_name": "kSilver"}, 4: {"effect_amounts": [{"name": "CritAmpPercent", "value": 225.0, "format_string": ""}, {"name": "CritChanceAmpPercent", "value": 30.0, "format_string": ""}], "min_units": 9, "max_units": 25000, "style_name": "kGold"}}}}
        for trait_iter in TFTTrait_initial:
            trait_id: str = trait_iter["trait_id"]
            conditional_trait_sets: dict[str, dict[str, Any]] = {}
            for conditional_trait_set in trait_iter["conditional_trait_sets"]:
                style_idx: str = conditional_trait_set["style_idx"]
                conditional_trait_sets[style_idx] = conditional_trait_set
            trait_iter["conditional_trait_sets"] = conditional_trait_sets
            TFTTraits_initial[trait_id] = trait_iter
        CherryAugments_initial = {int(CherryAugment_iter["id"]): CherryAugment_iter for CherryAugment_iter in CherryAugment_initial} #CherryAugments为嵌套字典，键为斗魂竞技场强化符文在LCU API上的表达形式，值为斗魂竞技场强化符文信息字典。一个键值对的实例如右：（Variable `CherryAugments` is a nested dictionary, whose keys are LCU API representation of Arena augments and values are Arena augment information dictionaries. An example of the key-value pairs is shown as follows: ）{205: {"nameTRA": "物理转魔法", "augmentSmallIconPath": "/lol-game-data/assets/ASSETS/UX/Cherry/Augments/Icons/ADAPt_small.png", "rarity": "kSilver"}}
    return (switch_language, False)

async def load_smurf(connection: Connection, current_puuid: str, infos: Optional[dict[str, dict[str, Any]]] = None) -> list[dict[str, Any]]:
    '''
    读取小号信息。<br>Load smurf information.
    
    用户可以手动输入小号，也可以选择从一个本地文件读取。读取完成后，也可以选择更新本地的小号信息。<br>Users may choose to manually input smurf information, or load smurfs from a local file. After loading, the user can decide whether to update the local smurf information.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    :param current_puuid: 主召唤师的玩家通用唯一识别码。<br>The main summoner's puuid.
    :type current_puuid: str
    :param infos: 召唤师信息缓存字典。键是玩家通用唯一识别码，值是召唤师信息字典。<br>Summoner information cache dictionary. Each key is a puuid, and each value is a summoner information dictionary.
    :type infos: dict[str, dict[str, Any]]
    :return: 小号信息列表。<br>A list of smurfs.
    :rtype: list[dict[str, Any]]
    '''
    if infos == None:
        infos = {}
    smurfs: list[dict[str, Any]] = []
    logPrint("是否导入其它账号？（输入任意非空字符串以导入，否则不导入。）\nImport other accounts? (Submit any non-empty string to import, or null to refuse importing.)")
    smurfMode_str: str = logInput()
    smurfMode: bool = bool(smurfMode_str)
    if smurfMode:
        smurf_header: dict[str, str] = {"displayName": "显示名", "gameName": "玩家名称", "tagLine": "名称编号", "summonerId": "召唤师序号", "puuid": "玩家通用唯一识别码"}
        smurf_df: pandas.DataFrame = pandas.DataFrame(data = smurf_header, index = [0])
        logPrint("请选择导入方式：\nPlease select an option to import:\n☆1\t读取文件（Read a file）\n2\t手动输入（Manually input）")
        smurf_option: str = logInput()
        if smurf_option != "" and smurf_option[0] == "2":
            smurf_option = "2"
        else:
            smurf_option = "1"
        #在下面的代码中，关键是列表`smurfs`中追加小号信息（The key point of the following code is to append smurf information into the list `smurfs`）
        smurf_file_read: bool = False #标记程序是否成功读取到含有小号信息的数据文件（Marks whether the smurf data file is read successfully）
        smurf_file: str = "Smurf Accounts.json"
        smurf_file_rename: str = "Smurf Accounts (Invalid).json"
        smurf_local: dict[str, dict[str, list[str]]] = {}
        if smurf_option == "1":
            while os.path.exists(smurf_file_rename): #确保下面的重命名操作不会引发报错（Ensure the following renaming operation won't cause an error）
                smurf_file_rename = os.path.splitext(smurf_file_rename)[0] + "(1)." + os.path.splitext(smurf_file_rename)[1]
            if os.path.exists(smurf_file):
                try:
                    with open(smurf_file, "r", encoding = "utf-8") as fp:
                        smurf_local = json.load(fp)
                except json.decoder.JSONDecodeError:
                    os.rename(smurf_file, smurf_file_rename) #上面的while循环保证这里重命名后的文件不可能存在（The above while-loop ensures the result file can't exist）
                    logPrint(f'''在同目录下发现了格式不正确的数据文件。该文件已重命名为“{smurf_file_rename}”。程序将转为手动输入。\nA smurf data file with invalid format is found under the same directory. This file has been renamed into "{smurf_file_rename}". You may need to input the smurfs' names manually.''')
                    smurf_option = "2"
                else:
                    if isinstance(smurf_local, dict) and all(map(lambda x: x in valid_platformIds, smurf_local.keys())) and all(map(lambda x: isinstance(x, dict), smurf_local.values())) and all(len(smurf_local_iter) == 0 or all(map(lambda x: isinstance(x, str) and verify_uuid(x), smurf_local_iter.keys())) and all(map(lambda x: isinstance(x, list) and all(map(lambda y: isinstance(y, str) and verify_uuid(y), x)), smurf_local_iter.values())) for smurf_local_iter in smurf_local.values()): #格式的严格校验（A serious verification of the format）
                        smurf_file_read = True
                        if platformId in smurf_local:
                            if current_puuid in smurf_local[platformId]:
                                count: int = 0 #标识小号的序号（Number the smurfs）
                                valid_puuid_count: int = 0 #记录能查询到玩家的玩家通用唯一识别码的数量（Record the number of puuids that can correspond to players）
                                for smurf_puuid in smurf_local[platformId][current_puuid]:
                                    count += 1
                                    logPrint(f"{count}.\t{smurf_puuid}")
                                    info: dict[str, Any] = await get_info(connection, smurf_puuid)
                                    if info["info_got"]:
                                        info_body: dict[str, Any] = info["body"]
                                        if not info_body["puuid"] in list(map(lambda x: x["puuid"], smurfs)):
                                            valid_puuid_count += 1
                                            logPrint(info_body)
                                            smurfs.append(info_body)
                                            smurf_record: dict[str, Any] = {key: info["body"][key] for key in smurf_header}
                                            smurf_df: pandas.DataFrame = pandas.concat([smurf_df, pandas.DataFrame([smurf_record])], ignore_index = True)
                                            print(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0], end = "\n\n")
                                            log.write(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0] + "\n\n")
                                            infos[info_body["puuid"]] = info_body
                                    else:
                                        logPrint(info["message"])
                                logPrint(f"从离线文件中读取到了{valid_puuid_count}个小号信息。是否继续输入更多小号？（输入任意键以继续添加小号，否则不添加。）\nThe program has detected {valid_puuid_count} smurf account(s) from the local file. Do you want to continue with more smurf accounts? (Submit any non-empty string to continue, or null to refuse adding.)")
                                input_more_smurf_str: str = logInput()
                                smurf_option = "2" if bool(input_more_smurf_str) else "1"
                            else:
                                logPrint("在同目录下发现了含有小号信息的数据文件，但是没有找到您的小号信息。程序将转为手动输入。\nThe smurf data file is found under the same directory, but without yours. You may need to input the smurfs' names manually.")
                                smurf_option = "2"
                        else:
                            logPrint("在同目录下发现了含有小号信息的数据文件，但是没有找到您的大区信息。如果您确认您的本地文件没有问题，请向作者反馈该问题。\nThe smurf data file is found under the same directory, but without your server's. If you're sure that there's not any problem in your local data file, please file the feedback to the author.\n一个可用的反馈链接：\nAn available feedback link:\nhttps://github.com/WordlessMeteor/LoL-DIY-Programs/issues/new \n程序将转为手动输入。\nYou may need to input the smurfs' names manually.")
                            smurf_option = "2"
                    else:
                        os.rename(smurf_file, smurf_file_rename)
                        logPrint(f'''在同目录下发现了格式不正确的数据文件。该文件已重命名为“{smurf_file_rename}”。程序将转为手动输入。\nA smurf data file with invalid format is found under the same directory. This file has been renamed into "{smurf_file_rename}". You may need to input the smurfs' names manually.''')
                        smurf_option = "2"
            else:
                logPrint("没有找到含有小号信息的数据文件。程序将转为手动输入。\nSmurf data file not found. You may need to input the smurfs' names manually.")
                smurf_option = "2"
        if smurf_option == "2":
            logPrint('请输入小号的召唤师名。输入“0”以清空已经输入的小号。输入-1以结束。\nPlease input the summoner names of the smurf accounts. Submit "0" to clear the entered smurfs. Submit "-1" to finish the importation.')
            while True:
                smurfName: str = logInput()
                if smurfName == "-1":
                    break
                elif smurfName == "0":
                    smurfs = []
                    smurf_df = pandas.DataFrame(data = smurf_header, index = [0])
                    logPrint("已清空小号。\nSmurfs cleared.")
                elif smurfName == "":
                    continue
                else:
                    info: dict[str, Any] = await get_info(connection, smurfName)
                    if info["info_got"]:
                        info_body: dict[str, Any] = info["body"]
                        if info_body["puuid"] == current_puuid:
                            logPrint("您不能把主账号作为小号！请添加其它账号。\nYou're not allowed to add your main account as a smurf account! Please try another account.")
                        elif info_body["puuid"] in list(map(lambda x: x["puuid"], smurfs)):
                            logPrint("您已经输入过该玩家了。\nYou've entered this player.")
                        else:
                            logPrint(info_body)
                            smurfs.append(info_body)
                            smurf_record: dict[str, Any] = {key: info["body"][key] for key in smurf_header}
                            smurf_df: pandas.DataFrame = pandas.concat([smurf_df, pandas.DataFrame([smurf_record])], ignore_index = True)
                            print(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0])
                            log.write(format_df(smurf_df, width_exceed_ask = False, direct_print = False, print_index = True)[0] + "\n")
                            infos[info_body["puuid"]] = info_body
                    else:
                        logPrint(info["message"])
        logPrint("是否需要将小号信息保存到本地，以便日后直接读取文件来添加小号？（输入任意键以确认，否则不保存。）\nDo you want to save the smurf information into a local data file, so that you may read this file directly to load the smurfs? (Submit any non-empty string to confirm, or null to refuse saving.)\n注意：程序会覆盖之前的小号信息，因此如果您不想丢失以前的小号信息，请在不输入任何字符的情况下直接按回车键不保存，然后将原来的小号信息做好备份。\nNote: The old smurf information will be overwritten, so if you expect the previous smurf information not to be lost, please directly press Enter without any other characters entered, and then make a backup of the original smurf information.")
        save_smurf_str: str = logInput()
        save_smurf: bool = bool(save_smurf_str)
        if save_smurf:
            if smurf_file_read:
                if platformId in smurf_local:
                    smurf_local[platformId][current_puuid] = list(map(lambda x: x["puuid"], smurfs)) #之所以考虑用玩家通用唯一识别码，而不用召唤师名或者召唤师序号作为小号信息存储介质的原因有两个方面的考量：从对人类友好的角度上，召唤师名的确更胜一筹，但是缺少唯一性。在调用get_info函数时，两个召唤师名如果只是差几个空格，就很有可能指向同一个召唤师。这样，上面和下面的代码在识别召唤师信息是否添加过时，就不太好实现；从存储格式的角度上来考虑，玩家通用唯一识别码服从通用唯一识别码的格式，相对比较统一，而且是全球统一的，这样在校验数据文件格式时比较方便。而召唤师序号只是整数，而且不同召唤师序号存在长短不一的情况，这样校验起来不够充分（The reason why I consider using puuid as the smurf data storing media, instead of the summoner name or summonerId, has two considerations. On the one hand, in terms of being human-friendly, a summoner name does far outweigh the puuid or summonerId. However, it lacks uniformity. When `get_info` function is called, if two parameters differ only in several spaces, the result might directs to a same summoner. In that case, it's not easy to implement the code to identify whether a summoner's information has been added to the list before, within the context. On the other hand, in terms of the format, a puuid obeys the format of uuids, so it's relatively general, let alone being "universally unqiue", which makes it convenient to verify the format of the smurf data file. Nevertheless, summonerId is just an integer, and different summonerIds may be of different lengths, so it's not sufficient to determine a summoner by summonerId）
                else:
                    smurf_local[platformId] = {current_puuid: list(map(lambda x: x["puuid"], smurfs))}
            else:
                smurf_local: dict[str, dict[str, list[str]]] = {platformId: {current_puuid: list(map(lambda x: x["puuid"], smurfs))}}
            with open(smurf_file, "w", encoding = "utf-8") as fp:
                json.dump(smurf_local, fp, indent = 4, ensure_ascii = False)
            logPrint(f"小号信息已保存到“{smurf_file}”中。\nSmurf information has been saved into {smurf_file}.")
    return smurfs

async def sort_basic_info(connection: Connection, puuid: str, remove_empty: bool = True) -> pandas.DataFrame:
    '''
    将**一名**召唤师的个人档案整理成一个表格。<br>Sort the personal profile of **a** summoner into a dataframe.
    
    有关个人档案包括的内容，请参阅召唤师模块中的`sort_summoner_info`函数。<br>To get the details of a summoner's profile, please refer to `sort_summoner_info` function in `src.utils.summoner` module.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    :param puuid: 要查询的召唤师的玩家通用唯一识别码。<br>The puuid of the player to query.
    :type puuid: str
    :param remove_empty: 是否从表格中移除内容为空的字段。默认为真。<br>Whether to remove empty fields from the table. True by default.
    :type remove_empty: bool
    :return: 召唤师档案数据框。<br>Summoner profile dataframe.
    :rtype: pandas.DataFrame
    '''
    info: dict[str, Any] = await get_info(connection, puuid)
    info_body: dict[str, Any] = info["body"]
    displayName: str = get_info_name(info_body)
    #先准备一些局部数据资源（First, prepare some local data resources）
    ##召唤师图标（Summoner icon）
    summonerIcons_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/summoner-icons.json")).json()
    summonerIcons: dict[int, dict[str, Any]] = {int(summonerIcon_iter["id"]): summonerIcon_iter for summonerIcon_iter in summonerIcons_source}
    ##英雄（LoL champion）
    LoLChampions_source: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-data/assets/v1/champion-summary.json")).json()
    LoLChampions: dict[int, dict[str, Any]] = {int(LoLChampion_iter["id"]): LoLChampion_iter for LoLChampion_iter in LoLChampions_source}
    ##旗帜（Regalia banner）
    regaliaBanners = await (await connection.request("GET", "/lol-regalia/v3/inventory/REGALIA_BANNER")).json()
    ##排位（Ranked）
    ranked: dict[str, Any] = await (await connection.request("GET", f"/lol-ranked/v1/ranked-stats/{puuid}")).json()
    if "errorCode" in ranked: #很久以前，国服体验服的排位数据API未知。现在已经与正式服统一（Long ago, API of ranked stats on Chinese PBE was unknown. Now it accords with Live servers）
        logPrint(ranked)
        logPrint("玩家%s的排位信息获取失败。\nRanked information of Player %s capture failed." %(displayName, displayName))
    ##冠军杯赛旗帜（Tournament flag）
    banner: dict[str, int | str] = await (await connection.request("GET", f"/lol-banners/v1/players/{puuid}/flags/equipped")).json()
    if isinstance(banner, dict) and "errorCode" in banner:
        logPrint(banner)
        if banner == {"errorCode": "RPC_ERROR", "httpStatus": 404, "implementationDetails": {}, "message": f"{puuid} has no valid banner flags."}:
            logPrint("玩家%s未装备冠军杯赛旗帜。\nPlayer %s doesn't equip any tournament flag." %(displayName, displayName))
        else:
            logPrint("玩家%s的冠军杯赛旗帜获取失败。\nTournament flag information of Player %s capture failed." %(displayName, displayName))
    ##成就（Challenge）
    challenge: dict[str, Any] = await (await connection.request("GET", f"/lol-challenges/v1/summary-player-data/player/{puuid}")).json()
    if "errorCode" in challenge:
        logPrint(challenge)
        logPrint("玩家%s的成就信息获取失败。\nChallenge information of Player %s capture failed." %(displayName, displayName))
    ##永恒星碑（Statstone）
    topStatstones: list[dict[str, Any]] = await (await connection.request("GET", f"/lol-statstones/v1/profile-summary/{puuid}")).json()
    if isinstance(topStatstones, dict) and "errorCode" in topStatstones:
        logPrint(topStatstones)
        logPrint("玩家%s的最高永恒星碑信息获取失败。\nTop statstone information of Player %s capture failed." %(displayName, displayName))
    #然后整理数据（Then organize data）
    unmapped_keys: dict[str, set[Any]] = {"summonerIcon": set(), "regaliaBanner": set(), "LoLChampion": set()}
    profile_header_keys: list[str] = list(profile_header.keys())
    profile_data: dict[str, list[Any]] = {"项目": [], "Items": [], "值": []}
    for i in range(len(profile_header_keys)):
        key: str = profile_header_keys[i]
        profile_data["项目"].append(profile_header[key])
        profile_data["Items"].append(key)
        if i <= 21: #召唤师信息（Summoner information）
            if i <= 16:
                if i >= 15: #召唤师图标相关键（Profile icon related keys）
                    profileIconId: int = info_body["profileIconId"]
                    if profileIconId in summonerIcons:
                        value: Any = summonerIcons[profileIconId].get(key.split("_")[1], "") #部分召唤师图标没有名称键（Some summoner icons don't have a "name" key）
                    else:
                        if not profileIconId in unmapped_keys["summonerIcon"]:
                            unmapped_keys["summonerIcon"].add(profileIconId)
                            logPrint("玩家%s的召唤师图标（%d）信息获取失败。\nSummoner icon information (%d) of Player %s capture failed." %(displayName, profileIconId, profileIconId, displayName))
                        value = ""
                else:
                    value = info_body[key]
            else: #重随点子键（`rerollPoints`' subkeys）
                value = info_body["rerollPoints"][key]
        elif i <= 28: #段位（Rank）
            if "errorCode" in ranked:
                value = ""
            else:
                if key in ranked:
                    if i in {24, 25, 27}:
                        value = tiers[ranked[key]]
                    else:
                        value = ranked[key]
                else:
                    value = ""
        elif i <= 33: #冠军杯赛旗帜（Tournament flag）
            if isinstance(banner, dict) and "errorCode" in banner:
                value = ""
            else:
                value = banner[key.split()[1]]
        elif i <= 90: #成就和头衔（Challenge and title）
            if "errorCode" in challenge:
                value = ""
            else:
                if i == 36 or i == 37: #整数字符串（Integer string）
                    value = "" if challenge[key.split()[1]] == "" else int(challenge[key.split()[1]])
                elif i == 39: #总成就等级（`challenge overallChallengeLevel`）
                    value = challengeCrystalLevels[challenge["overallChallengeLevel"]]
                elif i == 44: #天梯更新时间（`challenge apexLadderUpdateDate`）
                    value = getISOTime(challenge["apexLadderUpdateTime"] / 1000)
                elif i == 45 or i == 46: #身份旗帜子键（Info banner's subkeys）
                    bannerId: str = challenge["bannerId"]
                    if bannerId in regaliaBanners:
                        value = regaliaBanners[bannerId]["items"][0][key.split("_")[1]]
                    else:
                        if not bannerId in unmapped_keys["regaliaBanner"]:
                            unmapped_keys["regaliaBanner"].add(bannerId)
                            if bannerId == "":
                                logPrint("玩家%s未装备身份旗帜。\nPlayer %s doesn't equip any info banner." %(displayName, displayName))
                            else:
                                logPrint("玩家%s的身份旗帜（%s）信息获取失败。\nInfo banner information (%s) of Player %s capture failed." %(displayName, bannerId, bannerId, displayName))
                        value = ""
                elif i == 47: #排位徽章名称（`challenge crestName`）
                    value = "" #目前尚不明确排位徽章的本地化内容（Localized content of crests haven't figured out yet）
                elif i >= 48 and i <= 72: #分类进度子键（`categoryProgress`' subkeys）
                    challengeCategoryIndex: int = (i - 48) // 5
                    subIndex: int = (i - 48) % 5
                    subkey: str = key.split()[2]
                    if challengeCategoryIndex < len(challenge["categoryProgress"]):
                        if subIndex == 0: #名称（`category`）
                            value = challengeCategories[challenge["categoryProgress"][challengeCategoryIndex]["category"]]
                        elif subIndex == 2: #等级（`level`）
                            value = challengeCrystalLevels[challenge["categoryProgress"][challengeCategoryIndex]["level"]]
                        else:
                            value = challenge["categoryProgress"][challengeCategoryIndex][subkey]
                    else:
                        value = ""
                elif i >= 73 and i <= 84: #最佳成就子键（`topChallenges`' subkeys）
                    topChallengeIndex: int = (i - 73) // 4
                    subIndex: int = (i - 73) % 4
                    subkey = key.split()[1]
                    if topChallengeIndex < len(challenge["topChallenges"]):
                        if subIndex == 2: #当前等级（`currentLevel`）
                            value = challengeCrystalLevels[challenge["topChallenges"][topChallengeIndex]["currentLevel"]]
                        else:
                            value = challenge["topChallenges"][topChallengeIndex][subkey]
                    else:
                        value = ""
                elif i >= 85 and i <= 87: #头衔子键（`title`'s subkeys）
                    if i == 87: #头衔等级（`challenge title titleAcquisitionType`）
                        value = titleAcquisitionTypes[challenge["title"]["titleAcquisitionType"]]
                    else:
                        value = challenge["title"][key.split()[2]]
                elif i >= 88: #头衔成就数据子键（`challengeTitleData`'s subkeys）
                    if challenge["title"]["challengeTitleData"] == None:
                        value = ""
                    else:
                        if i == 90: #头衔成就等级（`challenge title challengeTitleData level`）
                            value = challengeCrystalLevels[challenge["title"]["challengeTitleData"]["level"]]
                        else:
                            value = challenge["title"]["challengeTitleData"][key.split()[3]]
                        if value == None:
                            value = ""
                else:
                    value = challenge[key.split()[1]]
        else: #最高永恒星碑相关键（Top statstone related keys）
            if isinstance(topStatstones, dict) and "errorCode" in topStatstones:
                value = ""
            else:
                topStatstoneIndex: int = (i - 91) // 8
                subIndex: int = (i - 91) % 8
                subkey: str = key.split()[1]
                if topStatstoneIndex < len(topStatstones):
                    if subIndex >= 5: #英雄子键（Champion's subkeys）
                        championId: int = topStatstones[topStatstoneIndex]["championId"]
                        if championId in LoLChampions:
                            value = LoLChampions[championId][key.split("_")[1]]
                        else:
                            if not championId in unmapped_keys["LoLChampion"]:
                                unmapped_keys["LoLChampion"].add(championId)
                                logPrint("玩家%s的第%d永恒星碑的英雄信息（%d）获取失败。\nStatstone No. %d champion information (%d) of Player %s capture failed." %(displayName, topStatstoneIndex + 1, championId, topStatstoneIndex + 1, championId, displayName))
                            value = ""
                    else:
                        value = topStatstones[topStatstoneIndex][subkey]
                else:
                    value = ""
        profile_data["值"].append(value)
    profile_statistics_output_order: list[int] = [0, 1, 3, 2, 11, 12, 4, 9, 8, 6, 7, 15, 16, 10, 13, 14, 5, 19, 18, 17, 20, 21, 24, 22, 25, 26, 27, 28, 23, 30, 33, 32, 31, 29, 38, 34, 44, 35, 36, 45, 46, 37, 47, 42, 43, 39, 40, 41, 85, 86, 87, 88, 89, 90, 48, 50, 49, 51, 52, 53, 55, 54, 56, 57, 58, 60, 59, 61, 62, 63, 65, 64, 66, 67, 68, 70, 69, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 94, 91, 95, 93, 92, 96, 97, 98, 102, 99, 103, 101, 100, 104, 105, 106, 110, 107, 111, 109, 108, 112, 113, 114]
    profile_data_organized: dict[str, list[Any]] = {"项目": [], "Items": [], "值": []}
    for i in profile_statistics_output_order:
        key: str = profile_header_keys[i]
        value: Any = profile_data["值"][i]
        profile_data_organized["项目"].append(profile_header[key])
        profile_data_organized["Items"].append(key)
        profile_data_organized["值"].append(value)
    info_df: pandas.DataFrame = pandas.DataFrame(data = profile_data_organized)
    if remove_empty:
        info_df = info_df[info_df["值"] != ""]
    return info_df

async def sort_champion_mastery(connection: Connection, puuid: str, LoLChampions: dict[int, dict[str, Any]], unmapped_keys: Optional[dict[str, set[Any]]] = None) -> pandas.DataFrame:
    '''
    将一名玩家的英雄成就整理成一张表格。<br>Sort the champion mastery of a player into a table.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    :param puuid: 要查询的召唤师的玩家通用唯一识别码。<br>The puuid of the player to query.
    :type puuid: str
    :param LoLChampions: 整理后的英雄数据资源。键是英雄序号，值是英雄信息字典。<br>Organized champion data resource. Each key is a championId, and each value is a champion information dictionary.
    
        原始英雄数据资源可通过以下链接获取：<br>The raw champion data resource can be obtained through the following links:
        - https://raw.communitydragon.org/pbe/plugins/rcp-be-lol-game-data/global/default/v1/champion-summary.json
        - https://raw.communitydragon.org/pbe/plugins/rcp-be-lol-game-data/global/default/v1/champions/{championId}.json
        
        也可以通过以下LCU接口获取：<br>It can also be obtained from the following LCU endpoints:
        - `GET /lol-game-data/assets/v1/champion-summary.json`
        - `GET /lol-game-data/assets/v1/champions/{championId}.json`
        - `GET /lol-champions/v1/inventories/{summonerId}/champions`
    :type LoLChampions: dict[int, dict[str, Any]]
    :param unmapped_keys: 各数据资源未找到的键。用于控制数据未找到匹配记录的提示最多输出一次。<br>Unmapped keys in all data resources, used to control the hint about data not found to be printed at most once.
    :type unmapped_keys: dict[str, set[int]]
    :return: 英雄成就数据框。<br>Champion mastery dataframe.
    '''
    if unmapped_keys == None:
        unmapped_keys = {"LoLChampion": set()}
    mastery: list[dict[str, Any]] = await (await connection.request("GET", f"/lol-champion-mastery/v1/{puuid}/champion-mastery")).json()
    mastery_header_keys: list[str] = list(mastery_header.keys())
    mastery_data: dict[str, list[Any]] = {key: [] for key in mastery_header_keys}
    for mastery_iter in mastery:
        for i in range(len(mastery_header)):
            key: str = mastery_header_keys[i]
            if i <= 16:
                if i == 6: #已赚取海克斯宝箱（`chestGranted`）
                    to_append: Any = mastery_iter.get("chestGranted", False) #外服的英雄成就接口中没有“chestGranted”这个键（Champion mastery API in Riot servers don't include the key "chestGranted"）
                elif i >= 13 and i <= 15: #英雄（`champion`）
                    subkey: str = "name" if i == 13 else "alias" if i == 14 else "squarePortraitPath"
                    championId: int = mastery_iter["championId"]
                    if championId in LoLChampions:
                        if i == 13 or i == 14:
                            to_append = LoLChampions[championId][subkey] if championId in LoLChampions else "" #在2026年4月22日，测试服的客户端数据中删除了堕落天使 莫甘娜的数据（On Apr. 22nd, 2026, Morgana was deleted from PBE League Client plugins data）
                        else: #英雄方块图像（`championSquarePortrait`）
                            to_append = urljoin(connection.address, LoLChampions[championId]["squarePortraitPath"]) if championId in LoLChampions else ""
                    else:
                        if not championId in unmapped_keys["LoLChampion"]:
                            unmapped_keys["LoLChampion"].add(championId)
                            logPrint("英雄成就（%d）获取失败。\nChampion mastery (%d) capture failed." %(championId, championId))
                        to_append = ""
                elif i == 16: #上次使用时间（`lastPlayDate`） #这里需要将时间戳转换为标准格式的时间（Here the timestamp is going to be converted into time in standard format）
                    to_append = getISOTime(mastery_iter["lastPlayTime"] / 1000) #英雄联盟中的时间戳精确到毫秒，也就是放大了1000倍（Timestamps in LCU API are accurate to milliseconds, namely multiplied by 1000）
                else:
                    to_append = mastery_iter[key]
            elif i >= 17 and i <= 19:
                if i == 17: #已达到Ⅳ级里程碑（`nextSeasonMilestoneBonus`）
                    to_append = mastery_iter["nextSeasonMilestone"]["bonus"]
                elif i == 18: #下个里程点所需对局评价（`nextSeasonMilestoneRequireGrade`）
                    to_append = mastery_iter["nextSeasonMilestone"]["requireGradeCounts"]
                else: #下个里程点奖励英雄成就标记个数（`nextSeasonMilestoneRewardMarks`）
                    to_append = mastery_iter["nextSeasonMilestone"]["rewardMarks"]
            else:
                if i == 20: #下个里程点最大奖励次数（`nextSeasonMilestoneMaximumReward`）
                    to_append = mastery_iter["nextSeasonMilestone"]["rewardConfig"]["maximumReward"]
                else: #下个里程点奖励物品序号（`nextSeasonMilestoneRewardValue`）
                    to_append = mastery_iter["nextSeasonMilestone"]["rewardConfig"]["rewardValue"]
            mastery_data[key].append(to_append)
    mastery_statistics_output_order: list[int] = [13, 14, 1, 2, 3, 4, 9, 12, 6, 7, 5, 17, 10, 18, 19, 20, 21, 16]
    mastery_data_organized: dict[str, list[Any]] = {mastery_header_keys[i]: mastery_data[mastery_header_keys[i]] for i in mastery_statistics_output_order}
    mastery_df: pandas.DataFrame = pandas.DataFrame(data = mastery_data_organized)
    optimize_bool_display(mastery_df)
    mastery_df = pandas.concat([pandas.DataFrame([mastery_header])[mastery_df.columns], mastery_df], ignore_index = True)
    return mastery_df
    mastery_web_display_order: list[int] = [15, 1, 2, 3, 4, 9, 12, 6, 7, 5, 17, 10, 18, 19, 20, 21, 16]
    mastery_data_organized_web: dict[str, list[Any]] = {mastery_header_keys[i]: mastery_data[mastery_header_keys[i]] for i in mastery_web_display_order}
    mastery_df_web: pandas.DataFrame = pandas.DataFrame(data = mastery_data_organized_web)
    optimize_bool_display(mastery_df_web)
    mastery_df_web = pandas.concat([pandas.DataFrame([mastery_header])[mastery_df_web.columns], mastery_df_web], ignore_index = True)
    mastery_htmltable: str = mastery_df_web.to_html(escape = False)

async def sort_ranked_data(connection: Connection, puuid: str) -> pandas.DataFrame:
    '''
    将一名玩家的排位数据整理成一张表格。<br>Organize a player's ranked data into a table.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    :param puuid: 要查询的召唤师的玩家通用唯一识别码。<br>The puuid of the player to query.
    :type puuid: str
    :return: 排位数据框。<br>Ranked dataframe.
    :rtype: pandas.DataFrame
    '''
    ranked: dict[str, Any] = await (await connection.request("GET", f"/lol-ranked/v1/ranked-stats/{puuid}")).json()
    ranked_header_keys: list[str] = list(ranked_header.keys())
    ranked_data: dict[str, list[Any]] = {key: [] for key in ranked_header_keys}
    for queue in ranked["queues"]:
        for i in range(len(ranked_header_keys)):
            key: str = ranked_header_keys[i]
            if i in {2, 3, 9, 11}: #段位分级相关键（Division-related keys）
                to_append: Any = "" if queue[key] == "NA" else queue[key]
            elif i in {4, 10, 12, 19}: #段位相关键（Tier-related keys）
                to_append = tiers[queue[key]]
            elif i == 16: #对局类型（`queueType`）
                to_append = queueTypes_ranked[queue[key]]
            elif i == 18: #云顶之弈狂暴模式段位（`ratedTier`）
                to_append = ratedTiers[queue[key]]
            elif i == 22 or i == 23:
                if i == 22: #综合段位（`tier / ratedTier`）
                    to_append = ratedTiers[queue["ratedTier"]] if queue["queueType"] == "RANKED_TFT_TURBO" else tiers[queue["tier"]]
                else: #综合胜点（`leaguePoints / ratedRating`）
                    to_append = queue["ratedRating"] if queue["queueType"] == "RANKED_TFT_TURBO" else queue["leaguePoints"]
            else:
                to_append = queue[key]
            ranked_data[key].append(to_append)
    ranked_statistics_output_order: list[int] = [16, 22, 2, 23, 21, 7, 5, 0, 14, 15, 8, 4, 3, 1, 10, 9, 12, 11, 13, 20]
    ranked_data_organized: dict[str, list[Any]] = {ranked_header_keys[i]: ranked_data[ranked_header_keys[i]] for i in ranked_statistics_output_order}
    ranked_df: pandas.DataFrame = pandas.DataFrame(data = ranked_data_organized)
    optimize_bool_display(ranked_df)
    ranked_df = pandas.concat([pandas.DataFrame([ranked_header])[ranked_df.columns], ranked_df], ignore_index = True)
    return ranked_df

async def sort_ranked_ladders(connection: Connection, puuid: str, fetch_summoner_info: bool = True) -> pandas.DataFrame:
    '''
    将一名玩家的排位天梯信息整理成一张表格。<br>Organize a player's ranked ladder data into a table.
    
    :param connection: 通过lcu-driver库创建的用于访问LCU API的连接对象。<br>A Connection object created through lcu-driver library, meant to access LCU API.
    :type connection: Connection
    :param puuid: 要查询的召唤师的玩家通用唯一识别码。<br>The puuid of the player to query.
    :type puuid: str
    :param fetch_summoner_info: 是否获取每名天梯玩家的召唤师信息。默认为真。<br>Whether to get the summoner information of each apex player. True by default.
    :type fetch_summoner_info: bool
    :return: 排位天梯数据框。<br>Ranked ladder dataframe.
    :rtype: pandas.DataFrame
    '''
    ladders: list[dict[str, Any]] = await (await connection.request("GET", f"/lol-ranked/v1/league-ladders/{puuid}")).json()
    # ladder_summoner_infos: dict[str, dict[str, Any]] = await get_infos(connection, puuids = [standing["puuid"] for ladder in ladders for division in ladder["divisions"] for standing in division["standings"]]) if fetch_summoner_info else {}
    ladder_header_keys: list[str] = list(ladder_header.keys())
    ladder_data: dict[str, list[Any]] = {key: [] for key in ladder_header_keys}
    for i in range(len(ladders)):
        ladder: dict[str, Any] = ladders[i]
        logPrint("顶级%s%s玩家信息整理进度（Top %s %s player information organization process）：" %(queueTypes_ranked[ladder["queueType"]], tiers_all[ladder["tier"]], ladder["queueType"], ladder["tier"]))
        for j in range(len(ladder["divisions"])):
            division: dict[str, Any] = ladder["divisions"][j]
            for k in range(len(division["standings"])):
                standing: dict[str, Any] = division["standings"][k]
                #准备召唤师信息（Prepare summoner information）
                if fetch_summoner_info:
                    # info_got: bool = standing["puuid"] in ladder_summoner_infos
                    # standing_summoner: dict[str, Any] = ladder_summoner_infos.get(standing["puuid"], {})
                    standing_summoner_recapture: int = 0
                    standing_summoner: dict[str, Any] = await get_info(connection, standing["puuid"])
                    while not standing_summoner["info_got"] and standing_summoner["body"]["httpStatus"] != 404 and standing_summoner_recapture < 3:
                        logPrint(standing_summoner["message"])
                        standing_summoner_recapture += 1
                        logPrint("顶级%s%s玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of top %s %s player (puuid: %s) capture failed! Recapturing this player's information ... Times tried: %d" %(queueTypes_ranked[ladder["queueType"]], tiers_all[ladder["tier"]], standing["puuid"], standing_summoner_recapture, ladder["queueType"], ladder["tier"], standing["puuid"], standing_summoner_recapture))
                        standing_summoner = await get_info(connection, standing["puuid"])
                    info_got: bool = standing_summoner["info_got"]
                    if not info_got:
                        logPrint(standing_summoner["message"])
                        logPrint("顶级%s%s玩家信息（玩家通用唯一识别码：%s）获取失败！\nInformation of top %s %s player (puuid: %s) capture failed!" %(queueTypes_ranked[ladder["queueType"]], tiers_all[ladder["tier"]], standing["puuid"], ladder["queueType"], ladder["tier"], standing["puuid"]))
                else:
                    info_got = False
                #整理数据（Organize data）
                for l in range(len(ladder_header_keys)):
                    key = ladder_header_keys[l]
                    if l == 0:
                        to_append: Any = queueTypes_ranked[ladder["queueType"]]
                    elif l <= 20:
                        if l == 1 or l == 12:
                            to_append = "" if standing[key] == "NA" else standing[key]
                        elif l == 13 or l == 19:
                            to_append = tiers[standing[key]]
                        else:
                            to_append = standing[key]
                    elif l <= 22:
                        to_append = standing_summoner["body"][key] if info_got else ""
                    else:
                        to_append = "☆" if standing["puuid"] == puuid else ""
                    ladder_data[key].append(to_append)
                logPrint("[%d/%d][%d/%d][%d/%d]%s\t%s" %(i + 1, len(ladders), j + 1, len(ladder["divisions"]), k + 1, len(division["standings"]), standing["puuid"], get_info_name(standing_summoner["body"]) if info_got else ""), end = "\r")
        else:
            logPrint("已完成。 | Done.")
    ladder_statistics_output_order: list[int] = [0, 9, 11, 10, 17, 15, 18, 21, 22, 19, 1, 4, 3, 14, 8, 7, 6, 20, 5, 13, 12, 2, 16, 23]
    # ladder_web_display_order: list[int] = [0, 9, 11, 10, 17, 15, 18, 21, 22, 19, 1, 4, 3, 14, 8, 7, 6, 20, 5, 13, 12, 2, 16, 23]
    ladder_data_organized: dict[str, list[Any]] = {ladder_header_keys[i]: ladder_data[ladder_header_keys[i]] for i in ladder_statistics_output_order}
    ladder_df: pandas.DataFrame = pandas.DataFrame(data = ladder_data_organized)
    optimize_bool_display(ladder_df)
    ladder_df = pandas.concat([pandas.DataFrame([ladder_header])[ladder_df.columns], ladder_df], ignore_index = True)
    return ladder_df

def analyze_recently_played_summoners(search_LoL: bool, search_TFT: bool, recent_LoLPlayer_df: pandas.DataFrame, recent_TFTPlayer_df: pandas.DataFrame, gameQueues: dict[int, dict[str, Any]], displayName: str, export_folder: str) -> None:
    '''
    生成近期一起玩过的玩家图表和元数据工作簿。<br>Generate graphs and a metadata workbook of recently played summoners.
    
    :param search_LoL: 是否搜索过英雄联盟对局记录。<br>Whether LoL match history has been searched.
    :type search_LoL: bool
    :param search_TFT: 是否搜索过云顶之弈对局记录。<br>Whether TFT match history has been searched.
    :type search_TFT: bool
    :param recent_LoLPlayer_df: 近期一起玩过的英雄联盟玩家数据框。<br>Recently played LoL summoner dataframe.
    :type recent_LoLPlayer_df: pandas.DataFrame
    :param recent_TFTPlayer_df: 近期一起玩过的英雄联盟玩家数据框。<br>Recently played LoL summoner dataframe.
    :type recent_TFTPlayer_df: pandas.DataFrame
    :param gameQueues: 整理后的队列数据资源。键是队列序号，值是游戏模式信息字典。<br>Organized queue data resource. Each key is a queueId, and each value is a game mode information dictionary.
    
        原始队列数据资源可以通过以下LCU接口获取：<br>The raw queue data resource can be obtained through the following LCU endpoint:
        - `GET /lol-game-queues/v1/queues`
    :type gameQueues: dict[int, dict[str, Any]]
    :param displayName: 用于图表和工作簿命名的召唤师名称。<br>The summoner name used as a part of the workbook's and graph's name.
    :type displayName: str
    :param export_folder: 图表和工作簿的导出目录。<br>The export directory of the graphs and workbooks.
    :type export_folder: str
    '''
    recent_players_metadata: dict[str, dict[str, Any]] = {} #这里另外设置元数据是为了整理出用于可视化的数据（Here the metadata is designed to sort out data for visualization）
    if search_LoL:
        #logPrint("用于可视化的元数据创建进度（Creating process of metadata for visualization）：")
        for i in range(1, len(recent_LoLPlayer_df)): #第0行是中文表头，所以要从第1行开始（The 0th line contains the Chinese headers, so the iteration should start from the first line）
            puuid_iter: str = recent_LoLPlayer_df["puuid"][i]
            if use_sgp:
                summonerName_iter: str = recent_LoLPlayer_df["riotIdGameName"][i] + "#" + recent_LoLPlayer_df["riotIdTagline"][i]
            else:
                summonerName_iter: str = recent_LoLPlayer_df["gameName"][i] + "#" + recent_LoLPlayer_df["tagLine"][i]
            if summonerName_iter == "#":
                summonerName_iter = puuid_iter
            matchId_iter: int = recent_LoLPlayer_df["gameId"][i]
            LoLGameCreation_iter: float = datetime.datetime.fromisoformat(recent_LoLPlayer_df["gameCreationDate"][i].replace("Z", "+00:00")).timestamp()
            LoLGameDuration_iter: float = recent_LoLPlayer_df["gameDuration"][i]
            isPvP_iter: bool = True if recent_LoLPlayer_df["queueId"][i] in gameQueues and gameQueues[recent_LoLPlayer_df["queueId"][i]]["category"] == "PvP" else False #添加是否玩家对战的信息，以便单独统计一同进行玩家对战的总时间。下同（Added the information whether a match is PvP, so that the total time of only PvP matches can be calculated. So do the following two variables）
            isPvE_iter: bool = True if recent_LoLPlayer_df["queueId"][i] in gameQueues and gameQueues[recent_LoLPlayer_df["queueId"][i]]["category"] == "VersusAi" else False
            isCustom_iter: bool = True if recent_LoLPlayer_df["queueId"][i] in gameQueues and gameQueues[recent_LoLPlayer_df["queueId"][i]]["category"] == "Custom" else False
            if not puuid_iter in recent_players_metadata:
                recent_players_metadata[puuid_iter] = {}
                recent_players_metadata[puuid_iter]["name"] = summonerName_iter #该语句不会在else部分出现。这是考虑到如果召唤师改过名字，那么呈现在频数直方图上的横轴的召唤师名应当是最新的（This statement won't appear in the else-part, considering if a summoner has changed its name, then the summonerName near the horizontal axis of the frequency histogram should be latest）
                recent_players_metadata[puuid_iter]["puuid"] = puuid_iter
                recent_players_metadata[puuid_iter]["gameCount"] = 1
                recent_players_metadata[puuid_iter]["matches"] = [matchId_iter]
                recent_players_metadata[puuid_iter]["createTimestamps"] = [LoLGameCreation_iter]
                recent_players_metadata[puuid_iter]["durations"] = [LoLGameDuration_iter]
                recent_players_metadata[puuid_iter]["isPvP"] = [isPvP_iter]
                recent_players_metadata[puuid_iter]["isPvE"] = [isPvE_iter]
                recent_players_metadata[puuid_iter]["isCustom"] = [isCustom_iter]
                recent_players_metadata[puuid_iter]["PvPCount"] = int(isPvP_iter)
                recent_players_metadata[puuid_iter]["PvECount"] = int(isPvE_iter)
                recent_players_metadata[puuid_iter]["CustomCount"] = int(isCustom_iter)
                recent_players_metadata[puuid_iter]["totalTime"] = LoLGameDuration_iter
                recent_players_metadata[puuid_iter]["totalPvPTime"] = LoLGameDuration_iter * isPvP_iter
                recent_players_metadata[puuid_iter]["totalPvETime"] = LoLGameDuration_iter * isPvE_iter
                recent_players_metadata[puuid_iter]["totalCustomTime"] = LoLGameDuration_iter * isCustom_iter
            else:
                recent_players_metadata[puuid_iter]["gameCount"] += 1
                recent_players_metadata[puuid_iter]["matches"].append(matchId_iter)
                recent_players_metadata[puuid_iter]["createTimestamps"].append(LoLGameCreation_iter)
                recent_players_metadata[puuid_iter]["durations"].append(LoLGameDuration_iter)
                recent_players_metadata[puuid_iter]["isPvP"].append(isPvP_iter)
                recent_players_metadata[puuid_iter]["isPvE"].append(isPvE_iter)
                recent_players_metadata[puuid_iter]["isCustom"].append(isCustom_iter)
                recent_players_metadata[puuid_iter]["PvPCount"] += isPvP_iter
                recent_players_metadata[puuid_iter]["PvECount"] += isPvE_iter
                recent_players_metadata[puuid_iter]["CustomCount"] += isCustom_iter
                recent_players_metadata[puuid_iter]["totalTime"] += LoLGameDuration_iter
                recent_players_metadata[puuid_iter]["totalPvPTime"] += LoLGameDuration_iter * isPvP_iter
                recent_players_metadata[puuid_iter]["totalPvETime"] += LoLGameDuration_iter * isPvE_iter
                recent_players_metadata[puuid_iter]["totalCustomTime"] += LoLGameDuration_iter * isCustom_iter
            #logPrint("[%d/%d]%d\t%s\t%s" %(i, len(recent_LoLPlayer_df) - 1, matchId_iter, puuid_iter, summonerName_iter), end = "\r")
    if search_TFT:
        #logPrint("用于可视化的元数据创建进度（Creating process of metadata for visualization）：")
        for i in range(1, len(recent_TFTPlayer_df)):
            puuid_iter = recent_TFTPlayer_df["puuid"][i]
            summonerName_iter = recent_TFTPlayer_df["riotIdGameName"][i] + "#" + recent_TFTPlayer_df["riotIdTagline"][i]
            if summonerName_iter == "#":
                summonerName_iter = puuid_iter
            matchId_iter = recent_TFTPlayer_df["game_id"][i]
            TFTGameCreation_iter: float = datetime.datetime.fromisoformat(recent_TFTPlayer_df["gameCreationDate"][i].replace("Z", "+00:00")).timestamp()
            TFTGameDuration_iter: float = recent_TFTPlayer_df["time_eliminated"][i]
            isPvP_iter = True if recent_TFTPlayer_df["queue_id"][i] in gameQueues and gameQueues[recent_TFTPlayer_df["queue_id"][i]]["category"] == "PvP" else False
            isPvE_iter = True if recent_TFTPlayer_df["queue_id"][i] in gameQueues and gameQueues[recent_TFTPlayer_df["queue_id"][i]]["category"] == "VersusAi" else False
            isCustom_iter = True if recent_TFTPlayer_df["queue_id"][i] in gameQueues and gameQueues[recent_TFTPlayer_df["queue_id"][i]]["category"] == "Custom" else False
            if not puuid_iter in recent_players_metadata:
                recent_players_metadata[puuid_iter] = {}
                recent_players_metadata[puuid_iter]["name"] = summonerName_iter
                recent_players_metadata[puuid_iter]["puuid"] = puuid_iter
                recent_players_metadata[puuid_iter]["gameCount"] = 1
                recent_players_metadata[puuid_iter]["matches"] = [matchId_iter]
                recent_players_metadata[puuid_iter]["createTimestamps"] = [TFTGameCreation_iter]
                recent_players_metadata[puuid_iter]["durations"] = [TFTGameDuration_iter]
                recent_players_metadata[puuid_iter]["isPvP"] = [isPvP_iter]
                recent_players_metadata[puuid_iter]["isPvE"] = [isPvE_iter]
                recent_players_metadata[puuid_iter]["isCustom"] = [isCustom_iter]
                recent_players_metadata[puuid_iter]["PvPCount"] = int(isPvP_iter)
                recent_players_metadata[puuid_iter]["PvECount"] = int(isPvE_iter)
                recent_players_metadata[puuid_iter]["CustomCount"] = int(isCustom_iter)
                recent_players_metadata[puuid_iter]["totalTime"] = TFTGameDuration_iter
                recent_players_metadata[puuid_iter]["totalPvPTime"] = TFTGameDuration_iter * isPvP_iter
                recent_players_metadata[puuid_iter]["totalPvETime"] = TFTGameDuration_iter * isPvE_iter
                recent_players_metadata[puuid_iter]["totalCustomTime"] = TFTGameDuration_iter * isCustom_iter
            else:
                recent_players_metadata[puuid_iter]["gameCount"] += 1
                recent_players_metadata[puuid_iter]["matches"].append(matchId_iter)
                recent_players_metadata[puuid_iter]["createTimestamps"].append(TFTGameCreation_iter)
                recent_players_metadata[puuid_iter]["durations"].append(TFTGameDuration_iter)
                recent_players_metadata[puuid_iter]["isPvP"].append(isPvP_iter)
                recent_players_metadata[puuid_iter]["isPvE"].append(isPvE_iter)
                recent_players_metadata[puuid_iter]["isCustom"].append(isCustom_iter)
                recent_players_metadata[puuid_iter]["PvPCount"] += isPvP_iter
                recent_players_metadata[puuid_iter]["PvECount"] += isPvE_iter
                recent_players_metadata[puuid_iter]["CustomCount"] += isCustom_iter
                recent_players_metadata[puuid_iter]["totalTime"] += TFTGameDuration_iter
                recent_players_metadata[puuid_iter]["totalPvPTime"] += TFTGameDuration_iter * isPvP_iter
                recent_players_metadata[puuid_iter]["totalPvETime"] += TFTGameDuration_iter * isPvE_iter
                recent_players_metadata[puuid_iter]["totalCustomTime"] += TFTGameDuration_iter * isCustom_iter
            #logPrint("[%d/%d]%d\t%s\t%s" %(i, len(recent_TFTPlayer_df) - 1, matchId_iter, puuid_iter, summonerName_iter), end = "\r")
    #进一步计算游玩热度——陪伴得分（Further calculate the company score）
    lambda_decay: float = 0.002 #时间衰减系数（Time decay coefficient）
    scale_factor: int = 100 #缩放因子（Scale factor）
    maxDuration: int = max(map(lambda x: max(x["durations"]), recent_players_metadata.values()))
    for puuid_iter in recent_players_metadata:
        recent_player: dict[str, Any] = recent_players_metadata[puuid_iter]
        sumRecency: float = 0
        sumPvPRecency: float = 0
        sumPvERecency: float = 0
        sumCustomRecency: float = 0
        #声明：以下算法由DeepSeek-V3.2模型生成（Declaration: The following algorithm is generated by DeepSeek-V3.2 model）
        #算法（Algorithm）：热度分数（Company score） = 100 × Σ[wi × (di / Tmax) × exp(-λ × (Tnow - ti))]，其中（where）
        #wi：游戏模式权重（Weight of game modes）
        #di：第i场对局的持续时间（Duration of the i-th game）
        #Tmax：最大持续时间（Max duration among all matches）
        #λ：时间衰减系数（Time decay coefficient）
        #Tnow：当前时间戳（Current unix timestamp）
        #ti：第i场对局的创建时间戳（Unix create timestamp of the i-th match）
        for i in range(recent_player["gameCount"]):
            total_weight: float = 1.5 if recent_player["isPvP"] else 1 if recent_player["isPvE"] else 0.5 #根据需要自行修改（Modify on demand）
            # delta: float = scale_factor * (recent_player["durations"][i] / maxDuration) * math.exp(-lambda_decay * (time.time() - recent_player["createTimestamps"][i]))
            delta: float = recent_player["durations"][i] / (time.time() - recent_player["createTimestamps"][i])
            sumRecency += delta * total_weight
            sumPvPRecency += delta * recent_player["isPvP"][i]
            sumPvERecency += delta * recent_player["isPvE"][i]
            sumCustomRecency += delta * recent_player["isCustom"][i]
        recent_player["totalRecency"] = sumRecency * 1000 #后来添加的修饰因子（Later added modifier）
        recent_player["totalPvPRecency"] = sumPvPRecency * 1000
        recent_player["totalPvERecency"] = sumPvERecency * 1000
        recent_player["totalCustomRecency"] = sumCustomRecency * 1000
    #pyperclip.copy(json.dumps(recent_players_metadata, ensure_ascii = False))
    json01name: str = "Recently Played Summoners - %s.json" %displayName
    json01path: str = os.path.join(export_folder, json01name).replace("\\", "/")
    while True:
        try:
            with open(json01path, "w", encoding = "utf-8") as jsonfile:
                jsonfile.write(json.dumps(recent_players_metadata, indent = 4, ensure_ascii = False))
        except FileNotFoundError:
            os.makedirs(export_folder, exist_ok = True)
        except UnicodeEncodeError:
            logPrint("近期一起玩过的玩家元数据文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nRecently played summoner metadata text generation failure! Please check if the summoner name includes any abnormal characters!\n")
            break
        else:
            break
    recent_players_metadata_list: list[dict[str, Any]] = sorted(recent_players_metadata.values(), key = lambda x: x["gameCount"], reverse = True)
    recent_players_metadata_header: dict[str, str] = {"name": "召唤师名", "puuid": "玩家通用唯一识别码", "gameCount": "共同作战局数", "matches": "共同对局序号", "durations": "对局持续时间列表", "isPvP": "玩家对战逻辑值列表", "isPvE": "人机对战逻辑值列表", "isCustom": "自定义对战逻辑值列表", "PvPCount": "玩家对战局数", "PvECount": "人机对战局数", "CustomCount": "自定义对战局数", "totalTime": "共同作战时长（秒）", "totalPvPTime": "共同玩家对战时长（秒）", "totalPvETime": "共同人机对战时长（秒）", "totalCustomTime": "共同自定义对战时长（秒）"}
    recent_players_metadata_header_keys: list[str] = list(recent_players_metadata_header.keys())
    recent_players_metadata_statistics_output_order: list[int] = [0, 1, 3, 2, 8, 9, 10, 11, 12, 13, 14]
    recent_players_metadata_organized: dict[str, list[Any]] = {recent_players_metadata_header_keys[i]: list(map(lambda x: x[recent_players_metadata_header_keys[i]], recent_players_metadata_list)) for i in recent_players_metadata_statistics_output_order}
    recent_players_metaDf: pandas.DataFrame = pandas.concat([pandas.DataFrame(data = recent_players_metadata_organized)])
    recent_players_metaDf = pandas.concat([pandas.DataFrame([recent_players_metadata_header])[recent_players_metaDf.columns], recent_players_metaDf], ignore_index = True)
    #默认导出玩家对局数量统计表（Export recent played summoner count table by default）
    wb01Name: str = f"Recently Played Summoner Count - {displayName}.xlsx"
    wb01Path: str = os.path.join(export_folder, wb01Name).replace("\\", "/")
    if not os.path.exists(wb01Path):
        wb01CreateFlag: bool = create_workbook_win32(os.path.abspath(wb01Path), log = log)
    os.makedirs(export_folder, exist_ok = True)
    while True:
        try:
            with (pandas.ExcelWriter(path = wb01Path, mode = "a", if_sheet_exists = "replace") if os.path.exists(wb01Path) else pandas.ExcelWriter(path = wb01Path)) as writer:
                addDefaultStyle(recent_players_metaDf).to_excel(excel_writer = writer)
        except PermissionError:
            logPrint("近期一起玩过的玩家对局数量统计表导出失败！请检查文件的权限以及是否被占用！按回车键重试，或者输入任意非空字符串以放弃导出。\nRecently played summoner count table export failure! Please check the permission and if the file is occupied! Press Enter to try again, or submit any non-empty string to give up exporting.")
            gameCount_export_str = logInput()
            gameCount_export = not bool(gameCount_export_str)
            if not gameCount_export:
                break
        else:
            break
    
    #针对元数据中记录的每个玩家的累计游戏时长和游戏对局数输出条形图（Output the bar chart of each summoner's total time and game counts in the metadata）
    totalTime: dict[str, float] = {}
    PvPTime: dict[str, float] = {}
    PvETime: dict[str, float] = {}
    CustomTime: dict[str, float] = {}
    totalCount: dict[str, int] = {}
    PvPCount: dict[str, int] = {}
    PvECount: dict[str, int] = {}
    CustomCount: dict[str, int] = {}
    totalRecency: dict[str, float] = {}
    PvPRecency: dict[str, float] = {}
    PvERecency: dict[str, float] = {}
    CustomRecency: dict[str, float] = {}
    for player in recent_players_metadata.values():
        totalTime[player["name"]] = player["totalTime"]
        PvPTime[player["name"]] = player["totalPvPTime"]
        PvETime[player["name"]] = player["totalPvETime"]
        CustomTime[player["name"]] = player["totalCustomTime"]
        totalCount[player["name"]] = player["gameCount"]
        PvPCount[player["name"]] = player["PvPCount"]
        PvECount[player["name"]] = player["PvECount"]
        CustomCount[player["name"]] = player["CustomCount"]
        totalRecency[player["name"]] = player["totalRecency"]
        PvPRecency[player["name"]] = player["totalPvPRecency"]
        PvERecency[player["name"]] = player["totalPvERecency"]
        CustomRecency[player["name"]] = player["totalCustomRecency"]
    totalTime_sorted: list[tuple[str, float]] = sorted(totalTime.items(), key = lambda x: x[1], reverse = True)
    PvPTime_sorted: list[tuple[str, float]] = sorted(PvPTime.items(), key = lambda x: x[1], reverse = True)
    PvETime_sorted: list[tuple[str, float]] = sorted(PvETime.items(), key = lambda x: x[1], reverse = True)
    CustomTime_sorted: list[tuple[str, float]] = sorted(CustomTime.items(), key = lambda x: x[1], reverse = True)
    totalCount_sorted: list[tuple[str, int]] = sorted(totalCount.items(), key = lambda x: x[1], reverse = True)
    PvPCount_sorted: list[tuple[str, int]] = sorted(PvPCount.items(), key = lambda x: x[1], reverse = True)
    PvECount_sorted: list[tuple[str, int]] = sorted(PvECount.items(), key = lambda x: x[1], reverse = True)
    CustomCount_sorted: list[tuple[str, int]] = sorted(CustomCount.items(), key = lambda x: x[1], reverse = True)
    totalRecency_sorted: list[tuple[str, float]] = sorted(totalRecency.items(), key = lambda x: x[1], reverse = True)
    PvPRecency_sorted: list[tuple[str, float]] = sorted(PvPRecency.items(), key = lambda x: x[1], reverse = True)
    PvERecency_sorted: list[tuple[str, float]] = sorted(PvERecency.items(), key = lambda x: x[1], reverse = True)
    CustomRecency_sorted: list[tuple[str, float]] = sorted(CustomRecency.items(), key = lambda x: x[1], reverse = True)
    logPrint("您希望条形图中显示游戏时长最长的前几名玩家？（默认为前20名）\nHow many players of the longest game time do you want to display in the bar chart? (20 by default)")
    while True:
        try:
            topN_str: str = logInput()
            if topN_str == "":
                topN: int = 20
                break
            else:
                topN = int(topN_str)
        except ValueError:
            logPrint("请输入整数！\nPlease input an integer!")
        else:
            if topN < 0:
                logPrint("请输入自然数！\nPlease input a non-negative integer!")
            else:
                break
    if topN > 0:
        topN = min(topN, len(set(recent_LoLPlayer_df["puuid"][1:])) + len(set(recent_TFTPlayer_df["puuid"][1:])))
        plt.rcParams["font.sans-serif"] = ["Microsoft YaHei"] #设置默认字体为微软雅黑（Set the default font Microsoft YaHei）
        valuefont: dict[str, str | int] = {"family": "Times New Roman", "weight": "normal", "size": 9} #指定柱上显示的数据的字体格式（Determines the font of the values above the bars）
        chart_data: list[tuple[list[tuple[str, Any]], str, str, str, int]] = [
            (totalTime_sorted, "总游戏时间\nGame Time: All Modes", "游戏时长（秒）\ntotalGameTime (s)", "Time_total", 0),
            (PvPTime_sorted, "玩家对战时间\nGame Time: PvP", "游戏时长（秒）\ntotalGameTime (s)", "Time_PvP", 0),
            (PvETime_sorted, "人机对战时间\nGame Time: PvE", "游戏时长（秒）\ntotalGameTime (s)", "Time_PvE", 0),
            (CustomTime_sorted, "自定义对战时间\nGame Time: Custom", "游戏时长（秒）\ntotalGameTime (s)", "Time_Custom", 0),
            (totalCount_sorted, "总游戏对局数\nGame Count: All Modes", "对局数\ntotalGameCount", "Count_total", 0),
            (PvPCount_sorted, "玩家对战局数\nGame Count: PvP", "对局数\ntotalGameCount", "Count_PvP", 0),
            (PvECount_sorted, "人机对战局数\nGame Count: PvE", "对局数\ntotalGameCount", "Count_PvE", 0),
            (CustomCount_sorted, "自定义对战局数\nGame Count: Custom", "对局数\ntotalGameCount", "Count_Custom", 0),
            (totalRecency_sorted, "总游玩热度\nRecency: All Modes", "陪伴得分\ncompanion score", "Recency_total", 0),
            (PvPRecency_sorted, "玩家对战热度\nRecency: PvP", "陪伴得分\ncompanion score", "Recency_PvP", 0),
            (PvERecency_sorted, "人机对战热度\nRecency: PvE", "陪伴得分\ncompanion score", "Recency_PvE", 0),
            (CustomRecency_sorted, "自定义对战热度\nRecency: Custom", "陪伴得分\ncompanion score", "Recency_Custom", 0)
        ]
        logPrint("您想要将所有图表合并为一张图表，还是分别生成？（输入任意非空字符串以分别生成，否则合并为一张图表。）\nDo you want to merge all charts into one, or generate them separately? (Input any non-empty string to generate them separately, or null to merge them into one chart.)")
        separate_str: str = logInput()
        separate: bool = bool(separate_str)
        if separate:
            for data, title, ylabel, file_suffix, ndigits in chart_data:
                plt.figure(figsize = (max(topN / 2, 6), 12))
                players: list[str] = [data[j][0] for j in range(topN)]
                values: list[int | float] = [data[j][1] for j in range(topN)]
                plt.bar(players, values)
                plt.xticks(rotation = 45, ha = "right")
                plt.ylabel(ylabel)
                plt.yticks(fontproperties = "Calibri", size = 12)
                for player, playtime in data[:topN]:
                    plt.text(player, round(playtime, ndigits), round(playtime, ndigits), ha = "center", va = "bottom", fontdict = valuefont)
                plt.title(title)
                plt.savefig(os.path.join(export_folder, "Recently Played Summoners - %s - %s.png" % (displayName, file_suffix)), bbox_inches = "tight")
                plt.clf()
        else:
            fig, axes = plt.subplots(nrows = 3, ncols = 4, figsize = (max(topN * 2, 10), 24))
            axes = axes.flatten()
            for i, (data, title, ylabel, file_suffix, ndigits) in enumerate(chart_data):
                ax = axes[i]
                players = [data[j][0] for j in range(topN)]
                values = [data[j][1] for j in range(topN)]
                bars = ax.bar(players, values)
                ax.set_title(title)
                ax.set_ylabel(ylabel)
                ax.set_xticks(range(len(players)))
                ax.set_xticklabels(players, rotation = 45, ha = "right")
                ax.tick_params(axis = "y", labelsize = 12)
                for label in ax.get_yticklabels():
                    label.set_fontfamily("Calibri")
                for player, playtime in data[:topN]:
                    ax.text(player, round(playtime, ndigits), round(playtime, ndigits), ha = "center", va = "bottom", fontdict = valuefont)
            plt.tight_layout(pad = 3.0)
            plt.savefig(os.path.join(export_folder, "Recently Played Summoners - %s.png" %displayName))
        plt.close("all")

async def search_profile(connection: Connection) -> None:
    # logPrint("是否将部分数据框用于网页展示？（输入任意键展示，否则不生成网页）\nDo you want to display some dataframes in a webpage? (Input anything to display them in a web, or null to skip generating the web.)")
    # web_display = logInput()
    # if bool(web_display):
    #     app = Flask()
    global platformId
    current_party: dict[str, Any] = await (await connection.request("GET", "/lol-lobby/v1/parties/player")).json()
    platformId = current_party["platformId"]
    riot_client_info: list[str] = await (await connection.request("GET", "/riotclient/command-line-args")).json()
    client_info: dict[str, str] = {}
    for i in range(len(riot_client_info)):
        try:
            client_info[riot_client_info[i].split("=")[0]] = riot_client_info[i].split("=")[1]
        except IndexError:
            pass
    region: str = client_info["--region"]
    platform_folder: str = set_platform_folder(region, platformId)
    match_folder: str = os.path.join(platform_folder, "1. MatchIDs")
    logPrint("请选择召唤师技能和装备的输出语言【默认为中文（中国）】：\nPlease select a language to output the summoner spells and items (the default option is zh_CN):")
    language_df: pandas.DataFrame = pandas.DataFrame(language_dict)
    print(format_df(language_df)[0])
    log.write(format_df(language_df, width_exceed_ask = False, direct_print = False)[0] + "\n")
    while True:
        language_option: str = logInput()
        if language_option == "" or language_option in [str(i) for i in range(1, 31)]:
            if language_option == "":
                language_option = "29"
            language_code: str = list(language_ddragon.keys())[int(language_option) - 1]
            switch_language, exit_flag = prepare_data_resources(platformId, language_code)
            if exit_flag:
                return
            if switch_language:
                continue
            else:
                break
        elif language_option[0] == "0":
            return
        else:
            logPrint("语言选项输入错误！请重新输入：\nERROR input of language option! Please try again:")
    # current_info: dict[str, Any] = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json() #一定要注意，在本脚本中，`current_puuid`和`current_info["puuid"]`不是一回事（Pay attention that `current_puuid` and `current_info["puuid"]` aren't the same thing in this program）
    #获取在线游戏模式数据（Get online game mode data）
    gameQueues_initial: list[dict[str, Any]] = await (await connection.request("GET", "/lol-game-queues/v1/queues")).json()
    gameQueues: dict[int, dict[str, Any]] = {queue["id"]: queue for queue in gameQueues_initial}
    #下面创建一个嵌套字典，用来判断所有版本的各种数据是否曾经获取过（The following code create a nested dictionary to judge whether all kinds of data of a patch is once recaptured）
    TemplateBoolList: list[bool] = [False for i in range(len(bigPatches))] #为什么想到起个template作为后面字典的构成，是为了致敬后续出现的模板羁绊（The reason why I choose a name containing "template" to compose the following dictionary is in honor of the following "TemplateTrait"）
    recaptured_header: list[str] = ["bigPatch", "spell", "LoLChampion", "LoLItem", "summonerIcon", "perk", "perkstyle", "TFTAugment", "TFTChampion", "TFTItem", "TFTCompanion", "TFTTrait", "CherryAugment"]
    recaptured: dict[str, dict[str, bool]] = {str(bigPatch): {_: False for _ in recaptured_header} for bigPatch in bigPatches}
    #实际上，目前recaptured并未投入使用。原本打算使用这个字典，是因为有些时候在获取连续的几场版本相同的对局时，如果都没能正确地把数据对应到其名称，那么每一局都会提示将原始数据填充至单元格。但是后来想到，这样虽然会使得输出减少，但是一旦代码完成英雄联盟对局记录的数据整理，要开始整理具体每一场对局了，那么回归到最近的对局的获取时，由于这场场对局的数据可能标记为“曾经获取过”，那么程序可能不再获取这场对局的版本的数据。此时，程序刚完成对局记录的整理，而对局记录最后几场对局可能是老版本，有些新版本的数据是没有的。这样的话，本来可以通过重新获取新版本的数据来将原始数据对应到其名称，现在却因为新版本被标记为已获取过数据的版本，而导致其原始数据被保存下来（Actually, `recaptured` isn't used currently. The original plan on using this dictionary is due to that if the data of several continuous matches of the same gameVersion fail to be mapped to their names, then the prompt like `the original data will be adopted` will pop up for every match to be captured. But then I come to realize that the use of `recaptured` may reduce the output, but under the circumstance of finishing the data organization of LoL match history, when the program is about to capture the latest specific game information and timeline, then the program may never fetch data of this patch. At that time, the program has just finished organizing the match history. Maybe the data version then is an old version, and it doesn't include some new data. In that case, the program could have recaptured data of the latest patch to map data to the corresponding names, but because of the use of `recapture`, this latest patch is marked as "a patch that has been recaptured", and hence the original data instead of their corresponding labels are saved）
    #下面创建一个字典，用来存储程序正在使用的各数据资源的版本（The following code create a dictionary to store the versions of data resources that the program currently uses）
    current_versions: dict[str, str] = {"queue": URLPatch, "spell": URLPatch, "LoLChampion": URLPatch, "LoLItem": URLPatch, "summonerIcon": URLPatch, "perk": URLPatch, "perkstyle": URLPatch, "TFTAugment": URLPatch, "TFTChampion": URLPatch, "TFTItem": URLPatch, "TFTCompanion": URLPatch, "TFTTrait": URLPatch, "CherryAugment": URLPatch}
    #下面创建一个字典，用来存储程序正在使用的各数据资源的版本下发生错误的键。当某个数据资源更换版本时，其出错的键会被清空（The following code create a dictionary to store the keys that fail to map to the constant dictionaries under certain versions of each kind of data resource. Once the version of a data resource changes, its unmapped keys will be cleared）
    unmapped_keys: dict[str, set[Any]] = {"queue": set(), "spell": set(), "LoLChampion": set(), "LoLItem": set(), "summonerIcon": set(), "perk": set(), "perkstyle": set(), "TFTAugment": set(), "TFTChampion": set(), "TFTItem": set(), "TFTCompanion": set(), "TFTTrait": set(), "CherryAugment": set()}
    #控制只输出一遍的提示（Control the hint to be displayed only once）
    puuid_change_warning_printed: bool = False
    #定义全局对局信息缓存（Define global match information caches）
    LoLGame_summary_cache_lcu: dict[int, dict[str, Any]] = {} #将对局概要保存到缓存中，以减少网络请求次数。下同（Save match summaries into the cache to reduce the number of times of web request. Same applies below）
    LoLGame_timeline_cache_lcu: dict[int, dict[str, Any]] = {}
    LoLGame_summary_cache_sgp: dict[int, dict[str, Any]] = {}
    LoLGame_timeline_cache_sgp: dict[int, dict[str, Any]] = {}
    TFTGame_summary_cache_sgp: dict[int, dict[str, Any]] = {}
    #logPrint('''在腾讯代理的服务器上，如果查询某名玩家的对局记录，请尝试以下操作：\nTo search for the match history of a player on Tencent servers, try out the following operations:\n1. 在浏览器中打开本地主机网络协议：%s\n   Open the localhost IP in any browser: %s\n2. 尝试用以下用户名和密码登录：\n   Try logining in with the following username and password:\n   用户名（Username）：riot\n   密码（Password）：%s\n3. （如果可以立即知道一位玩家的玩家通用唯一识别码，则可以跳过第3和4步）在浏览器的地址栏中的地址最后，添加“lol-summoner/v1/summoners?name={name}”，其中{name}指的是召唤师名称编码后的字符串。当召唤师名称只包含英文字母和阿拉伯数字时，直接以召唤师名称去空格后的字符串代入{name}即可；当召唤师名称存在非美国标准信息交换代码时，以召唤师名称编码后的字符串代入{name}。\n(If a summoner's puuid can be immediately known, the user may skip Steps 3 and 4) Add to following the last character of the address in the browser's address bar "lol-summoner/v1/summoners?name={name}", where {name} refers to strings encoded from summonerName. When summonerName contains only English letters and Arabic numbers, simply substitute {name} with the strings with the spaces removed from summonerName. When a non-ASCII character exists in summonerName, substitute {name} by encoded summonerName.\n3.1 对于包含非美国标准信息交换代码的召唤师名称，如果可以得到该召唤师的精确名称（如通过复制到剪贴板），那么在Python中可以得知其编码后的字符串。在Python中使用from urllib.parse import quote命令引入quote函数，再使用quote(x)函数获取字符串x编码后的字符串。\nFor summonerNames that include non-ASCII characters, if the exact summonerName can be obtained (e. g. by copying to clipboard), then its encoded string can be returned in Python. In Python console, use "from urllib.parse import quote" to introduce the "quote" function. Then use quote(x) function to get the string encoded from the string x.\n4. 在lol-summoner/v1/summoners?name={name}返回的结果中找到puuid并复制。\n   Find "puuid" in the result returned by "lol-summoner/v1/summoners?name={name}" and copy it.\n5. 将地址栏中4位IP地址后的斜杠后的内容删除，再添加“lol-match-history/v1/products/lol/{puuid}/matches?begIndex=0&endIndex=20”或“lol-match-history/v1/products/tft/{puuid}/matches?begin=0&count=20”，其中{puuid}是事先获知的玩家通用唯一识别码，或者是第4步复制到剪贴板的puuid。\nDelete the content following the slash after the 4-bit IP address in the address bar and then add to the end "lol-match-history/v1/products/lol/{puuid}/matches?begIndex=0&endIndex=20" or "lol-match-history/v1/products/tft/{puuid}/matches?begin=0&count=20", where {puuid} refers to the puuid previously known, or copied to clipboard in Step 4.\n6. 尝试将上一步输入的地址中的“endIndex=”或“count=”后的数字依次替换成21、199、200和500，观察每次替换后返回的网页结果有没有变多。\nTry changing the number following "endIndex=" or "count=" in the last step into 21, 199, 200 and 500 one by one, and observe whether the returned webpage contains more information after each change.\n7. 教程完成，请继续执行本脚本……\n   Instruction finished. Please continue to run this program ...''' %(connection.address, connection.address, connection.auth_key))
    while True:
        #初始化所有数据资源（Initialize all data resources）
        logPrint("\n正在初始化所有数据资源……\nInitializing all data resources ...\n")
        patches: list[str] = patches_initial.copy()
        queues: dict[int, dict[str, Any]] = queues_initial.copy()
        spells: dict[int, dict[str, Any]] = spells_initial.copy()
        LoLChampions: dict[int, dict[str, Any]] = LoLChampions_initial.copy()
        LoLItems: dict[int, dict[str, Any]] = LoLItems_initial.copy()
        summonerIcons: dict[int, dict[str, Any]] = summonerIcons_initial.copy()
        perks: dict[int, dict[str, Any]] = perks_initial.copy()
        perkstyles: dict[int, dict[str, Any]] = perkstyles_initial.copy()
        TFTAugments: dict[str, dict[str, Any]] = TFTAugments_initial.copy()
        TFTChampions: dict[str, dict[str, Any]] = TFTChampions_initial.copy()
        TFTItems: dict[str, dict[str, Any]] = TFTItems_initial.copy()
        TFTCompanions: dict[str, dict[str, Any]] = TFTCompanions_initial.copy()
        TFTTraits: dict[str, dict[str, Any]] = TFTTraits_initial.copy()
        CherryAugments: dict[int, dict[str, Any]] = CherryAugments_initial.copy()
        current_versions = {"queue": URLPatch, "spell": URLPatch, "LoLChampion": URLPatch, "LoLItem": URLPatch, "summonerIcon": URLPatch, "perk": URLPatch, "perkstyle": URLPatch, "TFTAugment": URLPatch, "TFTChampion": URLPatch, "TFTItem": URLPatch, "TFTCompanion": URLPatch, "TFTTrait": URLPatch, "CherryAugment": URLPatch}
        unmapped_keys = {"queue": set(), "spell": set(), "LoLChampion": set(), "LoLItem": set(), "summonerIcon": set(), "perk": set(), "perkstyle": set(), "TFTAugment": set(), "TFTChampion": set(), "TFTItem": set(), "TFTCompanion": set(), "TFTTrait": set(), "CherryAugment": set()}
        infos: dict[str, dict[str, Any]] = {} #存储程序运行过程中遇到的玩家信息，防止后续程序反复获取已经获取过的玩家信息（Store the summoner information fetched during the program execution, in case the program would keep capturing the summoner information already fetched before）
        match_notbelonging_warning_printed: bool = False #标记在整理一名玩家的对局记录时是否已经打印过某场对局中不包含主召唤师的提示。在该变量为假时，通过用户的输入决定`match_reserve`的取值（Marks whether the hint that a match doesn't contain the main summoner has been printed when the program is organizing a player's match history. When this variable is False, the value of `match_reserve` is determined by user input）
        match_reserve: bool = False #是否保留不包含主召唤师的对局。每次切换召唤师时初始化一次（Whether to reserve the matches that don't contain the main summoner. Initialized once every time the user switches the summoner）
        #处理主召唤师（Handle the main summoner）
        logPrint('''请输入要查询的主召唤师名称。输入“0”以退出程序。\nPlease input the main summoner's name to query. Submit "0" to exit.''')
        while True:
            main_summoner_name: str = logInput()
            if main_summoner_name == "0":
                return
            elif main_summoner_name == "":
                logPrint("请输入非空字符串！\nPlease input a string instead of null!")
                continue
            else:
                main_info: dict[str, Any] = await get_info(connection, main_summoner_name)
                if not main_info["info_got"]:
                    logPrint(main_info["message"])
                else:
                    main_info_body: dict[str, Any] = main_info["body"]
                    displayName: str = get_info_name(main_info_body) #用于文件名命名（For use of file naming）
                    current_puuid: str = main_info_body["puuid"] #用于核验对局是否包含该召唤师。此外，还用于扫描模式从对局的所有玩家信息中定位到该玩家（For use of checking whether the searched matches include this summoner. In addition, it's used for localization of this player from all players in a match in "scan" mode）
                    current_summonerName: str = "" if main_info_body["gameName"] == "" and main_info_body["tagLine"] == "" else main_info_body["gameName"] + "#" + main_info_body["tagLine"] #作用同上，用于模糊定位，主要应用于玩家通用唯一识别码发生变动的大区且在名称编号引入后注册的主召唤师的对局记录扫描模式（Acts as the same role as the above variable for a rough localization. It's mainly designed for Scan Mode on players that signed up after tagLine was introduced on servers that changed the players' puuids）
                    infos[current_puuid] = main_info_body
                    if main_info_body["privacy"] == "PUBLIC":
                        logPrint(f"玩家{displayName}的生涯是公开的。您可以在客户端内搜索其召唤师名以查看其信息。\nPlayer {displayName}'s profile is public. You can search for his/her summoner name in the League Client to view it.")
                    elif main_info_body["privacy"] == "PRIVATE":
                        logPrint(f"玩家{displayName}的生涯不公开。\nPlayer {displayName}'s profile is private.")
                    else:
                        logPrint(f"在获取玩家{displayName}的生涯公开性时出现了一个问题。\nThere's an issue about player {displayName}'s profile privacy.")
                    break
        #处理小号（Handle the smurfs）
        smurfs: list[dict[str, Any]] = await load_smurf(connection, current_puuid = current_puuid, infos = infos)
        #整理账号信息（Organize accounts）
        AllAccounts: list[dict[str, Any]] = [main_info_body] + smurfs
        current_puuid_list: list[str] = list(map(lambda x: x["puuid"], AllAccounts))
        current_summonerName_list: list[str] = list(map(get_info_name, AllAccounts))
                
        #下面检测本地已保存的召唤师信息是否包含已改名的主召唤师（Detect whether the local summoner information contain the main summoner that has changed its name）
        folderNames: list[str] = [folder for folder in os.listdir(platform_folder) if os.path.isdir(os.path.join(platform_folder, folder))]
        synonym_hint_printed: bool = False
        oldName_counter: int = 0
        if len(folderNames) > 0:
            logPrint("正在检查该召唤师是否改过名（Checking if this summoner has changed the name）：")
            for i in range(len(folderNames)):
                logPrint("[%d/%d]" %(i + 1, len(folderNames)), end = "\r")
                folderName: str = folderNames[i]
                summonerInfo_dir: str = os.path.join(platform_folder, folderName)
                json01path: str = os.path.join(summonerInfo_dir, f"Summoner Profile - {folderName}.json")
                if os.path.exists(json01path):
                    try:
                        with open(json01path, "r", encoding = "utf-8") as jsonfile01:
                            test_info_body: dict[str, Any] = json.load(jsonfile01)
                    except:
                        pass
                    else:
                        if isinstance(test_info_body, dict) and "puuid" in test_info_body and test_info_body["puuid"] == current_puuid:
                            if folderName != displayName:
                                if not synonym_hint_printed:
                                    logPrint(f"警告：检测到同大区文件夹下该召唤师存在其它显示名。\nWarning: Another displayName of this summoner is detected in this platform folder.")
                                    synonym_hint_printed = True
                                oldName_counter += 1
                                logPrint("[%d] %s" %(oldName_counter, folderName))
                                new_summonerInfo_dir: str = os.path.join(os.path.dirname(summonerInfo_dir), displayName)
                                try:
                                    os.rename(summonerInfo_dir, new_summonerInfo_dir)
                                except PermissionError:
                                    logPrint(f"重命名以下文件夹时遇到一个权限错误。请检查文件占用情况，然后尝试手动修改该文件夹及其内文件的名称。\nA PermissionError occurred when the program was trying to rename the following folder. Please check the file occupation situation and then try changing the names of this folder and files within manually.\n异常文件夹（Error folder）： {summonerInfo_dir}")
                                except FileExistsError:
                                    logPrint(f"新文件夹已存在。请对比文件并进行适当的删除和合并。\nThe new folder already exists. Please compare files and perform approriate deletion and merge operations.")
                                else: #只有在正常修改文件夹名称时，才去尝试修改其内的文件名（Only when the name of the folder is changed successfully will the program try changing the names of files within）
                                    logPrint(f"已重命名文件夹（Renamed folder）： {summonerInfo_dir} → {new_summonerInfo_dir}")
                                    new_summonerInfo_files: list[str] = os.listdir(new_summonerInfo_dir)
                                    permissionError_files: list[str] = [] #记录因权限不足而命名失败的文件（Record files that fail to be renamed due to permission error）
                                    for file in new_summonerInfo_files:
                                        if folderName in file:
                                            old_file_relpath: str = os.path.join(new_summonerInfo_dir, file)
                                            new_file_name: str = file.replace(folderName, displayName)
                                            new_file_relpath: str = os.path.join(new_summonerInfo_dir, new_file_name)
                                            try:
                                                os.rename(old_file_relpath, new_file_relpath)
                                            except PermissionError:
                                                permissionError_files.append(file)
                                            else:
                                                logPrint(f"已重命名文件（Renamed file）： {file} → {new_file_name}")
                                    else:
                                        if len(permissionError_files) > 0:
                                            logPrint("重命名以下文件时遇到一个权限错误。请检查文件占用情况，然后尝试手动修改改文件的名称。\nA PermissionError occurred when the program was trying to rename the following file(s). Please check the file occupation situation and then try changing the name of this file manually.\n异常文件（Error files）：" + "\n".join(permissionError_files))
        
        # logPrint("主召唤师信息如下：\nMain summoner information is as follows:")
        # logPrint(main_info_body)
        #下面设置输出文件的位置（The following code determines the output files' location）
        folder: str = set_summonerInfo_folder(region, platformId, main_info_body)
        
        #下面读取已保存的对局序号列表（The following code read the saved matchId lists）
        ##英雄联盟（LoL）
        LoLMatches_exported: list[int] = [] #这个列表代表的是导出到工作簿中的对局的对局序号（This list represents the id of the matches exported into the workbook）
        saved_LoLMatchIDs: list[int] = []
        json02name: str = f"Matches Saved (LoL) - {displayName}.json"
        json02path: str = os.path.join(folder, json02name)
        os.makedirs(folder, exist_ok = True)
        if os.path.exists(json02path):
            try:
                with open(json02path, "r", encoding = "utf-8") as jsonfile02:
                    saved_LoLMatchIDs = json.load(jsonfile02)
            except:
                logPrint("已存储的英雄联盟对局数据格式错误！程序将在导出到工作簿时覆盖写此文件。\nSaved LoL match data format error! The program will overwrite to this file when dataframes are exported to the workbook.")
            else:
                if isinstance(saved_LoLMatchIDs, list) and all(map(lambda x: isinstance(x, int), saved_LoLMatchIDs)):
                    LoLMatches_exported += saved_LoLMatchIDs
                else:
                    logPrint("已存储的英雄联盟对局数据格式错误！程序将在导出到工作簿时覆盖写此文件。\nSaved LoL match data format error! The program will overwrite to this file when dataframes are exported to the workbook.")
        saved_LoLMatchIDs.sort()
        ##云顶之弈（TFT）
        TFTMatches_exported: list[int] = []
        saved_TFTMatchIDs: list[int] = []
        json03name: str = f"Matches Saved (TFT) - {displayName}.json"
        json03path: str = os.path.join(folder, json03name)
        os.makedirs(folder, exist_ok = True)
        if os.path.exists(json03path):
            try:
                with open(json03path, "r", encoding = "utf-8") as jsonfile03:
                    saved_TFTMatchIDs: list[int] = json.load(jsonfile03)
            except:
                logPrint("已存储的云顶之弈对局数据格式错误！程序将在导出到工作簿时覆盖写此文件。\nSaved TFT match data format error! The program will overwrite to this file when dataframes are exported to the workbook.")
            else:
                if isinstance(saved_TFTMatchIDs, list) and all(map(lambda x: isinstance(x, int), saved_TFTMatchIDs)):
                    TFTMatches_exported += saved_TFTMatchIDs
                else:
                    logPrint("已存储的云顶之弈对局数据格式错误！程序将在导出到工作簿时覆盖写此文件。\nSaved TFT match data format error! The program will overwrite to this file when dataframes are exported to the workbook.")
        saved_TFTMatchIDs.sort()
        
        #保存主召唤师信息（Save the main summoner information）
        json04name: str = f"Summoner Profile - {displayName}.json"
        json04path: str = os.path.join(folder, json04name)
        os.makedirs(folder, exist_ok = True)
        try:
            with open(json04path, "w", encoding = "utf-8") as jsonfile04:
                jsonfile04.write(json.dumps(main_info_body, indent = 4, ensure_ascii = False))
        except UnicodeEncodeError:
            logPrint("召唤师信息文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nSummoner information text generation failure! Please check if the summoner name includes any abnormal characters!\n")
        else:
            logPrint('召唤师信息已保存为“%s”。\nSummoner information is saved as "%s".\n' %(json04path, json04path))
        # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
        # pkl1name: str = f"Intermediate Object - main_info (Summoner Profile) - {displayName} ({currentTime}).pkl"
        # with open(os.path.join(folder, pkl1name), "wb") as IntObj1:
        #     pickle.dump(main_info_body, IntObj1)
        info_df: pandas.DataFrame = await sort_basic_info(connection, current_puuid)
        info_htmlTable: str = info_df.to_html(escape = False)
        
        #整理英雄成就数据（Organize champion mastery data）
        #logPrint("召唤师英雄成就如下：\nSummoner champion mastery is as follows:")
        mastery: list[dict[str, Any]] = await (await connection.request("GET", f"/lol-champion-mastery/v1/{current_puuid}/champion-mastery")).json()
        #logPrint(mastery)
        json05name: str = f"Champion Mastery - {displayName}.json"
        json05path: str = os.path.join(folder, json05name)
        os.makedirs(folder, exist_ok = True)
        try:
            with open(json05path, "w", encoding = "utf-8") as jsonfile05:
                jsonfile05.write(json.dumps(mastery, indent = 4, ensure_ascii = False))
        except UnicodeEncodeError:
            logPrint("召唤师英雄成就文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nSummoner champion mastery text generation failure! Please check if the summoner name includes any abnormal characters!\n")
        else:
            logPrint('召唤师英雄成就已保存为“%s”。\nSummoner champion mastery is saved as "%s".\n' %(json05path, json05path))
        # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
        # pkl2name: str = f"Intermediate Object - mastery (Champion Mastery) - {displayName} ({currentTime}).pkl"
        # with open(os.path.join(folder, pkl2name), "wb") as IntObj2:
        #     pickle.dump(mastery, IntObj2)
        mastery_df: pandas.DataFrame = await sort_champion_mastery(connection, current_puuid, LoLChampions)
        
        #整理排位数据（Organize ranked data）
        #logPrint("召唤师排位数据如下：\nSummoner ranked data are as follows:") #排位赛部分数据位于召唤师信息中（Part of ranked data are in Profile Sheet）
        ranked: dict[str, Any] = await (await connection.request("GET", f"/lol-ranked/v1/ranked-stats/{current_puuid}")).json()
        #logPrint(ranked)
        if "errorCode" in ranked and ranked["httpStatus"] == 404: #从13.15版本开始，国服体验服的排位信息和对局记录可以正常查询（From Patch 13.15 on, rank data and match history can be searched on Chinese PBE server）
            logPrint("该服务器暂不支持排位数据和对局记录查询！\nThis server doesn't support ranked data and match history lookup!")
            logPrint("是否导出以上召唤师数据至Excel中？（输入任意键导出，否则不导出）\nDo you want to export the above data into Excel? (Press any key to export or null to refuse exporting)")
            export_str: str = logInput()
            export: bool = bool(export_str)
            if export:
                wbName: str = f"Summoner Profile - {displayName}.xlsx"
                wbPath: str = os.path.join(folder, wbName)
                os.makedirs(folder, exist_ok = True)
                if not os.path.exists(wbPath):
                    wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath), log = log)
                workbook_exist: bool = os.path.exists(wbPath)
                while True:
                    try:
                        with (pandas.ExcelWriter(path = wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = wbPath)) as writer:
                            addDefaultStyle(info_df).to_excel(excel_writer = writer, sheet_name = "Profile")
                            addDefaultStyle(mastery_df).to_excel(excel_writer = writer, sheet_name = "Champion Mastery")
                    except PermissionError:
                        logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                        logInput()
                    else:
                        break
            continue
        json06name: str = f"Ranked Data - {displayName}.json"
        json06path: str = os.path.join(folder, json06name)
        os.makedirs(folder, exist_ok = True)
        try:
            with open(json06path, "w", encoding = "utf-8") as jsonfile06:
                jsonfile06.write(json.dumps(ranked, indent = 4, ensure_ascii = False))
        except UnicodeEncodeError:
            logPrint("召唤师排位数据文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nSummoner ranked data text generation failure! Please check if the summoner name includes any abnormal characters!\n")
        else:
            logPrint('召唤师排位数据已保存为“%s”。\nSummoner ranked data are saved as "%s".\n' %(json06path, json06path))
        # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
        # pkl3name: str = f"Intermediate Object - ranked (Rank) - {displayName} ({currentTime}).pkl"
        # with open(os.path.join(folder, pkl3name), "wb") as IntObj3:
        #     pickle.dump(ranked, IntObj3)
        ranked_df: pandas.DataFrame = await sort_ranked_data(connection, current_puuid)
        ranked_htmltable: str = ranked_df.to_html(escape = False)
        
        #整理天梯数据（Organize ranked apex data）
        #logPrint("召唤师所在赛段天梯数据如下：\nSummoner league ladders data are as follows:")
        ladders: list[dict[str, Any]] = await (await connection.request("GET", f"/lol-ranked/v1/league-ladders/{current_puuid}")).json()
        json07name: str = "Ranked Ladders - " + displayName + ".json"
        json07path: str = os.path.join(folder, json07name)
        os.makedirs(folder, exist_ok = True)
        try:
            with open(json07path, "w", encoding = "utf-8") as jsonfile07:
                jsonfile07.write(json.dumps(ladders, indent = 4, ensure_ascii = False))
        except UnicodeEncodeError:
            logPrint("召唤师排位天梯数据文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nSummoner league ladder data text generation failure! Please check if the summoner name includes any abnormal characters!\n")
        else:
            logPrint('召唤师排位天梯数据已保存为“%s”。\nSummoner league ladder data are saved as "%s".\n' %(json07path, json07path))
        # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
        # pkl4name: str = f"Intermediate Object - ranked (Rank) - {displayName} ({currentTime}).pkl"
        # with open(os.path.join(folder, pkl4name), "wb") as IntObj3:
        #     pickle.dump(ladders, IntObj3)
        standings_count: int = 0
        for ladder in ladders:
            for division in ladder["divisions"]:
                standings_count += len(division["standings"])
        if standings_count > 1000:
            logPrint(f"即将获取{standings_count}名玩家的召唤师信息。是否继续？（输入任意非空字符串继续，否则跳过召唤师信息的获取。）\nSummoner information of {standings_count} players is going to be fetched. Do you want to continue? (Submit any non-empty string to continue or null to skip getting summoner information.)")
            ladder_fetch_summoner_info_str: str = logInput()
            ladder_fetch_summoner_info: bool = bool(ladder_fetch_summoner_info_str)
        else:
            ladder_fetch_summoner_info = True
        ladder_df: pandas.DataFrame = await sort_ranked_ladders(connection, current_puuid, fetch_summoner_info = ladder_fetch_summoner_info)
        ladder_htmltable: str = ladder_df.to_html(escape = False)
        
        #初始化对局记录相关变量（Initialize match history related variables）
        scan_lol: bool = False #用于将扫描获取的历史记录保存为后缀为“ - Scan”的工作表，防止后续【一键查询】时会把【本地重查】辛辛苦苦得到的对局记录覆盖掉。这样也有利于手动重整，即每次【一键查询】后，可手动将新增的对局记录加到后缀为“ - Scan”的工作表中（Determines whether to save the match histories to a sheet postfixxed with " - Scan", in case the subsequent [One-Key Query] overwrites the match histories fetched and organized hard by [Local Recheck]. It also helps manual arrangement. That is, after each [One-Key Query], the user may manually add the new match histories to the sheet postfixxed with " - Scan"）
        scan_tft: bool = False
        LoLHistory_df_all: pandas.DataFrame = pandas.DataFrame() #英雄联盟对局记录数据框（LoL match history dataframe）
        TFTHistory_df_all: pandas.DataFrame = pandas.DataFrame() #云顶之弈对局记录数据框（TFT match history dataframe）
        game_leaderboard_dfs: dict[int, pandas.DataFrame] = {} #对局排行榜数据框字典（Match leaderboard dataframe dictionary）
        game_summary_dfs: dict[int, pandas.DataFrame] = {} #对局概要数据框字典（Match summary dataframe dictionary）
        game_timeline_dfs: dict[int, pandas.DataFrame] = {} #对局时间轴数据框字典（Match timeline dataframe dictionary）
        game_event_dfs: dict[int, pandas.DataFrame] = {} #对局事件数据框字典（Match event dataframe dictionary）
        LoLGame_stat_df_export: bool = False #是否导出英雄联盟战绩（Whether to export LoL game stats）
        # TFTGame_stat_df_export: bool = False #云顶之弈战绩就是云顶之弈对局记录（TFT game stat dataframe is exactly the TFT match history dataframe）
        LoLGame_stat_df: pandas.DataFrame = pandas.DataFrame() #英雄联盟对局战绩数据框（LoL match stats dataframe）
        LoLGame_stat_self_df: pandas.DataFrame = pandas.DataFrame() #玩家英雄联盟战绩数据框（Player LoL stats dataframe）
        recent_LoLPlayer_df: pandas.DataFrame = pandas.DataFrame() #近期一起玩过的英雄联盟召唤师数据框（Recently played LoL summoner dataframe）
        TFTGame_stat_df: pandas.DataFrame = pandas.DataFrame() #云顶之弈对局战绩数据框（TFT match stats dataframe）
        TFTGame_stat_self_df: pandas.DataFrame = pandas.DataFrame() #玩家云顶之弈战绩数据框（Player TFT stats dataframe）
        recent_TFTPlayer_df: pandas.DataFrame = pandas.DataFrame() #近期一起玩过的云顶之弈召唤师数据框（Recently played TFT summoner dataframe）
        info_exist_error: dict[int, bool] = {} #当获取对局记录反复出现异常时，为了保证第二次没有获取到的报错信息在导出时不会覆盖上一次使用该程序时导出的正确工作表，设置该列表。列表中的某个元素为True，代表对应的对局记录将能正常导出。由于对局概要往往比对局时间轴更受关注，这里只以LoLGame_summary的完整性作为exist_error的追加依据（When the match history service encounters errors frequently, to make sure the error information won't overlay the normally captured match summary in the last time using this program, this list is declared here. When some element in this list is True, the corresponding match summary / timeline can be exported as usual. Because the LoLGame_summary is basically more focused on than LoLGame_timeline, True/False is appended to exist_error only based on the integrity of LoLGame_summary）
        timeline_exist_error: dict[int, bool] = {}
        main_player_included: dict[int, bool] = {} #当通过列表来查询对局记录时，有可能某场对局并不包含该召唤师（When searching the match history using a list, maybe the summoner isn't present in some match）
        match_reserve_strategy: dict[int, bool] = {} #当某场对局不包含该召唤师，或者对局数据异常时，决定将该对局概要和时间轴导出到Excel中（Decides whether to export the summary and timeline of a match into Excel when this match doesn't include the searched summoner at present or the match data are lost）
        isLoL: dict[int, bool] = {}
        isTFT: dict[int, bool] = {}
        
        #整理英雄联盟对局记录（Organize LoL match history）
        logPrint("是否查询英雄联盟对局记录？（输入任意键查询，否则不查询）\nSearch LoL matches? (Input anything to search or null to skip searching LoL matches)")
        search_LoL_str: str = logInput()
        search_LoL: bool = bool(search_LoL_str)
        LoLMatchIDs: list[int] = [] #代表实际需要查询的对局序号（Represents the matchIds for query）
        old_LoLMatch_detected: bool = False #是否检测到旧对局（Whether any old match is detected）
        update_unsaved_only_lol: bool = False #决定在批量查询时是否只保存未保存过的对局（Decides whether to only save the unsaved matches during searching in batch）
        if search_LoL:
            LoLHistory_dfs: list[pandas.DataFrame] = []
            for i in range(len(AllAccounts)):
                queues = queues_initial.copy()
                spells = spells_initial.copy()
                LoLChampions = LoLChampions_initial.copy()
                LoLItems = LoLItems_initial.copy()
                summonerIcons = summonerIcons_initial.copy()
                perks = perks_initial.copy()
                perkstyles = perkstyles_initial.copy()
                CherryAugments = CherryAugments_initial.copy()
                current_versions["queue"] = current_versions["summonerIcon"] = current_versions["spell"] = current_versions["LoLChampion"] = current_versions["LoLItem"] = current_versions["perk"] = current_versions["perkstyle"] = current_versions["CherryAugment"] = URLPatch
                unmapped_keys["queue"], unmapped_keys["summonerIcon"], unmapped_keys["spell"], unmapped_keys["LoLChampion"], unmapped_keys["LoLItem"], unmapped_keys["perk"], unmapped_keys["perkstyle"], unmapped_keys["CherryAugment"] = set(), set(), set(), set(), set(), set(), set(), set()
                info_puuid: str = current_puuid_list[i]
                info_summonerName: str = current_summonerName_list[i]
                logPrint("[%d/%d]正在获取客户端内玩家%s的英雄联盟对局记录……\nGetting LoL match history of player %s in the client ..." %(i + 1, len(AllAccounts), info_summonerName, info_summonerName))
                LoLHistory_get, LoLHistory_lcu = await get_LoLHistory(connection, info_puuid, log = log)
                json08name: str = f"Match History (LoL) - {displayName}.json"
                json08path: str = os.path.join(folder, json08name)
                os.makedirs(folder, exist_ok = True)
                if info_puuid == current_puuid: #只保存主召唤师的对局记录（Only save the main summoner's match history）
                    try:
                        with open(json08path, "w", encoding = "utf-8") as jsonfile08:
                            jsonfile08.write(json.dumps(LoLHistory_lcu, indent = 4, ensure_ascii = False))
                    except UnicodeEncodeError:
                        logPrint("召唤师英雄联盟对局记录文本文档生成失败！请检查召唤师名称和所选语言是否包含不常用字符！\nSummoner LoL match history text generation failure! Please check if the summoner name and the chosen language include any abnormal characters!\n")
                # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
                # pkl5name: str = f"Intermediate Object - LoLHistory - {displayName} ({currentTime}).pkl"
                # with open(os.path.join(folder, pkl5name), "wb") as IntObj4:
                #     pickle.dump(LoLHistory_lcu, IntObj4)
                logPrint("[%d/%d]正在扩展玩家%s的英雄联盟对局概要和详细信息……\nExpanding LoL match summary and details of player %s ..." %(i + 1, len(AllAccounts), info_summonerName, info_summonerName))
                LoLHistory_get, LoLHistory_sgp = await get_matchSummary_sgp(connection, sgpSession, info_puuid, "LoL", begin = 0, count = 1000, log = log) #这里之所以把count参数写出来，是因为考虑到后续可能随时要调整这个参数。毕竟1000场数据是非常庞大的（Here the reason I write this `count` parameter is considering its value might be adjusted at some time later. After all, data of 1000 matches can be really big）
                for game in LoLHistory_sgp["games"]:
                    matchId: int = int(game["metadata"]["match_id"].split("_")[1])
                    if not matchId in LoLGame_summary_cache_sgp:
                        LoLGame_summary_cache_sgp[matchId] = game
                if use_sgp: #没必要将从SGP API获取的对局概要数据导出到json文件中，因为它实际上就是从SGP API获取的各对局概要数据的加和。因此，为了保证每次读取对局记录时都会在本地形成一个对局记录json文件，LCU API是无论如何都会访问一次的（It's unnecessary to export the match summary data obtained through SGP API into a json file, because it's actually the sum of all match summaries obtained from SGP API. Therefore, in order to make sure a json file will be generated every time the program fetches the match history, LCU API is always accessed）
                    LoLDetails_get, LoLDetails = await get_matchDetails_sgp(connection, sgpSession, info_puuid, "LoL", begin = 0, count = len(LoLHistory_sgp["games"]) if LoLHistory_get else 1000, log = log)
                    for game in LoLDetails["games"]:
                        matchId: int = int(game["metadata"]["match_id"].split("_")[1])
                        if not matchId in LoLGame_timeline_cache_sgp:
                            LoLGame_timeline_cache_sgp[matchId] = game
                if LoLHistory_get:
                    LoLGameCount: int = len(LoLHistory_sgp["games"]) if use_sgp else LoLHistory_lcu["games"]["gameCount"]
                    LoLGamePlayed_singleSummoner: bool = LoLGameCount != 0 #标记该玩家是否进行过英雄联盟对局（Mark whether this summoner has played any LoL game）
                    if LoLGamePlayed_singleSummoner:
                        logPrint(f"玩家{info_summonerName}共进行{LoLGameCount}场英雄联盟对局。\nPlayer {info_summonerName} has played {LoLGameCount} LoL matches.\n")
                    else:
                        logPrint(f"玩家{info_summonerName}从5月1日起就没有进行过任何英雄联盟对局。\nPlayer {info_summonerName} hasn't played any LoL game yet since May 1st.")
                    if use_sgp:
                        LoLHistory_df: pandas.DataFrame = (sort_LoLHistory_sgp(LoLHistory_sgp, info_puuid, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, session = session, log = log))[0]
                    else:
                        LoLHistory_df = (sort_LoLHistory(LoLHistory_lcu, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, session = session, log = log))[0]
                    LoLHistory_dfs.append(LoLHistory_df)
                    if LoLGamePlayed_singleSummoner:
                        logPrint(LoLHistory_df[:min(21, LoLGameCount + 1)], write_time = False)
            LoLHistory_df_all: pandas.DataFrame = pandas.concat([LoLHistory_dfs[0].iloc[:1]] + list(map(lambda x: x.iloc[1:], LoLHistory_dfs)), ignore_index = True) #需要注意数据框的中文表头占用了一行（Note that the Chinese header takes up a record）
            #对局序号去重（Drop duplicates from gameIds）
            gameIds_occurred: set[str | int] = {LoLHistory_df_all["gameId"][0]}
            lines_to_drop: list[int] = []
            for i in range(1, len(LoLHistory_df_all)):
                if LoLHistory_df_all["gameId"][i] in gameIds_occurred:
                    lines_to_drop.append(i)
                else:
                    gameIds_occurred.add(LoLHistory_df_all["gameId"][i])
            LoLHistory_df_all.drop(lines_to_drop, inplace = True)
            LoLHistory_df_all = LoLHistory_df_all.reset_index(drop = True)
            LoLHistory_df_all = pandas.concat([LoLHistory_df_all.iloc[:1], LoLHistory_df_all.iloc[1:].sort_values(by = "gameCreationDate", ascending = False)], ignore_index = True) #这里弃用了根据对局序号排序（Here gameId isn't used to sort the values）
            #警告：在按照对局创建日期排序后，通过SGP API获取的异常对局由于其时间戳是0，将被放到对局记录的最底部（Warning: After ordering by game creation date, those abnormal matches are put at the bottom because their game creation timestamps are 0）
            
            logPrint('请输入要查询的英雄联盟对局序号，批量查询对局请输入对局序号列表，批量查询全部对局请输入“3”，退出英雄联盟对局查询请输入“0”：\nPlease enter the LoL match ID to check. Submit a list containing matchIDs to search in batch. Submit "3" to search the currently stored history in batch. Submit "0" to quit searching for LoL matches.')
            LoLGameIDs: list[int] = LoLHistory_df_all["gameId"][1:].to_list() #代表对局记录中的所有对局序号（Represents all matchIds in the match history）
            old_LoLMatch_detected = len(saved_LoLMatchIDs) > 0
            while True:
                matchId_str: str = logInput()
                if matchId_str == "":
                    continue
                elif matchId_str == "0":
                    LoLMatchIDs = []
                    break
                else:
                    if matchId_str == "3":
                        if old_LoLMatch_detected:
                            latest_LoLMatchID: int = max(saved_LoLMatchIDs) #需要注意，对局序号最大的对局未必是最近进行的对局。而这种情况并不会引起数据的丢失。相反，最近进行的对局会被重新保存一次，从数据完整性的角度上讲无关紧要（Note that the match with the greatest matchId doesn't mean it's the latest match. Nevertheless, when this situation happens, there won't be any data loss. Conversely, the latest match will be saved again, which doesn't matter in terms of data integrity）
                            latest_LoLMatchID_index: int = LoLGameIDs.index(latest_LoLMatchID) if latest_LoLMatchID in LoLGameIDs else 2000
                            logPrint("检测到您以前曾经查询过该召唤师的英雄联盟对局记录。是否只保存该召唤师信息文件夹中不包含的英雄联盟对局？（输入空字符串以只保存未保存过文本文档的对局，否则自行指定对局索引上下限）\nThe program detected that you've searched for this summoner's LoL match history before. Do you want to only save the LoL matches not present in the current summoner folder? (Enter an empty string to saved only the matches whose json files haven't been saved, or any non-empty string to specify the begIndex and endIndex of the matches by yourself)\n即将使用的对局索引下界和上界（The match begIndex and endIndex to be used）：0 %d" %latest_LoLMatchID_index)
                            update_unsaved_only_lol_str: str = logInput()
                            update_unsaved_only_lol: bool = not bool(update_unsaved_only_lol_str)
                        if old_LoLMatch_detected and update_unsaved_only_lol:
                            LoLMatchIDs = LoLGameIDs[:]
                        else:
                            begIndex: int = 0
                            endIndex: int = 0
                            logPrint("请设置需要查询的对局索引下界和上界，以空格为分隔符（输入空字符以默认查询近20场对局）：\nPlease set the begIndex and endIndex of the matches to be searched, split by space (Enter an empty string to search for the recent 20 matches):") #在13.13版本以前，腾讯代理的服务器只支持近20场对局查询（Before Patch 13.13, Tencent servers only provide search of the latest 20 matches）
                            while True:
                                gameIndex: str = logInput()
                                if gameIndex == "":
                                    begIndex, endIndex = 0, 200 * len(AllAccounts)
                                    break
                                elif gameIndex == "0":
                                    break
                                else:
                                    try:
                                        begIndex, endIndex = map(int, gameIndex.split())
                                    except ValueError:
                                        logPrint("请以空格为分隔符输入对局索引的自然数类型的下界和上界！\nPlease enter the two nonnegative integers as the begIndex and endIndex of the matches split by space!")
                                        continue
                                    else:
                                        break
                            if gameIndex == "0":
                                search_LoL = False
                                LoLMatchIDs = []
                                break
                            LoLMatchIDs = LoLGameIDs[begIndex:endIndex]
                    elif matchId_str == "scan":
                        LoLMatchIDs = LoLGameIDs + saved_LoLMatchIDs #小号模式不适用于扫描模式（Smurf Mode doesn't apply under Scan Mode）
                        if len(LoLMatchIDs) == 0:
                            logPrint("尚未保存过该玩家的数据，且该玩家从5月1日起就没有进行过任何英雄联盟对局！\nYou haven't saved this summoner's matches yet, and this player hasn't played any LoL game yet since May 1st.\n")
                            break
                        else:
                            LoLMatchIDs = sorted(set(LoLMatchIDs), reverse = True)
                            logPrint("检测到%d场对局。是否继续？（输入任意键以重新输入要查询的对局序号，否则重新获取这些对局的数据）\nDetected %d matches. Continue? (Input any non-empty string to return to the last step of inputting the matchId, or null to recapture those matches' data)" %(len(LoLMatchIDs), len(LoLMatchIDs)))
                            recapture_str: str = logInput()
                            recapture: bool = bool(recapture_str)
                            if recapture:
                                LoLMatchIDs = [] #如果没有这句语句，那么当重新输入对局序号列表时，从本地文件中检测到的对局数量相比上次检测数的基础上会多出本地文件中包含的对局的数量（Without this assignment, when reinputting the matchId list, the number of matches detected from the local files will become more than that of the last time's check）
                                logPrint('请输入要查询的英雄联盟对局序号，批量查询对局请输入对局序号列表，批量查询全部对局请输入“3”，退出英雄联盟对局查询请输入“0”：\nPlease enter the LoL match ID to check. Submit a list containing matchIDs to search in batch. Submit "3" to search the currently stored history in batch. Submit "0" to quit searching for LoL matches.')
                                continue
                            scan_lol = True #不应直接放到matchID == "scan"语句下，因为有可能历史记录不是扫描获取的，而是一开始就获取的。比如“尚未保存过该玩家的数据”，或者提示“检测到若干场对局。是否继续”选择了否（This statement shouldn't follow closely after the statement `matchId == "scan"`, because the match history might be obtained in the beginning instead of by scanning. Cases are that a summoner's data has never been saved locally, and that the user inputs something in face of the hint "Detected some matches. Continue?"）
                            queues = queues_initial.copy()
                            spells = spells_initial.copy()
                            LoLChampions = LoLChampions_initial.copy()
                            LoLItems = LoLItems_initial.copy()
                            summonerIcons = summonerIcons_initial.copy()
                            perks = perks_initial.copy()
                            perkstyles = perkstyles_initial.copy()
                            CherryAugments = CherryAugments_initial.copy()
                            current_versions["queue"] = current_versions["summonerIcon"] = current_versions["spell"] = current_versions["LoLChampion"] = current_versions["LoLItem"] = current_versions["perk"] = current_versions["perkstyle"] = current_versions["CherryAugment"] = URLPatch
                            unmapped_keys["queue"], unmapped_keys["summonerIcon"], unmapped_keys["spell"], unmapped_keys["LoLChampion"], unmapped_keys["LoLItem"], unmapped_keys["perk"], unmapped_keys["perkstyle"], unmapped_keys["CherryAugment"] = set(), set(), set(), set(), set(), set(), set(), set()
                            #官方的历史记录最多保留200场对局的个人信息。这里要实现将待保存对局全部整理成一个类似于历史记录的布局的功能（要查看历史记录的原来的布局，可以先不使用scan选项，生成Excel文件后查看“Match History”工作表的布局），所以不再使用前面的历史记录，而是从每一局中提取信息，整合成一张历史记录表。因此，大部分代码复制自前面一部分的代码（Official match history holds personal history of at most 200 matches. Here I want to implement a function to organize the information of all matches into a table like the original match history table. (To check this format for the first time, please don't choose the "scan" option and view the "Match History" sheet of the generated xlsx file.) Therefore, the previous history_df is abandoned. Instead, information in the match history is extracted from all matches to form the table subsequently）
                            if use_sgp:
                                LoLHistory_df_all = (await reconstruct_LoLHistory_sgp(connection, sgpSession, LoLMatchIDs, current_puuid_list, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, LoLGame_summary_cache = LoLGame_summary_cache_sgp, session = session, log = log))[0]
                            else:
                                LoLHistory_df_all = (await reconstruct_LoLHistory(connection, LoLMatchIDs, current_puuid_list, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, LoLGame_summary_cache = LoLGame_summary_cache_lcu, session = session, log = log))[0]
                                #LoLHistory_df_all.apply(lambda x: pandas.Series([-3], index = ["K/D/A"]))
                            logPrint("是否一同保存每场对局的信息？（输入任意键保存，否则将只导出对局记录）\nSave each match? (Input anything to save each match, or null to only save the scanned match history)")
                            sort_gameInfo_sync_str: str = logInput()
                            sort_gameInfo_sync: bool = bool(sort_gameInfo_sync_str)
                            if not sort_gameInfo_sync:
                                break
                    else:
                        try:
                            tmp = eval(matchId_str)
                        except (SyntaxError, NameError):
                            logPrint("您的输入存在语法错误。请重新输入！\nSyntax ERROR detected in this input! Please try again!")
                            continue
                        else:
                            LoLMatchIDs = []
                            if isinstance(tmp, int):
                                LoLMatchIDs.append(tmp)
                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int), tmp)):
                                LoLMatchIDs += tmp
                            else:
                                logPrint("数据类型不符。请重新输入！\nData type not matched! Please try again!")
                                continue
                            if len(LoLMatchIDs) == 0:
                                logPrint("您输入的对局序号集无效！请重新输入。\nInvalid matchId set! Please try again.")
                                continue
                    break
        #整理英雄联盟对局信息（Organize LoL match information）
        LoLGamePlayed: bool = len(LoLMatchIDs) > 0
        if LoLGamePlayed:
            LoLMatches_not_found: list[int] = [] #在扫描模式下，当从本地文件获取的对局从API重新获取出现异常时，处理策略是输出异常信息并跳过该对局，而不是将其直接从对局序号列表中去除，因为这样会使循环乱套。而后面的info_exist_error、timeline_exist_error、main_player_included和match_reserve_strategy只会在该对局正常获取时才会统计。所以一旦出现数据获取失败的对局，在最后导出数据时，“if match_reserve_strategy[i]:”语句会出现“IndexError: list index out of range”报错（Under scan mode, when an exception occurred during crawling matches with LoLMatchIDs obtained from local files from API, the strategy is to print the exception and skip this match, instead of directly removing them from the matchId list, for the removal will disturb the loop. However, the variables info_exist_error, timeline_exist_error, main_player_included and match_reserve_strategy only work when the matches are crawled from the database as expected. So once a match fails to be captured, during xlsx file export at the end of the program, an "IndexError: list index out of range" exception will emerge from the statement "if match_reserve_strategy[i]:"）
            error_LoLMatchIDs: list[int] = [] #记录实际存在但未如期获取的对局序号（Records the LoL matchIDs that really exist but fail to be fetched）
            queues = queues_initial.copy()
            spells = spells_initial.copy()
            LoLChampions = LoLChampions_initial.copy()
            LoLItems = LoLItems_initial.copy()
            summonerIcons = summonerIcons_initial.copy()
            perks = perks_initial.copy()
            perkstyles = perkstyles_initial.copy()
            CherryAugments = CherryAugments_initial.copy()
            current_versions["queue"] = current_versions["summonerIcon"] = current_versions["spell"] = current_versions["LoLChampion"] = current_versions["LoLItem"] = current_versions["perk"] = current_versions["perkstyle"] = current_versions["CherryAugment"] = URLPatch
            unmapped_keys["queue"], unmapped_keys["summonerIcon"], unmapped_keys["spell"], unmapped_keys["LoLChampion"], unmapped_keys["LoLItem"], unmapped_keys["perk"], unmapped_keys["perkstyle"], unmapped_keys["CherryAugment"] = set(), set(), set(), set(), set(), set(), set(), set()
            logPrint("是否输出每场对局的文本文档？（输入任意键不输出，否则默认输出）\nExport text files of each match? (Input anything to refuse exporting, or null to export by default)")
            save_all_json_str: str = logInput()
            save_all_json: bool = not bool(save_all_json_str)
            LoLGame_stat_header_keys: list[str] = list(LoLGame_stat_header.keys())
            LoLGame_stat_data: dict[str, list[Any]] = {key: [] for key in LoLGame_stat_header_keys} #将主召唤师的信息单独导出到一个工作表中（Export the game stats of the main summoner into a single sheet）
            for matchId in LoLMatchIDs:
                match_id: str = f"{platformId}_{matchId}"
                LoLGame_summary_export: bool = not (old_LoLMatch_detected and update_unsaved_only_lol and matchId in saved_LoLMatchIDs) #标记是否导出对局概要。如果是在批量查询全部对局的情况下仅保存本地没有的对局，且该对局已在本地，则不保存本场对局（Marks whether to export the match summary. If the user submits "3" to search matches in batch and selected to update the matches that don't exist locally, while the current match already exists, then the program won't export this match）
                LoLGame_leaderboard_export: bool = args.export_leaderboard #标记是否导出对局排行榜。仅可通过命令行变量指定（Marks whether to export the match leaderboard. Can only be specified by the command line argument）
                LoLGame_timeline_export: bool = LoLGame_summary_export #标记是否导出对局时间轴。时间轴的整理依赖于概要，因此目前认为这两者的值相同（Marks whether to export the match timeline. Timeline data organization is based on the match summary, so its value is set the same as the above）
                #LoLGame_event_export: bool = LoLGame_timeline_export #标记是否导出对局事件信息。由于事件信息源于时间轴，因此这两者的值在任何情形下是相同的（Marks whether to export the match events. Because events are extracted from the timeline, these two values should be the same under any circumstance）
                info_text_saved = timeline_text_saved = False #标记对局概要和时间轴的文本文档是否保存（Marks whether the json files of match summary and timeline are saved）
                isLoL[matchId] = False #这里可以使用（This assignment can be replaced by）：`isLoL[matchId] = isTFT[matchId] = False`
                
                #获取数据（Get data）
                ##信息/概要（Information / Summary） #即使不导出对局概要，对局概要也要用来制作玩家战绩表，因此仍然要获取对局概要。如果不需要加载多的对局，用户需要在前面指定对局上下限来控制LoLMatchIDs（Although the user doesn't want to export the match summary, it's still needed for match stats table, so match summary is always necessary. If the user doesn't want to load extra matches, he/she needs to control `LoLMatchIDs` by specifying the begIndex and the endIndex above）
                if use_sgp:
                    if matchId in LoLGame_summary_cache_sgp:
                        LoLGame_summary: dict[str, Any] = LoLGame_summary_cache_sgp[matchId]
                        status: int = 200
                    else:
                        status, LoLGame_summary = await get_game_summary_sgp(connection, sgpSession, match_id, skipTFT = True, log = log)
                        if status == 200:
                            LoLGame_summary_cache_sgp[matchId] = LoLGame_summary
                else:
                    if matchId in LoLGame_summary_cache_lcu:
                        LoLGame_summary = LoLGame_summary_cache_lcu[matchId]
                        status = 200
                    else:
                        status, LoLGame_summary = await get_LoLGame_summary(connection, matchId, log = log)
                        if status == 200:
                            LoLGame_summary_cache_lcu[matchId] = LoLGame_summary
                if status == 200 and (not use_sgp or "json" in LoLGame_summary and bool(LoLGame_summary["json"])):
                    info_exist_error[matchId] = False
                    isLoL[matchId] = True
                    isTFT[matchId] = False #这里假设不可能有一场英雄联盟对局和一场云顶之弈对局共用一个对局序号。这个假设显然是成立的（Here we assume there can't be a LoL match and a TFT match sharing the same matchId. This assumption obviously holds）
                    save_one_json = LoLGame_summary_export #决定是否保存单场对局的文本文档。match_reserve_strategy变量决定的是是否将不包含主召唤师的对局记录导出到Excel中。由于保存文本文档的同时往往意味着需要导出到Excel中，因此这两者总是相同的。当然也有例外，如他人的对局时（Decides whether to save this match into a json file. The variable match_reserve_strategy decides whether to export the matches which don't include the main summoner into Excel. Because saving the text file always means to export data into Excel, these two variables are usually equal. Exceptions are another one's matches）
                    if use_sgp:
                        participant_puuid: list[str] = list(map(lambda x: x["puuid"], LoLGame_summary["json"]["participants"]))
                        #participant_summonerName: list[str] = list(map(lambda x: x["summonerName"], LoLGame_summary["json"]["participants"])) #这个变量已弃用（This variable is deprecated）
                        participant_gameName: list[str] = list(map(lambda x: "%s#%s" %(x.get("riotIdGameName", x.get("riotIdName", "")), x.get("riotIdTagline", "")), LoLGame_summary["json"]["participants"]))
                    else:
                        participant_puuid: list[str] = list(map(lambda x: x["player"]["puuid"], LoLGame_summary["participantIdentities"]))
                        #participant_summonerName: list[str] = list(map(lambda x: x["player"]["summonerName"], LoLGame_summary["participantIdentities"])) #这个变量已弃用（This variable is deprecated）
                        participant_gameName: list[str] = list(map(lambda x: "%s#%s" %(x["player"]["gameName"], x["player"]["tagLine"]), LoLGame_summary["participantIdentities"]))
                    if len(set(current_puuid_list) & set(participant_puuid)) > 0: #之所以使用玩家通用唯一识别码，而不是用召唤师名称来识别对局是否包含主玩家，是因为该玩家可能使用过改名卡。这里也没有选择帐户序号，这是因为保存在对局中的各玩家的帐户序号竟然是0！（The reason why the puuid instead of the displayName or summonerName is used to identify whether the matches contain the main player is that the player may have used name changing card. AccountId isn't chosen here, because all players' accountIds saved in the match fetched from 127 API is 0, to my surprise!）
                        main_player_included[matchId] = True
                        match_reserve_strategy[matchId] = True
                    elif len(set(current_summonerName_list) & set(participant_gameName)) > 0: #在玩家通用唯一识别码发生变动的大区，要识别变动之前的对局是否包含主玩家，最好的办法是依据显示名。因为在引入名称编号后，显示名就固定下来，没有办法变动了，玩家只能通过改名卡修改玩家名称和名称编号。也就是说，显示名可视为玩家的另一种“身份识别码”。对于在引入名称编号后注册的玩家，其显示名是空字符串，所以在模糊定位时用玩家名称代替（On servers that changed the players' puuids once, to identify whether the matches before this change include this player, the best strategy is to refer to the displayName. This is because after tagLine is introduced, displayName is locked and there's no way of changing it. What the player can change through the Summmoner Name Change is gameName and tagLine. That is to say, displayName may be regarded as another ID of a player. For those who signed up after tagLine was introduced, their displayNames are empty strings. So gameName is taken for the rough localization）
                        main_player_included[matchId] = True
                        match_reserve_strategy[matchId] = True
                        if not puuid_change_warning_printed:
                            logPrint("警告：该大区的玩家通用唯一识别码曾发生变动！请检查保存的各对局是否属于该玩家。\nWarning: The puuids of players on this server have been changed! Please check if the saved matches really belong to this player.")
                            puuid_change_warning_printed = True
                    else:
                        main_player_included[matchId] = False
                        save_one_json = args.reserve_text #由于从文本文件中可以提取该召唤师的对局序号，所以需要保证保留下来的文本文件都包含该召唤师。因此，如果一场对局不包含该召唤师，就不应该把这场对局保存下来，除非用户出于特殊目的需要保留文本文件（Because a summoner's matchIDs can be extracted from the saved json files, it needs to be guaranteed that all saved json files belong to this summoner. Therefore, if a match doesn't include this summoner, then it shouldn't be saved into json files, unless the user must save it with special purposes）
                        if args.reserve:
                            match_reserve: bool = True
                            logPrint("[%d/%d]对局%d不包含该玩家！已保持该对局。\nMatch %d doesn't include the current player but is reserved." %(LoLMatchIDs.index(matchId) + 1, len(LoLMatchIDs), matchId, matchId))
                        else:
                            if not match_notbelonging_warning_printed:
                                logPrint("警告：对局%d不包含该玩家！是否仍要保持该对局？（输入任意键以保留该对局，否则舍弃该对局）\nWarning: The Match %d doesn't include the current player! Continue? (Input any nonempty string to reserve this match, or null to abandon it.)\n注意：此改动对于后续情形也生效。\nNote: This decision takes effect in similar situations later." %(matchId, matchId))
                                match_reserve_str: str = logInput()
                                match_reserve = bool(match_reserve_str)
                                match_notbelonging_warning_printed = True
                            elif match_reserve:
                                logPrint("[%d/%d]对局%d不包含该玩家！已保持该对局。\nMatch %d doesn't include the current player but is reserved." %(LoLMatchIDs.index(matchId) + 1, len(LoLMatchIDs), matchId, matchId))
                            else:
                                logPrint("[%d/%d]对局%d不包含该玩家！已舍弃该对局。\nMatch %d doesn't include the current player and is decrepated." %(LoLMatchIDs.index(matchId) + 1, len(LoLMatchIDs), matchId, matchId))
                        match_reserve_strategy[matchId] = match_reserve
                else:
                    LoLGame_summary_export = False
                    LoLGame_leaderboard_export = False
                    info_exist_error[matchId] = True
                    save_one_json: bool = False
                ##时间轴/详细信息（Timeline / Details）
                if LoLGame_timeline_export:
                    if use_sgp:
                        if matchId in LoLGame_timeline_cache_sgp:
                            LoLGame_timeline: dict[str, Any] = LoLGame_timeline_cache_sgp[matchId]
                            status: int = 200
                        else:
                            status, LoLGame_timeline = await get_game_timeline_sgp(connection, sgpSession, match_id, checkTFT = False, log = log)
                            if status == 200:
                                LoLGame_timeline_cache_sgp[matchId] = LoLGame_timeline
                    else:
                        if matchId in LoLGame_timeline_cache_lcu:
                            LoLGame_timeline = LoLGame_timeline_cache_lcu[matchId]
                            status: int = 200
                        else:
                            status, LoLGame_timeline = await get_LoLGame_timeline(connection, matchId, log = log)
                            if status == 200:
                                LoLGame_timeline_cache_lcu[matchId] = LoLGame_timeline
                    if "errorCode" in LoLGame_timeline:
                        LoLGame_timeline_export = False
                        timeline_exist_error[matchId] = True
                    elif "errorCode" in LoLGame_summary:
                        LoLGame_timeline_export = False
                        timeline_exist_error[matchId] = False
                    else: #在整理时间轴数据时，需要使用`LoLGame_summary`中的一些数据（While organizing the timeline data, some data in `LoLGame_summary` are needed）
                        timeline_exist_error[matchId] = False
                else:
                    LoLGame_timeline = {}
                
                #提示（Prompt）
                process_header: str = "保存进度（Saving process）" if save_all_json and save_one_json else "加载进度（Loading process）"
                if LoLGame_timeline_export:
                    timeline_note: str = ""
                else:
                    if old_LoLMatch_detected and update_unsaved_only_lol:
                        timeline_note = " (Match timeline skipped)"
                    elif save_all_json:
                        timeline_note = " (Exceptional match bypassed)"
                    else:
                        timeline_note = ""
                logPrint("%s：%d/%d\t对局序号（MatchID）： %d%s" %(process_header, LoLMatchIDs.index(matchId) + 1, len(LoLMatchIDs), matchId, timeline_note), print_time = True)
                
                #导出数据（Export data）
                ##信息/概要（Information / Summary）
                if status == 200 and (not use_sgp or "json" in LoLGame_summary):
                    if save_all_json and save_one_json:
                        json10name: str = f"Match Summary (LoL) - {platformId}-{matchId} (SGP).json" if use_sgp else f"Match Summary (LoL) - {platformId}-{matchId}.json"
                        os.makedirs(match_folder, exist_ok = True)
                        try:
                            with open(os.path.join(match_folder, json10name), "w", encoding = "utf-8") as jsonfile10: #如果有两个人存在于同一场对局中，那么保存第二个人的对局概要时，将重新写一遍这个文件（If two players exist in one match, then to save the second player's match summary, the same json file will be written twice）
                                jsonfile10.write(json.dumps(LoLGame_summary, indent = 4, ensure_ascii = False))
                        except UnicodeEncodeError:
                            logPrint("对局%d概要文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nMatch %d summary text generation failure! Please check if the summoner name includes any abnormal characters!" %(matchId, matchId))
                        else:
                            info_text_saved = True
                            LoLMatches_exported.append(matchId)
                        # currentTime = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
                        # pkl7name = f"Intermediate Object - Match Summary (LoL) - {platformId}-{matchId}.pkl"
                        # with open(os.path.join(match_folder, pkl7name), "wb") as IntObj6:
                        #     pickle.dump(LoLGame_summary, IntObj6)
                    if use_sgp:
                        LoLGame_summary_df, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments = sort_LoLGame_summary_sgp(LoLGame_summary, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments, gameIndex = LoLMatchIDs.index(matchId) + 1, current_puuid = current_puuid_list, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, session = session, sortStats = True, LoLGame_stat_data = LoLGame_stat_data, save_self = True, save_other = True, log = log)
                    else:
                        LoLGame_summary_df, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments = sort_LoLGame_summary(LoLGame_summary, queues, summonerIcons, LoLChampions, spells, LoLItems, perks, perkstyles, CherryAugments, gameIndex = LoLMatchIDs.index(matchId) + 1, current_puuid = current_puuid_list, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, session = session, sortStats = True, LoLGame_stat_data = LoLGame_stat_data, save_self = True, save_other = True, log = log)
                
                    #社交排行榜（Social leaderboard）
                    if args.export_leaderboard:
                        if use_sgp:
                            participant_puuid: list[str] = list(map(lambda x: x["puuid"], LoLGame_summary["json"]["participants"]))
                        else:
                            participant_puuid: list[str] = list(map(lambda x: x["player"]["puuid"], LoLGame_summary["participantIdentities"]))
                        LoLGame_leaderboard_df: pandas.DataFrame = await sort_game_leaderboard(connection, puuids = participant_puuid, log = log)
                    else:
                        LoLGame_leaderboard_df = pandas.DataFrame()
                else:
                    if status == 404: #由于时间轴的整理依赖于概要，仅根据对局概要的获取情况来判断对局是否存在即可（Because organizing timeline relies on the summary, the status of getting the match summary is enough for judging whether a match file exists）
                        LoLMatches_not_found.append(matchId)
                    else:
                        error_LoLMatchIDs.append(matchId)
                    LoLGame_summary_error: dict[str, list[str]] = {"项目": list(error_header.values()), "items": list(error_header.keys()), "值": [LoLGame_summary.get(key, "") for key in error_header_keys]}
                    LoLGame_summary_df: pandas.DataFrame = pandas.DataFrame(data = LoLGame_summary_error)
                    LoLGame_leaderboard_df = pandas.DataFrame()
                
                ##时间轴/详细信息（Timeline / Details）
                LoLGame_timeline_df: pandas.DataFrame = pandas.DataFrame()
                LoLGame_event_df: pandas.DataFrame = pandas.DataFrame()
                if LoLGame_timeline_export:
                    if "errorCode" in LoLGame_timeline:
                        LoLGame_timeline_error: dict[str, list[Any]] = {"项目": list(error_header.values()), "items": list(error_header.keys()), "值": [LoLGame_timeline.get(key, "") for key in error_header_keys]}
                        LoLGame_timeline_df = pandas.DataFrame(data = LoLGame_timeline_error)
                        LoLGame_event_df = pandas.DataFrame(data = LoLGame_timeline_error)
                    elif not "errorCode" in LoLGame_summary:
                        if save_all_json and save_one_json: #时间轴的单场文本文档保存策略继承了对局概要的单场文本文档保存策略（`save_one_json` of match timeline inherits from that of match summary）
                            json11name: str = f"Match Timeline (LoL) - {platformId}-{matchId} (SGP).json" if use_sgp else f"Match Timeline (LoL) - {platformId}-{matchId}.json"
                            os.makedirs(match_folder, exist_ok = True)
                            try:
                                with open(os.path.join(match_folder, json11name), "w", encoding = "utf-8") as jsonfile11:
                                    jsonfile11.write(json.dumps(LoLGame_timeline, indent = 4, ensure_ascii = False))
                            except UnicodeEncodeError:
                                logPrint("对局%d时间轴文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nMatch %d timeline text generation failure! Please check if the summoner name includes any abnormal characters!" %(matchId, matchId))
                            else:
                                timeline_text_saved = True
                            # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
                            # pkl8name: str = f"Intermediate Object - Match Timeline (LoL) - {platformId}-{matchId}.pkl"
                            # with open(os.path.join(match_folder, pkl8name), "wb") as IntObj7:
                            #     pickle.dump(LoLGame_timeline, IntObj7)
                        if use_sgp:
                            LoLGame_timeline_df, LoLGame_event_df, LoLItems = await sort_LoLGame_timeline_sgp(connection, LoLGame_timeline, LoLGame_summary, LoLChampions, LoLItems, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, useInfoDict = True, infos = infos, log = log)
                        else:
                            LoLGame_timeline_df, LoLGame_event_df, LoLItems = sort_LoLGame_timeline(LoLGame_timeline, LoLGame_summary, LoLChampions, LoLItems, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, log = log)
                
                if LoLGame_leaderboard_export:
                    game_leaderboard_dfs[matchId] = LoLGame_leaderboard_df.copy(deep = True)
                if LoLGame_summary_export:
                    game_summary_dfs[matchId] = LoLGame_summary_df.copy(deep = True) #这里添加的LoLGame_summary_df会在下一次循环中发生改变，这是数据框类型的特性。因此这里采用深复制，将原有内容克隆到另外一个地址，这样能保证每次添加的是不同的对局概要（The added LoLGame_summary_df will be modified next time in the loop, which belongs to the characteristics of DataFrame data type. Therefore a deep copy is used here to clone the original content to another address, so that each time the appended content is different）
                if LoLGame_timeline_export: #即使这里显示是导出的，由于后面取对局序号时是按照对局概要来取的，所以如果对局概要获取异常而对局时间轴获取正常，最终对局时间轴也不会导出到Excel中（Although it seems that a timeline dataframe is to be exported, because the final matchIDs are obtained from `game_summary_dfs`, if the program fails to get match summary but succeeds to get match timeline, the timeline information won't be exported into Excel）
                    game_timeline_dfs[matchId] = LoLGame_timeline_df.copy(deep = True)
                    game_event_dfs[matchId] = LoLGame_event_df.copy(deep = True)
            
            if use_sgp:
                LoLGame_stat_statistics_output_order: list[int] = [0, 112, 148, 131, 132, 146, 128, 147, 68, 21, 16, 13, 25, 26, 11, 18, 22, 14, 29, 15, 20, 30, 19, 24, 227, 218, 628, 184, 54, 625, 626, 96, 133, 125, 82, 152, 135, 52, 51, 55, 223, 224, 186, 187, 188, 189, 190, 191, 192, 221, 200, 212, 201, 213, 202, 214, 203, 215, 204, 216, 205, 217, 95, 64, 45, 229, 230, 231, 234, 235, 98, 94, 99, 49, 72, 71, 74, 73, 66, 167, 129, 113, 174, 153, 164, 157, 115, 102, 169, 156, 114, 101, 168, 97, 61, 60, 58, 59, 161, 162, 166, 158, 159, 116, 103, 170, 62, 176, 179, 178, 136, 177, 65, 79, 232, 80, 233, 93, 57, 163, 105, 155, 160, 171, 172, 83, 84, 106, 108, 173, 85, 107, 67, 47, 109, 110, 100, 48, 56, 78, 130, 127, 111, 43, 44, 104, 69, 70, 175, 63, 46, 81, 154, 165, 137, 139, 141, 142, 228, 144, 145, 602, 616, 608, 604, 609, 605, 610, 606, 611, 607, 620, 618, 621, 619, 598, 596, 597, 50, 149, 150, 75, 76, 77, 182, 181, 180, 236, 117, 143, 672, 658, 643, 728, 674, 671, 675, 647, 660, 715, 692, 687, 721, 701, 712, 705, 689, 678, 717, 704, 688, 677, 716, 673, 655, 654, 652, 653, 709, 710, 714, 706, 707, 690, 679, 718, 656, 723, 726, 725, 694, 724, 659, 665, 666, 670, 651, 711, 681, 703, 708, 729, 719, 720, 668, 669, 682, 683, 661, 645, 684, 685, 676, 646, 650, 664, 693, 691, 686, 641, 642, 680, 662, 663, 722, 657, 644, 667, 702, 713, 695, 696, 697, 698, 727, 699, 700, 802, 750, 749, 773, 759, 744, 830, 831, 833, 775, 772, 776, 748, 761, 817, 793, 788, 823, 803, 814, 807, 790, 779, 819, 806, 789, 778, 818, 774, 756, 755, 753, 754, 811, 812, 816, 808, 809, 791, 780, 820, 757, 825, 828, 827, 795, 826, 760, 766, 767, 834, 771, 752, 813, 782, 810, 805, 832, 821, 822, 769, 770, 762, 746, 785, 786, 783, 784, 777, 747, 751, 765, 794, 792, 787, 742, 743, 781, 763, 764, 824, 758, 745, 768, 804, 815, 796, 797, 798, 799, 829, 800, 801, 238, 239, 240, 241, 242, 243, 244, 245, 246, 247, 248, 249, 250, 251, 252, 253, 254, 255, 256, 257, 258, 259, 260, 261, 262, 263, 264, 265, 266, 267, 268, 269, 270, 381, 271, 272, 273, 274, 275, 276, 277, 278, 279, 280, 281, 282, 382, 283, 284, 285, 383, 286, 287, 288, 289, 290, 291, 292, 293, 294, 295, 296, 297, 298, 299, 300, 301, 302, 303, 304, 305, 306, 307, 308, 309, 310, 311, 312, 313, 314, 315, 316, 317, 318, 319, 384, 320, 321, 322, 323, 324, 325, 326, 327, 328, 329, 330, 385, 331, 332, 333, 334, 335, 336, 337, 338, 339, 340, 341, 342, 343, 344, 345, 346, 347, 348, 349, 350, 351, 352, 353, 354, 355, 356, 357, 358, 359, 360, 361, 362, 363, 364, 365, 366, 367, 368, 369, 370, 371, 372, 373, 374, 375, 376, 377, 378, 379, 380, 386, 387, 388, 389, 390, 391, 392, 393, 394, 395, 396, 397, 398, 399, 400, 401, 402, 403, 404, 405, 406, 407, 408, 409, 410, 411, 412, 413, 414, 415, 416, 417, 418, 419, 420, 421, 422, 423, 424, 425, 426, 427, 428, 429, 430, 431, 432, 433, 434, 435, 436, 437, 438, 439, 440, 441, 442, 443, 444, 445, 446, 447, 448, 449, 450, 451, 452, 453, 454, 455, 456, 457, 458, 459, 460, 461, 462, 463, 464, 465, 466, 467, 468, 469, 470, 471, 472, 473, 474, 475, 476, 477, 478, 479, 480, 481, 482, 483, 484, 485, 486, 487, 488, 489, 490, 491, 492, 493, 494, 495, 496, 497, 498, 499, 500, 501, 502, 503, 504, 505, 506, 507, 508, 509, 510, 511, 512, 513, 514, 515, 516, 517, 518, 519, 520, 521, 522, 523, 524, 525, 526, 527, 528, 529, 530, 531, 532, 533, 534, 535, 536, 537, 538, 539, 540, 541, 542, 543, 544, 545, 546, 547, 548, 549, 550, 551, 552, 553, 554, 555, 556, 557, 558, 559, 560, 561, 562, 563, 564, 565, 566, 237]
            else:
                LoLGame_stat_statistics_output_order = [0, 16, 26, 20, 27, 25, 24, 31, 5, 3, 13, 4, 11, 6, 14, 10, 15, 9, 42, 214, 231, 35, 36, 226, 227, 229, 230, 46, 38, 39, 160, 161, 162, 163, 164, 165, 166, 215, 196, 208, 197, 209, 198, 210, 199, 211, 200, 212, 201, 213, 74, 51, 43, 217, 218, 219, 222, 223, 47, 144, 145, 76, 73, 77, 55, 54, 59, 58, 57, 56, 52, 148, 133, 86, 153, 138, 146, 140, 114, 80, 150, 139, 113, 79, 149, 75, 49, 48, 142, 147, 141, 115, 81, 151, 50, 154, 157, 156, 135, 155, 63, 220, 64, 221, 143, 82, 84, 83, 152, 65, 78, 192, 194, 180, 174, 181, 175, 182, 176, 183, 177, 184, 178, 185, 179, 44, 53, 137, 45, 60, 61, 62, 158, 224, 136, 243, 237, 232, 290, 233, 277, 245, 242, 246, 238, 280, 269, 255, 285, 271, 278, 273, 257, 249, 282, 272, 256, 248, 281, 244, 235, 234, 275, 279, 274, 258, 250, 283, 236, 286, 289, 288, 270, 287, 239, 240, 276, 251, 253, 252, 291, 284, 241, 247, 293, 304, 298, 292, 351, 352, 354, 294, 338, 306, 303, 307, 299, 341, 330, 316, 346, 332, 339, 334, 318, 310, 343, 333, 317, 309, 342, 305, 296, 295, 336, 340, 335, 319, 311, 344, 297, 347, 350, 349, 331, 348, 300, 301, 355, 337, 312, 313, 314, 353, 345, 302, 308]
            LoLGame_stat_data_organized: dict[str, list[Any]] = {LoLGame_stat_header_keys[i]: LoLGame_stat_data[LoLGame_stat_header_keys[i]] for i in LoLGame_stat_statistics_output_order}
            LoLGame_stat_df: pandas.DataFrame = pandas.DataFrame(data = LoLGame_stat_data_organized)
            optimize_bool_display(LoLGame_stat_df)
            LoLGame_stat_df_export = True
            
            LoLGame_stat_self_df = pandas.concat([pandas.DataFrame([LoLGame_stat_header])[LoLGame_stat_df.columns], LoLGame_stat_df[LoLGame_stat_df["puuid"].isin(current_puuid_list)]], ignore_index = True)
            recent_LoLPlayer_df = pandas.concat([pandas.DataFrame([LoLGame_stat_header])[LoLGame_stat_df.columns], LoLGame_stat_df[~(LoLGame_stat_df["puuid"].isin(current_puuid_list))]], ignore_index = True)
            
            if len(LoLMatches_not_found) > 0:
                logPrint("警告：以下%d场对局不存在。\nWarning: The following %d match(es) aren't found." %(len(LoLMatches_not_found), len(LoLMatches_not_found)))
                logPrint(LoLMatches_not_found)
            if len(error_LoLMatchIDs) > 0:
                logPrint("警告：以下%d场对局获取失败。\nWarning: The following %d match(es) fail to be fetched." %(len(error_LoLMatchIDs), len(error_LoLMatchIDs)))
                logPrint(error_LoLMatchIDs)
            matches_to_remove: list[int] = LoLMatches_not_found + error_LoLMatchIDs
            for match_to_remove in matches_to_remove: #在去除获取异常的对局后，需要在对局序号列表中将这些对局也一并移除（After removing matches that fail to be captured, we need to remove them in matchId list, too）
                if match_to_remove in LoLMatchIDs:
                    LoLMatchIDs.remove(match_to_remove)
        else:
            if use_sgp:
                LoLGame_summary_header_keys: list[str] = list(LoLGame_stat_header.keys())
                recent_LoLPlayer_statistics_output_order: list[int] = [0, 112, 148, 131, 132, 146, 128, 147, 68, 21, 16, 13, 25, 26, 11, 18, 22, 14, 29, 15, 20, 30, 19, 24, 227, 218, 628, 184, 54, 625, 626, 96, 133, 125, 82, 152, 135, 52, 51, 55, 223, 224, 186, 187, 188, 189, 190, 191, 192, 221, 200, 212, 201, 213, 202, 214, 203, 215, 204, 216, 205, 217, 95, 64, 45, 229, 230, 231, 234, 235, 98, 94, 99, 49, 72, 71, 74, 73, 66, 167, 129, 113, 174, 153, 164, 157, 115, 102, 169, 156, 114, 101, 168, 97, 61, 60, 58, 59, 161, 162, 166, 158, 159, 116, 103, 170, 62, 176, 179, 178, 136, 177, 65, 79, 232, 80, 233, 93, 57, 163, 105, 155, 160, 171, 172, 83, 84, 106, 108, 173, 85, 107, 67, 47, 109, 110, 100, 48, 56, 78, 130, 127, 111, 43, 44, 104, 69, 70, 175, 63, 46, 81, 154, 165, 137, 139, 141, 142, 228, 144, 145, 602, 616, 608, 604, 609, 605, 610, 606, 611, 607, 620, 618, 621, 619, 598, 596, 597, 50, 149, 150, 75, 76, 77, 182, 181, 180, 236, 117, 143, 672, 658, 643, 728, 674, 671, 675, 647, 660, 715, 692, 687, 721, 701, 712, 705, 689, 678, 717, 704, 688, 677, 716, 673, 655, 654, 652, 653, 709, 710, 714, 706, 707, 690, 679, 718, 656, 723, 726, 725, 694, 724, 659, 665, 666, 670, 651, 711, 681, 703, 708, 729, 719, 720, 668, 669, 682, 683, 661, 645, 684, 685, 676, 646, 650, 664, 693, 691, 686, 641, 642, 680, 662, 663, 722, 657, 644, 667, 702, 713, 695, 696, 697, 698, 727, 699, 700, 802, 750, 749, 773, 759, 744, 830, 831, 833, 775, 772, 776, 748, 761, 817, 793, 788, 823, 803, 814, 807, 790, 779, 819, 806, 789, 778, 818, 774, 756, 755, 753, 754, 811, 812, 816, 808, 809, 791, 780, 820, 757, 825, 828, 827, 795, 826, 760, 766, 767, 834, 771, 752, 813, 782, 810, 805, 832, 821, 822, 769, 770, 762, 746, 785, 786, 783, 784, 777, 747, 751, 765, 794, 792, 787, 742, 743, 781, 763, 764, 824, 758, 745, 768, 804, 815, 796, 797, 798, 799, 829, 800, 801, 238, 239, 240, 241, 242, 243, 244, 245, 246, 247, 248, 249, 250, 251, 252, 253, 254, 255, 256, 257, 258, 259, 260, 261, 262, 263, 264, 265, 266, 267, 268, 269, 270, 381, 271, 272, 273, 274, 275, 276, 277, 278, 279, 280, 281, 282, 382, 283, 284, 285, 383, 286, 287, 288, 289, 290, 291, 292, 293, 294, 295, 296, 297, 298, 299, 300, 301, 302, 303, 304, 305, 306, 307, 308, 309, 310, 311, 312, 313, 314, 315, 316, 317, 318, 319, 384, 320, 321, 322, 323, 324, 325, 326, 327, 328, 329, 330, 385, 331, 332, 333, 334, 335, 336, 337, 338, 339, 340, 341, 342, 343, 344, 345, 346, 347, 348, 349, 350, 351, 352, 353, 354, 355, 356, 357, 358, 359, 360, 361, 362, 363, 364, 365, 366, 367, 368, 369, 370, 371, 372, 373, 374, 375, 376, 377, 378, 379, 380, 386, 387, 388, 389, 390, 391, 392, 393, 394, 395, 396, 397, 398, 399, 400, 401, 402, 403, 404, 405, 406, 407, 408, 409, 410, 411, 412, 413, 414, 415, 416, 417, 418, 419, 420, 421, 422, 423, 424, 425, 426, 427, 428, 429, 430, 431, 432, 433, 434, 435, 436, 437, 438, 439, 440, 441, 442, 443, 444, 445, 446, 447, 448, 449, 450, 451, 452, 453, 454, 455, 456, 457, 458, 459, 460, 461, 462, 463, 464, 465, 466, 467, 468, 469, 470, 471, 472, 473, 474, 475, 476, 477, 478, 479, 480, 481, 482, 483, 484, 485, 486, 487, 488, 489, 490, 491, 492, 493, 494, 495, 496, 497, 498, 499, 500, 501, 502, 503, 504, 505, 506, 507, 508, 509, 510, 511, 512, 513, 514, 515, 516, 517, 518, 519, 520, 521, 522, 523, 524, 525, 526, 527, 528, 529, 530, 531, 532, 533, 534, 535, 536, 537, 538, 539, 540, 541, 542, 543, 544, 545, 546, 547, 548, 549, 550, 551, 552, 553, 554, 555, 556, 557, 558, 559, 560, 561, 562, 563, 564, 565, 566, 237]
                recent_LoLPlayer_data_organized: dict[str, list[Any]] = {LoLGame_summary_header_keys[i]: [LoLGame_stat_header[LoLGame_summary_header_keys[i]]] for i in recent_LoLPlayer_statistics_output_order}
            else:
                LoLGame_summary_header_keys: list[str] = list(LoLGame_stat_header.keys())
                recent_LoLPlayer_statistics_output_order: list[int] = [0, 16, 26, 20, 27, 25, 24, 31, 5, 3, 13, 4, 11, 6, 14, 10, 15, 9, 42, 214, 231, 35, 36, 226, 227, 229, 230, 46, 38, 39, 160, 161, 162, 163, 164, 165, 166, 215, 196, 208, 197, 209, 198, 210, 199, 211, 200, 212, 201, 213, 74, 51, 43, 217, 218, 219, 222, 223, 47, 144, 145, 76, 73, 77, 55, 54, 59, 58, 57, 56, 52, 148, 133, 86, 153, 138, 146, 140, 114, 80, 150, 139, 113, 79, 149, 75, 49, 48, 142, 147, 141, 115, 81, 151, 50, 154, 157, 156, 135, 155, 63, 220, 64, 221, 143, 82, 84, 83, 152, 65, 78, 192, 194, 180, 174, 181, 175, 182, 176, 183, 177, 184, 178, 185, 179, 44, 53, 137, 45, 60, 61, 62, 158, 224, 136, 243, 237, 232, 290, 233, 277, 245, 242, 246, 238, 280, 269, 255, 285, 271, 278, 273, 257, 249, 282, 272, 256, 248, 281, 244, 235, 234, 275, 279, 274, 258, 250, 283, 236, 286, 289, 288, 270, 287, 239, 240, 276, 251, 253, 252, 291, 284, 241, 247, 293, 304, 298, 292, 351, 352, 354, 294, 338, 306, 303, 307, 299, 341, 330, 316, 346, 332, 339, 334, 318, 310, 343, 333, 317, 309, 342, 305, 296, 295, 336, 340, 335, 319, 311, 344, 297, 347, 350, 349, 331, 348, 300, 301, 355, 337, 312, 313, 314, 353, 345, 302, 308]
                recent_LoLPlayer_data_organized: dict[str, list[Any]] = {LoLGame_summary_header_keys[i]: [LoLGame_stat_header[LoLGame_summary_header_keys[i]]] for i in recent_LoLPlayer_statistics_output_order}
            recent_LoLPlayer_df: pandas.DataFrame = pandas.DataFrame(data = recent_LoLPlayer_data_organized)
        
        #整理云顶之弈对局记录（Organize TFT match history）
        logPrint("是否查询云顶之弈对局记录？（输入任意键查询，否则不查询）\nSearch TFT matches? (Input anything to search or null to export data or switch for another summoner)")
        search_TFT_str: str = logInput()
        search_TFT: bool = bool(search_TFT_str)
        TFTMatchIDs: list[int] = []
        old_TFTMatch_detected: bool = len(saved_TFTMatchIDs) > 0
        update_unsaved_only_tft: bool = False
        if search_TFT:
            TFTHistory_dfs: list[pandas.DataFrame] = []
            for i in range(len(AllAccounts)):
                queues = queues_initial.copy()
                TFTAugments = TFTAugments_initial.copy()
                TFTChampions = TFTChampions_initial.copy()
                TFTItems = TFTItems_initial.copy()
                TFTCompanions = TFTCompanions_initial.copy()
                TFTTraits = TFTTraits_initial.copy()
                current_versions["queue"] = current_versions["TFTAugment"] = current_versions["TFTChampion"] = current_versions["TFTItem"] = current_versions["TFTCompanion"] = current_versions["TFTTrait"] = URLPatch
                unmapped_keys["queue"], unmapped_keys["TFTAugment"], unmapped_keys["TFTChampion"], unmapped_keys["TFTItem"], unmapped_keys["TFTCompanion"], unmapped_keys["TFTTrait"] = set(), set(), set(), set(), set(), set()
                info_puuid: str = current_puuid_list[i]
                info_summonerName: str = current_summonerName_list[i]
                logPrint("[%d/%d]正在获取客户端内玩家%s的云顶之弈对局记录……\nGetting TFT match history of player %s in the client ..." %(i + 1, len(AllAccounts), info_summonerName, info_summonerName))
                TFTHistory_get, TFTHistory = await get_TFTHistory(connection, info_puuid, log = log)
                json09name: str = "Match History (TFT) - " + displayName + ".json"
                json09path: str = os.path.join(folder, json09name)
                os.makedirs(folder, exist_ok = True)
                try:
                    with open(json09path, "w", encoding = "utf-8") as jsonfile09:
                        jsonfile09.write(json.dumps(TFTHistory, indent = 4, ensure_ascii = False))
                except UnicodeEncodeError:
                    logPrint("召唤师云顶之弈对局记录文本文档生成失败！请检查召唤师名称和所选语言是否包含不常用字符！\nSummoner TFT match history text generation failure! Please check if the summoner name and the chosen language include any abnormal characters!\n")
                # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
                # pkl6name: str = f"Intermediate Object - TFTHistory - {displayName} ({currentTime}).pkl"
                # with open(os.path.join(folder, pkl6name), "wb") as IntObj5:
                #     pickle.dump(TFTHistory, IntObj5)
                logPrint("[%d/%d]正在扩展玩家%s的云顶之弈对局记录……\nExpanding TFT match history of player %s ..." %(i + 1, len(AllAccounts), info_summonerName, info_summonerName))
                TFTHistory_get, TFTHistory = await get_matchSummary_sgp(connection, sgpSession, info_puuid, "TFT", begin = 0, count = 1000, log = log)
                for game in TFTHistory["games"]:
                    matchId: int = int(game["metadata"]["match_id"].split("_")[1])
                    if not matchId in TFTGame_summary_cache_sgp: #由于云顶之弈的对局记录包含所有玩家的信息，所以如果多个玩家的对局记录包含同一场对局，则这些对局的信息一定是相同的（Because TFT match history includes all players' information, if a match is included in multiple players' match histories, then information of the matches recorded in different players' match histories must be the same）
                        TFTGame_summary_cache_sgp[matchId] = game
                if TFTHistory_get:
                    TFTGamePlayed_singleSummoner: bool = (TFTGameCount := len(TFTHistory["games"])) > 0 #标记该玩家是否进行过云顶之弈对局（Mark whether this summoner has played any TFT game）
                    if TFTGamePlayed_singleSummoner:
                        logPrint(f"玩家{info_summonerName}共进行{TFTGameCount}场云顶之弈对局。\nPlayer {info_summonerName} has played {TFTGameCount} TFT matches.\n")
                    else:
                        logPrint(f"玩家{info_summonerName}从5月1日起就没有进行过任何云顶之弈对局。\nPlayer {info_summonerName} hasn't played any TFT game yet since May 1st.")
                    TFTHistory_df: pandas.DataFrame = (await sort_TFTHistory(connection, TFTHistory, info_puuid, queues, TFTAugments, TFTChampions, TFTItems, TFTCompanions, TFTTraits, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, session = session, useInfoDict = True, infos = infos, log = log))[0]
                    TFTHistory_dfs.append(TFTHistory_df)
                    if TFTGamePlayed_singleSummoner:
                        logPrint(TFTHistory_df[:min(21, TFTGameCount + 1)], write_time = False)
            #由于云顶之弈的对局记录包含所有玩家的信息，所以这里考虑先整合所有账号的对局记录，再对总对局记录进行整理。如果先整理再整合，后续排序时玩家顺序的信息会丢失，因为在这种情形下根据对局序号排序，而数据框中不包含玩家序号键，无法按照玩家序号进行升序排列（Because TFT match history includes all players' information, here the program first merges all accounts' match history, and then sort out the aggregate match history. Otherwise, if the program first organize the match history respectively and then merge the result dataframe, the participantId order may be lost during the subsequent ordering, for gameId is taken to arrange the aggregate dataframe, but the key `participantId` isn't in the dataframe, and therefore the dataframe can't be arranged in the ascending order of participantId）
            # TFTHistory_all: dict[str, str | list[dict[str, Any]]] = {"active_puuid": "", "games": list(map(lambda x: TFTGame_summary_cache_sgp[x], TFTGameIDs))}
            #构建云顶之弈对局记录数据框（Construct TFT match history dataframe）
            TFTHistory_df_all: pandas.DataFrame = pandas.concat([TFTHistory_dfs[0].iloc[:1]] + list(map(lambda x: x.iloc[1:], TFTHistory_dfs)), ignore_index = True) #需要注意数据框的中文表头占用了一行（Note that the Chinese header takes up a record）
            gameIds_occurred: set[str | int] = {TFTHistory_df_all["game_id"][0]}
            lines_to_drop: list[int] = []
            for i in range(1, len(TFTHistory_df_all)):
                if TFTHistory_df_all["game_id"][i] in gameIds_occurred:
                    lines_to_drop.append(i)
                else:
                    gameIds_occurred.add(TFTHistory_df_all["game_id"][i])
            TFTHistory_df_all.drop(lines_to_drop, inplace = True)
            TFTHistory_df_all = TFTHistory_df_all.reset_index(drop = True)
            TFTHistory_df_all = pandas.concat([TFTHistory_df_all.iloc[:1], TFTHistory_df_all.iloc[1:].sort_values(by = "gameCreationDate", ascending = False)], ignore_index = True) #这里弃用了根据对局序号排序（Here gameId isn't used to sort the values）
            
            logPrint('请输入要查询的云顶之弈对局序号，批量查询对局请输入对局序号列表，批量查询全部对局请输入“3”，退出云顶之弈对局查询请输入“0”：\nPlease enter the TFT match ID to check. Submit a list containing matchIDs to search in batch. Submit "3" to search the currently stored history in batch. Submit "0" to quit searching for TFT matches.')
            TFTGameIDs: list[int] = TFTHistory_df_all["game_id"][1:].to_list()
            old_TFTMatch_detected: bool = len(saved_TFTMatchIDs) > 0
            while True:
                matchId_str: str = logInput()
                if matchId_str == "":
                    continue
                elif matchId_str == "0":
                    TFTMatchIDs = []
                    break
                else:
                    if matchId_str == "3":
                        if old_TFTMatch_detected:
                            latest_TFTMatchID: int = max(saved_TFTMatchIDs)
                            latest_TFTMatchID_index: int = TFTGameIDs.index(latest_TFTMatchID) if latest_TFTMatchID in TFTGameIDs else 500
                            logPrint("检测到您以前曾经查询过该召唤师的云顶之弈对局记录。是否只保存该召唤师信息文件夹中不包含的云顶之弈对局？（输入空字符串以只保存未保存过文本文档的对局，否则自行指定对局索引上下限）\nThe program detected that you've searched for this summoner's TFT match history before. Do you want to only save the TFT matches not present in the current summoner folder? (Enter an empty string to saved only the matches whose json files haven't been saved, or any non-empty string to specify the begIndex and endIndex of the matches by yourself)\n即将使用的起始索引和总对局数（The beginning index and total game count to be used）：0 %d" %latest_TFTMatchID_index)
                            update_unsaved_only_tft_str: str = logInput()
                            update_unsaved_only_tft: bool = not bool(update_unsaved_only_tft_str)
                        if old_TFTMatch_detected and update_unsaved_only_tft:
                            TFTMatchIDs = TFTGameIDs[:]
                        else:
                            begIndex: int = 0
                            endIndex: int = 10000
                            logPrint("请设置需要查询的对局索引下界和上界，以空格为分隔符（输入空字符以默认查询近20场对局）：\nPlease set the begIndex and endIndex of the matches to be searched, split by space (Enter an empty string to search for the recent 20 matches):") #在13.13版本以前，腾讯代理的服务器只支持近20场对局查询（Before Patch 13.13, Tencent servers only provide search of the latest 20 matches）
                            while True:
                                gameIndex: str = logInput()
                                if gameIndex == "":
                                    begIndex, endIndex = 0, 200 * len(AllAccounts)
                                    break
                                elif gameIndex == "0":
                                    break
                                else:
                                    try:
                                        begIndex, endIndex = map(int, gameIndex.split())
                                    except ValueError:
                                        logPrint("请以空格为分隔符输入对局索引的自然数类型的下界和上界！\nPlease enter the two nonnegative integers as the begIndex and endIndex of the matches split by space!")
                                        continue
                                    else:
                                        break
                            if gameIndex == "0":
                                search_TFT = False
                                TFTMatchIDs = []
                                break
                            TFTMatchIDs = TFTGameIDs[begIndex:endIndex]
                    elif matchId_str == "scan":
                        TFTMatchIDs = TFTGameIDs + saved_TFTMatchIDs
                        if TFTMatchIDs == list():
                            logPrint("尚未保存过该玩家的数据！\nYou haven't saved this summoner's matches yet!\n")
                            break
                        else:
                            TFTMatchIDs = sorted(set(TFTMatchIDs), reverse = True)
                            logPrint("检测到%d场对局。是否继续？（输入任意键以重新输入要查询的对局序号，否则重新获取这些对局的数据）\nDetected %d matches. Continue? (Input any non-empty string to return to the last step of inputting the matchId, or null to recapture those matches' data)" %(len(TFTMatchIDs), len(TFTMatchIDs)))
                            recapture_str: str = logInput()
                            recapture: bool = bool(recapture_str)
                            if recapture:
                                TFTMatchIDs = []
                                logPrint('请输入要查询的云顶之弈对局序号，批量查询对局请输入对局序号列表，批量查询全部对局请输入“3”，退出云顶之弈对局查询请输入“0”：\nPlease enter the TFT match ID to check. Submit a list containing matchIDs to search in batch. Submit "3" to search the currently stored history in batch. Submit "0" to quit searching for TFT matches.')
                                continue
                            scan_tft = True
                            TFTAugments = TFTAugments_initial.copy()
                            TFTChampions = TFTChampions_initial.copy()
                            TFTItems = TFTItems_initial.copy()
                            TFTCompanions = TFTCompanions_initial.copy()
                            TFTTraits = TFTTraits_initial.copy()
                            current_versions["TFTAugment"] = current_versions["TFTChampion"] = current_versions["TFTItem"] = current_versions["TFTCompanion"] = current_versions["TFTTrait"] = URLPatch
                            unmapped_keys["TFTAugment"], unmapped_keys["TFTChampion"], unmapped_keys["TFTItem"], unmapped_keys["TFTCompanion"], unmapped_keys["TFTTrait"] = set(), set(), set(), set(), set()
                            TFTHistory_df_all = (await reconstruct_TFTHistory(connection, sgpSession, TFTMatchIDs, current_puuid, queues, TFTAugments, TFTChampions, TFTItems, TFTCompanions, TFTTraits, useAllVersions = True, versionList = bigPatches, locale = language_code, TFTGame_summary_cache = TFTGame_summary_cache_sgp, log = log))[0]
                            logPrint("是否一同保存每场对局的信息？（输入任意键保存，否则将只导出对局记录）\nSave each match? (Input anything to save each match, or null to only save the scanned match history)")
                            sort_gameInfo_sync_str: str = logInput()
                            sort_gameInfo_sync: bool = bool(sort_gameInfo_sync_str)
                            if not sort_gameInfo_sync:
                                break
                    else:
                        try:
                            tmp = eval(matchId_str)
                            TFTMatchIDs = []
                            if isinstance(tmp, int):
                                TFTMatchIDs.append(tmp)
                            elif isinstance(tmp, list) and all(map(lambda x: isinstance(x, int), tmp)):
                                TFTMatchIDs += tmp
                            else:
                                logPrint("数据类型不符。请重新输入！\nData type not matched! Please try again!")
                                continue
                            if len(TFTMatchIDs) == 0:
                                logPrint("您输入的对局序号集无效！请重新输入。\nInvalid matchId set! Please try again.")
                                continue
                        except (SyntaxError, NameError):
                            logPrint("您的输入存在语法错误。请重新输入！\nSyntax ERROR detected in this input! Please try again!")
                            continue
                    break
        #整理云顶之弈对局信息（Organize TFT match information）
        TFTGamePlayed: bool = len(TFTMatchIDs) > 0
        if TFTGamePlayed:
            TFTMatches_not_found: list[int] = []
            error_TFTMatchIDs: list[int] = []
            queues = queues_initial.copy()
            TFTAugments = TFTAugments_initial.copy()
            TFTChampions = TFTChampions_initial.copy()
            TFTItems = TFTItems_initial.copy()
            TFTCompanions = TFTCompanions_initial.copy()
            TFTTraits = TFTTraits_initial.copy()
            current_versions["queue"] = current_versions["TFTAugment"] = current_versions["TFTChampion"] = current_versions["TFTItem"] = current_versions["TFTCompanion"] = current_versions["TFTTrait"] = URLPatch
            unmapped_keys["queue"], unmapped_keys["TFTAugment"], unmapped_keys["TFTChampion"], unmapped_keys["TFTItem"], unmapped_keys["TFTCompanion"], unmapped_keys["TFTTrait"] = set(), set(), set(), set(), set(), set()
            logPrint("是否输出每场对局的文本文档？（输入任意键不输出，否则默认输出）\nExport text files of each match? (Input anything to cancel, or null to export by default)")
            save_all_json_str: str = logInput()
            save_all_json: bool = not bool(save_all_json_str)
            TFTGame_stat_header_keys: list[str] = list(TFTGame_stat_header.keys())
            TFTGame_stat_data: dict[str, list[Any]] = {key: [] for key in TFTGame_stat_header_keys}
            for matchId in TFTMatchIDs:
                match_id: str = f"{platformId}_{matchId}"
                TFTGame_summary_export: bool = not (old_TFTMatch_detected and update_unsaved_only_tft and matchId in saved_TFTMatchIDs)
                TFTGame_leaderboard_export: bool = args.export_leaderboard
                info_text_saved: bool = False
                isTFT[matchId] = False #前面部分对局即使添加到此字典中，其值也是False（Even if some matches are added into this dictionary previously, their values are still False）
                
                #获取数据（Get data）
                if matchId in TFTGame_summary_cache_sgp:
                    TFTGame_summary: dict[str, Any] = TFTGame_summary_cache_sgp[matchId]
                    status: int = 200
                else:
                    status, TFTGame_summary = await get_game_summary_sgp(connection, sgpSession, match_id, checkLoL = False, checkTFT = True, log = log) #通过LCU API和SGP API获取到的云顶之弈对局记录和对局概要是相同的（TFT match history and TFT game summary obtained through LCU API and SGP API are the same）
                    if status == 200:
                        TFTGame_summary_cache_sgp[matchId] = TFTGame_summary
                if status == 200 and "json" in TFTGame_summary and bool(TFTGame_summary["json"]):
                    isTFT[matchId] = True
                    info_exist_error[matchId] = False
                    save_one_json: bool = TFTGame_summary_export
                    participant_puuid: list[str] = []
                    participant_gameName: list[str] = []
                    for participant in TFTGame_summary["json"]["participants"]:
                        participant_puuid.append(participant["puuid"])
                        if participant["puuid"] != "00000000-0000-0000-0000-000000000000":
                            if "riotIdGameName" in participant and "riotIdTagline" in participant:
                                TFTPlayer_summonerName: str = "%s#%s" %(participant["riotIdGameName"], participant["riotIdTagline"])
                            else:
                                if participant["puuid"] in infos:
                                    TFTPlayer_info_body = infos[participant["puuid"]]
                                    TFTPlayer_summonerName = get_info_name(TFTPlayer_info_body)
                                    TFTPlayer_info_got: bool = True
                                else:
                                    TFTPlayer_info_recapture = 0
                                    TFTPlayer_info = await get_info(connection, participant["puuid"])
                                    while not TFTPlayer_info["info_got"] and TFTPlayer_info["body"]["httpStatus"] != 404 and TFTPlayer_info_recapture < 3:
                                        logPrint(TFTPlayer_info["message"])
                                        TFTPlayer_info_recapture += 1
                                        logPrint("对局%d玩家信息（玩家通用唯一识别码：%s）获取失败！正在第%d次尝试重新获取该玩家信息……\nInformation of player (puuid: %s) in Match %d capture failed! Recapturing this player's information ... Times tried: %d." %(TFTGame_summary["json"]["game_id"], participant["puuid"], TFTPlayer_info_recapture, participant["puuid"], TFTGame_summary["json"]["game_id"], TFTPlayer_info_recapture))
                                        TFTPlayer_info = await get_info(connection, participant["puuid"])
                                    if TFTPlayer_info["info_got"]:
                                        TFTPlayer_info_body = TFTPlayer_info["body"]
                                        infos[participant["puuid"]] = TFTPlayer_info_body
                                        TFTPlayer_summonerName = get_info_name(TFTPlayer_info_body)
                                    else:
                                        logPrint(TFTPlayer_info["message"])
                                        logPrint("对局%d玩家信息（玩家通用唯一识别码：%s）获取失败！\nInformation of player (puuid: %s) in Match %d capture failed!" %(TFTGame_summary["json"]["game_id"], participant["puuid"], participant["puuid"], TFTGame_summary["json"]["game_id"]))
                                        TFTPlayer_summonerName = participant["puuid"] 
                                    TFTPlayer_info_got = TFTPlayer_info["info_got"]
                        else:
                            TFTPlayer_summonerName = participant["puuid"] #注意到新玩家的召唤师名是空字符串（Note that a new player's summoner name would also be an empty string）
                        participant_gameName.append(TFTPlayer_summonerName)
                    if len(set(current_puuid_list) & set(participant_puuid)) > 0:
                        main_player_included[matchId] = True
                        match_reserve_strategy[matchId] = True
                    elif len(set(current_summonerName_list) & set(participant_gameName)) > 0:
                        main_player_included[matchId] = True
                        match_reserve_strategy[matchId] = True
                        if not puuid_change_warning_printed:
                            logPrint("警告：该大区的玩家通用唯一识别码曾发生变动！请检查保存的各对局是否属于该玩家。\nWarning: The puuids of players on this server have been changed! Please check if the saved matches really belong to this player.")
                            puuid_change_warning_printed = True
                    else:
                        main_player_included[matchId] = False
                        save_one_json: bool = args.reserve_text #由于从文本文件中可以提取该召唤师的对局序号，所以需要保证保留下来的文本文件都包含该召唤师。因此，如果一场对局不包含该召唤师，就不应该把这场对局保存下来，除非用户出于特殊目的需要保留文本文件（Because a summoner's matchIDs can be extracted from the saved json files, it needs to be guaranteed that all saved json files belong to this summoner. Therefore, if a match doesn't include this summoner, then it shouldn't be saved into json files, unless the user must save it with special purposes）
                        if args.reserve:
                            match_reserve: bool = True
                            logPrint("[%d/%d]对局%d不包含该玩家！已保持该对局。\nMatch %d doesn't include the current player but is reserved." %(TFTMatchIDs.index(matchId) + 1, len(TFTMatchIDs), matchId, matchId))
                        else:
                            if not match_notbelonging_warning_printed:
                                logPrint("警告：对局%d不包含该玩家！是否仍要保持该对局？（输入任意键以保留该对局，否则舍弃该对局）\nWarning: The Match %d doesn't include the current player! Continue? (Input any nonempty string to reserve this match, or null to abandon it.)\n注意：此改动对于后续情形也生效。\nNote: This decision takes effect in similar situations later." %(matchId, matchId))
                                match_reserve_str: str = logInput()
                                match_reserve = bool(match_reserve_str)
                                match_notbelonging_warning_printed = True
                            elif match_reserve:
                                logPrint("[%d/%d]对局%d不包含该玩家！已保持该对局。\nMatch %d doesn't include the current player but is reserved." %(TFTMatchIDs.index(matchId) + 1, len(TFTMatchIDs), matchId, matchId))
                            else:
                                logPrint("[%d/%d]对局%d不包含该玩家！已舍弃该对局。\nMatch %d doesn't include the current player and is decrepated." %(TFTMatchIDs.index(matchId) + 1, len(TFTMatchIDs), matchId, matchId))
                        match_reserve_strategy[matchId] = match_reserve
                else:
                    TFTGame_summary_export: bool = False
                    TFTGame_leaderboard_export: bool = False
                    participant_puuid = []
                    info_exist_error[matchId] = True
                    save_one_json = False
                
                #提示（Prompt）
                info_note: str = "" if "json" in TFTGame_summary and bool(TFTGame_summary["json"]) else " (Match data deleted from API!)"
                process_header: str = "保存进度（Saving process）" if save_all_json and TFTGame_summary_export and save_one_json else "加载进度（Loading process）"
                logPrint("%s：%d/%d\t对局序号（MatchID）： %d%s" %(process_header, TFTMatchIDs.index(matchId) + 1, len(TFTMatchIDs), matchId, info_note), print_time = True)
                
                #导出数据（Export data）
                if "json" in TFTGame_summary and bool(TFTGame_summary["json"]):
                    if save_all_json and save_one_json:
                        json12name: str = f"Match Summary (TFT) - {platformId}-{matchId}.json"
                        os.makedirs(match_folder, exist_ok = True)
                        try:
                            with open(os.path.join(match_folder, json12name), "w", encoding = "utf-8") as jsonfile12:
                                jsonfile12.write(json.dumps(TFTGame_summary, indent = 4, ensure_ascii = False))
                        except UnicodeDecodeError:
                            logPrint("对局%d概要文本文档生成失败！请检查召唤师名称是否包含不常用字符！\nMatch %d summary text generation failure! Please check if the summoner name includes any abnormal characters!" %(matchId, matchId))
                        else:
                            info_text_saved = True
                            TFTMatches_exported.append(matchId)
                        # currentTime: str = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime())
                        # pkl9name: str = f"Intermediate Object - Match Summary (TFT) - {platformId}-{matchId}.pkl"
                        # with open(os.path.join(match_folder, pkl9name), "wb") as IntObj8:
                        #     pickle.dump(TFTGame_summary, IntObj8)
                    
                    TFTGame_summary_df, queues, TFTAugments, TFTChampions, TFTItems, TFTCompanions, TFTTraits = await sort_TFTGame_summary(connection, TFTGame_summary, queues, TFTAugments, TFTChampions, TFTItems, TFTCompanions, TFTTraits, gameIndex = TFTMatchIDs.index(matchId), current_puuid = current_puuid, useAllVersions = True, versionList = bigPatches, locale = language_code, current_versions = current_versions, unmapped_keys = unmapped_keys, session = session, useInfoDict = True, infos = infos, sortStats = True, TFTGame_stat_data = TFTGame_stat_data, save_self = True, save_other = True, log = log)
                    
                    #社交排行榜（Social leaderboard）
                    if args.export_leaderboard:
                        TFTGame_leaderboard_df: pandas.DataFrame = await sort_game_leaderboard(connection, puuids = participant_puuid, log = log)
                    else:
                        TFTGame_leaderboard_df = pandas.DataFrame()
                else:
                    TFTMatches_not_found.append(matchId)
                    TFTGame_summary_error: dict[str, list[str]] = {"项目": list(error_header.values()), "items": list(error_header.keys()), "值": [TFTGame_summary.get(key, "") for key in error_header_keys]}
                    TFTGame_summary_df: pandas.DataFrame = pandas.DataFrame(data = TFTGame_summary_error)
                    TFTGame_leaderboard_df = pandas.DataFrame()
                #注意到对于对局出现异常的情况，英雄联盟对局概要数据框的构建方式和云顶之弈有所不同。这是因为当云顶之弈对局概要获取异常时，往往是其“json”键为空或者无“json”键，而没有详细报错信息（Note that when an error occurs to a match, the method of creating the LoL match summary dataframe is different from that of creating the TFT match summary dataframe. This is because when a TFT match summary fails to be loaded, either its "json" value is null, or its "json" key is missing, without a detailed error information）
                
                #云顶之弈无时间轴（TFT games don't have timeline）
                timeline_exist_error[matchId] = True #云顶之弈对局中没有时间轴信息，因此每个云顶之弈对局的时间轴标记为异常获取（There's no timeline information in each TFT match, so each TFT match's timeline is labeled as "error" captured）

                if TFTGame_leaderboard_export:
                    game_leaderboard_dfs[matchId] = TFTGame_leaderboard_df.copy(deep = True)
                if TFTGame_summary_export:
                    game_summary_dfs[matchId] = TFTGame_summary_df.copy(deep = True)
            
            TFTGame_stat_statistics_output_order: list[int] = [0, 19, 46, 47, 43, 5, 14, 15, 16, 6, 10, 18, 7, 13, 11, 12, 307, 305, 40, 55, 33, 34, 35, 38, 52, 53, 49, 36, 50, 42, 54, 41, 39, 44, 45, 23, 24, 25, 150, 148, 149, 203, 206, 209, 155, 153, 154, 212, 215, 218, 160, 158, 159, 221, 224, 227, 165, 163, 164, 230, 233, 236, 170, 168, 169, 239, 242, 245, 175, 173, 174, 248, 251, 254, 180, 178, 179, 257, 260, 263, 185, 183, 184, 266, 269, 272, 190, 188, 189, 275, 278, 281, 195, 193, 194, 284, 287, 290, 200, 198, 199, 293, 296, 299, 61, 57, 58, 59, 60, 68, 64, 65, 66, 67, 75, 71, 72, 73, 74, 82, 78, 79, 80, 81, 89, 85, 86, 87, 88, 96, 92, 93, 94, 95, 103, 99, 100, 101, 102, 110, 106, 107, 108, 109, 117, 113, 114, 115, 116, 124, 120, 121, 122, 123, 131, 127, 128, 129, 130, 138, 134, 135, 136, 137, 145, 141, 142, 143, 144]
            TFTGame_stat_data_organized: dict[str, list[Any]] = {TFTGame_stat_header_keys[i]: TFTGame_stat_data[TFTGame_stat_header_keys[i]] for i in TFTGame_stat_statistics_output_order}
            TFTGame_stat_df: pandas.DataFrame = pandas.DataFrame(data = TFTGame_stat_data_organized)
            optimize_bool_display(TFTGame_stat_df)
            
            TFTGame_stat_self_df = pandas.concat([pandas.DataFrame([TFTGame_stat_header])[TFTGame_stat_df.columns], TFTGame_stat_df[TFTGame_stat_df["puuid"].isin(current_puuid_list)]], ignore_index = True)
            recent_TFTPlayer_df = pandas.concat([pandas.DataFrame([TFTGame_stat_header])[TFTGame_stat_df.columns], TFTGame_stat_df[~(TFTGame_stat_df["puuid"].isin(current_puuid_list))]], ignore_index = True)
            
            if len(TFTMatches_not_found) > 0:
                logPrint("警告：以下%d场对局不存在。\nWarning: The following %d match(es) aren't found." %(len(TFTMatches_not_found), len(TFTMatches_not_found)))
                logPrint(TFTMatches_not_found)
            if len(error_TFTMatchIDs) > 0:
                logPrint("警告：以下%d场对局获取失败。\nWarning: The following %d match(es) fail to be fetched." %(len(error_TFTMatchIDs), len(error_TFTMatchIDs)))
                logPrint(error_TFTMatchIDs)
            matches_to_remove: list[int] = TFTMatches_not_found + error_TFTMatchIDs
            for match_to_remove in matches_to_remove:
                TFTMatchIDs.remove(match_to_remove)
        else:
            TFTGame_stat_header_keys: list[str] = list(TFTGame_stat_header.keys())
            recent_TFTPlayer_statistics_output_order: list[int] = [0, 19, 46, 47, 43, 5, 14, 15, 16, 6, 10, 18, 7, 13, 11, 12, 307, 305, 40, 55, 33, 34, 35, 38, 52, 53, 49, 36, 50, 42, 54, 41, 39, 44, 45, 23, 24, 25, 150, 148, 149, 203, 206, 209, 155, 153, 154, 212, 215, 218, 160, 158, 159, 221, 224, 227, 165, 163, 164, 230, 233, 236, 170, 168, 169, 239, 242, 245, 175, 173, 174, 248, 251, 254, 180, 178, 179, 257, 260, 263, 185, 183, 184, 266, 269, 272, 190, 188, 189, 275, 278, 281, 195, 193, 194, 284, 287, 290, 200, 198, 199, 293, 296, 299, 61, 57, 58, 59, 60, 68, 64, 65, 66, 67, 75, 71, 72, 73, 74, 82, 78, 79, 80, 81, 89, 85, 86, 87, 88, 96, 92, 93, 94, 95, 103, 99, 100, 101, 102, 110, 106, 107, 108, 109, 117, 113, 114, 115, 116, 124, 120, 121, 122, 123, 131, 127, 128, 129, 130, 138, 134, 135, 136, 137, 145, 141, 142, 143, 144]
            recent_TFTPlayer_data_organized: dict[str, list[Any]] = {TFTGame_stat_header_keys[i]: [TFTGame_stat_header[TFTGame_stat_header_keys[i]]] for i in recent_TFTPlayer_statistics_output_order}
            recent_TFTPlayer_df: pandas.DataFrame = pandas.DataFrame(data = recent_TFTPlayer_data_organized)
        
        if search_LoL and LoLGamePlayed or search_TFT and TFTGamePlayed:
            analyze_recently_played_summoners(search_LoL, search_TFT, recent_LoLPlayer_df, recent_TFTPlayer_df, gameQueues, displayName, folder)
        
        #计算每场对局要保存的工作表数量（Calculate the number of sheets to be saved for each match）
        matchIDs: list[int] = list(game_summary_dfs.keys())
        matchIDs.sort()
        logPrint("正在计算每场对局要保存的工作表数量……\nCalculating the number of sheets to be saved for each match ...\n")
        sheetNumber: dict[int, int] = {}
        for i in range(len(matchIDs)):
            if not match_reserve_strategy[matchIDs[i]]:
                sheetNumber[matchIDs[i]] = 0
            else:
                sheetNumber[matchIDs[i]] = (1 - info_exist_error[matchIDs[i]]) + 2 * (1 - timeline_exist_error[matchIDs[i]])
        
        #导出数据（Export data）
        logPrint("是否导出以上召唤师数据至Excel中？（输入任意键导出，否则不导出）\nDo you want to export the above data into Excel? (Press any key to export or null to refuse exporting)")
        export_str: str = logInput()
        export: bool = bool(export_str)
        if export:
            if len(matchIDs) > 0 and sum(sheetNumber.values()) > 0:
                logPrint("是否导出所有对局的详细信息？注意，这可能需要一定时间。（输入任意键导出，否则不导出。）\nDo you want to export detailed information of each match? Note that this may take some time. (Submit any non-empty string to export, or null to refuse exporting.)")
                detail_export_str: str = logInput()
                detail_export: bool = bool(detail_export_str)
            else:
                detail_export = False
            wbName: str = f"Summoner Profile - {displayName}.xlsx"
            wbPath: str = os.path.join(folder, wbName)
            wbName_sorted: str = f"Summoner Profile - {displayName} (sorted).xlsx"
            os.makedirs(folder, exist_ok = True)
            if not os.path.exists(wbPath):
                wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath), sheet1_name = "Profile", log = log) #通过使用系统自带的Excel应用创建工作簿，使得默认字体为西文字体——等线（Use the built-in Excel application to create a workbook, so that its default font is of English style - SimHei）
            else:
                wbCreateFlag = False
            workbook_exist: bool = os.path.exists(wbPath)
            while True:
                try:
                    with (pandas.ExcelWriter(path = wbPath, engine = "openpyxl", mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = wbPath, engine = "openpyxl")) as writer:
                        addDefaultStyle(info_df).to_excel(excel_writer = writer, sheet_name = "Profile")
                        logPrint("召唤师生涯导出完成！\nSummoner profile exported!\n")
                        addDefaultStyle(ranked_df).to_excel(excel_writer = writer, sheet_name = "Rank")
                        logPrint("召唤师排位数据导出完成！\nSummoner ranked data exported!\n")
                        addDefaultStyle(ladder_df).to_excel(excel_writer = writer, sheet_name = "Ladders")
                        logPrint("召唤师排位天梯数据导出完成！\nSummoner league ladder data exported!\n")
                        addDefaultStyle(mastery_df).to_excel(excel_writer = writer, sheet_name = "Champion Mastery")
                        logPrint("召唤师英雄成就导出完成！\nSummoner champion mastery exported!\n")
                        addDefaultStyle(recent_LoLPlayer_df).to_excel(excel_writer = writer, sheet_name = "Recently Played Summoners (LoL)")
                        worksheet: Worksheet = writer.sheets["Recently Played Summoners (LoL)"]
                        worksheet.conditional_formatting.rules = [] #读取时清空原规则（Clear original rules when reading）
                        if len(recent_LoLPlayer_df) > 1:
                            max_numPlayersPerTeam_lol = 5 if len(recent_LoLPlayer_df) <= 1 else max(map(lambda x: 5 if x == 0 or not x in gameQueues else 2 if gameQueues[x]["gameMode"] == "CHERRY" else gameQueues[x]["numPlayersPerTeam"], recent_LoLPlayer_df["queueId"][1:])) #自定义对局的队伍规模视为5；斗魂竞技场的队伍规模虽然在API中记录为16，但这里应该考虑的是子阵营（The team size of any custom game is regarded as 5; although the team size of an Arena game is recorded as in LCU API, the subteam has more reference value）
                            addFormat_LoLGame_summary_wb(worksheet, recent_LoLPlayer_df, numColorScale_order = max_numPlayersPerTeam_lol)
                            logPrint("近期一起玩过的英雄联盟玩家数据导出完成！\nRecently played summoner data (LoL) exported!\n")
                        addDefaultStyle(recent_TFTPlayer_df).to_excel(excel_writer = writer, sheet_name = "Recently Played Summoners (TFT)")
                        logPrint("近期一起玩过的云顶之弈玩家数据导出完成！\nRecently played summoner data (TFT) exported!\n")
                        if search_LoL:
                            if scan_lol:
                                if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match History")
                                addDefaultStyle(LoLHistory_df_all).to_excel(excel_writer = writer, sheet_name = "LoL Match History - Scan")
                                if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match History - Manual")
                                worksheet = writer.sheets["LoL Match History - Scan"]
                            else:
                                addDefaultStyle(LoLHistory_df_all).to_excel(excel_writer = writer, sheet_name = "LoL Match History")
                                if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match History - Scan")
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match History - Manual")
                                worksheet = writer.sheets["LoL Match History"]
                            worksheet.conditional_formatting.rules = []
                            if len(LoLHistory_df_all) > 1:
                                addFormat_LoLHistory_wb(worksheet, LoLHistory_df_all)
                            logPrint("召唤师英雄联盟对局记录导出完成！\nSummoner LoL match history exported!\n")
                            if LoLGame_stat_df_export:
                                if scan_lol:
                                    if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                        pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats")
                                    addDefaultStyle(LoLGame_stat_self_df).to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Scan")
                                    if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                        pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Manual")
                                    worksheet = writer.sheets["LoL Match Stats - Scan"]
                                else:
                                    addDefaultStyle(LoLGame_stat_self_df).to_excel(excel_writer = writer, sheet_name = "LoL Match Stats")
                                    if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                        pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Scan")
                                        pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Manual")
                                    worksheet = writer.sheets["LoL Match Stats"]
                                worksheet.conditional_formatting.rules = []
                                if len(LoLGame_stat_self_df) > 1:
                                    max_numPlayersPerTeam_lol = 5 if len(LoLGame_stat_self_df) <= 1 else max(map(lambda x: 5 if x == 0 or not x in gameQueues else 2 if gameQueues[x]["gameMode"] == "CHERRY" else gameQueues[x]["numPlayersPerTeam"], LoLGame_stat_self_df["queueId"][1:]))
                                    addFormat_LoLGame_summary_wb(worksheet, LoLGame_stat_self_df, numColorScale_order = max_numPlayersPerTeam_lol)
                                logPrint("召唤师英雄联盟战绩导出完成！\nSummoner LoL game stats exported!\n")
                            elif (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats")
                                pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Scan")
                                pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Manual")
                                logPrint("已创建英雄联盟战绩的空白数据表。\nCreated an empty sheet for LoL game stats!\n")
                        elif (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match History")
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match History - Scan")
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match History - Manual")
                            logPrint("已创建英雄联盟对局记录的空白数据表！\nCreated an empty sheet for LoL match history!\n")
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats")
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Scan")
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "LoL Match Stats - Manual")
                            logPrint("已创建英雄联盟战绩的空白数据表。\nCreated an empty sheet for LoL game stats!\n")
                        if search_TFT:
                            if scan_tft:
                                if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "TFT Match History")
                                addDefaultStyle(TFTHistory_df_all).to_excel(excel_writer = writer, sheet_name = "TFT Match History - Scan")
                                if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "TFT Match History - Manual")
                            else:
                                addDefaultStyle(TFTHistory_df_all).to_excel(excel_writer = writer, sheet_name = "TFT Match History")
                                if (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "TFT Match History - Scan")
                                    pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "TFT Match History - Manual")
                            logPrint("召唤师云顶之弈对局记录导出完成！\nSummoner TFT match history exported!\n")
                        elif (not workbook_exist or wbCreateFlag) and not args.deny_empty_sheet_creation:
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "TFT Match History")
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "TFT Match History - Scan")
                            pandas.DataFrame().to_excel(excel_writer = writer, sheet_name = "TFT Match History - Manual")
                            logPrint("已创建云顶之弈对局记录的空白工作表！\nCreated an empty sheet for TFT match history!\n")
                        if detail_export:
                            #logPrint(len(info_exist_error), len(timeline_exist_error), len(main_player_included), len(match_reserve_strategy))
                            runTimes: list[float] = [] #记录保存一场对局的所有数据所花费的时间（Records the time spent in saving all data of a match）
                            total_used: float = 0
                            match_reserved: int = 0
                            for i in range(len(matchIDs)):
                                start: float = time.time()
                                if not main_player_included[matchIDs[i]]:
                                    if not match_reserve_strategy[matchIDs[i]]:
                                        logPrint("对局概要和时间轴导出进度（Match summary and timeline export process）：%d/%d (Excluding this summoner and not exported!)" %(i + 1, len(matchIDs)))
                                    else:
                                        logPrint("对局概要和时间轴导出进度（Match summary and timeline export process）：%d/%d (Excluding this summoner but yet exported!)" %(i + 1, len(matchIDs)))
                                else:
                                    if info_exist_error[matchIDs[i]] and not timeline_exist_error[matchIDs[i]]:
                                        logPrint("对局概要和时间轴导出进度（Match summary and timeline export process）：%d/%d (Match summary capture failure!)" %(i + 1, len(matchIDs)))
                                    elif not info_exist_error[matchIDs[i]] and timeline_exist_error[matchIDs[i]]:
                                        if isTFT[matchIDs[i]]:
                                            logPrint("对局概要和时间轴导出进度（Match summary and timeline export process）：%d/%d" %(i + 1, len(matchIDs)))
                                        else:
                                            logPrint("对局概要和时间轴导出进度（Match summary and timeline export process）：%d/%d (Match timeline capture failure!)" %(i + 1, len(matchIDs)))
                                    elif info_exist_error[matchIDs[i]] and timeline_exist_error[matchIDs[i]]:
                                        logPrint("对局概要和时间轴导出进度（Match summary and timeline export process）：%d/%d (Match summary & timeline capture Failure!)" %(i + 1, len(matchIDs)))
                                    else:
                                        logPrint("对局概要和时间轴导出进度（Match summary and timeline export process）：%d/%d" %(i + 1, len(matchIDs)))
                                logPrint("对局序号（MatchID）： %d" %matchIDs[i])
                                if match_reserve_strategy[matchIDs[i]]:
                                    match_reserved += 1
                                    if not info_exist_error[matchIDs[i]]:
                                        if args.export_leaderboard:
                                            addDefaultStyle(game_leaderboard_dfs[matchIDs[i]]).to_excel(excel_writer = writer, sheet_name = "Match %d - Leaderboard" %(matchIDs[i]))
                                            logPrint("对局段位排行榜导出完成。\nMatch leaderboard exported.")
                                        game_summary_df: pandas.DataFrame = game_summary_dfs[matchIDs[i]]
                                        game_summary_df_dense: pandas.DataFrame = eliminate_empty_fields(game_summary_df)
                                        addDefaultStyle(game_summary_df_dense.transpose()).to_excel(excel_writer = writer, sheet_name = "Match %d - Summary" %(matchIDs[i]))
                                        if isLoL.get(matchIDs[i], False) and args.info_color:
                                            worksheet = writer.sheets["Match %d - Summary" %(matchIDs[i])]
                                            worksheet.conditional_formatting.rules = []
                                            participantId_teamId_map: dict[str, list[int]] = {}
                                            participantId_subteamId_map: dict[str, list[int]] = {}
                                            for j in range(1, len(game_summary_df)):
                                                participantId: int = game_summary_df["participantId"][j]
                                                team = game_summary_df["team_color"][j]
                                                playerSubteam = game_summary_df["playerSubteamColor"][j]
                                                if not team in participantId_teamId_map:
                                                    participantId_teamId_map[team] = []
                                                participantId_teamId_map[team].append(participantId)
                                                if not playerSubteam in participantId_subteamId_map:
                                                    participantId_subteamId_map[playerSubteam] = []
                                                participantId_subteamId_map[playerSubteam].append(participantId)
                                            max_numPlayersPerTeam_lol = max(map(len, participantId_teamId_map.values())) if all(map(lambda x: game_summary_df["playerSubteamColor"][x] == "", list(range(1, len(game_summary_df))))) else max(map(len, participantId_subteamId_map.values()))
                                            addFormat_LoLGame_summary_wb_transpose(worksheet, game_summary_df_dense.transpose(), numColorScale_order = max_numPlayersPerTeam_lol)
                                        logPrint("对局概要导出完成。\nMatch summary exported.")
                                    if not timeline_exist_error[matchIDs[i]]:
                                        game_timeline_df: pandas.DataFrame = eliminate_empty_fields(game_timeline_dfs[matchIDs[i]])
                                        addDefaultStyle(game_timeline_df).to_excel(excel_writer = writer, sheet_name = "Match %d - Timeline" %(matchIDs[i]))
                                        logPrint("对局时间轴导出完成。\nMatch timeline exported.")
                                        game_event_df: pandas.DataFrame = eliminate_empty_fields(game_event_dfs[matchIDs[i]])
                                        addDefaultStyle(game_event_df).to_excel(excel_writer = writer, sheet_name = "Match %d - Events" %(matchIDs[i]))
                                        logPrint("对局事件导出完成。\nMatch events exported.")
                                end: float = time.time()
                                unit: float = end - start
                                total_used += unit
                                if match_reserve_strategy[matchIDs[i]]:
                                    runTimes.append((sheetNumber[matchIDs[i]], unit))
                                    total_remaining = 0 if sum([j[0] for j in runTimes[:match_reserved + 1]]) == 0 else sum([j[1] for j in runTimes[:match_reserved + 1]]) / sum([j[0] for j in runTimes[:match_reserved + 1]]) * sum([sheetNumber[matchIDs[j]] for j in range(i + 1, len(matchIDs))]) #需要考虑除数为0的情况（The case where the divisor is 0 needs considering）
                                    logPrint("保存本场对局所花费的时间（Time spent in saving this match）： %s" %(format_runtime(unit)))
                                    logPrint("已花费的总时间（Total time used）                          ： %s" %(format_runtime(total_used)))
                                    logPrint("剩余时间（Time remaining）                                 ： %s" %(format_runtime(total_remaining)))
                                    logPrint("预计总时间（Expected total time）                          ： %s" %(format_runtime(total_used + total_remaining)), end = "\n\n")
                            if len(matchIDs) != 0:
                                logPrint("对局概要和时间轴导出完成！\nMatch summary and timeline exported!")
                except PermissionError:
                    logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                    logInput()
                else:
                    if detail_export and len(smurfs) == 0: #在小号模式下不更新对局序号列表，以防扫描模式扫描时出现不必要的询问（MatchId lists aren't updated under Smurf Mode, in case unnecessary questions would pop up under Scan Mode）
                        with open(json02path, "w", encoding = "utf-8") as jsonfile02:
                            json.dump(sorted(LoLMatches_exported), jsonfile02, ensure_ascii = False)
                        logPrint("已更新英雄联盟对局序号列表！\nUpdated LoL matchId list!")
                        with open(json03path, "w", encoding = "utf-8") as jsonfile03:
                            json.dump(sorted(TFTMatches_exported), jsonfile03, ensure_ascii = False)
                        logPrint("已更新云顶之弈对局序号列表！\nUpdated TFT matchId list!")
                    del worksheet #清除工作表对象，以断开与内存中的工作簿数据的引用（Delete the worksheet object to disconnect it from the workbook data in memory）
                    del writer #清除工作簿对象，以断开与内存中的工作簿数据的引用（Delete the workbook object to disconnect it from the workbook data in memory）
                    break
            if workbook_exist:
                logPrint("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致对局序号顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of matchId order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
                sort_str: str = logInput()
                sort: bool = bool(sort_str)
                if sort: #所有工作表分为基础信息类和对局信息类，排列顺序为前者在前、后者在后。基础信息工作表类按顺序依次为人物简介、排位信息、英雄成就和对局记录。对局信息类工作表包括对局排行榜、对局概要和对局时间轴，按照对局序号排序（All sheets are divided into the basic data class and match information class, the former arranged in front of the latter. The basic data class includes profile, rank, champion mastery and match history in turn. The match information class includes match leaderboard, match summary and match timeline ordered by matchIDs）
                    logPrint("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
                    wb: Workbook = Workbook()
                    while True:
                        try:
                            wb = load_workbook(wbPath)
                        except FileNotFoundError:
                            logPrint('召唤师生涯工作簿读取失败！请确保“%s”文件夹内含有名为“%s”的工作簿。输入“0”以重试。\nERROR reading the summoner profile workbook! Please make sure the workbook "%s" is in the folder "%s". Submit "0" to try again.' %(folder, wbName, wbName, folder))
                            profile_reload_str: str = logInput()
                            if profile_reload_str == "0":
                                break
                        else:
                            sheetnames: list[str] = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                            #下面锁定基础信息类的工作表顺序（The following code lock the order of sheets in basic data class）
                            logPrint("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                            basic_info_list: list[str] = ["Profile", "Rank", "Ladders", "Champion Mastery", "Recently Played Summoners (LoL)", "Recently Played Summoners (TFT)", "LoL Match History", "LoL Match History - Scan", "LoL Match History - Manual", "LoL Match Stats", "LoL Match Stats - Scan", "LoL Match Stats - Manual", "TFT Match History", "TFT Match History - Manual", "TFT Match History - Scan"]
                            match_dict: dict[int, dict[str, str]] = {}
                            for sheet_iter in sheetnames:
                                if sheet_iter.startswith("Match "):
                                    matchId: int = int(sheet_iter.split()[1]) #目前暂不需要考虑对局序号因工作表名长度限制而被截断的问题（Currently the issue that matchId may be cut off due to the sheet name length limit doesn't need to be considered）
                                    key: str = sheet_iter.split()[3][0] #以工作表名的内容部分的首字母为排序依据（Sort the sheetnames by the initial letter of the content part of the sheet name）
                                    if not matchId in match_dict:
                                        match_dict[matchId] = {}
                                    match_dict[matchId][key] = sheet_iter
                            sheetnames_sorted: list[str] = [] #所有工作表的期望顺序存储在sheetnames_sorted变量中（The ordered result of all sheets is stored in the variable `sheetnames_sorted`）
                            for sheet_iter in basic_info_list:
                                if sheet_iter in sheetnames:
                                    sheetnames_sorted.append(sheet_iter)
                            for matchId in sorted(match_dict.keys()):
                                if "L" in match_dict[matchId]: #对局排行榜（Match leaderboard）
                                    sheetnames_sorted.append(match_dict[matchId]["L"])
                                if "I" in match_dict[matchId]: #对局信息（已弃用）【Match information (deprecated)】
                                    sheetnames_sorted.append(match_dict[matchId]["I"])
                                if "S" in match_dict[matchId]: #对局概要（Match summary）
                                    sheetnames_sorted.append(match_dict[matchId]["S"])
                                if "T" in match_dict[matchId]: #对局时间轴（Match timeline）
                                    sheetnames_sorted.append(match_dict[matchId]["T"])
                                if "E" in match_dict[matchId]: #对局事件（Match event）
                                    sheetnames_sorted.append(match_dict[matchId]["E"])
                            #下面排列所有工作表（The following code arrange all sheets）
                            logPrint("正在排序……\nOrdering ...")
                            sort_worksheet(wb, sheetnames_sorted)
                            logPrint('正在保存中……\nSaving the ordered workbook ...')
                            wb.save(os.path.join(folder, wbName_sorted))
                            logPrint('排序完成！排好序的工作簿已保存为“%s”。\nOrdering finished! The ordered workbook is saved as "%s".\n' %(wbName_sorted, wbName_sorted))
                            wb.close()
                            del wb #显式清除工作簿对象以释放内存（Explicitly delete the workbook object to release memory）
                            break
            logPrint("正在清理工作簿占用内存……\nCleaning up the memory occupied by the workbook ...")
            gc.collect() #显式回收垃圾。只有在当前运行环境中没有任何工作簿数据的引用时，工作簿所占用的内存才会被释放（Explicitly collect garbage. Only when there're no references to the workbook data in the current runtime environment will the memory occupied by the workbook be released）

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection: Connection) -> None:
    global sgpSession, log, logInput, logPrint
    log_folder: str = "日志（Logs）/Customized Program 05 - Search Summoner Profile"
    os.makedirs(log_folder, exist_ok = True)
    currentTime: str = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    log = LogManager(os.path.join(log_folder, currentTime + ".log"), mode = "a+", encoding = "utf-8")
    logInput = log.logInput
    logPrint = log.logPrint
    sgpSession.setLog(log)
    await sgpSession.init(connection)
    await print_summoner_info(connection)
    await save_platform_info(connection)
    await search_profile(connection)
    log.write("\n[Program terminated and returned status 0.]\n")
    log.close()

@connector.close
async def disconnect(connection: Connection) -> None:
    print("已从英雄联盟客户端断开连接。\nDisconnected from the League Client.")

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

connector.start()
