from lcu_driver import Connector
from lcu_driver.connection import Connection
import os, pandas, re, time
from openpyxl import load_workbook
from typing import Any
from src.utils.summoner import get_summoner_data, get_info_name
from src.utils.format import getISOTime, optimize_bool_display, format_df
from src.core.config.servers import set_summonerInfo_folder, save_platform_info
from src.core.config.headers import mission_header, objective_group_header, objective_category_header
from src.core.config.localization import celebrationTypes, clientNotifyLevels, displayTypes, missionTypes, metadataMissionTypes, objectiveStatus_dict, objectiveTypes, rewardGroupStrategies, rewardTypes, missionStatus_dict, gameTypes_mission, objectivesTypes, categoryTypes, lolEventHubTypes, objectiveCategoryFilter_dict, tftPassTypes

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2026/03/02
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

connector: Connector = Connector()

#-----------------------------------------------------------------------------
# 获取英雄联盟中的所有游戏类型信息（Get all game types' information in League of Legends）
#-----------------------------------------------------------------------------
async def get_mission_info(connection: Connection) -> None:
    platformId: str = await (await connection.request("GET", "/lol-platform-config/v1/namespaces/LoginDataPacket/platformId")).json()
    riot_client_info: list[str] = await (await connection.request("GET", "/riotclient/command-line-args")).json()
    client_info: dict[str, str] = {}
    for i in range(len(riot_client_info)):
        try:
            client_info[riot_client_info[i].split("=")[0]] = riot_client_info[i].split("=")[1]
        except IndexError:
            pass
    region: str = client_info["--region"]
    #设置输出路径（Set the output directory）
    current_info: dict[str, Any] = await (await connection.request("GET", "/lol-summoner/v1/current-summoner")).json()
    displayName: str = get_info_name(current_info)
    folder: str = set_summonerInfo_folder(region, platformId, current_info)
    #获取数据资源（Get data resources）
    missions: list[dict[str, Any]] = await (await connection.request("GET", "/lol-missions/v1/missions")).json()
    lolObjectives: list[dict[str, Any]] = await (await connection.request("GET", "/lol-objectives/v1/objectives/lol")).json() #该列表中的两个字典的“objectives”键的值相同（Values of the "objectives" key of both dictionaries in this list are the same）
    tftObjectives: list[dict[str, Any]] = await (await connection.request("GET", "/lol-objectives/v1/objectives/tft")).json()
    objectives: list[dict[str, Any]] = lolObjectives + tftObjectives
    objectiveGroup_objectiveCategory_map: dict[str, str] = {}
    objectiveCategoryId_order: dict[str, int] = {} #用于后续目标分组数据框排序——第一关键字（Used as the first keyword for the subsequent objective group dataframe sorting）
    mission_objectiveGroup_map: dict[tuple[str, int], str] = {}
    objectiveGroupId_order: dict[str, int] = {} #用于后续任务数据框排序——第一关键字（Used as the first keyword for the subsequent mission dataframe sorting）
    weight1: int = 0
    weight2: int = 0
    for objective in objectives:
        for objectiveCategory in objective["objectivesCategories"]:
            objectiveCategoryId_order[objectiveCategory["id"]] = weight1
            weight1 += 1
            for objectiveGroup in objectiveCategory["objectives"]:
                objectiveGroup_objectiveCategory_map[objectiveGroup["id"]] = {"id": objectiveCategory["id"], "categoryName": objectiveCategory["categoryName"]}
                objectiveGroupId_order[objectiveGroup["id"]] = weight2
                weight2 += 1
                for mission in objectiveGroup["missions"]: #任务所属目标信息在`/lol-missions/v1/missions`接口中无法体现，因此需要提前准备这方面数据（Belonging objective of a mission isn't reflected by the endpoint `/lol-missions/v1/missions`, so this information needs preparing in advance）
                    mission_objectiveGroup_map[(mission["id"], mission["sequence"])] = {"id": objectiveGroup["id"], "localizedTag": objectiveGroup["localizedTag"], "localizedTitle": objectiveGroup["localizedTitle"]}
    #整理数据（Organize data）
    ##任务（Mission）
    mission_header_keys: list[str] = list(mission_header.keys())
    mission_data: dict[str, list[Any]] = {key: [] for key in mission_header_keys}
    for mission in missions:
        for objective in mission["objectives"]:
            for i in range(len(mission_header_keys)):
                key: str = mission_header_keys[i]
                if i <= 34:
                    if i == 1: #庆祝类型（`celebrationType`）
                        to_append: Any = celebrationTypes[mission["celebrationType"]]
                    elif i == 2: #客户端通知类型（`clientNotifyLevel`）
                        to_append = clientNotifyLevels[mission["clientNotifyLevel"]]
                    elif i == 7: #显示类型（`displayType`）
                        to_append = displayTypes[mission["displayType"]]
                    elif i == 20: #任务类型（`missionType`）
                        to_append = missionTypes[mission["missionType"]]
                    elif i == 26: #状态（`status`）
                        to_append = missionStatus_dict[mission["status"]]
                    elif i >= 29 and i <= 33: #时间类键（Time-type keys）
                        subkey_dict: dict[int, str] = {29: "completedDate", 30: "earnedDate", 31: "endTime", 32: "lastUpdatedTimestamp", 33: "startTime"}
                        try:
                            timeStr: str = "" if mission[subkey_dict[i]] == -1 else getISOTime(mission[subkey_dict[i]] / 1000)
                        except OSError:
                            timeStr = mission[subkey_dict[i]]
                        to_append = timeStr
                    elif i == 34: #任务奖励描述（`rewardDescriptions`）
                        rewardDescriptions: list[str] = []
                        for reward in mission["rewards"]:
                            if not reward["description"] in rewardDescriptions: #同一项奖励描述可能被拆成多项奖励类型不同的奖励（One description can be distributed into multiple rewards with different rewardTypes）
                                rewardDescriptions.append(reward["description"])
                        to_append = rewardDescriptions
                    else:
                        to_append = mission[key]
                elif i <= 57: #该代码框架适用于纯嵌套字典类型的值（This code frame applies to pure nested dictionary-type values）
                    tmpObj_ptr: Any = mission
                    for subkey_iter in key.split():
                        tmpObj_ptr = tmpObj_ptr[subkey_iter]
                    if i == 40: #元数据：任务类型（`metadata missionType`）
                        to_append = metadataMissionTypes[tmpObj_ptr]
                    elif i == 55: #任务奖励分组（`rewardStrategy groupStrategy`）
                        to_append = rewardGroupStrategies[tmpObj_ptr]
                    else:
                        to_append = tmpObj_ptr
                elif i <= 67:
                    tmpObj_ptr: Any = objective
                    for j in range(1, len(key.split())):
                        tmpObj_ptr = tmpObj_ptr[key.split()[j]]
                    if i == 63: #目标状态（`objective status`）
                        to_append = objectiveStatus_dict[tmpObj_ptr]
                    elif i == 64: #目标类型（`objective type`）
                        to_append = objectiveTypes[tmpObj_ptr]
                    else:
                        to_append = tmpObj_ptr
                else:
                    subkey = key.split()[1]
                    if (mission["id"], mission["sequence"]) in mission_objectiveGroup_map:
                        to_append = mission_objectiveGroup_map[(mission["id"], mission["sequence"])][subkey]
                    else:
                        to_append = ""
                mission_data[key].append(to_append)
    mission_statistics_output_order: list[int] = [70, 69, 68, 13, 27, 14, 6, 19, 11, 24, 23, 21, 35, 36, 58, 62, 60, 64, 66, 65, 67, 63, 59, 61, 26, 34, 22, 55, 56, 57, 33, 31, 32, 30, 4, 29, 5, 20, 7, 1, 2, 15, 28, 18, 10, 0, 12, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54]
    mission_data_organized: dict[str, list[Any]] = {mission_header_keys[i]: mission_data[mission_header_keys[i]] for i in mission_statistics_output_order}
    mission_df: pandas.DataFrame = pandas.DataFrame(data = mission_data_organized)
    optimize_bool_display(mission_df)
    mission_df = pandas.concat([pandas.DataFrame([mission_header])[mission_df.columns], mission_df], ignore_index = True)
    mission_df_data: pandas.DataFrame = mission_df.loc[1:, :]
    mission_df_data_sorted: pandas.DataFrame = mission_df_data.sort_values(by = "objectiveGroup id", key = lambda x: x.map(objectiveGroupId_order), ascending = True)
    mission_df_sorted: pandas.DataFrame = pandas.concat([mission_df.iloc[:1, :], mission_df_data_sorted])
    ##目标和目标分组（Objective and objective group）
    objective_group_header_keys: list[str] = list(objective_group_header.keys())
    objective_group_data: dict[str, list[Any]] = {key: [] for key in objective_group_header_keys}
    objective_category_header_keys: list[str] = list(objective_category_header.keys())
    objective_category_data: dict[str, list[Any]] = {key: [] for key in objective_category_header_keys}
    for objective in objectives:
        for objectiveCategory in objective["objectivesCategories"]:
            for i in range(len(objective_category_header_keys)):
                key: str = objective_category_header_keys[i]
                if i == 0: #游戏类型（`gameType`）
                    to_append: Any = gameTypes_mission[objective["gameType"]]
                else:
                    if i == 3: #类型（`categoryType`）
                        to_append = categoryTypes[objectiveCategory["categoryType"]]
                    elif i == 6: #英雄联盟活动专题类型（`lolEventHubType`）
                        to_append = lolEventHubTypes[objectiveCategory["lolEventHubType"]]
                    elif i == 7: #大类（`objectiveCategoryFilter`）
                        to_append = objectiveCategoryFilter_dict[objectiveCategory["objectiveCategoryFilter"]]
                    elif i == 11: #云顶之弈通行证类型（`tftPassType`）
                        to_append = tftPassTypes[objectiveCategory["tftPassType"]]
                    elif i >= 12: #时间类键（Time-type keys）
                        subkey_dict: dict[int, str] = {12: "endDate", 13: "progressEndDate", 14: "startDate"}
                        timeStr: str = getISOTime(objectiveCategory[subkey_dict[i]] / 1000)
                        to_append = timeStr
                    else:
                        to_append = objectiveCategory[key]
                objective_category_data[key].append(to_append)
            for objectiveGroup in objectiveCategory["objectives"]:
                for i in range(len(objective_group_header_keys)):
                    key: str = objective_group_header_keys[i]
                    if i <= 14:
                        if i == 8: #类型（`objectivesType`）
                            to_append: Any = objectivesTypes[objectiveGroup["objectivesType"]]
                        elif i >= 13: #时间类键（Time-type keys）
                            subkey_dict: dict[int, str] = {13: "endDate", 14: "startDate"}
                            timeStr = getISOTime(objectiveGroup[subkey_dict[i]] / 1000)
                            to_append = timeStr
                        else:
                            to_append = objectiveGroup[key]
                    else:
                        subkey: str = key.split()[1]
                        if objectiveGroup["id"] in objectiveGroup_objectiveCategory_map:
                            to_append = objectiveGroup_objectiveCategory_map[objectiveGroup["id"]][subkey]
                        else:
                            to_append = ""
                    objective_group_data[key].append(to_append)
    objective_group_statistics_output_order: list[int] = [16, 15, 6, 5, 2, 9, 8, 3, 4, 14, 13, 10, 7, 0]
    objective_group_data_organized: dict[str, list[Any]] = {objective_group_header_keys[i]: objective_group_data[objective_group_header_keys[i]] for i in objective_group_statistics_output_order}
    objective_group_df: pandas.DataFrame = pandas.DataFrame(data = objective_group_data_organized)
    optimize_bool_display(objective_group_df)
    objective_group_df = pandas.concat([pandas.DataFrame([objective_group_header])[objective_group_df.columns], objective_group_df], ignore_index = True)
    objective_group_df_data: pandas.DataFrame = objective_group_df.loc[1:, :]
    objective_group_df_data_sorted: pandas.DataFrame = objective_group_df_data.sort_values(by = "objectiveCategory id", key = lambda x: x.map(objectiveCategoryId_order), ascending = True)
    objective_group_df_sorted: pandas.DataFrame = pandas.concat([objective_group_df.iloc[:1, :], objective_group_df_data_sorted])
    objective_category_statistics_output_order: list[int] = [0, 1, 5, 7, 3, 2, 6, 11, 14, 13, 12, 8]
    objective_category_data_organized: dict[str, list[Any]] = {objective_category_header_keys[i]: objective_category_data[objective_category_header_keys[i]] for i in objective_category_statistics_output_order}
    objective_category_df: pandas.DataFrame = pandas.DataFrame(data = objective_category_data_organized)
    optimize_bool_display(objective_category_df)
    objective_category_df = pandas.concat([pandas.DataFrame([objective_category_header])[objective_category_df.columns], objective_category_df], ignore_index = True)
    #保存文件（Save the files）
    excel_name: str = f"Player Mission - {displayName}.xlsx" 
    excel_name_sorted: str = f"Player Mission - {displayName} (sorted).xlsx"
    wbPath: str = os.path.join(folder, excel_name)
    workbook_exist: bool = os.path.exists(wbPath)
    while True:
        try:
            with (pandas.ExcelWriter(path = wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = wbPath)) as writer:
                currentTime: str = time.strftime("%Y-%m-%d", time.localtime(time.time()))
                mission_df_sorted.to_excel(excel_writer = writer, sheet_name = f"Missions - {currentTime}_{platformId}")
                objective_group_df_sorted.to_excel(excel_writer = writer, sheet_name = f"Objectives - {currentTime}_{platformId}")
                objective_category_df.to_excel(excel_writer = writer, sheet_name = f"Categories - {currentTime}_{platformId}")
        except PermissionError:
            print("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
            input()
        else:
            print('玩家目标和任务信息已保存为“%s”！\nPlayer objective and mission information is saved as "%s"!' %(wbPath, wbPath))
            break
    #工作表排序（Worksheet ordering）
    if workbook_exist:
        print("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致工作表顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of worksheet order. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
        sort: bool = bool(input())
        if sort:
            mission_loaded: bool = True
            print("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
            while True:
                try:
                    wb = load_workbook(wbPath)
                except FileNotFoundError:
                    print('任务信息工作簿读取失败！请确保“%s”文件夹内含有名为“%s”的工作簿。如果需要退出程序，请输入“0”。\nERROR reading the Missions workbook! Please make sure the workbook "%s" is in the folder "%s". If you want to exit the program, please submit "0".' %(folder, excel_name, excel_name, folder))
                    mission_reload: str = input()
                    if mission_reload == "0":
                        mission_loaded = False
                        break
                else:
                    break
            if mission_loaded:
                sheetnames: list[str] = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                print("请选择排序方式：\nPlease select an ordering pattern:\n1\t时间优先（默认）【Time in priority (by default)】\n2\t类别优先（Type in priority）")
                op: str = input()
                print("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                date_re = re.compile(r"\d{4}-\d{2}-\d{2}") #设置正则表达式识别
                if op == "" or op[0] != "2": #按照时间优先的原则对工作表进行排序，时间相同则任务工作表在前，目标工作表在后（Sort the sheets by time in priority. If the times are the same, then the mission sheet is arranged in front of the objective sheet）
                    sheetname_date_list: list[str] = list(map(lambda x: date_re.search(x).group(), sheetnames)) #从工作表名称提取日期信息形成列表（Extract the dates from the sheetnames to form a list）
                    sheetname_type_list: list[str] = list(map(lambda x: x.split()[0], sheetnames)) #从工作表名称提取数据类型信息形成列表（Extract the data types from the sheetnames to form a list）
                    sheetname_platform_list: list[str] = list(map(lambda x: x.split("_")[1], sheetnames)) #从工作表名称提取大区信息形成列表（Extract the platformId from the sheetnames to form a list）
                    sheetname_tmpDf: pandas.DataFrame = pandas.DataFrame(data = [sheetnames, sheetname_date_list, sheetname_type_list, sheetname_platform_list]).transpose() #创建一个四列数据框，各列分别是完整工作表名、日期信息、数据类型信息和大区信息（Create a 4-column dataframe whose columns are the complete sheetname, date, data type and platformId）
                    sheetnames_sorted: list[str] = sheetname_tmpDf.sort_values(by = [1, 2, 3], ascending = [True, True, True]).iloc[:, 0].tolist() #将工作表名按照第一关键字——日期信息正序排列，第二关键字——数据类型信息正序排列（先任务后目标），第三关键字——大区信息正序排列（Order the sheetnames according to the ascending order of the first keyword - date, the ascending order of the second keyword - data type and the ascending order of the third keyword - platformId）
                else:
                    sheets_Missions: list[str] = [sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Missions")] #提取任务类型的工作表名称（Extract the names of the sheets containing mission data）
                    sheets_Objectives: list[str] = [sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Objectives")] #提取目标类型的工作表名称（Extract the names of the sheets containing objective data）
                    sheets_Missions = sorted(sheets_Missions, key = lambda x: date_re.search(x).group()) #按照日期正序排列任务类型的工作表名称（Order the mission sheetnames according to the ascending order of dates）
                    sheets_Objectives = sorted(sheets_Objectives, key = lambda x: date_re.search(x).group()) #按照日期正序排列目标类型的工作表名称（Order the objective sheetnames according to the ascending order of dates）
                    sheetnames_sorted = sheets_Missions + sheets_Objectives #合并列表得到先按类别排列、再按日期排列的工作表名称（Combine the lists to get the sheetname list ordered firstly by data type and secondly by date）
                #下面排列所有工作表（The following code arrange all sheets）
                print("正在排序……\nOrdering ...")
                for i in range(len(sheetnames_sorted)): #排序的思路是每次将一个工作表根据其在原工作表列表中的索引和在顺序工作表列表中的索引的差值进行移动（The main idea of sheets' sorting is to move each sheet according to the difference of the indices between in the original sheet list and in the ordered sheet list）
                    sheetnames = wb.sheetnames #因为一次移动可能导致很多其它工作表的位置发生变化，所以必须每次都重新获取工作表列表（Because a moving event may result in location change of many other sheets, the sheet list must be obtained each time）
                    sheetname_iter: str = sheetnames_sorted[i] #这里以顺序工作表为迭代器进行遍历，因为顺序工作表是固定不变的（Here the ordered sheet list acts as the iterator to be traversed, for the ordered sheet list is fixed）
                    if sheetnames[i] != sheetname_iter:
                        preIndex: int = sheetnames.index(sheetname_iter)
                        wb.move_sheet(sheetname_iter, i - preIndex) #注意移动距离数应当是排序后的索引减去排序前的索引（Note that the moving offset should be the index in the ordered list subtracted by that in the original list）
                    #print("排序进度（Ordering process）：%d/%d\t工作表名称（Sheet name）： %s" %(i + 1, len(sheetnames_sorted), sheetname_iter))
                print('正在保存中……\nSaving the ordered workbook ...')
                wb.save(os.path.join(folder, excel_name_sorted))
                print('排序完成！排好序的工作簿已保存为“%s”。\nOrdering finished! The ordered workbook is saved as "%s".\n' %(excel_name_sorted, excel_name_sorted))

async def check_repeating_missions(connection: Connection) -> None:
    #查看可重复任务的刷新状态（Check repeating missions' cooldown status）
    while True:
        missions: list[dict[str, Any]] = await (await connection.request("GET", "/lol-missions/v1/missions")).json()
        repeating_missions: list[dict[str, Any]] = [mission for mission in missions if mission["cooldownTimeMillis"] != -1]
        completed_repeating_missions: list[dict[str, Any]] = [mission for mission in repeating_missions if mission["status"] == "COMPLETED"]
        pending_repeating_missions: list[dict[str, Any]] = [mission for mission in repeating_missions if mission["status"] == "PENDING"]
        if len(repeating_missions) != 0:
            print("在该服务器上检测到可重复任务。\nDetected repeating missions on this server.")
            if len(completed_repeating_missions) != 0:
                print("以下任务已完成：\nThe following missions are completed:")
                for mission in completed_repeating_missions:
                    rewardDescriptions: list[str] = []
                    for reward in mission["rewards"]:
                        if not reward["description"] in rewardDescriptions:
                            rewardDescriptions.append(reward["description"]) #奖励（Reward）
                    completedTime: str = getISOTime(mission["completedDate"] / 1000) #完成时间（Completed time）
                    cooldown: float = mission["cooldownTimeMillis"] / 1000
                    cooldown_hour: int = int(cooldown) // 3600
                    cooldown_minute: int = int(cooldown) // 3600 % 60
                    cooldown_second: int = int(cooldown) % 60
                    cooldown_str: str = "%d:%02d:%02d" %(cooldown_hour, cooldown_minute, cooldown_second) #刷新间隔（Cooldown）
                    refreshDate: float = mission["completedDate"] + mission["cooldownTimeMillis"]
                    refreshTime: str = getISOTime(refreshDate / 1000) #刷新时间（Refresh time）
                    cooldown_remaining: float = refreshDate // 1000 - time.time()
                    cooldown_remaining_hour: int = int(cooldown_remaining) // 3600
                    cooldown_remaining_minute: int = int(cooldown_remaining) % 3600 // 60
                    cooldown_remaining_second: int = int(cooldown_remaining) % 60
                    cooldown_remaining_str: str = "%d:%02d:%02d" %(cooldown_remaining_hour, cooldown_remaining_minute, cooldown_remaining_second) #剩余时间（Cooldown remaining）
                    missionData_simple: dict[str, list[Any]] = {"项目": ["标题", "识别码", "序列号", "描述", "奖励", "完成时间", "刷新间隔", "刷新时间", "剩余时间"], "Items": ["title", "id", "sequence", "description", "rewards", "completedTime", "cooldown", "refreshTime", "cooldownRemaining"], "值": [mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions, completedTime, cooldown_str, refreshTime, cooldown_remaining_str]}
                    #print("标题（Title）： %s\n序号（Id）： %s\n序列号（Sequence）： %d\n描述（Description）： %s\n奖励（Reward）： %s\n完成时间（Completed time）： %s\n刷新间隔（Cooldown）： %s\n刷新时间（Refresh time）： %s\n剩余时间：（Cooldown remaining）： %s\n" %(mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions, completedTime, cooldown_str, refreshTime, cooldown_remaining_str))
                    missionDf_simple: pandas.DataFrame = pandas.DataFrame(data = missionData_simple)
                    print(format_df(missionDf_simple, print_index = True, header_align = "^", align = "^^>")[0], end = "\n\n")
            if len(pending_repeating_missions) != 0:
                print("\n以下任务等待完成：\nThe following missions are pending:")
                for mission in pending_repeating_missions:
                    rewardDescriptions = []
                    for reward in mission["rewards"]:
                        if not reward["description"] in rewardDescriptions:
                            rewardDescriptions.append(reward["description"]) #奖励（Reward）
                    missionData_simple = {"项目": ["标题", "识别码", "序列号", "描述", "奖励"], "Items": ["title", "id", "sequence", "description", "rewards"], "值": [mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions]}
                    #print("标题（Title）： %s\n序号（Id）： %s\n序列号（Sequence）： %d\n描述（Description）： %s\n奖励（Reward）： %s\n" %(mission["title"], mission["id"], mission["sequence"], mission["description"], rewardDescriptions))
                    missionDf_simple = pandas.DataFrame(data = missionData_simple)
                    print(format_df(missionDf_simple, print_index = True, header_align = "^", align = "^^>")[0], end = "\n\n")
            print("\n是否更新任务状态？（输入任意键以更新，否则退出程序。）\nUpdate the missions? (Submit any non-empty string to update, or null to exit the program.)")
            if not bool(input()):
                break
        else:
            break

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection: Connection) -> None:
    await get_summoner_data(connection)
    await save_platform_info(connection)
    print("是否导出所有任务信息？（输入任意键不导出，否则导出。）\nDo you want to export all missions' information? (Submit any non-empty string to refuse exporting, or null to export.)")
    if not bool(input()):
        await get_mission_info(connection)
    await check_repeating_missions(connection)

@connector.close
async def disconnect(connection: Connection) -> None:
    print("已从英雄联盟客户端断开连接。\nDisconnected from the League Client.")

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

connector.start()
