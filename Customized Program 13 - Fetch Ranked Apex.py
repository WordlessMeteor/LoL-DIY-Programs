from lcu_driver import Connector
from lcu_driver.connection import Connection
from openpyxl import load_workbook, Workbook
import json, os, pandas, time
from typing import Any
from src.utils.summoner import print_summoner_info, get_info, get_infos, get_info_name
from src.utils.logger import LogManager
from src.utils.format import optimize_bool_display, addDefaultStyle, format_runtime
from src.utils.excel_workbook import create_workbook_win32, sort_worksheet
from src.core.config.servers import platform_TENCENT, platform_RIOT, platform_GARENA
from src.core.config.headers import challenger_ladder_metadata_header, challenger_ladder_header, topRated_ladder_header

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： XHXIAIEIN
# 更新（Last update）：     2026/08/16
#=============================================================================

#-----------------------------------------------------------------------------
# 工具库（Tool library）
#-----------------------------------------------------------------------------
#  - lcu-driver 
#    https://github.com/sousa-andre/lcu-driver
#-----------------------------------------------------------------------------

log: LogManager = LogManager()

connector: Connector = Connector()

#-----------------------------------------------------------------------------
#  获取最强王者段位信息（Get challenger tier league information）
#-----------------------------------------------------------------------------
async def get_challenger_tier(connection: Connection) -> None:
    current_party: dict[str, Any] = await (await connection.request("GET", "/lol-lobby/v1/parties/player")).json()
    platformId: str = current_party["platformId"]
    #下面设置输出文件的位置（The following code determines the output files' location）
    riot_client_info: list[str] = await (await connection.request("GET", "/riotclient/command-line-args")).json()
    client_info: dict[str, str] = {}
    for i in range(len(riot_client_info)):
        try:
            client_info[riot_client_info[i].split("=")[0]] = riot_client_info[i].split("=")[1]
        except IndexError:
            pass
    region: str = client_info["--region"]
    currentLoLSeason: dict[str, Any] = await (await connection.request("GET", "/lol-seasons/v1/season/product/LOL")).json() #API中记录的赛季与平常所说的赛季有所不同（The season recorded in API is different from the often mentioned season）
    currentLoLSeasonId: int = currentLoLSeason["seasonId"]
    currentLoLSeasonSplitId: int = currentLoLSeason["metadata"]["currentSplit"]
    # currentSplit: int = int(platform_config["LeagueConfig"]["CurrentSplit"]) + 1 #API中记录的赛段序号从0开始，平常所说的赛季序号从1开始，因此要加1（The split recorded in API counts from 0, while the split that people usually talk about counts from 1, so 1 should be added here）
    if region == "TENCENT":
        folder: str = "顶尖排位玩家（Ranked Apex）/国服（TENCENT）/%s/第%d赛季 - 第%d赛段（SEASON %d - Split %d）" %(platform_TENCENT[platformId], currentLoLSeasonId, currentLoLSeasonSplitId, currentLoLSeasonId, currentLoLSeasonSplitId)
    elif region == "GARENA":
        folder = "顶尖排位玩家（Ranked Apex）/竞舞（GARENA）/%s/第%d赛季 - 第%d赛段（SEASON %d - Split %d）" %(platform_GARENA[platformId], currentLoLSeasonId, currentLoLSeasonSplitId, currentLoLSeasonId, currentLoLSeasonSplitId)
    else: #拳头公司与竞舞娱乐公司的合同于2023年1月终止（In January 2023, Riot Games ended its contract with Garena）
        folder = "顶尖排位玩家（Ranked Apex）/外服（RIOT）/%s/第%d赛季 - 第%d赛段（SEASON %d - Split %d）" %((platform_RIOT | platform_GARENA)[platformId], currentLoLSeasonId, currentLoLSeasonSplitId, currentLoLSeasonId, currentLoLSeasonSplitId)
    
    # splitsConfig: dict[str, Any] = await (await connection.request("GET", "/lol-ranked/v1/splits-config")).json()
    # json1name: str = "SplitsConfig (Season %d).json" %currentLoLSeasonId
    # while True:
    #     try:
    #         with open(os.path.join(folder, json1name), "w", encoding = "utf-8") as jsonfile1:
    #             json.dump(splitsConfig, jsonfile1, indent = 4, ensure_ascii = False)
    #     except FileNotFoundError:
    #         os.makedirs(folder, exist_ok = True)
    #     except UnicodeEncodeError:
    #         logPrint("\n赛季信息文本文档生成失败！请检查内容是否包含不常用字符！\nSplit config text generation failure! Please check if the content includes any abnormal characters!\n")
    #         break
    #     else:
    #         logPrint('\n赛季信息已保存为“%s”。\nSplit config is saved as "%s".\n' %(os.path.join(folder, json1name), os.path.join(folder, json1name)))
    #         break
    # splits_info_header_keys: list[str] = list(splits_info_header.keys())
    # splits_info_data: dict[str, Any] = {key: [] for key in splits_info_header_keys}
    # logPrint("赛季信息整理进度（Split config organization process）：")
    # for i in range(len(splitsConfig["splits"])):
    #     split: dict[str, Any] = splitsConfig["splits"][i]
    #     for j in range(len(splits_info_header_keys)):
    #         key: str = splits_info_header_keys[j]
    #         if j <= 5:
    #             if j >= 4: #时间相关键（Time-related keys）
    #                 to_append: Any = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime(split[key + "Millis"] // 1000))
    #             else:
    #                 to_append = split[key]
    #         elif j == 6: #胜利系列皮肤奖励：物品识别码（`victoriousSkinReward: itemInstanceId`）
    #             to_append = split["victoriousSkinRewardGroup"]["itemInstanceId"]
    #         else:
    #             to_append = split["victoriousSkinRewardGroup"]["splitPointsByHighestSeasonEndTier"][key[27:]] if key[27:] in split["victoriousSkinRewardGroup"]["splitPointsByHighestSeasonEndTier"] else 0
    #         player_loot_data[key].append(to_append)
    #     logPrint("[%d/%d]Season %s - Split %s" %(i + 1, len(splitsConfig["splits"]), split["seasonId"], split["splitId"]), end = "\r", print_time = True)
    # else:
    #     logPrint("已完成。 | Done.")
    # splits_info_statistics_output_order: list[int] = [1, 2, 3, 5, 0, 4, 6, 13, 7, 16, 11, 15, 10, 9, 14, 12, 8]
    # splits_info_data_organized: dict[str, list[Any]] = {splits_info_header_keys[i]: splits_info_data[splits_info_header_keys[i]] for i in splits_info_statistics_output_order}
    # splits_info_df: pandas.DataFrame = pandas.DataFrame(data = splits_info_data_organized)
    # splits_info_df = pandas.concat([pandas.DataFrame([splits_info_header])[splits_info_df.columns], splits_info_df], ignore_index = True)
    
    # rewardTypes: dict[str, str] = {"CHAMPION_TOKEN": "成就代币", "EMOTE": "永久表情", "ETERNALS_CAPSULE": "永恒星碑魔法引擎", "HEXTECH_CHEST": "海克斯科技宝箱", "HEXTECH_KEY": "海克斯科技钥匙", "HEXTECH_KEY_FRAGMENT": "海克斯科技钥匙碎片", "MASTERWORK_CHEST": "杰作宝箱", "MYSTERY_EMOTE": "神秘表情", "ORANGE_ESSENCE": "橙色精萃", "SUMMONER_ICON": "召唤师图标", "WARD_SHARD": "守卫碎片"}
    # #下面整理奖品识别码和奖品名称的对应关系（The following code sorts out the relationship between the itemInstanceIds and the rewardNames）
    # rewardNames: dict[str, str] = {}
    # inventoryTypes: list[str] = ["ACHIEVEMENT_BANNER_ACCENT", "ACHIEVEMENT_TITLE", "ANNOUNCER_PACK", "AUGMENT", "AUGMENT_SLOT", "BOOST", "BUNDLES", "CHAMPION", "CHAMPION_SKIN", "CHERRY_BOON", "COMPANION", "CURRENCY", "EMOTE", "EVENT_PASS", "FANPASS", "GIFT", "HEXTECH_CRAFTING", "MODE_PROGRESSION_REWARD", "MYSTERY", "NEXUS_FINISHER", "PREMIUM_CLUB_MEMBERSHIP", "PROVIEW_PASS", "PVE_RELIC", "PVE_SUMMONER_PACKAGE", "PVE_UPGRADE", "QUEUE_ENTRY", "REGALIA_BANNER", "REGALIA_BORDER", "REGALIA_CREST", "RP", "RUNE", "SKIN_AUGMENT", "SKIN_BORDER", "SKIN_UPGRADE_GEAR", "SKIN_UPGRADE_HOME_GUARD", "SKIN_UPGRADE_RECALL", "SKIN_UPGRADE_SPAWN", "SPELL_BOOK_PAGE", "STATSTONE", "STRAWBERRY_BOON", "STRAWBERRY_LOADOUT_ITEM", "STRAWBERRY_MAP", "SUMMONER_CUSTOMIZATION", "SUMMONER_ICON", "TEAMPASS", "TEAM_SKIN_PURCHASE", "TFT_DAMAGE_SKIN", "TFT_EVENT_SKILLS", "TFT_MAP_SKIN", "TFT_PLAYBOOK", "TFT_ZOOM_SKIN", "TOURNAMENT_FLAG", "TOURNAMENT_FRAME", "TOURNAMENT_LOGO", "TOURNAMENT_TROPHY", "TRANSFER", "WARD_SKIN"]
    # for inventoryType in inventoryTypes:
    #     items: list[dict[str, Any]] = await (await connection.request("GET", "/lol-catalog/v1/items/" + inventoryType)).json()
    #     rewardNames |= {item["itemInstanceId"]: item["name"] for item in items}
    # rewardTrack_header_keys: list[str] = list(rewardTrack_header.keys())
    # rewardTrack_data: dict[str, list[Any]] = {key: [] for key in rewardTrack_header_keys}
    # logPrint("奖励里程整理进度（Reward track organization process）：")
    # for i in range(len(splitsConfig["splits"])):
    #     split: dict[str, Any] = splitsConfig["splits"][i]
    #     for j in range(len(split["rewardTrack"])):
    #         rewards: list[dict[str, Any]] = split["rewardTrack"][j]["rewards"]
    #         for k in range(len(rewards)):
    #             reward: dict[str, Any] = rewards[k]
    #             for l in range(len(rewardTrack_header_keys)):
    #                 key = rewardTrack_header_keys[l]
    #                 if l == 0: #赛季序号（`seasonId`）
    #                     to_append: Any = split["seasonId"]
    #                 elif l == 1: #奖励顺序（`rewardTrackId`）
    #                     to_append = j
    #                 else:
    #                     if l == 8: #奖品类型（`rewardType`）
    #                         to_append = rewardTypes[reward["rewardType"]]
    #                     elif l == 10: #奖品名称（`name`）
    #                         to_append = rewardNames.get(reward["id"], "")
    #                     else:
    #                         to_append = reward[key]
    #                 rewardTrack_data[key].append(to_append)
    #             logPrint("[%d/%d][%d/%d][%d/%d]%s\t%s" %(i + 1, len(splitsConfig["splits"]), j + 1, len(split["rewardTrack"]), k + 1, len(rewards), reward["id"], rewardNames.get(reward["id"], "")), end = "\r", print_time = True)
    # else:
    #     logPrint("已完成。 | Done.")
    # rewardTrack_statistics_output_order: list[int] = [0, 9, 1, 10, 4, 8, 2, 6, 3, 5, 7]
    # rewardTrack_data_organized: dict[str, list[Any]] = {rewardTrack_header_keys[i]: rewardTrack_data[rewardTrack_header_keys[i]] for i in rewardTrack_statistics_output_order}
    # rewardTrack_df: pandas.DataFrame = pandas.DataFrame(data = rewardTrack_data_organized)
    # rewardTrack_df = pandas.concat([pandas.DataFrame([rewardTrack_header])[rewardTrack_df.columns], rewardTrack_df], ignore_index = True)
    
    tiers_zh: dict[str, str] = {"": "", "NONE": "没有段位", "IRON": "坚韧黑铁", "BRONZE": "英勇黄铜", "SILVER": "不屈白银", "GOLD": "荣耀黄金", "PLATINUM": "华贵铂金", "EMERALD": "流光翡翠", "DIAMOND": "璀璨钻石", "MASTER": "超凡大师", "GRANDMASTER": "傲世宗师", "CHALLENGER": "最强王者", "SALT": "SALT", "WOOD": "WOOD"}
    tiers_en: dict[str, str] = {"": "", "NONE": "NONE", "IRON": "IRON", "BRONZE": "BRONZE", "SILVER": "SILVER", "GOLD": "GOLD", "PLATINUM": "PLATINUM", "EMERALD": "EMERALD", "DIAMOND": "DIAMOND", "MASTER": "MASTER", "GRANDMASTER": "GRANDMASTER", "CHALLENGER": "CHALLENGER", "SALT": "SALT", "WOOD": "WOOD"}
    ratedTiers_turbo: dict[str, str] = {"": "", "NONE": "没有段位", "GRAY": "灰白", "GREEN": "翠绿", "BLUE": "天蓝", "PURPLE": "绛紫", "ORANGE": "耀橙"}
    ratedTiers_cherry: dict[str, str] = {"": "", "NONE": "没有段位", "GRAY": "木木角斗士", "GREEN": "青铜角斗士", "BLUE": "白银角斗士", "PURPLE": "黄金角斗士", "ORANGE": "王者角斗士"}
    #ratedTiers: dict[str, str] = {"": "", "NONE": "NONE", "GRAY": "GRAY", "GREEN": "GREEN", "BLUE": "BLUE", "PURPLE": "PURPLE", "ORANGE": "ORANGE"}
    queueTypes_zh: dict[str, str] = {"RANKED_SOLO_5x5": "单人/双人", "RANKED_FLEX_SR": "灵活 5V5", "RANKED_TFT": "云顶之弈", "RANKED_TFT_PAIRS": "2V0", "RANKED_TFT_DOUBLE_UP": "双人作战", "RANKED_TFT_TURBO": "狂暴模式", "CHERRY": "斗魂竞技场", "RANKED_PREMADE_5x5": "5V5"} #2V0模式仅美测服可用（RANKED_TFT_PAIRS is only available on PBE）
    queueTypes_en: dict[str, str] = {"RANKED_SOLO_5x5": "Ranked Solo/Duo", "RANKED_FLEX_SR": "Ranked Flex", "RANKED_TFT": "Ranked TFT", "RANKED_TFT_PAIRS": "2V0", "RANKED_TFT_DOUBLE_UP": "Double Up", "RANKED_TFT_TURBO": "Hyper Roll", "CHERRY": "Arena", "RANKED_PREMADE_5x5": "5V5"}
    challenger_ladder_queueTypes: list[str] = await (await connection.request("GET", "/lol-ranked/v1/challenger-ladders-enabled")).json()
    
    challenger_ladder_metadata_header_keys: list[str] = list(challenger_ladder_metadata_header.keys())
    challenger_ladder_metadata: dict[str, list[Any]] = {key: [] for key in challenger_ladder_metadata_header_keys}
    challenger_ladder_header_keys: list[str] = list(challenger_ladder_header.keys())
    ladder_data: dict[str, dict[str, list[Any]]] = {"challenger_ladder": {}, "topRated_ladder": {}}
    ladder_dfs: dict[str, dict[str, pandas.DataFrame]] = {"challenger_ladder": {}, "topRated_ladder": {}}
    for queueType in challenger_ladder_queueTypes:
        ladder_data["challenger_ladder"][queueType] = {key: [] for key in challenger_ladder_header_keys}
        queue_ladder_data: dict[str, list[Any]] = ladder_data["challenger_ladder"][queueType] #注意字典赋值的原理（Pay attention to the principle of assigning a dictionary）
        for tier in ["CHALLENGER", "GRANDMASTER", "MASTER"]:
            ladders: dict[str, Any] = await (await connection.request("GET", f"/lol-ranked/v1/apex-leagues/{queueType}/{tier}")).json()
            json2name = "Apex-%s-%s(%s).json" %(tier.capitalize(), queueType, runTime_day)
            while True:
                try:
                    with open(os.path.join(folder, json2name), "w", encoding = "utf-8") as jsonfile2:
                        json.dump(ladders, jsonfile2, indent = 4, ensure_ascii = False)
                except FileNotFoundError:
                    os.makedirs(folder, exist_ok = True)
                except UnicodeEncodeError:
                    logPrint("\n顶级%s%s玩家信息文本文档生成失败！请检查内容是否包含不常用字符！\nTop %s %s player information generation failure! Please check if the content includes any abnormal characters!\n" %(queueTypes_zh[queueType], tiers_zh[tier], queueTypes_en[queueType], tiers_en[tier]))
                    break
                else:
                    logPrint('\n顶级%s%s玩家信息已保存为“%s”。\nTop %s %s player information is saved as "%s".\n' %(queueTypes_zh[queueType], tiers_zh[tier], os.path.join(folder, json2name), queueTypes_en[queueType], tiers_en[tier], os.path.join(folder, json2name)))
                    break
            logPrint("顶级%s%s玩家信息整理进度（Top %s %s player information organization process）：" %(queueTypes_zh[queueType], tiers_zh[tier], queueTypes_en[queueType], tiers_en[tier]))
            # ladder_summoner_infos: dict[str, dict[str, Any]] = await get_infos(connection, puuids = [standing["puuid"] for ladder in ladders for division in ladders["divisions"]])
            for i in range(len(ladders["divisions"])):
                division: dict[str, Any] = ladders["divisions"][i]
                for j in range(len(challenger_ladder_metadata_header_keys)):
                    key: str = challenger_ladder_metadata_header_keys[j]
                    if j <= 4:
                        if j == 2: #队列类型（`queueType`）
                            to_append: Any = queueTypes_zh[ladders[key]]
                        elif j == 4: #下次天梯更新时间（`nextApexUpdateTime`）
                            to_append = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime(ladders["nextApexUpdateMillis"] // 1000))
                        else:
                            to_append = ladders[key]
                    else:
                        if j == 6: #段位分级（`division`）
                            to_append = "" if division[key] == "" else division[key]
                        elif j == 9: #段位（`tier`）
                            to_append = tiers_zh[division[key]]
                        elif j == 11: #天梯解锁时间（`apexUnlockTime`）
                            to_append = time.strftime("%Y年%m月%d日%H时%M分%S秒", time.localtime(division["apexUnlockTimeMillis"] // 1000))
                        else:
                            to_append = division[key]
                    challenger_ladder_metadata[key].append(to_append)
                for j in range(len(division["standings"])):
                    standing: dict[str, Any] = division["standings"][j]
                    # info_got: bool = standing["puuid"] in ladder_summoner_infos
                    # standing_summoner: dict[str, Any] = ladder_summoner_infos.get(standing["puuid"], {})
                    standing_summoner_recapture: int = 0
                    standing_summoner: dict[str, Any] = await get_info(connection, standing["puuid"])
                    while not standing_summoner["info_got"] and standing_summoner["body"]["httpStatus"] != 404 and standing_summoner_recapture < 3:
                        logPrint(standing_summoner["body"])
                        standing_summoner_recapture += 1
                        logPrint("第%d/%d名顶级%s%s%s玩家（玩家通用唯一识别码：%s）信息获取失败！正在第%d次尝试重新获取该玩家信息……\n[%d/%d] Information of Player (Puuid: %s) in the %s %s %s apex capture failed! Recapturing this player's information ... Times tried: %d" %(j + 1, len(division["standings"]), queueTypes_zh[queueType], tiers_zh[tier], division["division"], standing["puuid"], standing_summoner_recapture, j + 1, len(division["standings"]), standing["puuid"], queueTypes_zh[queueType], tiers_zh[tier], division["division"], standing_summoner_recapture))
                        standing_summoner = await get_info(connection, standing["puuid"])
                    info_got: bool = standing_summoner["info_got"]
                    if not info_got:
                        logPrint(standing_summoner["body"])
                        logPrint("第%d/%d名顶级%s%s%s玩家（玩家通用唯一识别码：%s）信息获取失败！\n[%d/%d] Information of Player (Puuid: %s) in the %s %s %s apex capture failed!" %(j + 1, len(division["standings"]), queueTypes_zh[queueType], tiers_zh[tier], division["division"], standing["puuid"], j + 1, len(division["standings"]), standing["puuid"], queueTypes_zh[queueType], tiers_zh[tier], division["division"]))
                    for k in range(len(challenger_ladder_header_keys)):
                        key: str = challenger_ladder_header_keys[k]
                        if k == 0 or k == 11:
                            to_append: Any = "" if standing[key] == "NA" else standing[key]
                        elif k == 12 or k == 18:
                            to_append = tiers_zh[standing[key]]
                        elif k <= 19:
                            to_append = standing[key]
                        else:
                            to_append = standing_summoner["body"][key] if info_got else ""
                        queue_ladder_data[key].append(to_append)
                    logPrint("[%d/%d][%d/%d]%s\t%s" %(i + 1, len(ladders["divisions"]), j + 1, len(division["standings"]), standing["puuid"], get_info_name(standing_summoner["body"]) if info_got else ""), end = "\r", print_time = True)
            else:
                logPrint("已完成。 | Done.")
        challenger_ladder_statistics_output_order: list[int] = [8, 10, 9, 16, 14, 17, 20, 21, 18, 0, 3, 2, 13, 7, 6, 5, 19, 4, 12, 11, 1, 15]
        queue_ladder_data_organized: dict[str, list[Any]] = {challenger_ladder_header_keys[i]: queue_ladder_data[challenger_ladder_header_keys[i]] for i in challenger_ladder_statistics_output_order}
        ladder_data["challenger_ladder"][queueType] = queue_ladder_data_organized #注意字典赋值的原理。该语句其实无关紧要（Pay attention to the principle of assigning a dictionary. This statement is actually unnecessary）
        ladder_dfs["challenger_ladder"][queueType] = pandas.DataFrame(data = queue_ladder_data_organized)
        logPrint("正在优化逻辑值显示……\nOptimizing the display of boolean values ...")
        optimize_bool_display(ladder_dfs["challenger_ladder"][queueType])
        logPrint("逻辑值显示优化完成！\nBoolean value display optimization finished!")
        ladder_dfs["challenger_ladder"][queueType] = pandas.concat([pandas.DataFrame([challenger_ladder_header])[ladder_dfs["challenger_ladder"][queueType].columns], ladder_dfs["challenger_ladder"][queueType]], ignore_index = True)
    challenger_ladder_metadata_statistics_output_order: list[int] = [2, 9, 6, 1, 7, 10, 8, 3, 0, 4, 5, 11]
    challenger_ladder_metadata_organized: dict[str, list[Any]] = {challenger_ladder_metadata_header_keys[i]: challenger_ladder_metadata[challenger_ladder_metadata_header_keys[i]] for i in challenger_ladder_metadata_statistics_output_order}
    challenger_ladder_metadata_df = pandas.DataFrame(data = challenger_ladder_metadata_organized)
    challenger_ladder_metadata_df = pandas.concat([pandas.DataFrame([challenger_ladder_metadata_header])[challenger_ladder_metadata_df.columns], challenger_ladder_metadata_df], ignore_index = True)
    
    topRated_ladder_queueTypes: list[str] = await (await connection.request("GET", "/lol-ranked/v1/top-rated-ladders-enabled")).json()
    topRated_ladder_header_keys: list[str] = list(topRated_ladder_header.keys())
    for queueType in topRated_ladder_queueTypes:
        ladders: dict[str, Any] = await (await connection.request("GET", f"/lol-ranked/v1/rated-ladder/{queueType}")).json()
        json3name = "RatedApex-%s(%s).json" %(queueType, runTime_hour)
        while True:
            try:
                with open(os.path.join(folder, json3name), "w", encoding = "utf-8") as jsonfile3:
                    json.dump(ladders, jsonfile3, indent = 4, ensure_ascii = False)
            except FileNotFoundError:
                os.makedirs(folder, exist_ok = True)
            except UnicodeEncodeError:
                logPrint("\n顶级%s玩家信息文本文档生成失败！请检查内容是否包含不常用字符！\nTop %s player information generation failure! Please check if the content includes any abnormal characters!\n" %(queueTypes_zh[queueType], queueTypes_en[queueType]))
                break
            else:
                logPrint('\n顶级%s玩家信息已保存为“%s”。\nTop %s player information is saved as "%s".\n' %(queueTypes_zh[queueType], os.path.join(folder, json3name), queueTypes_en[queueType], os.path.join(folder, json3name)))
                break
        ladder_data["topRated_ladder"][queueType] = {key: [] for key in topRated_ladder_header_keys}
        queue_ladder_data: dict[str, list[Any]] = ladder_data["topRated_ladder"][queueType]
        logPrint("顶级%s玩家信息整理进度（Top %s player information organization process）：" %(queueTypes_zh[queueType], queueTypes_en[queueType]))
        # ladder_summoner_infos: dict[str, dict[str, Any]] = await get_infos(connection, puuids = [standing["puuid"] for ladder in ladders for division in ladders["divisions"]])
        for i in range(len(ladders["standings"])):
            standing: dict[str, Any] = ladders["standings"][i]
            # info_got: bool = standing["puuid"] in ladder_summoner_infos
            # standing_summoner: dict[str, Any] = ladder_summoner_infos.get(standing["puuid"], {})
            standing_summoner_recapture: int = 0
            standing_summoner: dict[str, Any] = await get_info(connection, standing["puuid"])
            while not standing_summoner["info_got"] and standing_summoner["body"]["httpStatus"] != 404 and standing_summoner_recapture < 3:
                logPrint(standing_summoner["body"])
                standing_summoner_recapture += 1
                logPrint("第%d/%d名顶级%s玩家（玩家通用唯一识别码：%s）信息获取失败！正在第%d次尝试重新获取该玩家信息……\n[%d/%d] Information of Player (Puuid: %s) in the %s apex capture failed! Recapturing this player's information ... Times tried: %d" %(i + 1, len(division["standings"]), queueTypes_zh[queueType], standing["puuid"], standing_summoner_recapture, i + 1, len(division["standings"]), standing["puuid"], queueTypes_zh[queueType], standing_summoner_recapture))
                standing_summoner = await get_info(connection, standing["puuid"])
            info_got: bool = standing_summoner["info_got"]
            if not info_got:
                logPrint(standing_summoner["body"])
                logPrint("第%d/%d名顶级%s玩家（玩家通用唯一识别码：%s）信息获取失败！\n[%d/%d] Information of Player (Puuid: %s) in the %s apex capture failed!" %(i + 1, len(division["standings"]), queueTypes_zh[queueType], standing["puuid"], i + 1, len(division["standings"]), standing["puuid"], queueTypes_zh[queueType]))
            for j in range(len(topRated_ladder_header_keys)):
                key: str = topRated_ladder_header_keys[j]
                if j <= 8:
                    if j == 5:
                        to_append: Any = ratedTiers_cherry[standing[key]] if queueType == "CHERRY" else ratedTiers_turbo[standing[key]]
                    else:
                        to_append = standing[key]
                else:
                    to_append = standing_summoner["body"][key] if info_got else ""
                queue_ladder_data[key].append(to_append)
            logPrint("[%d/%d]%s\t%s" %(i + 1, len(ladders["standings"]), standing["puuid"], get_info_name(standing_summoner["body"]) if info_got else ""), end = "\r", print_time = True)
        else:
            logPrint("已完成。 | Done.")
        topRated_ladder_statistics_output_order: list[int] = [1, 3, 2, 6, 4, 7, 9, 10, 5, 0, 8]
        queue_ladder_data_organized: dict[str, list[Any]] = {topRated_ladder_header_keys[i]: queue_ladder_data[topRated_ladder_header_keys[i]] for i in topRated_ladder_statistics_output_order}
        ladder_data["topRated_ladder"][queueType] = queue_ladder_data_organized
        ladder_dfs["topRated_ladder"][queueType] = pandas.DataFrame(data = queue_ladder_data_organized)
        logPrint("正在优化逻辑值显示……\nOptimizing the display of boolean values ...")
        optimize_bool_display(ladder_dfs["topRated_ladder"][queueType])
        logPrint("逻辑值显示优化完成！\nBoolean value display optimization finished!")
        ladder_dfs["topRated_ladder"][queueType] = pandas.concat([pandas.DataFrame([topRated_ladder_header])[ladder_dfs["topRated_ladder"][queueType].columns], ladder_dfs["topRated_ladder"][queueType]], ignore_index = True)
    
    logPrint("是否导出以上天梯数据至Excel中？（输入任意键导出，否则不导出）\nDo you want to export the above data into Excel? (Press any key to export or null to refuse exporting)")
    export_str: str = logInput()
    export: bool = bool(export_str)
    if export:
        excel_name: str = f"Ranked Apex - {platformId} ({currentLoLSeasonId}-{currentLoLSeasonSplitId}).xlsx"
        excel_name_sorted: str = f"Ranked Apex - {platformId} ({currentLoLSeasonId}-{currentLoLSeasonSplitId}) (sorted).xlsx"
        wbPath: str = os.path.join(folder, excel_name)
        os.makedirs(folder, exist_ok = True)
        if not os.path.exists(wbPath):
            wbCreateFlag: bool = create_workbook_win32(os.path.abspath(wbPath), sheet1_name = f"Tier Apex Metadata - Season {currentLoLSeasonId}")
        workbook_exist: bool = os.path.exists(wbPath)
        while True:
            try:
                with (pandas.ExcelWriter(path = wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(path = wbPath)) as writer:
                    # addDefaultStyle(splits_info_df).to_excel(excel_writer = writer, sheet_name = f"Split Config - Season {currentLoLSeasonId}")
                    # logPrint("赛季信息导出完成！\nSplit config exported!\n")
                    # addDefaultStyle(rewardTrack_df).to_excel(excel_writer = writer, sheet_name = f"Reward Track - Season {currentLoLSeasonId}")
                    # logPrint("奖励里程导出完成！\nReward milestones exported!\n")
                    addDefaultStyle(challenger_ladder_metadata_df).to_excel(excel_writer = writer, sheet_name = f"Tier Apex Metadata - Season {currentLoLSeasonId}")
                    logPrint("胜点系列段位天梯元数据导出完成！\nLP apex metadata exported!\n")
                    # addDefaultStyle(topRated_ladder_metadata_df).to_excel(excel_writer = writer, sheet_name = f"Rating Apex Metadata - Season {currentLoLSeasonId}")
                    # logPrint("排名分系列段位天梯元数据导出完成！\nRating apex metadata exported!\n")
                    runTimes: list[float] = [] #记录保存每个队列的顶级玩家信息所花费的时间（Records the time spent in saving the top player information of each queue）
                    total_used: float = 0
                    ladders_reserved: int = 0
                    for ladderType in ladder_dfs:
                        for queueType in ladder_dfs[ladderType]:
                            start: float = time.time()
                            logPrint("正在导出顶级%s玩家信息……\nExporting top %s player information ..." %(queueTypes_zh[queueType], queueTypes_en[queueType]))
                            addDefaultStyle(ladder_dfs[ladderType][queueType]).to_excel(excel_writer = writer, sheet_name = queueType + " " + (runTime_day if ladderType == "challenger_ladder" else runTime_hour))
                            ladders_reserved += 1
                            end: float = time.time()
                            unit: float = end - start
                            total_used += unit
                            runTimes.append(unit)
                            total_remaining: float = 0 if sum([i for i in runTimes[:ladders_reserved + 1]]) == 0 else sum([i for i in runTimes[:ladders_reserved + 1]]) / ladders_reserved * (len(ladder_dfs["challenger_ladder"]) + len(ladder_dfs["topRated_ladder"]) - ladders_reserved)
                            logPrint("保存该段位排位天梯所花费的时间（Time spent in saving this match）： %s" %(format_runtime(unit)))
                            logPrint("已花费的总时间（Total time used）                                ： %s" %(format_runtime(total_used)))
                            logPrint("剩余时间（Time remaining）                                       ： %s" %(format_runtime(total_remaining)))
                            logPrint("预计总时间（Expected total time）                                ： %s" %(format_runtime(total_used + total_remaining)), end = "\n\n")
            except PermissionError:
                logPrint("无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试。\nPermission denied! Please ensure the file isn't opened right now or read-only! Press any key to try again.")
                logInput()
            else:
                logPrint("各队列顶级玩家信息导出完成！\nTop player information of all queues exported!\n")
                break
        if workbook_exist:
            logPrint("警告：由于该文件已存在，本次导出已追加新工作表到工作簿的末尾。这可能导致队列和时间顺序的错乱。是否需要对工作表进行排序？（输入任意键排序，否则不排序）\nWarning: Because the excel workbook has existed, new sheets are appended to the last of the original sheet list. This may result in the disarrangement of queue and time orders. Do you want to sort the sheets? (Input anything to sort the sheets, or null to skip sorting)")
            sort_str: str = logInput()
            sort: bool = bool(sort_str)
            if sort:
                logPrint("正在读取刚刚创建的工作表……\nLoading the workbook just created ...")
                while True:
                    try:
                        wb: Workbook = load_workbook(wbPath)
                    except FileNotFoundError:
                        logPrint('排位天梯工作簿读取失败！请确保“%s”文件夹内含有名为“%s”的工作簿。如果需要退出程序，请输入“0”。\nERROR reading the ranked apex workbook! Please make sure the workbook "%s" is in the folder "%s". If you want to exit the program, please submit "0".' %(folder, excel_name, excel_name, folder))
                        apex_reload = logInput()
                        if apex_reload == "0":
                            break
                    else:
                        sheetnames: list[str] = wb.sheetnames #第一次获取原工作簿的工作表名称列表（The first time to get the sheet name list of the original workbook）
                        #下面锁定工作表顺序（The following code determine the sheet order）
                        logPrint("正在创建顺序工作表列表……\nCreating the ordered sheet list ...")
                        ##第一部分：赛季信息类工作表（Part 1: Split config sheets）
                        split_config_dict: dict[int, str] = {int(sheet_iter.split()[-1]): sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Split Config")}
                        reward_track_dict: dict[int, str] = {int(sheet_iter.split()[-1]): sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Reward Track")}
                        ##第二部分：天梯元数据工作表（Part 2: Apex metadata sheets）
                        tier_apex_metadata_dict: dict[int, str] = {int(sheet_iter.split()[-1]): sheet_iter for sheet_iter in sheetnames if sheet_iter.startswith("Tier Apex Metadata")}
                        ##第三部分：天梯工作表（Part 3: Apex sheets）
                        challenger_ladders_dict: dict[str, dict[str, str]] = {}
                        topRated_ladders_dict: dict[str, dict[str, str]] = {}
                        for sheet_iter in sheetnames:
                            if any(sheet_iter.split(maxsplit = 1)[0] == queueType for queueType in challenger_ladder_queueTypes):
                                queueType_tmp = sheet_iter.split(maxsplit = 1)[0] #以工作表名的队列部分为排序依据（Sort the sheetnames by the queueType part of the sheet name）
                                time_str = sheet_iter.split(maxsplit = 1)[1] #目前暂不需要考虑时间因工作表名长度限制而被截断的问题（Currently the issue that the time may be cut off due to the sheet name length limit doesn't need to be considered）
                                if not queueType_tmp in challenger_ladders_dict:
                                    challenger_ladders_dict[queueType_tmp] = {}
                                challenger_ladders_dict[queueType_tmp][time_str] = sheet_iter
                            elif any(sheet_iter.split(maxsplit = 1)[0] == queueType for queueType in topRated_ladder_queueTypes):
                                queueType_tmp = sheet_iter.split(maxsplit = 1)[0] #以工作表名的队列部分为排序依据（Sort the sheetnames by the queueType part of the sheet name）
                                time_str = sheet_iter.split(maxsplit = 1)[1] #目前暂不需要考虑时间因工作表名长度限制而被截断的问题（Currently the issue that the time may be cut off due to the sheet name length limit doesn't need to be considered）
                                if not queueType_tmp in topRated_ladders_dict:
                                    topRated_ladders_dict[queueType_tmp] = {}
                                topRated_ladders_dict[queueType_tmp][time_str] = sheet_iter
                        sheetnames_sorted: list[str] = [] #所有工作表的期望顺序存储在sheetnames_sorted变量中（The ordered result of all sheets is stored in the variable `sheetnames_sorted`）
                        for season in sorted(set(split_config_dict.keys()) | set(reward_track_dict.keys())): #第一部分：赛季信息类工作表（Part 1: Split config sheets）
                            if split_config_dict[season] in sheetnames:
                                sheetnames_sorted.append(split_config_dict[season])
                            if reward_track_dict[season] in sheetnames:
                                sheetnames_sorted.append(reward_track_dict[season])
                        for season in sorted(tier_apex_metadata_dict.keys()): #第二部分：天梯元数据工作表（Part 2: Apex metadata sheets）
                            if tier_apex_metadata_dict[season] in sheetnames:
                                sheetnames_sorted.append(tier_apex_metadata_dict[season])
                        for queueType_iter in challenger_ladder_queueTypes + topRated_ladder_queueTypes: #第三部分：天梯工作表（Part 3: Apex sheets）
                            if queueType_iter in challenger_ladders_dict:
                                for time_iter in sorted(challenger_ladders_dict[queueType_iter].keys()): #队列顺序以API中记录的队列顺序为准。下同（The queueType order of sheets adopts that recorded in API. So does the following）
                                    sheetnames_sorted.append(challenger_ladders_dict[queueType_iter][time_iter])
                            if queueType_iter in topRated_ladders_dict:
                                for time_iter in sorted(topRated_ladders_dict[queueType_iter].keys()):
                                    sheetnames_sorted.append(topRated_ladders_dict[queueType_iter][time_iter])
                        #下面排列所有工作表（The following code arrange all sheets）
                        logPrint("正在排序……\nOrdering ...")
                        sort_worksheet(wb, sheetnames_sorted)
                        logPrint('正在保存中……\nSaving the ordered workbook ...')
                        wb.save(os.path.join(folder, excel_name_sorted))
                        logPrint('排序完成！排好序的工作簿已保存为“%s”。\nOrdering finished! The ordered workbook is saved as "%s".\n' %(excel_name_sorted, excel_name_sorted))
                        wb.close()
                        break

#-----------------------------------------------------------------------------
# websocket
#-----------------------------------------------------------------------------
@connector.ready
async def connect(connection: Connection) -> None:
    global logInput, logPrint
    log_folder: str = "日志（Logs）/Customized Program 13 - Fetch Ranked Apex"
    os.makedirs(log_folder, exist_ok = True)
    currentTime: str = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    log: LogManager = LogManager(path = os.path.join(log_folder, currentTime + ".log"), mode = "a+", encoding = "utf-8")
    logInput = log.logInput
    logPrint = log.logPrint
    await print_summoner_info(connection)
    await get_challenger_tier(connection)
    log.write("\n[Program terminated and returned status 0.]\n")
    log.close()

@connector.close
async def disconnect(connection: Connection) -> None:
    print("已从英雄联盟客户端断开连接。\nDisconnected from the League Client.")

#-----------------------------------------------------------------------------
# Main
#-----------------------------------------------------------------------------

runTime_day: str = time.strftime("%Y-%m-%d", time.localtime())
runTime_hour: str = time.strftime("%Y-%m-%d %Hh", time.localtime())
connector.start()
