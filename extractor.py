import argparse, copy, json, os, pandas, re, requests, time
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Iterable, Literal, Optional
from src.utils.logger import LogManager
from src.utils.patch import Patch, get_cdragon_patchList
from src.utils.webRequest import requestUrl
from src.utils.format import format_df, addDefaultStyle, eliminate_empty_fields
from src.utils.runtimeDebug import subscope
from src.utils.excel_workbook import create_workbook_win32, sort_worksheet
from src.core.config.localization import language_ddragon, language_dict
from src.core.extractor.base import verifyDictHeterogeneity, syncListOrder, traverse_keyPath, getBinaryKeys, LoLDataExtractor
from src.core.extractor.types import *

#=============================================================================
# * 声明（Declaration）
#=============================================================================
# 作者（Author）：          WordlessMeteor
# 主页（Home page）：       https://github.com/WordlessMeteor/LoL-DIY-Programs/
# 鸣谢（Acknowledgement）： Morilli, Le poussin, Moga
# 更新（Last update）：     2026/08/15
#=============================================================================

#定义模式覆盖文本描述函数（Define the overriden data tooltip function）
def modeOverrideTooltipTransform(binData: dict[str, Any], dType: Literal["champion", "item"], objectType: str, keyPaths: str | list[str], gameModeName: Literal["TUTORIAL", "TUTORIAL_MODULE_1", "TUTORIAL_MODULE_2", "TUTORIAL_MODULE_3", "SWIFTPLAY", "PRACTICETOOL", "FIRSTBLOOD", "ARSR", "ARAM", "{bffdf499}", "KINGPORO", "URF", "SNOWURF", "ONEFORALL", "{6462680f}", "NEXUSBLITZ", "TFT", "ASSASSINATE", "ULTBOOK", "cherry", "STRAWBERRY", "{a110bc47}", "Ruby", "DOOMBOTSTEEMO", "{b0cea932}", "{afcea79f}", "{aecea60c}", "{9cf6bf22}"], strtable: dict[str, int | dict[str, str]]) -> str: #这个函数只用于制作英雄平衡表格，不用于本程序（This function is only designed for making the balance table, not for this program）
    '''
    遍历二进制描述数据中的模式覆盖数据并输出。<br>Traverse through the mode overriden values in binary description data and output them.
    
    :param binData: 完整的二进制描述数据，通常通过形如session.get(url).json()的代码直接获得。注意，二进制描述的预处理已在本函数内完成，所以传入该参数的二进制描述数据不能预先被处理过。<br>The complete binary description data, which is often obtained through code like `session.get(url).json()`. Note that the pre-processing of a binary description is finished within this function, so the value to pass to this parameter shouldn't be pre-processed in advance.
    :type binData: dict[str, Any]
    :param dType: 数据类型。决定输出格式。<br>Data type. It determines the output format.
    
        如果指定为“champion”，函数会构建从法术对象到技能热键的映射，于是输出中会带有技能热键。<br>If it's specified as "champion", this function will construct a map from SpellObjects to spell hotkeys, so the output will display the ability hotkey.
    :type dType: Literal["champion", "item"]
    :param objectType: 要遍历的对象类型，通常为某个一级对象的__type键的值。例如，在英雄二进制描述数据中，一般指定该参数为“CharacterRecord”。<br>Object type to traverse through the binData, usually the value of `__type` key of a Level-1 object. For example, when it comes to traversing champion binary description data, this parameter is usually specified as "CharacterRecord".
    :type objectType: str
    :param keyPaths: 键路径，由键通过竖线组成的字符串，详见getBinaryKeys函数的文档字符串。<br>Paths of keys, concatenated by "|" from keys. Detailed are in the DocString of `getBinaryKeys` function.<br>参考数据覆盖键路径：<br>Some known mode override keyPaths corresponding to objectTypes:
        <pre>
        **objectType**      **keyPath**                         **description**<br>
        SpellObject     mSpell|{f9c2333e}               模式覆盖冷却时间（Mode overriden cooldown time）<br>
        SpellObject     mSpell|{b08bc498}               模式覆盖充能时间（Mode overriden charge time）<br>
        SpellObject     mSpell|DataValuesModeOverride   模式覆盖数值（Mode overriden data values）<br>
        ItemData        DataValuesModeOverride          模式覆盖数值（Mode overriden data values）
        </pre>
    :type keyPaths: str
    :param gameModeName: 游戏模式代号，在地图二进制描述中可找到。<br>The game mode name that can be found in the map binary description.<br>目前已知的游戏模式代号及其名称：<br>Currently known gameModeNames and the localized names:
        <pre>
        **gameModeName**      **description**<br>
        TUTORIAL            新手教程（Tutorial）<br>
        TUTORIAL_MODULE_2   新手教程 第二部分（Tutorial Part 2）<br>
        TUTORIAL_MODULE_3   新手教程 第三部分（Tutorial Part 3）<br>
        TUTORIAL_MODULE_1   新手教程 第一部分（Tutorial Part 1）<br>
        SWIFTPLAY           快速模式（Swiftplay）<br>
        PRACTICETOOL        训练模式（Practice Tool）<br>
        FIRSTBLOOD          大对决（Showdown）<br>
        ARSR                峡谷大乱斗（ARSR）<br>
        ARAM                极地大乱斗（ARAM）<br>
        {bffdf499}          海克斯大乱斗（ARAM: Mayhem）<br>
        KINGPORO            魄罗大乱斗（Poro King）<br>
        URF                 无限火力（Ultra Rapid Fire）<br>
        SNOWURF             冰雪无限火力（Snow Ultra Rapid Fire）<br>
        ONEFORALL           克隆大作战（One for All）<br>
        {6462680f}          {c706490e}<br>
        NEXUSBLITZ          极限闪击（Nexus Blitz）<br>
        TFT                 云顶之弈（Teamfight Tactics）<br>
        ASSASSINATE         红月决（Blood Moon Hunt）<br>
        ULTBOOK             终极魔典（Ultimate Spellbook）<br>
        cherry              斗魂竞技场（Arena）<br>
        STRAWBERRY          无尽狂潮（Swarm）<br>
        {a110bc47}          神木之门（Brawl）<br>
        Ruby                末日人工智能（Doom Bots）<br>
        DOOMBOTSTEEMO       末日人工智能（Doom Bots Teemo）<br>
        {b0cea932}          末日人工智能：维迦的诅咒！（Doom Bots - Veigar's Curse!）<br>
        {afcea79f}          末日人工智能：维迦的邪咒！（Doom Bots - Veigar's Evil!）<br>
        {aecea60c}          末日人工智能：维迦的末日厄咒！（Doom Bots - Veigar's Doom!）<br>
        {9cf6bf22}          WASD<br>
        {ad33a648}          海克斯大乱斗 经典模式版（ARAM: Mayhem Classic-ish）<br>
        {5358c483}          BASELINESR<br>
        {20426d6f}          英雄联盟经典模式（League Classic）
        </pre>
    :type gameModeName: str
    :param strtable: 字符串常量池。<br>Stringtable.
    :type strtable: dict[str, int | dict[str, str]]
    :return: 二进制描述binData中的每个objectType项目在gameModeName中的某些符合keyPaths的属性覆盖值的说明文本。<br>Tooltip of some `gameModeName` overriden values that correspond to `keyPaths` in each `objectType` object in the binary description `binData`.
    :rtype: str
    '''
    #预处理参数（Pre-process parameters）
    binData = LoLDataExtractor.normalizeBinData(binData) #因此，传入的binData不能预先被normalizeSpellData函数处理过（Hence, the passed `binData` should be in raw format and not pre-processed by `normalizeBinData` function）
    if isinstance(keyPaths, str):
        keyPathList: list[list[str]] = [keyPaths.split("|")]
    elif isinstance(keyPaths, list) and all(map(lambda x: isinstance(x, str), keyPaths)):
        keyPathList = list(map(lambda x: x.split("|"), keyPaths))
    else:
        print(f"警告：您传入的键路径有误。请检查。函数将返回空字符串。\nWarning: Invalid `keyPath`! Please check it. The function will return an empty string instead.")
        return ""
    if dType != "champion":
        dType = "item"
    #构建热键映射（Build the hotkey map）
    hotkey_map: dict[str, str] = {} #从指令主键到技能热键的映射（A map from spell keys to ability hotkeys）
    characterRecordKey_spellKey_map: dict[str, str] = {} #从指令主键到角色记录主键的映射（A map from spell keys to character record keys）
    if dType == "champion":
        abilityKey_rootSpellKey_map: dict[str, str] = {} #从根指令主键到技能对象主键的映射（A map from root spell keys to ability object keys）
        for (key, value) in binData.items():
            if key != "__linked" and value["__type"] == "AbilityObject":
                abilityKey_rootSpellKey_map[value["mRootSpell"]] = key
        for (key, value) in binData.items():
            if key != "__linked" and value["__type"] == "CharacterRecord":
                if "mCharacterPassiveSpell" in value:
                    if value["mCharacterPassiveSpell"] in abilityKey_rootSpellKey_map and value["mCharacterPassiveSpell"] in abilityKey_rootSpellKey_map and "mChildSpells" in binData[abilityKey_rootSpellKey_map[value["mCharacterPassiveSpell"]]]:
                        for spellKey in binData[abilityKey_rootSpellKey_map[value["mCharacterPassiveSpell"]]]["mChildSpells"]:
                            hotkey_map[spellKey] = "P" #这里假设同一个技能只可能属于一个英雄，即不存在某个指令是一个英雄的被动技能，但是是另一个英雄的Q技能的情况。下同（Here we assume that a spell can only belong to one champion, i.e. a spell can't be both a champion's passive but also another champion's Q ability. Same in below）
                            characterRecordKey_spellKey_map[spellKey] = key
                    else:
                        hotkey_map[value["mCharacterPassiveSpell"]] = "P"
                        characterRecordKey_spellKey_map[value["mCharacterPassiveSpell"]] = key
                hotkey_list: list[str] = ["Q", "W", "E", "R"]
                if "spells" in value:
                    for i in range(len(value["spells"])): #事先已知每个角色记录对象都有“spells”键，且其值的长度都是4（It's known in advance that each CharacterRecord object has a "spells" key and its value is always a list of length 4）
                        spellKey: str = value["spells"][i]
                        if spellKey in abilityKey_rootSpellKey_map and "mChildSpells" in binData[abilityKey_rootSpellKey_map[spellKey]]:
                            for spellKey in binData[abilityKey_rootSpellKey_map[spellKey]]["mChildSpells"]:
                                hotkey_map[spellKey] = hotkey_list[i]
                                characterRecordKey_spellKey_map[spellKey] = key
                        else:
                            hotkey_map[spellKey] = hotkey_list[i]
                            characterRecordKey_spellKey_map[spellKey] = key
    #遍历二进制描述数据（Traverse binary description data）
    s: str = "" #初始化结果字符串（Initialize the result string）
    for (key, value) in binData.items():
        if key != "__linked" and value["__type"] == objectType:
            tmp_ptr: Any = value
            for keyPath in keyPathList:
                for tmp_key in keyPath:
                    if tmp_key in tmp_ptr:
                        tmp_ptr = tmp_ptr[tmp_key]
                    else:
                        break
                else:
                    endKey: str = keyPath[-1]
                    value = copy.deepcopy(value)
                    if value["__type"] == "SpellObject":
                        value["mSpell"] = LoLDataExtractor.normalizeBinData(value["mSpell"])
                    elif value["__type"] == "ItemData":
                        value = LoLDataExtractor.normalizeBinData(value)
                    if endKey in {"{f9c2333e}", "{b08bc498}", "DataValuesModeOverride"}:
                        if gameModeName in tmp_ptr:
                            #获取技能名称（Get ability name）
                            if value["__type"] == "SpellObject":
                                keyNamePath: list[str] = ["mSpell", "mClientData", "mTooltipData", "mLocKeys", "keyName"]
                            elif value["__type"] == "ItemData":
                                keyNamePath = ["mDisplayName"]
                                # keyNamePath = ["mItemDataClient", "mTooltipData", "mLocKeys", "keyName"]
                            else:
                                keyNamePath = []
                            if keyNamePath == []:
                                keyName_content: str = key
                            else:
                                tmp_ptr1: Any = value
                                for tmp_key1 in keyNamePath:
                                    if tmp_key1 in tmp_ptr1:
                                        tmp_ptr1 = tmp_ptr1[tmp_key1]
                                    else:
                                        keyName_content = key
                                        break
                                else:
                                    keyName = tmp_ptr1
                                    keyName_content = LoLDataExtractor.get_strtable_value(strtable, keyName, default = keyName)
                            #获取技能热键（Get ability hotkey）
                            if value["__type"] == "SpellObject" and dType == "champion": #只有在数据类型为英雄时才获取其角色名称和技能热键（Only when the data type is champion will the character name and hotkey be obtained）
                                characterName: str = binData[characterRecordKey_spellKey_map[key]]["mCharacterName"] if key in characterRecordKey_spellKey_map else ""
                                mHotKey: str = hotkey_map.get(key, "")
                            else:
                                characterName = mHotKey = ""
                            #获取缺省值和覆盖值（Get the default and overriden value）
                            if endKey == "{f9c2333e}" or endKey == "{b08bc498}": #基础冷却时间和基础充能时间（Basic cooldown and basic charge time）
                                if endKey == "{f9c2333e}": #基础冷却时间（Basic cooldown）
                                    if "cooldownTime" in value["mSpell"]:
                                        if mHotKey == "R":
                                            time_list: list[float] = value["mSpell"]["cooldownTime"][1:4]
                                        else:
                                            time_list = value["mSpell"]["cooldownTime"][1:6]
                                        defaultValue: str = LoLDataExtractor.burnValueList(time_list)
                                    else:
                                        defaultValue = "φ"
                                else: #基础充能时间（Basic charge time）
                                    if "mAmmoRechargeTime" in value["mSpell"]:
                                        if mHotKey == "R":
                                            time_list = value["mSpell"]["mAmmoRechargeTime"][1:4]
                                        else:
                                            time_list = value["mSpell"]["mAmmoRechargeTime"][1:6]
                                        defaultValue = LoLDataExtractor.burnValueList(time_list)
                                    else:
                                        defaultValue = "φ"
                                if mHotKey == "R":
                                    overrideValue: str = LoLDataExtractor.burnValueList(tmp_ptr[gameModeName]["value"][1:4])
                                else:
                                    overrideValue = LoLDataExtractor.burnValueList(tmp_ptr[gameModeName]["value"][1:6])
                                if endKey == "{f9c2333e}":
                                    s += "{%s}【%s】（%s）：基础冷却时间：%s秒 → %s秒\n" %(characterName, keyName_content, mHotKey, defaultValue, overrideValue)
                                else:
                                    s += "{%s}【%s】（%s）：基础充能时间：%s秒 → %s秒\n" %(characterName, keyName_content, mHotKey, defaultValue, overrideValue)
                            elif endKey == "DataValuesModeOverride":
                                overrideValues: dict[str, Any] = tmp_ptr[gameModeName]
                                if overrideValues["__type"] == "SpellDataValueVector":
                                    s += "{%s}【%s】（%s）：" %(characterName, keyName_content, mHotKey)
                                    for i in range(len(overrideValues["SpellDataValues"])):
                                        spellDataValue: dict[str, Any] = overrideValues["SpellDataValues"][i]
                                        var: str = spellDataValue["name"] if "name" in spellDataValue else spellDataValue["mName"]
                                        if var.lower() in value["mSpell"]["DataValues"]:
                                            varData: dict[str, Any] = value["mSpell"]["DataValues"][var.lower()]
                                            if mHotKey == "R":
                                                value_list: list[float] = varData["values"][1:4] if "values" in varData else varData["mValues"][1:4]
                                            else:
                                                value_list = varData["values"][1:6] if "values" in varData else varData["mValues"][1:6]
                                            defaultValue: str = LoLDataExtractor.burnValueList(value_list)
                                        else:
                                            varData = {}
                                            defaultValue = "φ"
                                        if mHotKey == "R":
                                            overrideValue = LoLDataExtractor.burnValueList(spellDataValue["values"][1:4]) if "values" in spellDataValue else varData["mValues"][1:4] if "mValues" in varData else "φ"
                                        else:
                                            overrideValue = LoLDataExtractor.burnValueList(spellDataValue["values"][1:6]) if "values" in spellDataValue else varData["mValues"][1:6] if "mValues" in varData else "φ"
                                        s += "{%s}：%s → %s" %(var, defaultValue, overrideValue)
                                        if i < len(overrideValues["SpellDataValues"]) - 1:
                                            s += "；"
                                        else:
                                            s += "\n"
                                elif overrideValues["__type"] == "ItemDataValues":
                                    s += "【%s】：" %(keyName_content)
                                    for i in range(len(overrideValues["DataValues"])):
                                        itemDataValue = overrideValues["DataValues"][i]
                                        var: str = itemDataValue["name"] if "name" in itemDataValue else itemDataValue["mName"]
                                        if var.lower() in value["mDataValues"]:
                                            defaultValue: str = str(LoLDataExtractor.aRound(value["mDataValues"][var.lower()]["mValue"], 5))
                                        else:
                                            defaultValue = "φ"
                                        overrideValue = str(LoLDataExtractor.aRound(itemDataValue["mValue"], 5))
                                        s += "{%s}：%s → %s" %(var, defaultValue, overrideValue)
                                        if i < len(overrideValues["DataValues"]) - 1:
                                            s += "；"
                                        else:
                                            s += "\n"
    return s

#定义音效提取类（Define the sfx extractor class）
class LoLSfxExtractor:
    def __init__(self, log: Optional[LogManager] = None) -> None:
        '''
        初始化音效提取器类对象。<br>Initialize a `LoLSfxExtractor` class object.
        
        这个类目前只是一个初步构想。未来预期设计成一个建立音频库单元事件和封装后的音频文件中每个音频文件名之间的对应关系的类，并且有可能会放到其它模块中。<br>This class is a very preemptive concept. I plan to design it as a class that can build a map between bank unit events and names of audio files in encapsulated files. This class may be moved to another module some day.
        '''
        self.log: LogManager = LogManager() if log == None else log
    
    @classmethod
    def compute_bankEvent_hash(cls, s: str) -> int:
        '''
        使用FNV-1算法计算某个音频库单元事件字符串的hash值。<br>Compute the hash value of a bank unit string using FNV-1 algorithm.
        
        :return: 字符串的hash值。<br>Hash value of the string.
        :rtype: int
        '''
        basis: int = 0x811c9dc5 #偏移基准（Offset basis）
        hash_int: int = basis
        for b in s.lower().encode("ascii"):
            hash_int = ((hash_int * 0x01000193) ^ b) % 0x100000000
        return hash_int

if __name__ == "__main__":
    parser: argparse.ArgumentParser = argparse.ArgumentParser(formatter_class = argparse.RawTextHelpFormatter)
    parser.add_argument("--sfx", help = "启用音频库单元hash计算调试。\nEnable bank unit hash calculation debugging.", action = "store_true")
    args: argparse.Namespace = parser.parse_args()
    
    cwd: str = os.getcwd().replace("\\", "/")
    if cwd.endswith("src/core"): #允许用户直接双击脚本（Users are allowed to double click this program）
        os.chdir("../..")
    elif cwd.endswith("src"):
        os.chdir("..")
    log_folder: str = "日志（Logs）/游戏数据提取脚本/"
    os.makedirs(log_folder, exist_ok = True)
    currentTime: str = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
    log: LogManager = LogManager(os.path.join(log_folder, currentTime + ".log"), mode = "a+", encoding = "utf-8")
    logInput = log.logInput
    logPrint = log.logPrint
    #定义语言设置过程（Define the process of setting language）
    def set_locale(initial_launch: bool = True, old_locale: str = "") -> str:
        '''
        设置全局语言环境。<br>Set the global locale.
        
        :param initial_launch: 该函数是否在程序刚开始运行时调用。默认为真。<br>Whether this function is called at the beginning of the program execution. True by default.
        
            如果是刚开始运行时调用，那么输入“0”直接退出程序。否则输入“0”取消语言更改操作。<br>If it is, then submitting "0" will directly exit the program. Otherwise, submitting "0" simply cancels the language changing operation.
        :type initial_launch: bool
        :param old_locale: 旧语言文化代码。仅在`initial_launch`参数为假时有用。默认为空字符串。<br>Old language code. It makes a difference only when `initial_launch` is False. An empty string by default.
        :type old_locale: str
        :return: 语言文化代码。<br>Language code.
        :rtype: str
        '''
        logPrint("请选择说明文本的输出语言【默认为中文（中国）】：\nPlease select a language for tooltips (the default option is zh_CN):")
        language_df: pandas.DataFrame = pandas.DataFrame(language_dict)
        logPrint(format_df(language_df)[0], write_time = False)
        while True:
            language_option = logInput()
            if language_option == "" or language_option in set(map(str, range(1, 31))):
                if language_option == "":
                    language_option = "29"
                language_code = list(language_ddragon.keys())[int(language_option) - 1]
                break
            elif language_option[0] == "0":
                language_code = "" if initial_launch else old_locale
                break
            else:
                logPrint("语言选项输入错误！请重新输入：\nERROR input of language option! Please try again:")
        return language_code

    #定义版本设置过程（Define the process of setting version）
    def set_version(session: Optional[requests.Session] = None) -> tuple[list[str], requests.Session]:
        '''
        设置游戏数据版本。<br>Set the game data version.
        
        :param session: 网络请求会话。如果没有指定，则内部新建一个会话，对外不可见。<br>Web request session. If unspecified, a new session will be created, which isn't visible to outside.
        :type session: requests.Session | None
        :return: 大版本号列表和网络请求会话。<br>List of major version numbers and web request session.
        :rtype: tuple[list[str], requests.Session]
        '''
        if session == None:
            session = requests.Session()
        logPrint("请在以下版本号中选择并输入完整的版本号：\nPlease select a version and then enter it entirely:")
        patches_cdragon, patchList_fetched = get_cdragon_patchList(session = session, log = log)
        if not patchList_fetched:
            return ([], session)
        logPrint(json.dumps(patches_cdragon, ensure_ascii = False))
        while True:
            version: str = logInput()
            if version == "":
                versions: list[str] = ["pbe"]
                break
            elif version == "all":
                versions = patches_cdragon
                break
            elif version[0] == "0":
                versions = []
                break
            elif version in patches_cdragon:
                versions = [version]
                break
            else:
                version = version.replace("both", '["latest", "pbe"]') #相当于将“both”视为一个输入时的保留字，可参与计算，如`both * 2`。当然，事先已知合法的单个版本不包含“both”（Equivalent to regarding "both" as a reserved word, enabling it to take part in calculation like `both * 2`. Of course, we know in advance that a legal version can't contain "both"）
                try:
                    versions = eval(version)
                except:
                    logPrint("您的输入有误，请重新输入！\nERROR input! Please try again.")
                else:
                    if isinstance(versions, list) and all(map(lambda x: x in patches_cdragon, versions)):
                        break
                    else:
                        logPrint("您的输入有误，请重新输入！\nERROR input! Please try again.")
        return (versions, session)

    #定义工作表排序函数（Define worksheet sorting function）
    def sort_workbook_sheets(wbPath: str, naming_pattern: int) -> None:
        '''
        对游戏数据提取工作簿的工作表进行排序。<br>Sort the sheets of Game Data Extraction workbook.
        
        :param wbPath: 游戏数据提取工作簿路径。<br>Game Data Extraction workbook path.
        :type wbPath: str
        :param naming_pattern: 命名模式代号。有以下两个取值：<br>Naming pattern code, which has the following two values:
        
            - 1: 中文备注英文。如“斗魂竞技场强化符文（Cherry Augments）”。<br>Chinese with English note. E.g. "斗魂竞技场强化符文（Cherry Augments）".
            - 2: 版本号和英文。如“16.8.7638450 CherryAugments”。<br>Patch number and English. E.g. "16.8.7638450 CherryAugments".
        :type naming_pattern: int
        '''
        if not os.path.exists(wbPath):
            logPrint("文件不存在。不执行任何操作。\nFile not found. No operation will be performed.")
            return
        elif not os.path.isfile(wbPath):
            logPrint("参数不是文件。不执行任何操作。\nThe parameter isn't a file. No operation will be performed.")
            return
        elif not os.path.splitext(wbPath)[1] == ".xlsx":
            logPrint("文件格式错误。不执行任何操作。\nFile format error. No operation will be performed.")
            return
        if naming_pattern != 1 and naming_pattern != 2:
            logPrint("命名模式代号有误。不执行任何操作。\nNaming pattern code error. No operation will be performed.")
            return
        wbPath = wbPath.replace("\\", "/")
        wbPath_sorted: str = " (sorted)".join(os.path.splitext(wbPath))
        #读取工作簿（Read the workbook）
        try:
            wb: Workbook = load_workbook(wbPath)
        except Exception as e:
            logPrint(e)
            logPrint("工作簿读取失败。不执行任何操作。\nWorkbook reading failed. No operation will be performed.")
            return
        #首先整理出所有工作表的排列顺序（First, sort out the order of all sheets）
        sheetnames: list[str] = wb.sheetnames
        logPrint("正在创建顺序工作表列表……\nCreating the ordered sheet list ...", print_time = True)
        if naming_pattern == 1:
            sheetnames_order: list[str] = [
                "地图（Map）",
                "指令集（CheatSet）",
                "指令（Cheat）",
                "召唤师技能（Summoner Spells）",
                "符文系（PerkStyles）",
                "符文（Perks）",
                "英雄（Champions）",
                "英雄技能（Champion Spells）",
                "角色（Characters）",
                "角色技能（Character Spells）",
                "装备（Items）",
                "装备分组（Item Groups）",
                "装备修饰（Item Modifiers）",
                "斗魂竞技场强化符文（Cherry Augments）",
                "无尽狂潮强化（Swarm Augments）",
                "海克斯大乱斗强化符文（Kiwi Augments）",
                "海克斯大乱斗强化符文套装（Kiwi Augment Set）",
                "海克斯大乱斗任务线（Kiwi Questlines）",
                "海克斯大乱斗经典强化符文（KiwiJade Augments）",
                "斗魂竞技场锻造器（Cherry Anvils）",
                "海克斯大乱斗锻造器（Kiwi Anvils）",
                "斗魂竞技场回合列表（Cherry Round List）",
                "斗魂竞技场回合（Cherry Round）",
                "斗魂竞技场阶段（Cherry Phase）",
                "斗魂竞技场回合阶段（Cherry Round Phase）",
                "斗魂竞技场场景英雄（Cherry Cameos）",
                "斗魂竞技场荣誉嘉宾（Cherry Guests）",
                "云顶之弈赛季（TFT Set）",
                "云顶之弈商店（TFT Shop）",
                "云顶之弈商店内容（TFT Shop Content）",
                "云顶之弈掉率表（TFT Drop Rate）",
                "云顶之弈回合阶段（TFT Stage Round）",
                "云顶之弈回合（TFT Round）",
                "云顶之弈传送门（TFT Portal）",
                "云顶之弈开场奇遇（TFT Encounter Distribu",
                "云顶之弈开场奇遇（TFT Encounter Distribution）",
                "云顶之弈奇遇（TFT Encounter）",
                "云顶之弈单位属性（TFT Unit Property）",
                "云顶之弈角色定位（TFT Character Role）",
                "云顶之弈装备列表（TFT Item List）",
                "云顶之弈装备（TFT Item）",
                "云顶之弈羁绊列表（TFT Trait List）",
                "云顶之弈羁绊（TFT Trait）",
                "云顶之弈电脑玩家英雄（TFT PVE NPC）",
                "云顶之弈脚本（TFT Script）",
                "云顶之弈通告（TFT Announcement）",
                "字体描述（Font Description）",
                "字体类型（Font Types）",
                "字体分辨率（Font Resolution）",
                "字体样式（Font Style）",
                "CSS样式（CSS Style）",
                "内嵌图标（Inline Icons）"
            ]
        else:
            pVersion_dataType: re.Pattern[str] = re.compile(r"\d+(\.\d+)*\s\w+") #定义正则表达式来检验工作表名称是否符合整合工作簿中的工作表格式——版本号+数据类型（Define a regular expression to verify whether a sheet name obeys the format of sheets in an integrated workbook: version number + data type）
            version_order: list[Patch] = sorted(set(Patch(name.split()[0]) for name in sheetnames if pVersion_dataType.fullmatch(name))) #提取工作表的版本部分，整理形成正序版本列表（Extract the version part of sheet names and organize them into a ascending list）
            dataType_order: list[str] = [
                "Map",
                "CheatSet",
                "Cheat",
                "SummonerSpells",
                "PerkStyles",
                "Perks",
                "Champions",
                "ChampionSpells",
                "Characters",
                "CharacterSpells",
                "Items",
                "ItemGroups",
                "ItemModifiers",
                "CherryAugments",
                "SwarmAugments",
                "KiwiAugments",
                "KiwiAugmentSet",
                "KiwiQuestline",
                "KiwiJadeAugments",
                "CherryAnvils",
                "KiwiAnvils",
                "CherryRoundList",
                "CherryRound",
                "CherryPhase",
                "CherryRoundPhase",
                "CherryCameos",
                "CherryGuests",
                "TFTSet",
                "TFTShop",
                "TFTShopContent",
                "TFTDropRate",
                "TFTStageRound",
                "TFTRound",
                "TFTPortal",
                "TFTEncounterDistri", #游戏版本的第二位是个位数的情形（The case where the second number of game version is 1-digit）
                "TFTEncounterDistr",
                "TFTEncounterDistribution",
                "TFTEncounter",
                "TFTUnitProperty",
                "TFTCharacterRole",
                "TFTItemList",
                "TFTItem",
                "TFTTraitList",
                "TFTTrait",
                "TFTPVENPC",
                "TFTScript",
                "TFTAnnouncement",
                "FontDescription",
                "FontTypes",
                "FontResolution",
                "FontStyle",
                "FontCSSStyle",
                "InlineIcons",
            ]
            tmpDf: pandas.DataFrame = pandas.DataFrame(data = [{"name": name, "version_weight": version_order.index(Patch(name.split()[0])), "type_weight": dataType_order.index(name.split()[1])} for name in sheetnames if pVersion_dataType.fullmatch(name)]) #忽略名称不合法的工作表（Bypass sheets with illegal names）
            tmpDf_sorted: pandas.DataFrame = tmpDf.sort_values(by = ["version_weight", "type_weight"]) #工作表名称按照版本的正序和数据类型的正序进行排列（Arrange sheet names in the ascending orders of versions and data types）
            sheetnames_order = tmpDf_sorted["name"].to_list()
        sheetnames_sorted: list[str] = [] #所有工作表的期望顺序存储在sheetnames_sorted变量中（The ordered result of all sheets is stored in the variable `sheetnames_sorted`）
        for sheet_iter in sheetnames_order:
            if sheet_iter in sheetnames:
                sheetnames_sorted.append(sheet_iter)
        #然后排列所有工作表（Then, arrange all sheets）
        logPrint("正在排序……\nOrdering ...", print_time = True)
        sort_worksheet(wb, sheetnames_sorted)
        logPrint("正在保存中……\nSaving the ordered workbook ...", print_time = True)
        wb.save(wbPath_sorted)
        logPrint("排序完成！排好序的工作簿已保存为以下文件。\nOrdering finished! The ordered workbook is saved as the following file.\n%s" %(wbPath_sorted), print_time = True)
        wb.close()

    #定义主函数（Define the main function）
    def main() -> int:
        '''
        主函数，用于发行版。<br>Main function for the release version.
        
        主要读取在线数据资源。<br>Mainly loads online data resources.
        
        :return: 状态码。<br>Status code.
        :rtype: int
        '''
        #初始化网络请求会话（Initialize the web request session）
        session = requests.Session()
        # session.trust_env = False #忽略系统代理设置（Bypass system proxy）

        #设置语言（Set the language）
        language_code: str = set_locale(initial_launch = True)
        if language_code == "":
            return 1

        #设置版本（Set the version）
        versions, session = set_version(session = session)
        if len(versions) == 0:
            return 2
        versions_conf: list[str] = [] #初始化具有相同配置的版本列表临时变量。`versions`根据元素的重复情况可分成多个批次。例如`["latest", "pbe", "latest", "pbe"]`可以分成两批`["latest", "pbe"]`，`["latest", "pbe", "latest", "latest"]`可以分成`["latest", "pbe"]`、`["latest"]`和`["latest"]`共三批。该变量存储的是具有相同配置的一个批次，主要用于模拟版本回溯（Initialize a version list temporary variable whose elements have the same configuration. `versions` can be divided into multiple batches. For example, `["latest", "pbe", "latest", "pbe"]` can be divided into two batches of `["latest", "pbe"]`, and `["latest", "pbe", "latest", "latest"]` can be divided into 3 batches, namely `["latest", "pbe"]`, `["latest"]` and `["latest"]`. This variable stores one batch where all versions have the same configuration and is mainly designed to simulate version backtrack）
        
        #设置工作表集成（Determine whether to integrate sheets in different patches into one workbook）
        logPrint("是否将不同版本的工作表集成到一个工作簿中？（输入任意非空字符串以确认集成，否则分不同版本保存。）\nDo you want to integrate sheets of different versions into a single workbook? (Input any non-empty string to confirm integration, or null to save data into multiple workbooks of the different version.)")
        integrate_str: str = logInput()
        integrate: bool = bool(integrate_str)
        
        #设置自动化操作（Set automatic operations）
        preset_data_options: list[int] = [] #保留第一个版本的导出数据类型（Reserve data types to export in the first version）
        if len(versions) > 1:
            logPrint('''是否启用一键式导出？（输入任意非空字符串以禁用，否则启用。启用时，用户在第一个版本按下“-1”清空队列时，后续版本将只会整理和导出第一个版本获取过的数据。）\nDo you want to enable one-click export? (Submit any non-empty string to disable it, otherwise enable it. If it's enabled, when the user submits "-1" to empty the dataframe queue, the subsequent versions will only organize and export data of the same types as of the first version.)''')
            one_click_str: str = logInput()
            one_click: bool = not bool(one_click_str)
            if one_click:
                logPrint("你看，他们像柱子一样！\nColumn like you see 'em.")
            else:
                logPrint("已禁用一键式导出。您将需要手动设置每个版本要导出的数据类型。\nOne-click has been disabled. You will need to manually set data types for each version.")
        else:
            one_click = False
        
        #设置默认导出行为（Set the default export behavior）
        wb_export: bool = True
        web_export: bool = False
        logPrint('数据将只导出为Excel工作簿。如果需要导出为网页，请在选择数据类型的步骤输入“-2”以设置导出选项。\nData will only be exported to Excel workbooks. If you want to export data into web, please input "-2" in the data type selection step to set export options.')
        single_export: bool = False
        logPrint('''数据在完成整理后不会立刻导出为Excel工作簿。如果需要在整理后立刻导出，请在选择数据类型的步骤输入“-2”以设置导出选项。\nData won't be exported into Excel workbooks immediately after being organized. If you want to export data immediately after they're organized, please input "-2" in the data type selection step to set export options.''')
        
        #设置样式保留行为（Set CSS retention behavior）
        LoLDataExtractor.set_tooltip_layout(False)
        logPrint('''说明文本变量代换过程默认不保留CSS样式。如果需要保留，请在选择数据类型的步骤输入“-2”以调整样式选项。\nCSS styles aren't retained during the variable substitution process of tooltips by default. If you want to retain them, please input "-2" in the data type selection step to set the CSS retention option.''')
        
        #设置变量代换过程中的变量名保留行为（Set the variable name retention behavior in the variable substitution process）
        LoLDataExtractor.set_variable_reserve_strategy(False)
        logPrint('''说明文本变量代换过程默认不保留变量名。如果需要保留，请在选择数据类型的步骤输入“-2”以调整变量代换选项。\nVariable names aren't retained during the variable substitution process of tooltips by default. If you want to retain them, please input "-2" in the data type selection step to set the variable name retention option.''')
        
        #设置hash值解析深度（Set the hash resolution depth）
        LoLDataExtractor.set_resolution_depth(False)
        logPrint('程序默认只解析hash值。如果需要统一不同版本间的字符串大小写，请在选择数据类型的步骤输入“-2”以设置hash值解析深度。\nThe program only resolves hash values by default. If you want to unify the string cases among different versions, please input "-2" in the data type selection step to set the hash resolution depth.')
        
        #设置等级计算的等级上限（Set the level cap for level scaling calculations）
        LoLDataExtractor.set_levelScaling_cap(18)
        logPrint('等级计算的等级上限默认为18级。如果需要调整，请在选择数据类型的步骤输入“-2”以调整等级上限。\nThe level cap for level scaling calculations is 18 by default. If you want to adjust it, please input "-2" in the data type selection step to adjust the level cap.')
        
        #设置数据框导出密度（Set dataframe export density）
        LoLDataExtractor.set_export_density(True)
        logPrint('程序默认消除空字段。如果需要保留所有数据框的可导出的列，请在选择数据类型的步骤输入“-2”以设置导出密度。\nThe program removes empty fields. If you want to reserve all dataframe columns that can be exported, please input "-2" in the data type selection step to set the export density.')
        
        for i in range(len(versions)):
            version: str = versions[i]
            if version in versions_conf:
                versions_conf.clear() #每当程序遍历到一个在此列表中已经存在的版本时，将此列表清空（Every time the program traverses a version already in this list, clear this list）
            versions_conf.append(version)
            logPrint("[%d/%d]开始处理%s版本的游戏数据。\nStart to process game data of Version %s." %(i + 1, len(versions), version, version))
            extractor = LoLDataExtractor(version, language_code, session = session, log = log)
            if integrate:
                extractor.encapsulate()
            else:
                extractor.decapsulate()
            #加载二进制描述数据的字符串散列表（Load the string hashtable for binary description data）
            if i == 0: #字符串散列表是静态的，适用于所有版本，只需加载一次即可（The string hashtable is static and applicable to all versions, so it only needs to be loaded once）
                logPrint("正在加载二进制描述数据的字符串散列表……\nLoading the string hashtable for binary description data ...", print_time = True)
                extractor.get_bin_hashes()
            #加载版本数据（Load version data）
            logPrint(f"正在加载完整的游戏版本号……\nLoading the complete version number ...", print_time = True)
            extractor.get_version()
            if extractor.patch == "" or extractor.patch_number == "":
                continue
            print(f"当前版本号（Current patch）： {extractor.patch}")
            #创建保存目录（Create saving directory）
            extractor.make_dir()
            #加载文件导出列表（Load file export list）
            logPrint(f"正在加载文件导出列表……\nLoading the file export list ...", print_time = True)
            extractor.get_exported_files()
            if not extractor.fileExportList_ready:
                continue
            #加载中英文字符串常量池（Load the stringtable in Chinese and English）
            logPrint("正在加载字符串常量池……\nLoading stringtables ...", print_time = True)
            extractor.get_strtable()
            if not (extractor.strtable_organize_manner == 1 and extractor.strtables_ready["lol_target"] and extractor.strtables_ready["lol_default"] and extractor.strtables_ready["tft_target"] and extractor.strtables_ready["tft_default"]) and not (extractor.strtable_organize_manner == 2 and extractor.strtables_ready["target"] and extractor.strtables_ready["default"]):
                continue
            #加载共享数据（Load shared data）
            logPrint("正在加载共享数据……\nLoading shared data ...", print_time = True)
            extractor.init_mSpells()
            if not extractor.shared_ready:
                logPrint("共享数据获取失败。将忽略该数据。\nShared data capture failure! The program will ignore them.")
                # continue
            #初始化计数器（Initialize counter）
            nDataOptions: int = 0
            nDataOption_iter: int = 0
            step: int = 1
            #设置要提取的数据类型（Set the type of data to extract）
            while True:
                logPrint("请选择您要提取的数据：\nPlease select the type of data you want to extract:\n-2\t设置（Settings）\n0\t退出当前版本（Quit this version）\n1\t地图（Maps）\n2\t作弊指令（Cheat sheet）\n3\t召唤师技能（Summoner Spells）\n4\t符文（Perks）\n5\t英雄（Champions）\n6\t角色（Characters）\n7\t装备（Items）\n8\t强化符文（Augments）\n9\t锻造器（Anvils）\n10\t斗魂竞技场回合阶段（Arena Round Phase）\n11\t场景英雄（Cameo）\n12\t荣誉嘉宾（Guests of Honor）\n13\t云顶之弈赛季、装备和羁绊（TFT Sets, Items and Traits）\n14\t字体（Fonts）\nall\t所有（All）" + ("" if single_export else "\n-1\t批量导出所有数据框并清空队列（Batch export all dataframes and clear queue）"))
                if one_click and len(versions_conf) > 1: #新批次开始时，`versions_conf`中只有一个元素。从此批次开始重新配置（When a new batch starts, there's only one element in `versions_conf`. Reconfigure from this batch）
                    if step == 1:
                        mode: str = str(preset_data_options)
                    elif step == 2:
                        mode = "-1"
                    elif step == 3:
                        mode = "0"
                    else:
                        logPrint("异常步骤！\nStep error!")
                        break
                    logPrint(mode)
                    step += 1
                else:
                    mode: str = logInput()
                if mode == "":
                    continue
                elif mode == "-2":
                    logPrint("请选择一个配置：\nPlease select an configuration option:\n0\t返回上一层（Return to the last step）\n1\t导出为工作簿（Export to workbook）\n2\t单类数据导出（Single-type data export）\n3\t导出为网页（Export to web）\n4\t切换语言（Switch language）\n5\t说明文本样式（Tooltip style）\n6\t变量替换样式（Variable substitution style）\n7\thash值解析深度（Hash value resolution depth）\n8\t等级计算上限（Level scaling cap）\n9\t切换数据框导出密度（Switch dataframe export density）")
                    while True:
                        option: str = logInput()
                        if option == "":
                            continue
                        elif option == "-1":
                            return 0
                        elif option[0] == "0":
                            break
                        elif option[0] == "1":
                            logPrint("是否将数据导出为Excel工作簿？（输入任意非空字符串以将数据导出为Excel工作簿，否则不导出。）\nDo you want to export data into Excel workbooks? (Submit any non-empty string to export data into Excel workbooks, or null to refuse.)")
                            wb_export_str: str = logInput()
                            wb_export = bool(wb_export_str)
                            if wb_export:
                                logPrint("数据将导出为Excel工作簿。\nData will be exported into Excel workbooks.")
                            else:
                                logPrint("数据将不会导出为Excel工作簿。\nData won't be exported into Excel workbooks.")
                        elif option[0] == "2":
                            logPrint('是否选择在整理数据后立刻将其导出到Excel中？（输入任意非空字符串以选择单项导出并清空数据框队列，否则选择批量导出，即在主界面输入“-1”后将数据框队列中的所有数据框一次性导出到Excel工作簿中。）\nDo you want to export data to Excel as soon as data organization finishes? (Submit any non-empty string to select Single Export and clear the dataframe queue, or null to select Batch Export, which means to export all dataframes in the dataframe queue to an Excel workbook at one time after submitting "-1" at the home screen.)')
                            single_export_str: str = logInput()
                            single_export = bool(single_export_str)
                            if single_export:
                                extractor.df_queue.clear() #避免数据框被重复导出，降低效率（Avoid dataframes of same types from being exported over and over again, which reduces the efficiency）
                                logPrint("每个类型的数据在整理完成后将直接导出到Excel工作簿中，而不会添加到数据框队列中。数据框队列已清空，且批量导出选项已禁用。\nData of each type will be exported to an Excel workbook directly after data organization finishes, but won't be added into the dataframe queue. The dataframe queue has been cleared, and Batch Export option has been disabled.")
                            else:
                                logPrint('每个类型的数据将只用来构建数据框，而不会立刻导出。您可以输入“-1”以导出队列中的所有数据框。批量导出选项已启用。\nData of each type will only be used to build dataframes but not be exported immediately. You may submit "-1" to export all dataframes in the queue. Batch Export option has been enabled.')
                        elif option[0] == "3":
                            logPrint("是否将数据导出为网页？（输入任意非空字符串以将数据导出为网页，否则不导出。）\nDo you want to export data into web pages? (Submit any non-empty string to export data into web pages, or null to refuse.)")
                            web_export_str: str = logInput()
                            web_export = bool(web_export_str)
                            if web_export:
                                logPrint("数据将导出为网页。\nData will be exported into web pages.")
                            else:
                                logPrint("数据将不会导出为网页。\nData won't be exported into web pages.")
                        elif option[0] == "4":
                            old_locale: str = language_code
                            language_code = set_locale(initial_launch = False, old_locale = old_locale)
                            if language_code != old_locale:
                                logPrint("说明文本将使用%s。\nTooltips will be in %s." %(language_ddragon[language_code]["desc_zh"], language_ddragon[language_code]["desc_en"]))
                                extractor.set_language(language_code)
                                logPrint("正在加载字符串常量池……\nLoading stringtables ...", print_time = True)
                                extractor.init_strtable_readiness()
                                extractor.get_strtable()
                                if not (extractor.strtable_organize_manner == 1 and extractor.strtables_ready["lol_target"] and extractor.strtables_ready["lol_default"] and extractor.strtables_ready["tft_target"] and extractor.strtables_ready["tft_default"]) and not (extractor.strtable_organize_manner == 2 and extractor.strtables_ready["target"] and extractor.strtables_ready["default"]):
                                    continue
                        elif option[0] == "5":
                            logPrint("是否保留说明文本的原始样式？（输入任意非空字符串以保留原始CSS样式；否则移除所有CSS样式，用统一的标点符号进行强调。）\nDo you want to reserve the original style of tooltips? (Input any non-empty string to reserve the original CSS style; otherwise, remove all CSS styles and use the unified punctuation marks for emphasis.)")
                            reserve_CSS_str: str = logInput()
                            reserve_CSS: bool = bool(reserve_CSS_str)
                            extractor.set_tooltip_layout(reserve_CSS = reserve_CSS)
                            if reserve_CSS:
                                logPrint("说明文本将保留原始CSS标签。\nCSS tags will be reserved in the tooltips.")
                            else:
                                logPrint("说明文本将移除所有CSS标签。\nCSS tags will be removed from the tooltips.")
                        elif option[0] == "6":
                            logPrint('是否在数值替换的同时保留原变量？（输入任意非空字符串以将转换后的变量写成“[{变量名}] = {值}”的形式，否则只保留值。）\nDo you want to reserve the original variable when variable substitution is being performed? (Input any non-empty string to transform the variable into the form "[{Var_name}] = {Value}", or null to reserve the value only.)')
                            reserve_variable_str: str = logInput()
                            reserve_variable: bool = bool(reserve_variable_str)
                            extractor.set_variable_reserve_strategy(reserve_variable)
                            if reserve_variable:
                                logPrint("说明文本在完成变量代换后将同时显示变量名和值。\nBoth the name and the value of variables will appear in the tooltip after variable substitution.")
                            else:
                                logPrint("说明文本在完成变量代换后将只显示值。\nOnly the value of variables will appear in the tooltip after variable substitution.")
                        elif option[0] == "7":
                            logPrint("是否启用hash值深度解析模式？（输入任意非空字符串以重新计算一段二进制描述数据中所有字符串的hash值并寻找其原始字符串以统一大小写，否则只对数据中已有的hash值进行解析。）\nDo you want to enable the deep resolution mode of hash value? (Input any non-empty string to recompute the hash values of all strings in a piece of binary description data and find their original strings to unify the cases, or null to only resolve the hash values already in the data.)")
                            deep_resolve_hash_str: str = logInput()
                            deep_resolve_hash: bool = bool(deep_resolve_hash_str)
                            if deep_resolve_hash != LoLDataExtractor.deep_resolve_hash:
                                extractor.set_resolution_depth(deep_resolve_hash) #修改对象的类属性可以应用到其它对象，因此不需要在`preset_settings`中保存这个设置（Modifying the class attribute of the object can be applied to other objects, so there's no need to save this setting in `preset_settings`）
                                extractor.clear_cache(clear_online = True) #由于需要切换hash值的解析程度，数据资源的内容会发生变化，所以这里应清除在线缓存（Because the degree of hash resolve is changed, the content of data resources is changed, and therefore the online cache should be cleared）
                                logPrint("已清空缓存。\nCache cleared.")
                                if one_click:
                                    preset_data_options.clear()
                                    logPrint("已清空应用到后续版本的数据类型设置。\nCleared types of data to be exported for subsequent versions.")
                            if deep_resolve_hash:
                                logPrint("已启用hash值深度解析模式。\nEnabled deep resolution mode of hash value.")
                            else:
                                logPrint("已禁用hash值深度解析模式。\nDisabled deep resolution mode of hash value.")
                        elif option[0] == "8":
                            logPrint(f"请设置等级计算的等级上限。输入空字符串以取消更改。\nPlease set the level cap for level scaling calculations. Submit an empty string to cancel the change.\n当前等级上限（Current level cap）：{extractor.levelScaling_cap}")
                            levelScaling_cap_str: str = logInput()
                            if levelScaling_cap_str.isdigit():
                                levelScaling_cap: int = int(levelScaling_cap_str)
                                extractor.set_levelScaling_cap(levelScaling_cap)
                                logPrint("等级上限已修改。\nLevel cap changed.")
                        elif option[0] == "9":
                            logPrint("是否启用密集导出？（输入任意非空字符串以密集导出，从而消除空字段；否则保留所有可导出的字段。）\nDo you want to enable dense export? (Submit any non-empty string to export dataframes in a dense manner to remove all empty fields, or null to reserve all fields that can be exported.)")
                            dense_export_str: str = logInput()
                            dense_export: bool = bool(dense_export_str)
                            extractor.set_export_density(dense_export)
                            if dense_export:
                                logPrint("数据框在导出前将消除空字段。\nEmpty fields will be removed before dataframes are exported.")
                            else:
                                logPrint("数据框的所有可用字段将保留。\nAll available fields will be reserved when dataframes are exported.")
                        else:
                            logPrint("您的输入有误！请重新输入。\nERROR input. Please try again.")
                            continue
                        logPrint("请选择一个配置：\nPlease select an configuration option:\n0\t返回上一层（Return to the last step）\n1\t导出为工作簿（Export to workbook）\n2\t单类数据导出（Single-type data export）\n3\t导出为网页（Export to web）\n4\t切换语言（Switch language）\n5\t说明文本样式（Tooltip style）\n6\t变量替换样式（Variable substitution style）\n7\thash值解析深度（Hash value resolution depth）\n8\t等级计算上限（Level scaling cap）\n9\t切换数据框导出密度（Switch dataframe export density）")
                elif not single_export and mode == "-1":
                    df_queue: list[dict[str, Any]] = sorted(extractor.df_queue, key = lambda x: x["order"])
                    if len(df_queue) > 0:
                        logPrint("正在导出数据……\nExporting data ...", print_time = True)
                        while True:
                            try:
                                if not os.path.exists(extractor.wbPath):
                                    wbCreateFlag: bool = create_workbook_win32(os.path.abspath(extractor.wbPath))
                                workbook_exist: bool = os.path.exists(extractor.wbPath)
                                with (pandas.ExcelWriter(extractor.wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(extractor.wbPath, mode = "w")) as writer:
                                    for j in range(len(df_queue)):
                                        df_struct: dict[str, Any] = df_queue[j]
                                        df: pandas.DataFrame = df_struct["sheet"]
                                        sheet_name: str = df_struct["sheet_name"]
                                        export_note: str = " (Skipped.)" if len(df) == 1 else ""
                                        logPrint("[%d/%d]%s%s" %(j + 1, len(df_queue), sheet_name, export_note), end = "\r", print_time = True)
                                        if len(df) > 1: #只导出非空数据框。每个数据框有一行中文表头（Only non-empty dataframes are exported. Each dataframe has a Chinese header）
                                            columns_to_drop: list[str] = [column for column in df.columns if df[column].astype(str).str.len().max() > 32767] #存储单元格长度超过Excel限制的列（Store columns with cell length exceeding Excel limit）
                                            df = df.drop(labels = columns_to_drop, axis = 1)
                                            if extractor.dense_export:
                                                df = eliminate_empty_fields(df)
                                            if df_struct.get("T", False):
                                                df = df.transpose()
                                            addDefaultStyle(df).to_excel(excel_writer = writer, sheet_name = sheet_name[:31])
                                            worksheet: Worksheet = writer.sheets[sheet_name[:31]]
                                            if worksheet.calculate_dimension() != "A1:A1":
                                                worksheet.cell(row = 1, column = 1, value = extractor.patch) #在A1单元格填充数据所在版本（Fill in A0 cell with the data version）
                                    else:
                                        logPrint("已完成。 | Done.")
                            except PermissionError:
                                logPrint('''无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试，或者输入“0”以放弃导出。\nPermission denied! Please ensure the file isn't opened right now or read-only! Submit any string to try again, or submit "0" to quit exporting.''')
                                cont: str = logInput()
                                if cont != "" and cont[0] == "0":
                                    break
                            else:
                                logPrint(f"数据已导出到{extractor.wbPath}。\nData have been exported to {extractor.wbPath}.", print_time = True)
                                break
                        del df_queue
                        extractor.df_queue.clear()
                    else:
                        logPrint("没有等待导出的数据。\nNo data waiting for export.")
                elif mode[0] == "0":
                    extractor.clear_cache()
                    if one_click:
                        if len(versions_conf) == 1:
                            preset_data_options = sorted(set(preset_data_options))
                        else: #这个分支不可能包含`len(versions_conf) == 0`的情形，因为在循环的开头执行了一步追加操作。所以这个分支的含义是在一键式导出的过程中一个批次即将结束的场景（This function can't contain the case where `len(versions_conf) == 0`, for an `append`` method is called at the beginning of the loop. Therefore, the meaning of this condition is the case where a batch is about to end when one-click export is enabled）
                            preset_data_options.clear() #在下一个批次开始时，重新设置要导出的数据选项（When the next batch begins, set the types of data to export again）
                    break
                else:
                    data_options: list[int] = []
                    if mode == "all":
                        data_options = list(range(1, 15))
                    else:
                        try:
                            tmp = eval(mode)
                        except Exception as e:
                            logPrint(e, write_time = False)
                            logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                        else:
                            if isinstance(tmp, int):
                                if tmp >= 1 and tmp <= 14:
                                    data_options = [tmp]
                                else:
                                    logPrint("您输入的正整数不在合法范围内。请重新输入。\nThe integer you input doesn't fall within a legal range. Please try again.")
                            elif isinstance(tmp, Iterable) and all(map(lambda x: isinstance(x, int), tmp)):
                                data_options = [_ for _ in tmp if _ >= 1 and _ <= 14]
                            else:
                                logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                    if one_click and len(versions_conf) == 1:
                        preset_data_options.extend(data_options)
                    nDataOptions += len(data_options)
                    for j in range(len(data_options)):
                        nDataOption_iter += 1
                        dOption: int = data_options[j]
                        if dOption == 1:
                            logPrint("[%d/%d][%d/%d]正在整理地图数据……\nOrganizing map data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            mapExtractor: MapExtractor = MapExtractor(extractor)
                            mapExtractor.build_map_dataframe()
                            if wb_export:
                                if single_export:
                                    mapExtractor.export_map_data()
                                else:
                                    mapExtractor.enqueue_map_dataframe()
                        elif dOption == 2:
                            logPrint("[%d/%d][%d/%d]正在整理作弊指令数据……\nOrganizing cheat data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            cheatExtractor: CheatExtractor = CheatExtractor(extractor)
                            cheatExtractor.build_cheat_dataframe()
                            if wb_export:
                                if single_export:
                                    cheatExtractor.export_cheat_data()
                                else:
                                    cheatExtractor.enqueue_cheat_dataframe()
                        elif dOption == 3:
                            logPrint("[%d/%d][%d/%d]正在整理召唤师技能数据……\nOrganizing summoner spell data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            summonerSpellExtractor: SummonerSpellExtractor = SummonerSpellExtractor(extractor)
                            summonerSpellExtractor.build_summonerSpell_dataframe()
                            if wb_export:
                                if single_export:
                                    summonerSpellExtractor.export_summonerSpell_data()
                                else:
                                    summonerSpellExtractor.enqueue_summonerSpell_dataframe()
                            if web_export:
                                summonerSpellExtractor.to_html()
                        elif dOption == 4:
                            logPrint("[%d/%d][%d/%d]正在整理符文数据……\nOrganizing perk data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            perkExtractor: PerkExtractor = PerkExtractor(extractor)
                            perkExtractor.build_perk_dataframe()
                            if wb_export:
                                if single_export:
                                    perkExtractor.export_perk_data()
                                else:
                                    perkExtractor.enqueue_perk_dataframe()
                            if web_export:
                                perkExtractor.to_html()
                        elif dOption == 5:
                            logPrint("[%d/%d][%d/%d]正在整理英雄数据……\nOrganizing champion data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            championExtractor1: ChampionExtractor = ChampionExtractor(extractor)
                            championExtractor1.set_mode(False)
                            championExtractor1.build_champion_dataframe()
                            if wb_export:
                                if single_export:
                                    championExtractor1.export_champion_data()
                                else:
                                    championExtractor1.enqueue_champion_dataframe()
                            if web_export:
                                championExtractor1.to_html()
                        elif dOption == 6:
                            logPrint("[%d/%d][%d/%d]正在整理角色数据……\nOrganizing character data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            championExtractor2: ChampionExtractor = ChampionExtractor(extractor)
                            championExtractor2.set_mode(True)
                            championExtractor2.build_champion_dataframe()
                            if wb_export:
                                if single_export:
                                    championExtractor2.export_champion_data()
                                else:
                                    championExtractor2.enqueue_champion_dataframe()
                            if web_export:
                                championExtractor2.to_html()
                        elif dOption == 7:
                            logPrint("[%d/%d][%d/%d]正在整理装备数据……\nOrganizing item data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            itemExtractor: ItemExtractor = ItemExtractor(extractor)
                            itemExtractor.build_item_dataframe()
                            if wb_export:
                                if single_export:
                                    itemExtractor.export_item_data()
                                else:
                                    itemExtractor.enqueue_item_dataframe()
                            if web_export:
                                itemExtractor.to_html()
                        elif dOption == 8:
                            logPrint("[%d/%d][%d/%d]正在整理强化符文数据……\nOrganizing augment data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            augmentExtractor: AugmentExtractor = AugmentExtractor(extractor)
                            augmentExtractor.build_augment_dataframe()
                            if wb_export:
                                if single_export:
                                    augmentExtractor.export_augment_data()
                                else:
                                    augmentExtractor.enqueue_augment_dataframe()
                            if web_export:
                                augmentExtractor.to_html()
                        elif dOption == 9:
                            logPrint("[%d/%d][%d/%d]正在整理锻造器数据……\nOrganizing anvil data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            anvilExtractor: AnvilExtractor = AnvilExtractor(extractor)
                            anvilExtractor.build_anvil_dataframe()
                            if wb_export:
                                if single_export:
                                    anvilExtractor.export_anvil_data()
                                else:
                                    anvilExtractor.enqueue_anvil_dataframe()
                            if web_export:
                                anvilExtractor.to_html()
                        elif dOption == 10:
                            logPrint("[%d/%d][%d/%d]正在整理斗魂竞技场回合数据……\nOrganizing Arena round data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            cherryRoundExtractor: CherryRoundExtractor = CherryRoundExtractor(extractor)
                            cherryRoundExtractor.build_CherryRound_dataframe()
                            if wb_export:
                                if single_export:
                                    cherryRoundExtractor.export_CherryRound_data()
                                else:
                                    cherryRoundExtractor.enqueue_CherryRound_dataframe()
                            if web_export:
                                cherryRoundExtractor.to_html()
                        elif dOption == 11:
                            logPrint("[%d/%d][%d/%d]正在整理场景英雄数据……\nOrganizing Cameo data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            cameoExtractor: CameoExtractor = CameoExtractor(extractor)
                            cameoExtractor.build_cameo_dataframe()
                            if wb_export:
                                if single_export:
                                    cameoExtractor.export_cameo_data()
                                else:
                                    cameoExtractor.enqueue_cameo_dataframe()
                        elif dOption == 12:
                            logPrint("[%d/%d][%d/%d]正在整理荣誉嘉宾数据……\nOrganizing Guest of Honor data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            gohExtractor: GoHExtractor = GoHExtractor(extractor)
                            gohExtractor.build_GoH_dataframe()
                            if wb_export:
                                if single_export:
                                    gohExtractor.export_GoH_data()
                                else:
                                    gohExtractor.enqueue_GoH_dataframe()
                            if web_export:
                                gohExtractor.to_html()
                        elif dOption == 13:
                            logPrint("[%d/%d][%d/%d]正在整理云顶之弈数据……\nOrganizing TFT data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            tftExtractor: TFTExtractor = TFTExtractor(extractor)
                            tftExtractor.build_tft_dataframe()
                            if wb_export:
                                if single_export:
                                    tftExtractor.export_tft_data()
                                else:
                                    tftExtractor.enqueue_tft_dataframe()
                            if web_export:
                                tftExtractor.to_html()
                        elif dOption == 14:
                            logPrint("[%d/%d][%d/%d]正在整理字体数据……\nOrganizing font data ..." %(i + 1, len(versions), nDataOption_iter, nDataOptions))
                            fontExtractor: FontExtractor = FontExtractor(extractor)
                            fontExtractor.build_font_dataframe()
                            if wb_export:
                                if single_export:
                                    fontExtractor.export_font_data()
                                else:
                                    fontExtractor.enqueue_font_dataframe()
        else:
            print("是否排序工作表？（输入任意非空字符串以排序，否则不排序。）\nDo you want to sort the worksheets? (Submit any non-empty string to sort, or null to refuse sorting.)")
            sort_sheet_str: str = logInput()
            sort_sheet: bool = bool(sort_sheet_str)
            if sort_sheet:
                wbPaths: list[str] = []
                if integrate:
                    logPrint("请输入一个含有多版本数据的工作簿路径。\nPlease the path of a workbook that contains data of multiple patches.")
                else:
                    logPrint('请依次输入每个版本的工作簿路径。输入“-1”以结束。\nPlease input the paths of workbooks of single patches one by one. Enter "-1" to cancel.')
                while True:
                    wbPath: str = logInput()
                    if wbPath == "":
                        continue
                    elif wbPath == "-1" and not integrate:
                        break
                    elif os.path.exists(wbPath) and os.path.isfile(wbPath) and os.path.splitext(wbPath)[1] == ".xlsx":
                        wbPaths.append(wbPath.replace("\\", "/"))
                        if integrate:
                            break
                    elif not os.path.exists(wbPath):
                        logPrint("文件未找到。请重新输入。\nFile not found. Please try again.")
                    elif not os.path.isfile(wbPath):
                        logPrint("请输入一个文件。\nPlease provide a file.")
                    else:
                        logPrint('文件格式错误。请输入以“.xlsx”为后缀的文件。\nFile format error. Please provide a file with ".xlsx" extension.')
                for i in range(len(wbPaths)):
                    wbPath: str = wbPaths[i]
                    logPrint("[%d/%d]正在排序（Ordering）： %s" %(i + 1, len(wbPaths), wbPath), print_time = True)
                    sort_workbook_sheets(wbPath, 2 if integrate else 1)
                else:
                    if len(wbPaths) > 0:
                        logPrint("排序完成。程序即将退出。\nOrder finished. The program will exit soon.")
        return 0

    #定义调试函数。开发者可在其中随时修改代码（Define the debug function. Code in this function may be modified at will）
    def debug(dir_type: Literal["repo", "extract"] = "repo", locale: str = "zh_CN") -> int:
        '''
        调试函数，用于测试。<br>Debug function for the beta version.

        主要读取离线数据资源。<br>Mainly loads offline data resources.

        本人设备上默认使用的数据资源目录：<br>The default data resource directory used on my device:<br>C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe
        
        本人设备上经过LoL-Wad-Extract-Tencent存储库提取的测试服的游戏文本文件的默认目录：<br>The default directory for text files of PBE extracted by LoL-Wad-Extract-Tencent repository:<br>D:/Workspace/LoL-Wad-Extract-Riot/pbe-text
        
        :param dir_type: 目录类型。有以下两个取值：<br>Type of directory, which has two values:
        
            - **repo**: LoL-Dragon-Change-S16存储库中的测试服文件夹。<br>PBE folder under LoL-Dragon-Change-S16 repository.
            - **extract**: 通过LoL-Wad-Extract-Tencent存储库提取测试服时指定的目的文件夹。<br>The destination / target folder specified when extracting PBE data using LoL-Wad-Extract-Tencent repository.
        :type dir_type: str
        :param locale: 字符串常量池的目标语言文化代码。默认为简体中文。<br>Target language code of stringtables. Chinese Simplified by default.
        :type locale: str
        :return: 状态码。<br>Status code.
        :rtype: int
        '''
        extract_game_dir: Path = Path("D:/Workspace/LoL-Wad-Extract-Riot/pbe-text/Game/DATA/FINAL/")
        extract_plugins_dir: Path = Path("D:/Workspace/LoL-Wad-Extract-Riot/pbe-text/Plugins/")
        repo_game_dir: Path = Path("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/")
        repo_plugins_dir: Path = Path("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/plugins/")
        #设置语言（Set the language）
        DEFAULT_LOCALE: str = LoLDataExtractor.DEFAULT_LOCALE
        if not locale in language_ddragon:
            return 1
        
        #设置版本（Set the version）
        version = "pbe"
        
        #设置默认导出行为（Set the default export behavior）
        wb_export: bool = True
        web_export: bool = False
        # logPrint('数据将只导出为Excel工作簿。如果需要导出为网页，请在选择数据类型的步骤输入“-2”以设置导出选项。\nData will only be exported to Excel workbooks. If you want to export data into web, please input "-2" in the data type selection step to set export options.')
        single_export: bool = False
        # logPrint('''数据在完成整理后不会立刻导出为Excel工作簿。如果需要在整理后立刻导出，请在选择数据类型的步骤输入“-2”以设置导出选项。\nData won't be exported into Excel workbooks immediately after being organized. If you want to export data immediately after they're organized, please input "-2" in the data type selection step to set export options.''')
        
        #设置样式保留行为（Set CSS retention behavior）
        LoLDataExtractor.set_tooltip_layout(False)
        # logPrint('''说明文本变量代换过程默认不保留CSS样式。如果需要保留，请在选择数据类型的步骤输入“-2”以调整样式选项。\nCSS styles aren't retained during the variable substitution process of tooltips by default. If you want to retain them, please input "-2" in the data type selection step to set the CSS retention option.''')
        
        #设置变量代换过程中的变量名保留行为（Set the variable name retention behavior in the variable substitution process）
        LoLDataExtractor.set_variable_reserve_strategy(False)
        # logPrint('''说明文本变量代换过程默认不保留变量名。如果需要保留，请在选择数据类型的步骤输入“-2”以调整变量代换选项。\nVariable names aren't retained during the variable substitution process of tooltips by default. If you want to retain them, please input "-2" in the data type selection step to set the variable name retention option.''')
        
        #设置hash值解析深度（Set the hash resolution depth）
        LoLDataExtractor.set_resolution_depth(False)
        # logPrint('程序默认只解析hash值。如果需要统一不同版本间的字符串大小写，请在选择数据类型的步骤输入“-2”以设置hash值解析深度。\nThe program only resolves hash values by default. If you want to unify the string cases among different versions, please input "-2" in the data type selection step to set the hash resolution depth.')
        
        #设置等级计算的等级上限（Set the level cap for level scaling calculations）
        LoLDataExtractor.set_levelScaling_cap(18)
        # logPrint('等级计算的等级上限默认为18级。如果需要调整，请在选择数据类型的步骤输入“-2”以调整等级上限。\nThe level cap for level scaling calculations is 18 by default. If you want to adjust it, please input "-2" in the data type selection step to adjust the level cap.')
        
        #设置数据框导出密度（Set dataframe export density）
        LoLDataExtractor.set_export_density(True)
        # logPrint('程序默认消除空字段。如果需要保留所有数据框的可导出的列，请在选择数据类型的步骤输入“-2”以设置导出密度。\nThe program removes empty fields. If you want to reserve all dataframe columns that can be exported, please input "-2" in the data type selection step to set the export density.')
        
        #设置工作表集成（Determine whether to integrate sheets in different patches into one workbook）
        logPrint("是否将不同版本的工作表集成到一个工作簿中？（输入任意非空字符串以确认集成，否则分不同版本保存。）\nDo you want to integrate sheets of different versions into a single workbook? (Input any non-empty string to confirm integration, or null to save data into multiple workbooks of the different version.)")
        integrate_str: str = logInput()
        integrate: bool = bool(integrate_str)
        
        logPrint(f"开始处理%s版本的游戏数据。\nStart to process game data of Version %s." %(version, version))
        extractor = LoLDataExtractor(version, locale, log = log)
        if integrate:
            extractor.encapsulate()
        else:
            extractor.decapsulate()
        #加载二进制描述数据的字符串散列表（Load the string hashtable for binary description data）
        logPrint("正在加载二进制描述数据的字符串散列表……\nLoading the string hashtable for binary description data ...", print_time = True)
        bin_hash_paths: list[str] = [
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.binentries.txt",
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.binfields.txt",
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.binhashes.txt",
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.bintypes.txt",
        ]
        extractor.read_bin_hashes(bin_hash_paths = bin_hash_paths)
        #加载版本数据（Load version data）
        logPrint(f"正在加载完整的游戏版本号……\nLoading the complete version number ...", print_time = True)
        if dir_type == "extract":
            game_version_path: str = "D:/Workspace/LoL-Wad-Extract-Riot/pbe-text/Game/content-metadata.json"
        else:
            game_version_path = "C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/compat-version-metadata.json"
        extractor.read_version(game_version_path)
        if extractor.patch == "" or extractor.patch_number == "":
            return 0
        print(f"当前版本号（Current patch）： {extractor.patch}")
        #创建保存目录（Create saving directory）
        extractor.make_dir()
        #加载文件导出列表（Load file export list）
        logPrint(f"正在加载文件导出列表……\nLoading the file export list ...", print_time = True)
        extractor.read_exported_files("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/cdragon/files.exported.txt")
        # if not extractor.fileExportList_ready:
        #     return 0
        #加载中英文字符串常量池（Load the stringtable in Chinese and English）
        logPrint("正在加载字符串常量池……\nLoading stringtables ...", print_time = True)
        if dir_type == "extract":
            strtable_paths: list[Path] = [
                extract_game_dir / locale.lower() / "data/menu/en_us/lol.stringtable.json",
                extract_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/lol.stringtable.json",
                extract_game_dir / locale.lower() / "data/menu/en_us/tft.stringtable.json",
                extract_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/tft.stringtable.json",
            ]
        else:
            strtable_paths = [
                repo_game_dir / locale.lower() / "data/menu/en_us/lol.stringtable.json",
                repo_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/lol.stringtable.json",
                repo_game_dir / locale.lower() / "data/menu/en_us/tft.stringtable.json",
                repo_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/tft.stringtable.json",
            ]
        extractor.read_strtable(strtable_paths = list(map(lambda x: x.as_posix(), strtable_paths)))
        if not (extractor.strtable_organize_manner == 1 and extractor.strtables_ready["lol_target"] and extractor.strtables_ready["lol_default"] and extractor.strtables_ready["tft_target"] and extractor.strtables_ready["tft_default"]) and not (extractor.strtable_organize_manner == 2 and extractor.strtables_ready["target"] and extractor.strtables_ready["default"]):
            return 0
        #加载共享数据（Load shared data）
        logPrint("正在加载共享数据……\nLoading shared data ...", print_time = True)
        if dir_type == "extract":
            shared_bin_path: Path = extract_game_dir / "shared.cdtb.bin.json"
        else:
            shared_bin_path = repo_game_dir / "shared.cdtb.bin.json"
        extractor.init_mSpells(debug = True, path = shared_bin_path.as_posix())
        if not extractor.shared_ready:
            # logPrint("共享数据获取失败。将忽略该数据。\nShared data capture failure! The program will ignore them.")
            return 0
        #初始化计数器（Initialize counter）
        nDataOptions: int = 0
        nDataOption_iter: int = 0
        #设置要提取的数据类型（Set the type of data to extract）
        while True:
            logPrint("请选择您要提取的数据：\nPlease select the type of data you want to extract:\n-3\t调试（Debug）\n-2\t设置（Settings）\n0\t退出当前版本（Quit this version）\n1\t地图（Maps）\n2\t作弊指令（Cheat sheet）\n3\t召唤师技能（Summoner Spells）\n4\t符文（Perks）\n5\t英雄（Champions）\n6\t角色（Characters）\n7\t装备（Items）\n8\t强化符文（Augments）\n9\t锻造器（Anvils）\n10\t斗魂竞技场回合阶段（Arena Round Phase）\n11\t场景英雄（Cameo）\n12\t荣誉嘉宾（Guests of Honor）\n13\t云顶之弈赛季、装备和羁绊（TFT Sets, Items and Traits）\n14\t字体（Fonts）\nall\t所有（All）" + ("" if single_export else "\n-1\t批量导出所有数据框并清空队列（Batch export all dataframes and clear queue）"))
            mode: str = logInput()
            if mode == "":
                continue
            elif mode == "-3":
                logPrint("请选择草稿选项：\nPlease select a draft option:\n0\t退出调试（Quit debug）\n1\t启动子环境（Start a sub-environment）")
                while True:
                    draft_option = logInput()
                    if draft_option == "":
                        continue
                    elif draft_option[0] == "0":
                        break
                    elif draft_option[0] == "1":
                        scope: dict[str, Any] = {
                            "LogManager": LogManager,
                            "Patch": Patch,
                            "requestUrl": requestUrl,
                            "format_df": format_df,
                            "verifyDictHeterogeneity": verifyDictHeterogeneity,
                            "syncListOrder": syncListOrder,
                            "traverse_keyPath": traverse_keyPath,
                            "getBinaryKeys": getBinaryKeys,
                            "LoLDataExtractor": LoLDataExtractor,
                            "MapExtractor": MapExtractor,
                            "CheatExtractor": CheatExtractor,
                            "PerkExtractor": PerkExtractor,
                            "ChampionExtractor": ChampionExtractor,
                            "ItemExtractor": ItemExtractor,
                            "AugmentExtractor": AugmentExtractor,
                            "AnvilExtractor": AnvilExtractor,
                            "CherryRoundExtractor": CherryRoundExtractor,
                            "CameoExtractor": CameoExtractor,
                            "GoHExtractor": GoHExtractor,
                            "TFTExtractor": TFTExtractor,
                            "FontExtractor": FontExtractor,
                            "modeOverrideTooltipTransform": modeOverrideTooltipTransform,
                            "extractor": extractor
                        }
                        if "mapExtractor" in dir():
                            scope["mapExtractor"] = mapExtractor
                        if "cheatExtractor" in dir():
                            scope["cheatExtractor"] = cheatExtractor
                        if "perkExtractor" in dir():
                            scope["perkExtractor"] = perkExtractor
                        if "summonerSpellExtractor" in dir():
                            scope["summonerSpellExtractor"] = summonerSpellExtractor
                        if "championExtractor1" in dir():
                            scope["championExtractor"] = championExtractor1
                        if "championExtractor2" in dir():
                            scope["championExtractor"] = championExtractor2
                        if "itemExtractor" in dir():
                            scope["itemExtractor"] = itemExtractor
                        if "augmentExtractor" in dir():
                            scope["augmentExtractor"] = augmentExtractor
                        if "anvilExtractor" in dir():
                            scope["anvilExtractor"] = anvilExtractor
                        if "cherryRoundExtractor" in dir():
                            scope["cherryRoundExtractor"] = cherryRoundExtractor
                        if "cameoExtractor" in dir():
                            scope["cameoExtractor"] = cameoExtractor
                        if "gohExtractor" in dir():
                            scope["gohExtractor"] = gohExtractor
                        if "tftExtractor" in dir():
                            scope["tftExtractor"] = tftExtractor
                        if "fontExtractor" in dir():
                            scope["fontExtractor"] = fontExtractor
                        logPrint('示例（Examples）：\nprint(dir())\nlog: LogManager = LogManager()\nlogInput = log.logInput\nlogPrint = log.logPrint\nlogPrint(format_df(mapExtractor.map_df)[0], write_time = False)\n输入“-1”以退出调试。\nSubmit "-1" to quit debug.')
                        subscope(scope, log = log)
                    else:
                        logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                    logPrint("请选择草稿选项：\nPlease select a draft option:\n0\t退出调试（Quit debug）\n1\t启动子环境（Start a sub-environment）")
            elif mode == "-2":
                logPrint("请选择一个配置：\nPlease select an configuration option:\n0\t返回上一层（Return to the last step）\n1\t导出为工作簿（Export to workbook）\n2\t单类数据导出（Single-type data export）\n3\t导出为网页（Export to web）\n4\t切换语言（Switch language）\n5\t说明文本样式（Tooltip style）\n6\t变量替换样式（Variable substitution style）\n7\thash值解析深度（Hash value resolution depth）\n8\t等级计算上限（Level scaling cap）\n9\t切换数据框导出密度（Switch dataframe export density）")
                while True:
                    option: str = logInput()
                    if option == "":
                        continue
                    elif option == "-1":
                        return 0
                    elif option[0] == "0":
                        break
                    elif option[0] == "1":
                        logPrint("是否将数据导出为Excel工作簿？（输入任意非空字符串以将数据导出为Excel工作簿，否则不导出。）\nDo you want to export data into Excel workbooks? (Submit any non-empty string to export data into Excel workbooks, or null to refuse.)")
                        wb_export_str: str = logInput()
                        wb_export = bool(wb_export_str)
                        if wb_export:
                            logPrint("数据将导出为Excel工作簿。\nData will be exported into Excel workbooks.")
                        else:
                            logPrint("数据将不会导出为Excel工作簿。\nData won't be exported into Excel workbooks.")
                    elif option[0] == "2":
                        logPrint('是否选择在整理数据后立刻将其导出到Excel中？（输入任意非空字符串以选择单项导出并清空数据框队列，否则选择批量导出，即在主界面输入“-1”后将数据框队列中的所有数据框一次性导出到Excel工作簿中。）\nDo you want to export data to Excel as soon as data organization finishes? (Submit any non-empty string to select Single Export and clear the dataframe queue, or null to select Batch Export, which means to export all dataframes in the dataframe queue to an Excel workbook at one time after submitting "-1" at the home screen.)')
                        single_export_str: str = logInput()
                        single_export = bool(single_export_str)
                        if single_export:
                            extractor.df_queue.clear() #避免数据框被重复导出，降低效率（Avoid dataframes of same types from being exported over and over again, which reduces the efficiency）
                            logPrint("每个类型的数据在整理完成后将直接导出到Excel工作簿中，而不会添加到数据框队列中。数据框队列已清空，且批量导出选项已禁用。\nData of each type will be exported to an Excel workbook directly after data organization finishes, but won't be added into the dataframe queue. The dataframe queue has been cleared, and Batch Export option has been disabled.")
                        else:
                            logPrint('每个类型的数据将只用来构建数据框，而不会立刻导出。您可以输入“-1”以导出队列中的所有数据框。批量导出选项已启用。\nData of each type will only be used to build dataframes but not be exported immediately. You may submit "-1" to export all dataframes in the queue. Batch Export option has been enabled.')
                    elif option[0] == "3":
                        logPrint("是否将数据导出为网页？（输入任意非空字符串以将数据导出为网页，否则不导出。）\nDo you want to export data into web pages? (Submit any non-empty string to export data into web pages, or null to refuse.)")
                        web_export_str: str = logInput()
                        web_export = bool(web_export_str)
                        if web_export:
                            logPrint("数据将导出为网页。\nData will be exported into web pages.")
                        else:
                            logPrint("数据将不会导出为网页。\nData won't be exported into web pages.")
                    elif option[0] == "4":
                        old_locale: str = locale
                        locale = set_locale(initial_launch = False, old_locale = old_locale)
                        if locale != old_locale:
                            logPrint("说明文本将使用%s。\nTooltips will be in %s." %(language_ddragon[locale]["desc_zh"], language_ddragon[locale]["desc_en"]))
                            extractor.set_language(locale)
                            logPrint("正在加载字符串常量池……\nLoading stringtables ...", print_time = True)
                            extractor.init_strtable_readiness()
                            if dir_type == "extract":
                                strtable_paths: list[Path] = [
                                    extract_game_dir / locale.lower() / "data/menu/en_us/lol.stringtable.json",
                                    extract_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/lol.stringtable.json",
                                    extract_game_dir / locale.lower() / "data/menu/en_us/tft.stringtable.json",
                                    extract_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/tft.stringtable.json",
                                ]
                            else:
                                strtable_paths = [
                                    repo_game_dir / locale.lower() / "data/menu/en_us/lol.stringtable.json",
                                    repo_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/lol.stringtable.json",
                                    repo_game_dir / locale.lower() / "data/menu/en_us/tft.stringtable.json",
                                    repo_game_dir / DEFAULT_LOCALE.lower() / "data/menu/en_us/tft.stringtable.json",
                                ]
                            extractor.read_strtable(strtable_paths = list(map(lambda x: x.as_posix(), strtable_paths)))
                            if not (extractor.strtable_organize_manner == 1 and extractor.strtables_ready["lol_target"] and extractor.strtables_ready["lol_default"] and extractor.strtables_ready["tft_target"] and extractor.strtables_ready["tft_default"]) and not (extractor.strtable_organize_manner == 2 and extractor.strtables_ready["target"] and extractor.strtables_ready["default"]):
                                return 0
                    elif option[0] == "5":
                        logPrint("是否保留说明文本的原始样式？（输入任意非空字符串以保留原始CSS样式；否则移除所有CSS样式，用统一的标点符号进行强调。）\nDo you want to reserve the original style of tooltips? (Input any non-empty string to reserve the original CSS style; otherwise, remove all CSS styles and use the unified punctuation marks for emphasis.)")
                        reserve_CSS_str: str = logInput()
                        reserve_CSS: bool = bool(reserve_CSS_str)
                        extractor.set_tooltip_layout(reserve_CSS = reserve_CSS)
                        if reserve_CSS:
                            logPrint("说明文本将保留原始CSS标签。\nCSS tags will be reserved in the tooltips.")
                        else:
                            logPrint("说明文本将移除所有CSS标签。\nCSS tags will be removed from the tooltips.")
                    elif option[0] == "6":
                        logPrint('是否在数值替换的同时保留原变量？（输入任意非空字符串以将转换后的变量写成“[{变量名}] = {值}”的形式，否则只保留值。）\nDo you want to reserve the original variable when variable substitution is being performed? (Input any non-empty string to transform the variable into the form "[{Var_name}] = {Value}", or null to reserve the value only.)')
                        reserve_variable_str: str = logInput()
                        reserve_variable: bool = bool(reserve_variable_str)
                        extractor.set_variable_reserve_strategy(reserve_variable = reserve_variable)
                        if reserve_variable:
                            logPrint("说明文本在完成变量代换后将同时显示变量名和值。\nBoth the name and the value of variables will appear in the tooltip after variable substitution.")
                        else:
                            logPrint("说明文本在完成变量代换后将只显示值。\nOnly the value of variables will appear in the tooltip after variable substitution.")
                    elif option[0] == "7":
                        logPrint("是否启用hash值深度解析模式？（输入任意非空字符串以重新计算一段二进制描述数据中所有字符串的hash值并寻找其原始字符串以统一大小写，否则只对数据中已有的hash值进行解析。）\nDo you want to enable the deep resolution mode of hash value? (Input any non-empty string to recompute the hash values of all strings in a piece of binary description data and find their original strings to unify the cases, or null to only resolve the hash values already in the data.)")
                        deep_resolve_hash_str: str = logInput()
                        deep_resolve_hash: bool = bool(deep_resolve_hash_str)
                        if deep_resolve_hash != LoLDataExtractor.deep_resolve_hash:
                            extractor.set_resolution_depth(deep_resolve_hash)
                            extractor.clear_cache(clear_online = True)
                            logPrint("已清空缓存。\nCache cleared.")
                        if deep_resolve_hash:
                            logPrint("已启用hash值深度解析模式。\nEnabled deep resolution mode of hash value.")
                        else:
                            logPrint("已禁用hash值深度解析模式。\nDisabled deep resolution mode of hash value.")
                    elif option[0] == "8":
                        logPrint(f"请设置等级计算的等级上限。输入空字符串以取消更改。\nPlease set the level cap for level scaling calculations. Submit an empty string to cancel the change.\n当前等级上限（Current level cap）：{extractor.levelScaling_cap}")
                        levelScaling_cap_str: str = logInput()
                        if levelScaling_cap_str.isdigit():
                            levelScaling_cap: int = int(levelScaling_cap_str)
                            extractor.set_levelScaling_cap(levelScaling_cap)
                            logPrint("等级上限已修改。\nLevel cap changed.")
                    elif option[0] == "9":
                        logPrint("是否启用密集导出？（输入任意非空字符串以密集导出，从而消除空字段；否则保留所有可导出的字段。）\nDo you want to enable dense export? (Submit any non-empty string to export dataframes in a dense manner to remove all empty fields, or null to reserve all fields that can be exported.)")
                        dense_export_str: str = logInput()
                        dense_export: bool = bool(dense_export_str)
                        extractor.set_export_density(dense_export)
                        if dense_export:
                            logPrint("数据框在导出前将消除空字段。\nEmpty fields will be removed before dataframes are exported.")
                        else:
                            logPrint("数据框的所有可用字段将保留。\nAll available fields will be reserved when dataframes are exported.")
                    else:
                        logPrint("您的输入有误！请重新输入。\nERROR input. Please try again.")
                        continue
                    logPrint("请选择一个配置：\nPlease select an configuration option:\n0\t返回上一层（Return to the last step）\n1\t导出为工作簿（Export to workbook）\n2\t单类数据导出（Single-type data export）\n3\t导出为网页（Export to web）\n4\t切换语言（Switch language）\n5\t说明文本样式（Tooltip style）\n6\t变量替换样式（Variable substitution style）\n7\thash值解析深度（Hash value resolution depth）\n8\t等级计算上限（Level scaling cap）\n9\t切换数据框导出密度（Switch dataframe export density）")
            elif not single_export and mode == "-1":
                df_queue: list[dict[str, Any]] = sorted(extractor.df_queue, key = lambda x: x["order"])
                if len(df_queue) > 0:
                    logPrint("正在导出数据……\nExporting data ...", print_time = True)
                    while True:
                        try:
                            if not os.path.exists(extractor.wbPath):
                                wbCreateFlag: bool = create_workbook_win32(os.path.abspath(extractor.wbPath))
                            workbook_exist: bool = os.path.exists(extractor.wbPath)
                            with (pandas.ExcelWriter(extractor.wbPath, mode = "a", if_sheet_exists = "replace") if workbook_exist else pandas.ExcelWriter(extractor.wbPath, mode = "w")) as writer:
                                for j in range(len(df_queue)):
                                    df_struct: dict[str, Any] = df_queue[j]
                                    df: pandas.DataFrame = df_struct["sheet"]
                                    sheet_name: str = df_struct["sheet_name"]
                                    export_note: str = " (Skipped.)" if len(df) == 1 else ""
                                    logPrint("[%d/%d]%s%s" %(j + 1, len(df_queue), sheet_name, export_note), end = "\r", print_time = True)
                                    if len(df) > 1:
                                        columns_to_drop: list[str] = [column for column in df.columns if df[column].astype(str).str.len().max() > 32767]
                                        df = df.drop(labels = columns_to_drop, axis = 1)
                                        if extractor.dense_export:
                                            df = eliminate_empty_fields(df)
                                        if df_struct.get("T", False):
                                            df = df.transpose()
                                        addDefaultStyle(df).to_excel(excel_writer = writer, sheet_name = sheet_name[:31])
                                        worksheet: Worksheet = writer.sheets[sheet_name[:31]]
                                        if worksheet.calculate_dimension() != "A1:A1":
                                            worksheet.cell(row = 1, column = 1, value = extractor.patch) #在A1单元格填充数据所在版本（Fill in A0 cell with the data version）
                                else:
                                    logPrint("已完成。 | Done.")
                        except PermissionError:
                            logPrint('''无写入权限！请确保文件未被打开且非只读状态！输入任意键以重试，或者输入“0”以放弃导出。\nPermission denied! Please ensure the file isn't opened right now or read-only! Submit any string to try again, or submit "0" to quit exporting.''')
                            cont: str = logInput()
                            if cont != "" and cont[0] == "0":
                                break
                        else:
                            logPrint(f"数据已导出到{extractor.wbPath}。\nData have been exported to {extractor.wbPath}.", print_time = True)
                            break
                    del df_queue
                    extractor.df_queue.clear()
                else:
                    logPrint("没有等待导出的数据。\nNo data waiting for export.")
            elif mode[0] == "0":
                extractor.clear_cache()
                break
            else:
                data_options: list[int] = []
                if mode == "all":
                    data_options = list(range(1, 15))
                else:
                    try:
                        tmp = eval(mode)
                    except Exception as e:
                        logPrint(e, write_time = False)
                        logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                    else:
                        if isinstance(tmp, int):
                            if tmp >= 1 and tmp <= 14:
                                data_options = [tmp]
                            else:
                                logPrint("您输入的正整数不在合法范围内。请重新输入。\nThe integer you input doesn't fall within a legal range. Please try again.")
                        elif isinstance(tmp, Iterable) and all(map(lambda x: isinstance(x, int), tmp)):
                            data_options = [_ for _ in tmp if _ >= 1 and _ <= 14]
                        else:
                            logPrint("您的输入有误！请重新输入。\nERROR input! Please try again.")
                nDataOptions += len(data_options)
                for i in range(len(data_options)):
                    nDataOption_iter += 1
                    dOption: int = data_options[i]
                    if dOption == 1:
                        logPrint("[%d/%d]正在调试地图数据……\nDebugging map data ..." %(nDataOption_iter, nDataOptions))
                        mapExtractor: MapExtractor = MapExtractor(extractor)
                        if dir_type == "extract":
                            map_paths: list[Path] = [
                                extract_game_dir / "data/maps/shipping/map11/map11.bin.json",
                                extract_game_dir / "data/maps/shipping/map12/map12.bin.json",
                                extract_game_dir / "data/maps/shipping/map21/map21.bin.json",
                                extract_game_dir / "data/maps/shipping/map22/map22.bin.json",
                                extract_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                extract_game_dir / "data/maps/shipping/map33/map33.bin.json",
                                extract_game_dir / "data/maps/shipping/map35/map35.bin.json",
                                extract_game_dir / "data/maps/shipping/map453/map453.bin.json"
                            ]
                        else:
                            map_paths = [
                                repo_game_dir / "data/maps/shipping/map11/map11.bin.json",
                                repo_game_dir / "data/maps/shipping/map12/map12.bin.json",
                                repo_game_dir / "data/maps/shipping/map21/map21.bin.json",
                                repo_game_dir / "data/maps/shipping/map22/map22.bin.json",
                                repo_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                repo_game_dir / "data/maps/shipping/map33/map33.bin.json",
                                repo_game_dir / "data/maps/shipping/map35/map35.bin.json",
                                repo_game_dir / "data/maps/shipping/map453/map453.bin.json"
                            ]
                        mapExtractor.build_map_dataframe(debug = True, paths = list(map(lambda x: x.as_posix(), map_paths)))
                        if wb_export:
                            if single_export:
                                mapExtractor.export_map_data()
                            else:
                                mapExtractor.enqueue_map_dataframe()
                    elif dOption == 2:
                        logPrint("[%d/%d]正在调试作弊指令数据……\nDebugging cheat data ..." %(nDataOption_iter, nDataOptions))
                        cheatExtractor: CheatExtractor = CheatExtractor(extractor)
                        if dir_type == "extract":
                            cheat_path: Path = extract_game_dir / "cheats.cdtb.bin.json"
                        else:
                            cheat_path = repo_game_dir / "cheats.cdtb.bin.json"
                        cheatExtractor.build_cheat_dataframe(debug = True, path = cheat_path.as_posix())
                        if wb_export:
                            if single_export:
                                cheatExtractor.export_cheat_data()
                            else:
                                cheatExtractor.enqueue_cheat_dataframe()
                    elif dOption == 3:
                        logPrint("[%d/%d]正在调试召唤师技能数据……\nDebugging summoner spell data ..." %(nDataOption_iter, nDataOptions))
                        summonerSpellExtractor: SummonerSpellExtractor = SummonerSpellExtractor(extractor)
                        if dir_type == "extract":
                            summonerSpell_path: Path = extract_game_dir / "shared.cdtb.bin.json"
                        else:
                            summonerSpell_path = repo_game_dir / "shared.cdtb.bin.json"
                        summonerSpellExtractor.build_summonerSpell_dataframe(debug = True, path = summonerSpell_path.as_posix())
                        if wb_export:
                            if single_export:
                                summonerSpellExtractor.export_summonerSpell_data()
                            else:
                                summonerSpellExtractor.enqueue_summonerSpell_dataframe()
                        if web_export:
                            summonerSpellExtractor.to_html()
                    elif dOption == 4:
                        logPrint("[%d/%d]正在调试符文数据……\nDebugging perk data ..." %(nDataOption_iter, nDataOptions))
                        perkExtractor: PerkExtractor = PerkExtractor(extractor)
                        if dir_type == "extract":
                            perk_path: Path = extract_game_dir / "perks.cdtb.bin.json"
                        else:
                            perk_path = repo_game_dir / "perks.cdtb.bin.json"
                        perkExtractor.build_perk_dataframe(debug = True, path = perk_path.as_posix())
                        if wb_export:
                            if single_export:
                                perkExtractor.export_perk_data()
                            else:
                                perkExtractor.enqueue_perk_dataframe()
                        if web_export:
                            perkExtractor.to_html()
                    elif dOption == 5:
                        logPrint("[%d/%d]正在调试英雄数据……\nDebugging champion data ..." %(nDataOption_iter, nDataOptions))
                        championExtractor1: ChampionExtractor = ChampionExtractor(extractor)
                        championExtractor1.set_mode(False)
                        if dir_type == "extract":
                            champion_paths: list[Path] = [
                                extract_plugins_dir / "rcp-be-lol-game-data/global/zh_cn/v1/champion-summary.json", #这里的语言文化代码可为任意值，因为在英雄提取器的读取英雄数据方法中，这个文件只是用来提取别名进一步确定各英雄二进制描述文件的路径的（Here the locale can be any value, for in `ChampionExtractor.read_champion_data`, this file is only used to determine paths of each champion's binary description file path）
                                extract_game_dir / "data/characters"
                            ]
                        else:
                            champion_paths = [
                                repo_plugins_dir / "rcp-be-lol-game-data/global/zh_cn/v1/champion-summary.json",
                                repo_game_dir / "data/characters"
                            ]
                        championExtractor1.build_champion_dataframe(debug = True, paths = list(map(lambda x: x.as_posix(), champion_paths)))
                        if wb_export:
                            if single_export:
                                championExtractor1.export_champion_data()
                            else:
                                championExtractor1.enqueue_champion_dataframe()
                        if web_export:
                            championExtractor1.to_html()
                    elif dOption == 6:
                        logPrint("[%d/%d]正在调试角色数据……\nDebugging character data ..." %(nDataOption_iter, nDataOptions))
                        championExtractor2: ChampionExtractor = ChampionExtractor(extractor)
                        championExtractor2.set_mode(True)
                        if dir_type == "extract":
                            character_paths: list[Path] = [
                                extract_plugins_dir / "rcp-be-lol-game-data/global/zh_cn/v1/champion-summary.json",
                                extract_game_dir / "data/maps/shipping/map22/map22.bin.json",
                                extract_game_dir / "data/characters",
                                extract_game_dir / "characters"
                            ]
                        else:
                            character_paths = [
                                repo_plugins_dir / "rcp-be-lol-game-data/global/zh_cn/v1/champion-summary.json",
                                repo_game_dir / "data/maps/shipping/map22/map22.bin.json",
                                repo_game_dir / "data/characters",
                                repo_game_dir / "characters"
                            ]
                        championExtractor2.build_champion_dataframe(debug = True, paths = list(map(lambda x: x.as_posix(), character_paths)))
                        if wb_export:
                            if single_export:
                                championExtractor2.export_champion_data()
                            else:
                                championExtractor2.enqueue_champion_dataframe()
                        if web_export:
                            championExtractor2.to_html()
                    elif dOption == 7:
                        logPrint("[%d/%d]正在调试装备数据……\nDebugging item data ..." %(nDataOption_iter, nDataOptions))
                        itemExtractor: ItemExtractor = ItemExtractor(extractor)
                        if dir_type == "extract":
                            item_path: Path = extract_game_dir / "items.cdtb.bin.json"
                        else:
                            item_path = repo_game_dir / "items.cdtb.bin.json"
                        itemExtractor.build_item_dataframe(debug = True, path = item_path.as_posix())
                        if wb_export:
                            if single_export:
                                itemExtractor.export_item_data()
                            else:
                                itemExtractor.enqueue_item_dataframe()
                        if web_export:
                            itemExtractor.to_html()
                    elif dOption == 8:
                        logPrint("[%d/%d]正在调试强化符文数据……\nDebugging augment data ..." %(nDataOption_iter, nDataOptions))
                        augmentExtractor: AugmentExtractor = AugmentExtractor(extractor)
                        if dir_type == "extract":
                            augment_paths: list[Path] = [
                                extract_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                extract_game_dir / "maps/modespecificdata/cherry.bin.json",
                                extract_game_dir / "data/maps/shipping/map33/map33.bin.json",
                                extract_game_dir / "data/maps/shipping/map12/map12.bin.json",
                                extract_game_dir / "maps/modespecificdata/kiwi.bin.json",
                                extract_game_dir / "maps/modespecificdata/kiwi_jade.bin.json"
                            ]
                        else:
                            augment_paths = [
                                repo_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                repo_game_dir / "maps/modespecificdata/cherry.bin.json",
                                repo_game_dir / "data/maps/shipping/map33/map33.bin.json",
                                repo_game_dir / "data/maps/shipping/map12/map12.bin.json",
                                repo_game_dir / "maps/modespecificdata/kiwi.bin.json",
                                repo_game_dir / "maps/modespecificdata/kiwi_jade.bin.json"
                            ]
                        augmentExtractor.build_augment_dataframe(debug = True, paths = list(map(lambda x: x.as_posix(), augment_paths)))
                        if wb_export:
                            if single_export:
                                augmentExtractor.export_augment_data()
                            else:
                                augmentExtractor.enqueue_augment_dataframe()
                        if web_export:
                            augmentExtractor.to_html()
                    elif dOption == 9:
                        logPrint("[%d/%d]正在调试锻造器数据……\nDebugging anvil data ..." %(nDataOption_iter, nDataOptions))
                        anvilExtractor: AnvilExtractor = AnvilExtractor(extractor)
                        if dir_type == "extract":
                            anvil_paths: list[Path] = [
                                extract_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                extract_game_dir / "data/maps/shipping/map12/map12.bin.json"
                            ]
                        else:
                            anvil_paths = [
                                repo_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                repo_game_dir / "data/maps/shipping/map12/map12.bin.json"
                            ]
                        anvilExtractor.build_anvil_dataframe(debug = True, paths = list(map(lambda x: x.as_posix(), anvil_paths)))
                        if wb_export:
                            if single_export:
                                anvilExtractor.export_anvil_data()
                            else:
                                anvilExtractor.enqueue_anvil_dataframe()
                        if web_export:
                            anvilExtractor.to_html()
                    elif dOption == 10:
                        logPrint("[%d/%d]正在调试斗魂竞技场回合数据……\nDebugging Arena round data ..." %(nDataOption_iter, nDataOptions))
                        cherryRoundExtractor: CherryRoundExtractor = CherryRoundExtractor(extractor)
                        if dir_type == "extract":
                            CherryRound_path: Path = extract_game_dir / "data/maps/shipping/map30/map30.bin.json"
                        else:
                            CherryRound_path = repo_game_dir / "data/maps/shipping/map30/map30.bin.json"
                        cherryRoundExtractor.build_CherryRound_dataframe(debug = True, path = CherryRound_path.as_posix())
                        if wb_export:
                            if single_export:
                                cherryRoundExtractor.export_CherryRound_data()
                            else:
                                cherryRoundExtractor.enqueue_CherryRound_dataframe()
                        if web_export:
                            cherryRoundExtractor.to_html()
                    elif dOption == 11:
                        logPrint("[%d/%d]正在调试场景英雄数据……\nDebugging Cameo data ..." %(nDataOption_iter, nDataOptions))
                        cameoExtractor: CameoExtractor = CameoExtractor(extractor)
                        if dir_type == "extract":
                            cameoPath: Path = extract_game_dir / "data/maps/shipping/map30/map30.bin.json"
                        else:
                            cameoPath = repo_game_dir / "data/maps/shipping/map30/map30.bin.json"
                        cameoExtractor.build_cameo_dataframe(debug = True, path = cameoPath.as_posix())
                        if wb_export:
                            if single_export:
                                cameoExtractor.export_cameo_data()
                            else:
                                cameoExtractor.enqueue_cameo_dataframe()
                    elif dOption == 12:
                        logPrint("[%d/%d]正在调试荣誉嘉宾数据……\nDebugging Guest of Honor data ..." %(nDataOption_iter, nDataOptions))
                        gohExtractor: GoHExtractor = GoHExtractor(extractor)
                        if dir_type == "extract":
                            GoHPaths: list[Path] = [
                                extract_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                extract_game_dir / "maps/modespecificdata/cherry.bin.json"
                            ]
                        else:
                            GoHPaths = [
                                repo_game_dir / "data/maps/shipping/map30/map30.bin.json",
                                repo_game_dir / "maps/modespecificdata/cherry.bin.json"
                            ]
                        gohExtractor.build_GoH_dataframe(debug = True, paths = list(map(lambda x: x.as_posix(), GoHPaths)))
                        if wb_export:
                            if single_export:
                                gohExtractor.export_GoH_data()
                            else:
                                gohExtractor.enqueue_GoH_dataframe()
                        if web_export:
                            gohExtractor.to_html()
                    elif dOption == 13:
                        logPrint("[%d/%d]正在调试云顶之弈数据……\nDebugging TFT data ..." %(nDataOption_iter, nDataOptions))
                        tftExtractor: TFTExtractor = TFTExtractor(extractor)
                        if dir_type == "extract":
                            map22_path: Path = extract_game_dir / "data/maps/shipping/map22/map22.bin.json"
                        else:
                            map22_path = repo_game_dir / "data/maps/shipping/map22/map22.bin.json"
                        tftExtractor.build_tft_dataframe(debug = True, path = map22_path.as_posix())
                        if wb_export:
                            if single_export:
                                tftExtractor.export_tft_data()
                            else:
                                tftExtractor.enqueue_tft_dataframe()
                        if web_export:
                            tftExtractor.to_html()
                    elif dOption == 14:
                        logPrint("[%d/%d]正在调试字体数据……\nDebugging font data ..." %(nDataOption_iter, nDataOptions))
                        fontExtractor: FontExtractor = FontExtractor(extractor)
                        if dir_type == "extract":
                            font_path: Path = extract_game_dir / "ux/font.cdtb.bin.json"
                        else:
                            font_path = repo_game_dir / "ux/fonts.cdtb.bin.json"
                        fontExtractor.build_font_dataframe(debug = True, path = font_path.as_posix())
                        if wb_export:
                            if single_export:
                                fontExtractor.export_font_data()
                            else:
                                fontExtractor.enqueue_font_dataframe()
        return 0

    #个性化函数（Personalized function）
    def DIY() -> int: #该函数将会分为数据资源、数据准备和说明文本转换部分，并将随时补充。在VSCode中，按Ctrl-Q以注释或解除注释其中的部分（This function is basically divided into three parts: data resource, data preparation and tooltip transformation, and always receives supplement. In VSCode, press Ctrl—Q to comment or uncomment out regions）
        '''
        个性化函数，用于进一步调试。<br>Personalized function for further debug use.
        
        :return: 状态码。总是0。<br>Status code. Always return 0.
        :rtype: int
        '''
        locale: str = "zh_CN"
        #数据资源（Data resource）
        ##二进制条目散列表（Binary entry hash table）
        bin_hash_paths: list[str] = [
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.binentries.txt",
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.binfields.txt",
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.binhashes.txt",
            "C:/Users/19250/AppData/Local/cdragon/data/hashes/lol/hashes.bintypes.txt",
        ]
        LoLDataExtractor("", locale).read_bin_hashes(bin_hash_paths = bin_hash_paths)
        ##字符串常量池（Stringtable）
        lolstringtable_zh_path = "C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/zh_cn/data/menu/en_us/lol.stringtable.json"
        with open(lolstringtable_zh_path, "r", encoding = "utf-8") as fp:
            lolstringtable_zh: dict[str, int | dict[str, str]] = json.load(fp)
        lolstringtable_en_path = "C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/en_us/data/menu/en_us/lol.stringtable.json"
        with open(lolstringtable_en_path, "r", encoding = "utf-8") as fp:
            lolstringtable_en: dict[str, int | dict[str, str]] = json.load(fp)
        tftstringtable_zh_path = "C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/zh_cn/data/menu/en_us/tft.stringtable.json"
        with open(tftstringtable_zh_path, "r", encoding = "utf-8") as fp:
            tftstringtable_zh: dict[str, int | dict[str, str]] = json.load(fp)
        tftstringtable_en_path = "C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/en_us/data/menu/en_us/tft.stringtable.json"
        with open(tftstringtable_en_path, "r", encoding = "utf-8") as fp:
            tftstringtable_en: dict[str, int | dict[str, str]] = json.load(fp)
        ##地图（Map）
        # with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/data/maps/shipping/map22/map22.bin.json", "r", encoding = "utf-8") as fp:
        #     map22_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # map22_bin = LoLDataExtractor.resolve_bin_hash(map22_bin)
        # with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/data/maps/shipping/map30/map30.bin.json", "r", encoding = "utf-8") as fp:
        #     map30_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # map30_bin = LoLDataExtractor.resolve_bin_hash(map30_bin)
        # with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/data/maps/shipping/map33/map33.bin.json", "r", encoding = "utf-8") as fp:
        #     map33_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # map33_bin = LoLDataExtractor.resolve_bin_hash(map33_bin)
        ##装备（Item）
        with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/items.cdtb.bin.json", "r", encoding = "utf-8") as fp:
            items_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # items_bin = LoLDataExtractor.resolve_bin_hash(items_bin)
        ##共享数据（Shared data）
        # with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/shared.cdtb.bin.json", "r", encoding = "utf-8") as fp:
        #     shared_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # shared_bin = LoLDataExtractor.resolve_bin_hash(shared_bin)
        ##符文（Perk）
        # with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/perks.cdtb.bin.json", "r", encoding = "utf-8") as fp:
        #     perks_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # perks_bin = LoLDataExtractor.resolve_bin_hash(perks_bin)
        ##强化符文和荣誉嘉宾（Augment and Guest of Honor）
        # with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/maps/modespecificdata/cherry.bin.json", "r", encoding = "utf-8") as fp:
        #     cherry_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # cherry_bin = LoLDataExtractor.resolve_bin_hash(cherry_bin)
        # with open("C:/Users/19250/Documents/GitHub/LoL-Dragon-Change-S16/Data/cdragon/pbe/game/maps/modespecificdata/kiwi.bin.json", "r", encoding = "utf-8") as fp:
        #     kiwi_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # kiwi_bin = LoLDataExtractor.resolve_bin_hash(kiwi_bin)
        ##整合后的数据（Merged data）
        with open("C:/Users/19250/Documents/Workspace/JupyterLab/英雄联盟数据提取/champions_bin.json", "r", encoding = "utf-8") as fp:
            champions_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # champions_bin = LoLDataExtractor.resolve_bin_hash(champions_bin)
        # with open("C:/Users/19250/Documents/Workspace/JupyterLab/英雄联盟数据提取/characters_bin.json", "r", encoding = "utf-8") as fp:
        #     characters_bin: dict[str, list[str] | dict[str, Any]] = json.load(fp)
        # characters_bin = LoLDataExtractor.resolve_bin_hash(characters_bin)
        
        #数据准备（Data preparation）
        # for (key, value) in shared_bin.items():
        #     if key != "__linked" and value["__type"] == "SpellObject":
        #         LoLDataExtractor.mSpells[value["mScriptName"]] = value
        # for (key, value) in characters_bin.items():
        #     if key != "__linked" and value["__type"] == "SpellObject":
        #         LoLDataExtractor.mSpells[value["mScriptName"]] = value
        # for (key, value) in map22_bin.items():
        #     if key != "__linked" and value["__type"] == "TftUnitPropertyDefinition":
        #         LoLDataExtractor.TFTUnitPropertyMap[value["name"]] = value
        #     elif key != "__linked" and value["__type"] == "TftTraitData":
        #         LoLDataExtractor.TFTTraitMap[value["mName"]] = value
        #     elif key != "__linked" and value["__type"] == "ScriptDataObject":
        #         LoLDataExtractor.TFTScriptDataMap[value["mName"]] = value
        
        #总结数据结构（Summarize the data structure）
        # keyDict: dict[str, dict[str, int]] = getBinaryKeys(map33_bin, isBin = True, keyPaths = None, objectTypes = "AugmentData")[1]
        # logPrint(json.dumps(keyDict, indent = 4, ensure_ascii = False))
        # import pyperclip
        # s: str = ""
        # for key in keyDict["{fa33a427}"]:
        #     s += key + "\n"
        # pyperclip.copy(s)
        
        #输出键的hash值（Output hash value of a key）
        # logPrint(LoLDataExtractor.compute_rsthash("Spell_TFT17_PykeSpell_Name", 5))
        
        #键对应（Key map）
        # mDisplayName_key = "Item_2523_Name"
        # logPrint(LoLDataExtractor.get_strtable_value(lolstringtable_zh, mDisplayName_key, default = "获取失败。"))
        
        #说明文本转换（Tooltip transformation）
        # logPrint("说明文本测试样例：")
        # tests: list[dict[str, Any]] = [
        #     {
        #         "tooltip": "<rules>对野怪的百分比生命值伤害的上限为<magicDamage>@SuperQMonsterMaxDamageTotal@</magicDamage>。</rules>",
        #         "binData": champions_bin["Characters/Galio/Spells/GalioQAbility/GalioQ"]["mSpell"],
        #         "reservedVars": None
        #     },
        # ]
        # for i in range(len(tests)):
        #     LoLDataExtractor.calculatedVariables.clear()
        #     logPrint("*" * 20)
        #     logPrint("样例%d：" %(i + 1))
        #     tooltip_raw: str = tests[i]["tooltip"]
        #     logPrint("原始说明文本：\n" + tooltip_raw)
        #     binData: dict[str, Any] = tests[i]["binData"]
        #     reservedVars: Optional[dict[str, str]] = tests[i].get("reservedVars")
        #     logPrint("----")
        #     logPrint("转换文本：")
        #     logPrint(LoLDataExtractor.tooltipTransform(tooltip_raw, lolstringtable_zh, binData, locale, enableModeOverride = True, reservedVars = reservedVars, reserve_variable = False))
        #     # logPrint(LoLDataExtractor.tooltipTransform(tooltip_raw, lolstringtable_zh, binData, locale, enableModeOverride = True, reservedVars = reservedVars, reserve_variable = True))
        #     # logPrint(LoLDataExtractor.tooltipSubstitute(tooltip_raw, lolstringtable_zh, binData, locale, enableModeOverride = True, reservedVars = reservedVars, reserve_variable = False))
        #     # logPrint(LoLDataExtractor.tooltipSubstitute(tooltip_raw, lolstringtable_zh, binData, locale, enableModeOverride = True, reservedVars = reservedVars, reserve_variable = True))
        # else:
        #     logPrint("*" * 20)
        
        #模式重载（Mode override）
        logPrint(modeOverrideTooltipTransform(champions_bin, dType = "champion", objectType = "SpellObject", keyPaths = "mSpell|DataValuesModeOverride", gameModeName = "URF", strtable = lolstringtable_zh))
        logPrint(modeOverrideTooltipTransform(items_bin, dType = "item", objectType = "ItemData", keyPaths = "DataValuesModeOverride", gameModeName = "URF", strtable = lolstringtable_zh))
        
        return 0

    def bankUnit_test() -> int:
        logPrint('请输入需要计算FNV-1 hash值的字符串。输入“-1”以退出程序。\nPlease input the string you want to calculate the FNV-1 hash value for. Submit "-1" to exit the program.')
        while True:
            event_name: str = logInput("> ")
            if event_name == "-1":
                break
            hash_result: int = LoLSfxExtractor.compute_bankEvent_hash(event_name)
            logPrint(hash_result)
        return 0

    if args.sfx:
        status: int = bankUnit_test()
    else:
        status: int = main() #供用户使用（For user use）
        # status: int = debug(dir_type = "repo") #供开发者使用（For developer use）
        # status: int = DIY()
    #结束日志输入流（Cancel the log input stream）
    log.write(f"\n[Program terminated and returned status {status}.]\n")
    log.close()
